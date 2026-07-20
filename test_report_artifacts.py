from __future__ import annotations

import os
import shutil
import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from cryptography.fernet import Fernet
from openpyxl import load_workbook


TEST_ROOT = Path(tempfile.mkdtemp(prefix="sellerpro-report-tests-"))
os.environ.setdefault("TELEGRAM_BOT_TOKEN", f"{123456}:{'A' * 35}")
os.environ.setdefault("ENCRYPTION_KEY", Fernet.generate_key().decode())
os.environ.setdefault("DB_PATH", str(TEST_ROOT / "reports.db"))
os.environ.setdefault("PDF_FONT_PATH", str(Path(__file__).with_name("DejaVuSans.ttf")))

import main  # noqa: E402
from seller_pdf_report import build_seller_pdf_report  # noqa: E402


def sample_report_payload() -> dict:
    generated = datetime(2026, 7, 20, 9, 30)
    stats = {
        "orders": 28,
        "rows": 34,
        "units": 37,
        "cancelled": 3,
        "returns": 2,
        "revenue": 12_500_000,
        "commission": 1_250_000,
        "logistics": 540_000,
        "payout_total": 10_710_000,
        "cancellation_rate": 0.081,
    }
    business = {
        "revenue": 12_500_000,
        "commission": 1_250_000,
        "logistics": 540_000,
        "payout_total": 10_710_000,
        "cost_total": 5_200_000,
        "known_profit": 4_710_000,
        "tax_expense": 500_000,
        "uzum_expense_total": 180_000,
        "advertising_expense": 200_000,
        "storage_expense": 80_000,
        "other_expense": 50_000,
        "net_profit": 3_700_000,
        "net_margin": 29.6,
        "coverage": 0.84,
        "missing_count": 1,
        "uzum_expenses_available": True,
        "complete": False,
    }
    products = [
        {
            "title": "Кофейная кружка с очень длинным названием для проверки переноса текста",
            "sku": "MUG-BLACK-500",
            "qty": 25,
            "revenue": 8_000_000,
            "commission": 800_000,
            "logistics": 300_000,
            "payout": 6_900_000,
            "cost_per_unit": 160_000,
            "cost_total": 4_000_000,
            "profit": 2_900_000,
            "margin": 36.25,
        },
        {
            "title": "Органайзер / Tartiblagich",
            "sku": "ORG-02",
            "qty": 12,
            "revenue": 4_500_000,
            "commission": 450_000,
            "logistics": 240_000,
            "payout": 3_810_000,
            "cost_per_unit": None,
            "cost_total": 0,
            "profit": None,
            "margin": None,
        },
    ]
    stock = [
        {
            "product_id": "p-1",
            "sku_id": "1001",
            "barcode": "478000000001",
            "title": products[0]["title"],
            "sku": products[0]["sku"],
            "price": 320_000,
            "purchase_price": 160_000,
            "ikpu": "12345678901234567",
            "paid_storage": True,
            "paid_storage_price_item": 250,
            "paid_storage_amount": 12_500,
            "fbo": 3,
            "fbs": 1,
            "total": 4,
            "sold_7": 14,
            "days_left": 2,
            "missing": 0,
            "defected": 1,
            "status": "LOW",
            "low_stock_threshold": 5,
            "action_ru": "Подготовить поставку",
            "action_uz": "Yetkazib berishni tayyorlash",
        },
        {
            "product_id": "p-2",
            "sku_id": "1002",
            "barcode": "478000000002",
            "title": products[1]["title"],
            "sku": products[1]["sku"],
            "price": 375_000,
            "purchase_price": None,
            "ikpu": "",
            "paid_storage": False,
            "fbo": 0,
            "fbs": 0,
            "total": 0,
            "sold_7": 3,
            "days_left": 0,
            "missing": 2,
            "defected": 0,
            "status": "OUT",
            "low_stock_threshold": 5,
            "action_ru": "Пополнить остаток и проверить purchasePrice",
            "action_uz": "Qoldiqni to‘ldirish va purchasePrice ni tekshirish",
        },
    ]
    sales = [
        {
            "date": generated - timedelta(days=1),
            "kind": "sale",
            "status": "DELIVERED",
            "order_id": "ORD-1",
            "title": products[0]["title"],
            "sku": products[0]["sku"],
            "qty": 2,
            "revenue": 640_000,
            "commission": 64_000,
            "logistics": 24_000,
            "payout": 552_000,
            "withdrawn": 0,
        },
        {
            "date": generated - timedelta(days=2),
            "kind": "cancel",
            "status": "CANCELED",
            "order_id": "ORD-2",
            "title": products[1]["title"],
            "sku": products[1]["sku"],
            "qty": 1,
            "revenue": 375_000,
            "commission": 0,
            "logistics": 0,
            "payout": 0,
            "withdrawn": 0,
            "reason": "Причина не передана Uzum",
        },
        {
            "date": generated - timedelta(days=3),
            "kind": "return",
            "status": "RETURNED",
            "order_id": "ORD-3",
            "title": products[0]["title"],
            "sku": products[0]["sku"],
            "qty": 1,
            "revenue": 320_000,
            "commission": 0,
            "logistics": 0,
            "payout": 0,
            "withdrawn": 0,
            "reason": "Возврат покупателя",
        },
    ]
    actions = [
        {
            "priority": "critical",
            "category_ru": "Остатки",
            "category_uz": "Qoldiq",
            "title": products[1]["title"],
            "sku": products[1]["sku"],
            "problem_ru": "Товар закончился",
            "problem_uz": "Tovar tugagan",
            "recommendation_ru": "Подготовить поставку сегодня",
            "recommendation_uz": "Bugun yetkazib berishni tayyorlash",
            "title_ru": "Срочно пополнить остаток",
            "title_uz": "Qoldiqni tez to‘ldirish",
            "body_ru": "Нулевой остаток уже останавливает продажи.",
            "body_uz": "Nol qoldiq savdoni to‘xtatmoqda.",
            "amount": 375_000,
            "source": "Uzum API",
        },
        {
            "priority": "warning",
            "category_ru": "Себестоимость",
            "category_uz": "Tannarx",
            "title": products[1]["title"],
            "sku": products[1]["sku"],
            "problem_ru": "Uzum не передал purchasePrice",
            "problem_uz": "Uzum purchasePrice bermadi",
            "recommendation_ru": "Проверить карточку в Uzum",
            "recommendation_uz": "Uzum kartasini tekshirish",
            "title_ru": "Проверить purchasePrice",
            "title_uz": "purchasePrice ni tekshirish",
            "body_ru": "Бот не подставляет приблизительную себестоимость.",
            "body_uz": "Bot taxminiy tannarxni qo‘ymaydi.",
            "amount": 0,
            "source": "Uzum purchasePrice",
        },
    ]
    daily = [
        {
            "date": generated - timedelta(days=6 - index),
            "label": f"{14 + index:02d}.07",
            "orders": 3 + index,
            "units": 4 + index,
            "cancelled": index % 2,
            "returns": 0,
            "revenue": 1_000_000 + index * 180_000,
            "commission": 100_000 + index * 18_000,
            "logistics": 40_000,
            "payout": 860_000 + index * 162_000,
        }
        for index in range(7)
    ]
    return {
        "shop_id": 41000,
        "generated_at": generated,
        "period_key": "7d",
        "period_days": 7,
        "period_label": "14.07.2026–20.07.2026",
        "stats": stats,
        "previous_stats": {**stats, "orders": 22, "revenue": 10_800_000},
        "comparison_available": True,
        "profit": {
            "cost_total": 5_200_000,
            "profit": 4_710_000,
            "coverage": 0.84,
            "missing_count": 1,
            "known_count": 1,
        },
        "business_profit": business,
        "finance_settings": {
            "tax_percent": 4,
            "advertising_monthly": 200_000,
            "storage_monthly": 80_000,
            "other_monthly": 50_000,
        },
        "periods": [
            {"key": "7d", **stats, "profit": 3_700_000, "cost_coverage": 0.84},
            {"key": "30d", **stats, "profit": 3_700_000, "cost_coverage": 0.84},
        ],
        "sales": sales,
        "daily": daily,
        "products": products,
        "stock": stock,
        "stock_data_available": True,
        "problems": [row for row in sales if row["kind"] != "sale"],
        "defect_events": [
            {
                "product_title": products[0]["title"],
                "sku_id": "1001",
                "defected_delta": 1,
                "defected_qty": 1,
                "estimated_value": 320_000,
            }
        ],
        "cumulative_defects": [{"defected": 1}],
        "loss_data_available": True,
        "actions": actions,
        "lost": [
            {
                **stock[1],
                "archived": False,
                "estimated_loss": 750_000,
            }
        ],
        "cost_coverage": 0.84,
        "notes": ["Один SKU без purchasePrice; итог показан только по известным данным."],
    }


class ReportArtifactSmokeTests(unittest.TestCase):
    @classmethod
    def tearDownClass(cls) -> None:
        shutil.rmtree(TEST_ROOT, ignore_errors=True)

    def test_management_workbook_reopens_and_keeps_honest_labels(self) -> None:
        payload = sample_report_payload()
        for lang, profit_sheet, result_label in (
            ("ru", "Прибыль", "Результат по известным данным"),
            ("uz", "Foyda", "Ma’lum ma’lumotlar natijasi"),
        ):
            path = TEST_ROOT / f"management-{lang}.xlsx"
            main.build_premium_workbook(payload, path, lang=lang)
            workbook = load_workbook(path, data_only=False)
            try:
                self.assertIn(profit_sheet, workbook.sheetnames)
                profit_values = [cell.value for cell in workbook[profit_sheet]["A"]]
                self.assertIn(result_label, profit_values)
                product_sheet = workbook["Товары и прибыль" if lang == "ru" else "Tovar va foyda"]
                self.assertTrue(str(product_sheet["I3"].value).startswith("=IF("))
                self.assertIsNone(product_sheet["H4"].value)
                stock_sheet = workbook["Остатки" if lang == "ru" else "Qoldiqlar"]
                self.assertEqual(stock_sheet["G3"].value, 160_000)
                self.assertEqual(stock_sheet["H3"].value, "12345678901234567")
            finally:
                workbook.close()

    def test_pdf_builds_in_russian_and_uzbek_with_bundled_fonts(self) -> None:
        payload = sample_report_payload()
        regular = Path(__file__).with_name("DejaVuSans.ttf")
        bold = Path(__file__).with_name("DejaVuSans-Bold.ttf")
        for lang in ("ru", "uz"):
            path = TEST_ROOT / f"management-{lang}.pdf"
            build_seller_pdf_report(
                payload,
                path,
                lang=lang,
                regular_font_path=regular,
                bold_font_path=bold,
            )
            data = path.read_bytes()
            self.assertTrue(data.startswith(b"%PDF-"))
            self.assertGreater(len(data), 20_000)


if __name__ == "__main__":
    unittest.main()
