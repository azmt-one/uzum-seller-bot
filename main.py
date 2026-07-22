from __future__ import annotations

import asyncio
import base64
import hashlib
import hmac
import json
import logging
import math
import os
import sqlite3
import tempfile
import time
import urllib.error
import urllib.request
import zipfile
from datetime import date, datetime, timedelta, timezone
from html import escape
from urllib.parse import urlencode
from pathlib import Path
from typing import Any, Iterable

from aiogram import BaseMiddleware, Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup, KeyboardButton, Message, ReplyKeyboardMarkup
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from db import Database, TokenCipher
from seller_pdf_report import build_seller_pdf_report
from market_plus_reports import (
    build_claim_docx,
    build_compensation_workbook,
    build_market_daily_pdf,
    build_market_daily_workbook,
    prepare_compensation_rows,
)
from subscription_automation import (
    build_reminder_draft,
    parse_reminder_days,
    select_milestone,
)
from formatters import (
    clean_num,
    compact_json_preview,
    excel_value,
    extract_items,
    flatten_sku_rows,
    format_order_line,
    format_product_line,
    format_shop_line,
    format_sku_stock_line,
    pick,
    safe,
    status_display,
)
from uzum_client import UzumClient
from uzum_finance import (
    build_stock_records,
    expense_items,
    parse_api_datetime,
    return_reminder_bucket,
    summarize_expenses,
    supply_reminder_bucket,
)

# --- Самостоятельный генератор управленческого Excel-отчёта ---
NAVY = "17365D"
BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
AMBER = "FFC000"
LIGHT_AMBER = "FFF2CC"
RED = "C00000"
LIGHT_RED = "FCE4D6"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="D9E1F2")


def _t(lang: str, ru: str, uz: str) -> str:
    return uz if str(lang).lower() == "uz" else ru


def _sheet_name(lang: str, ru: str, uz: str) -> str:
    return _t(lang, ru, uz)[:31]


def _safe_table_name(value: str) -> str:
    result = "".join(ch if ch.isalnum() else "_" for ch in value)
    if not result or result[0].isdigit():
        result = "T_" + result
    return result[:250]


def _title(ws, text: str, *, end_col: int = 12) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, text)
    cell.font = Font(size=18, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30


def _header(ws, row: int, start_col: int, end_col: int) -> None:
    for cell in ws[row][start_col - 1:end_col]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    ws.row_dimensions[row].height = 34


def _style_body(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    if end_row < start_row:
        return
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)


def _add_table(ws, start_row: int, end_row: int, end_col: int, name: str) -> None:
    if end_row <= start_row:
        return
    ref = f"A{start_row}:{ws.cell(start_row, end_col).column_letter}{end_row}"
    table = Table(displayName=_safe_table_name(name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _autowidth(ws, *, max_width: int = 42) -> None:
    for index, column in enumerate(ws.columns, start=1):
        letter = get_column_letter(index)
        length = 0
        for cell in column:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            length = max(length, max((len(x) for x in text.splitlines()), default=0))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _money(ws, columns: Iterable[str], start_row: int = 2) -> None:
    for column in columns:
        for cell in ws[column][start_row - 1:]:
            cell.number_format = '#,##0'


def _percent(ws, columns: Iterable[str], start_row: int = 2) -> None:
    for column in columns:
        for cell in ws[column][start_row - 1:]:
            cell.number_format = '0.0%'


def _date_format(ws, columns: Iterable[str], start_row: int = 2) -> None:
    for column in columns:
        for cell in ws[column][start_row - 1:]:
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = 'dd.mm.yyyy hh:mm'


def _excel_datetime(value: Any) -> Any:
    """OpenPyXL rejects timezone-aware datetimes; keep report wall time safely."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _finalize(ws, *, freeze: str = "A2", max_width: int = 42) -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    _autowidth(ws, max_width=max_width)


def _period_label(key: str, lang: str) -> str:
    labels = {
        "today": ("Сегодня", "Bugun"),
        "yesterday": ("Вчера", "Kecha"),
        "7d": ("7 дней", "7 kun"),
        "30d": ("30 дней", "30 kun"),
        "prev30d": ("Предыдущие 30 дней", "Oldingi 30 kun"),
    }
    ru, uz = labels.get(key, (key, key))
    return _t(lang, ru, uz)


def _kind_label(kind: str, lang: str) -> str:
    labels = {
        "sale": ("Продажа", "Savdo"),
        "cancel": ("Отмена", "Bekor qilish"),
        "return": ("Возврат", "Qaytarish"),
    }
    ru, uz = labels.get(kind, (kind, kind))
    return _t(lang, ru, uz)


def _priority_label(priority: str, lang: str) -> str:
    labels = {
        "critical": ("Критично", "Jiddiy"),
        "warning": ("Внимание", "Diqqat"),
        "info": ("Рекомендация", "Tavsiya"),
    }
    ru, uz = labels.get(priority, (priority, priority))
    return _t(lang, ru, uz)


def build_premium_workbook(
    payload: dict[str, Any],
    output_path: str | Path,
    *,
    lang: str = "ru",
) -> Path:
    """Build the user-facing management workbook from normalized report data."""

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    periods = list(payload.get("periods") or [])
    sales = list(payload.get("sales") or [])
    daily = list(payload.get("daily") or [])
    products = list(payload.get("products") or [])
    stock = list(payload.get("stock") or [])
    actions = list(payload.get("actions") or [])
    lost = list(payload.get("lost") or [])
    notes = list(payload.get("notes") or [])
    business_profit = dict(payload.get("business_profit") or {})
    finance_settings = dict(payload.get("finance_settings") or {})

    # 1. Executive dashboard
    ws = wb.active
    ws.title = _sheet_name(lang, "Сводка", "Xulosa")
    _title(ws, _t(lang, "Управленческий отчёт Uzum Seller", "Uzum Seller boshqaruv hisoboti"), end_col=12)
    generated_at = _excel_datetime(payload.get("generated_at"))
    ws["A2"] = _t(lang, "Магазин", "Do‘kon")
    ws["B2"] = str(payload.get("shop_id") or "—")
    ws["D2"] = _t(lang, "Сформирован", "Yaratildi")
    ws["E2"] = generated_at
    ws["G2"] = _t(lang, "Покрытие себестоимостью", "Tannarx bilan qamrov")
    ws["H2"] = float(payload.get("cost_coverage") or 0)
    ws["H2"].number_format = "0.0%"
    ws["J2"] = _t(lang, "Строк продаж загружено", "Yuklangan savdo qatorlari")
    ws["K2"] = len(sales)
    if isinstance(generated_at, datetime):
        ws["E2"].number_format = "dd.mm.yyyy hh:mm"

    summary_headers = [
        _t(lang, "Период", "Davr"),
        _t(lang, "Заказы", "Buyurtmalar"),
        _t(lang, "Позиций", "Pozitsiyalar"),
        _t(lang, "Товаров", "Tovarlar"),
        _t(lang, "Отмены", "Bekor qilish"),
        _t(lang, "Возвраты", "Qaytarish"),
        _t(lang, "Выручка", "Tushum"),
        _t(lang, "Комиссия", "Komissiya"),
        _t(lang, "Логистика", "Logistika"),
        _t(lang, "К выплате", "To‘lovga"),
        _t(lang, "Расчётная прибыль", "Hisobiy foyda"),
        _t(lang, "Покрытие затрат", "Xarajatlar qamrovi"),
    ]
    ws.append([])
    ws.append(summary_headers)
    summary_start = ws.max_row
    for period in periods:
        ws.append([
            _period_label(str(period.get("key") or ""), lang),
            int(period.get("orders") or 0),
            int(period.get("rows") or 0),
            float(period.get("units") or 0),
            int(period.get("cancelled") or 0),
            float(period.get("returns") or 0),
            float(period.get("revenue") or 0),
            float(period.get("commission") or 0),
            float(period.get("logistics") or 0),
            float(period.get("payout_total") or 0),
            period.get("profit"),
            float(period.get("cost_coverage") or 0),
        ])
    summary_end = ws.max_row
    _header(ws, summary_start, 1, 12)
    _style_body(ws, summary_start + 1, summary_end, 1, 12)
    _money(ws, ("G", "H", "I", "J", "K"), summary_start + 1)
    _percent(ws, ("L",), summary_start + 1)
    _add_table(ws, summary_start, summary_end, 12, "DashboardPeriods")

    action_row = summary_end + 3
    ws.cell(action_row, 1, _t(lang, "Что требует внимания", "Nimaga e'tibor kerak"))
    ws.cell(action_row, 1).font = Font(size=14, bold=True, color=NAVY)
    action_headers = [
        _t(lang, "Приоритет", "Ustuvorlik"),
        _t(lang, "Категория", "Toifa"),
        _t(lang, "Проблема", "Muammo"),
        _t(lang, "Что сделать", "Nima qilish kerak"),
        _t(lang, "Сумма/эффект", "Summa/ta'sir"),
    ]
    ws.append(action_headers)
    action_header = ws.max_row
    for action in actions[:12]:
        ws.append([
            _priority_label(str(action.get("priority") or "info"), lang),
            action.get("category_uz") if lang == "uz" else action.get("category_ru"),
            action.get("problem_uz") if lang == "uz" else action.get("problem_ru"),
            action.get("recommendation_uz") if lang == "uz" else action.get("recommendation_ru"),
            float(action.get("amount") or 0),
        ])
    action_end = ws.max_row
    _header(ws, action_header, 1, 5)
    _style_body(ws, action_header + 1, action_end, 1, 5)
    _money(ws, ("E",), action_header + 1)
    if action_end > action_header:
        for row in range(action_header + 1, action_end + 1):
            priority = str(ws.cell(row, 1).value or "")
            fill = LIGHT_RED if priority in {"Критично", "Jiddiy"} else LIGHT_AMBER if priority in {"Внимание", "Diqqat"} else LIGHT_GREEN
            for col in range(1, 6):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=fill)

    note_row = action_end + 3
    ws.cell(note_row, 1, _t(lang, "Важно", "Muhim"))
    ws.cell(note_row, 1).font = Font(bold=True, color=RED)
    default_note = _t(
        lang,
        (
            "Результат учитывает налог, расходы Uzum и внешние расходы. Себестоимость берётся "
            "только из Uzum purchasePrice; при неполных данных итог не считается окончательной чистой прибылью."
            if business_profit
            else "Расчётная прибыль учитывает комиссию, логистику и себестоимость Uzum purchasePrice. Налог и дополнительные расходы в неё не входят."
        ),
        (
            "Natija soliq, Uzum xarajatlari va tashqi xarajatlarni hisobga oladi. Tannarx faqat "
            "Uzum purchasePrice’dan olinadi; ma’lumot to‘liq bo‘lmasa, yakuniy sof foyda deb ko‘rsatilmaydi."
            if business_profit
            else "Hisobiy foyda komissiya, logistika va Uzum purchasePrice tannarxini hisobga oladi. Soliq va qo‘shimcha xarajatlar kiritilmagan."
        ),
    )
    ws.cell(note_row + 1, 1, default_note)
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=12)
    ws.cell(note_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    for index, note in enumerate(notes, start=note_row + 2):
        ws.cell(index, 1, str(note))
        ws.merge_cells(start_row=index, start_column=1, end_row=index, end_column=12)
        ws.cell(index, 1).alignment = Alignment(wrap_text=True, vertical="top")

    if summary_end > summary_start:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = _t(lang, "Выручка и выплаты", "Tushum va to‘lov")
        chart.y_axis.title = _t(lang, "сум", "so‘m")
        chart.x_axis.title = _t(lang, "Период", "Davr")
        data = Reference(ws, min_col=7, max_col=10, min_row=summary_start, max_row=summary_end)
        cats = Reference(ws, min_col=1, min_row=summary_start + 1, max_row=summary_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws.add_chart(chart, "N4")
    _finalize(ws, freeze=f"A{summary_start + 1}", max_width=48)

    # 2. Business result with Uzum ledger and seller-entered external expenses
    if business_profit:
        result_is_complete = bool(business_profit.get("complete"))
        result_name = _t(
            lang,
            "Чистая прибыль" if result_is_complete else "Результат по известным данным",
            "Sof foyda" if result_is_complete else "Ma’lum ma’lumotlar natijasi",
        )
        ws = wb.create_sheet(_sheet_name(lang, "Прибыль", "Foyda"))
        _title(ws, f"{result_name} — 30 {_t(lang, 'дней', 'kun')}", end_col=4)
        ws.append([
            _t(lang, "Шаг расчёта", "Hisob bosqichi"),
            _t(lang, "Знак", "Belgi"),
            _t(lang, "Сумма", "Summa"),
            _t(lang, "Что означает", "Nimani anglatadi"),
        ])
        calculation_revenue = float(
            business_profit.get("calculation_revenue", business_profit.get("revenue")) or 0
        )
        calculation_payout = float(
            business_profit.get("calculation_payout", business_profit.get("payout_total")) or 0
        )
        calculation_commission = float(
            business_profit.get("calculation_commission", business_profit.get("commission")) or 0
        )
        calculation_logistics = float(
            business_profit.get("calculation_logistics", business_profit.get("logistics")) or 0
        )
        payout_residual = calculation_revenue - calculation_payout - calculation_commission - calculation_logistics
        signed_uzum_expenses = float(business_profit.get("uzum_expense_total") or 0)
        uzum_deductions = float(
            business_profit.get("uzum_expense_deductions", max(0.0, signed_uzum_expenses)) or 0
        )
        uzum_refunds = float(
            business_profit.get("uzum_expense_refunds", max(0.0, -signed_uzum_expenses)) or 0
        )
        ws.append([
            _t(lang, "Выручка для расчёта", "Hisob uchun tushum"),
            "+",
            calculation_revenue,
            _t(lang, "Все продажи; при неполном покрытии — только SKU с purchasePrice", "Barcha savdo; qamrov to‘liq bo‘lmasa — faqat purchasePrice bor SKU"),
        ])
        ws.append([_t(lang, "Комиссия Uzum", "Uzum komissiyasi"), "−", calculation_commission, _t(lang, "Уже учтена в выплате", "To‘lovda allaqachon hisobga olingan")])
        ws.append([_t(lang, "Логистика", "Logistika"), "−", calculation_logistics, _t(lang, "Уже учтена в выплате", "To‘lovda allaqachon hisobga olingan")])
        ws.append([_t(lang, "Другие удержания внутри выплаты", "To‘lov ichidagi boshqa ushlanmalar"), "−", float(business_profit.get("other_payout_deductions", max(0.0, payout_residual)) or 0), "Uzum Finance API"])
        ws.append([_t(lang, "Корректировка выплаты", "To‘lov tuzatishi"), "+", float(business_profit.get("payout_adjustment", max(0.0, -payout_residual)) or 0), "Uzum Finance API"])
        payout_row = ws.max_row + 1
        ws.append([_t(lang, "К выплате", "To‘lovga"), "=", f"=C3-C4-C5-C6+C7", _t(lang, "Комиссия и логистика второй раз не вычитаются", "Komissiya va logistika ikkinchi marta ayrilmaydi")])
        ws.append([_t(lang, "Себестоимость", "Tannarx"), "−", float(business_profit.get("cost_total") or 0), "Uzum purchasePrice"])
        profit_before_tax_row = ws.max_row + 1
        ws.append([_t(lang, "Прибыль до налога", "Soliqdan oldingi foyda"), "=", f"=C{payout_row}-C{payout_row + 1}", _t(lang, "К выплате минус себестоимость", "To‘lovga minus tannarx")])
        ws.append([_t(lang, "Налог", "Soliq"), "−", float(business_profit.get("tax_expense") or 0), f"{float(finance_settings.get('tax_percent') or 0):.2f}%"])
        ws.append([_t(lang, "Доп. расходы Uzum", "Uzum qo‘shimcha xarajatlari"), "−", uzum_deductions, _t(lang, "Без комиссии и логистики", "Komissiya va logistikasiz")])
        ws.append([_t(lang, "Возвраты от Uzum", "Uzum qaytargan mablag‘"), "+", uzum_refunds, "Uzum Finance API"])
        ws.append([_t(lang, "Внешняя реклама", "Tashqi reklama"), "−", float(business_profit.get("advertising_expense") or 0), _t(lang, "Настройка продавца", "Sotuvchi sozlamasi")])
        ws.append([_t(lang, "Внешнее хранение", "Tashqi saqlash"), "−", float(business_profit.get("storage_expense") or 0), _t(lang, "Настройка продавца", "Sotuvchi sozlamasi")])
        ws.append([_t(lang, "Другие внешние расходы", "Boshqa tashqi xarajat"), "−", float(business_profit.get("other_expense") or 0), _t(lang, "Настройка продавца", "Sotuvchi sozlamasi")])
        result_row = ws.max_row + 1
        ws.append([
            result_name,
            "=",
            f"=C{profit_before_tax_row}-C{profit_before_tax_row + 1}-C{profit_before_tax_row + 2}+C{profit_before_tax_row + 3}-C{profit_before_tax_row + 4}-C{profit_before_tax_row + 5}-C{profit_before_tax_row + 6}",
            _t(lang, "Итог по показанной формуле", "Ko‘rsatilgan formula bo‘yicha yakun"),
        ])
        end = ws.max_row
        _header(ws, 2, 1, 4)
        _style_body(ws, 3, end, 1, 4)
        _money(ws, ("C",), 3)
        for row_number in (payout_row, profit_before_tax_row, result_row):
            for col in range(1, 5):
                ws.cell(row_number, col).font = Font(bold=True)
                ws.cell(row_number, col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for col in range(1, 5):
            ws.cell(end, col).font = Font(bold=True, color=WHITE)
            ws.cell(end, col).fill = PatternFill("solid", fgColor=GREEN if float(business_profit.get("net_profit") or 0) >= 0 else RED)
        note_row = end + 2
        coverage = float(business_profit.get("coverage") or 0) * 100
        ws.cell(note_row, 1, _t(lang, "Полнота себестоимости", "Tannarx to‘liqligi"))
        ws.cell(note_row, 2, coverage / 100)
        ws.cell(note_row, 2).number_format = "0.0%"
        ws.cell(note_row + 1, 1, _t(
            lang,
            "Себестоимость берётся только из Uzum purchasePrice. Комиссия и логистика находятся внутри суммы «К выплате» и не вычитаются повторно. Если покрытие ниже 100% или API расходов недоступен, показан только результат по известным данным.",
            "Tannarx faqat Uzum purchasePrice’dan olinadi. Komissiya va logistika «To‘lovga» summasida va qayta ayrilmaydi. Qamrov 100% dan past yoki xarajatlar API mavjud bo‘lmasa, faqat ma’lum ma’lumotlar natijasi ko‘rsatiladi.",
        ))
        ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=4)
        _finalize(ws, freeze="A3", max_width=55)

    # Complete sales ledger
    ws = wb.create_sheet(_sheet_name(lang, "Продажи", "Savdolar"))
    _title(ws, _t(lang, "Все операции за 30 дней", "30 kunlik barcha operatsiyalar"), end_col=12)
    headers = [
        _t(lang, "Дата", "Sana"), _t(lang, "Тип", "Turi"), _t(lang, "Статус", "Holat"),
        _t(lang, "ID заказа", "Buyurtma ID"), _t(lang, "Товар", "Tovar"), "SKU",
        _t(lang, "Количество", "Soni"), _t(lang, "Выручка", "Tushum"),
        _t(lang, "Комиссия", "Komissiya"), _t(lang, "Логистика", "Logistika"),
        _t(lang, "К выплате", "To‘lovga"), _t(lang, "Выведено", "Chiqarilgan"),
    ]
    ws.append(headers)
    for row in sales:
        ws.append([
            _excel_datetime(row.get("date")), _kind_label(str(row.get("kind") or "sale"), lang), row.get("status"),
            str(row.get("order_id") or "—"), row.get("title"), str(row.get("sku") or "—"),
            float(row.get("qty") or 0), float(row.get("revenue") or 0),
            float(row.get("commission") or 0), float(row.get("logistics") or 0),
            float(row.get("payout") or 0), float(row.get("withdrawn") or 0),
        ])
    end = ws.max_row
    _header(ws, 2, 1, 12)
    _style_body(ws, 3, end, 1, 12)
    _date_format(ws, ("A",), 3)
    _money(ws, ("H", "I", "J", "K", "L"), 3)
    _add_table(ws, 2, end, 12, "SalesLedger")
    _finalize(ws, freeze="A3", max_width=50)

    # 3. Daily trend
    ws = wb.create_sheet(_sheet_name(lang, "Динамика", "Dinamika"))
    _title(ws, _t(lang, "Динамика продаж по дням", "Kunlik savdo dinamikasi"), end_col=9)
    headers = [
        _t(lang, "Дата", "Sana"), _t(lang, "Заказы", "Buyurtmalar"),
        _t(lang, "Товаров", "Tovarlar"), _t(lang, "Отмены", "Bekor qilish"),
        _t(lang, "Возвраты", "Qaytarish"), _t(lang, "Выручка", "Tushum"),
        _t(lang, "Комиссия", "Komissiya"), _t(lang, "Логистика", "Logistika"),
        _t(lang, "К выплате", "To‘lovga"),
    ]
    ws.append(headers)
    for row in daily:
        ws.append([
            _excel_datetime(row.get("date")), int(row.get("orders") or 0), float(row.get("units") or 0),
            int(row.get("cancelled") or 0), float(row.get("returns") or 0),
            float(row.get("revenue") or 0), float(row.get("commission") or 0),
            float(row.get("logistics") or 0), float(row.get("payout") or 0),
        ])
    end = ws.max_row
    _header(ws, 2, 1, 9)
    _style_body(ws, 3, end, 1, 9)
    _date_format(ws, ("A",), 3)
    _money(ws, ("F", "G", "H", "I"), 3)
    _add_table(ws, 2, end, 9, "DailyTrend")
    if end >= 3:
        chart = LineChart()
        chart.style = 13
        chart.title = _t(lang, "Выручка по дням", "Kunlik tushum")
        chart.y_axis.title = _t(lang, "сум", "so‘m")
        data = Reference(ws, min_col=6, max_col=9, min_row=2, max_row=end)
        cats = Reference(ws, min_col=1, min_row=3, max_row=end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 18
        ws.add_chart(chart, "K2")
    _finalize(ws, freeze="A3", max_width=24)

    # 4. Product profitability; formulas keep calculations auditable.
    ws = wb.create_sheet(_sheet_name(lang, "Товары и прибыль", "Tovar va foyda"))
    _title(ws, _t(lang, "Аналитика товаров и расчётная прибыль", "Tovarlar va hisobiy foyda tahlili"), end_col=12)
    headers = [
        _t(lang, "Товар", "Tovar"), "SKU", _t(lang, "Продано", "Sotildi"),
        _t(lang, "Выручка", "Tushum"), _t(lang, "Комиссия", "Komissiya"),
        _t(lang, "Логистика", "Logistika"), _t(lang, "К выплате", "To‘lovga"),
        _t(lang, "Себестоимость/шт", "Tannarx/dona"), _t(lang, "Себестоимость всего", "Jami tannarx"),
        _t(lang, "Расчётная прибыль", "Hisobiy foyda"), _t(lang, "Маржа", "Marja"),
        _t(lang, "Себестоимость заполнена", "Tannarx kiritilgan"),
    ]
    ws.append(headers)
    for product in products:
        row_number = ws.max_row + 1
        cost = product.get("cost_per_unit")
        ws.append([
            product.get("title"), str(product.get("sku") or "—"), float(product.get("qty") or 0),
            float(product.get("revenue") or 0), float(product.get("commission") or 0),
            float(product.get("logistics") or 0), float(product.get("payout") or 0),
            float(cost) if cost is not None else None,
            f'=IF(H{row_number}="","",C{row_number}*H{row_number})',
            f'=IF(I{row_number}="","",G{row_number}-I{row_number})',
            f'=IFERROR(J{row_number}/D{row_number},"")',
            _t(lang, "Да", "Ha") if cost is not None else _t(lang, "Нет", "Yo‘q"),
        ])
    end = ws.max_row
    _header(ws, 2, 1, 12)
    _style_body(ws, 3, end, 1, 12)
    _money(ws, ("D", "E", "F", "G", "H", "I", "J"), 3)
    _percent(ws, ("K",), 3)
    _add_table(ws, 2, end, 12, "ProductProfitability")
    if end >= 3:
        ws.conditional_formatting.add(f"J3:J{end}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=LIGHT_RED)))
        ws.conditional_formatting.add(f"K3:K{end}", ColorScaleRule(start_type="min", start_color=RED, mid_type="percentile", mid_value=50, mid_color=AMBER, end_type="max", end_color=GREEN))
    _finalize(ws, freeze="A3", max_width=50)

    # 5. Cancellations and returns
    ws = wb.create_sheet(_sheet_name(lang, "Отмены и возвраты", "Bekor va qaytarish"))
    _title(ws, _t(lang, "Отмены и возвраты — контроль потерь", "Bekor qilish va qaytarish nazorati"), end_col=9)
    headers = [
        _t(lang, "Дата", "Sana"), _t(lang, "Тип", "Turi"), _t(lang, "Статус", "Holat"),
        _t(lang, "ID заказа", "Buyurtma ID"), _t(lang, "Товар", "Tovar"), "SKU",
        _t(lang, "Количество", "Soni"), _t(lang, "Сумма", "Summa"), _t(lang, "Потеря выплаты", "Yo‘qotilgan to‘lov"),
    ]
    ws.append(headers)
    for row in (x for x in sales if x.get("kind") in {"cancel", "return"}):
        ws.append([
            _excel_datetime(row.get("date")), _kind_label(str(row.get("kind")), lang), row.get("status"),
            str(row.get("order_id") or "—"), row.get("title"), str(row.get("sku") or "—"),
            float(row.get("qty") or 0), float(row.get("revenue") or 0), float(row.get("payout") or 0),
        ])
    end = ws.max_row
    _header(ws, 2, 1, 9)
    _style_body(ws, 3, end, 1, 9)
    _date_format(ws, ("A",), 3)
    _money(ws, ("H", "I"), 3)
    _add_table(ws, 2, end, 9, "CancellationsReturns")
    _finalize(ws, freeze="A3", max_width=50)

    # 6. Stock with sell-through forecast
    ws = wb.create_sheet(_sheet_name(lang, "Остатки", "Qoldiqlar"))
    _title(ws, _t(lang, "Остатки и прогноз пополнения", "Qoldiq va to‘ldirish prognozi"), end_col=20)
    headers = [
        "Product ID", "SKU ID", "Barcode", _t(lang, "Товар", "Tovar"), "SKU",
        _t(lang, "Цена продажи", "Sotuv narxi"), _t(lang, "Себестоимость Uzum", "Uzum tannarxi"),
        "ИКПУ / МХИК", _t(lang, "Платное хранение", "Pullik saqlash"),
        _t(lang, "Тариф хранения", "Saqlash tarifi"), _t(lang, "Начислено хранения", "Saqlash hisoblandi"),
        "FBO", "FBS/DBS", _t(lang, "Итого", "Jami"),
        _t(lang, "Продано за 7 дней", "7 kunda sotildi"), _t(lang, "Хватит, дней", "Yetadi, kun"),
        _t(lang, "Потеряно", "Yo‘qolgan"), _t(lang, "Брак", "Yaroqsiz"),
        _t(lang, "Статус", "Holat"), _t(lang, "Рекомендуемое действие", "Tavsiya etilgan amal"),
    ]
    ws.append(headers)
    for row in stock:
        ws.append([
            str(row.get("product_id") or ""), str(row.get("sku_id") or ""), str(row.get("barcode") or ""),
            row.get("title"), str(row.get("sku") or ""), float(row.get("price") or 0),
            row.get("purchase_price"), str(row.get("ikpu") or ""),
            _t(lang, "Да", "Ha") if row.get("paid_storage") else _t(lang, "Нет", "Yo‘q"),
            row.get("paid_storage_price_item"), row.get("paid_storage_amount"),
            float(row.get("fbo") or 0), float(row.get("fbs") or 0), float(row.get("total") or 0),
            float(row.get("sold_7") or 0), row.get("days_left"), float(row.get("missing") or 0),
            float(row.get("defected") or 0), row.get("status"),
            row.get("action_uz") if lang == "uz" else row.get("action_ru"),
        ])
    end = ws.max_row
    _header(ws, 2, 1, 20)
    _style_body(ws, 3, end, 1, 20)
    _money(ws, ("F", "G", "J", "K"), 3)
    _add_table(ws, 2, end, 20, "StockForecast")
    if end >= 3:
        ws.conditional_formatting.add(f"N3:N{end}", ColorScaleRule(start_type="min", start_color=RED, mid_type="percentile", mid_value=50, mid_color=AMBER, end_type="max", end_color=GREEN))
        ws.conditional_formatting.add(f"P3:P{end}", ColorScaleRule(start_type="min", start_color=RED, mid_type="percentile", mid_value=50, mid_color=AMBER, end_type="max", end_color=GREEN))
    _finalize(ws, freeze="A3", max_width=48)

    # 7. Full action list
    ws = wb.create_sheet(_sheet_name(lang, "Требует внимания", "E'tibor kerak"))
    _title(ws, _t(lang, "Готовый список действий для селлера", "Seller uchun tayyor amallar ro‘yxati"), end_col=8)
    headers = [
        _t(lang, "Приоритет", "Ustuvorlik"), _t(lang, "Категория", "Toifa"),
        _t(lang, "Товар", "Tovar"), "SKU", _t(lang, "Проблема", "Muammo"),
        _t(lang, "Что сделать", "Nima qilish kerak"), _t(lang, "Сумма/эффект", "Summa/ta'sir"),
        _t(lang, "Источник", "Manba"),
    ]
    ws.append(headers)
    for action in actions:
        ws.append([
            _priority_label(str(action.get("priority") or "info"), lang),
            action.get("category_uz") if lang == "uz" else action.get("category_ru"),
            action.get("title"), str(action.get("sku") or ""),
            action.get("problem_uz") if lang == "uz" else action.get("problem_ru"),
            action.get("recommendation_uz") if lang == "uz" else action.get("recommendation_ru"),
            float(action.get("amount") or 0), action.get("source"),
        ])
    end = ws.max_row
    _header(ws, 2, 1, 8)
    _style_body(ws, 3, end, 1, 8)
    _money(ws, ("G",), 3)
    _add_table(ws, 2, end, 8, "SellerActions")
    if end >= 3:
        for row in range(3, end + 1):
            priority = str(ws.cell(row, 1).value or "")
            fill = LIGHT_RED if priority in {"Критично", "Jiddiy"} else LIGHT_AMBER if priority in {"Внимание", "Diqqat"} else LIGHT_GREEN
            for col in range(1, 9):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=fill)
    _finalize(ws, freeze="A3", max_width=52)

    # 8. All-time lost and defective stock reported by Uzum per SKU
    ws = wb.create_sheet(_sheet_name(lang, "Потери за весь период", "Barcha davr yo‘qotish"))
    _title(ws, _t(lang, "Потерянные товары и брак за весь период", "Barcha davrdagi yo‘qolgan va yaroqsiz tovarlar"), end_col=13)
    headers = [
        "Product ID", "SKU ID", "Barcode", _t(lang, "Товар", "Tovar"), "SKU", _t(lang, "Цена продажи", "Sotuv narxi"),
        _t(lang, "Потеряно", "Yo‘qolgan"), _t(lang, "Брак", "Yaroqsiz"),
        _t(lang, "Всего проблемных", "Jami muammoli"),
        _t(lang, "Ориентировочная стоимость", "Taxminiy qiymat"),
        _t(lang, "Текущий остаток", "Joriy qoldiq"),
        _t(lang, "Карточка", "Kartochka"), _t(lang, "Статус", "Holat"),
    ]
    ws.append(headers)
    for row in lost:
        missing = float(row.get("missing") or 0)
        defected = float(row.get("defected") or 0)
        ws.append([
            str(row.get("product_id") or ""), str(row.get("sku_id") or ""), str(row.get("barcode") or ""),
            row.get("title"), str(row.get("sku") or ""), float(row.get("price") or 0),
            missing, defected, missing + defected, float(row.get("estimated_loss") or 0),
            float(row.get("total") or 0),
            _t(lang, "В архиве", "Arxivda") if row.get("archived") else _t(lang, "Активная", "Faol"),
            row.get("status"),
        ])
    end = ws.max_row
    _header(ws, 2, 1, 13)
    _style_body(ws, 3, end, 1, 13)
    _money(ws, ("F", "J"), 3)
    _add_table(ws, 2, end, 13, "LostDefectiveAllTime")
    _finalize(ws, freeze="A3", max_width=50)

    wb.save(output)
    return output

load_dotenv()
logging.basicConfig(level=logging.INFO)
logging.getLogger("httpx").setLevel(logging.WARNING)
APP_BUILD = "2026.07.17-stock-truth-forecast-v1"

TELEGRAM_BOT_TOKEN = (
    os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    or os.getenv("TELEGRAM_TOKEN", "").strip()
    or os.getenv("TOKEN", "").strip()
)
UZUM_API_BASE_URL = os.getenv(
    "UZUM_API_BASE_URL", "https://api-seller.uzum.uz/api/seller-openapi"
).strip()
DB_PATH = (
    os.getenv("DB_PATH", "").strip()
    or os.getenv("BOT_DB_PATH", "").strip()
    or "bot.db"
)
ENCRYPTION_KEY = os.getenv("ENCRYPTION_KEY", "").strip()
ORDER_CHECK_INTERVAL_SECONDS = int(os.getenv("ORDER_CHECK_INTERVAL_SECONDS", "900") or "900")
NEW_ORDER_NOTIFICATIONS = (
    os.getenv("NEW_ORDER_NOTIFICATIONS", "0").strip().lower()
    not in {"0", "false", "no", "off"}
)
LOW_STOCK_NOTIFICATIONS = (
    os.getenv("LOW_STOCK_NOTIFICATIONS", "1").strip().lower()
    not in {"0", "false", "no", "off"}
)
LOW_STOCK_CHECK_INTERVAL_SECONDS = int(
    os.getenv("LOW_STOCK_CHECK_INTERVAL_SECONDS", "1800") or "1800"
)
LOW_STOCK_THRESHOLD = int(os.getenv("LOW_STOCK_THRESHOLD", "5") or "5")
OUT_OF_STOCK_NOTIFICATIONS = (
    os.getenv("OUT_OF_STOCK_NOTIFICATIONS", "1").strip().lower()
    not in {"0", "false", "no", "off"}
)
OUT_OF_STOCK_CHECK_INTERVAL_SECONDS = int(
    os.getenv("OUT_OF_STOCK_CHECK_INTERVAL_SECONDS", "1800") or "1800"
)
SALE_NOTIFICATIONS = (
    os.getenv("SALE_NOTIFICATIONS", "1").strip().lower()
    not in {"0", "false", "no", "off"}
)
SALE_CHECK_INTERVAL_SECONDS = int(os.getenv("SALE_CHECK_INTERVAL_SECONDS", "300") or "300")
SALES_DIGEST_INTERVAL_SECONDS = max(
    900,
    int(os.getenv("SALES_DIGEST_INTERVAL_SECONDS", "3600") or "3600"),
)
INSTANT_SALE_BURST_LIMIT = max(
    1,
    min(50, int(os.getenv("INSTANT_SALE_BURST_LIMIT", "10") or "10")),
)
STOCK_CHANGE_NOTIFICATIONS = (
    os.getenv("STOCK_CHANGE_NOTIFICATIONS", "0").strip().lower()
    not in {"0", "false", "no", "off"}
)
STOCK_CHANGE_CHECK_INTERVAL_SECONDS = int(
    os.getenv("STOCK_CHANGE_CHECK_INTERVAL_SECONDS", "900") or "900"
)
LOSS_DEFECT_CHECK_INTERVAL_SECONDS = int(
    os.getenv("LOSS_DEFECT_CHECK_INTERVAL_SECONDS", "900") or "900"
)
FBO_ACCEPTANCE_CHECK_INTERVAL_SECONDS = int(
    os.getenv("FBO_ACCEPTANCE_CHECK_INTERVAL_SECONDS", "600") or "600"
)
FBO_ACCEPTANCE_INVOICE_PAGES = max(
    1,
    min(20, int(os.getenv("FBO_ACCEPTANCE_INVOICE_PAGES", "5") or "5")),
)
LOGISTICS_REMINDER_CHECK_INTERVAL_SECONDS = max(
    300,
    int(os.getenv("LOGISTICS_REMINDER_CHECK_INTERVAL_SECONDS", "1800") or "1800"),
)
LOGISTICS_REMINDER_INVOICE_PAGES = max(
    1,
    min(20, int(os.getenv("LOGISTICS_REMINDER_INVOICE_PAGES", "5") or "5")),
)
LOGISTICS_REMINDER_RETURN_PAGES = max(
    1,
    min(20, int(os.getenv("LOGISTICS_REMINDER_RETURN_PAGES", "5") or "5")),
)
UZUM_API_MIN_INTERVAL_SECONDS = max(
    0.1,
    float(os.getenv("UZUM_API_MIN_INTERVAL_SECONDS", "1.0") or "1.0"),
)
UZUM_API_429_RETRIES = max(
    0,
    min(5, int(os.getenv("UZUM_API_429_RETRIES", "3") or "3")),
)
UZUM_API_429_BASE_DELAY_SECONDS = max(
    2.0,
    float(os.getenv("UZUM_API_429_BASE_DELAY_SECONDS", "10") or "10"),
)
UZUM_API_429_MAX_DELAY_SECONDS = max(
    UZUM_API_429_BASE_DELAY_SECONDS,
    float(os.getenv("UZUM_API_429_MAX_DELAY_SECONDS", "60") or "60"),
)
WATCH_STOCK_CACHE_SECONDS = max(
    30,
    int(os.getenv("WATCH_STOCK_CACHE_SECONDS", "240") or "240"),
)
SALES_WATCH_FAST_LOOKBACK_HOURS = max(
    24,
    int(os.getenv("SALES_WATCH_FAST_LOOKBACK_HOURS", "36") or "36"),
)
SALES_WATCH_FULL_SCAN_INTERVAL_SECONDS = max(
    900,
    int(os.getenv("SALES_WATCH_FULL_SCAN_INTERVAL_SECONDS", "3600") or "3600"),
)
FBO_ACCEPTANCE_FAST_PAGES = max(
    1,
    min(
        FBO_ACCEPTANCE_INVOICE_PAGES,
        int(os.getenv("FBO_ACCEPTANCE_FAST_PAGES", "2") or "2"),
    ),
)
FBO_ACCEPTANCE_FULL_SCAN_INTERVAL_SECONDS = max(
    1800,
    int(os.getenv("FBO_ACCEPTANCE_FULL_SCAN_INTERVAL_SECONDS", "3600") or "3600"),
)

# Uzum периодически отвечает 429, если несколько фоновых задач одновременно
# листают продажи, остатки и поставки. Этот ограничитель действует на все
# экземпляры UzumClient внутри main.py: запросы стартуют с безопасным интервалом,
# а после 429 весь бот выдерживает общий cooldown и повторяет запрос.
_ORIGINAL_UZUM_REQUEST = UzumClient._request
_UZUM_API_GATE_LOCK: asyncio.Lock | None = None
_UZUM_API_GATE_LOOP: Any = None
_UZUM_API_LAST_REQUEST_AT = 0.0
_UZUM_API_COOLDOWN_UNTIL = 0.0


def _is_uzum_rate_limit_error(error: BaseException) -> bool:
    text = str(error).lower()
    return (
        "error 429" in text
        or " 429:" in text
        or "status 429" in text
        or "too many requests" in text
        or "rate limit" in text
    )


def _uzum_access_error_kind(error: BaseException) -> str | None:
    """Classify expected Uzum access failures without exposing tokens."""
    text = str(error).lower()
    if not ("403" in text or "forbidden" in text):
        return None
    if "token inactive" in text:
        return "token_inactive"
    if "shop is not available" in text:
        return "shop_unavailable"
    return "forbidden"


_WATCHER_ERROR_LOGGED_AT: dict[tuple[str, int, tuple[int, ...], str], float] = {}
WATCHER_ACCESS_ERROR_LOG_INTERVAL_SECONDS = max(
    300,
    int(os.getenv("WATCHER_ACCESS_ERROR_LOG_INTERVAL_SECONDS", "3600") or "3600"),
)


def _uzum_api_gate_lock() -> asyncio.Lock:
    global _UZUM_API_GATE_LOCK, _UZUM_API_GATE_LOOP
    loop = asyncio.get_running_loop()
    if _UZUM_API_GATE_LOCK is None or _UZUM_API_GATE_LOOP is not loop:
        _UZUM_API_GATE_LOCK = asyncio.Lock()
        _UZUM_API_GATE_LOOP = loop
    return _UZUM_API_GATE_LOCK


async def _wait_for_uzum_api_slot() -> None:
    global _UZUM_API_LAST_REQUEST_AT
    async with _uzum_api_gate_lock():
        now = time.monotonic()
        wait_for = max(
            0.0,
            _UZUM_API_COOLDOWN_UNTIL - now,
            UZUM_API_MIN_INTERVAL_SECONDS - (now - _UZUM_API_LAST_REQUEST_AT),
        )
        if wait_for > 0:
            await asyncio.sleep(wait_for)
        _UZUM_API_LAST_REQUEST_AT = time.monotonic()


def _set_uzum_api_cooldown(delay_seconds: float) -> None:
    global _UZUM_API_COOLDOWN_UNTIL
    _UZUM_API_COOLDOWN_UNTIL = max(
        _UZUM_API_COOLDOWN_UNTIL,
        time.monotonic() + max(0.0, float(delay_seconds)),
    )


async def _rate_limited_uzum_request(
    client: UzumClient,
    method: str,
    path: str,
    **kwargs: Any,
) -> Any:
    for attempt in range(UZUM_API_429_RETRIES + 1):
        await _wait_for_uzum_api_slot()
        try:
            return await _ORIGINAL_UZUM_REQUEST(client, method, path, **kwargs)
        except Exception as error:
            if not _is_uzum_rate_limit_error(error):
                raise
            delay = min(
                UZUM_API_429_MAX_DELAY_SECONDS,
                UZUM_API_429_BASE_DELAY_SECONDS * (2 ** attempt),
            )
            _set_uzum_api_cooldown(delay)
            safe_path = str(path).split("?", 1)[0][:120]
            if attempt >= UZUM_API_429_RETRIES:
                logging.warning(
                    "Uzum API 429: retries exhausted method=%s path=%s; "
                    "global cooldown %.0fs remains active",
                    method,
                    safe_path,
                    delay,
                )
                raise
            logging.warning(
                "Uzum API 429: retry %s/%s in %.0fs method=%s path=%s",
                attempt + 1,
                UZUM_API_429_RETRIES,
                delay,
                method,
                safe_path,
            )
    raise RuntimeError("Unreachable Uzum API retry state")


UzumClient._request = _rate_limited_uzum_request


def _log_watcher_api_failure(
    watcher: str,
    error: BaseException,
    *,
    shop_id: int,
    telegram_ids: list[int],
) -> None:
    if _is_uzum_rate_limit_error(error):
        logging.warning(
            "%s: Uzum API rate limit for shop=%s users=%s; "
            "snapshot kept, retry on next cycle",
            watcher,
            shop_id,
            telegram_ids,
        )
        return

    access_kind = _uzum_access_error_kind(error)
    if access_kind:
        users_key = tuple(sorted(int(value) for value in telegram_ids))
        throttle_key = (str(watcher), int(shop_id), users_key, access_kind)
        now_mono = time.monotonic()
        last_logged = float(_WATCHER_ERROR_LOGGED_AT.get(throttle_key, 0.0))
        if now_mono - last_logged < WATCHER_ACCESS_ERROR_LOG_INTERVAL_SECONDS:
            return
        _WATCHER_ERROR_LOGGED_AT[throttle_key] = now_mono

        if access_kind == "token_inactive":
            reason = "API token is inactive; user must reconnect with a new Uzum API key"
        elif access_kind == "shop_unavailable":
            reason = "shop is unavailable for this token; reconnect or select an accessible shop"
        else:
            reason = "Uzum denied access to this shop or endpoint"
        logging.warning(
            "%s: access unavailable shop=%s users=%s reason=%s; "
            "watcher state kept, retry will continue without traceback spam",
            watcher,
            shop_id,
            telegram_ids,
            reason,
        )
        return

    logging.exception(
        "%s: failed shop=%s users=%s",
        watcher,
        shop_id,
        telegram_ids,
    )
TRIAL_DAYS = int(os.getenv("TRIAL_DAYS", "3") or "3")
SUBSCRIPTION_PRICE_TEXT = os.getenv("SUBSCRIPTION_PRICE_TEXT", "300 000 сум / 1 месяц").strip()
PAYMENT_TEXT = os.getenv(
    "PAYMENT_TEXT",
    "Нажмите кнопку ниже, напишите администратору и отправьте чек. После проверки доступ будет продлён."
).strip()
PAYMENT_REQUISITES = os.getenv("PAYMENT_REQUISITES", "").strip().replace("\\n", "\n")
PAYMENT_PLANS: dict[int, dict[str, Any]] = {
    1: {"months": 1, "days": 30, "amount": 300_000, "ru": "1 месяц", "uz": "1 oy"},
    3: {"months": 3, "days": 90, "amount": 800_000, "ru": "3 месяца", "uz": "3 oy"},
    6: {"months": 6, "days": 180, "amount": 1_500_000, "ru": "6 месяцев", "uz": "6 oy"},
}
SUBSCRIPTION_PLANS_TEXT = os.getenv(
    "SUBSCRIPTION_PLANS_TEXT",
    "1 месяц — 300 000 сум\n3 месяца — 800 000 сум\n6 месяцев — 1 500 000 сум\n\nБез ограничений по количеству магазинов продавца"
).strip()
SUBSCRIPTION_PLANS_TEXT_UZ = os.getenv(
    "SUBSCRIPTION_PLANS_TEXT_UZ",
    "1 oy — 300 000 so‘m\n3 oy — 800 000 so‘m\n6 oy — 1 500 000 so‘m\n\nSotuvchining do‘konlari soni cheklanmagan"
).strip()
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "azmt_one").strip().lstrip("@")
ADMIN_CONTACT_URL = os.getenv("ADMIN_CONTACT_URL", "").strip()
VIDEO_INSTRUCTION_URL = os.getenv("VIDEO_INSTRUCTION_URL", "https://t.me/uzum_assist_bot/2").strip()

# Telegram ↔ веб-кабинет.
# WEB_SYNC_SECRET должен полностью совпадать с BOT_SYNC_SECRET на сайте.
WEB_APP_URL = os.getenv("WEB_APP_URL", "").strip().rstrip("/")
WEB_SYNC_SECRET = os.getenv("WEB_SYNC_SECRET", "").strip()
WEB_SYNC_TIMEOUT_SECONDS = max(5, min(60, int(os.getenv("WEB_SYNC_TIMEOUT_SECONDS", "25") or "25")))

# Подключение через сотрудника было экспериментом и отключено.
# Основной официальный способ: API-ключ продавца через /connect.
STAFF_UZUM_TOKEN = (
    os.getenv("STAFF_UZUM_TOKEN", "").strip()
    or os.getenv("TECHNICAL_UZUM_TOKEN", "").strip()
    or os.getenv("MASTER_UZUM_TOKEN", "").strip()
)
STAFF_PHONE = (
    os.getenv("STAFF_PHONE", "").strip()
    or os.getenv("BOT_STAFF_PHONE", "").strip()
    or os.getenv("TECHNICAL_STAFF_PHONE", "").strip()
)
STAFF_CONNECT_ENABLED = False
FINANCE_REPORT_MAX_PAGES = max(
    10, min(500, int(os.getenv("FINANCE_REPORT_MAX_PAGES", "100") or "100"))
)
SMART_LOW_STOCK_DAYS = int(os.getenv("SMART_LOW_STOCK_DAYS", "3") or "3")
TOP_PRODUCTS_DAYS = int(os.getenv("TOP_PRODUCTS_DAYS", "30") or "30")
DEAD_STOCK_DAYS = int(os.getenv("DEAD_STOCK_DAYS", "30") or "30")
LOW_MARGIN_THRESHOLD_PERCENT = float(os.getenv("LOW_MARGIN_THRESHOLD_PERCENT", "10") or "10")
DAILY_REPORTS = (
    os.getenv("DAILY_REPORTS", "0").strip().lower()
    not in {"0", "false", "no", "off"}
)
DAILY_REPORT_HOUR_UZT = int(os.getenv("DAILY_REPORT_HOUR_UZT", "9") or "9")
SUBSCRIPTION_REMINDERS = (
    os.getenv("SUBSCRIPTION_REMINDERS", "1").strip().lower()
    not in {"0", "false", "no", "off"}
)
SUBSCRIPTION_REMINDER_DAYS = parse_reminder_days(
    os.getenv("SUBSCRIPTION_REMINDER_DAYS", "7,3,1")
)
SUBSCRIPTION_EXPIRED_QUEUE_DAYS = max(
    1, int(os.getenv("SUBSCRIPTION_EXPIRED_QUEUE_DAYS", "7") or "7")
)
SUBSCRIPTION_ADMIN_DIGEST_HOUR_UZT = max(
    0, min(23, int(os.getenv("SUBSCRIPTION_ADMIN_DIGEST_HOUR_UZT", "9") or "9"))
)

# Печать SKU-этикеток через официальный Uzum Seller OpenAPI.
# Swagger разрешает до 100 SKU и до 100 этикеток на один SKU.
BARCODE_MAX_SKUS = min(100, max(1, int(os.getenv("BARCODE_MAX_SKUS", "100") or "100")))
BARCODE_MAX_AMOUNT_PER_SKU = 100
# Дополнительный безопасный лимит, чтобы PDF не получился слишком большим для Telegram.
BARCODE_MAX_TOTAL_LABELS = min(5000, max(1, int(os.getenv("BARCODE_MAX_TOTAL_LABELS", "1000") or "1000")))
COST_IMPORT_MAX_FILE_BYTES = max(
    1_048_576,
    min(
        20 * 1_048_576,
        int(
            os.getenv("COST_IMPORT_MAX_FILE_BYTES", str(5 * 1_048_576))
            or str(5 * 1_048_576)
        ),
    ),
)
COST_IMPORT_MAX_UNCOMPRESSED_BYTES = max(
    COST_IMPORT_MAX_FILE_BYTES,
    min(
        100 * 1_048_576,
        int(
            os.getenv(
                "COST_IMPORT_MAX_UNCOMPRESSED_BYTES",
                str(50 * 1_048_576),
            )
            or str(50 * 1_048_576)
        ),
    ),
)
COST_IMPORT_MAX_ROWS = max(
    100,
    min(50_000, int(os.getenv("COST_IMPORT_MAX_ROWS", "10000") or "10000")),
)
MAX_CONCURRENT_USER_HANDLERS = max(
    5,
    min(
        500,
        int(os.getenv("MAX_CONCURRENT_USER_HANDLERS", "40") or "40"),
    ),
)

# Администраторы. Владельцы имеют полный доступ, менеджеры — без выгрузки базы.
BUILTIN_OWNER_IDS: set[int] = {
    445354240,
    938965878,
}
BUILTIN_MANAGER_IDS: set[int] = {
    8046815224,
}


def _parse_telegram_ids(*keys: str) -> set[int]:
    """Read Telegram IDs from BotHost variables despite spaces/quotes/newlines."""
    import re

    ids: set[int] = set()
    for key in keys:
        raw = str(os.getenv(key, "") or "")
        for value in re.findall(r"(?<!\d)\d{5,20}(?!\d)", raw):
            try:
                ids.add(int(value))
            except (TypeError, ValueError):
                continue
    return ids


OWNER_IDS = BUILTIN_OWNER_IDS | _parse_telegram_ids(
    "OWNER_IDS",
    "OWNER_TELEGRAM_ID",
    "OWNER_ID",
)
ADMIN_IDS = (
    OWNER_IDS
    | BUILTIN_MANAGER_IDS
    | _parse_telegram_ids("ADMIN_IDS", "MANAGER_IDS")
)
MANAGER_IDS = ADMIN_IDS - OWNER_IDS
logging.info(
    "ADMIN_ROLES_LOADED owners=%s managers=%s",
    sorted(OWNER_IDS),
    sorted(MANAGER_IDS),
)

if not TELEGRAM_BOT_TOKEN:
    raise RuntimeError(
        "TELEGRAM_BOT_TOKEN is empty. Set it in BotHost environment variables."
    )

# База и шифрование Uzum API-токена
db = Database(DB_PATH)
SQLITE_BUSY_TIMEOUT_MS = max(
    1000, min(60000, int(os.getenv("SQLITE_BUSY_TIMEOUT_MS", "10000") or "10000"))
)
SQLITE_WAL = (
    os.getenv("SQLITE_WAL", "1").strip().lower()
    not in {"0", "false", "no", "off"}
)
_ORIGINAL_DB_CONNECT = db.connect


def _db_connect_release_safe():
    conn = _ORIGINAL_DB_CONNECT()
    try:
        conn.execute(f"PRAGMA busy_timeout={SQLITE_BUSY_TIMEOUT_MS}")
    except Exception:
        pass
    return conn


db.connect = _db_connect_release_safe  # type: ignore[method-assign]
if SQLITE_WAL:
    try:
        with db.connect() as _conn:
            _conn.execute("PRAGMA journal_mode=WAL")
            _conn.execute("PRAGMA synchronous=NORMAL")
    except Exception:
        logging.exception("SQLite WAL could not be enabled; continuing with default mode")
cipher = TokenCipher(ENCRYPTION_KEY)

bot = Bot(
    TELEGRAM_BOT_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML),
)
dp = Dispatcher()

# --- Подписки / trial ---
def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _dt_to_db(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.astimezone(timezone.utc).isoformat(timespec="seconds")


def _dt_from_db(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        text = str(value).replace("Z", "+00:00")
        dt = datetime.fromisoformat(text)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


def _fmt_dt(value: Any) -> str:
    dt = value if isinstance(value, datetime) else _dt_from_db(value)
    if not dt:
        return "—"
    return dt.astimezone(timezone(timedelta(hours=5))).strftime("%d.%m.%Y %H:%M")


def is_admin(telegram_id: int | None) -> bool:
    if telegram_id is None:
        return False
    return int(telegram_id) in ADMIN_IDS


def is_owner(telegram_id: int | None) -> bool:
    if telegram_id is None:
        return False
    return int(telegram_id) in OWNER_IDS


def init_subscription_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS subscriptions (
                telegram_id INTEGER PRIMARY KEY,
                trial_started_at TEXT,
                trial_until TEXT,
                subscription_until TEXT,
                blocked INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def get_subscription_row(telegram_id: int) -> dict[str, Any] | None:
    with db.connect() as conn:
        row = conn.execute("SELECT * FROM subscriptions WHERE telegram_id = ?", (int(telegram_id),)).fetchone()
    return dict(row) if row else None


def ensure_subscription(telegram_id: int) -> dict[str, Any]:
    row = get_subscription_row(telegram_id)
    if row:
        return row
    now = _utc_now()
    trial_until = now + timedelta(days=TRIAL_DAYS)
    with db.connect() as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO subscriptions
            (telegram_id, trial_started_at, trial_until, subscription_until, blocked, created_at, updated_at)
            VALUES (?, ?, ?, NULL, 0, ?, ?)
            """,
            (int(telegram_id), _dt_to_db(now), _dt_to_db(trial_until), _dt_to_db(now), _dt_to_db(now)),
        )
        conn.commit()
    return get_subscription_row(telegram_id) or {}


def subscription_active_until(row: dict[str, Any] | None) -> datetime | None:
    if not row:
        return None
    dates = [_dt_from_db(row.get("trial_until")), _dt_from_db(row.get("subscription_until"))]
    dates = [d for d in dates if d is not None]
    return max(dates) if dates else None


def has_active_subscription(telegram_id: int) -> bool:
    return subscription_access_level(telegram_id) in {"admin", "paid", "trial"}


# Trial is intentionally useful but limited: it demonstrates the two core
# outcomes of Seller Pro without providing the full paid product.
TRIAL_ALLOWED_FEATURES = frozenset(
    {
        "sales_notifications",
        "sales_today",
        "morning_report",
    }
)


def subscription_access_level(telegram_id: int) -> str:
    """Return one stable access level for UI, handlers and background jobs."""
    telegram_id = int(telegram_id)
    if is_admin(telegram_id):
        return "admin"

    row = ensure_subscription(telegram_id)
    if int(row.get("blocked") or 0) == 1:
        return "blocked"

    now = _utc_now()
    paid_until = _dt_from_db(row.get("subscription_until"))
    if paid_until and paid_until > now:
        return "paid"

    trial_until = _dt_from_db(row.get("trial_until"))
    if trial_until and trial_until > now:
        return "trial"

    return "expired"


def has_paid_subscription(telegram_id: int) -> bool:
    return subscription_access_level(telegram_id) in {"admin", "paid"}


def feature_access_allowed(telegram_id: int, feature: str) -> bool:
    level = subscription_access_level(telegram_id)
    if level in {"admin", "paid"}:
        return True
    return level == "trial" and str(feature) in TRIAL_ALLOWED_FEATURES


def subscription_status_text(telegram_id: int) -> str:
    row = ensure_subscription(telegram_id)
    if is_admin(telegram_id):
        return "👑 Админ-доступ: без ограничений"
    if int(row.get("blocked") or 0) == 1:
        return "⛔ Пользователь заблокирован"
    now = _utc_now()
    trial_until = _dt_from_db(row.get("trial_until"))
    paid_until = _dt_from_db(row.get("subscription_until"))
    until = subscription_active_until(row)
    if until and until > now:
        if paid_until and paid_until == until:
            return f"✅ Подписка активна до {_fmt_dt(paid_until)}"
        return f"🎁 Trial активен до {_fmt_dt(trial_until)}"
    return "⛔ Подписка закончилась"


def subscription_full_text(telegram_id: int) -> str:
    row = ensure_subscription(telegram_id)
    status = subscription_status_text(telegram_id)

    if is_admin(telegram_id):
        return (
            "💎 <b>Моя подписка</b>\n\n"
            f"Telegram ID: <code>{telegram_id}</code>\n"
            f"Статус: {status}\n\n"
            "Trial и дата оплаты для администратора не важны — доступ всегда открыт.\n\n"
            "Команды администратора:\n"
            "• <code>/users</code> — пользователи\n"
            "• <code>/renewals</code> — очередь продлений и черновики\n"
            "• <code>/extend ID 30</code> — продлить доступ\n"
            "• <code>/block ID</code> — заблокировать\n"
            "• <code>/unblock ID</code> — разблокировать\n"
            "• <code>/paid ID сумма дни</code> — записать оплату\n"
            "• <code>/payments</code> — история оплат\n"
            "• <code>/backup_db</code> — скачать базу"
        )

    return (
        "💎 <b>Моя подписка</b>\n\n"
        f"Telegram ID: <code>{telegram_id}</code>\n"
        f"Статус: {status}\n"
        f"Trial до: <b>{_fmt_dt(row.get('trial_until'))}</b>\n"
        f"Оплачено до: <b>{_fmt_dt(row.get('subscription_until'))}</b>\n\n"
        "Тарифы:\n"
        f"<b>{escape(SUBSCRIPTION_PLANS_TEXT)}</b>\n\n"
        f"{escape(PAYMENT_TEXT)}\n\n"
        "История оплат: <code>/my_payments</code>\n"
        "Поддержка: <code>/support</code>\n"
        "Заменить API-ключ: <code>/reconnect</code>\n"
        "Удалить API-ключ: <code>/disconnect</code>"
    )


async def require_active_subscription(message: Message, telegram_id: int | None = None) -> bool:
    if telegram_id is None:
        telegram_id = upsert_from_message(message)
    ensure_subscription(int(telegram_id))
    if has_active_subscription(int(telegram_id)):
        return True
    await message.answer(
        tr_user(int(telegram_id), "access_limited"),
        reply_markup=menu_for_message(message),
    )
    return False


async def require_premium_subscription(
    message: Message,
    telegram_id: int | None = None,
) -> bool:
    """Allow only a paid subscription or admin access.

    Trial users receive the plan selector, while expired/blocked users retain
    the existing access-ended screen.  The helper is also used for entry points
    such as the web cabinet that must never be opened by a saved old button.
    """
    if telegram_id is None:
        telegram_id = upsert_from_message(message)
    telegram_id = int(telegram_id)
    level = subscription_access_level(telegram_id)
    if level in {"admin", "paid"}:
        return True
    if level == "trial":
        await send_trial_premium_locked(message, telegram_id)
        return False
    await message.answer(
        tr_user(telegram_id, "access_limited"),
        reply_markup=menu_for_message(message),
    )
    return False


def admin_only(telegram_id: int) -> bool:
    return is_admin(int(telegram_id))


def owner_only(telegram_id: int) -> bool:
    return is_owner(int(telegram_id))


def admin_contact_link() -> str | None:
    if ADMIN_CONTACT_URL:
        return ADMIN_CONTACT_URL
    if ADMIN_USERNAME:
        return f"https://t.me/{ADMIN_USERNAME}"
    return None


def admin_contact_markup() -> InlineKeyboardMarkup | None:
    url = admin_contact_link()
    if not url:
        return None
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="✍️ Написать администратору", url=url)]]
    )


def video_instruction_markup(lang: str = "ru") -> InlineKeyboardMarkup | None:
    if not VIDEO_INSTRUCTION_URL:
        return None
    button_text = "▶️ Videoni ko‘rish" if normalize_lang(lang) == "uz" else "▶️ Смотреть видео"
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text=button_text, url=VIDEO_INSTRUCTION_URL)]]
    )


def help_links_markup(lang: str = "ru") -> InlineKeyboardMarkup | None:
    rows = []
    if VIDEO_INSTRUCTION_URL:
        rows.append([InlineKeyboardButton(
            text=("🎥 API ulash videosi" if normalize_lang(lang) == "uz" else "🎥 Видео подключения API"),
            url=VIDEO_INSTRUCTION_URL,
        )])
    admin_url = admin_contact_link()
    if admin_url:
        rows.append([InlineKeyboardButton(
            text=("✍️ Administratorga yozish" if normalize_lang(lang) == "uz" else "✍️ Написать администратору"),
            url=admin_url,
        )])
    if not rows:
        return None
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_contact_text() -> str:
    if ADMIN_USERNAME:
        return f"@{escape(ADMIN_USERNAME)}"
    if ADMIN_CONTACT_URL:
        return escape(ADMIN_CONTACT_URL)
    return "администратору"


def extend_subscription_days(telegram_id: int, days: int) -> datetime:
    ensure_subscription(telegram_id)
    row = get_subscription_row(telegram_id) or {}
    now = _utc_now()
    candidates = [now, _dt_from_db(row.get("subscription_until")), _dt_from_db(row.get("trial_until"))]
    base = max([d for d in candidates if d is not None])
    new_until = base + timedelta(days=max(1, int(days)))
    with db.connect() as conn:
        conn.execute(
            "UPDATE subscriptions SET subscription_until = ?, blocked = 0, updated_at = ? WHERE telegram_id = ?",
            (_dt_to_db(new_until), _dt_to_db(now), int(telegram_id)),
        )
        conn.commit()
    return new_until


def set_trial_days(telegram_id: int, days: int) -> datetime:
    ensure_subscription(telegram_id)
    row = get_subscription_row(telegram_id) or {}
    now = _utc_now()
    current_until = _dt_from_db(row.get("trial_until"))
    base = max(now, current_until) if current_until is not None else now
    new_until = base + timedelta(days=max(1, int(days)))
    with db.connect() as conn:
        conn.execute(
            "UPDATE subscriptions SET trial_until = ?, blocked = 0, updated_at = ? WHERE telegram_id = ?",
            (_dt_to_db(new_until), _dt_to_db(now), int(telegram_id)),
        )
        conn.commit()
    return new_until


def set_blocked(telegram_id: int, blocked: bool) -> None:
    ensure_subscription(telegram_id)
    now = _utc_now()
    with db.connect() as conn:
        conn.execute(
            "UPDATE subscriptions SET blocked = ?, updated_at = ? WHERE telegram_id = ?",
            (1 if blocked else 0, _dt_to_db(now), int(telegram_id)),
        )
        conn.commit()




def init_business_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS payment_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER NOT NULL,
                amount INTEGER NOT NULL DEFAULT 0,
                days INTEGER NOT NULL DEFAULT 0,
                admin_id INTEGER,
                comment TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def init_subscription_automation_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS subscription_reminder_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER NOT NULL,
                active_until TEXT NOT NULL,
                subscription_kind TEXT NOT NULL,
                milestone TEXT NOT NULL,
                days_remaining INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'pending',
                draft_text TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                reviewed_by INTEGER,
                reviewed_at TEXT,
                sent_at TEXT,
                last_error TEXT,
                UNIQUE (telegram_id, active_until, milestone)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_subscription_reminder_queue_status
            ON subscription_reminder_queue (status, active_until)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS subscription_automation_state (
                state_key TEXT PRIMARY KEY,
                state_value TEXT,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def init_payment_request_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS payment_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER NOT NULL,
                plan_months INTEGER NOT NULL,
                plan_days INTEGER NOT NULL,
                amount INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'awaiting_receipt',
                receipt_file_id TEXT,
                receipt_file_unique_id TEXT,
                receipt_type TEXT,
                reviewed_by INTEGER,
                reviewed_at TEXT,
                payment_history_id INTEGER,
                rejection_reason TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_payment_requests_status
            ON payment_requests (status, created_at)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_payment_requests_user
            ON payment_requests (telegram_id, created_at)
            """
        )
        conn.commit()


def record_payment(telegram_id: int, amount: int, days: int, admin_id: int | None = None, comment: str = "") -> int:
    init_business_tables()
    with db.connect() as conn:
        cur = conn.execute(
            """
            INSERT INTO payment_history (telegram_id, amount, days, admin_id, comment, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (int(telegram_id), int(amount), int(days), int(admin_id) if admin_id else None, comment.strip(), _dt_to_db(_utc_now())),
        )
        conn.commit()
        return int(cur.lastrowid)


def list_payments(telegram_id: int | None = None, limit: int = 20) -> list[dict[str, Any]]:
    init_business_tables()
    with db.connect() as conn:
        if telegram_id is None:
            rows = conn.execute(
                """
                SELECT id, telegram_id, amount, days, admin_id, comment, created_at
                FROM payment_history
                ORDER BY id DESC
                LIMIT ?
                """,
                (int(limit),),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT id, telegram_id, amount, days, admin_id, comment, created_at
                FROM payment_history
                WHERE telegram_id = ?
                ORDER BY id DESC
                LIMIT ?
                """,
                (int(telegram_id), int(limit)),
            ).fetchall()
    return [dict(row) for row in rows]


def payment_line(row: dict[str, Any]) -> str:
    comment = (row.get("comment") or "").strip()
    comment_part = f" | {escape(comment)}" if comment else ""
    amount_text = f"{int(row.get('amount') or 0):,}".replace(",", " ")
    return (
        f"#{row.get('id')} | <code>{int(row.get('telegram_id') or 0)}</code> | "
        f"{amount_text} сум | {int(row.get('days') or 0)} дней | {_fmt_dt(row.get('created_at'))}{comment_part}"
    )


def _payment_amount_text(amount: int) -> str:
    return f"{int(amount):,}".replace(",", " ")


def payment_plan(months: int) -> dict[str, Any] | None:
    plan = PAYMENT_PLANS.get(int(months))
    return dict(plan) if plan else None


def create_payment_request(telegram_id: int, months: int) -> tuple[dict[str, Any], bool]:
    init_payment_request_tables()
    plan = payment_plan(months)
    if not plan:
        raise ValueError("Unknown payment plan")
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        pending = conn.execute(
            """
            SELECT * FROM payment_requests
            WHERE telegram_id = ? AND status = 'pending_review'
            ORDER BY id DESC LIMIT 1
            """,
            (int(telegram_id),),
        ).fetchone()
        if pending:
            return dict(pending), False
        conn.execute(
            """
            UPDATE payment_requests
            SET status = 'superseded', updated_at = ?
            WHERE telegram_id = ? AND status = 'awaiting_receipt'
            """,
            (now_text, int(telegram_id)),
        )
        cursor = conn.execute(
            """
            INSERT INTO payment_requests
            (telegram_id, plan_months, plan_days, amount, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, 'awaiting_receipt', ?, ?)
            """,
            (
                int(telegram_id),
                int(plan["months"]),
                int(plan["days"]),
                int(plan["amount"]),
                now_text,
                now_text,
            ),
        )
        request_id = int(cursor.lastrowid)
        conn.commit()
        row = conn.execute("SELECT * FROM payment_requests WHERE id = ?", (request_id,)).fetchone()
    return dict(row), True


def get_payment_request(request_id: int) -> dict[str, Any] | None:
    init_payment_request_tables()
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT p.*, u.username, u.first_name
            FROM payment_requests p
            LEFT JOIN users u ON u.telegram_id = p.telegram_id
            WHERE p.id = ?
            """,
            (int(request_id),),
        ).fetchone()
    return dict(row) if row else None


def latest_awaiting_payment_request(telegram_id: int) -> dict[str, Any] | None:
    init_payment_request_tables()
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT * FROM payment_requests
            WHERE telegram_id = ? AND status = 'awaiting_receipt'
            ORDER BY id DESC LIMIT 1
            """,
            (int(telegram_id),),
        ).fetchone()
    return dict(row) if row else None


def latest_payment_request_by_status(telegram_id: int, status: str) -> dict[str, Any] | None:
    init_payment_request_tables()
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT * FROM payment_requests
            WHERE telegram_id = ? AND status = ?
            ORDER BY id DESC LIMIT 1
            """,
            (int(telegram_id), str(status)),
        ).fetchone()
    return dict(row) if row else None


def list_payment_requests(status: str = "pending_review", limit: int = 50) -> list[dict[str, Any]]:
    init_payment_request_tables()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT p.*, u.username, u.first_name
            FROM payment_requests p
            LEFT JOIN users u ON u.telegram_id = p.telegram_id
            WHERE p.status = ?
            ORDER BY p.id DESC
            LIMIT ?
            """,
            (str(status), max(1, int(limit))),
        ).fetchall()
    return [dict(row) for row in rows]


def attach_payment_receipt(
    request_id: int,
    telegram_id: int,
    *,
    file_id: str,
    file_unique_id: str,
    receipt_type: str,
) -> tuple[dict[str, Any] | None, str | None]:
    init_payment_request_tables()
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        duplicate = conn.execute(
            """
            SELECT id FROM payment_requests
            WHERE receipt_file_unique_id = ?
              AND id <> ?
              AND status IN ('pending_review', 'approved')
            LIMIT 1
            """,
            (str(file_unique_id), int(request_id)),
        ).fetchone()
        if duplicate:
            return None, "duplicate"
        cursor = conn.execute(
            """
            UPDATE payment_requests
            SET receipt_file_id = ?, receipt_file_unique_id = ?, receipt_type = ?,
                status = 'pending_review', updated_at = ?
            WHERE id = ? AND telegram_id = ? AND status = 'awaiting_receipt'
            """,
            (
                str(file_id),
                str(file_unique_id),
                str(receipt_type),
                now_text,
                int(request_id),
                int(telegram_id),
            ),
        )
        conn.commit()
    if not cursor.rowcount:
        return None, "not_available"
    return get_payment_request(request_id), None


def cancel_payment_request(request_id: int, telegram_id: int) -> bool:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        cursor = conn.execute(
            """
            UPDATE payment_requests
            SET status = 'cancelled', updated_at = ?
            WHERE id = ? AND telegram_id = ? AND status = 'awaiting_receipt'
            """,
            (now_text, int(request_id), int(telegram_id)),
        )
        conn.commit()
    return bool(cursor.rowcount)


def approve_payment_request(request_id: int, admin_id: int) -> dict[str, Any] | None:
    request = get_payment_request(request_id)
    if not request or request.get("status") != "pending_review":
        return None
    telegram_id = int(request["telegram_id"])
    ensure_subscription(telegram_id)
    now = _utc_now()
    now_text = _dt_to_db(now) or ""

    with db.connect() as conn:
        conn.execute("BEGIN IMMEDIATE")
        current = conn.execute(
            "SELECT * FROM payment_requests WHERE id = ?",
            (int(request_id),),
        ).fetchone()
        if not current or current["status"] != "pending_review":
            conn.rollback()
            return None

        subscription = conn.execute(
            "SELECT * FROM subscriptions WHERE telegram_id = ?",
            (telegram_id,),
        ).fetchone()
        subscription_row = dict(subscription) if subscription else {}
        candidates = [
            now,
            _dt_from_db(subscription_row.get("subscription_until")),
            _dt_from_db(subscription_row.get("trial_until")),
        ]
        base = max(value for value in candidates if value is not None)
        new_until = base + timedelta(days=int(current["plan_days"]))
        new_until_text = _dt_to_db(new_until) or ""

        conn.execute(
            """
            UPDATE subscriptions
            SET subscription_until = ?, blocked = 0, updated_at = ?
            WHERE telegram_id = ?
            """,
            (new_until_text, now_text, telegram_id),
        )
        payment_cursor = conn.execute(
            """
            INSERT INTO payment_history (telegram_id, amount, days, admin_id, comment, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                telegram_id,
                int(current["amount"]),
                int(current["plan_days"]),
                int(admin_id),
                f"чек #{int(request_id)}, {int(current['plan_months'])} мес.",
                now_text,
            ),
        )
        payment_id = int(payment_cursor.lastrowid)
        conn.execute(
            """
            UPDATE payment_requests
            SET status = 'approved', reviewed_by = ?, reviewed_at = ?,
                payment_history_id = ?, updated_at = ?
            WHERE id = ? AND status = 'pending_review'
            """,
            (int(admin_id), now_text, payment_id, now_text, int(request_id)),
        )
        conn.execute(
            """
            UPDATE subscription_reminder_queue
            SET status = 'superseded', updated_at = ?
            WHERE telegram_id = ? AND status = 'pending'
            """,
            (now_text, telegram_id),
        )
        conn.commit()

    result = dict(current)
    result["new_until"] = new_until_text
    result["payment_history_id"] = payment_id
    return result


def reject_payment_request(request_id: int, admin_id: int, reason: str) -> dict[str, Any] | None:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        cursor = conn.execute(
            """
            UPDATE payment_requests
            SET status = 'rejected', reviewed_by = ?, reviewed_at = ?,
                rejection_reason = ?, updated_at = ?
            WHERE id = ? AND status = 'pending_review'
            """,
            (int(admin_id), now_text, str(reason)[:500], now_text, int(request_id)),
        )
        conn.commit()
    return get_payment_request(request_id) if cursor.rowcount else None


def payment_request_user_label(row: dict[str, Any]) -> str:
    username = str(row.get("username") or "").strip()
    first_name = str(row.get("first_name") or "").strip()
    return f"@{username}" if username else (first_name or "без имени")


def payment_request_caption(row: dict[str, Any]) -> str:
    telegram_id = int(row.get("telegram_id") or 0)
    return (
        "🧾 <b>Новый чек на проверку</b>\n\n"
        f"Заявка: <b>#{int(row.get('id') or 0)}</b>\n"
        f"Клиент: <a href=\"tg://user?id={telegram_id}\">{escape(payment_request_user_label(row))}</a>\n"
        f"Telegram ID: <code>{telegram_id}</code>\n"
        f"Тариф: <b>{int(row.get('plan_months') or 0)} мес.</b>\n"
        f"Сумма: <b>{_payment_amount_text(int(row.get('amount') or 0))} сум</b>\n"
        f"Отправлен: <b>{_fmt_dt(row.get('updated_at'))}</b>\n\n"
        "Проверьте сумму и получателя на чеке. Доступ изменится только после подтверждения."
    )



# --- Юнит-экономика / себестоимость SKU ---
def init_unit_economy_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS unit_costs (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                sku_key TEXT NOT NULL,
                title TEXT,
                cost REAL NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, sku_key)
            )
            """
        )
        # The legacy ``unit_costs`` table is intentionally preserved so an
        # upgrade never destroys seller data.  New profit calculations read
        # exclusively from this Uzum-sourced cache.
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS uzum_sku_financials (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                sku_key TEXT NOT NULL,
                aliases_json TEXT NOT NULL DEFAULT '[]',
                sku_id TEXT,
                barcode TEXT,
                seller_item_code TEXT,
                sku_title TEXT,
                product_title TEXT,
                purchase_price REAL,
                ikpu TEXT,
                paid_storage_price_item REAL,
                paid_storage_amount REAL,
                paid_storage INTEGER NOT NULL DEFAULT 0,
                fetched_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, sku_key)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_uzum_sku_financials_cost
            ON uzum_sku_financials (telegram_id, shop_id, purchase_price)
            """
        )
        conn.commit()


def _unit_sku_key(value: Any) -> str:
    return str(value or "").strip().lower()


def save_unit_cost(telegram_id: int, shop_id: int, sku: str, cost: float, title: str = "") -> None:
    init_unit_economy_tables()
    sku_key = _unit_sku_key(sku)
    if not sku_key:
        return
    with db.connect() as conn:
        conn.execute(
            """
            INSERT INTO unit_costs (telegram_id, shop_id, sku_key, title, cost, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id, sku_key) DO UPDATE SET
                title = excluded.title,
                cost = excluded.cost,
                updated_at = excluded.updated_at
            """,
            (int(telegram_id), int(shop_id), sku_key, str(title or "").strip(), float(cost), _dt_to_db(_utc_now()) or ""),
        )
        conn.commit()


def delete_unit_cost(telegram_id: int, shop_id: int, sku: str) -> bool:
    init_unit_economy_tables()
    sku_key = _unit_sku_key(sku)
    with db.connect() as conn:
        cur = conn.execute(
            "DELETE FROM unit_costs WHERE telegram_id = ? AND shop_id = ? AND sku_key = ?",
            (int(telegram_id), int(shop_id), sku_key),
        )
        conn.commit()
        return bool(cur.rowcount)


def get_unit_cost_map(telegram_id: int, shop_id: int) -> dict[str, dict[str, Any]]:
    """Return only costs fetched from Uzum; legacy manual rows are ignored."""
    init_unit_economy_tables()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT sku_key, aliases_json, sku_id, barcode, seller_item_code,
                   sku_title, product_title, purchase_price, ikpu,
                   paid_storage_price_item, paid_storage_amount, paid_storage,
                   fetched_at
            FROM uzum_sku_financials
            WHERE telegram_id = ? AND shop_id = ? AND purchase_price > 0
            """,
            (int(telegram_id), int(shop_id)),
        ).fetchall()
    result: dict[str, dict[str, Any]] = {}
    for raw_row in rows:
        row = dict(raw_row)
        entry = {
            **row,
            "cost": float(row.get("purchase_price") or 0),
            "title": str(row.get("product_title") or row.get("sku_title") or ""),
            "source": "uzum",
            "updated_at": row.get("fetched_at"),
        }
        aliases: list[str] = []
        try:
            decoded = json.loads(str(row.get("aliases_json") or "[]"))
            if isinstance(decoded, list):
                aliases.extend(str(value) for value in decoded)
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
        aliases.extend(
            str(row.get(field) or "")
            for field in ("sku_key", "sku_id", "barcode", "seller_item_code", "sku_title")
        )
        for alias in aliases:
            key = _unit_sku_key(alias)
            if key and key not in {"-", "—"}:
                result[key] = entry
    return result


def list_uzum_sku_financials(
    telegram_id: int,
    shop_id: int,
) -> list[dict[str, Any]]:
    init_unit_economy_tables()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT * FROM uzum_sku_financials
            WHERE telegram_id = ? AND shop_id = ?
            ORDER BY product_title COLLATE NOCASE, sku_title COLLATE NOCASE
            """,
            (int(telegram_id), int(shop_id)),
        ).fetchall()
    return [dict(row) for row in rows]


def list_unit_costs(telegram_id: int, shop_id: int, limit: int = 50) -> list[dict[str, Any]]:
    init_unit_economy_tables()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT sku_key, title, cost, updated_at
            FROM unit_costs
            WHERE telegram_id = ? AND shop_id = ?
            ORDER BY updated_at DESC
            LIMIT ?
            """,
            (int(telegram_id), int(shop_id), int(limit)),
        ).fetchall()
    return [dict(r) for r in rows]


def _parse_cost_command_args(text: str) -> tuple[str, float] | None:
    import re
    raw = parse_args(text or "").strip()
    # Формат: /cost SKU 60000 или /cost SKU 60 000
    m = re.match(r"^(.+?)\s+([0-9][0-9\s.,]*)$", raw)
    if not m:
        return None
    sku = m.group(1).strip()
    money_raw = m.group(2)
    digits = re.sub(r"[^0-9]", "", money_raw)
    if not sku or not digits:
        return None
    return sku, float(digits)

def list_subscription_users(limit: int = 30) -> list[dict[str, Any]]:
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT u.telegram_id, u.username, u.first_name, u.default_shop_id,
                   s.trial_until, s.subscription_until, s.blocked
            FROM users u
            LEFT JOIN subscriptions s ON s.telegram_id = u.telegram_id
            ORDER BY u.updated_at DESC
            LIMIT ?
            """,
            (int(limit),),
        ).fetchall()
    return [dict(r) for r in rows]


def subscription_compact_line(row: dict[str, Any]) -> str:
    telegram_id = int(row.get("telegram_id"))
    username = row.get("username") or ""
    name = row.get("first_name") or ""
    label = f"@{username}" if username else (name or "без имени")
    if row.get("blocked"):
        status_label = "⛔ block"
    else:
        untils = [_dt_from_db(row.get("subscription_until")), _dt_from_db(row.get("trial_until"))]
        untils = [d for d in untils if d]
        until = max(untils) if untils else None
        status_label = "✅" if until and until > _utc_now() else "❌"
    until_value = row.get("subscription_until") or row.get("trial_until")
    return f"{status_label} <code>{telegram_id}</code> — {escape(str(label))} | до: {_fmt_dt(until_value)}"


def _subscription_until_for_row(row: dict[str, Any]) -> datetime | None:
    return subscription_active_until(row)


def get_admin_stats() -> dict[str, int]:
    now = _utc_now()
    with db.connect() as conn:
        total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        connected = conn.execute("SELECT COUNT(*) FROM users WHERE uzum_token_encrypted IS NOT NULL").fetchone()[0]
        rows = conn.execute("SELECT telegram_id, trial_until, subscription_until, blocked FROM subscriptions").fetchall()
        payments_today = conn.execute(
            "SELECT COALESCE(SUM(amount), 0) FROM payment_history WHERE created_at >= ?",
            (_dt_to_db(now.replace(hour=0, minute=0, second=0, microsecond=0)),),
        ).fetchone()[0]
        payments_30 = conn.execute(
            "SELECT COALESCE(SUM(amount), 0) FROM payment_history WHERE created_at >= ?",
            (_dt_to_db(now - timedelta(days=30)),),
        ).fetchone()[0]
        pending_receipts = conn.execute(
            "SELECT COUNT(*) FROM payment_requests WHERE status = 'pending_review'"
        ).fetchone()[0]
    active = paid = trial = expired = blocked = 0
    for r in rows:
        row = dict(r)
        if int(row.get("blocked") or 0) == 1:
            blocked += 1
            continue
        until = subscription_active_until(row)
        if until and until > now:
            active += 1
            paid_until = _dt_from_db(row.get("subscription_until"))
            if paid_until and paid_until == until:
                paid += 1
            else:
                trial += 1
        else:
            expired += 1
    return {
        "total_users": int(total_users or 0),
        "connected": int(connected or 0),
        "active": active,
        "paid": paid,
        "trial": trial,
        "expired": expired,
        "blocked": blocked,
        "payments_today": int(payments_today or 0),
        "payments_30": int(payments_30 or 0),
        "pending_receipts": int(pending_receipts or 0),
    }


def list_expiring_users(days: int = 3, limit: int = 50) -> list[dict[str, Any]]:
    now = _utc_now()
    until_limit = now + timedelta(days=max(1, int(days)))
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT u.telegram_id, u.username, u.first_name, u.default_shop_id,
                   s.trial_until, s.subscription_until, s.blocked
            FROM subscriptions s
            LEFT JOIN users u ON u.telegram_id = s.telegram_id
            WHERE s.blocked = 0
            ORDER BY COALESCE(s.subscription_until, s.trial_until) ASC
            LIMIT ?
            """,
            (int(limit),),
        ).fetchall()
    result: list[dict[str, Any]] = []
    for r in rows:
        row = dict(r)
        if is_admin(int(row.get("telegram_id") or 0)):
            continue
        until = subscription_active_until(row)
        if until and now < until <= until_limit:
            result.append(row)
    return result


def _subscription_kind_and_until(row: dict[str, Any]) -> tuple[str, datetime | None]:
    trial_until = _dt_from_db(row.get("trial_until"))
    paid_until = _dt_from_db(row.get("subscription_until"))
    dates = [value for value in (trial_until, paid_until) if value is not None]
    if not dates:
        return "trial", None
    active_until = max(dates)
    kind = "paid" if paid_until and paid_until == active_until else "trial"
    return kind, active_until


def refresh_subscription_reminder_queue(now: datetime | None = None) -> dict[str, int]:
    """Create reviewable drafts without sending or changing subscriptions."""
    init_subscription_automation_tables()
    now = now or _utc_now()
    now_text = _dt_to_db(now) or ""
    created = 0

    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT s.telegram_id, s.trial_until, s.subscription_until, s.blocked
            FROM subscriptions s
            ORDER BY s.telegram_id
            """
        ).fetchall()

    for raw_row in rows:
        row = dict(raw_row)
        telegram_id = int(row.get("telegram_id") or 0)
        kind, active_until = _subscription_kind_and_until(row)

        if not telegram_id or is_admin(telegram_id) or int(row.get("blocked") or 0) == 1 or not active_until:
            with db.connect() as conn:
                conn.execute(
                    """
                    UPDATE subscription_reminder_queue
                    SET status = 'superseded', updated_at = ?
                    WHERE telegram_id = ? AND status = 'pending'
                    """,
                    (now_text, telegram_id),
                )
                conn.commit()
            continue

        active_until_text = _dt_to_db(active_until) or ""
        milestone = select_milestone(
            active_until,
            now,
            SUBSCRIPTION_REMINDER_DAYS,
            SUBSCRIPTION_EXPIRED_QUEUE_DAYS,
        )

        with db.connect() as conn:
            # Оплата или ручное продление меняют дату. Старые черновики больше не актуальны.
            conn.execute(
                """
                UPDATE subscription_reminder_queue
                SET status = 'superseded', updated_at = ?
                WHERE telegram_id = ? AND status = 'pending' AND active_until <> ?
                """,
                (now_text, telegram_id, active_until_text),
            )

            if milestone is None:
                if active_until <= now:
                    conn.execute(
                        """
                        UPDATE subscription_reminder_queue
                        SET status = 'superseded', updated_at = ?
                        WHERE telegram_id = ? AND status = 'pending'
                        """,
                        (now_text, telegram_id),
                    )
                conn.commit()
                continue

            lang = get_user_language(telegram_id)
            plans_text = SUBSCRIPTION_PLANS_TEXT_UZ if lang == "uz" else SUBSCRIPTION_PLANS_TEXT
            draft_text = build_reminder_draft(
                lang=lang,
                active_until_text=_fmt_dt(active_until),
                subscription_kind=kind,
                milestone=milestone,
                plans_text=escape(plans_text),
            )
            cursor = conn.execute(
                """
                INSERT OR IGNORE INTO subscription_reminder_queue
                (telegram_id, active_until, subscription_kind, milestone, days_remaining,
                 status, draft_text, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, 'pending', ?, ?, ?)
                """,
                (
                    telegram_id,
                    active_until_text,
                    kind,
                    milestone.key,
                    milestone.days_remaining,
                    draft_text,
                    now_text,
                    now_text,
                ),
            )
            created += max(0, int(cursor.rowcount or 0))
            conn.execute(
                """
                UPDATE subscription_reminder_queue
                SET draft_text = ?, days_remaining = ?, updated_at = ?, last_error = NULL
                WHERE telegram_id = ? AND active_until = ? AND milestone = ? AND status = 'pending'
                """,
                (
                    draft_text,
                    milestone.days_remaining,
                    now_text,
                    telegram_id,
                    active_until_text,
                    milestone.key,
                ),
            )
            # На экране действий оставляем только самый свежий этап для одной даты.
            conn.execute(
                """
                UPDATE subscription_reminder_queue
                SET status = 'superseded', updated_at = ?
                WHERE telegram_id = ? AND active_until = ? AND status = 'pending' AND milestone <> ?
                """,
                (now_text, telegram_id, active_until_text, milestone.key),
            )
            conn.commit()

    pending_rows = list_pending_subscription_reminders(500)
    summary = {"created": created, "pending": len(pending_rows), "d7": 0, "d3": 0, "d1": 0, "expired": 0}
    for row in pending_rows:
        key = str(row.get("milestone") or "")
        summary[key] = int(summary.get(key, 0)) + 1
    return summary


def list_pending_subscription_reminders(limit: int = 100) -> list[dict[str, Any]]:
    init_subscription_automation_tables()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT q.*, u.username, u.first_name, u.default_shop_id
            FROM subscription_reminder_queue q
            LEFT JOIN users u ON u.telegram_id = q.telegram_id
            WHERE q.status = 'pending'
            ORDER BY
                CASE q.milestone
                    WHEN 'expired' THEN 0
                    WHEN 'd1' THEN 1
                    WHEN 'd3' THEN 2
                    WHEN 'd7' THEN 3
                    ELSE 4
                END,
                q.active_until ASC,
                q.id ASC
            LIMIT ?
            """,
            (max(1, int(limit)),),
        ).fetchall()
    return [dict(row) for row in rows]


def get_subscription_reminder(queue_id: int, pending_only: bool = False) -> dict[str, Any] | None:
    init_subscription_automation_tables()
    where_pending = "AND q.status = 'pending'" if pending_only else ""
    with db.connect() as conn:
        row = conn.execute(
            f"""
            SELECT q.*, u.username, u.first_name, u.default_shop_id
            FROM subscription_reminder_queue q
            LEFT JOIN users u ON u.telegram_id = q.telegram_id
            WHERE q.id = ? {where_pending}
            """,
            (int(queue_id),),
        ).fetchone()
    return dict(row) if row else None


def claim_subscription_reminder(queue_id: int, admin_id: int) -> bool:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        cursor = conn.execute(
            """
            UPDATE subscription_reminder_queue
            SET status = 'sending', reviewed_by = ?, reviewed_at = ?, updated_at = ?, last_error = NULL
            WHERE id = ? AND status = 'pending'
            """,
            (int(admin_id), now_text, now_text, int(queue_id)),
        )
        conn.commit()
    return bool(cursor.rowcount)


def finish_subscription_reminder(queue_id: int, *, sent: bool, error: str = "") -> None:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        if sent:
            conn.execute(
                """
                UPDATE subscription_reminder_queue
                SET status = 'sent', sent_at = ?, updated_at = ?, last_error = NULL
                WHERE id = ? AND status = 'sending'
                """,
                (now_text, now_text, int(queue_id)),
            )
        else:
            conn.execute(
                """
                UPDATE subscription_reminder_queue
                SET status = 'pending', updated_at = ?, last_error = ?
                WHERE id = ? AND status = 'sending'
                """,
                (now_text, str(error or "Ошибка отправки")[:1000], int(queue_id)),
            )
        conn.commit()


def dismiss_subscription_reminder(queue_id: int, admin_id: int) -> bool:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        cursor = conn.execute(
            """
            UPDATE subscription_reminder_queue
            SET status = 'dismissed', reviewed_by = ?, reviewed_at = ?, updated_at = ?
            WHERE id = ? AND status = 'pending'
            """,
            (int(admin_id), now_text, now_text, int(queue_id)),
        )
        conn.commit()
    return bool(cursor.rowcount)


def _subscription_automation_state_get(key: str) -> str | None:
    init_subscription_automation_tables()
    with db.connect() as conn:
        row = conn.execute(
            "SELECT state_value FROM subscription_automation_state WHERE state_key = ?",
            (str(key),),
        ).fetchone()
    return str(row["state_value"]) if row and row["state_value"] is not None else None


def _subscription_automation_state_set(key: str, value: str) -> None:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        conn.execute(
            """
            INSERT INTO subscription_automation_state (state_key, state_value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(state_key) DO UPDATE SET
                state_value = excluded.state_value,
                updated_at = excluded.updated_at
            """,
            (str(key), str(value), now_text),
        )
        conn.commit()


def _renewal_user_label(row: dict[str, Any]) -> str:
    username = str(row.get("username") or "").strip()
    first_name = str(row.get("first_name") or "").strip()
    return f"@{username}" if username else (first_name or "без имени")


def subscription_action_line(row: dict[str, Any]) -> str:
    milestone = str(row.get("milestone") or "")
    emoji = {"expired": "🚨", "d1": "🔴", "d3": "🟠", "d7": "🟡"}.get(milestone, "▫️")
    kind = "оплата" if row.get("subscription_kind") == "paid" else "trial"
    return (
        f"{emoji} <b>#{int(row.get('id') or 0)}</b> | <code>{int(row.get('telegram_id') or 0)}</code> — "
        f"{escape(_renewal_user_label(row))}\n"
        f"{kind}, до <b>{_fmt_dt(row.get('active_until'))}</b> · открыть: <code>/reminder {int(row.get('id') or 0)}</code>"
    )


def build_subscription_action_digest(rows: list[dict[str, Any]], created: int = 0) -> str:
    counts = {"expired": 0, "d1": 0, "d3": 0, "d7": 0}
    for row in rows:
        key = str(row.get("milestone") or "")
        if key in counts:
            counts[key] += 1
    lines = [subscription_action_line(row) for row in rows[:25]]
    more = len(rows) - len(lines)
    body = "\n\n".join(lines) if lines else "— сейчас действий нет"
    if more > 0:
        body += f"\n\nЕщё записей: <b>{more}</b>"
    return (
        "🔐 <b>Контроль подписок</b>\n\n"
        f"🚨 Уже закончились: <b>{counts['expired']}</b>\n"
        f"🔴 До 1 дня: <b>{counts['d1']}</b>\n"
        f"🟠 До 3 дней: <b>{counts['d3']}</b>\n"
        f"🟡 До 7 дней: <b>{counts['d7']}</b>\n"
        f"🆕 Новых черновиков: <b>{created}</b>\n\n"
        f"{body}\n\n"
        "<b>Что делать:</b>\n"
        "1. Откройте черновик командой <code>/reminder НОМЕР</code>.\n"
        "2. Проверьте текст и дату.\n"
        "3. Отправьте его кнопкой подтверждения или отклоните.\n\n"
        "Без вашего нажатия клиенту ничего не отправляется, подписка не продлевается и не блокируется."
    )


def list_blocked_users(limit: int = 50) -> list[dict[str, Any]]:
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT u.telegram_id, u.username, u.first_name, u.default_shop_id,
                   s.trial_until, s.subscription_until, s.blocked
            FROM subscriptions s
            LEFT JOIN users u ON u.telegram_id = s.telegram_id
            WHERE s.blocked = 1
            ORDER BY s.updated_at DESC
            LIMIT ?
            """,
            (int(limit),),
        ).fetchall()
    return [dict(r) for r in rows]



def init_staff_connect_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS staff_shop_connections (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                error TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id)
            )
            """
        )
        conn.commit()


def save_staff_shop_status(telegram_id: int, shop_id: int, status: str, error: str = "") -> None:
    init_staff_connect_tables()
    now = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        conn.execute(
            """
            INSERT INTO staff_shop_connections (telegram_id, shop_id, status, error, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id) DO UPDATE SET
                status = excluded.status,
                error = excluded.error,
                updated_at = excluded.updated_at
            """,
            (int(telegram_id), int(shop_id), str(status), str(error)[:1000], now, now),
        )
        conn.commit()


def list_staff_shop_connections(limit: int = 30) -> list[dict[str, Any]]:
    init_staff_connect_tables()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT telegram_id, shop_id, status, error, created_at, updated_at
            FROM staff_shop_connections
            ORDER BY updated_at DESC
            LIMIT ?
            """,
            (int(limit),),
        ).fetchall()
    return [dict(r) for r in rows]


# --- Персональные настройки и доказуемая бизнес-ценность ---
PRODUCT_SETTING_FIELDS = {
    "notify_orders",
    "notify_sales",
    "sales_notification_mode",
    "notify_low_stock",
    "notify_out_of_stock",
    "notify_cancellations",
    "notify_reviews",
    "notify_stock_change",
    "notify_losses",
    "notify_defects",
    "notify_fbo_acceptance",
    "notify_supply_reminders",
    "notify_return_pickup",
    "daily_enabled",
    "daily_hour",
    "weekly_enabled",
    "weekly_weekday",
    "weekly_hour",
    "monthly_enabled",
    "monthly_day",
    "monthly_hour",
    "low_stock_threshold",
    "lead_time_days",
    "safety_days",
    "target_cover_days",
}

FINANCE_SETTING_FIELDS = {
    "tax_percent",
    "advertising_monthly",
    "storage_monthly",
    "other_monthly",
}


def init_product_value_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS product_settings (
                telegram_id INTEGER PRIMARY KEY,
                notify_orders INTEGER NOT NULL DEFAULT 0,
                notify_sales INTEGER NOT NULL DEFAULT 1,
                sales_notification_mode TEXT NOT NULL DEFAULT 'hourly',
                sales_mode_explicit INTEGER NOT NULL DEFAULT 0,
                notify_low_stock INTEGER NOT NULL DEFAULT 1,
                notify_out_of_stock INTEGER NOT NULL DEFAULT 1,
                notify_cancellations INTEGER NOT NULL DEFAULT 1,
                notify_reviews INTEGER NOT NULL DEFAULT 0,
                notify_stock_change INTEGER NOT NULL DEFAULT 0,
                notify_losses INTEGER NOT NULL DEFAULT 1,
                notify_defects INTEGER NOT NULL DEFAULT 1,
                notify_fbo_acceptance INTEGER NOT NULL DEFAULT 1,
                notify_supply_reminders INTEGER NOT NULL DEFAULT 1,
                notify_return_pickup INTEGER NOT NULL DEFAULT 1,
                daily_enabled INTEGER NOT NULL DEFAULT 0,
                daily_hour INTEGER NOT NULL DEFAULT 9,
                weekly_enabled INTEGER NOT NULL DEFAULT 0,
                weekly_weekday INTEGER NOT NULL DEFAULT 0,
                weekly_hour INTEGER NOT NULL DEFAULT 9,
                monthly_enabled INTEGER NOT NULL DEFAULT 0,
                monthly_day INTEGER NOT NULL DEFAULT 1,
                monthly_hour INTEGER NOT NULL DEFAULT 9,
                low_stock_threshold INTEGER NOT NULL DEFAULT 5,
                lead_time_days INTEGER NOT NULL DEFAULT 3,
                safety_days INTEGER NOT NULL DEFAULT 5,
                target_cover_days INTEGER NOT NULL DEFAULT 30,
                updated_at TEXT NOT NULL
            )
            """
        )
        existing_product_columns = {
            str(row[1]) for row in conn.execute("PRAGMA table_info(product_settings)").fetchall()
        }
        for column in (
            "notify_losses",
            "notify_defects",
            "notify_fbo_acceptance",
            "notify_supply_reminders",
            "notify_return_pickup",
        ):
            if column not in existing_product_columns:
                conn.execute(
                    f"ALTER TABLE product_settings ADD COLUMN {column} INTEGER NOT NULL DEFAULT 1"
                )
        if "sales_notification_mode" not in existing_product_columns:
            conn.execute(
                "ALTER TABLE product_settings "
                "ADD COLUMN sales_notification_mode TEXT NOT NULL DEFAULT 'hourly'"
            )
        if "sales_mode_explicit" not in existing_product_columns:
            conn.execute(
                "ALTER TABLE product_settings "
                "ADD COLUMN sales_mode_explicit INTEGER NOT NULL DEFAULT 0"
            )
        migrated_hourly = conn.execute(
            """
            UPDATE product_settings
            SET sales_notification_mode = 'hourly', updated_at = ?
            WHERE notify_sales = 1
              AND sales_mode_explicit = 0
              AND LOWER(COALESCE(sales_notification_mode, '')) IN ('', 'instant')
            """,
            (_dt_to_db(_utc_now()) or "",),
        ).rowcount
        if migrated_hourly:
            logging.info(
                "Sales notification default migrated to hourly users=%s",
                migrated_hourly,
            )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS business_finance_settings (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                tax_percent REAL NOT NULL DEFAULT 0,
                advertising_monthly REAL NOT NULL DEFAULT 0,
                storage_monthly REAL NOT NULL DEFAULT 0,
                other_monthly REAL NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS business_actions (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                action_key TEXT NOT NULL,
                category TEXT NOT NULL,
                title TEXT,
                amount REAL NOT NULL DEFAULT 0,
                state TEXT NOT NULL DEFAULT 'active',
                snoozed_until TEXT,
                first_seen TEXT NOT NULL,
                last_seen TEXT NOT NULL,
                resolved_at TEXT,
                PRIMARY KEY (telegram_id, shop_id, action_key)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS business_value_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                action_key TEXT,
                event_type TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                description TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS scheduled_report_delivery (
                telegram_id INTEGER NOT NULL,
                report_kind TEXT NOT NULL,
                period_key TEXT NOT NULL,
                sent_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, report_kind, period_key)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS operational_watcher_baseline (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                watcher_kind TEXT NOT NULL,
                initialized_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, watcher_kind)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS product_loss_snapshot (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                sku_key TEXT NOT NULL,
                product_title TEXT,
                sku_title TEXT,
                sku_id TEXT,
                barcode TEXT,
                missing_qty INTEGER NOT NULL DEFAULT 0,
                defected_qty INTEGER NOT NULL DEFAULT 0,
                price REAL NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, sku_key)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS loss_defect_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                event_key TEXT NOT NULL,
                sku_key TEXT NOT NULL,
                product_title TEXT,
                sku_id TEXT,
                barcode TEXT,
                missing_delta INTEGER NOT NULL DEFAULT 0,
                defected_delta INTEGER NOT NULL DEFAULT 0,
                missing_qty INTEGER NOT NULL DEFAULT 0,
                defected_qty INTEGER NOT NULL DEFAULT 0,
                estimated_value REAL NOT NULL DEFAULT 0,
                detected_at TEXT NOT NULL,
                UNIQUE (telegram_id, shop_id, event_key)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_loss_defect_events_period
            ON loss_defect_events (telegram_id, shop_id, detected_at)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS fbo_acceptance_watch (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                invoice_key TEXT NOT NULL,
                invoice_id TEXT,
                invoice_number TEXT,
                status TEXT,
                planned_qty REAL NOT NULL DEFAULT 0,
                accepted_qty REAL NOT NULL DEFAULT 0,
                terminal INTEGER NOT NULL DEFAULT 0,
                first_seen TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                notified_at TEXT,
                PRIMARY KEY (telegram_id, shop_id, invoice_key)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_fbo_acceptance_pending
            ON fbo_acceptance_watch (telegram_id, shop_id, terminal, notified_at)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_digest_queue (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                event_key TEXT NOT NULL,
                identity_key TEXT NOT NULL,
                order_id TEXT,
                product_title TEXT,
                sku_title TEXT,
                quantity REAL NOT NULL DEFAULT 0,
                revenue REAL NOT NULL DEFAULT 0,
                commission REAL NOT NULL DEFAULT 0,
                logistics REAL NOT NULL DEFAULT 0,
                payout REAL NOT NULL DEFAULT 0,
                sold_at TEXT,
                detected_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, event_key)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_sales_digest_queue_pending
            ON sales_digest_queue (telegram_id, shop_id, detected_at)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_digest_state (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                last_sent_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sales_digest_user_state (
                telegram_id INTEGER PRIMARY KEY,
                last_sent_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def ensure_product_settings(telegram_id: int) -> dict[str, Any]:
    init_product_value_tables()
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO product_settings (
                telegram_id, notify_orders, notify_sales,
                sales_notification_mode, sales_mode_explicit, notify_low_stock,
                notify_out_of_stock, notify_cancellations, notify_reviews,
                notify_stock_change, daily_enabled, daily_hour,
                low_stock_threshold, updated_at
            ) VALUES (?, ?, ?, ?, 0, ?, ?, 1, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(telegram_id),
                1 if NEW_ORDER_NOTIFICATIONS else 0,
                1 if SALE_NOTIFICATIONS else 0,
                "hourly",
                1 if LOW_STOCK_NOTIFICATIONS else 0,
                1 if OUT_OF_STOCK_NOTIFICATIONS else 0,
                1 if REVIEW_NOTIFICATIONS else 0,
                1 if STOCK_CHANGE_NOTIFICATIONS else 0,
                1 if DAILY_REPORTS else 0,
                max(0, min(23, int(DAILY_REPORT_HOUR_UZT))),
                max(0, int(LOW_STOCK_THRESHOLD)),
                now_text,
            ),
        )
        row = conn.execute(
            "SELECT * FROM product_settings WHERE telegram_id = ?",
            (int(telegram_id),),
        ).fetchone()
        conn.commit()
    return dict(row) if row else {}


def update_product_setting(telegram_id: int, field: str, value: Any) -> dict[str, Any]:
    if field not in PRODUCT_SETTING_FIELDS:
        raise ValueError(f"Unsupported product setting: {field}")
    ensure_product_settings(telegram_id)
    explicit_clause = (
        ", sales_mode_explicit = 1"
        if field == "sales_notification_mode"
        else ""
    )
    with db.connect() as conn:
        conn.execute(
            f"UPDATE product_settings SET {field} = ?{explicit_clause}, updated_at = ? WHERE telegram_id = ?",
            (value, _dt_to_db(_utc_now()) or "", int(telegram_id)),
        )
        conn.commit()
    return ensure_product_settings(telegram_id)


def product_setting_enabled(telegram_id: int, field: str) -> bool:
    return bool(int(ensure_product_settings(telegram_id).get(field) or 0))


SALES_NOTIFICATION_MODES = {"instant", "hourly", "off"}


def _sales_notification_mode_from_settings(settings: dict[str, Any]) -> str:
    if not bool(int(settings.get("notify_sales") or 0)):
        return "off"
    mode = str(settings.get("sales_notification_mode") or "hourly").strip().lower()
    return mode if mode in {"instant", "hourly"} else "hourly"


def get_sales_notification_mode(telegram_id: int) -> str:
    return _sales_notification_mode_from_settings(ensure_product_settings(telegram_id))


def set_sales_notification_mode(telegram_id: int, mode: str) -> dict[str, Any]:
    normalized = str(mode or "").strip().lower()
    if normalized not in SALES_NOTIFICATION_MODES:
        raise ValueError(f"Unsupported sales notification mode: {mode}")
    ensure_product_settings(telegram_id)
    enabled = 0 if normalized == "off" else 1
    stored_mode = "hourly" if normalized == "off" else normalized
    with db.connect() as conn:
        conn.execute(
            """
            UPDATE product_settings
            SET notify_sales = ?, sales_notification_mode = ?,
                sales_mode_explicit = 1, updated_at = ?
            WHERE telegram_id = ?
            """,
            (
                enabled,
                stored_mode,
                _dt_to_db(_utc_now()) or "",
                int(telegram_id),
            ),
        )
        conn.commit()
    return ensure_product_settings(telegram_id)


def ensure_finance_settings(telegram_id: int, shop_id: int) -> dict[str, Any]:
    init_product_value_tables()
    with db.connect() as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO business_finance_settings
                (telegram_id, shop_id, updated_at)
            VALUES (?, ?, ?)
            """,
            (int(telegram_id), int(shop_id), _dt_to_db(_utc_now()) or ""),
        )
        row = conn.execute(
            """
            SELECT * FROM business_finance_settings
            WHERE telegram_id = ? AND shop_id = ?
            """,
            (int(telegram_id), int(shop_id)),
        ).fetchone()
        conn.commit()
    return dict(row) if row else {}


def update_finance_setting(
    telegram_id: int,
    shop_id: int,
    field: str,
    value: float,
) -> dict[str, Any]:
    if field not in FINANCE_SETTING_FIELDS:
        raise ValueError(f"Unsupported finance setting: {field}")
    ensure_finance_settings(telegram_id, shop_id)
    with db.connect() as conn:
        conn.execute(
            f"""
            UPDATE business_finance_settings
            SET {field} = ?, updated_at = ?
            WHERE telegram_id = ? AND shop_id = ?
            """,
            (float(value), _dt_to_db(_utc_now()) or "", int(telegram_id), int(shop_id)),
        )
        conn.commit()
    return ensure_finance_settings(telegram_id, shop_id)


def _business_action_key(action: dict[str, Any]) -> str:
    raw = "|".join(
        str(action.get(key) or "").strip().lower()
        for key in ("category_ru", "sku", "title", "source")
    )
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def sync_business_actions(
    telegram_id: int,
    shop_id: int,
    actions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    init_product_value_tables()
    now = _utc_now()
    now_text = _dt_to_db(now) or ""
    with db.connect() as conn:
        conn.execute(
            """
            UPDATE business_actions
            SET state = 'active', snoozed_until = NULL
            WHERE telegram_id = ? AND shop_id = ?
              AND snoozed_until IS NOT NULL AND snoozed_until <= ?
            """,
            (int(telegram_id), int(shop_id), now_text),
        )
        for action in actions:
            key = str(action.get("action_key") or _business_action_key(action))
            action["action_key"] = key
            conn.execute(
                """
                INSERT INTO business_actions (
                    telegram_id, shop_id, action_key, category, title, amount,
                    state, first_seen, last_seen
                ) VALUES (?, ?, ?, ?, ?, ?, 'active', ?, ?)
                ON CONFLICT (telegram_id, shop_id, action_key) DO UPDATE SET
                    category = excluded.category,
                    title = excluded.title,
                    amount = excluded.amount,
                    last_seen = excluded.last_seen
                """,
                (
                    int(telegram_id),
                    int(shop_id),
                    key,
                    str(action.get("category_ru") or ""),
                    str(action.get("title") or ""),
                    max(0.0, float(action.get("amount") or 0)),
                    now_text,
                    now_text,
                ),
            )
        rows = conn.execute(
            """
            SELECT action_key, state, snoozed_until
            FROM business_actions
            WHERE telegram_id = ? AND shop_id = ?
            """,
            (int(telegram_id), int(shop_id)),
        ).fetchall()
        conn.commit()
    states = {str(row["action_key"]): dict(row) for row in rows}
    visible: list[dict[str, Any]] = []
    for action in actions:
        state = states.get(str(action.get("action_key") or ""), {})
        if str(state.get("state") or "active") != "active":
            continue
        visible.append(action)
    return visible


def update_business_action_state(
    telegram_id: int,
    shop_id: int,
    action_key: str,
    state: str,
) -> bool:
    if state not in {"resolved", "snoozed"}:
        return False
    init_product_value_tables()
    now = _utc_now()
    until = now + timedelta(days=7 if state == "resolved" else 3)
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT category, title, amount, state, snoozed_until
            FROM business_actions
            WHERE telegram_id = ? AND shop_id = ? AND action_key = ?
            """,
            (int(telegram_id), int(shop_id), str(action_key)),
        ).fetchone()
        if not row:
            return False
        previous_until = _dt_from_db(row["snoozed_until"])
        already_suppressed = (
            str(row["state"] or "") == state
            and previous_until is not None
            and previous_until > now
        )
        conn.execute(
            """
            UPDATE business_actions
            SET state = ?, snoozed_until = ?, resolved_at = ?, last_seen = last_seen
            WHERE telegram_id = ? AND shop_id = ? AND action_key = ?
            """,
            (
                state,
                _dt_to_db(until),
                _dt_to_db(now) if state == "resolved" else None,
                int(telegram_id),
                int(shop_id),
                str(action_key),
            ),
        )
        if state == "resolved" and not already_suppressed:
            conn.execute(
                """
                INSERT INTO business_value_events (
                    telegram_id, shop_id, action_key, event_type, amount,
                    description, created_at
                ) VALUES (?, ?, ?, 'resolved_action', ?, ?, ?)
                """,
                (
                    int(telegram_id),
                    int(shop_id),
                    str(action_key),
                    max(0.0, float(row["amount"] or 0)),
                    f"{row['category']}: {row['title']}",
                    _dt_to_db(now) or "",
                ),
            )
        conn.commit()
    return True


def resolved_business_value(telegram_id: int, shop_id: int) -> float:
    now = datetime.now(UZT)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT COALESCE(SUM(amount), 0) AS total
            FROM business_value_events
            WHERE telegram_id = ? AND shop_id = ?
              AND event_type = 'resolved_action' AND created_at >= ?
            """,
            (
                int(telegram_id),
                int(shop_id),
                _dt_to_db(month_start.astimezone(timezone.utc)) or "",
            ),
        ).fetchone()
    return float(row["total"] or 0) if row else 0.0


def scheduled_report_was_sent(telegram_id: int, kind: str, period_key: str) -> bool:
    init_product_value_tables()
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT 1 FROM scheduled_report_delivery
            WHERE telegram_id = ? AND report_kind = ? AND period_key = ?
            """,
            (int(telegram_id), str(kind), str(period_key)),
        ).fetchone()
    return row is not None


def mark_scheduled_report_sent(telegram_id: int, kind: str, period_key: str) -> None:
    init_product_value_tables()
    with db.connect() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO scheduled_report_delivery
                (telegram_id, report_kind, period_key, sent_at)
            VALUES (?, ?, ?, ?)
            """,
            (int(telegram_id), str(kind), str(period_key), _dt_to_db(_utc_now()) or ""),
        )
        conn.commit()


def operational_watcher_initialized(telegram_id: int, shop_id: int, kind: str) -> bool:
    init_product_value_tables()
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT 1 FROM operational_watcher_baseline
            WHERE telegram_id = ? AND shop_id = ? AND watcher_kind = ?
            """,
            (int(telegram_id), int(shop_id), str(kind)),
        ).fetchone()
    return row is not None


def mark_operational_watcher_initialized(telegram_id: int, shop_id: int, kind: str) -> None:
    init_product_value_tables()
    with db.connect() as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO operational_watcher_baseline
                (telegram_id, shop_id, watcher_kind, initialized_at)
            VALUES (?, ?, ?, ?)
            """,
            (int(telegram_id), int(shop_id), str(kind), _dt_to_db(_utc_now()) or ""),
        )
        conn.commit()


def load_product_loss_snapshot(
    telegram_id: int,
    shop_id: int,
) -> dict[str, dict[str, Any]]:
    init_product_value_tables()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT sku_key, product_title, sku_title, sku_id, barcode,
                   missing_qty, defected_qty, price
            FROM product_loss_snapshot
            WHERE telegram_id = ? AND shop_id = ?
            """,
            (int(telegram_id), int(shop_id)),
        ).fetchall()
    return {str(row["sku_key"]): dict(row) for row in rows}


def save_product_loss_snapshot(
    telegram_id: int,
    shop_id: int,
    snapshot: dict[str, dict[str, Any]],
    *,
    reset_absent: bool,
) -> None:
    init_product_value_tables()
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        if reset_absent:
            conn.execute(
                """
                UPDATE product_loss_snapshot
                SET missing_qty = 0, defected_qty = 0, updated_at = ?
                WHERE telegram_id = ? AND shop_id = ?
                """,
                (now_text, int(telegram_id), int(shop_id)),
            )
        for key, item in snapshot.items():
            conn.execute(
                """
                INSERT INTO product_loss_snapshot (
                    telegram_id, shop_id, sku_key, product_title, sku_title,
                    sku_id, barcode, missing_qty, defected_qty, price, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (telegram_id, shop_id, sku_key) DO UPDATE SET
                    product_title = excluded.product_title,
                    sku_title = excluded.sku_title,
                    sku_id = excluded.sku_id,
                    barcode = excluded.barcode,
                    missing_qty = excluded.missing_qty,
                    defected_qty = excluded.defected_qty,
                    price = excluded.price,
                    updated_at = excluded.updated_at
                """,
                (
                    int(telegram_id),
                    int(shop_id),
                    str(key),
                    str(item.get("product_title") or ""),
                    str(item.get("sku_title") or ""),
                    str(item.get("sku_id") or ""),
                    str(item.get("barcode") or ""),
                    max(0, int(item.get("missing_qty") or 0)),
                    max(0, int(item.get("defected_qty") or 0)),
                    max(0.0, float(item.get("price") or 0)),
                    now_text,
                ),
            )
        conn.commit()


def record_loss_defect_events(
    telegram_id: int,
    shop_id: int,
    changes: list[dict[str, Any]],
) -> None:
    """Persist watcher deltas so period PDF reports can show new defects/losses."""
    if not changes:
        return
    init_product_value_tables()
    detected_at = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        for item in changes:
            sku_key = str(item.get("sku_key") or "unknown")
            event_source = "|".join(
                (
                    sku_key,
                    str(max(0, int(item.get("missing_qty") or 0))),
                    str(max(0, int(item.get("defected_qty") or 0))),
                )
            )
            event_key = hashlib.sha256(event_source.encode("utf-8")).hexdigest()
            conn.execute(
                """
                INSERT OR IGNORE INTO loss_defect_events (
                    telegram_id, shop_id, event_key, sku_key, product_title,
                    sku_id, barcode, missing_delta, defected_delta,
                    missing_qty, defected_qty, estimated_value, detected_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(telegram_id),
                    int(shop_id),
                    event_key,
                    sku_key,
                    str(item.get("product_title") or ""),
                    str(item.get("sku_id") or ""),
                    str(item.get("barcode") or ""),
                    max(0, int(item.get("missing_delta") or 0)),
                    max(0, int(item.get("defected_delta") or 0)),
                    max(0, int(item.get("missing_qty") or 0)),
                    max(0, int(item.get("defected_qty") or 0)),
                    max(0.0, float(item.get("estimated_value") or 0)),
                    detected_at,
                ),
            )
        conn.commit()


def list_loss_defect_events(
    telegram_id: int,
    shop_id: int,
    date_from_ms: int,
    date_to_ms: int,
) -> list[dict[str, Any]]:
    """Return recorded deltas detected inside the selected Uzbekistan period."""
    init_product_value_tables()
    date_from = datetime.fromtimestamp(date_from_ms / 1000, timezone.utc)
    date_to = datetime.fromtimestamp(date_to_ms / 1000, timezone.utc)
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT product_title, sku_id, barcode, missing_delta,
                   defected_delta, missing_qty, defected_qty,
                   estimated_value, detected_at
            FROM loss_defect_events
            WHERE telegram_id = ? AND shop_id = ?
              AND detected_at >= ? AND detected_at <= ?
            ORDER BY detected_at DESC, id DESC
            """,
            (
                int(telegram_id),
                int(shop_id),
                _dt_to_db(date_from) or "",
                _dt_to_db(date_to) or "",
            ),
        ).fetchall()
    return [dict(row) for row in rows]


def get_fbo_acceptance_watch_state(
    telegram_id: int,
    shop_id: int,
    invoice_key: str,
) -> dict[str, Any] | None:
    init_product_value_tables()
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT * FROM fbo_acceptance_watch
            WHERE telegram_id = ? AND shop_id = ? AND invoice_key = ?
            """,
            (int(telegram_id), int(shop_id), str(invoice_key)),
        ).fetchone()
    return dict(row) if row else None


def save_fbo_acceptance_watch_state(
    telegram_id: int,
    shop_id: int,
    invoice_key: str,
    *,
    invoice_id: Any,
    invoice_number: str,
    status: str,
    planned_qty: float,
    accepted_qty: float,
    terminal: bool,
    baseline_notified: bool = False,
) -> None:
    init_product_value_tables()
    now_text = _dt_to_db(_utc_now()) or ""
    notified_at = now_text if baseline_notified else None
    with db.connect() as conn:
        conn.execute(
            """
            INSERT INTO fbo_acceptance_watch (
                telegram_id, shop_id, invoice_key, invoice_id, invoice_number,
                status, planned_qty, accepted_qty, terminal, first_seen,
                updated_at, notified_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (telegram_id, shop_id, invoice_key) DO UPDATE SET
                invoice_id = excluded.invoice_id,
                invoice_number = excluded.invoice_number,
                status = excluded.status,
                planned_qty = excluded.planned_qty,
                accepted_qty = excluded.accepted_qty,
                terminal = excluded.terminal,
                updated_at = excluded.updated_at,
                notified_at = COALESCE(fbo_acceptance_watch.notified_at, excluded.notified_at)
            """,
            (
                int(telegram_id),
                int(shop_id),
                str(invoice_key),
                str(invoice_id or ""),
                str(invoice_number or ""),
                str(status or ""),
                max(0.0, float(planned_qty or 0)),
                max(0.0, float(accepted_qty or 0)),
                1 if terminal else 0,
                now_text,
                now_text,
                notified_at,
            ),
        )
        conn.commit()


def mark_fbo_acceptance_notified(
    telegram_id: int,
    shop_id: int,
    invoice_key: str,
) -> None:
    init_product_value_tables()
    with db.connect() as conn:
        conn.execute(
            """
            UPDATE fbo_acceptance_watch
            SET notified_at = ?, updated_at = ?
            WHERE telegram_id = ? AND shop_id = ? AND invoice_key = ?
            """,
            (
                _dt_to_db(_utc_now()) or "",
                _dt_to_db(_utc_now()) or "",
                int(telegram_id),
                int(shop_id),
                str(invoice_key),
            ),
        )
        conn.commit()


init_subscription_tables()
init_business_tables()
init_subscription_automation_tables()
init_payment_request_tables()
init_unit_economy_tables()
init_staff_connect_tables()
init_product_value_tables()

# --- Языки интерфейса ---
# Основной код отчётов остаётся совместимым с русскими командами, но клиент может выбрать язык меню и основных экранов.
SUPPORTED_LANGUAGES = {"ru", "uz"}


def init_language_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS user_language (
                telegram_id INTEGER PRIMARY KEY,
                lang TEXT NOT NULL DEFAULT 'ru',
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def normalize_lang(value: Any) -> str:
    lang = str(value or "ru").strip().lower()
    if lang.startswith("uz"):
        return "uz"
    return "ru"


def get_user_language(telegram_id: int | None) -> str:
    if telegram_id is None:
        return "ru"
    try:
        with db.connect() as conn:
            row = conn.execute(
                "SELECT lang FROM user_language WHERE telegram_id = ?",
                (int(telegram_id),),
            ).fetchone()
        if row:
            return normalize_lang(row[0] if not isinstance(row, dict) else row.get("lang"))
    except Exception:
        logging.exception("Failed to read user language")
    return "ru"


def set_user_language(telegram_id: int, lang: str) -> None:
    lang = normalize_lang(lang)
    now = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        conn.execute(
            """
            INSERT INTO user_language (telegram_id, lang, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(telegram_id) DO UPDATE SET lang = excluded.lang, updated_at = excluded.updated_at
            """,
            (int(telegram_id), lang, now),
        )
        conn.commit()



# Создаём таблицу языков после определения функции, чтобы не было NameError при запуске.

def language_title(lang: str) -> str:
    return "O‘zbekcha" if normalize_lang(lang) == "uz" else "Русский"


I18N: dict[str, dict[str, str]] = {
    "ru": {
        "choose_action": "Выберите действие",
        "choose_section": "Выберите раздел 👇",
        "main_menu": "Главное меню 👇",
        "cancelled": "Действие отменено.",
        "language_title": "🌐 <b>Язык интерфейса</b>",
        "language_body": "Выберите язык, на котором бот будет показывать меню и основные подсказки.",
        "language_set": "✅ Язык изменён: <b>Русский</b>",
        "language_button_ru": "🇷🇺 Русский",
        "language_button_uz": "🇺🇿 O‘zbekcha",
        "admin_only": "⛔ Админ-панель доступна только владельцу бота.",
        "access_limited": "⛔ <b>Доступ ограничен</b>\n\nTrial или подписка закончились.\nВаш Uzum-токен и настройки сохранены — после продления всё снова заработает.\n\nПроверить подписку: <code>/my_subscription</code>\nОплата: <code>/subscribe</code>",
        "connect_first": "Сначала подключите Uzum API-токен: <code>/connect</code>",
        "connection_deleted": "✅ Подключение к Uzum API удалено. Можно подключить заново через <code>/connect</code>.",
        "token_instruction_title": "🔑 <b>Где взять Uzum Seller API-ключ</b>",
        "support_title": "🆘 <b>Поддержка</b>",
        "security_title": "🔐 <b>Безопасность API-ключа</b>",
    },
    "uz": {
        "choose_action": "Amalni tanlang",
        "choose_section": "Bo‘limni tanlang 👇",
        "main_menu": "Asosiy menyu 👇",
        "cancelled": "Amal bekor qilindi.",
        "language_title": "🌐 <b>Interfeys tili</b>",
        "language_body": "Bot menyu va asosiy ko‘rsatmalarni qaysi tilda ko‘rsatishini tanlang.",
        "language_set": "✅ Til o‘zgartirildi: <b>O‘zbekcha</b>",
        "language_button_ru": "🇷🇺 Русский",
        "language_button_uz": "🇺🇿 O‘zbekcha",
        "admin_only": "⛔ Admin panel faqat bot egasi uchun.",
        "access_limited": "⛔ <b>Kirish cheklangan</b>\n\nTrial yoki obuna muddati tugagan.\nUzum tokeningiz va sozlamalaringiz saqlanadi — obuna uzaytirilgach hammasi yana ishlaydi.\n\nObunani tekshirish: <code>/my_subscription</code>\nTo‘lov: <code>/subscribe</code>",
        "connect_first": "Avval Uzum API-kalitini ulang: <code>/connect</code>",
        "connection_deleted": "✅ Uzum API ulanishi o‘chirildi. Qayta ulash uchun <code>/connect</code> buyrug‘idan foydalaning.",
        "token_instruction_title": "🔑 <b>Uzum Seller API-kalitini qayerdan olish mumkin</b>",
        "support_title": "🆘 <b>Yordam</b>",
        "security_title": "🔐 <b>API-kalit xavfsizligi</b>",
    },
}


def tr(lang: str, key: str) -> str:
    lang = normalize_lang(lang)
    return I18N.get(lang, I18N["ru"]).get(key, I18N["ru"].get(key, key))


def tr_user(telegram_id: int | None, key: str) -> str:
    return tr(get_user_language(telegram_id), key)



# --- Автоперевод сообщений с данными на узбекский ---
def translate_runtime_text_to_uz(text: str) -> str:
    """Лёгкий пост-процессор: переводит основные русские ответы и отчёты на узбекский.

    Меню уже переключается отдельными клавиатурами. Этот слой нужен для старых
    функций, где текст отчётов был собран на русском внутри бизнес-логики.
    Числа, ID, SKU, статусы и суммы не меняются.
    """
    if not isinstance(text, str) or not text:
        return text

    replacements = [
        # waiting / service
        ("⌛ Считаю баланс за 30 дней...", "⌛ 30 kunlik balans hisoblanmoqda..."),
        ("⌛ Считаю баланс по всем магазинам за 30 дней...", "⌛ Barcha do‘konlar bo‘yicha 30 kunlik balans hisoblanmoqda..."),
        ("⌛ Считаю продажи за сегодня...", "⌛ Bugungi sotuvlar hisoblanmoqda..."),
        ("⌛ Считаю продажи за вчера...", "⌛ Kechagi sotuvlar hisoblanmoqda..."),
        ("⌛ Считаю продажи за 7 дней...", "⌛ 7 kunlik sotuvlar hisoblanmoqda..."),
        ("⏳ Считаю продажи за сегодня, 7 и 30 дней...", "⏳ Bugun, 7 kun va 30 kunlik sotuvlar hisoblanmoqda..."),
        ("⏳ Считаю заказы по статусам...", "⏳ Buyurtmalar statuslar bo‘yicha hisoblanmoqda..."),
        ("⌛ Считаю топ товаров", "⌛ Top tovarlar hisoblanmoqda"),
        ("⌛ Считаю, на сколько дней хватит остатков...", "⌛ Qoldiq necha kunga yetishi hisoblanmoqda..."),
        ("⏳ Готовлю Excel-отчёт...", "⏳ Excel hisobot tayyorlanmoqda..."),
        ("⏳ Собираю утренний отчёт...", "⏳ Ertalabki hisobot tayyorlanmoqda..."),

        # titles
        ("💰 <b>Баланс Uzum FBO за 30 дней</b>", "💰 <b>Uzum FBO balansi 30 kun uchun</b>"),
        ("💰 <b>Продажи Uzum FBO/FBS за сегодня</b>", "💰 <b>Bugungi Uzum FBO/FBS sotuvlari</b>"),
        ("💰 <b>Продажи Uzum FBO/FBS за вчера</b>", "💰 <b>Kechagi Uzum FBO/FBS sotuvlari</b>"),
        ("💰 <b>Продажи Uzum FBO/FBS за 7 дней</b>", "💰 <b>7 kunlik Uzum FBO/FBS sotuvlari</b>"),
        ("💰 <b>Продажи Uzum FBO/FBS за 30 дней</b>", "💰 <b>30 kunlik Uzum FBO/FBS sotuvlari</b>"),
        ("🌐 <b>Баланс по всем магазинам за 30 дней</b>", "🌐 <b>Barcha do‘konlar bo‘yicha 30 kunlik balans</b>"),
        ("📊 <b>Сводка продаж</b>", "📊 <b>Sotuvlar xulosasi</b>"),
        ("📊 <b>Сводка заказов</b>", "📊 <b>Buyurtmalar xulosasi</b>"),
        ("📦 <b>Остатки</b>", "📦 <b>Qoldiq</b>"),
        ("⚠️ <b>Умное 'Заканчивается'</b>", "⚠️ <b>Qoldiq prognozi</b>"),
        ("🏆 <b>Топ товаров", "🏆 <b>Top tovarlar"),
        ("🐢 <b>Товары без продаж", "🐢 <b>Sotilmayotgan tovarlar"),
        ("🏪 <b>Ваши магазины</b>", "🏪 <b>Do‘konlaringiz</b>"),
        ("📄 <b>FBO-накладные поставки</b>", "📄 <b>FBO yuk xatlari</b>"),
        ("📦 <b>Состав FBO-накладной</b>", "📦 <b>FBO yuk xati tarkibi</b>"),
        ("🌙 <b>Утренний отчёт Uzum</b>", "🌙 <b>Uzum ertalabki hisoboti</b>"),
        ("🛒 <b>Новая продажа Uzum FBO</b>", "🛒 <b>Yangi Uzum FBO sotuvi</b>"),
        ("⚠️ <b>Заканчиваются товары</b>", "⚠️ <b>Tovarlar tugayapti</b>"),
        ("❌ <b>Товары закончились</b>", "❌ <b>Tovarlar tugagan</b>"),
        ("💎 <b>Моя подписка</b>", "💎 <b>Mening obunam</b>"),
        ("👑 <b>Админ-панель</b>", "👑 <b>Admin panel</b>"),

        # labels, finance
        ("Магазинов найдено:", "Topilgan do‘konlar:"),
        ("Магазинов:", "Do‘konlar soni:"),
        ("Магазин:", "Do‘kon:"),
        ("Текущий магазин:", "Joriy do‘kon:"),
        ("Активный магазин:", "Faol do‘kon:"),
        ("Позиции продаж:", "Sotuv pozitsiyalari:"),
        ("Кол-во товаров:", "Tovarlar soni:"),
        ("Возвраты:", "Qaytarishlar:"),
        ("Выручка:", "Tushum:"),
        ("Комиссия Uzum:", "Uzum komissiyasi:"),
        ("Комиссия:", "Komissiya:"),
        ("Логистика:", "Logistika:"),
        ("К выплате всего:", "Jami to‘lovga:"),
        ("К выплате:", "To‘lovga:"),
        ("Уже выведено:", "Allaqachon chiqarilgan:"),
        ("Остаток к выплате:", "To‘lovga qoldi:"),
        ("Статусы:", "Statuslar:"),
        ("Цена продажи:", "Sotuv narxi:"),
        ("ID заказа:", "Buyurtma ID:"),
        ("ID продажи:", "Sotuv ID:"),
        ("Статус:", "Status:"),
        ("Дата:", "Sana:"),
        ("Товар:", "Tovar:"),
        ("Кол-во:", "Soni:"),
        ("Кол-во товаров", "Tovarlar soni"),
        ("Позиции продаж", "Sotuv pozitsiyalari"),
        ("Возвраты", "Qaytarishlar"),
        ("Выручка", "Tushum"),
        ("Логистика", "Logistika"),
        ("Комиссия", "Komissiya"),
        ("К выплате", "To‘lovga"),
        ("Уже выведено", "Allaqachon chiqarilgan"),
        ("Остаток к выплате", "To‘lovga qoldi"),
        ("Позиций продаж", "Sotuv pozitsiyalari"),

        # products / stock
        ("Всего товаров:", "Jami tovarlar:"),
        ("Всего:", "Jami:"),
        ("Остаток:", "Qoldiq:"),
        ("Итого:", "Jami:"),
        ("Разница:", "Farq:"),
        ("Проверить остатки:", "Qoldiqni tekshirish:"),
        ("Уменьшился остаток по SKU", "SKU bo‘yicha qoldiq kamaydi"),
        ("Это может быть продажа, резерв, списание или изменение склада.", "Bu sotuv, rezerv, hisobdan chiqarish yoki ombor o‘zgarishi bo‘lishi mumkin."),
        ("Товары, которые заканчиваются", "Tugayotgan tovarlar"),
        ("Остаток меньше или равен", "Qoldiq kam yoki teng"),
        ("Товар закончился", "Tovar tugagan"),
        ("Потерянные товары", "Yo‘qolgan tovarlar"),
        ("Потеряно:", "Yo‘qolgan:"),
        ("Примерная сумма:", "Taxminiy summa:"),
        ("Продано:", "Sotilgan:"),
        ("Продаж не найдено.", "Sotuvlar topilmadi."),
        ("Не нашёл товаров с остатком и нулевыми продажами.", "Qoldig‘i bor, lekin sotuvi yo‘q tovarlar topilmadi."),
        ("Расчёт примерный", "Hisob-kitob taxminiy"),
        ("хватит примерно на", "taxminan yetadi"),
        ("дней", "kun"),
        ("дня", "kun"),
        ("день", "kun"),
        ("шт.", "dona"),
        ("шт", "dona"),
        ("сум", "so‘m"),

        # invoices / excel
        ("Накладная:", "Yuk xati:"),
        ("Накладная №", "Yuk xati №"),
        ("Создана:", "Yaratilgan:"),
        ("Окно поставки:", "Yetkazib berish oynasi:"),
        ("К поставке:", "Yetkazishga:"),
        ("Принято:", "Qabul qilingan:"),
        ("Сумма:", "Summa:"),
        ("Состав:", "Tarkibi:"),
        ("По накладной:", "Yuk xati bo‘yicha:"),
        ("Расхождение:", "Farq:"),
        ("Закупочная цена:", "Xarid narxi:"),
        ("Excel-отчёт готов", "Excel hisobot tayyor"),
        ("Отчёт готов", "Hisobot tayyor"),

        # explanations / errors
                ("Finance API пока не вернул строки продаж за сегодня. Если в кабинете продажа уже есть, она может появиться здесь позже.",
         "Finance API bugungi sotuvlarni hali qaytarmadi. Agar kabinetda sotuv ko‘rinsa, bu yerda biroz keyin paydo bo‘lishi mumkin."),
        ("за выбранный период", "tanlangan davr uchun"),
        ("за сегодня", "bugun uchun"),
        ("за вчера", "kecha uchun"),
        ("за 7 дней", "7 kun uchun"),
        ("за 30 дней", "30 kun uchun"),
        ("за последние 30 дней", "oxirgi 30 kun uchun"),
        ("Ничего не найдено", "Hech narsa topilmadi"),
        ("Данных нет", "Ma’lumot yo‘q"),
        ("ошибка", "xatolik"),
        ("Ошибка", "Xatolik"),
        ("Попробуйте позже", "Keyinroq urinib ko‘ring"),
    ]
    for old, new in replacements:
        text = text.replace(old, new)
    return text


_ORIGINAL_MESSAGE_ANSWER = Message.answer
_ORIGINAL_BOT_SEND_MESSAGE = Bot.send_message


async def _answer_with_runtime_translation(self: Message, text: Any = None, *args: Any, **kwargs: Any) -> Any:
    try:
        telegram_id = self.from_user.id if self.from_user else None
        if isinstance(text, str) and get_user_language(telegram_id) == "uz":
            text = translate_runtime_text_to_uz(text)
    except Exception:
        pass
    return await _ORIGINAL_MESSAGE_ANSWER(self, text, *args, **kwargs)


async def _send_message_with_runtime_translation(self: Bot, chat_id: Any, text: Any, *args: Any, **kwargs: Any) -> Any:
    try:
        if isinstance(text, str) and get_user_language(int(chat_id)) == "uz":
            text = translate_runtime_text_to_uz(text)
    except Exception:
        pass
    return await _ORIGINAL_BOT_SEND_MESSAGE(self, chat_id, text, *args, **kwargs)


Message.answer = _answer_with_runtime_translation

Bot.send_message = _send_message_with_runtime_translation

# --- Чистка узбекского текста ---
# Первый переводчик выше специально не трогает бизнес-логику, а делает замену текста на лету.
# Здесь финальный слой: убирает смешанные русско-узбекские фразы вроде "Продажи за 30 kun".
_LEGACY_TRANSLATE_RUNTIME_TEXT_TO_UZ = translate_runtime_text_to_uz


def translate_runtime_text_to_uz(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text
    text = _LEGACY_TRANSLATE_RUNTIME_TEXT_TO_UZ(text)

    fixes = [
        # Заголовки продаж — без смеси русского и узбекского
        ("💰 <b>Продажи сегодня</b>", "💰 <b>Bugungi savdo</b>"),
        ("💰 <b>Продажи bugun uchun</b>", "💰 <b>Bugungi savdo</b>"),
        ("💰 <b>Продажи за 7 kun</b>", "💰 <b>7 kunlik savdo</b>"),
        ("💰 <b>Продажи 7 kun uchun</b>", "💰 <b>7 kunlik savdo</b>"),
        ("💰 <b>Продажи за 30 kun</b>", "💰 <b>30 kunlik savdo</b>"),
        ("💰 <b>Продажи 30 kun uchun</b>", "💰 <b>30 kunlik savdo</b>"),
        ("💰 <b>Продажи tanlangan davr uchun</b>", "💰 <b>Tanlangan davr savdosi</b>"),
        ("💰 <b>Продажи kecha uchun</b>", "💰 <b>Kechagi savdo</b>"),
        ("💰 <b>Продажи за выбранный период</b>", "💰 <b>Tanlangan davr savdosi</b>"),

        # Частые поля в финансовых отчётах
        ("Проданных строк/позиций:", "Sotuv pozitsiyalari:"),
        ("Sotilgan строк/позitsiyalar:", "Sotuv pozitsiyalari:"),
        ("Sotilgan qator/pozitsiyalar:", "Sotuv pozitsiyalari:"),
        ("Позиций/строк:", "Sotuv pozitsiyalari:"),
        ("строк:", "qatorlar:"),
        ("Строк:", "Qatorlar:"),
        ("Штук:", "Soni:"),
        ("штук:", "soni:"),
        ("Средняя строка:", "O‘rtacha savdo:"),
        ("O‘rtacha строка:", "O‘rtacha savdo:"),
        ("Отменённых строк:", "Bekor qilinganlar:"),
        ("Отмененных строк:", "Bekor qilinganlar:"),
        ("Bekor qilingan qatorlar:", "Bekor qilinganlar:"),
        ("Топ товаров по so‘mme:", "Summa bo‘yicha top tovarlar:"),
        ("Топ товаров по so‘mме:", "Summa bo‘yicha top tovarlar:"),
        ("Топ товаров по сумме:", "Summa bo‘yicha top tovarlar:"),
        ("Юнит-экономика", "Unit iqtisodiyot"),
        ("Себестоимость", "Tannarx"),
        ("Прибыль", "Foyda"),
        ("Маржа", "Marja"),
        ("Top товаров по so‘mme:", "Summa bo‘yicha top tovarlar:"),
        ("Top товаров по сумме:", "Summa bo‘yicha top tovarlar:"),
        ("Top tovarlar по so‘mme:", "Summa bo‘yicha top tovarlar:"),
        ("Top tovarlar по сумме:", "Summa bo‘yicha top tovarlar:"),
        ("по so‘mme", "summa bo‘yicha"),
        ("по so‘mме", "summa bo‘yicha"),
        ("по сумме", "summa bo‘yicha"),
        ("Без названия", "Nomsiz"),

        # Периоды и служебные фразы
        ("за 30 kun", "30 kun uchun"),
        ("за 7 kun", "7 kun uchun"),
        ("за 1 kun", "1 kun uchun"),
        ("Сегодня", "Bugun"),
        ("Вчера", "Kecha"),
        ("7 дней", "7 kun"),
        ("30 дней", "30 kun"),
        ("Ответ Finance API пришёл, но строки продаж не найдены.", "Finance API javob berdi, lekin savdo qatorlari topilmadi."),
        ("Фрагмент ответа:", "Javobdan parcha:"),
        ("Подробно:", "Batafsil:"),
        ("Показаны первые", "Birinchi"),
        ("позиций из", "pozitsiya ko‘rsatildi, jami"),

        # Остатки, накладные, общие поля
        ("Товары без продаж", "Sotilmayotgan tovarlar"),
        ("Не продаётся", "Sotilmayapti"),
        ("Прогноз остатков", "Qoldiq prognozi"),
        ("Все магазины", "Barcha do‘konlar"),
        ("Накладные FBO", "FBO yuk xatlari"),
        ("Состав накладной", "Yuk xati tarkibi"),
        ("Потерянные", "Yo‘qolganlar"),
        ("Заканчивается", "Tugayapti"),
        ("Заканчиваются", "Tugayapti"),
        ("Остатки", "Qoldiq"),
        ("Остаток", "Qoldiq"),
        ("Продано", "Sotilgan"),
        ("Возврат", "Qaytarilgan"),
        ("Возвраты", "Qaytarilganlar"),
        ("Комиссия Uzum", "Uzum komissiyasi"),
        ("Комиссия", "Komissiya"),
        ("Логистика", "Logistika"),
        ("Выручка", "Tushum"),
        ("К выплате всего", "Jami to‘lovga"),
        ("К выплате", "To‘lovga"),
        ("Уже выведено", "Chiqarilgan"),
        ("Остаток к выплате", "To‘lovga qoldi"),
        ("Статусы", "Statuslar"),
        ("Статус", "Status"),
        ("Магазин", "Do‘kon"),
        ("Товар", "Tovar"),
        ("Дата", "Sana"),
        ("Цена продажи", "Sotuv narxi"),
        ("ID заказа", "Buyurtma ID"),
        ("ID продажи", "Sotuv ID"),
        ("Кол-во", "Soni"),
        ("Итого", "Jami"),
        ("Сумма", "Summa"),
        ("Разница", "Farq"),
        ("Расхождение", "Farq"),
        ("Принято", "Qabul qilingan"),
        ("К поставке", "Yetkazishga"),
        ("Накладная", "Yuk xati"),
        ("Отзывы", "Sharhlar"),
        ("Последние отзывы", "Oxirgi sharhlar"),
        ("Новый отзыв", "Yangi sharh"),
        ("Отзыв", "Sharh"),
        ("Оценка", "Baho"),
        ("Клиент", "Mijoz"),
        ("Ответ продавца", "Sotuvchi javobi"),
    ]
    for old, new in fixes:
        text = text.replace(old, new)

    # Безопасная замена валюты/единиц: только как отдельные слова в суммах.
    import re
    text = re.sub(r"(?<=\d)\s*сум\b", " so‘m", text)
    text = re.sub(r"(?<=\d)\s*шт\.?\b", " dona", text)
    return text



# --- Дополнительная чистка узбекского перевода: магазины и уведомления ---
_CLEAN3_TRANSLATE_RUNTIME_TEXT_TO_UZ = translate_runtime_text_to_uz


def translate_runtime_text_to_uz(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text
    text = _CLEAN3_TRANSLATE_RUNTIME_TEXT_TO_UZ(text)

    fixes = [
        # Магазины
        ("🏪 <b>Ваши do‘konlari:</b>", "🏪 <b>Do‘konlaringiz:</b>"),
        ("🏪 <b>Ваши do‘konlar:</b>", "🏪 <b>Do‘konlaringiz:</b>"),
        ("🏪 <b>Ваши магазины:</b>", "🏪 <b>Do‘konlaringiz:</b>"),
        ("Текущий основной Do‘kon:", "Joriy asosiy do‘kon:"),
        ("Текущий основной магазин:", "Joriy asosiy do‘kon:"),
        ("Joriy основной do‘kon:", "Joriy asosiy do‘kon:"),
        ("Чтобы выбрать:", "Tanlash uchun:"),
        ("не выбран", "tanlanmagan"),

        # Уведомления: заголовки и статусы
        ("💸 <b>Уведомления о новых продажах</b>", "💸 <b>Yangi savdolar xabarnomalari</b>"),
        ("💸 <b>Yangi продажа xabarnomalari</b>", "💸 <b>Yangi savdolar xabarnomalari</b>"),
        ("🔔 <b>Уведомления</b>", "🔔 <b>Xabarnomalar</b>"),
        ("🔔 <b>Уведомления:</b>", "🔔 <b>Xabarnomalar:</b>"),
        ("Уведомления о новых продажах", "Yangi savdolar xabarnomalari"),
        ("Holat: ✅ включены", "Holat: ✅ yoqilgan"),
        ("Holat: ❌ выключены", "Holat: ❌ o‘chirilgan"),
        ("Status: ✅ включены", "Holat: ✅ yoqilgan"),
        ("Status: ❌ выключены", "Holat: ❌ o‘chirilgan"),
        ("Статус: ✅ включены", "Holat: ✅ yoqilgan"),
        ("Статус: ❌ выключены", "Holat: ❌ o‘chirilgan"),
        ("Проверка каждые:", "Tekshiruv har:"),
        ("Tekshiruv har: <b>", "Tekshiruv har <b>"),
        ("сек.", "soniya"),
        ("Состояние: продажи уже запомнены", "Holat: savdolar allaqachon eslab qolingan"),
        ("Состояние: инициализация при следующей проверке", "Holat: keyingi tekshiruvda ishga tushadi"),
        ("Holat: продажи уже запомнены", "Holat: savdolar allaqachon eslab qolingan"),
        ("Holat: инициализация при следующей проверке", "Holat: keyingi tekshiruvda ishga tushadi"),
        ("Бот смотрит Finance API bugun uchun.", "Bot bugungi savdolarni Finance API orqali tekshiradi."),
        ("Бот смотрит Finance API за сегодня.", "Bot bugungi savdolarni Finance API orqali tekshiradi."),
        ("Если Finance API отдаёт продажу с задержкой, уведомление тоже придёт с задержкой.", "Agar Finance API savdoni kechikib bersa, xabarnoma ham kechikib keladi."),
        ("Bot bugungi savdolarni Finance API orqali tekshiradi. Agar Finance API savdoni kechikib bersa, xabarnoma ham kechikib keladi.", "Bot bugungi savdolarni Finance API orqali tekshiradi. Agar savdo kechikib ko‘rinsa, xabarnoma ham biroz kechikib kelishi mumkin."),
        ("Do‘kon:", "Do‘kon:"),
    ]
    for old, new in fixes:
        text = text.replace(old, new)
    return text

def language_markup() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="🇷🇺 Русский", callback_data="set_lang:ru"),
                InlineKeyboardButton(text="🇺🇿 O‘zbekcha", callback_data="set_lang:uz"),
            ]
        ]
    )


MAIN_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔌 Подключить магазин")],
        [KeyboardButton(text="🎥 Как подключить")],
        [KeyboardButton(text="🌐 Язык"), KeyboardButton(text="ℹ️ Помощь")],
        [KeyboardButton(text="💎 Подписка")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Сначала подключите магазин",
)

MAIN_MENU_RU_ADMIN = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔌 Подключить магазин")],
        [KeyboardButton(text="🎥 Как подключить")],
        [KeyboardButton(text="🌐 Язык"), KeyboardButton(text="ℹ️ Помощь")],
        [KeyboardButton(text="💎 Подписка"), KeyboardButton(text="👑 Админ")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Сначала подключите магазин",
)

MAIN_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔌 Do‘konni ulash")],
        [KeyboardButton(text="🎥 Qanday ulash kerak")],
        [KeyboardButton(text="🌐 Til"), KeyboardButton(text="ℹ️ Yordam")],
        [KeyboardButton(text="💎 Obuna")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Avval do‘konni ulang",
)

MAIN_MENU_UZ_ADMIN = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔌 Do‘konni ulash")],
        [KeyboardButton(text="🎥 Qanday ulash kerak")],
        [KeyboardButton(text="🌐 Til"), KeyboardButton(text="ℹ️ Yordam")],
        [KeyboardButton(text="💎 Obuna"), KeyboardButton(text="👑 Admin")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Avval do‘konni ulang",
)

# Главное меню после подключения API: простая структура по разделам.
MAIN_MENU_RU_CONNECTED = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🏠 Обзор магазина")],
        [KeyboardButton(text="💰 Продажи"), KeyboardButton(text="📦 Склад")],
        [KeyboardButton(text="🚨 Важно сейчас"), KeyboardButton(text="📊 Отчёты")],
        [KeyboardButton(text="⚙️ Настройки")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Выберите раздел",
)

MAIN_MENU_RU_CONNECTED_ADMIN = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🏠 Обзор магазина")],
        [KeyboardButton(text="💰 Продажи"), KeyboardButton(text="📦 Склад")],
        [KeyboardButton(text="🚨 Важно сейчас"), KeyboardButton(text="📊 Отчёты")],
        [KeyboardButton(text="⚙️ Настройки"), KeyboardButton(text="👑 Админ")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Выберите раздел",
)

MAIN_MENU_UZ_CONNECTED = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🏠 Do‘kon holati")],
        [KeyboardButton(text="💰 Savdo"), KeyboardButton(text="📦 Ombor")],
        [KeyboardButton(text="🚨 Hozir muhim"), KeyboardButton(text="📊 Hisobotlar")],
        [KeyboardButton(text="⚙️ Sozlamalar")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Bo‘limni tanlang",
)

MAIN_MENU_UZ_CONNECTED_ADMIN = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🏠 Do‘kon holati")],
        [KeyboardButton(text="💰 Savdo"), KeyboardButton(text="📦 Ombor")],
        [KeyboardButton(text="🚨 Hozir muhim"), KeyboardButton(text="📊 Hisobotlar")],
        [KeyboardButton(text="⚙️ Sozlamalar"), KeyboardButton(text="👑 Admin")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Bo‘limni tanlang",
)

# Trial menu deliberately exposes only the value promised for the free period.
# Paid sections remain discoverable through one clear "full version" button.
TRIAL_MAIN_MENU_RU_CONNECTED = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="💰 Продажи"), KeyboardButton(text="🌙 Утренний отчёт")],
        [KeyboardButton(text="💎 Полная версия")],
        [KeyboardButton(text="⚙️ Настройки")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Пробный период: продажи и утренний отчёт",
)

TRIAL_MAIN_MENU_UZ_CONNECTED = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="💰 Savdo"), KeyboardButton(text="🌙 Ertalabki hisobot")],
        [KeyboardButton(text="💎 To‘liq versiya")],
        [KeyboardButton(text="⚙️ Sozlamalar")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Sinov: savdo va ertalabki hisobot",
)

TRIAL_SALES_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📊 Сегодня")],
        [KeyboardButton(text="💸 Уведомления о продажах")],
        [KeyboardButton(text="🌙 Утренний отчёт")],
        [KeyboardButton(text="💎 Полная версия"), KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Продажи в пробном периоде",
)

TRIAL_SALES_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📊 Bugun")],
        [KeyboardButton(text="💸 Savdo xabarlari")],
        [KeyboardButton(text="🌙 Ertalabki hisobot")],
        [KeyboardButton(text="💎 To‘liq versiya"), KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Sinov davridagi savdo",
)

TRIAL_SETTINGS_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="💸 Уведомления о продажах")],
        [KeyboardButton(text="🏪 Магазины"), KeyboardButton(text="🔐 Подключение Uzum")],
        [KeyboardButton(text="🌐 Язык"), KeyboardButton(text="ℹ️ Помощь")],
        [KeyboardButton(text="💎 Полная версия"), KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Настройки пробного периода",
)

TRIAL_SETTINGS_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="💸 Savdo xabarlari")],
        [KeyboardButton(text="🏪 Do‘konlar"), KeyboardButton(text="🔐 Uzum ulanishi")],
        [KeyboardButton(text="🌐 Til"), KeyboardButton(text="ℹ️ Yordam")],
        [KeyboardButton(text="💎 To‘liq versiya"), KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Sinov davri sozlamalari",
)

SALES_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📊 Сегодня"), KeyboardButton(text="📆 Вчера")],
        [KeyboardButton(text="🗓 7 дней"), KeyboardButton(text="📅 30 дней")],
        [KeyboardButton(text="💰 Прибыль"), KeyboardButton(text="🏆 Топ товаров")],
        [KeyboardButton(text="✨ Ещё по продажам")],
        [KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Продажи",
)

SALES_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📊 Bugun"), KeyboardButton(text="📆 Kecha")],
        [KeyboardButton(text="🗓 7 kun"), KeyboardButton(text="📅 30 kun")],
        [KeyboardButton(text="💰 Foyda"), KeyboardButton(text="🏆 Top tovarlar")],
        [KeyboardButton(text="✨ Yana savdo tahlili")],
        [KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Savdo",
)

SALES_MORE_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📊 Бизнес-сводка"), KeyboardButton(text="🌐 Все магазины")],
        [KeyboardButton(text="🐢 Не продаётся"), KeyboardButton(text="🧾 Юнит-экономика")],
        [KeyboardButton(text="🔄 Себестоимость Uzum"), KeyboardButton(text="🧮 Расходы и налоги")],
        [KeyboardButton(text="⬅️ Продажи"), KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Дополнительная аналитика",
)

SALES_MORE_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📊 Biznes xulosa"), KeyboardButton(text="🌐 Barcha do‘konlar")],
        [KeyboardButton(text="🐢 Sotilmayapti"), KeyboardButton(text="🧾 Unit iqtisodiyot")],
        [KeyboardButton(text="🔄 Uzum tannarxi"), KeyboardButton(text="🧮 Xarajat va soliq")],
        [KeyboardButton(text="⬅️ Savdo"), KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Qo‘shimcha tahlil",
)

STOCK_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📦 Все остатки"), KeyboardButton(text="🚚 Что заказать")],
        [KeyboardButton(text="🧭 Потери и брак"), KeyboardButton(text="🏷 Этикетки SKU")],
        [KeyboardButton(text="📑 Документы по потерям")],
        [KeyboardButton(text="🧾 ИКПУ / МХИК")],
        [KeyboardButton(text="⚙️ Срок поставки")],
        [KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Склад",
)

STOCK_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📦 Barcha qoldiq"), KeyboardButton(text="🚚 Nima buyurtma qilish")],
        [KeyboardButton(text="🧭 Yo‘qotish va brak"), KeyboardButton(text="🏷 SKU etiketkalari")],
        [KeyboardButton(text="📑 Yo‘qotish hujjatlari")],
        [KeyboardButton(text="🧾 IKPU / MXIK")],
        [KeyboardButton(text="⚙️ Yetkazish muddati")],
        [KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Ombor",
)

NOTIFY_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="⚙️ Настроить уведомления")],
        [KeyboardButton(text="⬅️ Настройки"), KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Уведомления",
)

NOTIFY_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="⚙️ Xabarnomalarni sozlash")],
        [KeyboardButton(text="⬅️ Sozlamalar"), KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Xabarnomalar",
)

REPORT_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📋 Дневной отчёт")],
        [KeyboardButton(text="📄 PDF-отчёт"), KeyboardButton(text="📊 Excel-отчёт")],
        [KeyboardButton(text="🌙 Краткий отчёт"), KeyboardButton(text="💰 Прибыль")],
        [KeyboardButton(text="📈 Эффект рекомендаций"), KeyboardButton(text="📅 Автоотчёты")],
        [KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Отчёты",
)

REPORT_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📋 Kunlik hisobot")],
        [KeyboardButton(text="📄 PDF hisobot"), KeyboardButton(text="📊 Excel hisobot")],
        [KeyboardButton(text="🌙 Qisqa hisobot"), KeyboardButton(text="💰 Foyda")],
        [KeyboardButton(text="📈 Tavsiyalar ta’siri"), KeyboardButton(text="📅 Avtohisobotlar")],
        [KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Hisobotlar",
)

SETTINGS_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔔 Уведомления")],
        [KeyboardButton(text="🏪 Магазины"), KeyboardButton(text="🌐 Язык")],
        [KeyboardButton(text="💰 Себестоимость и расходы"), KeyboardButton(text="🚚 Настройки поставки")],
        [KeyboardButton(text="🔐 API и подключение"), KeyboardButton(text="🌐 Веб-кабинет")],
        [KeyboardButton(text="ℹ️ Помощь"), KeyboardButton(text="💎 Подписка")],
        [KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Настройки",
)

SETTINGS_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔔 Xabarnomalar")],
        [KeyboardButton(text="🏪 Do‘konlar"), KeyboardButton(text="🌐 Til")],
        [KeyboardButton(text="💰 Tannarx va xarajat"), KeyboardButton(text="🚚 Yetkazish sozlamalari")],
        [KeyboardButton(text="🔐 API va ulanish"), KeyboardButton(text="🌐 Veb-kabinet")],
        [KeyboardButton(text="ℹ️ Yordam"), KeyboardButton(text="💎 Obuna")],
        [KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Sozlamalar",
)

FINANCE_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔄 Себестоимость Uzum"), KeyboardButton(text="🧾 Расходы Uzum")],
        [KeyboardButton(text="🧮 Расходы и налоги"), KeyboardButton(text="💰 Прибыль")],
        [KeyboardButton(text="🧾 Юнит-экономика")],
        [KeyboardButton(text="⬅️ Настройки"), KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Финансы",
)

FINANCE_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔄 Uzum tannarxi"), KeyboardButton(text="🧾 Uzum xarajatlari")],
        [KeyboardButton(text="🧮 Xarajat va soliq"), KeyboardButton(text="💰 Foyda")],
        [KeyboardButton(text="🧾 Unit iqtisodiyot")],
        [KeyboardButton(text="⬅️ Sozlamalar"), KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Moliya",
)

CONNECTION_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="✅ Проверить подключение")],
        [KeyboardButton(text="🔌 Обновить API-ключ"), KeyboardButton(text="🔐 Безопасность")],
        [KeyboardButton(text="⬅️ Настройки"), KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Подключение Uzum",
)

CONNECTION_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="✅ Ulanishni tekshirish")],
        [KeyboardButton(text="🔌 API-kalitni yangilash"), KeyboardButton(text="🔐 Xavfsizlik")],
        [KeyboardButton(text="⬅️ Sozlamalar"), KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Uzum ulanishi",
)

CONNECT_INPUT_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="❌ Отмена")]],
    resize_keyboard=True,
    input_field_placeholder="Вставьте API-ключ сюда",
)

CONNECT_INPUT_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="❌ Bekor qilish")]],
    resize_keyboard=True,
    input_field_placeholder="API-kalitni shu yerga kiriting",
)

ATTENTION_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔍 Проверить магазин")],
        [KeyboardButton(text="⚠️ Остатки"), KeyboardButton(text="🐢 Без продаж")],
        [KeyboardButton(text="🧾 Нет себестоимости"), KeyboardButton(text="📉 Низкая прибыль")],
        [KeyboardButton(text="❌ Отмены")],
        [KeyboardButton(text="🏠 Главное")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Что проверить",
)

ATTENTION_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔍 Do‘konni tekshirish")],
        [KeyboardButton(text="⚠️ Qoldiqlar"), KeyboardButton(text="🐢 Sotuv yo‘q")],
        [KeyboardButton(text="🧾 Tannarx yo‘q"), KeyboardButton(text="📉 Past foyda")],
        [KeyboardButton(text="❌ Bekor qilishlar")],
        [KeyboardButton(text="🏠 Asosiy")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Nimani tekshiramiz",
)

ADMIN_PANEL_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="👥 Пользователи"), KeyboardButton(text="🔑 Подключение API")],
        [KeyboardButton(text="💳 Оплаты")],
        [KeyboardButton(text="⏳ Скоро заканчиваются"), KeyboardButton(text="⛔ Заблокированные")],
        [KeyboardButton(text="✅ Проверить подключение"), KeyboardButton(text="🎥 Видеоинструкция")],
        [KeyboardButton(text="📦 Бэкап базы")],
        [KeyboardButton(text="📢 Рассылка"), KeyboardButton(text="⬅️ Главное меню")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Админ-панель",
)

ADMIN_PANEL_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="👥 Foydalanuvchilar"), KeyboardButton(text="🔑 API ulanishi")],
        [KeyboardButton(text="💳 To‘lovlar")],
        [KeyboardButton(text="⏳ Tugayotganlar"), KeyboardButton(text="⛔ Bloklanganlar")],
        [KeyboardButton(text="✅ Ulanishni tekshirish"), KeyboardButton(text="🎥 API ulash videosi")],
        [KeyboardButton(text="📦 Baza zaxirasi")],
        [KeyboardButton(text="📢 Xabar yuborish"), KeyboardButton(text="⬅️ Asosiy menyu")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Admin panel",
)

ADMIN_PANEL_MANAGER_MENU_RU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="👥 Пользователи"), KeyboardButton(text="🔑 Подключение API")],
        [KeyboardButton(text="💳 Оплаты")],
        [KeyboardButton(text="⏳ Скоро заканчиваются"), KeyboardButton(text="⛔ Заблокированные")],
        [KeyboardButton(text="✅ Проверить подключение"), KeyboardButton(text="🎥 Видеоинструкция")],
        [KeyboardButton(text="📢 Рассылка"), KeyboardButton(text="⬅️ Главное меню")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Админ-панель",
)

ADMIN_PANEL_MANAGER_MENU_UZ = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="👥 Foydalanuvchilar"), KeyboardButton(text="🔑 API ulanishi")],
        [KeyboardButton(text="💳 To‘lovlar")],
        [KeyboardButton(text="⏳ Tugayotganlar"), KeyboardButton(text="⛔ Bloklanganlar")],
        [KeyboardButton(text="✅ Ulanishni tekshirish"), KeyboardButton(text="🎥 API ulash videosi")],
        [KeyboardButton(text="📢 Xabar yuborish"), KeyboardButton(text="⬅️ Asosiy menyu")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Admin panel",
)

# Для совместимости: если где-то осталась статичная разметка, будет русский вариант.
MAIN_MENU = MAIN_MENU_RU
ADMIN_PANEL_MENU = ADMIN_PANEL_MENU_RU
ANALYTICS_MENU = MAIN_MENU_RU
PRODUCTS_MENU = MAIN_MENU_RU
ORDERS_MENU = MAIN_MENU_RU
NOTIFICATIONS_MENU = MAIN_MENU_RU


def _user_has_uzum_connection(telegram_id: int | None) -> bool:
    if not telegram_id:
        return False
    try:
        if hasattr(db, "has_uzum_connection"):
            return bool(db.has_uzum_connection(int(telegram_id)))
        user = db.get_user(int(telegram_id))
        return bool(user and user["uzum_token_encrypted"])
    except Exception:
        return False


def main_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    lang = get_user_language(telegram_id)
    admin = is_admin(telegram_id)
    connected = _user_has_uzum_connection(telegram_id)
    trial = bool(
        telegram_id is not None
        and subscription_access_level(int(telegram_id)) == "trial"
    )

    if lang == "uz":
        if connected:
            if trial:
                return TRIAL_MAIN_MENU_UZ_CONNECTED
            return MAIN_MENU_UZ_CONNECTED_ADMIN if admin else MAIN_MENU_UZ_CONNECTED
        return MAIN_MENU_UZ_ADMIN if admin else MAIN_MENU_UZ

    if connected:
        if trial:
            return TRIAL_MAIN_MENU_RU_CONNECTED
        return MAIN_MENU_RU_CONNECTED_ADMIN if admin else MAIN_MENU_RU_CONNECTED
    return MAIN_MENU_RU_ADMIN if admin else MAIN_MENU_RU


def sales_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    if telegram_id is not None and subscription_access_level(int(telegram_id)) == "trial":
        return TRIAL_SALES_MENU_UZ if get_user_language(telegram_id) == "uz" else TRIAL_SALES_MENU_RU
    return SALES_MENU_UZ if get_user_language(telegram_id) == "uz" else SALES_MENU_RU


def sales_more_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    return SALES_MORE_MENU_UZ if get_user_language(telegram_id) == "uz" else SALES_MORE_MENU_RU


def stock_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    return STOCK_MENU_UZ if get_user_language(telegram_id) == "uz" else STOCK_MENU_RU


def notify_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    return NOTIFY_MENU_UZ if get_user_language(telegram_id) == "uz" else NOTIFY_MENU_RU


def report_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    return REPORT_MENU_UZ if get_user_language(telegram_id) == "uz" else REPORT_MENU_RU


def settings_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    if telegram_id is not None and subscription_access_level(int(telegram_id)) == "trial":
        return TRIAL_SETTINGS_MENU_UZ if get_user_language(telegram_id) == "uz" else TRIAL_SETTINGS_MENU_RU
    return SETTINGS_MENU_UZ if get_user_language(telegram_id) == "uz" else SETTINGS_MENU_RU


def finance_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    return FINANCE_MENU_UZ if get_user_language(telegram_id) == "uz" else FINANCE_MENU_RU


def connection_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    return CONNECTION_MENU_UZ if get_user_language(telegram_id) == "uz" else CONNECTION_MENU_RU


def attention_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    return ATTENTION_MENU_UZ if get_user_language(telegram_id) == "uz" else ATTENTION_MENU_RU


def _message_user_id(message: Message) -> int | None:
    try:
        return message.from_user.id if message.from_user else None
    except Exception:
        return None


def sales_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return sales_menu_for_user(_message_user_id(message))


def sales_more_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return sales_more_menu_for_user(_message_user_id(message))


def stock_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return stock_menu_for_user(_message_user_id(message))


def notify_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return notify_menu_for_user(_message_user_id(message))


def report_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return report_menu_for_user(_message_user_id(message))


def settings_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return settings_menu_for_user(_message_user_id(message))


def finance_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return finance_menu_for_user(_message_user_id(message))


def connection_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return connection_menu_for_user(_message_user_id(message))


def attention_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    return attention_menu_for_user(_message_user_id(message))


def menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    try:
        telegram_id = message.from_user.id if message.from_user else None
    except Exception:
        telegram_id = None
    return main_menu_for_user(telegram_id)


def admin_menu_for_user(telegram_id: int | None) -> ReplyKeyboardMarkup:
    lang = get_user_language(telegram_id)
    if is_owner(telegram_id):
        return ADMIN_PANEL_MENU_UZ if lang == "uz" else ADMIN_PANEL_MENU_RU
    return ADMIN_PANEL_MANAGER_MENU_UZ if lang == "uz" else ADMIN_PANEL_MANAGER_MENU_RU


def admin_menu_for_message(message: Message) -> ReplyKeyboardMarkup:
    try:
        telegram_id = message.from_user.id if message.from_user else None
    except Exception:
        telegram_id = None
    return admin_menu_for_user(telegram_id)


# --- Универсальная постраничность для длинных списков ---
LIST_PAGE_SIZE = max(
    5,
    min(20, int(os.getenv("LIST_PAGE_SIZE", "10") or "10")),
)
PAGED_LIST_TTL_SECONDS = max(
    300,
    int(os.getenv("PAGED_LIST_TTL_SECONDS", "1800") or "1800"),
)
PAGED_LIST_MAX_SESSIONS = max(
    20,
    int(os.getenv("PAGED_LIST_MAX_SESSIONS", "200") or "200"),
)
_paged_list_cache: dict[tuple[int, str], dict[str, Any]] = {}


def _page_size_safe(value: Any = None) -> int:
    try:
        raw = int(value or LIST_PAGE_SIZE)
    except (TypeError, ValueError):
        raw = LIST_PAGE_SIZE
    return max(5, min(20, raw))


def _paged_kind_key(value: Any) -> str:
    safe = "".join(
        char
        for char in str(value or "list").lower()
        if char.isascii() and (char.isalnum() or char in {"_", "-"})
    )
    return safe[:24] or "list"


def _prune_paged_list_cache() -> None:
    now = time.monotonic()
    stale = [
        key
        for key, session in _paged_list_cache.items()
        if now - float(session.get("created_at") or 0) > PAGED_LIST_TTL_SECONDS
    ]
    for key in stale:
        _paged_list_cache.pop(key, None)

    overflow = len(_paged_list_cache) - PAGED_LIST_MAX_SESSIONS
    if overflow <= 0:
        return
    oldest = sorted(
        _paged_list_cache,
        key=lambda key: float(_paged_list_cache[key].get("created_at") or 0),
    )
    for key in oldest[:overflow]:
        _paged_list_cache.pop(key, None)


def _section_text_and_markup(
    section: str,
    telegram_id: int,
    lang: str,
) -> tuple[str, ReplyKeyboardMarkup]:
    if section == "sales":
        text = (
            "💰 <b>Savdo</b>\nKerakli davr yoki hisobotni tanlang 👇"
            if lang == "uz"
            else "💰 <b>Продажи</b>\nВыберите период или отчёт 👇"
        )
        return text, sales_menu_for_user(telegram_id)
    if section == "stock":
        text = (
            "📦 <b>Ombor</b>\nQoldiq, prognoz yoki yo‘qotishlarni tanlang 👇"
            if lang == "uz"
            else "📦 <b>Склад</b>\nВыберите остатки, прогноз или потери 👇"
        )
        return text, stock_menu_for_user(telegram_id)
    if section == "attention":
        text = (
            "🚨 <b>Hozir muhim</b>\nKerakli bo‘limni tanlang 👇"
            if lang == "uz"
            else "🚨 <b>Важно сейчас</b>\nВыберите нужный раздел 👇"
        )
        return text, attention_menu_for_user(telegram_id)
    if section == "reports":
        text = (
            "📊 <b>Hisobotlar</b>\nPDF, Excel, foyda va tayyor hisobotlar 👇"
            if lang == "uz"
            else "📊 <b>Отчёты</b>\nPDF, Excel, прибыль и готовые отчёты 👇"
        )
        return text, report_menu_for_user(telegram_id)
    text = "🏠 <b>Asosiy menyu</b>" if lang == "uz" else "🏠 <b>Главное меню</b>"
    return text, main_menu_for_user(telegram_id)


def _paged_markup(
    kind: str,
    page: int,
    total_pages: int,
    lang: str,
    section: str = "main",
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    if total_pages > 1:
        navigation: list[InlineKeyboardButton] = []
        if page > 0:
            navigation.append(
                InlineKeyboardButton(
                    text="⬅️ Oldingi" if lang == "uz" else "⬅️ Назад",
                    callback_data=f"pglist:{kind}:{page - 1}",
                )
            )
        navigation.append(
            InlineKeyboardButton(
                text=f"{page + 1}/{total_pages}",
                callback_data="pgnoop",
            )
        )
        if page < total_pages - 1:
            navigation.append(
                InlineKeyboardButton(
                    text="Keyingi ➡️" if lang == "uz" else "Далее ➡️",
                    callback_data=f"pglist:{kind}:{page + 1}",
                )
            )
        rows.append(navigation)
    rows.append(
        [
            InlineKeyboardButton(
                text="⬅️ Bo‘limga" if lang == "uz" else "⬅️ В раздел",
                callback_data=f"pgsection:{section}",
            ),
            InlineKeyboardButton(
                text="🏠 Menyu" if lang == "uz" else "🏠 Меню",
                callback_data="pgmain",
            ),
        ]
    )
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _paged_text(session: dict[str, Any], page: int) -> str:
    items = [str(item) for item in (session.get("items") or [])]
    page_size = _page_size_safe(session.get("page_size"))
    total = len(items)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = max(0, min(int(page), total_pages - 1))
    start_index = page * page_size
    chunk = items[start_index:start_index + page_size]
    start_number = start_index + 1 if total else 0
    end_number = start_index + len(chunk)
    lang = normalize_lang(session.get("lang"))
    if lang == "uz":
        meta = (
            f"Ko‘rsatilmoqda: <b>{start_number}–{end_number}</b> / <b>{total}</b>\n"
            f"Sahifa: <b>{page + 1}/{total_pages}</b>"
        )
    else:
        meta = (
            f"Показано: <b>{start_number}–{end_number}</b> из <b>{total}</b>\n"
            f"Страница: <b>{page + 1}/{total_pages}</b>"
        )
    parts = [
        str(session.get("title") or ""),
        *[str(value) for value in (session.get("summary") or []) if value],
        meta,
    ]
    if chunk:
        parts.append("\n\n".join(chunk))
    return "\n\n".join(part for part in parts if part)


async def send_paginated_list(
    message: Message,
    *,
    kind: str,
    title: str,
    items: list[str],
    summary: list[str] | None = None,
    empty_text: str | None = None,
    section: str = "main",
    page_size: int | None = None,
    reply_markup: ReplyKeyboardMarkup | None = None,
) -> None:
    user_id = message.from_user.id if message.from_user else 0
    lang = get_user_language(user_id)
    if not items:
        await message.answer(
            empty_text
            or ("Ma’lumot topilmadi." if lang == "uz" else "Данные не найдены."),
            reply_markup=reply_markup or menu_for_message(message),
        )
        return

    _prune_paged_list_cache()
    safe_kind = _paged_kind_key(kind)
    safe_section = section if section in {"sales", "stock", "attention", "reports"} else "main"
    session = {
        "title": title,
        "summary": list(summary or []),
        "items": list(items),
        "lang": lang,
        "section": safe_section,
        "page_size": _page_size_safe(page_size),
        "created_at": time.monotonic(),
    }
    _paged_list_cache[(int(user_id), safe_kind)] = session
    total_pages = max(
        1,
        (len(items) + int(session["page_size"]) - 1) // int(session["page_size"]),
    )
    await message.answer(
        _paged_text(session, 0),
        reply_markup=_paged_markup(
            safe_kind,
            0,
            total_pages,
            lang,
            safe_section,
        ),
    )


@dp.callback_query(F.data == "pgnoop")
async def paged_noop_callback(callback: CallbackQuery) -> None:
    await callback.answer()


@dp.callback_query(F.data == "pgmain")
async def paged_main_callback(callback: CallbackQuery) -> None:
    user_id = callback.from_user.id if callback.from_user else 0
    lang = get_user_language(user_id)
    text, markup = _section_text_and_markup("main", user_id, lang)
    if callback.message:
        await callback.message.answer(text, reply_markup=markup)
    await callback.answer()


@dp.callback_query(F.data.startswith("pgsection:"))
async def paged_section_callback(callback: CallbackQuery) -> None:
    user_id = callback.from_user.id if callback.from_user else 0
    lang = get_user_language(user_id)
    section = str(callback.data or "").split(":", 1)[-1]
    text, markup = _section_text_and_markup(section, user_id, lang)
    if callback.message:
        await callback.message.answer(text, reply_markup=markup)
    await callback.answer()


@dp.callback_query(F.data.startswith("pglist:"))
async def paged_list_callback(callback: CallbackQuery) -> None:
    user_id = callback.from_user.id if callback.from_user else 0
    lang = get_user_language(user_id)
    if not has_paid_subscription(user_id):
        if subscription_access_level(user_id) == "trial":
            await send_trial_premium_locked(callback, user_id)
        else:
            await callback.answer(
                "Obuna tugagan" if lang == "uz" else "Подписка закончилась",
                show_alert=True,
            )
        return
    try:
        _, raw_kind, raw_page = str(callback.data or "").split(":", 2)
        kind = _paged_kind_key(raw_kind)
        page = int(raw_page)
    except (TypeError, ValueError):
        await callback.answer()
        return

    _prune_paged_list_cache()
    session = _paged_list_cache.get((int(user_id), kind))
    if not session:
        await callback.answer(
            "Ro‘yxat eskirdi. Bo‘limni qayta oching."
            if lang == "uz"
            else "Список устарел. Откройте раздел заново.",
            show_alert=True,
        )
        return

    page_size = _page_size_safe(session.get("page_size"))
    total_pages = max(
        1,
        (len(session.get("items") or []) + page_size - 1) // page_size,
    )
    page = max(0, min(page, total_pages - 1))
    try:
        if callback.message:
            await callback.message.edit_text(
                _paged_text(session, page),
                reply_markup=_paged_markup(
                    kind,
                    page,
                    total_pages,
                    str(session.get("lang") or lang),
                    str(session.get("section") or "main"),
                ),
            )
        await callback.answer()
    except Exception:
        logging.exception(
            "Paged list callback failed user=%s kind=%s page=%s",
            user_id,
            kind,
            page,
        )
        await callback.answer(
            "Sahifani ochib bo‘lmadi"
            if lang == "uz"
            else "Не получилось открыть страницу",
            show_alert=True,
        )


# Переопределяем тексты подписки после инициализации языка, чтобы /my_subscription был на выбранном языке.
def subscription_status_text(telegram_id: int) -> str:
    row = ensure_subscription(telegram_id)
    lang = get_user_language(telegram_id)
    if is_admin(telegram_id):
        return "👑 Admin kirish: cheklovsiz" if lang == "uz" else "👑 Админ-доступ: без ограничений"
    if int(row.get("blocked") or 0) == 1:
        return "⛔ Foydalanuvchi bloklangan" if lang == "uz" else "⛔ Пользователь заблокирован"
    now = _utc_now()
    trial_until = _dt_from_db(row.get("trial_until"))
    paid_until = _dt_from_db(row.get("subscription_until"))
    until = subscription_active_until(row)
    if until and until > now:
        if paid_until and paid_until == until:
            return (f"✅ Obuna {_fmt_dt(paid_until)} gacha faol" if lang == "uz" else f"✅ Подписка активна до {_fmt_dt(paid_until)}")
        return (f"🎁 Trial {_fmt_dt(trial_until)} gacha faol" if lang == "uz" else f"🎁 Trial активен до {_fmt_dt(trial_until)}")
    return "⛔ Obuna muddati tugagan" if lang == "uz" else "⛔ Подписка закончилась"


def subscription_full_text(telegram_id: int) -> str:
    row = ensure_subscription(telegram_id)
    status = subscription_status_text(telegram_id)
    lang = get_user_language(telegram_id)

    if is_admin(telegram_id):
        if lang == "uz":
            return (
                "💎 <b>Mening obunam</b>\n\n"
                f"Telegram ID: <code>{telegram_id}</code>\n"
                f"Holat: {status}\n\n"
                "Admin uchun trial va to‘lov sanasi muhim emas — kirish doim ochiq.\n\n"
                "Admin buyruqlari:\n"
                "• <code>/admin</code> — admin panel\n"
                "• <code>/users</code> — foydalanuvchilar\n"
                "• <code>/extend ID 30</code> — kirishni uzaytirish\n"
                "• <code>/block ID</code> — bloklash\n"
                "• <code>/unblock ID</code> — blokdan chiqarish\n"
                "• <code>/paid ID summa kun</code> — to‘lovni yozish\n"
                "• <code>/payments</code> — to‘lovlar tarixi\n"
                "• <code>/backup_db</code> — baza zaxirasi"
            )
        return (
            "💎 <b>Моя подписка</b>\n\n"
            f"Telegram ID: <code>{telegram_id}</code>\n"
            f"Статус: {status}\n\n"
            "Trial и дата оплаты для администратора не важны — доступ всегда открыт.\n\n"
            "Команды администратора:\n"
            "• <code>/admin</code> — админ-панель\n"
            "• <code>/users</code> — пользователи\n"
            "• <code>/extend ID 30</code> — продлить доступ\n"
            "• <code>/block ID</code> — заблокировать\n"
            "• <code>/unblock ID</code> — разблокировать\n"
            "• <code>/paid ID сумма дни</code> — записать оплату\n"
            "• <code>/payments</code> — история оплат\n"
            "• <code>/backup_db</code> — скачать базу"
        )

    if lang == "uz":
        return (
            "💎 <b>Mening obunam</b>\n\n"
            f"Telegram ID: <code>{telegram_id}</code>\n"
            f"Holat: {status}\n"
            f"Trial tugash vaqti: <b>{_fmt_dt(row.get('trial_until'))}</b>\n"
            f"To‘langan muddat: <b>{_fmt_dt(row.get('subscription_until'))}</b>\n\n"
            "🎁 Trialda: bugungi savdo, yangi savdo xabarlari va ertalabki hisobot.\n"
            "💎 To‘liq obunada: botning barcha imkoniyatlari.\n\n"
            "Tariflar:\n"
            f"<b>{escape(SUBSCRIPTION_PLANS_TEXT_UZ)}</b>\n\n"
            "Tarifni tanlash va chek yuborish: <code>/subscribe</code>\n\n"
            "To‘lovlar tarixi: <code>/my_payments</code>\n"
            "Yordam: <code>/support</code>\n"
            "API-kalitni almashtirish: <code>/reconnect</code>\n"
            "API-kalitni o‘chirish: <code>/disconnect</code>"
        )
    return (
        "💎 <b>Моя подписка</b>\n\n"
        f"Telegram ID: <code>{telegram_id}</code>\n"
        f"Статус: {status}\n"
        f"Trial до: <b>{_fmt_dt(row.get('trial_until'))}</b>\n"
        f"Оплачено до: <b>{_fmt_dt(row.get('subscription_until'))}</b>\n\n"
        "🎁 В trial: продажи за сегодня, уведомления о новых продажах и утренний отчёт.\n"
        "💎 В полной подписке: все функции бота.\n\n"
        "Тарифы:\n"
        f"<b>{escape(SUBSCRIPTION_PLANS_TEXT)}</b>\n\n"
        "Выбрать тариф и отправить чек: <code>/subscribe</code>\n\n"
        "История оплат: <code>/my_payments</code>\n"
        "Поддержка: <code>/support</code>\n"
        "Заменить API-ключ: <code>/reconnect</code>\n"
        "Удалить API-ключ: <code>/disconnect</code>"
    )

class ConnectStates(StatesGroup):
    waiting_for_token = State()
    waiting_for_shop_id = State()


class CostImportStates(StatesGroup):
    waiting_for_file = State()


class BarcodePrintStates(StatesGroup):
    waiting_for_items = State()


class PaymentStates(StatesGroup):
    waiting_for_receipt = State()


class ProductSettingsStates(StatesGroup):
    waiting_for_value = State()


TRIAL_ALLOWED_COMMANDS = frozenset(
    {
        "start",
        "help",
        "menu",
        "cancel",
        "status",
        "connect_shop",
        "staff_connect",
        "addshop",
        "connect",
        "reconnect",
        "disconnect",
        "pinguzum",
        "shops",
        "setshop",
        "language",
        "lang",
        "til",
        "video",
        "api_video",
        "instruction",
        "api_token",
        "token_help",
        "how_token",
        "security",
        "privacy",
        "support",
        "subscribe",
        "my_payments",
        "my_subscription",
        "subscription",
        "today",
        "sales_today",
        "morning_report",
        "daily_report",
        "sales_notify_status",
        "sales_notify_mode",
        "sales_notifications",
    }
)

TRIAL_ALLOWED_BUTTONS = frozenset(
    {
        "💰 Продажи",
        "💰 Savdo",
        "📊 Сегодня",
        "📊 Bugun",
        "🌙 Утренний отчёт",
        "🌙 Ertalabki hisobot",
        "💸 Уведомления о продажах",
        "💸 Savdo xabarlari",
        "💎 Полная версия",
        "💎 To‘liq versiya",
        "💎 Подписка",
        "💎 Obuna",
        "⚙️ Настройки",
        "⚙️ Sozlamalar",
        "⬅️ Настройки",
        "⬅️ Sozlamalar",
        "🏠 Главное",
        "🏠 Asosiy",
        "🏪 Магазины",
        "🏪 Do‘konlar",
        "🔐 Подключение Uzum",
        "🔐 Uzum ulanishi",
        "🔐 API и подключение",
        "🔐 API va ulanish",
        "🔌 Подключить",
        "🔌 Подключить магазин",
        "🔌 Ulash",
        "🔌 Do‘konni ulash",
        "🔌 Обновить API-ключ",
        "🔌 API-kalitni yangilash",
        "✅ Проверить подключение",
        "✅ Ulanishni tekshirish",
        "🎥 Видеоинструкция",
        "🎥 Как подключить",
        "🎥 API ulash videosi",
        "🎥 Qanday ulash kerak",
        "🌐 Язык",
        "🌐 Til",
        "ℹ️ Помощь",
        "ℹ️ Yordam",
        "❓ Помощь",
        "🔐 Безопасность",
        "🔐 Xavfsizlik",
        "❌ Отмена",
        "❌ Bekor qilish",
    }
)

TRIAL_ALLOWED_CALLBACK_PREFIXES = (
    "set_lang:",
    "shop_select:",
    "payplan:",
    "paycancel:",
    "salesmode:",
)


async def send_trial_premium_locked(
    event: Message | CallbackQuery,
    telegram_id: int,
) -> None:
    """Show one consistent, actionable upgrade screen to a trial user."""
    lang = get_user_language(telegram_id)
    if lang == "uz":
        text = (
            "🔒 <b>Bu funksiya to‘liq obunada mavjud</b>\n\n"
            f"{TRIAL_DAYS} kunlik sinovda quyidagilar ochiq:\n"
            "✅ bugungi savdolar\n"
            "✅ yangi savdolar haqida xabarlar\n"
            "✅ ertalabki hisobot\n\n"
            "To‘liq obunada ombor, FBO/FBS tahlili, foyda, Excel, "
            "yo‘qotishlar, bekor qilishlar, avtohisobotlar va boshqa imkoniyatlar ochiladi.\n\n"
            "Tarifni tanlang 👇"
        )
        alert = "Bu funksiya faqat to‘liq obunada mavjud"
    else:
        text = (
            "🔒 <b>Эта функция доступна в полной подписке</b>\n\n"
            f"В пробном периоде на {TRIAL_DAYS} дня доступны:\n"
            "✅ продажи за сегодня\n"
            "✅ уведомления о новых продажах\n"
            "✅ утренний отчёт\n\n"
            "Полная подписка открывает склад, аналитику FBO/FBS, прибыль, Excel, "
            "потери, отмены, автоотчёты и остальные функции.\n\n"
            "Выберите тариф 👇"
        )
        alert = "Функция доступна только в полной подписке"

    if isinstance(event, CallbackQuery):
        await event.answer(alert, show_alert=True)
        if event.message:
            await event.message.answer(text, reply_markup=payment_plan_markup(lang))
        return
    await event.answer(text, reply_markup=payment_plan_markup(lang))


class TrialFeatureMiddleware(BaseMiddleware):
    """Prevent stale keyboards and direct commands from bypassing trial limits."""

    async def __call__(self, handler: Any, event: Any, data: dict[str, Any]) -> Any:
        user = getattr(event, "from_user", None)
        if user is None or subscription_access_level(int(user.id)) != "trial":
            return await handler(event, data)

        if isinstance(event, CallbackQuery):
            callback_data = str(event.data or "")
            if callback_data.startswith(TRIAL_ALLOWED_CALLBACK_PREFIXES):
                return await handler(event, data)
            await send_trial_premium_locked(event, int(user.id))
            return None

        if not isinstance(event, Message):
            return await handler(event, data)

        state = data.get("state")
        current_state = await state.get_state() if state is not None else None
        text = str(event.text or "").strip()
        if current_state in {
            ConnectStates.waiting_for_token.state,
            ConnectStates.waiting_for_shop_id.state,
        } and not text.startswith("/"):
            return await handler(event, data)
        if current_state == PaymentStates.waiting_for_receipt.state and (
            bool(event.photo) or event.document is not None
        ):
            return await handler(event, data)

        if text.startswith("/"):
            command = text.split(maxsplit=1)[0].split("@", 1)[0].lstrip("/").lower()
            if command in TRIAL_ALLOWED_COMMANDS:
                return await handler(event, data)
        elif text in TRIAL_ALLOWED_BUTTONS:
            return await handler(event, data)

        await send_trial_premium_locked(event, int(user.id))
        return None


_ACTIVE_USER_HANDLERS: set[int] = set()


class UserConcurrencyMiddleware(BaseMiddleware):
    """Drop duplicate in-flight work and cap global user-handler concurrency."""

    async def __call__(self, handler: Any, event: Any, data: dict[str, Any]) -> Any:
        user = getattr(event, "from_user", None)
        if user is None:
            return await handler(event, data)
        telegram_id = int(user.id)
        overloaded = (
            telegram_id in _ACTIVE_USER_HANDLERS
            or len(_ACTIVE_USER_HANDLERS) >= MAX_CONCURRENT_USER_HANDLERS
        )
        if overloaded:
            if isinstance(event, CallbackQuery):
                try:
                    await event.answer(
                        "⏳ Дождитесь завершения предыдущего запроса",
                        show_alert=False,
                    )
                except Exception:
                    pass
            return None

        _ACTIVE_USER_HANDLERS.add(telegram_id)
        try:
            return await handler(event, data)
        finally:
            _ACTIVE_USER_HANDLERS.discard(telegram_id)


dp.message.outer_middleware(UserConcurrencyMiddleware())
dp.callback_query.outer_middleware(UserConcurrencyMiddleware())
dp.message.outer_middleware(TrialFeatureMiddleware())
dp.callback_query.outer_middleware(TrialFeatureMiddleware())


def get_tg_id(message: Message) -> int:
    if not message.from_user:
        raise RuntimeError("Unknown Telegram user")
    return message.from_user.id


def upsert_from_message(message: Message) -> int:
    user = message.from_user
    if not user:
        raise RuntimeError("Unknown Telegram user")
    db.upsert_user(user.id, user.username, user.first_name)
    return user.id


def get_uzum_for_user(telegram_id: int) -> UzumClient | None:
    encrypted = db.get_encrypted_token(telegram_id)
    if not encrypted:
        return None
    token = cipher.decrypt(encrypted)
    return UzumClient(token, UZUM_API_BASE_URL)


def get_staff_uzum_client() -> UzumClient | None:
    if not STAFF_CONNECT_ENABLED or not STAFF_UZUM_TOKEN:
        return None
    return UzumClient(STAFF_UZUM_TOKEN, UZUM_API_BASE_URL)


def _shop_id_from_obj(shop: Any) -> int | None:
    if not isinstance(shop, dict):
        return None
    for key in ("id", "shopId", "shop_id", "storeId"):
        value = shop.get(key)
        if value is not None:
            try:
                return int(value)
            except Exception:
                pass
    return None


def _find_shop_in_list(shops: list[Any], shop_id: int) -> dict[str, Any] | None:
    for item in shops:
        if isinstance(item, dict) and _shop_id_from_obj(item) == int(shop_id):
            return item
    return None


def _fallback_shop_obj(shop_id: int) -> dict[str, Any]:
    return {
        "id": int(shop_id),
        "shopId": int(shop_id),
        "title": f"Магазин {int(shop_id)}",
        "name": f"Магазин {int(shop_id)}",
    }


async def notify_admins_staff_shop_connected(message: Message, telegram_id: int, shop_id: int) -> None:
    user = message.from_user
    username = f"@{user.username}" if user and user.username else "—"
    first_name = user.first_name if user and user.first_name else "—"
    text = (
        "🆕 <b>Магазин подключён через сотрудника</b>\n\n"
        f"Пользователь: {escape(first_name)} {escape(username)}\n"
        f"Telegram ID: <code>{telegram_id}</code>\n"
        f"Shop ID: <code>{shop_id}</code>\n"
        f"Доступ: {subscription_status_text(telegram_id)}"
    )
    for admin_id in ADMIN_IDS:
        if int(admin_id) == int(telegram_id):
            continue
        try:
            await bot.send_message(int(admin_id), text, reply_markup=admin_menu_for_user(int(admin_id)))
            await asyncio.sleep(0.1)
        except Exception:
            logging.exception("Failed to notify admin about staff shop connection")


async def connect_shop_by_staff(message: Message, shop_id_text: str, state: FSMContext | None = None) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    shop_id_text = (shop_id_text or "").strip()
    if not shop_id_text.isdigit():
        if lang == "uz":
            await message.answer(
                "Shop ID faqat raqamlardan iborat bo‘lishi kerak.\nMasalan: <code>113982</code>",
                reply_markup=menu_for_message(message),
            )
        else:
            await message.answer(
                "ID магазина должен состоять только из цифр.\nНапример: <code>113982</code>",
                reply_markup=menu_for_message(message),
            )
        return

    if not await require_active_subscription(message, telegram_id):
        return

    staff_client = get_staff_uzum_client()
    if staff_client is None:
        await message.answer(
            "⚠️ Простой способ подключения пока не настроен.\n\n"
            "Можно подключиться старым способом через API-ключ: <code>/connect</code>",
            reply_markup=menu_for_message(message),
        )
        if state:
            await state.clear()
        return

    shop_id = int(shop_id_text)
    save_staff_shop_status(telegram_id, shop_id, "pending")

    try:
        shops: list[Any] = []
        shop_obj: dict[str, Any] | None = None
        try:
            shops_data = await staff_client.get_shops()
            shops = extract_items(shops_data)
            shop_obj = _find_shop_in_list(shops, shop_id)
        except Exception:
            logging.exception("Staff connect: failed to load shops list")

        await staff_client.get_products(shop_id, page=0, size=1)

        # Проверяем доступ к Finance. Даже если продаж нет, метод должен ответить без 403.
        date_from, date_to = _days_range_ms(30)
        await _load_finance_orders(
            staff_client,
            shop_id,
            date_from_ms=date_from,
            date_to_ms=date_to,
            max_pages=1,
            page_size=1,
        )

        encrypted = cipher.encrypt(STAFF_UZUM_TOKEN)
        db.save_connection(telegram_id, encrypted, [shop_obj or _fallback_shop_obj(shop_id)])
        try:
            db.set_default_shop_id(telegram_id, shop_id)
        except Exception:
            pass
        save_staff_shop_status(telegram_id, shop_id, "connected")

        if state:
            await state.clear()

        if lang == "uz":
            text_ok = (
                "✅ <b>Do‘kon ulandi</b>\n\n"
                f"Shop ID: <code>{shop_id}</code>\n"
                "Xodim orqali kirish tasdiqlandi.\n\n"
                "Boshlash uchun <b>🏠 Do‘kon holati</b> tugmasini bosing."
            )
        else:
            text_ok = (
                "✅ <b>Магазин подключён</b>\n\n"
                f"Shop ID: <code>{shop_id}</code>\n"
                "Доступ через сотрудника подтверждён.\n\n"
                "Для начала нажмите <b>🏠 Обзор магазина</b>."
            )
        await message.answer(text_ok, reply_markup=menu_for_message(message))
        await notify_admins_staff_shop_connected(message, telegram_id, shop_id)

    except Exception as e:
        save_staff_shop_status(telegram_id, shop_id, "no_access", str(e))
        raw = str(e)
        low = raw.lower()
        if "403" in raw or "rbac" in low or "forbidden" in low or "access" in low:
            if lang == "uz":
                text = (
                    "⛔ <b>Do‘konga kirish topilmadi</b>\n\n"
                    "Ehtimol, xodim hali qo‘shilmagan yoki unga savdo/moliya/tovarlar bo‘yicha huquqlar berilmagan.\n\n"
                    "1. Uzum Seller kabinetida xodim qo‘shilganini tekshiring.\n"
                    "2. Savdo, moliya, tovarlar va qoldiq huquqlarini bering.\n"
                    "3. Keyin Shop ID ni qayta yuboring."
                )
            else:
                text = (
                    "⛔ <b>Доступ к магазину не найден</b>\n\n"
                    "Скорее всего, сотрудник ещё не добавлен или ему не дали права на продажи/финансы/товары.\n\n"
                    "1. Проверьте, что сотрудник добавлен в Uzum Seller.\n"
                    "2. Дайте права на продажи, финансы, товары и остатки.\n"
                    "3. После этого отправьте Shop ID ещё раз."
                )
            await message.answer(text, reply_markup=menu_for_message(message))
            return
        await send_api_error(message, e)


async def require_connection(message: Message) -> tuple[int, UzumClient, int] | None:
    telegram_id = upsert_from_message(message)
    if not await require_active_subscription(message, telegram_id):
        return None
    client = get_uzum_for_user(telegram_id)
    shop_id = db.get_default_shop_id(telegram_id)

    if client is None:
        lang = get_user_language(telegram_id)
        if lang == "uz":
            text = (
                "🔌 <b>Avval do‘konni ulang</b>\n\n"
                "Pastdagi <b>🎥 Qanday ulash kerak</b> videosini ko‘ring, so‘ng "
                "<b>🔌 Do‘konni ulash</b> tugmasini bosing."
            )
        else:
            text = (
                "🔌 <b>Сначала подключите магазин</b>\n\n"
                "Посмотрите <b>🎥 Как подключить</b>, затем нажмите "
                "<b>🔌 Подключить магазин</b>."
            )
        await message.answer(text, reply_markup=menu_for_message(message))
        return None

    if shop_id is None:
        lang = get_user_language(telegram_id)
        await message.answer(
            (
                "API-kalit ulangan, lekin faol do‘kon tanlanmagan.\n"
                "<b>⚙️ Sozlamalar → 🏪 Do‘konlar</b> bo‘limidan do‘konni tanlang."
                if lang == "uz"
                else "API-ключ подключён, но активный магазин не выбран.\n"
                "Откройте <b>⚙️ Настройки → 🏪 Магазины</b> и выберите магазин."
            ),
            reply_markup=menu_for_message(message),
        )
        return None

    return telegram_id, client, int(shop_id)


async def send_api_error(message: Message, error: Exception) -> None:
    raw = str(error)
    low = raw.lower()
    if "token inactive" in low:
        user_text = (
            "🔐 <b>API-ключ Uzum больше не активен</b>\n\n"
            "Создайте новый API-ключ в кабинете Uzum Seller и откройте "
            "<b>⚙️ Настройки → 🔐 API и подключение → 🔌 Обновить API-ключ</b>."
        )
    elif "shop is not available" in low:
        user_text = (
            "🏪 <b>Магазин недоступен для этого API-ключа</b>\n\n"
            "Ключ не имеет доступа к выбранному Shop ID. Обновите API-ключ либо выберите "
            "магазин, который доступен этому ключу: <b>⚙️ Настройки → 🏪 Магазины</b>."
        )
    elif "401" in raw or "unauthorized" in low:
        user_text = (
            "🔐 <b>Uzum API-ключ не принят</b>\n\n"
            "Возможно, ключ неверный, удалён или истёк.\n"
            "Создайте новый ключ и откройте <b>⚙️ Настройки → 🔐 API и подключение → "
            "🔌 Обновить API-ключ</b>."
        )
    elif "403" in raw or "rbac" in low or "forbidden" in low:
        user_text = (
            "⛔ <b>Uzum запретил доступ</b>\n\n"
            "Проверьте права API-ключа и доступ ключа к выбранному магазину."
        )
    elif "429" in raw or "too many" in low:
        user_text = (
            "⏳ <b>Uzum временно ограничил запросы</b>\n\n"
            "Слишком много запросов к Uzum API. Подождите несколько минут и попробуйте снова."
        )
    elif "500" in raw or "502" in raw or "503" in raw or "504" in raw:
        user_text = (
            "⚠️ <b>Uzum API временно недоступен</b>\n\n"
            "Это похоже на ошибку на стороне Uzum. Попробуйте позже."
        )
    else:
        text = escape(raw)
        if len(text) > 1200:
            text = text[:1200] + "\n..."
        user_text = f"⚠️ <b>Ошибка API</b>\n<code>{text}</code>"
    await message.answer(user_text, reply_markup=menu_for_message(message))


def parse_args(text: str) -> str:
    text = (text or "").strip()
    # Аргументы берём только у настоящих slash-команд.
    # Если пользователь нажал русскую кнопку вроде "📦 Товары",
    # это не должно уходить в Uzum API как поисковый запрос или статус заказа.
    if not text.startswith("/"):
        return ""
    parts = text.split(maxsplit=1)
    return parts[1].strip() if len(parts) > 1 else ""


async def load_products(
    client: UzumClient,
    shop_id: int,
    *,
    search_query: str = "",
    max_pages: int = 20,
    page_size: int = 100,
    product_filter: str = "ALL",
) -> list[Any]:
    all_products: list[Any] = []
    for page in range(max_pages):
        data = await client.get_products(
            shop_id,
            search_query=search_query,
            page=page,
            size=page_size,
            product_filter=product_filter,
        )
        items = extract_items(data)
        if not items:
            break
        all_products.extend(items)
        total_products: int | None = None
        if isinstance(data, dict):
            raw_total = data.get("totalProductsAmount")
            try:
                if raw_total is not None:
                    total_products = max(0, int(raw_total))
            except (TypeError, ValueError):
                total_products = None
        if total_products is not None and total_products > 0:
            if len(all_products) >= total_products:
                break
        elif len(items) < page_size:
            break
    return all_products


def _official_stock_number(raw: dict[str, Any], *names: str) -> float | None:
    """Read a numeric stock field without confusing it with another quantity.

    Uzum's current Product API exposes quantityActive (warehouse stock) and
    quantityFbs (seller warehouse) as separate values.  Older bot versions
    treated quantityActive as a combined total and then subtracted FBS from it,
    which could produce misleading FBO/total figures.
    """
    if not isinstance(raw, dict):
        return None
    lower_keys = {str(key).lower(): key for key in raw}
    for name in names:
        actual = name if name in raw else lower_keys.get(name.lower())
        if actual is None:
            continue
        value = _num_from_value(raw.get(actual))
        if value is not None:
            return value
    return None


def _stock_row_identity(row: dict[str, Any]) -> str:
    """Stable SKU identity used to prevent double counting API duplicates."""
    for field in ("sku_id", "barcode", "seller_item_code"):
        value = row.get(field)
        if value not in (None, "", 0, "0", "—"):
            return f"{field}:{str(value).strip().lower()}"
    return "fallback:" + "|".join(
        str(row.get(field) or "").strip().lower()
        for field in ("product_id", "sku_full_title", "sku_title", "product_title")
    )


def _normalize_current_stock_row(source: dict[str, Any]) -> dict[str, Any]:
    """Normalize one Product API row using the documented current fields."""
    row = dict(source)
    raw = row.get("raw") if isinstance(row.get("raw"), dict) else {}

    active = _official_stock_number(raw, "quantityActive")
    fbs = _official_stock_number(raw, "quantityFbs")
    if active is not None or fbs is not None:
        # Official Swagger descriptions:
        # quantityActive — active SKU units at the warehouse;
        # quantityFbs — units available under Fulfillment by Seller.
        fbo_units = max(0.0, float(active or 0.0))
        fbs_units = max(0.0, float(fbs or 0.0))
        row["active"] = fbo_units
        row["fbo"] = fbo_units
        row["fbs"] = fbs_units
        row["total"] = fbo_units + fbs_units
        row["stock_source"] = "quantityActive + quantityFbs"
    else:
        # Compatibility fallback for a different/older response shape.
        fbo_units = max(0.0, float(_num_from_value(row.get("fbo")) or 0.0))
        fbs_units = max(0.0, float(_num_from_value(row.get("fbs")) or 0.0))
        total = _num_from_value(row.get("total"))
        row["fbo"] = fbo_units
        row["fbs"] = fbs_units
        row["total"] = max(0.0, float(total)) if total is not None else fbo_units + fbs_units
        row["stock_source"] = "compatibility"

    # Uzum also exposes an official per-SKU average.  It is used only as a
    # fallback when a Finance API row cannot be matched to this SKU.
    row["avg_daily_sales_api"] = max(
        0.0,
        float(_official_stock_number(raw, "avgdsales", "avgDailySales") or 0.0),
    )
    row["avg_daily_active_api"] = max(
        0.0,
        float(_official_stock_number(raw, "avgdquantity", "avgDailyQuantity") or 0.0),
    )
    # Keep documented financial/catalog values distinct from sale price.
    if row.get("purchase_price") is None:
        row["purchase_price"] = _official_stock_number(raw, "purchasePrice")
    if not row.get("ikpu"):
        row["ikpu"] = raw.get("ikpu")
    if row.get("paid_storage_price_item") is None:
        row["paid_storage_price_item"] = _official_stock_number(raw, "paidStoragePriceItem")
    if row.get("paid_storage_amount") is None:
        row["paid_storage_amount"] = _official_stock_number(raw, "paidStorageAmount")
    row["paid_storage"] = bool(row.get("paid_storage") or raw.get("pstorage"))
    row["duplicate_count"] = max(1, int(row.get("duplicate_count") or 1))
    return row


def _normalize_current_stock_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Normalize and deduplicate current stock without summing repeated SKUs."""
    merged: dict[str, dict[str, Any]] = {}
    ordered_keys: list[str] = []
    quantity_fields = (
        "active",
        "fbo",
        "fbs",
        "total",
        "additional",
        "sold",
        "returned",
        "missing",
        "defected",
        "pending",
        "avg_daily_sales_api",
        "avg_daily_active_api",
    )
    descriptive_fields = (
        "product_id",
        "sku_id",
        "barcode",
        "seller_item_code",
        "product_title",
        "sku_title",
        "sku_full_title",
        "category",
        "price",
        "market_price",
        "purchase_price",
        "ikpu",
        "paid_storage_price_item",
        "paid_storage_amount",
        "paid_storage",
        "status",
    )

    for source in rows:
        row = _normalize_current_stock_row(source)
        key = _stock_row_identity(row)
        current = merged.get(key)
        if current is None:
            merged[key] = row
            ordered_keys.append(key)
            continue

        current["duplicate_count"] = int(current.get("duplicate_count") or 1) + 1
        # A repeated API row is the same SKU, not more physical stock.  Keep
        # the greatest observed counter instead of adding it twice.
        for field in quantity_fields:
            current_value = _num_from_value(current.get(field))
            row_value = _num_from_value(row.get(field))
            if row_value is not None and (current_value is None or row_value > current_value):
                current[field] = row_value
        for field in descriptive_fields:
            if current.get(field) in (None, "", "—") and row.get(field) not in (None, "", "—"):
                current[field] = row.get(field)

        # Rebuild total from the two documented live-stock components.
        if current.get("stock_source") == "quantityActive + quantityFbs":
            current["total"] = max(0.0, float(current.get("fbo") or 0.0)) + max(
                0.0, float(current.get("fbs") or 0.0)
            )

    result = [merged[key] for key in ordered_keys]
    duplicate_rows = sum(max(0, int(row.get("duplicate_count") or 1) - 1) for row in result)
    if duplicate_rows:
        logging.warning(
            "Stock normalization removed %s duplicate SKU rows (raw=%s unique=%s)",
            duplicate_rows,
            len(rows),
            len(result),
        )
    return result


async def load_sku_rows(
    client: UzumClient,
    shop_id: int,
    *,
    search_query: str = "",
    max_pages: int = 20,
) -> list[dict[str, Any]]:
    products = await load_products(
        client, shop_id, search_query=search_query, max_pages=max_pages, page_size=100
    )
    return _normalize_current_stock_rows(flatten_sku_rows(products))


UZUM_SKU_FINANCE_CACHE_SECONDS = max(
    60,
    min(3600, int(os.getenv("UZUM_SKU_FINANCE_CACHE_SECONDS", "600") or "600")),
)
_UZUM_SKU_FINANCE_SYNCED_AT: dict[tuple[int, int], float] = {}
_UZUM_SKU_FINANCE_LOCKS: dict[tuple[int, int], tuple[Any, asyncio.Lock]] = {}


def _uzum_sku_finance_lock(key: tuple[int, int]) -> asyncio.Lock:
    loop = asyncio.get_running_loop()
    stored = _UZUM_SKU_FINANCE_LOCKS.get(key)
    if stored is None or stored[0] is not loop:
        lock = asyncio.Lock()
        _UZUM_SKU_FINANCE_LOCKS[key] = (loop, lock)
        return lock
    return stored[1]


def _uzum_sku_finance_status(
    telegram_id: int,
    shop_id: int,
    *,
    stale: bool = False,
) -> dict[str, Any]:
    rows = list_uzum_sku_financials(telegram_id, shop_id)
    with_cost = sum(1 for row in rows if float(row.get("purchase_price") or 0) > 0)
    with_ikpu = sum(1 for row in rows if str(row.get("ikpu") or "").strip())
    paid_storage = sum(1 for row in rows if int(row.get("paid_storage") or 0))
    fetched_values = [str(row.get("fetched_at") or "") for row in rows if row.get("fetched_at")]
    return {
        "total": len(rows),
        "with_cost": with_cost,
        "missing_cost": max(0, len(rows) - with_cost),
        "with_ikpu": with_ikpu,
        "missing_ikpu": max(0, len(rows) - with_ikpu),
        "paid_storage": paid_storage,
        "fetched_at": max(fetched_values) if fetched_values else None,
        "stale": bool(stale),
    }


def _replace_uzum_sku_financials(
    telegram_id: int,
    shop_id: int,
    stock_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    records = build_stock_records(stock_rows)
    now_text = _dt_to_db(_utc_now()) or ""
    values = [
        (
            int(telegram_id),
            int(shop_id),
            str(record.get("sku_key") or ""),
            json.dumps(record.get("aliases") or [], ensure_ascii=False),
            str(record.get("sku_id") or ""),
            str(record.get("barcode") or ""),
            str(record.get("seller_item_code") or ""),
            str(record.get("sku_title") or ""),
            str(record.get("product_title") or ""),
            record.get("purchase_price"),
            str(record.get("ikpu") or ""),
            record.get("paid_storage_price_item"),
            record.get("paid_storage_amount"),
            1 if record.get("paid_storage") else 0,
            now_text,
        )
        for record in records
        if str(record.get("sku_key") or "")
    ]
    init_unit_economy_tables()
    with db.connect() as conn:
        conn.execute(
            "DELETE FROM uzum_sku_financials WHERE telegram_id = ? AND shop_id = ?",
            (int(telegram_id), int(shop_id)),
        )
        if values:
            conn.executemany(
                """
                INSERT INTO uzum_sku_financials (
                    telegram_id, shop_id, sku_key, aliases_json, sku_id,
                    barcode, seller_item_code, sku_title, product_title,
                    purchase_price, ikpu, paid_storage_price_item,
                    paid_storage_amount, paid_storage, fetched_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                values,
            )
        conn.commit()
    return _uzum_sku_finance_status(telegram_id, shop_id)


async def sync_uzum_sku_financials(
    client: UzumClient,
    telegram_id: int,
    shop_id: int,
    *,
    stock_rows: list[dict[str, Any]] | None = None,
    force: bool = False,
) -> dict[str, Any]:
    """Refresh purchase prices/catalog fields exclusively from Uzum.

    A temporary blank/error response never overwrites a non-empty cache.  This
    keeps reports stable during an Uzum outage while the UI marks cached data.
    """
    key = (int(telegram_id), int(shop_id))
    cached_rows = list_uzum_sku_financials(*key)
    last_sync = _UZUM_SKU_FINANCE_SYNCED_AT.get(key, 0.0)
    if not force and cached_rows and time.monotonic() - last_sync < UZUM_SKU_FINANCE_CACHE_SECONDS:
        return _uzum_sku_finance_status(*key)

    async with _uzum_sku_finance_lock(key):
        cached_rows = list_uzum_sku_financials(*key)
        last_sync = _UZUM_SKU_FINANCE_SYNCED_AT.get(key, 0.0)
        if not force and cached_rows and time.monotonic() - last_sync < UZUM_SKU_FINANCE_CACHE_SECONDS:
            return _uzum_sku_finance_status(*key)
        try:
            rows = stock_rows
            if rows is None:
                rows = await load_sku_rows(client, int(shop_id), max_pages=100)
            if not rows and cached_rows:
                logging.warning(
                    "Uzum SKU finance sync returned empty rows; keeping cache user=%s shop=%s",
                    telegram_id,
                    shop_id,
                )
                return _uzum_sku_finance_status(*key, stale=True)
            status = _replace_uzum_sku_financials(*key, list(rows or []))
            _UZUM_SKU_FINANCE_SYNCED_AT[key] = time.monotonic()
            logging.info(
                "Uzum SKU finance sync user=%s shop=%s sku=%s with_cost=%s missing=%s ikpu=%s",
                telegram_id,
                shop_id,
                status["total"],
                status["with_cost"],
                status["missing_cost"],
                status["with_ikpu"],
            )
            return status
        except Exception:
            if cached_rows:
                logging.exception(
                    "Uzum SKU finance sync failed; using cached data user=%s shop=%s",
                    telegram_id,
                    shop_id,
                )
                return _uzum_sku_finance_status(*key, stale=True)
            raise


async def connect_token(
    message: Message, token: str, state: FSMContext | None = None
) -> None:
    telegram_id = upsert_from_message(message)
    token = token.strip()
    if not token or len(token) < 20:
        lang = get_user_language(telegram_id)
        await message.answer(
            (
                "Bu API-kalitga o‘xshamaydi. To‘liq kalitni bitta xabar qilib yuboring."
                if lang == "uz"
                else "Похоже, это не API-ключ. Отправьте полный ключ одним сообщением."
            ),
            reply_markup=CONNECT_INPUT_MENU_UZ if lang == "uz" else CONNECT_INPUT_MENU_RU,
        )
        return

    try:
        client = UzumClient(token, UZUM_API_BASE_URL)
        data = await client.get_shops()
        shops = extract_items(data)
        if not shops:
            lang = get_user_language(telegram_id)
            await message.answer(
                (
                    "Kalit qabul qilindi, lekin unga bog‘langan do‘kon topilmadi. "
                    "Kalit huquqlarini tekshiring yoki boshqa kalit yuboring."
                    if lang == "uz"
                    else "Ключ принят, но доступных магазинов не найдено. "
                    "Проверьте права ключа или отправьте другой ключ."
                ),
                reply_markup=CONNECT_INPUT_MENU_UZ if lang == "uz" else CONNECT_INPUT_MENU_RU,
            )
            return

        encrypted = cipher.encrypt(token)
        default_shop_id = db.save_connection(telegram_id, encrypted, shops)

        try:
            await message.delete()
        except Exception:
            pass

        lang = get_user_language(telegram_id)
        if lang == "uz":
            text_ok = (
                "✅ <b>Do‘kon ulandi</b>\n\n"
                f"Topilgan do‘konlar: <b>{len(shops)}</b>\n"
                "Faol do‘kon: "
                + (f"<code>{default_shop_id}</code>" if default_shop_id else "tanlanmagan")
                + "\n\nTayyor. Boshlash uchun:\n"
                "🏠 <b>Do‘kon holati</b> — asosiy raqamlar\n"
                "🚨 <b>Hozir muhim</b> — birinchi navbatdagi muammolar\n"
                "💰 <b>Savdo</b> va 📦 <b>Ombor</b> — batafsil ma’lumot"
            )
        else:
            text_ok = (
                "✅ <b>Магазин подключён</b>\n\n"
                f"Найдено магазинов: <b>{len(shops)}</b>\n"
                "Активный магазин: "
                + (f"<code>{default_shop_id}</code>" if default_shop_id else "не выбран")
                + "\n\nГотово. Начните с разделов:\n"
                "🏠 <b>Обзор магазина</b> — главные цифры\n"
                "🚨 <b>Важно сейчас</b> — проблемы в первую очередь\n"
                "💰 <b>Продажи</b> и 📦 <b>Склад</b> — подробности"
            )
        await message.answer(text_ok, reply_markup=menu_for_message(message))
        if state:
            await state.clear()
    except Exception as e:
        await send_api_error(message, e)
        if state:
            lang = get_user_language(telegram_id)
            await message.answer(
                "Boshqa kalit yuboring yoki bekor qiling."
                if lang == "uz"
                else "Отправьте другой ключ или отмените подключение.",
                reply_markup=CONNECT_INPUT_MENU_UZ if lang == "uz" else CONNECT_INPUT_MENU_RU,
            )


@dp.message(Command("language", "lang", "til"))
async def language_command(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    await message.answer(
        f"{tr(lang, 'language_title')}\n\n{tr(lang, 'language_body')}",
        reply_markup=language_markup(),
    )


@dp.message(F.text == "🌐 Язык")
@dp.message(F.text == "🌐 Til")
async def language_button(message: Message) -> None:
    await language_command(message)


@dp.callback_query(F.data.startswith("set_lang:"))
async def set_language_callback(callback: CallbackQuery) -> None:
    if not callback.from_user:
        return
    telegram_id = callback.from_user.id
    db.upsert_user(telegram_id, callback.from_user.username, callback.from_user.first_name)
    lang = normalize_lang((callback.data or "").split(":", 1)[-1])
    set_user_language(telegram_id, lang)
    await callback.answer("OK")
    if callback.message:
        await callback.message.answer(tr(lang, "language_set"), reply_markup=main_menu_for_user(telegram_id))



# ---------------------------------------------------------------------------
# Telegram ↔ Seller Assistant Web (встроенная интеграция)
# ---------------------------------------------------------------------------
def _web_row_value(row: Any, name: str, default: Any = None) -> Any:
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(name, default)
    try:
        return row[name]
    except Exception:
        return getattr(row, name, default)


def _web_iso(value: Any) -> str | None:
    if not value:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        try:
            dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        except (TypeError, ValueError):
            return str(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).isoformat()


def _web_subscription_payload(telegram_id: int) -> tuple[str, str | None]:
    try:
        subscription = get_subscription_row(telegram_id)
    except Exception:
        logging.exception("WEB_BRIDGE: не удалось прочитать подписку")
        subscription = None
    if not subscription:
        return "unknown", None

    blocked = bool(int(_web_row_value(subscription, "blocked", 0) or 0))
    paid_until = _dt_from_db(_web_row_value(subscription, "subscription_until"))
    trial_until = _dt_from_db(_web_row_value(subscription, "trial_until"))
    now = datetime.now(timezone.utc)
    if blocked:
        state = "blocked"
        until = max(
            (value for value in (paid_until, trial_until) if value is not None),
            default=None,
        )
    elif is_admin(telegram_id) or (paid_until and paid_until > now):
        state = "active"
        until = paid_until
    elif trial_until and trial_until > now:
        state = "trial"
        until = trial_until
    else:
        state = "expired"
        until = max(
            (value for value in (paid_until, trial_until) if value is not None),
            default=None,
        )
    return state, _web_iso(until)


def _web_cost_rows(telegram_id: int, shop_id: int | None) -> list[dict[str, Any]]:
    if not shop_id:
        return []
    try:
        values = get_unit_cost_map(telegram_id, int(shop_id)) or {}
    except Exception:
        logging.exception("WEB_BRIDGE: не удалось прочитать себестоимость")
        return []

    rows: list[dict[str, Any]] = []
    if isinstance(values, dict):
        for sku, row in values.items():
            rows.append(
                {
                    "sku_id": str(sku),
                    "product_name": str(_web_row_value(row, "title", "") or ""),
                    "unit_cost": float(_web_row_value(row, "cost", 0) or 0),
                }
            )
            if len(rows) >= 3000:
                break
    return rows


def _web_payload_for(message: Message, *, sensitive: bool) -> dict[str, Any]:
    if not message.from_user:
        raise RuntimeError("Telegram не передал данные пользователя")

    telegram_id = int(message.from_user.id)
    db.upsert_user(telegram_id, message.from_user.username, message.from_user.first_name)
    try:
        user = db.get_user(telegram_id)
    except Exception:
        logging.exception("WEB_BRIDGE: не удалось прочитать пользователя")
        user = None

    locale = "uz" if get_user_language(telegram_id).lower().startswith("uz") else "ru"
    subscription_state, subscription_until = _web_subscription_payload(telegram_id)
    default_shop_id = _web_row_value(user, "default_shop_id")
    default_shop_id = int(default_shop_id) if str(default_shop_id or "").isdigit() else None

    payload: dict[str, Any] = {
        "telegram_id": telegram_id,
        "first_name": message.from_user.first_name or "",
        "last_name": message.from_user.last_name or "",
        "username": message.from_user.username or "",
        "locale": locale,
        "subscription_until": subscription_until,
        "subscription_state": subscription_state,
        "default_shop_id": default_shop_id,
        "iat": int(time.time()),
    }

    if sensitive:
        encrypted_token = _web_row_value(user, "uzum_token_encrypted", "") or _web_row_value(user, "encrypted_token", "")
        # Fallback for Database implementations that expose a dedicated getter.
        if not encrypted_token and hasattr(db, "get_encrypted_token"):
            try:
                encrypted_token = db.get_encrypted_token(telegram_id) or ""
            except Exception:
                logging.exception("WEB_BRIDGE: не удалось прочитать зашифрованный Uzum-токен")
        payload["encrypted_token"] = str(encrypted_token or "")
        payload["costs"] = _web_cost_rows(telegram_id, default_shop_id)

    return payload


def _web_validate_settings() -> None:
    if not WEB_APP_URL:
        raise RuntimeError("добавьте WEB_APP_URL в переменные BotHost бота")
    if not WEB_APP_URL.startswith("https://"):
        raise RuntimeError("WEB_APP_URL должен начинаться с https://")
    if len(WEB_SYNC_SECRET) < 32:
        raise RuntimeError("WEB_SYNC_SECRET должен содержать не менее 32 символов")


def _web_signed_url(message: Message) -> str:
    _web_validate_settings()
    raw = json.dumps(
        _web_payload_for(message, sensitive=False),
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    payload = base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")
    signature = hmac.new(
        WEB_SYNC_SECRET.encode("utf-8"),
        payload.encode("ascii"),
        hashlib.sha256,
    ).hexdigest()
    return f"{WEB_APP_URL}/auth/telegram/bridge?payload={payload}&sig={signature}"


def _web_sync_request(payload: dict[str, Any]) -> None:
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    signature = hmac.new(WEB_SYNC_SECRET.encode("utf-8"), raw, hashlib.sha256).hexdigest()
    request = urllib.request.Request(
        f"{WEB_APP_URL}/api/integrations/telegram/sync",
        data=raw,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Accept": "application/json",
            "X-Seller-Signature": signature,
            "User-Agent": "uzum-seller-assistant-bot/builtin-web-bridge",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=WEB_SYNC_TIMEOUT_SECONDS) as response:
            body = response.read()
            if response.status >= 300:
                raise RuntimeError(f"сайт вернул HTTP {response.status}: {body[:300].decode('utf-8', 'replace')}")
    except urllib.error.HTTPError as exc:
        try:
            detail = exc.read(500).decode("utf-8", "replace")
        except Exception:
            detail = ""
        raise RuntimeError(f"сайт вернул HTTP {exc.code}: {detail or exc.reason}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"сайт недоступен: {exc.reason}") from exc


async def sync_to_web_cabinet(message: Message) -> None:
    _web_validate_settings()
    await asyncio.to_thread(_web_sync_request, _web_payload_for(message, sensitive=True))


def web_cabinet_markup(message: Message) -> InlineKeyboardMarkup:
    locale = _web_payload_for(message, sensitive=False).get("locale")
    label = "🌐 Veb-kabinetni ochish" if locale == "uz" else "🌐 Открыть веб-кабинет"
    # URL-кнопка надёжнее WebApp-кнопки: не требует предварительной настройки
    # домена через BotFather, но всё равно выполняет безопасный вход через Telegram.
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text=label, url=_web_signed_url(message))]]
    )


@dp.message(Command("site", "web", "cabinet"))
async def open_web_cabinet_command(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not await require_premium_subscription(message, telegram_id):
        return
    locale = get_user_language(telegram_id)
    try:
        # Сначала передаём сайт-сайту зашифрованный Uzum-токен, подписку,
        # выбранный магазин и себестоимость. В URL секретные данные не попадают.
        await sync_to_web_cabinet(message)
        markup = web_cabinet_markup(message)
    except Exception as exc:
        logging.exception("WEB_BRIDGE: ошибка синхронизации")
        if locale == "uz":
            await message.answer(
                f"⚠️ Veb-kabinet hozircha sozlanmagan:\n<code>{escape(str(exc))}</code>"
            )
        else:
            await message.answer(
                f"⚠️ Веб-кабинет пока не настроен:\n<code>{escape(str(exc))}</code>"
            )
        return

    if locale == "uz":
        text = (
            "🌐 <b>Seller Assistant veb-kabineti</b>\n\n"
            "Bot va sayt bitta akkauntda ishlaydi. Til, obuna, do‘kon, "
            "Uzum ulanishi va tannarxlar sinxronlandi."
        )
    else:
        text = (
            "🌐 <b>Веб-кабинет Seller Assistant</b>\n\n"
            "Бот и сайт работают под одним аккаунтом. Язык, подписка, магазин, "
            "подключение Uzum и себестоимость синхронизированы."
        )
    await message.answer(text, reply_markup=markup)


@dp.message(F.text.in_({"🌐 Веб-кабинет", "🌐 Открыть сайт", "🌐 Veb-kabinet"}))
async def open_web_cabinet_button(message: Message) -> None:
    await open_web_cabinet_command(message)


@dp.message(Command("start", "help"))
async def start(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    ensure_subscription(telegram_id)
    lang = get_user_language(telegram_id)
    is_connected = db.has_uzum_connection(telegram_id)
    connected = "✅ подключён" if is_connected else "❌ не подключён"
    if lang == "uz":
        connected = "✅ ulangan" if is_connected else "❌ ulanmagan"
    sub_line = subscription_status_text(telegram_id)
    trial = subscription_access_level(telegram_id) == "trial"

    if lang == "uz":
        if is_connected:
            if trial:
                text = (
                    "👋 <b>Seller.pro.uz</b>\n\n"
                    f"Do‘kon: {connected}\nKirish: {sub_line}\n\n"
                    "🎁 <b>Sinov davrida ochiq:</b>\n"
                    "💰 bugungi savdolar\n💸 yangi savdo xabarlari\n🌙 ertalabki hisobot\n\n"
                    "Kerakli tugmani pastdan bosing 👇"
                )
            else:
                text = (
                    "👋 <b>Seller.pro.uz</b>\n\n"
                    f"Do‘kon: {connected}\n"
                    f"Kirish: {sub_line}\n\n"
                    "<b>Nimadan boshlash kerak?</b>\n"
                    "🏠 <b>Do‘kon holati</b> — asosiy raqamlarni bir joyda ko‘rish\n"
                    "🚨 <b>Hozir muhim</b> — nimaga e’tibor berish va nima qilish\n"
                    "💰 <b>Savdo</b> va 📦 <b>Ombor</b> — kundalik ishlar\n\n"
                    "Kerakli tugmani pastdan bosing 👇"
                )
        else:
            text = (
                "👋 <b>Seller.pro.uz</b>\n\n"
                "Uzum do‘koningizni Telegram orqali nazorat qilish uchun yordamchi.\n\n"
                f"Do‘kon: {connected}\n"
                f"Kirish: {sub_line}\n\n"
                "<b>Boshlash juda oson:</b>\n"
                "1. <b>🎥 Qanday ulash kerak</b> — qisqa videoni ko‘ring\n"
                "2. <b>🔌 Do‘konni ulash</b> — API-kalitni yuboring\n\n"
                "Qolgan bo‘limlar ulanishdan keyin avtomatik ochiladi."
            )
    else:
        if is_connected:
            if trial:
                text = (
                    "👋 <b>Seller.pro.uz</b>\n\n"
                    f"Магазин: {connected}\nДоступ: {sub_line}\n\n"
                    "🎁 <b>В пробном периоде открыты:</b>\n"
                    "💰 продажи за сегодня\n💸 уведомления о новых продажах\n🌙 утренний отчёт\n\n"
                    "Нажмите нужную кнопку внизу 👇"
                )
            else:
                text = (
                    "👋 <b>Seller.pro.uz</b>\n\n"
                    f"Магазин: {connected}\n"
                    f"Доступ: {sub_line}\n\n"
                    "<b>С чего начать?</b>\n"
                    "🏠 <b>Обзор магазина</b> — главные цифры на одном экране\n"
                    "🚨 <b>Важно сейчас</b> — что требует внимания и что сделать\n"
                    "💰 <b>Продажи</b> и 📦 <b>Склад</b> — ежедневная работа\n\n"
                    "Нажмите нужную кнопку внизу 👇"
                )
        else:
            text = (
                "👋 <b>Seller.pro.uz</b>\n\n"
                "Помощник для контроля магазина Uzum прямо в Telegram.\n\n"
                f"Магазин: {connected}\n"
                f"Доступ: {sub_line}\n\n"
                "<b>Начать очень просто:</b>\n"
                "1. <b>🎥 Как подключить</b> — посмотрите короткое видео\n"
                "2. <b>🔌 Подключить магазин</b> — отправьте API-ключ\n\n"
                "Остальные разделы откроются автоматически после подключения."
            )
    await message.answer(text, reply_markup=menu_for_message(message))


@dp.message(Command("menu"))
async def menu(message: Message) -> None:
    upsert_from_message(message)
    await message.answer(tr_user(upsert_from_message(message), "choose_section"), reply_markup=menu_for_message(message))


@dp.message(Command("cancel"))
async def cancel(message: Message, state: FSMContext) -> None:
    telegram_id = upsert_from_message(message)
    current_state = await state.get_state()
    if current_state == PaymentStates.waiting_for_receipt.state:
        data = await state.get_data()
        request_id = int(data.get("payment_request_id") or 0)
        if request_id:
            cancel_payment_request(request_id, telegram_id)
    await state.clear()
    await message.answer(tr_user(telegram_id, "cancelled"), reply_markup=menu_for_message(message))


@dp.message(Command("status"))
async def status(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    ensure_subscription(telegram_id)
    user = db.get_user(telegram_id)
    shops = db.list_shops(telegram_id)
    connected = bool(user and user["uzum_token_encrypted"])
    default_shop_id = user["default_shop_id"] if user else None
    await message.answer(
        "⚙️ <b>Статус</b>\n\n"
        f"Uzum API: {'✅ подключён' if connected else '❌ не подключён'}\n"
        f"Магазинов: {len(shops)}\n"
        f"Основной магазин: {f'<code>{default_shop_id}</code>' if default_shop_id else 'не выбран'}\n"
        f"Подписка: {subscription_status_text(telegram_id)}\n",
        reply_markup=menu_for_message(message),
    )


def api_already_connected_text(lang: str) -> str:
    if lang == "uz":
        return (
            "✅ <b>Do‘kon allaqachon ulangan</b>\n\n"
            "Amaldagi API-kalit xavfsiz saqlangan. Uni almashtirish kerak bo‘lsa:\n"
            "<b>⚙️ Sozlamalar → 🔐 Uzum ulanishi → 🔌 API-kalitni yangilash</b>."
        )
    return (
        "✅ <b>Магазин уже подключён</b>\n\n"
        "Текущий API-ключ сохранён и продолжает работать. Если его нужно заменить:\n"
        "<b>⚙️ Настройки → 🔐 Подключение Uzum → 🔌 Обновить API-ключ</b>."
    )


@dp.message(Command("connect_shop", "staff_connect", "addshop"))
async def connect_shop_command(message: Message, state: FSMContext) -> None:
    # Подключение только через API-ключ продавца.
    # Безопасность: если API уже подключён, обычная кнопка/команда не переводит пользователя
    # в режим замены и не трогает старый ключ. Для замены есть отдельная команда /reconnect.
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    if db.has_uzum_connection(telegram_id):
        await state.clear()
        await message.answer(api_already_connected_text(lang), reply_markup=menu_for_message(message))
        return

    await state.set_state(ConnectStates.waiting_for_token)
    if lang == "uz":
        text = (
            "🔌 <b>Do‘konni ulash</b>\n\n"
            "Uzum Seller kabinetida yaratilgan API-kalitni nusxalang va keyingi xabarda yuboring.\n\n"
            "🔐 Kalit Telegram chatida ko‘rsatilmaydi va bot uni ishlashdan oldin tekshiradi."
        )
    else:
        text = (
            "🔌 <b>Подключение магазина</b>\n\n"
            "Скопируйте API-ключ из кабинета Uzum Seller и отправьте его следующим сообщением.\n\n"
            "🔐 Ключ не показывается в интерфейсе и будет проверен до сохранения."
        )
    await message.answer(
        text,
        reply_markup=CONNECT_INPUT_MENU_UZ if lang == "uz" else CONNECT_INPUT_MENU_RU,
    )


@dp.message(F.text.in_({"🔌 Подключить", "🔌 Подключить магазин", "🔌 Ulash", "🔌 Do‘konni ulash"}))
async def connect_shop_button(message: Message, state: FSMContext) -> None:
    await connect_shop_command(message, state)


@dp.message(F.text.in_({"🔌 Обновить API-ключ", "🔌 API-kalitni yangilash"}))
async def reconnect_shop_button(message: Message, state: FSMContext) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    await state.set_state(ConnectStates.waiting_for_token)
    if lang == "uz":
        text = (
            "🔁 <b>API-kalitni yangilash</b>\n\n"
            "Yangi API-kalitni keyingi xabarda yuboring. Eski ulanish faqat yangi kalit "
            "muvaffaqiyatli tekshirilgandan keyin almashtiriladi."
        )
    else:
        text = (
            "🔁 <b>Обновление API-ключа</b>\n\n"
            "Отправьте новый API-ключ следующим сообщением. Старое подключение заменится "
            "только после успешной проверки нового ключа."
        )
    await message.answer(
        text,
        reply_markup=CONNECT_INPUT_MENU_UZ if lang == "uz" else CONNECT_INPUT_MENU_RU,
    )


@dp.message(ConnectStates.waiting_for_token, F.text.in_({"❌ Отмена", "❌ Bekor qilish"}))
async def cancel_connect_button(message: Message, state: FSMContext) -> None:
    telegram_id = upsert_from_message(message)
    await state.clear()
    await message.answer(
        "Bekor qilindi." if get_user_language(telegram_id) == "uz" else "Подключение отменено.",
        reply_markup=menu_for_message(message),
    )


@dp.message(ConnectStates.waiting_for_shop_id, F.text)
async def connect_waiting_shop_id(message: Message, state: FSMContext) -> None:
    await connect_token(message, message.text or "", state)


@dp.message(Command("connect", "reconnect"))
async def connect(message: Message, state: FSMContext) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    raw_text = (message.text or "").strip()
    command = raw_text.split()[0].lower() if raw_text.startswith("/") else ""
    is_reconnect = command.startswith("/reconnect")
    token = parse_args(raw_text)

    # Безопасность: /connect не заменяет уже подключённый ключ.
    # Замена разрешена только через явную команду /reconnect или если пользователь сразу передал токен в /reconnect <token>.
    if db.has_uzum_connection(telegram_id) and not is_reconnect:
        await state.clear()
        await message.answer(api_already_connected_text(lang), reply_markup=menu_for_message(message))
        return

    if token:
        await connect_token(message, token, state)
        return

    await state.set_state(ConnectStates.waiting_for_token)
    if lang == "uz":
        title = "🔁 <b>API-kalitni almashtirish</b>" if is_reconnect else "🔑 <b>API-kalitni ulash</b>"
        text_connect = (
            f"{title}\n\n"
            "Uzum Seller kabinetidan olingan API-kalitni keyingi xabarda yuboring.\n\n"
            "Kalitni qayerdan olish: <code>/api_token</code>\n\n"
            "Muhim:\n"
            "• API-kalit kabinet paroli emas;\n"
            "• eski kalit faqat yangi kalit muvaffaqiyatli tekshirilgandan keyin almashtiriladi;\n"
            "• tekshiruvdan so‘ng bot kalit yuborilgan xabarni o‘chirishga harakat qiladi;\n"
            "• bekor qilish: <code>/cancel</code>."
        )
    else:
        title = "🔁 <b>Замена API-ключа</b>" if is_reconnect else "🔑 <b>Подключение API-ключа</b>"
        text_connect = (
            f"{title}\n\n"
            "Отправьте следующим сообщением API-ключ из кабинета Uzum Seller.\n\n"
            "Где взять ключ: <code>/api_token</code>\n\n"
            "Важно:\n"
            "• API-ключ — это не пароль от кабинета;\n"
            "• старый ключ заменится только после успешной проверки нового;\n"
            "• после проверки бот постарается удалить сообщение с ключом;\n"
            "• отменить: <code>/cancel</code>."
        )
    await message.answer(text_connect, reply_markup=menu_for_message(message))


@dp.message(ConnectStates.waiting_for_token, F.text)
async def connect_waiting_token(message: Message, state: FSMContext) -> None:
    # Пока бот ждёт API-ключ, команды помощи не должны восприниматься как токен.
    # Иначе /video или /api_token попадают в проверку токена и пользователь видит ошибку.
    raw_text = (message.text or "").strip()
    command = raw_text.split()[0].lower() if raw_text.startswith("/") else ""

    if command in {"/video", "/api_video", "/instruction"}:
        await video_instruction(message)
        return

    if command in {"/api_token", "/token_help", "/how_token"}:
        await api_token_help(message)
        return

    if command == "/cancel":
        await cancel(message, state)
        return

    if command == "/menu":
        await state.clear()
        await menu(message)
        return

    if raw_text.startswith("/"):
        telegram_id = upsert_from_message(message)
        lang = get_user_language(telegram_id)
        if lang == "uz":
            await message.answer(
                "Hozir bot API-kalitni kutyapti.\n\n"
                "API-kalitni yuboring yoki bekor qilish uchun <code>/cancel</code> bosing.\n"
                "Yordam: <code>/video</code> yoki <code>/api_token</code>",
                reply_markup=menu_for_message(message),
            )
        else:
            await message.answer(
                "Сейчас бот ждёт API-ключ.\n\n"
                "Отправьте API-ключ или нажмите <code>/cancel</code>, чтобы отменить подключение.\n"
                "Помощь: <code>/video</code> или <code>/api_token</code>",
                reply_markup=menu_for_message(message),
            )
        return

    await connect_token(message, raw_text, state)


def disconnect_uzum_for_user(telegram_id: int) -> None:
    if hasattr(db, "disconnect_uzum"):
        db.disconnect_uzum(telegram_id)
        return
    with db.connect() as conn:
        conn.execute(
            "UPDATE users SET uzum_token_encrypted = NULL, default_shop_id = NULL, updated_at = ? WHERE telegram_id = ?",
            (_dt_to_db(_utc_now()), int(telegram_id)),
        )
        try:
            conn.execute("DELETE FROM shops WHERE telegram_id = ?", (int(telegram_id),))
        except Exception:
            pass
        conn.commit()


@dp.message(Command("disconnect"))
async def disconnect(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    disconnect_uzum_for_user(telegram_id)
    await message.answer(
        "✅ Подключение к Uzum API удалено. Можно подключить заново через <code>/connect</code>.",
        reply_markup=menu_for_message(message),
    )


@dp.message(Command("pinguzum"))
async def ping_uzum(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    client = get_uzum_for_user(telegram_id)
    if client is None:
        await message.answer(
            "Сначала подключите Uzum API-токен: <code>/connect</code>",
            reply_markup=menu_for_message(message),
        )
        return
    try:
        data = await client.get_shops()
        shops = extract_items(data)
        await message.answer(f"✅ Uzum API отвечает. Найдено магазинов: {len(shops)}", reply_markup=menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)




def shop_selection_markup(
    shops_list: list[dict[str, Any]],
    current_shop_id: int | None,
    lang: str = "ru",
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for shop in shops_list[:30]:
        shop_id = _shop_id_from_obj(shop)
        if shop_id is None:
            continue

        name = ""
        for key in ("title", "name", "shopName", "storeName", "displayName"):
            value = shop.get(key)
            if value not in (None, ""):
                name = str(value).strip()
                break

        if not name:
            name = f"Do‘kon {shop_id}" if normalize_lang(lang) == "uz" else f"Магазин {shop_id}"

        name = name[:42]
        marker = "✅" if current_shop_id and int(current_shop_id) == int(shop_id) else "▫️"
        rows.append([
            InlineKeyboardButton(
                text=f"{marker} {name} · {shop_id}",
                callback_data=f"shop_select:{shop_id}",
            )
        ])

    return InlineKeyboardMarkup(inline_keyboard=rows)


@dp.callback_query(F.data.startswith("shop_select:"))
async def select_shop_callback(callback: CallbackQuery) -> None:
    import re
    user = callback.from_user
    if not user:
        await callback.answer("Не удалось определить пользователя", show_alert=True)
        return

    db.upsert_user(user.id, user.username, user.first_name)
    raw_shop_id = str(callback.data or "").split(":", 1)[-1]
    if not raw_shop_id.isdigit():
        await callback.answer("Некорректный магазин", show_alert=True)
        return

    shop_id = int(raw_shop_id)
    if not db.set_default_shop_id(user.id, shop_id):
        lang = get_user_language(user.id)
        error_text = (
            "Bu do‘kon topilmadi. Do‘konlar ro‘yxatini qayta oching."
            if lang == "uz"
            else "Этот магазин не найден. Откройте список магазинов заново."
        )
        await callback.answer(error_text, show_alert=True)
        return

    try:
        if callback.message and callback.message.reply_markup:
            new_rows: list[list[InlineKeyboardButton]] = []
            for row in callback.message.reply_markup.inline_keyboard:
                new_row: list[InlineKeyboardButton] = []
                for button in row:
                    button_data = str(button.callback_data or "")
                    label = str(button.text or "")
                    label = re.sub(r"^(✅|▫️)\s*", "", label)
                    selected = button_data == f"shop_select:{shop_id}"
                    new_row.append(
                        InlineKeyboardButton(
                            text=f"{'✅' if selected else '▫️'} {label}",
                            callback_data=button.callback_data,
                        )
                    )
                new_rows.append(new_row)
            await callback.message.edit_reply_markup(
                reply_markup=InlineKeyboardMarkup(inline_keyboard=new_rows)
            )
    except Exception:
        logging.exception("Failed to refresh shop selection keyboard")

    lang = get_user_language(user.id)
    await callback.answer("Do‘kon tanlandi" if lang == "uz" else "Магазин выбран")

    if callback.message:
        confirmation = (
            f"✅ Asosiy do‘kon tanlandi: <code>{shop_id}</code>"
            if lang == "uz"
            else f"✅ Основной магазин выбран: <code>{shop_id}</code>"
        )
        await callback.message.answer(
            confirmation,
            reply_markup=main_menu_for_user(user.id),
        )


@dp.message(Command("shops"))
async def shops(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    client = get_uzum_for_user(telegram_id)
    if client is None:
        text_no_api = (
            "Avval Uzum API-kalitini ulang: <code>/connect</code>"
            if lang == "uz"
            else "Сначала подключите Uzum API-ключ: <code>/connect</code>"
        )
        await message.answer(text_no_api, reply_markup=menu_for_message(message))
        return

    try:
        data = await client.get_shops()
        items = extract_items(data)
        if not items:
            no_shops_text = (
                "API javob berdi, lekin do‘konlar topilmadi."
                if lang == "uz"
                else "API ответил, но магазины не найдены."
            )
            await message.answer(no_shops_text, reply_markup=menu_for_message(message))
            return

        encrypted = db.get_encrypted_token(telegram_id)
        if encrypted:
            db.save_connection(telegram_id, encrypted, items)

        current = db.get_default_shop_id(telegram_id)
        markup = shop_selection_markup(items, current, lang)

        if lang == "uz":
            screen_text = (
                "🏪 <b>Do‘konlaringiz</b>\n\n"
                "Kerakli do‘konni pastdagi tugma orqali tanlang.\n"
                "✅ belgisi hozirgi asosiy do‘konni ko‘rsatadi."
            )
        else:
            screen_text = (
                "🏪 <b>Ваши магазины</b>\n\n"
                "Выберите нужный магазин кнопкой ниже.\n"
                "✅ отмечен текущий основной магазин."
            )

        await message.answer(screen_text, reply_markup=markup)
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("setshop"))
async def setshop(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    arg = parse_args(message.text or "")
    if not arg.isdigit():
        await message.answer(
            "Напишите так: <code>/setshop SHOP_ID</code>\nНапример: <code>/setshop 12345</code>",
            reply_markup=menu_for_message(message),
        )
        return

    shop_id = int(arg)
    ok = db.set_default_shop_id(telegram_id, shop_id)
    if not ok:
        await message.answer("Этот магазин не найден среди подключённых. Сначала обновите список: <code>/shops</code>", reply_markup=menu_for_message(message))
        return

    await message.answer(f"✅ Основной магазин выбран: <code>{shop_id}</code>", reply_markup=menu_for_message(message))


async def send_stock_list(message: Message, *, mode: str = "all") -> None:
    """Load and display SKU-level stock for all, FBO or FBS inventory."""
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    safe_mode = mode if mode in {"all", "fbo", "fbs"} else "all"

    try:
        rows = await load_sku_rows(client, shop_id, max_pages=50)
        if safe_mode == "fbo":
            visible_rows = [row for row in rows if float(row.get("fbo") or 0) > 0]
        elif safe_mode == "fbs":
            visible_rows = [row for row in rows if float(row.get("fbs") or 0) > 0]
        else:
            visible_rows = rows

        title_by_mode = {
            "all": ("📦 <b>Остатки по SKU</b>", "📦 <b>SKU qoldiqlari</b>"),
            "fbo": ("🏭 <b>Остатки FBO</b>", "🏭 <b>FBO qoldiqlari</b>"),
            "fbs": ("🏠 <b>Остатки FBS/DBS</b>", "🏠 <b>FBS/DBS qoldiqlari</b>"),
        }
        ru_title, uz_title = title_by_mode[safe_mode]
        total_fbo = sum(max(0.0, float(row.get("fbo") or 0)) for row in visible_rows)
        total_fbs = sum(max(0.0, float(row.get("fbs") or 0)) for row in visible_rows)
        total_units = sum(max(0.0, float(row.get("total") or 0)) for row in visible_rows)
        if lang == "uz":
            summary = [
                f"🏪 Do‘kon: <code>{shop_id}</code>",
                f"🔖 SKU: <b>{len(visible_rows)}</b>",
                f"📦 FBO: <b>{clean_num(total_fbo)}</b> · FBS/DBS: <b>{clean_num(total_fbs)}</b> · Jami: <b>{clean_num(total_units)}</b>",
            ]
            empty_text = (
                "FBO omborida musbat qoldiqli SKU topilmadi."
                if safe_mode == "fbo"
                else "FBS/DBS omborida musbat qoldiqli SKU topilmadi."
                if safe_mode == "fbs"
                else "SKU qoldiqlari topilmadi."
            )
        else:
            summary = [
                f"🏪 Магазин: <code>{shop_id}</code>",
                f"🔖 SKU: <b>{len(visible_rows)}</b>",
                f"📦 FBO: <b>{clean_num(total_fbo)}</b> · FBS/DBS: <b>{clean_num(total_fbs)}</b> · Всего: <b>{clean_num(total_units)}</b>",
            ]
            empty_text = (
                "Не найдено SKU с положительным остатком FBO."
                if safe_mode == "fbo"
                else "Не найдено SKU с положительным остатком FBS/DBS."
                if safe_mode == "fbs"
                else "SKU-остатки не найдены."
            )

        items = [format_sku_stock_line(row, mode=safe_mode) for row in visible_rows]
        await send_paginated_list(
            message,
            kind=f"stock_{safe_mode}",
            title=uz_title if lang == "uz" else ru_title,
            items=items,
            summary=summary,
            empty_text=empty_text,
            section="stock",
            page_size=5,
            reply_markup=stock_menu_for_message(message),
        )
    except Exception as error:
        await send_api_error(message, error)


@dp.message(Command("products"))
async def products(message: Message) -> None:
    """Показывает варианты SKU с корректным разделением FBO и FBS/DBS."""
    await send_stock_list(message, mode="all")


@dp.message(Command("stock"))
async def stock(message: Message) -> None:
    await send_stock_list(message, mode="all")


@dp.message(Command("fbo"))
async def fbo(message: Message) -> None:
    await send_stock_list(message, mode="fbo")


@dp.message(Command("fbs"))
async def fbs(message: Message) -> None:
    await send_stock_list(message, mode="fbs")


@dp.message(Command("lowstock"))
async def lowstock(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    telegram_id = message.from_user.id if message.from_user else 0
    lang = get_user_language(telegram_id)
    arg = parse_args(message.text or "")
    settings = ensure_product_settings(telegram_id)
    threshold = (
        min(100_000, int(arg))
        if arg.isdigit()
        else max(0, int(settings.get("low_stock_threshold") or LOW_STOCK_THRESHOLD))
    )

    try:
        rows = await load_sku_rows(client, shop_id, max_pages=50)
        if not rows:
            await message.answer("SKU qoldiqlari topilmadi." if lang == "uz" else "SKU-остатки не найдены.", reply_markup=stock_menu_for_message(message))
            return

        low = [r for r in rows if r.get("total") is not None and r["total"] <= threshold]
        if not low:
            text = f"✅ Umumiy qoldiq ≤ {threshold} bo‘lgan SKU topilmadi." if lang == "uz" else f"✅ Товаров с общим остатком ≤ {threshold} не найдено."
            await message.answer(text, reply_markup=stock_menu_for_message(message))
            return

        items = [format_sku_stock_line(row, mode="all") for row in low]
        title = "⚠️ <b>Kam qoldiqdagi tovarlar</b>" if lang == "uz" else "⚠️ <b>Низкие остатки</b>"
        summary = [
            f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
            f"📉 Chegara: ≤ <b>{threshold}</b> dona" if lang == "uz" else f"📉 Порог: ≤ <b>{threshold}</b> шт.",
        ]
        await send_paginated_list(message, kind="lowstock", title=title, summary=summary, items=items, section="stock", reply_markup=stock_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("orders"))
async def orders(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    status = parse_args(message.text or "").upper() or "CREATED"

    try:
        data = await client.get_fbs_orders(shop_id, status=status, page=0, size=10)
        items = extract_items(data)
        if not items:
            await message.answer(f"Заказы со статусом <code>{escape(status)}</code> не найдены.", reply_markup=menu_for_message(message))
            return

        lines = [format_order_line(item) for item in items[:10]]
        await message.answer(
            f"🛒 <b>Заказы {escape(status)} для магазина</b> <code>{shop_id}</code>:\n\n"
            + "\n".join(lines),
            reply_markup=menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)



# --- Продажи / Финансы ---
# Используем официальный Finance endpoint Uzum Seller OpenAPI:
# GET /v1/finance/orders?shopIds=...&dateFrom=...&dateTo=...

def _epoch_ms(dt: datetime) -> int:
    return int(dt.timestamp() * 1000)


UZT = timezone(timedelta(hours=5))


def _today_range_ms() -> tuple[int, int]:
    # Считаем день по времени Узбекистана, а не по UTC.
    now = datetime.now(UZT)
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    return _epoch_ms(start), _epoch_ms(now)


def _days_range_ms(days: int) -> tuple[int, int]:
    # 7 дней = с начала дня 6 дней назад до текущего момента по Ташкенту.
    now = datetime.now(UZT)
    start_today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    start = start_today - timedelta(days=max(1, days) - 1)
    return _epoch_ms(start), _epoch_ms(now)


def _yesterday_range_ms() -> tuple[int, int]:
    # Вчера = прошлый полный день по времени Узбекистана.
    now = datetime.now(UZT)
    start_today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    start_yesterday = start_today - timedelta(days=1)
    return _epoch_ms(start_yesterday), _epoch_ms(start_today)


def _last_7_days_range_ms() -> tuple[int, int]:
    return _days_range_ms(7)


def _deep_items(obj: Any) -> list[dict[str, Any]]:
    """Достаём список строк из разных возможных форматов ответа Uzum."""
    direct = extract_items(obj)
    if direct:
        return [x for x in direct if isinstance(x, dict)]

    keys = (
        "orderItems",
        "orders",
        "items",
        "content",
        "data",
        "payload",
        "result",
        "list",
        "financeOrders",
        "sellerOrders",
    )
    found: list[dict[str, Any]] = []

    def walk(x: Any) -> None:
        if isinstance(x, dict):
            for k in keys:
                v = x.get(k)
                if isinstance(v, list):
                    for item in v:
                        if isinstance(item, dict):
                            found.append(item)
                elif isinstance(v, dict):
                    walk(v)
        elif isinstance(x, list):
            for item in x:
                if isinstance(item, dict):
                    found.append(item)

    walk(obj)
    # Убираем явные дубли по JSON-представлению.
    unique: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in found:
        sig = json.dumps(item, ensure_ascii=False, sort_keys=True, default=str)[:1000]
        if sig not in seen:
            seen.add(sig)
            unique.append(item)
    return unique


def _num_from_value(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = (
            value.replace(" ", "")
            .replace("\u00a0", "")
            .replace("сум", "")
            .replace("UZS", "")
            .replace(",", ".")
        )
        try:
            return float(cleaned)
        except Exception:
            return None
    return None


def _pick_number(item: dict[str, Any], names: tuple[str, ...]) -> float | None:
    for name in names:
        value = item.get(name)
        number = _num_from_value(value)
        if number is not None:
            return number
    return None


def _finance_status(item: dict[str, Any]) -> str:
    value = pick(
        item,
        "status",
        "orderStatus",
        "financeStatus",
        "state",
        "statusName",
        "statusTitle",
    )
    if isinstance(value, dict):
        value = pick(value, "title", "name", "value", "code")
    return str(value or "UNKNOWN")


def _finance_title(item: dict[str, Any]) -> str:
    value = pick(
        item,
        "skuTitle",
        "productTitle",
        "productName",
        "title",
        "name",
        "skuName",
        "offerName",
    )
    if isinstance(value, dict):
        value = pick(value, "title", "name")
    return str(value or "Без названия")


def _finance_qty(item: dict[str, Any]) -> float:
    return _pick_number(
        item,
        (
            "quantity",
            "amount",
            "count",
            "qty",
            "skuAmount",
            "productAmount",
            "quantityPurchased",
        ),
    ) or 1.0


def _finance_revenue(item: dict[str, Any]) -> float:
    # Пробуем готовые суммы.
    direct = _pick_number(
        item,
        (
            "totalPrice",
            "totalAmount",
            "totalSum",
            "sellerAmount",
            "sellerPrice",
            "totalSellerPrice",
            "purchasePrice",
            "priceWithDiscount",
            "orderItemPrice",
            "orderPrice",
            "amountToWithdraw",
            "accrual",
            "sum",
        ),
    )
    if direct is not None:
        return max(0.0, direct)

    # Если есть только цена за штуку — умножаем на количество.
    price = _pick_number(item, ("price", "itemPrice", "skuPrice", "sellPrice"))
    if price is not None:
        return max(0.0, price * _finance_qty(item))
    return 0.0


def _is_cancelled_status(status: str) -> bool:
    s = status.upper()
    return "CANCEL" in s or "ОТМЕН" in s


def _format_money(value: float) -> str:
    return f"{value:,.0f}".replace(",", " ") + " сум"


async def _finance_orders_request(
    client: UzumClient,
    shop_id: int,
    *,
    date_from_ms: int,
    date_to_ms: int,
    page: int = 0,
    size: int = 100,
) -> Any:
    params = [
        ("page", page),
        ("size", size),
        ("group", "false"),
        # Важно: рабочий Noorza Bot использует dateFrom в секундах, dateTo в миллисекундах.
        # Если отправить dateFrom в миллисекундах, Uzum Finance может вернуть 0 строк.
        ("dateFrom", int(date_from_ms / 1000)),
        ("dateTo", date_to_ms),
        ("shopIds", shop_id),
    ]
    path = "/v1/finance/orders?" + urlencode(params)
    return await client._request("GET", path)


async def _load_finance_orders(
    client: UzumClient,
    shop_id: int,
    *,
    date_from_ms: int,
    date_to_ms: int,
    max_pages: int = 10,
    page_size: int = 100,
) -> tuple[list[dict[str, Any]], Any | None]:
    rows: list[dict[str, Any]] = []
    first_response: Any | None = None
    for page in range(max_pages):
        data = await _finance_orders_request(
            client,
            shop_id,
            date_from_ms=date_from_ms,
            date_to_ms=date_to_ms,
            page=page,
            size=page_size,
        )
        if first_response is None:
            first_response = data
        items = _deep_items(data)
        if not items:
            break
        rows.extend(items)
        if len(items) < page_size:
            break
    return rows, first_response


def _build_sales_stats(rows: list[dict[str, Any]]) -> dict[str, Any]:
    total_rows = len(rows)
    cancelled_rows = 0
    revenue = 0.0
    units = 0.0
    statuses: dict[str, int] = {}
    products: dict[str, dict[str, float | str]] = {}

    for item in rows:
        status = _finance_status(item)
        statuses[status] = statuses.get(status, 0) + 1
        qty = _finance_qty(item)
        amount = _finance_revenue(item)
        if _is_cancelled_status(status):
            cancelled_rows += 1
            continue
        revenue += amount
        units += qty
        title = _finance_title(item)
        if title not in products:
            products[title] = {"title": title, "qty": 0.0, "revenue": 0.0}
        products[title]["qty"] = float(products[title]["qty"]) + qty
        products[title]["revenue"] = float(products[title]["revenue"]) + amount

    top_products = sorted(
        products.values(), key=lambda x: float(x.get("revenue") or 0), reverse=True
    )[:5]
    avg = revenue / max(1, (total_rows - cancelled_rows))
    return {
        "rows": total_rows,
        "cancelled": cancelled_rows,
        "active_rows": max(0, total_rows - cancelled_rows),
        "revenue": revenue,
        "units": units,
        "avg": avg,
        "statuses": statuses,
        "top_products": top_products,
    }


def _short_period_title(days: int) -> str:
    if days == 1:
        return "сегодня"
    return f"за {days} дней"


async def _sales_period_stats(
    client: UzumClient, shop_id: int, days: int
) -> tuple[dict[str, Any], Any | None]:
    if days == 1:
        date_from, date_to = _today_range_ms()
    else:
        date_from, date_to = _days_range_ms(days)
    rows, first = await _load_finance_orders(
        client, shop_id, date_from_ms=date_from, date_to_ms=date_to
    )
    return _build_sales_stats(rows), first


def _format_sales_summary_line(title: str, stats: dict[str, Any]) -> str:
    return (
        f"<b>{escape(title)}</b>\n"
        f"• Выручка: <b>{_format_money(float(stats['revenue']))}</b>\n"
        f"• Позиций/строк: <b>{stats['active_rows']}</b>\n"
        f"• Штук: <b>{float(stats['units']):.0f}</b>\n"
        f"• Средняя строка: <b>{_format_money(float(stats['avg']))}</b>"
    )


def _format_sales_details(days: int, shop_id: int, stats: dict[str, Any], first_response: Any | None) -> str:
    title = _short_period_title(days)
    lines = [
        f"💰 <b>Продажи {title}</b>",
        f"Магазин: <code>{shop_id}</code>",
        "",
        f"Выручка: <b>{_format_money(float(stats['revenue']))}</b>",
        f"Проданных строк/позиций: <b>{stats['active_rows']}</b>",
        f"Штук: <b>{float(stats['units']):.0f}</b>",
        f"Средняя строка: <b>{_format_money(float(stats['avg']))}</b>",
        f"Отменённых строк: <b>{stats['cancelled']}</b>",
    ]

    top_products = stats.get("top_products") or []
    if top_products:
        lines.append("")
        lines.append("<b>Топ товаров по сумме:</b>")
        for idx, item in enumerate(top_products, start=1):
            title_item = str(item.get("title") or "Без названия")
            if len(title_item) > 70:
                title_item = title_item[:67] + "..."
            lines.append(
                f"{idx}. {escape(title_item)} — "
                f"{float(item.get('qty') or 0):.0f} шт, "
                f"{_format_money(float(item.get('revenue') or 0))}"
            )

    if stats.get("rows") == 0 and first_response is not None:
        lines.append("")
        lines.append("Ответ Finance API пришёл, но строки продаж не найдены.")
        lines.append("Фрагмент ответа:")
        lines.append("<code>" + escape(compact_json_preview(first_response)) + "</code>")

    return "\n".join(lines)


# --- Короткие разделы в стиле Noorza Bot ---
# Блок "Сегодня" работает в стиле второго бота: берёт Finance API за текущий день
# и показывает выручку, комиссию, логистику и к выплате.

def _deep_pick_number(obj: Any, names: tuple[str, ...]) -> float | None:
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k in names:
                n = _num_from_value(v)
                if n is not None:
                    return n
        for v in obj.values():
            n = _deep_pick_number(v, names)
            if n is not None:
                return n
    elif isinstance(obj, list):
        for v in obj:
            n = _deep_pick_number(v, names)
            if n is not None:
                return n
    return None


def _finance_gross_revenue(item: dict[str, Any]) -> float:
    # В Finance API Uzum поле sellPrice обычно является ценой за 1 штуку.
    # Поэтому для выручки умножаем sellPrice на amount, как в рабочем Noorza Bot.
    direct_total = _deep_pick_number(
        item,
        (
            "totalPrice", "totalAmount", "totalSum", "totalSellerPrice",
            "orderItemPrice", "orderPrice", "sellerPrice", "sellerAmount",
        ),
    )
    if direct_total is not None:
        return max(0.0, direct_total)

    unit_price = _deep_pick_number(
        item,
        (
            "sellPrice", "soldPrice", "productPrice", "skuPrice",
            "priceWithDiscount", "purchasePrice", "price",
        ),
    )
    if unit_price is not None:
        return max(0.0, unit_price * _finance_qty(item))
    return _finance_revenue(item)


def _finance_commission(item: dict[str, Any]) -> float:
    value = _deep_pick_number(
        item,
        (
            "commission", "commissionAmount", "commissionSum", "uzumCommission",
            "marketplaceCommission", "sellerCommission", "fee", "feeAmount",
        ),
    )
    return abs(value or 0.0)


def _finance_logistics(item: dict[str, Any]) -> float:
    value = _deep_pick_number(
        item,
        (
            "logisticDeliveryFee", "logistics", "logistic", "logisticAmount", "logisticsAmount",
            "logisticsSum", "delivery", "deliveryAmount", "deliveryPrice",
            "deliveryCost", "shipping", "shippingAmount",
        ),
    )
    return abs(value or 0.0)


def _finance_payout_direct(item: dict[str, Any]) -> float | None:
    return _deep_pick_number(
        item,
        (
            "sellerProfit", "amountToWithdraw", "toWithdraw", "withdrawAmount", "sellerPayout",
            "payout", "payoutAmount", "sellerAmount", "accrual", "accrualAmount",
        ),
    )


def _finance_withdrawn(item: dict[str, Any]) -> float:
    value = _deep_pick_number(
        item,
        (
            "withdrawnProfit", "withdrawn", "withdrawnAmount", "paid", "paidAmount", "transferred",
            "transferredAmount", "alreadyWithdrawn",
        ),
    )
    return max(0.0, value or 0.0)


async def _finance_orders_request_extra(
    client: UzumClient,
    shop_id: int,
    *,
    date_from_ms: int,
    date_to_ms: int,
    extra_params: list[tuple[str, Any]] | None = None,
    page: int = 0,
    size: int = 100,
) -> Any:
    params: list[tuple[str, Any]] = [
        ("page", page),
        ("size", size),
        ("group", "false"),
        # Важно: рабочий Noorza Bot использует dateFrom в секундах, dateTo в миллисекундах.
        ("dateFrom", int(date_from_ms / 1000)),
        ("dateTo", date_to_ms),
        ("shopIds", shop_id),
    ]
    if extra_params:
        params.extend(extra_params)
    path = "/v1/finance/orders?" + urlencode(params)
    return await client._request("GET", path)


async def _load_today_finance_flexible(
    client: UzumClient, shop_id: int
) -> tuple[list[dict[str, Any]], Any | None, str]:
    date_from, date_to = _today_range_ms()
    return await _load_finance_range_flexible(client, shop_id, date_from, date_to)


async def _load_finance_range_flexible(
    client: UzumClient,
    shop_id: int,
    date_from_ms: int,
    date_to_ms: int,
    *,
    max_pages: int | None = None,
    page_size: int = 100,
) -> tuple[list[dict[str, Any]], Any | None, str]:
    """Load every Finance page using only documented query parameters.

    The old implementation requested only page 0, so Telegram and Excel silently
    understated totals when a period contained more than 100 rows.  Earlier
    fallback attempts also sent unsupported ``status``/``statuses`` parameters,
    which produced HTTP 400 responses when the selected period had no rows.
    """
    max_pages = max(1, int(max_pages or FINANCE_REPORT_MAX_PAGES))
    page_size = max(1, min(100, int(page_size or 100)))
    attempts: list[tuple[str, list[tuple[str, Any]]]] = [
        ("без дополнительных фильтров", []),
    ]
    all_rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    first_response: Any | None = None
    used_attempts: list[str] = []
    truncated = False
    for label, extra in attempts:
        try:
            attempt_rows = 0
            attempt_pages = 0
            for page in range(max_pages):
                data = await _finance_orders_request_extra(
                    client,
                    shop_id,
                    date_from_ms=date_from_ms,
                    date_to_ms=date_to_ms,
                    extra_params=extra,
                    page=page,
                    size=page_size,
                )
                attempt_pages += 1
                if first_response is None:
                    first_response = data
                rows = _deep_items(data)
                if not rows:
                    break
                attempt_rows += len(rows)
                for row in rows:
                    raw = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
                    sig = hashlib.sha256(raw.encode("utf-8")).hexdigest()
                    if sig not in seen:
                        seen.add(sig)
                        all_rows.append(row)
                if len(rows) < page_size:
                    break
                if page + 1 >= max_pages:
                    truncated = True
                await asyncio.sleep(0.04)
            used_attempts.append(f"{label}: {attempt_pages} стр./{attempt_rows} строк")
            if all_rows and label == "без дополнительных фильтров":
                break
            await asyncio.sleep(0.15)
        except Exception as e:
            logging.info("Finance attempt failed: %s: %s", label, e)
            continue
    if truncated:
        warning = f"Достигнут защитный лимит {max_pages * page_size} строк"
        used_attempts.append(warning)
        logging.warning(
            "Finance report reached row limit shop=%s from=%s to=%s limit=%s",
            shop_id,
            date_from_ms,
            date_to_ms,
            max_pages * page_size,
        )
    return all_rows, first_response, "; ".join(used_attempts)


UZUM_EXPENSE_MAX_PAGES = max(
    1,
    min(200, int(os.getenv("UZUM_EXPENSE_MAX_PAGES", "100") or "100")),
)
UZUM_EXPENSE_CACHE_SECONDS = max(
    30,
    min(1800, int(os.getenv("UZUM_EXPENSE_CACHE_SECONDS", "300") or "300")),
)
_UZUM_EXPENSE_CACHE: dict[tuple[str, int, int, int], tuple[float, dict[str, Any]]] = {}
_UZUM_EXPENSE_LOCKS: dict[tuple[str, int, int, int], tuple[Any, asyncio.Lock]] = {}


def _uzum_expense_cache_key(
    client: UzumClient,
    shop_id: int,
    date_from_ms: int,
    date_to_ms: int,
) -> tuple[str, int, int, int]:
    token = str((getattr(client, "headers", {}) or {}).get("Authorization") or "")
    fingerprint = hashlib.sha256(token.encode("utf-8")).hexdigest()[:20]
    return fingerprint, int(shop_id), int(date_from_ms), int(date_to_ms)


def _uzum_expense_lock(key: tuple[str, int, int, int]) -> asyncio.Lock:
    loop = asyncio.get_running_loop()
    stored = _UZUM_EXPENSE_LOCKS.get(key)
    if stored is None or stored[0] is not loop:
        lock = asyncio.Lock()
        _UZUM_EXPENSE_LOCKS[key] = (loop, lock)
        return lock
    return stored[1]


async def _load_uzum_expense_rows_attempt(
    client: UzumClient,
    shop_id: int,
    *,
    date_from: int,
    date_to: int,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in range(UZUM_EXPENSE_MAX_PAGES):
        data = await client.get_expenses(
            shop_id=int(shop_id),
            date_from=int(date_from),
            date_to=int(date_to),
            page=page,
            size=100,
        )
        rows = expense_items(data)
        if not rows:
            break
        for row in rows:
            raw = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
            signature = hashlib.sha256(raw.encode("utf-8")).hexdigest()
            if signature not in seen:
                seen.add(signature)
                result.append(row)
        if len(rows) < 100:
            break
        await asyncio.sleep(0.04)
    if len(result) >= UZUM_EXPENSE_MAX_PAGES * 100:
        logging.warning(
            "Uzum expenses reached safety limit shop=%s rows=%s",
            shop_id,
            len(result),
        )
    return result


async def load_uzum_expense_summary(
    client: UzumClient,
    shop_id: int,
    date_from_ms: int,
    date_to_ms: int,
    *,
    force: bool = False,
) -> dict[str, Any]:
    """Load booked Uzum expenses with timestamp compatibility and caching."""
    key = _uzum_expense_cache_key(client, shop_id, date_from_ms, date_to_ms)
    cached = _UZUM_EXPENSE_CACHE.get(key)
    if not force and cached and time.monotonic() - cached[0] < UZUM_EXPENSE_CACHE_SECONDS:
        return dict(cached[1])

    async with _uzum_expense_lock(key):
        cached = _UZUM_EXPENSE_CACHE.get(key)
        if not force and cached and time.monotonic() - cached[0] < UZUM_EXPENSE_CACHE_SECONDS:
            return dict(cached[1])
        errors: list[str] = []
        rows: list[dict[str, Any]] = []
        date_mode = "milliseconds"
        # Swagger declares int64 but deployments have historically differed on
        # dateFrom units. Try the documented millisecond range first, then the
        # compatibility form already required by Finance orders.
        attempts = (
            (int(date_from_ms), int(date_to_ms), "milliseconds"),
            (int(date_from_ms / 1000), int(date_to_ms), "seconds/milliseconds"),
        )
        for attempt_from, attempt_to, label in attempts:
            try:
                rows = await _load_uzum_expense_rows_attempt(
                    client,
                    shop_id,
                    date_from=attempt_from,
                    date_to=attempt_to,
                )
                date_mode = label
                if rows:
                    break
            except Exception as error:
                errors.append(f"{label}: {error}")
                logging.info(
                    "Uzum expenses attempt failed shop=%s mode=%s: %s",
                    shop_id,
                    label,
                    error,
                )
        if not rows and len(errors) == len(attempts):
            summary = {
                **summarize_expenses([]),
                "available": False,
                "date_mode": date_mode,
                "error": errors[-1] if errors else "Uzum expenses unavailable",
            }
            return summary
        summary = {
            **summarize_expenses(rows),
            "available": True,
            "date_mode": date_mode,
            "error": None,
        }
        _UZUM_EXPENSE_CACHE[key] = (time.monotonic(), dict(summary))
        if len(_UZUM_EXPENSE_CACHE) > 200:
            stale_before = time.monotonic() - UZUM_EXPENSE_CACHE_SECONDS * 3
            for cache_key, (saved_at, _) in list(_UZUM_EXPENSE_CACHE.items()):
                if saved_at < stale_before:
                    _UZUM_EXPENSE_CACHE.pop(cache_key, None)
                    _UZUM_EXPENSE_LOCKS.pop(cache_key, None)
        return summary


def _is_returned_status(status: str) -> bool:
    value = str(status or "").upper()
    return "RETURN" in value or "ВОЗВРАТ" in value or "QAYTAR" in value


def _finance_order_key_for_stats(item: dict[str, Any]) -> str:
    value = pick(
        item,
        "orderId",
        "order_id",
        "postingNumber",
        "orderNumber",
        "financeOrderId",
        "operationId",
        "id",
    )
    if isinstance(value, dict):
        value = pick(value, "id", "value", "number")
    if value not in (None, ""):
        return str(value)
    raw = json.dumps(item, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:20]


def _finance_sku_key_for_stats(item: dict[str, Any]) -> str:
    value = pick(
        item,
        "skuId",
        "sellerSku",
        "offerId",
        "barcode",
        "skuTitle",
        "skuName",
    )
    if isinstance(value, dict):
        value = pick(value, "id", "value", "title", "name")
    return str(value or _finance_title(item) or "—")


def _finance_return_qty(item: dict[str, Any], status: str | None = None) -> float:
    explicit = abs(
        _deep_pick_number(
            item,
            (
                "amountReturns",
                "returnAmount",
                "returnedAmount",
                "quantityReturns",
                "returnedQuantity",
            ),
        )
        or 0.0
    )
    if explicit > 0:
        return explicit
    if _is_returned_status(status or _finance_status(item)):
        return abs(_finance_qty(item))
    return 0.0


def _build_noorza_today_stats(rows: list[dict[str, Any]]) -> dict[str, Any]:
    # Логика максимально приближена к рабочему Noorza Bot:
    # sellPrice * amount, commission, logisticDeliveryFee, sellerProfit, withdrawnProfit.
    active_rows = 0
    cancelled_rows = 0
    returned_rows = 0
    returns = 0.0
    units = 0.0
    revenue = 0.0
    commission = 0.0
    logistics = 0.0
    payout_total = 0.0
    withdrawn = 0.0
    statuses: dict[str, int] = {}
    order_keys: set[str] = set()
    products: dict[str, dict[str, Any]] = {}
    for item in rows:
        status = _finance_status(item)
        qty = _finance_qty(item)
        statuses[status] = statuses.get(status, 0) + 1
        if _is_cancelled_status(status):
            cancelled_rows += 1
            continue
        returned_qty = _finance_return_qty(item, status)
        returns += returned_qty
        if _is_returned_status(status):
            returned_rows += 1
            continue

        active_rows += 1
        order_keys.add(_finance_order_key_for_stats(item))
        units += qty

        gross = _finance_gross_revenue(item)
        comm = _finance_commission(item)
        logi = _finance_logistics(item)
        payout_direct = _finance_payout_direct(item)
        payout = payout_direct if payout_direct is not None else max(0.0, gross - comm - logi)

        revenue += gross
        commission += comm
        logistics += logi
        payout_total += max(0.0, payout)
        withdrawn += _finance_withdrawn(item)
        sku = _finance_sku_key_for_stats(item)
        product = products.setdefault(
            sku,
            {
                "title": _finance_title(item),
                "sku": sku,
                "qty": 0.0,
                "revenue": 0.0,
                "payout": 0.0,
            },
        )
        product["qty"] = float(product["qty"]) + qty
        product["revenue"] = float(product["revenue"]) + gross
        product["payout"] = float(product["payout"]) + max(0.0, payout)
    orders = len(order_keys)
    top_products = sorted(
        products.values(),
        key=lambda value: float(value.get("revenue") or 0),
        reverse=True,
    )[:5]
    return {
        "source_rows": len(rows),
        "rows": active_rows,
        "orders": orders,
        "cancelled": cancelled_rows,
        "returned_rows": returned_rows,
        "units": units,
        "returns": returns,
        "revenue": revenue,
        "commission": commission,
        "logistics": logistics,
        "payout_total": payout_total,
        "withdrawn": withdrawn,
        "left_to_withdraw": max(0.0, payout_total - withdrawn),
        "average_order": revenue / max(1, orders),
        "average_unit": revenue / max(1.0, units),
        "commission_rate": commission / revenue if revenue > 0 else 0.0,
        "logistics_rate": logistics / revenue if revenue > 0 else 0.0,
        "cancellation_rate": cancelled_rows / max(1, active_rows + cancelled_rows + returned_rows),
        "statuses": statuses,
        "top_products": top_products,
    }


def _format_noorza_today(shop_id: int, stats: dict[str, Any], rows: list[dict[str, Any]]) -> str:
    extra = ""
    if not rows:
        extra = (
            "\n\n<i>Пока продаж за выбранный период не найдено. "
            "Если продажа только появилась в кабинете, она может отобразиться чуть позже.</i>"
        )
    return (
        "💰 <b>Продажи за сегодня</b>\n"
        f"🏪 Магазин: <code>{shop_id}</code>\n\n"
        f"🛒 Заказов: <b>{int(stats.get('orders') or 0)}</b>\n"
        f"🧾 Позиций: <b>{int(stats['rows'])}</b>\n"
        f"📦 Товаров продано: <b>{float(stats['units']):.0f} шт.</b>\n"
        f"❌ Отмен: <b>{int(stats.get('cancelled') or 0)}</b>\n"
        f"↩️ Возвратов: <b>{float(stats['returns']):.0f} шт.</b>\n\n"
        f"💵 Выручка: <b>{_format_money(float(stats['revenue']))}</b>\n"
        f"🏷 Комиссия Uzum: <b>{_format_money(float(stats['commission']))}</b> "
        f"({float(stats.get('commission_rate') or 0) * 100:.1f}%)\n"
        f"🚚 Логистика: <b>{_format_money(float(stats['logistics']))}</b> "
        f"({float(stats.get('logistics_rate') or 0) * 100:.1f}%)\n\n"
        f"✅ К выплате: <b>{_format_money(float(stats['payout_total']))}</b>\n"
        f"💳 Уже выведено: <b>{_format_money(float(stats['withdrawn']))}</b>\n"
        f"🧾 Остаток к выплате: <b>{_format_money(float(stats['left_to_withdraw']))}</b>\n"
        f"🛍 Средний чек: <b>{_format_money(float(stats.get('average_order') or 0))}</b>" + extra
    )


def _percent_change(current: float, previous: float) -> float | None:
    current = float(current or 0)
    previous = float(previous or 0)
    if abs(previous) < 0.000001:
        return None if abs(current) < 0.000001 else 1.0
    return (current - previous) / abs(previous)


def _format_change(value: float | None, *, lang: str = "ru") -> str:
    if value is None:
        return "—"
    if value >= 0:
        return f"📈 +{value * 100:.1f}%"
    return f"📉 {value * 100:.1f}%"


def _format_premium_period_report(
    title_ru: str,
    title_uz: str,
    shop_id: int,
    stats: dict[str, Any],
    rows: list[dict[str, Any]],
    previous_stats: dict[str, Any],
    profit_summary: dict[str, Any],
    business_profit: dict[str, Any] | None = None,
    *,
    lang: str = "ru",
) -> str:
    lang = normalize_lang(lang)
    revenue_change = _format_change(
        _percent_change(float(stats.get("revenue") or 0), float(previous_stats.get("revenue") or 0)),
        lang=lang,
    )
    payout_change = _format_change(
        _percent_change(float(stats.get("payout_total") or 0), float(previous_stats.get("payout_total") or 0)),
        lang=lang,
    )
    coverage = float(profit_summary.get("coverage") or 0)
    known_revenue = float(profit_summary.get("known_revenue") or 0)
    profit = float(profit_summary.get("profit") or 0)
    top_products = list(stats.get("top_products") or [])[:3]

    if lang == "uz":
        lines = [
            f"📊 <b>{escape(title_uz)}</b>",
            f"🏪 Do‘kon: <code>{shop_id}</code>",
            "",
            f"🛒 Buyurtmalar: <b>{int(stats.get('orders') or 0)}</b> | 🧾 Pozitsiyalar: <b>{int(stats.get('rows') or 0)}</b>",
            f"📦 Sotildi: <b>{float(stats.get('units') or 0):.0f} dona</b>",
            f"❌ Bekor qilish: <b>{int(stats.get('cancelled') or 0)}</b> | ↩️ Qaytarish: <b>{float(stats.get('returns') or 0):.0f} dona</b>",
            "",
            f"💵 Tushum: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
            f"🏷 Uzum komissiyasi: <b>{_format_money(float(stats.get('commission') or 0))}</b> ({float(stats.get('commission_rate') or 0) * 100:.1f}%)",
            f"🚚 Logistika: <b>{_format_money(float(stats.get('logistics') or 0))}</b> ({float(stats.get('logistics_rate') or 0) * 100:.1f}%)",
            f"✅ To‘lovga: <b>{_format_money(float(stats.get('payout_total') or 0))}</b>",
            f"🛍 O‘rtacha chek: <b>{_format_money(float(stats.get('average_order') or 0))}</b>",
            "",
            f"🔄 Oldingi shu davrga nisbatan: tushum {revenue_change}, to‘lov {payout_change}",
        ]
        if known_revenue > 0:
            if business_profit:
                lines.extend(["", *_format_profit_bridge_lines(business_profit, lang=lang)])
            else:
                lines.extend(["", f"💰 Hisobiy foyda: <b>{_format_money(profit)}</b>"])
            lines.append(f"📌 Tannarx bilan qamrov: <b>{coverage * 100:.1f}%</b>")
            if coverage < 0.999:
                lines.append("<i>Foyda faqat Uzum purchasePrice bergan savdolar bo‘yicha hisoblangan.</i>")
        else:
            lines.extend(["", "🧾 Uzum bu savdolar uchun purchasePrice bermadi. Bot tannarxni taxmin qilmaydi."])
        if top_products:
            lines.extend(["", "🏆 <b>Tushum bo‘yicha top-3:</b>"])
            for index, product in enumerate(top_products, start=1):
                lines.append(
                    f"{index}. {escape(_short_text(str(product.get('title') or product.get('sku') or '—'), 55))} — "
                    f"{float(product.get('qty') or 0):.0f} dona, {_format_money(float(product.get('revenue') or 0))}"
                )
    else:
        lines = [
            f"📊 <b>{escape(title_ru)}</b>",
            f"🏪 Магазин: <code>{shop_id}</code>",
            "",
            f"🛒 Заказов: <b>{int(stats.get('orders') or 0)}</b> | 🧾 Позиций: <b>{int(stats.get('rows') or 0)}</b>",
            f"📦 Продано: <b>{float(stats.get('units') or 0):.0f} шт.</b>",
            f"❌ Отмен: <b>{int(stats.get('cancelled') or 0)}</b> | ↩️ Возвратов: <b>{float(stats.get('returns') or 0):.0f} шт.</b>",
            "",
            f"💵 Выручка: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
            f"🏷 Комиссия Uzum: <b>{_format_money(float(stats.get('commission') or 0))}</b> ({float(stats.get('commission_rate') or 0) * 100:.1f}%)",
            f"🚚 Логистика: <b>{_format_money(float(stats.get('logistics') or 0))}</b> ({float(stats.get('logistics_rate') or 0) * 100:.1f}%)",
            f"✅ К выплате: <b>{_format_money(float(stats.get('payout_total') or 0))}</b>",
            f"🛍 Средний чек: <b>{_format_money(float(stats.get('average_order') or 0))}</b>",
            "",
            f"🔄 К предыдущему такому же периоду: выручка {revenue_change}, выплата {payout_change}",
        ]
        if known_revenue > 0:
            if business_profit:
                lines.extend(["", *_format_profit_bridge_lines(business_profit, lang=lang)])
            else:
                lines.extend(["", f"💰 Расчётная прибыль: <b>{_format_money(profit)}</b>"])
            lines.append(f"📌 Покрытие себестоимостью: <b>{coverage * 100:.1f}%</b>")
            if coverage < 0.999:
                lines.append("<i>Прибыль рассчитана только по продажам, для которых Uzum передал purchasePrice.</i>")
        else:
            lines.extend(["", "🧾 Uzum не передал purchasePrice для этих продаж. Бот не подставляет себестоимость приблизительно."])
        if top_products:
            lines.extend(["", "🏆 <b>Топ-3 по выручке:</b>"])
            for index, product in enumerate(top_products, start=1):
                lines.append(
                    f"{index}. {escape(_short_text(str(product.get('title') or product.get('sku') or '—'), 55))} — "
                    f"{float(product.get('qty') or 0):.0f} шт., {_format_money(float(product.get('revenue') or 0))}"
                )

    if not rows:
        lines.extend([
            "",
            "<i>Finance API пока не вернул продажи за выбранный период.</i>"
            if lang != "uz"
            else "<i>Finance API tanlangan davr uchun savdolarni hali qaytarmadi.</i>",
        ])
    return "\n".join(lines)


async def _send_premium_period_report(
    message: Message,
    *,
    title_ru: str,
    title_uz: str,
    date_from: int,
    date_to: int,
    comparison_shift_days: int,
    wait_ru: str,
    wait_uz: str,
) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    trial = subscription_access_level(telegram_id) == "trial"
    if trial:
        wait_text = "⌛ Bugungi savdolar yuklanmoqda..." if lang == "uz" else "⌛ Загружаю продажи за сегодня..."
    else:
        wait_text = wait_uz if lang == "uz" else wait_ru
    await message.answer(wait_text, reply_markup=sales_menu_for_message(message))
    try:
        rows, _, source_info = await _load_finance_range_flexible(
            client,
            shop_id,
            date_from,
            date_to,
        )
        stats = _build_noorza_today_stats(rows)
        if trial:
            # Trial demonstrates the core sales picture without exposing the
            # paid profit, comparison and top-product analytics.
            text = _format_noorza_today(shop_id, stats, rows)
            await message.answer(text, reply_markup=sales_menu_for_message(message))
            return

        shift_ms = max(1, int(comparison_shift_days)) * 24 * 60 * 60 * 1000
        previous_rows, _, _ = await _load_finance_range_flexible(
            client,
            shop_id,
            date_from - shift_ms,
            date_to - shift_ms,
        )
        previous_stats = _build_noorza_today_stats(previous_rows)
        await sync_uzum_sku_financials(client, telegram_id, shop_id)
        finance_settings = ensure_finance_settings(telegram_id, shop_id)
        costs = get_unit_cost_map(telegram_id, shop_id)
        unit_rows = _build_unit_rows_from_finance(
            rows,
            costs,
            tax_percent=float(finance_settings.get("tax_percent") or 0),
        )
        profit_summary = _profit_summary_from_unit_rows(unit_rows, stats)
        expense_summary = await load_uzum_expense_summary(
            client,
            shop_id,
            date_from,
            date_to,
        )
        business_profit = calculate_business_profit(
            profit_summary,
            stats,
            finance_settings,
            days=max(1, int(comparison_shift_days)),
            uzum_expenses=expense_summary,
        )
        text = _format_premium_period_report(
            title_ru,
            title_uz,
            shop_id,
            stats,
            rows,
            previous_stats,
            profit_summary,
            business_profit,
            lang=lang,
        )
        if "Достигнут защитный лимит" in source_info:
            warning = (
                "\n\n⚠️ <i>Достигнут технический лимит строк. Уменьшите период или увеличьте FINANCE_REPORT_MAX_PAGES.</i>"
                if lang != "uz"
                else "\n\n⚠️ <i>Qatorlar texnik limitiga yetildi. Davrni qisqartiring yoki FINANCE_REPORT_MAX_PAGES ni oshiring.</i>"
            )
            text += warning
        await message.answer(text, reply_markup=sales_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("today"))
async def today_sales(message: Message) -> None:
    date_from, date_to = _today_range_ms()
    await _send_premium_period_report(
        message,
        title_ru="Продажи за сегодня",
        title_uz="Bugungi savdolar",
        date_from=date_from,
        date_to=date_to,
        comparison_shift_days=1,
        wait_ru="⌛ Готовлю полный отчёт за сегодня...",
        wait_uz="⌛ Bugungi to‘liq hisobot tayyorlanmoqda...",
    )


def _format_noorza_period(title: str, shop_id: int, stats: dict[str, Any], rows: list[dict[str, Any]]) -> str:
    text = _format_noorza_today(shop_id, stats, rows)
    text = text.replace("💰 <b>Продажи за сегодня</b>", f"💰 <b>{escape(title)}</b>", 1)
    text = text.replace("💰 <b>Продажи Uzum FBO/FBS за сегодня</b>", f"💰 <b>{escape(title)}</b>", 1)
    if not rows:
        text = text.replace("за сегодня", "за выбранный период")
    return text


@dp.message(Command("yesterday"))
async def yesterday_sales(message: Message) -> None:
    date_from, date_to = _yesterday_range_ms()
    await _send_premium_period_report(
        message,
        title_ru="Продажи за вчера",
        title_uz="Kechagi savdolar",
        date_from=date_from,
        date_to=date_to,
        comparison_shift_days=1,
        wait_ru="⌛ Готовлю полный отчёт за вчера...",
        wait_uz="⌛ Kechagi to‘liq hisobot tayyorlanmoqda...",
    )


@dp.message(Command("week"))
@dp.message(Command("last7"))
async def week_sales(message: Message) -> None:
    date_from, date_to = _last_7_days_range_ms()
    await _send_premium_period_report(
        message,
        title_ru="Продажи за 7 дней",
        title_uz="7 kunlik savdolar",
        date_from=date_from,
        date_to=date_to,
        comparison_shift_days=7,
        wait_ru="⌛ Готовлю полный отчёт за 7 дней...",
        wait_uz="⌛ 7 kunlik to‘liq hisobot tayyorlanmoqda...",
    )


@dp.message(Command("balance"))
async def balance(message: Message) -> None:
    date_from, date_to = _days_range_ms(30)
    await _send_premium_period_report(
        message,
        title_ru="Управленческий отчёт за 30 дней",
        title_uz="30 kunlik boshqaruv hisoboti",
        date_from=date_from,
        date_to=date_to,
        comparison_shift_days=30,
        wait_ru="⌛ Готовлю управленческий отчёт за 30 дней...",
        wait_uz="⌛ 30 kunlik boshqaruv hisoboti tayyorlanmoqda...",
    )


LOSS_PRODUCT_FILTERS = ("ALL", "ARCHIVE", "DEFECTED")
LOSS_REPORT_MAX_PAGES = max(1, int(os.getenv("LOSS_REPORT_MAX_PAGES", "50") or "50"))


def _loss_qty(row: dict[str, Any], key: str) -> int:
    return max(0, int(_num_from_value(row.get(key)) or 0))


def _loss_row_key(row: dict[str, Any]) -> str:
    """Стабильный ключ SKU для удаления дублей между фильтрами Uzum."""
    for key in ("sku_id", "barcode", "seller_item_code"):
        value = row.get(key)
        if value not in (None, "", "—"):
            return f"{key}:{value}"
    return "fallback:" + "|".join(
        str(row.get(key) or "")
        for key in ("product_id", "product_title", "sku_full_title", "sku_title")
    )


async def _load_all_time_loss_rows_legacy(
    client: UzumClient,
    shop_id: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Накопительные недостачи и брак по всем доступным карточкам магазина.

    Uzum хранит quantityMissing/quantityDefected внутри skuList. Один и тот же
    SKU может прийти в ALL, ARCHIVE и DEFECTED, поэтому строки объединяются, а
    счётчики берутся по максимальному значению, чтобы не задвоить потери.
    """
    merged: dict[str, dict[str, Any]] = {}
    unavailable_filters: list[str] = []
    first_error: Exception | None = None
    successful_filters = 0

    for product_filter in LOSS_PRODUCT_FILTERS:
        try:
            products = await load_products(
                client,
                shop_id,
                max_pages=LOSS_REPORT_MAX_PAGES,
                page_size=100,
                product_filter=product_filter,
            )
            successful_filters += 1
        except Exception as exc:
            first_error = first_error or exc
            unavailable_filters.append(product_filter)
            logging.warning(
                "Loss report: Uzum filter %s unavailable for shop=%s: %s",
                product_filter,
                shop_id,
                exc,
            )
            continue

        for product in products:
            if not isinstance(product, dict):
                continue
            product_archived = bool(product.get("archived")) or product_filter == "ARCHIVE"
            for source_row in flatten_sku_rows([product]):
                row = dict(source_row)
                raw = row.get("raw") if isinstance(row.get("raw"), dict) else {}
                row["archived"] = product_archived or bool(raw.get("archived"))
                row["loss_all_time"] = True
                row["loss_only"] = True
                row["missing"] = _loss_qty(row, "missing")
                row["defected"] = _loss_qty(row, "defected")
                if row["missing"] <= 0 and row["defected"] <= 0:
                    continue

                key = _loss_row_key(row)
                current = merged.get(key)
                if current is None:
                    row["source_filters"] = {product_filter}
                    merged[key] = row
                    continue

                current["missing"] = max(_loss_qty(current, "missing"), row["missing"])
                current["defected"] = max(_loss_qty(current, "defected"), row["defected"])
                current["archived"] = bool(current.get("archived")) or bool(row.get("archived"))
                current.setdefault("source_filters", set()).add(product_filter)
                for field in (
                    "price",
                    "commission",
                    "total",
                    "active",
                    "fbo",
                    "fbs",
                    "status",
                    "product_id",
                    "sku_id",
                    "barcode",
                    "seller_item_code",
                ):
                    if current.get(field) in (None, "", "—") and row.get(field) not in (None, "", "—"):
                        current[field] = row.get(field)

    if successful_filters == 0 and first_error is not None:
        raise first_error

    rows = list(merged.values())
    rows.sort(
        key=lambda row: (
            -(_loss_qty(row, "missing") + _loss_qty(row, "defected")),
            str(row.get("product_title") or row.get("sku_full_title") or ""),
        )
    )
    return rows, unavailable_filters


def _short_text(value: Any, limit: int = 70) -> str:
    text = " ".join(str(value or "-").split())
    if len(text) <= limit:
        return text
    return text[: limit - 1] + "…"


def _split_long_message(text: str, limit: int = 3900) -> list[str]:
    parts: list[str] = []
    current = ""
    for block in text.split("\n\n"):
        candidate = block if not current else current + "\n\n" + block
        if len(candidate) <= limit:
            current = candidate
        else:
            if current:
                parts.append(current)
            current = block
    if current:
        parts.append(current)
    return parts


def _format_lost_sku_line(row: dict[str, Any], idx: int, lang: str) -> str:
    missing = _loss_qty(row, "missing")
    defected = _loss_qty(row, "defected")
    available = max(0, int(_num_from_value(row.get("total")) or 0))
    title = escape(_short_text(row.get("product_title") or row.get("sku_full_title") or row.get("sku_title") or "Без названия"))
    sku_value = row.get("sku_id") or row.get("barcode") or row.get("seller_item_code") or row.get("sku_full_title") or "-"
    sku = escape(_short_text(sku_value, limit=90))
    status_value = status_display(row.get("status")) if row.get("status") else "-"
    status = escape(str(status_value))
    price = _num_from_value(row.get("price")) or 0
    approx = (missing + defected) * price
    archived = bool(row.get("archived"))

    if lang == "uz":
        line = (
            f"{idx}. <b>{title}</b>\n"
            f"SKU: <code>{sku}</code>\n"
            f"Yo‘qolgan: <b>{missing} dona</b> | Yaroqsiz: <b>{defected} dona</b>\n"
            f"Qoldiq: {available} dona | Holat: {status}"
        )
        if archived:
            line += " | Arxivda"
        if price:
            line += f"\nSotuv narxi: {_format_money(price)} | Taxminiy qiymat: <b>{_format_money(approx)}</b>"
        return line

    line = (
        f"{idx}. <b>{title}</b>\n"
        f"SKU: <code>{sku}</code>\n"
        f"Потеряно: <b>{missing} шт.</b> | Брак: <b>{defected} шт.</b>\n"
        f"Остаток: {available} шт. | Статус: {status}"
    )
    if archived:
        line += " | В архиве"
    if price:
        line += f"\nЦена продажи: {_format_money(price)} | Ориентировочная стоимость: <b>{_format_money(approx)}</b>"
    return line


@dp.message(Command("lost"))
@dp.message(Command("missing"))
async def lost_goods(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    telegram_id = message.from_user.id if message.from_user else 0
    lang = get_user_language(telegram_id)

    await message.answer(
        "⌛ Barcha davr uchun yo‘qolgan va yaroqsiz tovarlar tekshirilmoqda..."
        if lang == "uz"
        else "⌛ Проверяю потерянные товары и брак за весь период...",
        reply_markup=stock_menu_for_message(message),
    )
    try:
        rows, unavailable_filters = await _load_all_time_loss_rows(
            client,
            shop_id,
            force_refresh=True,
        )

        if not rows:
            text = (
                f"🧭 <b>Barcha davrdagi yo‘qotishlar</b>\n🏪 Do‘kon: <code>{shop_id}</code>\n\n"
                "Uzum API yo‘qolgan yoki yaroqsiz SKUlarni qaytarmadi."
                if lang == "uz"
                else f"🧭 <b>Потерянные товары за весь период</b>\n🏪 Магазин: <code>{shop_id}</code>\n\n"
                "Uzum API не вернул SKU с накопленной недостачей или браком."
            )
            if unavailable_filters:
                filters_text = ", ".join(unavailable_filters)
                text += (
                    f"\n\n⚠️ Ayrim filtrlar vaqtincha ishlamadi: <code>{escape(filters_text)}</code>. "
                    "Natija to‘liq bo‘lmasligi mumkin."
                    if lang == "uz"
                    else f"\n\n⚠️ Часть фильтров временно недоступна: <code>{escape(filters_text)}</code>. "
                    "Нельзя достоверно утверждать, что потерь нет."
                )
            await message.answer(text, reply_markup=stock_menu_for_message(message))
            return

        total_missing = sum(_loss_qty(row, "missing") for row in rows)
        total_defected = sum(_loss_qty(row, "defected") for row in rows)
        approx_value = sum(
            (_loss_qty(row, "missing") + _loss_qty(row, "defected"))
            * (_num_from_value(row.get("price")) or 0)
            for row in rows
        )
        if lang == "uz":
            title = "🧭 <b>Barcha davrdagi yo‘qotishlar</b>"
            summary = [
                f"🏪 Do‘kon: <code>{shop_id}</code>",
                f"Yo‘qotishli SKU: <b>{len(rows)}</b>",
                f"Yo‘qolgan: <b>{total_missing}</b> dona | Yaroqsiz: <b>{total_defected}</b> dona",
                f"Sotuv narxi bo‘yicha taxminiy qiymat: <b>{_format_money(approx_value)}</b>",
                "ℹ️ Uzum alohida yo‘qotish sanalarini bermaydi; ko‘rsatkichlar SKU bo‘yicha jamlangan.",
            ]
            if unavailable_filters:
                summary.append("⚠️ Uzum ayrim arxiv filtrlarini vaqtincha bermadi; ro‘yxat to‘liq bo‘lmasligi mumkin.")
        else:
            title = "🧭 <b>Потерянные товары за весь период</b>"
            summary = [
                f"🏪 Магазин: <code>{shop_id}</code>",
                f"SKU с потерями: <b>{len(rows)}</b>",
                f"Потеряно: <b>{total_missing}</b> шт. | Брак: <b>{total_defected}</b> шт.",
                f"Ориентировочная стоимость по цене продажи: <b>{_format_money(approx_value)}</b>",
                "ℹ️ Uzum не передаёт даты отдельных потерь; показаны накопительные показатели по SKU.",
            ]
            if unavailable_filters:
                summary.append("⚠️ Часть архивных фильтров Uzum временно недоступна; список может быть неполным.")
        items = [_format_lost_sku_line(row, idx, lang) for idx, row in enumerate(rows, start=1)]
        await send_paginated_list(message, kind="lost", title=title, summary=summary, items=items, section="stock", reply_markup=stock_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("loss_documents", "claim_documents", "compensation_documents"))
@dp.message(F.text.in_({"📑 Документы по потерям", "📑 Yo‘qotish hujjatlari"}))
async def loss_claim_documents(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    if not await require_premium_subscription(message, telegram_id):
        return
    lang = get_user_language(telegram_id)
    await message.answer(
        "⌛ Yo‘qotish va brak bo‘yicha Excel reyestrlari hamda pretenziyalar tayyorlanmoqda..."
        if lang == "uz"
        else "⌛ Готовлю Excel-реестры и претензии по потерям и браку...",
        reply_markup=stock_menu_for_message(message),
    )
    paths: list[Path] = []
    try:
        stock_rows, unavailable_filters = await _load_all_time_loss_rows(
            client,
            shop_id,
            force_refresh=True,
        )
        loss_rows, loss_summary = prepare_compensation_rows(stock_rows, kind="loss")
        damage_rows, damage_summary = prepare_compensation_rows(stock_rows, kind="damage")
        if not loss_rows and not damage_rows:
            text = (
                "Uzum API yo‘qolgan yoki yaroqsiz tovarlarni qaytarmadi."
                if lang == "uz"
                else "Uzum API не вернул потерянные или повреждённые товары."
            )
            if unavailable_filters:
                text += (
                    " Ayrim filtrlar ishlamadi, natija to‘liq bo‘lmasligi mumkin."
                    if lang == "uz"
                    else " Часть фильтров недоступна, поэтому результат может быть неполным."
                )
            await message.answer(text, reply_markup=stock_menu_for_message(message))
            return

        stamp = datetime.now(UZT).strftime("%Y%m%d_%H%M%S")
        jobs: list[tuple[str, str, list[dict[str, Any]], Any]] = []
        if loss_rows:
            jobs.extend(
                [
                    ("loss", "xlsx", loss_rows, build_compensation_workbook),
                    ("loss", "docx", loss_rows, build_claim_docx),
                ]
            )
        if damage_rows:
            jobs.extend(
                [
                    ("damage", "xlsx", damage_rows, build_compensation_workbook),
                    ("damage", "docx", damage_rows, build_claim_docx),
                ]
            )

        for kind, extension, rows, builder in jobs:
            output = Path(tempfile.gettempdir()) / f"sellerpro_{kind}_{shop_id}_{stamp}.{extension}"
            if extension == "xlsx":
                path = await asyncio.to_thread(
                    builder,
                    rows,
                    output,
                    kind=kind,
                    lang=lang,
                )
            else:
                path = await asyncio.to_thread(
                    builder,
                    rows,
                    output,
                    kind=kind,
                    shop_id=shop_id,
                    lang=lang,
                )
            paths.append(Path(path))

        summary_parts: list[str] = []
        if loss_rows:
            summary_parts.append(
                (f"yo‘qolgan: {loss_summary['quantity']} dona" if lang == "uz" else f"потеряно: {loss_summary['quantity']} шт.")
            )
        if damage_rows:
            summary_parts.append(
                (f"brak: {damage_summary['quantity']} dona" if lang == "uz" else f"брак: {damage_summary['quantity']} шт.")
            )
        missing_values = int(loss_summary.get("missing_compensation") or 0) + int(damage_summary.get("missing_compensation") or 0)
        caption = (
            "📑 Hujjatlar tayyor · " + ", ".join(summary_parts)
            if lang == "uz"
            else "📑 Документы готовы · " + ", ".join(summary_parts)
        )
        if missing_values:
            caption += (
                f"\n⚠️ {missing_values} ta SKU bo‘yicha Uzum aniq komissiya yoki narx bermadi; summa bo‘sh qoldirildi."
                if lang == "uz"
                else f"\n⚠️ По {missing_values} SKU Uzum не передал точную комиссию или цену; сумма оставлена пустой."
            )
        await message.answer(caption, reply_markup=stock_menu_for_message(message))
        for path in paths:
            await message.answer_document(
                FSInputFile(str(path), filename=path.name),
                reply_markup=stock_menu_for_message(message),
            )
    except Exception as error:
        await send_api_error(message, error)
    finally:
        for path in paths:
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass


# --- FBO Invoice / накладные поставки ---
def _extract_list_any(data: Any) -> list[Any]:
    """Достаём список из разных форматов ответа Uzum API."""
    if isinstance(data, list):
        return data
    try:
        items = extract_items(data)
        if isinstance(items, list) and items:
            return items
    except Exception:
        pass
    if isinstance(data, dict):
        for key in (
            "payload", "content", "items", "data", "result", "results", "list", "records",
            "invoices", "invoiceList", "productList", "products", "returns", "returnList",
        ):
            value = data.get(key)
            if isinstance(value, list):
                return value
            if isinstance(value, dict):
                nested = _extract_list_any(value)
                if nested:
                    return nested
    return []


def _value_by_path(item: Any, *paths: str) -> Any:
    if not isinstance(item, dict):
        return None
    for path in paths:
        cur: Any = item
        ok = True
        for part in path.split("."):
            if isinstance(cur, dict) and part in cur:
                cur = cur.get(part)
            else:
                ok = False
                break
        if ok and cur not in (None, ""):
            return cur
    return None


def _status_text_any(value: Any) -> str:
    if isinstance(value, dict):
        return str(
            value.get("text")
            or value.get("title")
            or value.get("name")
            or value.get("value")
            or value.get("code")
            or value.get("status")
            or "—"
        )
    if value in (None, ""):
        return "—"
    return str(value)


def _date_text_any(value: Any) -> str:
    if not value:
        return "—"
    text = str(value).strip()
    parsed = _dt_from_db(text)
    if parsed:
        return _fmt_dt(parsed)
    return text[:19].replace("T", " ")


def _num_any(value: Any) -> float:
    n = _num_from_value(value)
    return float(n or 0)


def _fmt_qty(value: Any) -> str:
    n = _num_any(value)
    if abs(n - int(n)) < 0.00001:
        return str(int(n))
    return str(round(n, 2)).rstrip("0").rstrip(".")


async def _request_fbo_invoices(
    client: UzumClient,
    shop_id: int,
    *,
    page: int = 0,
    size: int = 20,
) -> Any:
    params = [("size", int(size)), ("page", int(page))]
    path = f"/v1/shop/{int(shop_id)}/invoice?" + urlencode(params)
    return await client._request("GET", path)


async def _load_fbo_invoices(
    client: UzumClient,
    shop_id: int,
    *,
    max_pages: int = 3,
    page_size: int = 20,
) -> tuple[list[dict[str, Any]], Any | None]:
    rows: list[dict[str, Any]] = []
    first_response: Any | None = None
    for page in range(max_pages):
        data = await _request_fbo_invoices(client, shop_id, page=page, size=page_size)
        if first_response is None:
            first_response = data
        items = _extract_list_any(data)
        if not items:
            break
        rows.extend([x for x in items if isinstance(x, dict)])
        if len(items) < page_size:
            break
    return rows, first_response


async def _request_fbo_invoice_products(
    client: UzumClient,
    shop_id: int,
    invoice_id: int,
) -> Any:
    params = [("invoiceId", int(invoice_id))]
    path = f"/v1/shop/{int(shop_id)}/invoice/products?" + urlencode(params)
    return await client._request("GET", path)


def _invoice_id(item: dict[str, Any]) -> Any:
    return _value_by_path(item, "id", "invoiceId", "invoice.id")


def _invoice_number(item: dict[str, Any]) -> str:
    return str(
        _value_by_path(item, "invoiceNumber", "number", "invoice.number", "deliveryCertificate")
        or _invoice_id(item)
        or "—"
    )


def _invoice_status(item: dict[str, Any]) -> str:
    return _status_text_any(
        _value_by_path(item, "invoiceStatus", "status", "state", "invoiceStatus.value")
    )


def _format_invoice_line(item: dict[str, Any], idx: int) -> str:
    invoice_id = _invoice_id(item)
    number = escape(_short_text(_invoice_number(item), 80))
    status = escape(_short_text(_invoice_status(item), 60))
    created = _date_text_any(_value_by_path(item, "dateCreated", "createdAt", "creationDate"))
    accepted_date = _date_text_any(_value_by_path(item, "dateAccepted", "acceptedAt", "acceptanceDate"))
    time_from = _date_text_any(_value_by_path(item, "timeSlotReservation.timeFrom", "timeFrom"))
    time_to = _date_text_any(_value_by_path(item, "timeSlotReservation.timeTo", "timeTo"))
    total_to_stock = _value_by_path(item, "totalToStock", "quantityToStock", "totalQuantity")
    total_accepted = _value_by_path(item, "totalAccepted", "quantityAccepted", "acceptedQuantity")
    full_price = _value_by_path(item, "fullPrice", "totalPrice", "price")

    lines = [f"{idx}. <b>Накладная №{number}</b>"]
    if invoice_id not in (None, ""):
        lines.append(f"ID: <code>{escape(str(invoice_id))}</code>")
    lines.append(f"Статус: <b>{status}</b>")
    if created != "—":
        lines.append(f"Создана: {escape(created)}")
    if time_from != "—" or time_to != "—":
        lines.append(f"Окно поставки: {escape(time_from)} — {escape(time_to)}")
    if accepted_date != "—":
        lines.append(f"Принята: {escape(accepted_date)}")
    if total_to_stock not in (None, "") or total_accepted not in (None, ""):
        lines.append(f"К поставке: <b>{_fmt_qty(total_to_stock)}</b> шт. | Принято: <b>{_fmt_qty(total_accepted)}</b> шт.")
    if _num_any(full_price):
        lines.append(f"Сумма: <b>{_format_money(_num_any(full_price))}</b>")
    if invoice_id not in (None, ""):
        lines.append(f"Состав: <code>/invoice {escape(str(invoice_id))}</code>")
    return "\n".join(lines)


async def fbo_invoices(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req

    await message.answer("⌛ Загружаю FBO-накладные поставки...", reply_markup=menu_for_message(message))
    try:
        invoices, first_response = await _load_fbo_invoices(client, shop_id, max_pages=3, page_size=20)
        title = "📄 <b>FBO-накладные поставки</b>"
        if not invoices:
            text = (
                f"{title}\n"
                f"Магазин: <code>{shop_id}</code>\n\n"
                "Накладные не найдены или Uzum API вернул пустой список."
            )
            if first_response is not None:
                text += "\n\nПервые данные API:\n<code>" + escape(compact_json_preview(first_response)) + "</code>"
            await message.answer(text, reply_markup=menu_for_message(message))
            return

        lines = [
            title,
            f"Магазин: <code>{shop_id}</code>",
            f"Найдено: <b>{len(invoices)}</b>",
            "Чтобы посмотреть состав, отправьте <code>/invoice ID</code>. Например: <code>/invoice 123456</code>",
        ]
        for idx, item in enumerate(invoices[:50], start=1):
            lines.append(_format_invoice_line(item, idx))
        if len(invoices) > 50:
            lines.append(f"Показаны первые 50 накладных из {len(invoices)}.")

        for part in _split_long_message("\n\n".join(lines)):
            await message.answer(part, reply_markup=menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


def _format_invoice_product_line(item: dict[str, Any], idx: int) -> str:
    product_title = _value_by_path(item, "productTitle", "title", "product.name")
    sku_title = _value_by_path(item, "skuTitle", "sku.title", "skuName")
    title = escape(_short_text(product_title or sku_title or "Без названия", 90))
    sku_text = escape(_short_text(sku_title or "—", 90))
    item_id = _value_by_path(item, "id", "skuId", "productId")
    to_stock = _value_by_path(item, "quantityToStock", "toStock", "quantity")
    accepted = _value_by_path(item, "quantityAccepted", "accepted", "acceptedQuantity")
    # Only the documented Uzum purchasePrice is a cost. ``price`` can be the
    # sale price and must never be used as a profitability fallback.
    purchase_price = _value_by_path(item, "purchasePrice")
    diff = _num_any(to_stock) - _num_any(accepted)

    lines = [f"{idx}. <b>{title}</b>"]
    if item_id not in (None, ""):
        lines.append(f"ID/SKU: <code>{escape(str(item_id))}</code>")
    if sku_text != "—":
        lines.append(f"SKU: {sku_text}")
    lines.append(f"По накладной: <b>{_fmt_qty(to_stock)}</b> шт. | Принято: <b>{_fmt_qty(accepted)}</b> шт.")
    if abs(diff) > 0.00001:
        sign = "−" if diff > 0 else "+"
        lines.append(f"Расхождение: <b>{sign}{_fmt_qty(abs(diff))}</b> шт.")
    if _num_any(purchase_price):
        lines.append(f"Закупочная цена: <b>{_format_money(_num_any(purchase_price))}</b>")

    sku_list = _value_by_path(item, "skuForInvoiceDtoList", "skuList", "skus")
    if isinstance(sku_list, list) and sku_list:
        sku_lines: list[str] = []
        for sku in sku_list[:3]:
            if not isinstance(sku, dict):
                continue
            sku_name = escape(_short_text(_value_by_path(sku, "skuTitle", "title", "name") or "SKU", 60))
            sku_to = _value_by_path(sku, "quantityToStock", "quantity", "toStock")
            sku_acc = _value_by_path(sku, "quantityAccepted", "accepted", "acceptedQuantity")
            sku_lines.append(f"• {sku_name}: {_fmt_qty(sku_to)} / принято {_fmt_qty(sku_acc)}")
        if sku_lines:
            if len(sku_list) > 3:
                sku_lines.append(f"• ещё {len(sku_list) - 3} SKU")
            lines.append("Внутри:\n" + "\n".join(sku_lines))
    return "\n".join(lines)


async def fbo_invoice_products(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    arg = parse_args(message.text or "")
    if not arg or not arg.split()[0].isdigit():
        await message.answer(
            "📄 <b>Состав FBO-накладной</b>\n\n"
            "Сначала откройте список накладных: <code>/invoices</code>\n"
            "Потом отправьте команду с ID накладной. Например:\n"
            "<code>/invoice 123456</code>",
            reply_markup=menu_for_message(message),
        )
        return
    invoice_id = int(arg.split()[0])

    await message.answer(f"⌛ Загружаю состав накладной <code>{invoice_id}</code>...", reply_markup=menu_for_message(message))
    try:
        data = await _request_fbo_invoice_products(client, shop_id, invoice_id)
        products = [x for x in _extract_list_any(data) if isinstance(x, dict)]
        title = "📦 <b>Состав FBO-накладной</b>"
        if not products:
            await message.answer(
                f"{title}\n"
                f"Магазин: <code>{shop_id}</code>\n"
                f"Накладная ID: <code>{invoice_id}</code>\n\n"
                "Товары не найдены или API вернул пустой состав.\n\n"
                "Ответ API:\n<code>" + escape(compact_json_preview(data)) + "</code>",
                reply_markup=menu_for_message(message),
            )
            return

        total_to_stock = sum(_num_any(_value_by_path(x, "quantityToStock", "toStock", "quantity")) for x in products)
        total_accepted = sum(_num_any(_value_by_path(x, "quantityAccepted", "accepted", "acceptedQuantity")) for x in products)
        diff = total_to_stock - total_accepted
        total_purchase = sum(
            _num_any(_value_by_path(x, "purchasePrice"))
            * _num_any(_value_by_path(x, "quantityToStock", "toStock", "quantity"))
            for x in products
        )

        lines = [
            title,
            f"Магазин: <code>{shop_id}</code>",
            f"Накладная ID: <code>{invoice_id}</code>",
            f"Позиций: <b>{len(products)}</b>",
            f"По накладной: <b>{_fmt_qty(total_to_stock)}</b> шт.",
            f"Принято: <b>{_fmt_qty(total_accepted)}</b> шт.",
        ]
        if abs(diff) > 0.00001:
            sign = "−" if diff > 0 else "+"
            lines.append(f"Расхождение: <b>{sign}{_fmt_qty(abs(diff))}</b> шт.")
        if total_purchase:
            lines.append(f"Сумма по закупочной цене: <b>{_format_money(total_purchase)}</b>")

        for idx, item in enumerate(products[:80], start=1):
            lines.append(_format_invoice_product_line(item, idx))
        if len(products) > 80:
            lines.append(f"Показаны первые 80 позиций из {len(products)}.")

        for part in _split_long_message("\n\n".join(lines)):
            await message.answer(part, reply_markup=menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("sales"))
async def sales(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    await message.answer("⏳ Считаю продажи за сегодня, 7 и 30 дней...", reply_markup=menu_for_message(message))
    try:
        today, _ = await _sales_period_stats(client, shop_id, 1)
        week, _ = await _sales_period_stats(client, shop_id, 7)
        month, _ = await _sales_period_stats(client, shop_id, 30)
        await message.answer(
            "💰 <b>Сводка продаж</b>\n"
            f"Магазин: <code>{shop_id}</code>\n\n"
            + _format_sales_summary_line("Сегодня", today)
            + "\n\n"
            + _format_sales_summary_line("7 дней", week)
            + "\n\n"
            + _format_sales_summary_line("30 дней", month)
            + "\n\nПодробно: <code>/sales_today</code>, <code>/sales_7</code>, <code>/sales_30</code>",
            reply_markup=menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)


async def _send_sales_details(message: Message, days: int) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    try:
        stats, first = await _sales_period_stats(client, shop_id, days)
        await message.answer(
            _format_sales_details(days, shop_id, stats, first), reply_markup=menu_for_message(message)
        )
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("sales_today"))
async def sales_today(message: Message) -> None:
    await _send_sales_details(message, 1)


@dp.message(Command("sales_7"))
async def sales_7(message: Message) -> None:
    await _send_sales_details(message, 7)


@dp.message(Command("sales_30"))
async def sales_30(message: Message) -> None:
    await _send_sales_details(message, 30)


def _problem_finance_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return normalized cancellation/return pieces, including partial cases."""
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in rows:
        reason: Any = _deep_pick_value(
            item,
            (
                "cancelReason",
                "cancellationReason",
                "returnReason",
                "reason",
                "reasonText",
            ),
        )
        if isinstance(reason, dict):
            reason = pick(reason, "title", "name", "text", "value", "description")
        for piece in _normalize_finance_rows([item]):
            if piece.get("kind") not in {"cancel", "return"}:
                continue
            row = {**piece, "reason": str(reason or "")}
            identity = "|".join(
                (
                    str(row.get("kind") or ""),
                    str(row.get("order_id") or ""),
                    str(row.get("sku") or ""),
                    str(row.get("qty") or 0),
                    str(row.get("status") or ""),
                )
            )
            if identity in seen:
                continue
            seen.add(identity)
            result.append(row)
    return result


@dp.message(Command("cancellations", "cancels", "returns"))
async def cancellations_report(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    await message.answer(
        "⌛ Bekor qilish va qaytarishlar tekshirilmoqda..."
        if lang == "uz"
        else "⌛ Проверяю отмены и возвраты за 30 дней...",
        reply_markup=attention_menu_for_message(message),
    )
    try:
        date_from, date_to = _days_range_ms(30)
        rows, _, _ = await _load_finance_range_flexible(client, shop_id, date_from, date_to)
        problem_rows = _problem_finance_rows(rows)
        cancelled = sum(1 for item in problem_rows if item.get("kind") == "cancel")
        returned = sum(1 for item in problem_rows if item.get("kind") == "return")
        amount = sum(float(item.get("revenue") or 0) for item in problem_rows)
        title = "🚫 <b>Bekor qilish va qaytarishlar — 30 kun</b>" if lang == "uz" else "🚫 <b>Отмены и возвраты — 30 дней</b>"
        summary = [
            f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
            f"❌ Bekor qilish: <b>{cancelled}</b> | ↩️ Qaytarish: <b>{returned}</b>" if lang == "uz" else f"❌ Отмен: <b>{cancelled}</b> | ↩️ Возвратов: <b>{returned}</b>",
            f"💵 Pozitsiyalar summasi: <b>{_format_money(amount)}</b>" if lang == "uz" else f"💵 Сумма проблемных позиций: <b>{_format_money(amount)}</b>",
        ]
        items: list[str] = []
        for index, item in enumerate(problem_rows, start=1):
            status = str(item.get("status") or "-")
            is_return = item.get("kind") == "return"
            kind = "Qaytarish" if is_return else "Bekor qilish"
            if lang != "uz":
                kind = "Возврат" if is_return else "Отмена"
            reason = item.get("reason")
            reason_line = f"\n💬 Sabab: {escape(_short_text(str(reason), 100))}" if reason and lang == "uz" else f"\n💬 Причина: {escape(_short_text(str(reason), 100))}" if reason else ""
            items.append(
                f"{index}. <b>{kind}</b> — {escape(_short_text(str(item.get('title') or 'Товар'), 70))}\n"
                f"🔖 SKU: <code>{escape(_short_text(str(item.get('sku') or '-'), 55))}</code>\n"
                f"🆔 <code>{escape(str(item.get('order_id') or '-'))}</code> | "
                f"{_format_money(float(item.get('revenue') or 0))}\n"
                f"📌 {escape(status)}{reason_line}"
            )
        if not items:
            items = [
                "✅ Oxirgi 30 kunda bekor qilish va qaytarish topilmadi."
                if lang == "uz"
                else "✅ За последние 30 дней отмен и возвратов не найдено."
            ]
        await send_paginated_list(
            message,
            kind="cancellations",
            title=title,
            summary=summary,
            items=items,
            section="attention",
            reply_markup=attention_menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)



# --- Общая сводка магазина / заказы по статусам ---
# Дополняет Finance API. Если Finance API вернул 0 продаж, сводка всё равно покажет
# текущие заказы FBS/DBS и состояние остатков.
FBS_STATUS_LABELS: dict[str, str] = {
    "CREATED": "Создан",
    "PACKING": "Сборка",
    "PENDING_DELIVERY": "Ожидает доставки",
    "DELIVERING": "В доставке",
    "DELIVERED": "Доставлен",
    "ACCEPTED_AT_DP": "Принят в ПВЗ/ДП",
    "DELIVERED_TO_CUSTOMER_DELIVERY_POINT": "В пункте выдачи",
    "COMPLETED": "Завершён",
    "CANCELED": "Отменён",
    "RETURNED": "Возврат",
}

# Минимальный набор статусов для сводки, чтобы не ловить 429 Too Many Requests.
# Полную детализацию статусов добавим позже, когда подберём лимиты Uzum API.
FBS_SUMMARY_STATUSES: tuple[str, ...] = (
    "CREATED",
    "PACKING",
    "COMPLETED",
    "CANCELED",
)
ORDER_SUMMARY_REQUEST_DELAY_SECONDS = float(os.getenv("ORDER_SUMMARY_REQUEST_DELAY_SECONDS", "0.45") or "0.45")


def _extract_count(data: Any) -> int:
    if isinstance(data, bool):
        return 0
    if isinstance(data, (int, float)):
        return int(data)
    if isinstance(data, str):
        try:
            return int(float(data))
        except Exception:
            return 0
    if isinstance(data, dict):
        for key in ("payload", "data", "result", "value", "count", "total", "totalElements", "totalAmount"):
            if key in data:
                value = data.get(key)
                if isinstance(value, dict):
                    nested = _extract_count(value)
                    if nested:
                        return nested
                else:
                    number = _num_from_value(value)
                    if number is not None:
                        return int(number)
        # Последняя попытка — ищем первое числовое поле.
        for value in data.values():
            if isinstance(value, (dict, list)):
                nested = _extract_count(value)
                if nested:
                    return nested
            else:
                number = _num_from_value(value)
                if number is not None:
                    return int(number)
    if isinstance(data, list):
        return len(data)
    return 0


async def _fbs_order_count(
    client: UzumClient,
    shop_id: int,
    status: str,
    *,
    date_from_ms: int,
    date_to_ms: int,
) -> int:
    params = [
        ("shopIds", shop_id),
        ("status", status),
        ("dateFrom", date_from_ms),
        ("dateTo", date_to_ms),
    ]
    path = "/v2/fbs/orders/count?" + urlencode(params)
    data = await client._request("GET", path)
    return _extract_count(data)


async def _orders_counts_for_days(client: UzumClient, shop_id: int, days: int) -> dict[str, int]:
    date_from, date_to = _today_range_ms() if days == 1 else _days_range_ms(days)
    counts: dict[str, int] = {}
    for status in FBS_SUMMARY_STATUSES:
        try:
            counts[status] = await _fbs_order_count(
                client, shop_id, status, date_from_ms=date_from, date_to_ms=date_to
            )
        except Exception as e:
            # Если Uzum вернул 429/403 по одному статусу, не валим всю сводку.
            logging.warning("FBS count failed status=%s days=%s: %s", status, days, e)
            counts[status] = 0

        # Маленькая пауза, чтобы не упираться в лимиты Uzum API.
        await asyncio.sleep(max(0.0, ORDER_SUMMARY_REQUEST_DELAY_SECONDS))
    return counts


def _format_orders_counts(title: str, counts: dict[str, int]) -> str:
    useful = {k: v for k, v in counts.items() if v}
    if not useful:
        return f"<b>{escape(title)}</b>\n• Заказов по основным статусам не найдено"
    lines = [f"<b>{escape(title)}</b>"]
    for status, count in useful.items():
        label = FBS_STATUS_LABELS.get(status, status)
        lines.append(f"• {escape(label)}: <b>{count}</b>")
    lines.append(f"• Итого: <b>{sum(useful.values())}</b>")
    return "\n".join(lines)


def _build_stock_stats(rows: list[dict[str, Any]]) -> dict[str, float | int]:
    total_skus = len(rows)
    product_ids = {
        str(r.get("product_id"))
        for r in rows
        if r.get("product_id") not in (None, "", "—")
    }
    total_units = 0.0
    fbo_units = 0.0
    fbs_units = 0.0
    low_count = 0
    zero_count = 0
    with_stock_count = 0
    stock_value = 0.0
    active_count = 0

    for r in rows:
        total = _num_from_value(r.get("total")) or 0.0
        fbo = _num_from_value(r.get("fbo")) or 0.0
        fbs = _num_from_value(r.get("fbs")) or 0.0
        price = _num_from_value(r.get("price")) or 0.0
        status = str(status_display(r.get("status")) if r.get("status") else r.get("status") or "").upper()

        total_units += total
        fbo_units += fbo
        fbs_units += fbs
        stock_value += max(0.0, total) * max(0.0, price)
        if total <= 0:
            zero_count += 1
        else:
            with_stock_count += 1
            if total <= LOW_STOCK_THRESHOLD:
                low_count += 1
        if "RUN_OUT" not in status and total > 0:
            active_count += 1

    return {
        "total_products": len(product_ids) if product_ids else total_skus,
        "total_skus": total_skus,
        "total_units": total_units,
        "fbo_units": fbo_units,
        "fbs_units": fbs_units,
        "with_stock_count": with_stock_count,
        "low_count": low_count,
        "zero_count": zero_count,
        "stock_value": stock_value,
        "active_count": active_count,
    }


@dp.message(Command("orders_summary"))
async def orders_summary(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    await message.answer("⏳ Считаю заказы по статусам...", reply_markup=menu_for_message(message))
    try:
        today = await _orders_counts_for_days(client, shop_id, 1)
        week = await _orders_counts_for_days(client, shop_id, 7)
        month = await _orders_counts_for_days(client, shop_id, 30)
        await message.answer(
            f"📊 <b>Сводка заказов</b>\nМагазин: <code>{shop_id}</code>\n\n"
            + _format_orders_counts("Сегодня", today)
            + "\n\n"
            + _format_orders_counts("7 дней", week)
            + "\n\n"
            + _format_orders_counts("30 дней", month),
            reply_markup=menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("dashboard", "summary"))
async def dashboard(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    await message.answer("⏳ Собираю общую сводку магазина...", reply_markup=menu_for_message(message))
    try:
        sales_7, _ = await _sales_period_stats(client, shop_id, 7)
        sales_30, _ = await _sales_period_stats(client, shop_id, 30)
        counts_7 = await _orders_counts_for_days(client, shop_id, 7)
        rows = await load_sku_rows(client, shop_id, max_pages=50)
        st = _build_stock_stats(rows)

        active_orders = sum(counts_7.get(s, 0) for s in ("CREATED", "PACKING"))
        completed_7 = counts_7.get("COMPLETED", 0)
        canceled_7 = counts_7.get("CANCELED", 0)
        returned_7 = counts_7.get("RETURNED", 0)

        await message.answer(
            f"📈 <b>Сводка магазина</b>\n"
            f"Магазин: <code>{shop_id}</code>\n\n"
            f"💰 <b>Продажи Finance API</b>\n"
            f"• 7 дней: <b>{_format_money(float(sales_7['revenue']))}</b> / строк: <b>{sales_7['active_rows']}</b>\n"
            f"• 30 дней: <b>{_format_money(float(sales_30['revenue']))}</b> / строк: <b>{sales_30['active_rows']}</b>\n\n"
            f"🛒 <b>Заказы FBS/DBS за 7 дней</b>\n"
            f"• Активные статусы: <b>{active_orders}</b>\n"
            f"• Завершено: <b>{completed_7}</b>\n"
            f"• Отменено: <b>{canceled_7}</b>\n"
            f"• Возвраты: <b>{returned_7}</b>\n\n"
            f"📦 <b>Остатки</b>\n"
            f"• Карточек / SKU-вариантов: <b>{int(st['total_products'])} / {int(st['total_skus'])}</b>\n"
            f"• Доступно к продаже: <b>{float(st['total_units']):.0f} шт.</b>\n"
            f"• Склад Uzum: <b>{float(st['fbo_units']):.0f}</b> шт. / ваш склад FBS/DBS: <b>{float(st['fbs_units']):.0f}</b> шт.\n"
            f"• SKU с остатком: <b>{int(st['with_stock_count'])}</b>\n"
            f"• Заканчиваются ≤ {LOW_STOCK_THRESHOLD}: <b>{int(st['low_count'])}</b>\n"
            f"• Нет в наличии: <b>{int(st['zero_count'])}</b>\n"
            f"• Примерная стоимость доступного остатка: <b>{_format_money(float(st['stock_value']))}</b>\n\n"
            "ℹ️ <i>Это сумма штук по всем SKU, а не количество карточек. "
            "Ожидающие товары, брак и потери сюда не входят.</i>\n\n"
            "Подробнее: <b>💰 Продажи</b>, <b>📦 Склад</b> или <b>🚨 Важно сейчас</b>.",
            reply_markup=menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)



# --- Отзывы покупателей ---
# В официальном Seller OpenAPI Uzum раздел отзывов может быть недоступен.
# Поэтому список/ответы сделаны безопасно: сначала используются переменные окружения
# REVIEWS_LIST_PATH и REVIEWS_REPLY_PATH, а если их нет — бот пробует несколько
# типовых путей и честно сообщает ошибку, если endpoint не найден.
REVIEWS_LIST_PATH = os.getenv("REVIEWS_LIST_PATH", "").strip()
REVIEWS_REPLY_PATH = os.getenv("REVIEWS_REPLY_PATH", "").strip()
REVIEWS_REPLY_METHOD = os.getenv("REVIEWS_REPLY_METHOD", "POST").strip().upper() or "POST"
REVIEWS_REPLY_BODY_FIELD = os.getenv("REVIEWS_REPLY_BODY_FIELD", "text").strip() or "text"
REVIEW_NOTIFICATIONS = (
    os.getenv("REVIEW_NOTIFICATIONS", "0").strip().lower()
    in {"1", "true", "yes", "on", "да"}
)
REVIEW_CHECK_INTERVAL_SECONDS = int(os.getenv("REVIEW_CHECK_INTERVAL_SECONDS", "600") or "600")


def _format_endpoint_template(template: str, *, shop_id: int, review_id: str = "", page: int = 0, size: int = 10) -> str:
    return (
        template.replace("{shop_id}", str(shop_id))
        .replace("{shopId}", str(shop_id))
        .replace("{review_id}", str(review_id))
        .replace("{reviewId}", str(review_id))
        .replace("{page}", str(page))
        .replace("{size}", str(size))
    )


def _review_candidates(shop_id: int, page: int = 0, size: int = 10) -> list[str]:
    if REVIEWS_LIST_PATH:
        return [_format_endpoint_template(REVIEWS_LIST_PATH, shop_id=shop_id, page=page, size=size)]
    return [
        f"/v1/reviews?shopId={shop_id}&page={page}&size={size}",
        f"/v1/reviews/shop/{shop_id}?page={page}&size={size}",
        f"/v1/shop/{shop_id}/reviews?page={page}&size={size}",
        f"/v1/feedbacks?shopId={shop_id}&page={page}&size={size}",
        f"/v1/shop/{shop_id}/feedbacks?page={page}&size={size}",
        f"/v1/comments?shopId={shop_id}&page={page}&size={size}",
        f"/v1/shop/{shop_id}/comments?page={page}&size={size}",
    ]


def _reply_candidates(shop_id: int, review_id: str) -> list[str]:
    if REVIEWS_REPLY_PATH:
        return [_format_endpoint_template(REVIEWS_REPLY_PATH, shop_id=shop_id, review_id=review_id)]
    return [
        f"/v1/reviews/{review_id}/reply",
        f"/v1/reviews/{review_id}/answer",
        f"/v1/review/{review_id}/reply",
        f"/v1/feedbacks/{review_id}/reply",
        f"/v1/feedback/{review_id}/answer",
        f"/v1/shop/{shop_id}/reviews/{review_id}/reply",
        f"/v1/shop/{shop_id}/feedbacks/{review_id}/reply",
        "/v1/reviews/reply",
        "/v1/feedbacks/reply",
    ]


def _find_review_lists(obj: Any) -> list[Any]:
    """Best-effort поиск списков отзывов в неизвестной структуре ответа."""
    direct = extract_items(obj)
    if direct:
        return direct
    found: list[Any] = []
    if isinstance(obj, dict):
        for key, value in obj.items():
            key_l = str(key).lower()
            if isinstance(value, list) and any(x in key_l for x in ("review", "feedback", "comment", "rating")):
                found.extend(value)
            elif isinstance(value, (dict, list)):
                found.extend(_find_review_lists(value))
    elif isinstance(obj, list):
        for value in obj:
            if isinstance(value, dict):
                found.append(value)
            elif isinstance(value, (dict, list)):
                found.extend(_find_review_lists(value))
    return found


def _review_id(review: Any) -> str:
    value = pick(review, "id", "reviewId", "feedbackId", "commentId", "uuid", "uid", default="—")
    return str(value)


def _review_answer(review: Any) -> str:
    return str(pick(review, "answer", "reply", "sellerAnswer", "sellerReply", "response", "commentAnswer", default="—"))


def format_review_line(review: Any) -> str:
    review_id = _review_id(review)
    product = pick(review, "productTitle", "productName", "title", "name", "skuTitle", default="—")
    rating = pick(review, "rating", "stars", "mark", "grade", "score", default="—")
    author = pick(review, "customerName", "userName", "buyerName", "clientName", "author", default="—")
    created = pick(review, "createdAt", "date", "createdDate", "publishedAt", default="—")
    text = pick(review, "text", "comment", "review", "content", "message", "description", default="—")
    answer = _review_answer(review)

    text_s = safe(text)
    if len(text_s) > 600:
        text_s = text_s[:600] + "..."

    answer_part = ""
    if answer not in (None, "", "—"):
        answer_s = safe(answer)
        if len(answer_s) > 300:
            answer_s = answer_s[:300] + "..."
        answer_part = f"\n💬 Ответ продавца: {answer_s}"

    return (
        f"• ID отзыва: <code>{safe(review_id)}</code>\n"
        f"Товар: {safe(product)}\n"
        f"Оценка: {safe(rating)} | Клиент: {safe(author)} | Дата: {safe(created)}\n"
        f"Отзыв: {text_s}"
        f"{answer_part}"
    )


async def get_reviews_from_uzum(client: UzumClient, shop_id: int, *, page: int = 0, size: int = 10) -> tuple[list[Any], str | None, str | None]:
    last_error: str | None = None
    last_path: str | None = None
    for path in _review_candidates(shop_id, page=page, size=size):
        try:
            data = await client._request("GET", path)
            items = _find_review_lists(data)
            return items, path, None
        except Exception as e:
            last_error = str(e)
            last_path = path
            continue
    return [], last_path, last_error


@dp.message(Command("reviews"))
async def reviews(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req

    items, path, error = await get_reviews_from_uzum(client, shop_id, page=0, size=10)
    if error and not items:
        await message.answer(
            "⭐ <b>Отзывы</b>\n\n"
            "Не смог получить отзывы через текущий Uzum API.\n"
            "Скорее всего, в вашем Seller OpenAPI нет открытого метода для отзывов.\n\n"
            "Что можно сделать:\n"
            "1. Если у вас есть endpoint отзывов из кабинета Uzum, добавьте в bothost переменные:\n"
            "<code>REVIEWS_LIST_PATH</code> и <code>REVIEWS_REPLY_PATH</code>.\n"
            "2. Потом перезапустите бота.\n\n"
            f"Последний путь: <code>{escape(str(path))}</code>\n"
            f"Ошибка: <code>{escape(error[:1000])}</code>",
            reply_markup=menu_for_message(message),
        )
        return

    if not items:
        await message.answer("⭐ Отзывы не найдены.", reply_markup=menu_for_message(message))
        return

    lines = [format_review_line(item) for item in items[:10]]
    await message.answer(
        "⭐ <b>Последние отзывы</b>\n\n"
        + "\n\n".join(lines)
        + "\n\nЧтобы ответить: <code>/reply ID_ОТЗЫВА ваш ответ</code>",
        reply_markup=menu_for_message(message),
    )


@dp.message(Command("reply"))
async def reply_review(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req

    arg = parse_args(message.text or "")
    if not arg or " " not in arg:
        await message.answer(
            "Напишите так:\n"
            "<code>/reply ID_ОТЗЫВА Спасибо за отзыв!</code>\n\n"
            "ID отзыва можно посмотреть через <code>/reviews</code>.",
            reply_markup=menu_for_message(message),
        )
        return

    review_id, answer_text = arg.split(maxsplit=1)
    review_id = review_id.strip()
    answer_text = answer_text.strip()
    if not review_id or not answer_text:
        await message.answer("Не вижу ID отзыва или текст ответа.", reply_markup=menu_for_message(message))
        return
    if len(answer_text) > 1000:
        await message.answer("Ответ слишком длинный. Сделайте до 1000 символов.", reply_markup=menu_for_message(message))
        return

    payloads = [
        {REVIEWS_REPLY_BODY_FIELD: answer_text},
        {"text": answer_text},
        {"reply": answer_text},
        {"answer": answer_text},
        {"message": answer_text},
        {"reviewId": review_id, "shopId": shop_id, "text": answer_text},
        {"feedbackId": review_id, "shopId": shop_id, "answer": answer_text},
    ]

    errors: list[str] = []
    tried = 0
    for path in _reply_candidates(shop_id, review_id):
        # Для кастомного REVIEWS_REPLY_PATH пробуем только первый payload, заданный полем REVIEWS_REPLY_BODY_FIELD.
        selected_payloads = payloads[:1] if REVIEWS_REPLY_PATH else payloads
        for payload in selected_payloads:
            tried += 1
            try:
                await client._request(REVIEWS_REPLY_METHOD, path, json=payload)
                await message.answer(
                    "✅ Ответ на отзыв отправлен.\n\n"
                    f"ID отзыва: <code>{escape(review_id)}</code>",
                    reply_markup=menu_for_message(message),
                )
                return
            except Exception as e:
                errors.append(f"{path}: {str(e)[:250]}")
                continue

    await message.answer(
        "⚠️ Не получилось отправить ответ на отзыв.\n\n"
        "Вероятно, ваш текущий Uzum Seller OpenAPI не поддерживает ответы на отзывы, "
        "или нужен точный endpoint из кабинета продавца.\n\n"
        "Можно добавить в bothost:\n"
        "<code>REVIEWS_REPLY_PATH=/точный/путь/{review_id}/reply</code>\n"
        "<code>REVIEWS_REPLY_BODY_FIELD=text</code>\n\n"
        f"Попыток: <b>{tried}</b>\n"
        f"Последняя ошибка:\n<code>{escape(errors[-1] if errors else '—')}</code>",
        reply_markup=menu_for_message(message),
    )


@dp.message(Command("reviews_check"))
async def reviews_check(message: Message) -> None:
    """Проверить, доступен ли endpoint отзывов для текущего API-ключа."""
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    items, path, error = await get_reviews_from_uzum(client, shop_id, page=0, size=5)
    if not error:
        if lang == "uz":
            await message.answer(
                "✅ <b>Sharhlar API tekshiruvi</b>\n\n"
                f"Do‘kon: <code>{shop_id}</code>\n"
                f"Ishlagan yo‘l: <code>{escape(str(path))}</code>\n"
                f"Topilgan sharhlar: <b>{len(items)}</b>\n\n"
                "Endi <code>/reviews</code> orqali sharhlarni ko‘rish mumkin.",
                reply_markup=menu_for_message(message),
            )
        else:
            await message.answer(
                "✅ <b>Проверка отзывов</b>\n\n"
                f"Магазин: <code>{shop_id}</code>\n"
                f"Рабочий путь: <code>{escape(str(path))}</code>\n"
                f"Найдено отзывов: <b>{len(items)}</b>\n\n"
                "Теперь можно смотреть отзывы через <code>/reviews</code>.",
                reply_markup=menu_for_message(message),
            )
        return

    if lang == "uz":
        await message.answer(
            "⚠️ <b>Sharhlar API hozircha ishlamadi</b>\n\n"
            "Uzum Seller OpenAPI sizning kalitingiz bilan sharhlarni bermayapti. "
            "Bu odatda endpoint yopiq yoki huquq yetarli emasligini bildiradi.\n\n"
            f"Oxirgi yo‘l: <code>{escape(str(path))}</code>\n"
            f"Xatolik: <code>{escape(str(error)[:1000])}</code>\n\n"
            "Agar Uzum aniq sharhlar endpointini bersa, bothost’da "
            "<code>REVIEWS_LIST_PATH</code> o‘zgaruvchisini qo‘shamiz.",
            reply_markup=menu_for_message(message),
        )
    else:
        await message.answer(
            "⚠️ <b>Отзывы пока не удалось получить</b>\n\n"
            "Uzum Seller OpenAPI не отдаёт отзывы по текущему API-ключу. "
            "Обычно это значит, что endpoint закрыт или не хватает прав.\n\n"
            f"Последний путь: <code>{escape(str(path))}</code>\n"
            f"Ошибка: <code>{escape(str(error)[:1000])}</code>\n\n"
            "Если Uzum даст точный endpoint отзывов, добавим в bothost "
            "переменную <code>REVIEWS_LIST_PATH</code>.",
            reply_markup=menu_for_message(message),
        )


@dp.message(Command("reviews_notify_status"))
async def reviews_notify_status(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    enabled = product_setting_enabled(telegram_id, "notify_reviews")
    if lang == "uz":
        text = (
            "⭐ <b>Sharhlar xabarnomalari</b>\n\n"
            f"Holat: {'✅ yoqilgan' if enabled else '❌ o‘chirilgan'}\n"
            f"Tekshiruv har <b>{REVIEW_CHECK_INTERVAL_SECONDS}</b> soniya\n\n"
            "Ishlashi uchun <code>/reviews_check</code> tekshiruvi muvaffaqiyatli bo‘lishi kerak."
        )
    else:
        text = (
            "⭐ <b>Уведомления о новых отзывах</b>\n\n"
            f"Статус: {'✅ включены' if enabled else '❌ выключены'}\n"
            f"Проверка каждые <b>{REVIEW_CHECK_INTERVAL_SECONDS}</b> сек.\n\n"
            "Будет работать только если <code>/reviews_check</code> найдёт рабочий метод отзывов."
        )
    await message.answer(text, reply_markup=menu_for_message(message))


def _review_seen_key(review: Any) -> str:
    rid = _review_id(review)
    if rid and rid != "—":
        return "id:" + str(rid)
    raw = json.dumps(review, ensure_ascii=False, sort_keys=True, default=str)
    return "hash:" + hashlib.sha1(raw.encode("utf-8")).hexdigest()


def build_review_notification(review: Any, shop_id: int, lang: str = "ru") -> str:
    lang = normalize_lang(lang)
    review_id = escape(_review_id(review))
    product = safe(pick(review, "productTitle", "productName", "title", "name", "skuTitle", default="—"))
    rating = safe(pick(review, "rating", "stars", "mark", "grade", "score", default="—"))
    author = safe(pick(review, "customerName", "userName", "buyerName", "clientName", "author", default="—"))
    created = safe(pick(review, "createdAt", "date", "createdDate", "publishedAt", default="—"))
    text = safe(pick(review, "text", "comment", "review", "content", "message", "description", default="—"))
    if len(text) > 700:
        text = text[:700] + "..."
    if lang == "uz":
        return (
            "⭐ <b>Yangi sharh</b>\n\n"
            f"🏪 Do‘kon: <code>{shop_id}</code>\n"
            f"📦 Tovar: <b>{product}</b>\n"
            f"⭐ Baho: <b>{rating}</b>\n"
            f"👤 Mijoz: {author}\n"
            f"🕒 Sana: {created}\n\n"
            f"💬 Sharh:\n{text}\n\n"
            f"ID: <code>{review_id}</code>\n"
            "Javob berish: <code>/reply ID matn</code>"
        )
    return (
        "⭐ <b>Новый отзыв</b>\n\n"
        f"🏪 Магазин: <code>{shop_id}</code>\n"
        f"📦 Товар: <b>{product}</b>\n"
        f"⭐ Оценка: <b>{rating}</b>\n"
        f"👤 Клиент: {author}\n"
        f"🕒 Дата: {created}\n\n"
        f"💬 Отзыв:\n{text}\n\n"
        f"ID: <code>{review_id}</code>\n"
        "Ответить: <code>/reply ID текст</code>"
    )


_seen_review_keys_by_scope: dict[tuple[int, int], set[str]] = {}
_reviews_watch_initialized_scopes: set[tuple[int, int]] = set()


async def check_new_reviews_once() -> None:
    for group in connected_watch_groups("notify_reviews"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(x) for x in group["telegram_ids"]]
        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            items, path, error = await get_reviews_from_uzum(client, shop_id, page=0, size=20)
            if error and not items:
                logging.warning("Reviews watcher: no access shop=%s path=%s error=%s", shop_id, path, str(error)[:300])
                await asyncio.sleep(2)
                continue
        except Exception as error:
            _log_watcher_api_failure(
                "Reviews watcher",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        keys_now = [_review_seen_key(item) for item in items]
        for telegram_id in telegram_ids:
            scope = (telegram_id, shop_id)
            known = _seen_review_keys_by_scope.setdefault(scope, set())
            if scope not in _reviews_watch_initialized_scopes:
                known.update(keys_now)
                _reviews_watch_initialized_scopes.add(scope)
                logging.info("Reviews watcher initialized for user=%s shop=%s reviews=%s", telegram_id, shop_id, len(keys_now))
                continue

            new_items = [item for item, key in zip(items, keys_now) if key not in known]
            known.update(keys_now)
            if len(known) > 2000:
                _seen_review_keys_by_scope[scope] = set(keys_now)
            for item in new_items[:10]:
                try:
                    await bot.send_message(
                        telegram_id,
                        build_review_notification(item, shop_id=shop_id, lang=get_user_language(telegram_id)),
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                    await asyncio.sleep(0.15)
                except Exception:
                    logging.exception("Reviews watcher: failed to send notification to %s", telegram_id)
        await asyncio.sleep(0.5)


async def reviews_watch_loop() -> None:
    logging.info("Reviews watcher started. Interval: %s seconds. Enabled: %s", REVIEW_CHECK_INTERVAL_SECONDS, REVIEW_NOTIFICATIONS)
    if not REVIEW_NOTIFICATIONS:
        logging.info("Reviews watcher disabled globally; no Uzum review API requests will be made")
        return
    while True:
        try:
            await check_new_reviews_once()
        except Exception:
            logging.exception("Reviews watcher crashed")
        await asyncio.sleep(max(60, REVIEW_CHECK_INTERVAL_SECONDS))


def payment_plan_markup(lang: str) -> InlineKeyboardMarkup:
    is_uz = normalize_lang(lang) == "uz"
    rows = []
    for months in (1, 3, 6):
        plan = PAYMENT_PLANS[months]
        label = plan["uz"] if is_uz else plan["ru"]
        amount = _payment_amount_text(int(plan["amount"]))
        rows.append([
            InlineKeyboardButton(
                text=f"{label} — {amount} {'so‘m' if is_uz else 'сум'}",
                callback_data=f"payplan:{months}",
            )
        ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def payment_cancel_markup(request_id: int, lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[[
            InlineKeyboardButton(
                text="❌ Bekor qilish" if normalize_lang(lang) == "uz" else "❌ Отменить оплату",
                callback_data=f"paycancel:{int(request_id)}",
            )
        ]]
    )


def payment_admin_action_markup(request_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Подтвердить", callback_data=f"payapprove:{int(request_id)}"),
                InlineKeyboardButton(text="❌ Отклонить", callback_data=f"payreject_menu:{int(request_id)}"),
            ]
        ]
    )


def payment_rejection_markup(request_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📷 Чек не читается", callback_data=f"payreject:{int(request_id)}:unreadable")],
            [InlineKeyboardButton(text="💰 Сумма не совпадает", callback_data=f"payreject:{int(request_id)}:amount")],
            [InlineKeyboardButton(text="🔎 Платёж не найден", callback_data=f"payreject:{int(request_id)}:not_found")],
            [InlineKeyboardButton(text="❌ Другая причина", callback_data=f"payreject:{int(request_id)}:other")],
        ]
    )


async def send_payment_request_to_admins(row: dict[str, Any]) -> int:
    sent = 0
    caption = payment_request_caption(row)
    request_id = int(row.get("id") or 0)
    only_admin = int(row.get("_only_admin") or 0)
    admin_ids = [only_admin] if only_admin else sorted(ADMIN_IDS)
    for admin_id in admin_ids:
        try:
            if row.get("receipt_type") == "document":
                await bot.send_document(
                    admin_id,
                    str(row.get("receipt_file_id") or ""),
                    caption=caption,
                    reply_markup=payment_admin_action_markup(request_id),
                )
            else:
                await bot.send_photo(
                    admin_id,
                    str(row.get("receipt_file_id") or ""),
                    caption=caption,
                    reply_markup=payment_admin_action_markup(request_id),
                )
            sent += 1
            await asyncio.sleep(0.15)
        except Exception:
            logging.exception("Payment receipt delivery failed: request=%s admin=%s", request_id, admin_id)
    return sent


@dp.message(Command("subscribe"))
async def subscribe(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    ensure_subscription(telegram_id)
    lang = get_user_language(telegram_id)
    pending = latest_payment_request_by_status(telegram_id, "pending_review")
    if pending:
        text = (
            "🧾 <b>Chekingiz tekshirilmoqda</b>\n\n"
            f"Ariza: <b>#{int(pending['id'])}</b>\n"
            f"Tarif: <b>{int(pending['plan_months'])} oy</b>\n"
            f"Summa: <b>{_payment_amount_text(int(pending['amount']))} so‘m</b>\n\n"
            "Administrator tekshirgach, sizga xabar keladi."
            if lang == "uz"
            else
            "🧾 <b>Ваш чек находится на проверке</b>\n\n"
            f"Заявка: <b>#{int(pending['id'])}</b>\n"
            f"Тариф: <b>{int(pending['plan_months'])} мес.</b>\n"
            f"Сумма: <b>{_payment_amount_text(int(pending['amount']))} сум</b>\n\n"
            "После проверки администратором вы получите уведомление."
        )
        await message.answer(text, reply_markup=menu_for_message(message))
        return
    if lang == "uz":
        text = (
            "💎 <b>Uzum Seller Assistant obunasi</b>\n\n"
            f"🎁 <b>{TRIAL_DAYS} kunlik sinov:</b> bugungi savdo, yangi savdo xabarlari va ertalabki hisobot.\n\n"
            "<b>To‘liq obunaga kiradi:</b>\n"
            "✅ bugun, kecha, 7 va 30 kunlik FBO/FBS savdolar\n"
            "✅ qoldiqlar va tugab borayotgan tovarlar\n"
            "✅ Uzum API bergan bo‘lsa, yo‘qolgan tovarlar\n"
            "✅ yangi savdolar haqida xabarlar\n"
            "✅ bir nechta do‘kon bilan ishlash\n"
            "✅ Excel, foyda, ertalabki va avtomatik hisobotlar\n\n"
            "💰 <b>Tarifni tanlang</b>\n\n"
            "Keyingi bosqichda to‘lov rekvizitlari chiqadi. To‘lovdan so‘ng chekni shu yerga yuborasiz."
        )
    else:
        text = (
            "💎 <b>Подписка Uzum Seller Assistant</b>\n\n"
            f"🎁 <b>Пробный период {TRIAL_DAYS} дня:</b> продажи за сегодня, уведомления о новых продажах и утренний отчёт.\n\n"
            "<b>В полную подписку входят:</b>\n"
            "✅ продажи FBO/FBS за сегодня, вчера, 7 и 30 дней\n"
            "✅ остатки и товары, которые заканчиваются\n"
            "✅ потерянные товары, если Uzum отдаёт их в API\n"
            "✅ уведомления о новых продажах\n"
            "✅ работа с несколькими магазинами\n"
            "✅ Excel, прибыль, утренние и автоматические отчёты\n\n"
            "💰 <b>Выберите тариф</b>\n\n"
            "На следующем шаге появятся реквизиты. После оплаты вы отправите чек прямо сюда."
        )
    await message.answer(text, reply_markup=payment_plan_markup(lang))


@dp.callback_query(F.data.startswith("payplan:"))
async def payment_plan_selected(callback: CallbackQuery, state: FSMContext) -> None:
    user = callback.from_user
    if not user:
        return
    telegram_id = int(user.id)
    db.upsert_user(telegram_id, user.username, user.first_name)
    lang = get_user_language(telegram_id)
    raw_months = str(callback.data or "").split(":", 1)[-1]
    plan = payment_plan(int(raw_months)) if raw_months.isdigit() else None
    if not plan:
        await callback.answer("Некорректный тариф", show_alert=True)
        return
    if not PAYMENT_REQUISITES:
        await callback.answer(
            "Реквизиты оплаты пока не настроены. Напишите администратору.",
            show_alert=True,
        )
        if callback.message:
            await callback.message.answer(
                ("Administratorga yozing: " if lang == "uz" else "Напишите администратору: ")
                + f"<b>{admin_contact_text()}</b>",
                reply_markup=admin_contact_markup() or main_menu_for_user(telegram_id),
            )
        return

    row, created = create_payment_request(telegram_id, int(plan["months"]))
    if not created and row.get("status") == "pending_review":
        await callback.answer(
            "Chekingiz allaqachon tekshirilmoqda" if lang == "uz" else "Ваш чек уже проверяется",
            show_alert=True,
        )
        return

    request_id = int(row["id"])
    await state.set_state(PaymentStates.waiting_for_receipt)
    await state.update_data(payment_request_id=request_id)
    await callback.answer()
    if not callback.message:
        return
    amount_text = _payment_amount_text(int(plan["amount"]))
    if lang == "uz":
        text = (
            "💳 <b>To‘lov</b>\n\n"
            f"Tarif: <b>{plan['uz']}</b>\n"
            f"To‘lov summasi: <b>{amount_text} so‘m</b>\n\n"
            f"<b>Rekvizitlar:</b>\n{escape(PAYMENT_REQUISITES)}\n\n"
            "To‘lovdan so‘ng chek skrinshotini yoki PDF faylini shu chatga yuboring."
        )
    else:
        text = (
            "💳 <b>Оплата подписки</b>\n\n"
            f"Тариф: <b>{plan['ru']}</b>\n"
            f"К оплате: <b>{amount_text} сум</b>\n\n"
            f"<b>Реквизиты:</b>\n{escape(PAYMENT_REQUISITES)}\n\n"
            "После оплаты отправьте в этот чат скриншот чека или PDF-файл."
        )
    await callback.message.answer(text, reply_markup=payment_cancel_markup(request_id, lang))


@dp.callback_query(F.data.startswith("paycancel:"))
async def payment_cancel_callback(callback: CallbackQuery, state: FSMContext) -> None:
    user = callback.from_user
    raw_id = str(callback.data or "").split(":", 1)[-1]
    if not user or not raw_id.isdigit():
        await callback.answer("Ошибка", show_alert=True)
        return
    cancelled = cancel_payment_request(int(raw_id), int(user.id))
    await state.clear()
    lang = get_user_language(int(user.id))
    await callback.answer("Bekor qilindi" if lang == "uz" else "Оплата отменена")
    if callback.message:
        await callback.message.answer(
            "To‘lov bekor qilindi." if lang == "uz" else "Оплата отменена. Вы можете выбрать тариф заново: /subscribe",
            reply_markup=main_menu_for_user(int(user.id)),
        )
    if not cancelled:
        logging.info("Payment cancel ignored: request=%s user=%s", raw_id, user.id)


@dp.message(PaymentStates.waiting_for_receipt, F.photo)
@dp.message(PaymentStates.waiting_for_receipt, F.document)
async def payment_receipt_received(message: Message, state: FSMContext) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    receipt_type = "photo"
    file_id = ""
    file_unique_id = ""
    if message.photo:
        photo = message.photo[-1]
        file_id = photo.file_id
        file_unique_id = photo.file_unique_id
    elif message.document:
        mime = str(message.document.mime_type or "").lower()
        if not (mime.startswith("image/") or mime == "application/pdf"):
            await message.answer(
                "Rasm yoki PDF chek yuboring." if lang == "uz" else "Отправьте чек как изображение или PDF-файл."
            )
            return
        receipt_type = "document"
        file_id = message.document.file_id
        file_unique_id = message.document.file_unique_id

    data = await state.get_data()
    request_id = int(data.get("payment_request_id") or 0)
    if not request_id:
        latest = latest_awaiting_payment_request(telegram_id)
        request_id = int(latest["id"]) if latest else 0
    if not request_id or not file_id:
        await state.clear()
        await message.answer(
            "Avval tarifni tanlang: /subscribe" if lang == "uz" else "Сначала выберите тариф: /subscribe",
            reply_markup=main_menu_for_user(telegram_id),
        )
        return

    row, error = attach_payment_receipt(
        request_id,
        telegram_id,
        file_id=file_id,
        file_unique_id=file_unique_id,
        receipt_type=receipt_type,
    )
    if error == "duplicate":
        await message.answer(
            "Bu chek avval yuborilgan." if lang == "uz" else "Этот чек уже использовался в другой заявке."
        )
        return
    if not row:
        await state.clear()
        await message.answer(
            "Ariza topilmadi. Qaytadan boshlang: /subscribe"
            if lang == "uz"
            else "Заявка не найдена. Начните заново: /subscribe",
            reply_markup=main_menu_for_user(telegram_id),
        )
        return

    await state.clear()
    if lang == "uz":
        client_text = (
            "✅ <b>Chek qabul qilindi</b>\n\n"
            f"Ariza: <b>#{int(row['id'])}</b>\n"
            "Administrator chekni tekshiradi. Natija haqida shu yerda xabar beramiz."
        )
    else:
        client_text = (
            "✅ <b>Чек принят</b>\n\n"
            f"Заявка: <b>#{int(row['id'])}</b>\n"
            "Администратор проверит чек. Результат придёт в этот чат."
        )
    await message.answer(client_text, reply_markup=main_menu_for_user(telegram_id))
    sent = await send_payment_request_to_admins(row)
    if not sent:
        logging.error("Payment request saved but no admin notification sent: request=%s", row["id"])


@dp.message(PaymentStates.waiting_for_receipt)
async def payment_receipt_wrong_format(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    await message.answer(
        "Chekni rasm yoki PDF shaklida yuboring."
        if lang == "uz"
        else "Отправьте чек фотографией, скриншотом или PDF-файлом."
    )


@dp.callback_query(F.data.startswith("payview:"))
async def payment_request_view_callback(callback: CallbackQuery) -> None:
    admin_id = int(callback.from_user.id) if callback.from_user else 0
    if not admin_only(admin_id):
        await callback.answer("Только для администратора", show_alert=True)
        return
    raw_id = str(callback.data or "").split(":", 1)[-1]
    row = get_payment_request(int(raw_id)) if raw_id.isdigit() else None
    if not row or row.get("status") != "pending_review":
        await callback.answer("Заявка уже обработана или не найдена", show_alert=True)
        return
    await callback.answer()
    await send_payment_request_to_admins({**row, "_only_admin": admin_id})


@dp.callback_query(F.data.startswith("payapprove:"))
async def payment_request_approve_callback(callback: CallbackQuery) -> None:
    admin_id = int(callback.from_user.id) if callback.from_user else 0
    if not admin_only(admin_id):
        await callback.answer("Только для администратора", show_alert=True)
        return
    raw_id = str(callback.data or "").split(":", 1)[-1]
    if not raw_id.isdigit():
        await callback.answer("Некорректная заявка", show_alert=True)
        return
    request_id = int(raw_id)
    result = approve_payment_request(request_id, admin_id)
    if not result:
        await callback.answer("Заявка уже обработана", show_alert=True)
        return

    target = int(result["telegram_id"])
    amount = _payment_amount_text(int(result["amount"]))
    new_until = _fmt_dt(result.get("new_until"))
    lang = get_user_language(target)
    if lang == "uz":
        client_text = (
            "✅ <b>To‘lov tasdiqlandi</b>\n\n"
            f"Tarif: <b>{int(result['plan_months'])} oy</b>\n"
            f"Summa: <b>{amount} so‘m</b>\n"
            f"Obuna muddati: <b>{new_until}</b> gacha.\n\n"
            "Rahmat! Botning barcha funksiyalari ishlaydi."
        )
    else:
        client_text = (
            "✅ <b>Оплата подтверждена</b>\n\n"
            f"Тариф: <b>{int(result['plan_months'])} мес.</b>\n"
            f"Сумма: <b>{amount} сум</b>\n"
            f"Подписка активна до: <b>{new_until}</b>\n\n"
            "Спасибо! Все функции бота доступны."
        )
    try:
        await bot.send_message(target, client_text, reply_markup=main_menu_for_user(target))
    except Exception:
        logging.exception("Approved payment client notification failed: request=%s user=%s", request_id, target)

    await callback.answer("Оплата подтверждена")
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        await callback.message.answer(
            "✅ <b>Оплата подтверждена</b>\n\n"
            f"Заявка: <b>#{request_id}</b>\n"
            f"Пользователь: <code>{target}</code>\n"
            f"Продлено до: <b>{new_until}</b>\n"
            f"Запись оплаты: <b>#{int(result['payment_history_id'])}</b>",
            reply_markup=admin_menu_for_user(admin_id),
        )


@dp.callback_query(F.data.startswith("payreject_menu:"))
async def payment_request_reject_menu_callback(callback: CallbackQuery) -> None:
    admin_id = int(callback.from_user.id) if callback.from_user else 0
    if not admin_only(admin_id):
        await callback.answer("Только для администратора", show_alert=True)
        return
    raw_id = str(callback.data or "").split(":", 1)[-1]
    row = get_payment_request(int(raw_id)) if raw_id.isdigit() else None
    if not row or row.get("status") != "pending_review":
        await callback.answer("Заявка уже обработана", show_alert=True)
        return
    await callback.answer()
    if callback.message:
        await callback.message.answer(
            f"Почему отклонить чек <b>#{int(raw_id)}</b>?",
            reply_markup=payment_rejection_markup(int(raw_id)),
        )


@dp.callback_query(F.data.startswith("payreject:"))
async def payment_request_reject_callback(callback: CallbackQuery) -> None:
    admin_id = int(callback.from_user.id) if callback.from_user else 0
    if not admin_only(admin_id):
        await callback.answer("Только для администратора", show_alert=True)
        return
    parts = str(callback.data or "").split(":", 2)
    if len(parts) != 3 or not parts[1].isdigit():
        await callback.answer("Некорректная заявка", show_alert=True)
        return
    request_id = int(parts[1])
    reason_code = parts[2]
    reasons = {
        "unreadable": "Чек не читается",
        "amount": "Сумма на чеке не совпадает с тарифом",
        "not_found": "Платёж не найден",
        "other": "Чек не прошёл проверку",
    }
    reason = reasons.get(reason_code, reasons["other"])
    row = reject_payment_request(request_id, admin_id, reason)
    if not row:
        await callback.answer("Заявка уже обработана", show_alert=True)
        return

    target = int(row["telegram_id"])
    lang = get_user_language(target)
    if lang == "uz":
        reason_uz = {
            "unreadable": "Chek aniq ko‘rinmayapti",
            "amount": "Chekdagi summa tarifga mos kelmaydi",
            "not_found": "To‘lov topilmadi",
            "other": "Chek tekshiruvdan o‘tmadi",
        }.get(reason_code, "Chek tekshiruvdan o‘tmadi")
        client_text = (
            "❌ <b>Chek tasdiqlanmadi</b>\n\n"
            f"Sabab: <b>{reason_uz}</b>\n\n"
            "Ma’lumotlarni tekshirib, yangi chek yuboring: /subscribe"
        )
    else:
        client_text = (
            "❌ <b>Чек не подтверждён</b>\n\n"
            f"Причина: <b>{escape(reason)}</b>\n\n"
            "Проверьте данные и отправьте новый чек: /subscribe"
        )
    try:
        await bot.send_message(target, client_text, reply_markup=main_menu_for_user(target))
    except Exception:
        logging.exception("Rejected payment client notification failed: request=%s user=%s", request_id, target)

    await callback.answer("Чек отклонён")
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        await callback.message.answer(
            f"❌ Чек <b>#{request_id}</b> отклонён. Причина: <b>{escape(reason)}</b>",
            reply_markup=admin_menu_for_user(admin_id),
        )


@dp.message(Command("video", "api_video", "instruction"))
async def video_instruction(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    if lang == "uz":
        text = (
            "🎥 <b>API-kalitni ulash bo‘yicha video</b>\n\n"
            "Videoda qisqa ko‘rsatilgan:\n"
            "1. Uzum Seller kabinetida API kalit qayerda joylashgan.\n"
            "2. Yangi kalit qanday yaratiladi.\n"
            "3. Kalit <b>🔌 Do‘konni ulash</b> tugmasi orqali qanday ulanadi.\n\n"
            "Videoni ko‘rish uchun pastdagi tugmani bosing 👇"
        )
    else:
        text = (
            "🎥 <b>Видеоинструкция по подключению API</b>\n\n"
            "В видео коротко показано:\n"
            "1. Где в кабинете Uzum Seller находятся ключи API.\n"
            "2. Как создать новый ключ.\n"
            "3. Как подключить ключ кнопкой <b>🔌 Подключить магазин</b>.\n\n"
            "Нажмите кнопку ниже, чтобы открыть видео 👇"
        )
    await message.answer(text, reply_markup=video_instruction_markup(lang) or menu_for_message(message))


@dp.message(Command("api_token", "token_help", "how_token"))
async def api_token_help(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    if lang == "uz":
        text = (
            "🔑 <b>Uzum API-kalitini botga ulash</b>\n\n"
            "🎥 Videoqo‘llanma pastdagi tugma orqali ochiladi.\n\n"
            "API-kalitni faqat Uzum Seller kabinetidan olasiz.\n"
            "Bu kabinet paroli emas, uni istalgan vaqtda o‘chirishingiz mumkin.\n\n"
            "API-kalit qayerda:\n"
            "1. Yuqori o‘ng burchakdagi profil / avatarni bosing.\n"
            "2. <b>Mening profilim</b> bo‘limini oching.\n"
            "3. <b>API kalitlari</b> ni bosing.\n"
            "4. <b>Kalit yaratish</b> ni bosing.\n"
            "5. API-kalitni nusxa oling.\n"
            "6. Botga qayting va <b>🔌 Do‘konni ulash</b> tugmasini bosing.\n"
            "7. Kalitni bitta xabar qilib yuboring.\n\n"
            "⚠️ API-kalitni begonalarga yubormang. Bot uni himoyalangan ko‘rinishda saqlaydi va xabarni o‘chirishga harakat qiladi."
        )
    else:
        text = (
            "🔑 <b>Как подключить Uzum API к боту</b>\n\n"
            "🎥 Видеоинструкция открывается кнопкой ниже.\n\n"
            "API-ключ создаётся только в вашем кабинете Uzum Seller.\n"
            "Это не пароль от кабинета, ключ можно удалить в любой момент.\n\n"
            "Где взять API-ключ:\n"
            "1. Нажмите на профиль / аватарку в правом верхнем углу.\n"
            "2. Откройте <b>Мой профиль</b>.\n"
            "3. Нажмите <b>Ключи API</b>.\n"
            "4. Нажмите <b>Создать ключ</b>.\n"
            "5. Скопируйте API-ключ.\n"
            "6. Вернитесь в бот и нажмите <b>🔌 Подключить магазин</b>.\n"
            "7. Отправьте ключ одним сообщением.\n\n"
            "⚠️ Не отправляйте ключ посторонним. Бот хранит его защищённо и старается удалить сообщение с ключом после проверки."
        )
    await message.answer(text, reply_markup=video_instruction_markup(lang) or menu_for_message(message))


@dp.message(Command("security", "privacy"))
async def security(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    if lang == "uz":
        text = (
            "🔐 <b>API-kalit xavfsizligi</b>\n\n"
            "Uzum API-kalitingiz botda ko‘rsatilmaydi va sizga qayta xabar qilib yuborilmaydi.\n"
            "Ulangandan keyin bot kalit yuborilgan xabarni o‘chirishga harakat qiladi.\n"
            "Bazaga faqat himoyalangan versiya saqlanadi.\n\n"
            "Kalitni almashtirish uchun <b>⚙️ Sozlamalar → 🔐 Uzum ulanishi → "
            "🔌 API-kalitni yangilash</b> bo‘limidan foydalaning."
        )
    else:
        text = (
            "🔐 <b>Безопасность API-ключа</b>\n\n"
            "Ваш Uzum API-ключ не показывается в боте и не отправляется обратно сообщением.\n"
            "После подключения бот старается удалить сообщение, где был отправлен ключ.\n"
            "В базе хранится только защищённая версия ключа.\n\n"
            "Чтобы заменить ключ, откройте <b>⚙️ Настройки → 🔐 Подключение Uzum → "
            "🔌 Обновить API-ключ</b>."
        )
    await message.answer(text, reply_markup=connection_menu_for_message(message))


@dp.message(Command("support"))
async def support(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    user = db.get_user(telegram_id)
    shops = db.list_shops(telegram_id)
    connected = bool(user and user["uzum_token_encrypted"])
    if lang == "uz":
        text = (
            "🛟 <b>Uzum Seller Assistant yordami</b>\n\n"
            f"Telegram ID: <code>{telegram_id}</code>\n"
            f"Uzum API: {'✅ ulangan' if connected else '❌ ulanmagan'}\n"
            f"Topilgan do‘konlar: <b>{len(shops)}</b>\n"
            f"Obuna: {subscription_status_text(telegram_id)}\n\n"
            "Agar bot ma’lumotlarni ko‘rsatmasa, tekshiring:\n"
            "1. API-kalit Uzum Seller kabinetida faol.\n"
            "2. Kalit kerakli do‘konga ruxsatga ega.\n"
            "3. Tanlangan davr bo‘yicha Uzum kabinetida savdolar bor.\n"
            "4. API-kalitni o‘zgartirgan bo‘lsangiz — <code>/reconnect</code> ni bosing.\n\n"
            f"Administrator bilan bog‘lanish: <b>{admin_contact_text()}</b>"
        )
    else:
        text = (
            "🛟 <b>Поддержка Uzum Seller Assistant</b>\n\n"
            f"Ваш Telegram ID: <code>{telegram_id}</code>\n"
            f"Uzum API: {'✅ подключён' if connected else '❌ не подключён'}\n"
            f"Магазинов найдено: <b>{len(shops)}</b>\n"
            f"Подписка: {subscription_status_text(telegram_id)}\n\n"
            "Если бот не показывает данные, проверьте:\n"
            "1. API-ключ активен в кабинете Uzum Seller.\n"
            "2. У ключа есть доступ к нужному магазину.\n"
            "3. В кабинете Uzum есть продажи за выбранный период.\n"
            "4. Если меняли API-ключ — нажмите <code>/reconnect</code>.\n\n"
            f"Связаться с администратором: <b>{admin_contact_text()}</b>"
        )
    await message.answer(text, reply_markup=admin_contact_markup() or menu_for_message(message))


@dp.message(Command("my_payments"))
async def my_payments(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    pending = latest_payment_request_by_status(telegram_id, "pending_review")
    rows = list_payments(telegram_id, 10)
    if not rows and not pending:
        await message.answer(
            "💳 To‘lovlar tarixi hozircha bo‘sh." if lang == "uz" else "💳 История оплат пока пустая.",
            reply_markup=menu_for_message(message),
        )
        return
    parts = ["💳 <b>Mening to‘lovlarim</b>" if lang == "uz" else "💳 <b>Мои оплаты</b>"]
    if pending:
        status_text = (
            f"🕐 Chek <b>#{int(pending['id'])}</b> tekshirilmoqda"
            if lang == "uz"
            else f"🕐 Чек <b>#{int(pending['id'])}</b> находится на проверке"
        )
        parts.append(status_text)
    if rows:
        parts.append("\n".join(payment_line(row) for row in rows))
    await message.answer("\n\n".join(parts), reply_markup=menu_for_message(message))

@dp.message(Command("my_subscription", "subscription"))
async def my_subscription(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    await message.answer(subscription_full_text(telegram_id), reply_markup=menu_for_message(message))


@dp.message(Command("admin"))
async def admin_panel(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        await message.answer("⛔ Админ-панель доступна только владельцу бота.", reply_markup=menu_for_message(message))
        return
    init_business_tables()
    stats = get_admin_stats()
    money_today = f"{stats['payments_today']:,}".replace(",", " ")
    money_30 = f"{stats['payments_30']:,}".replace(",", " ")
    backup_hint = (
        "\n• <code>/backup_db</code> — скачать базу"
        if owner_only(admin_id)
        else ""
    )
    await message.answer(
        "👑 <b>Админ-панель</b>\n\n"
        f"👥 Пользователей всего: <b>{stats['total_users']}</b>\n"
        f"🔑 Подключили Uzum API: <b>{stats['connected']}</b>\n"
        f"✅ Активных доступов: <b>{stats['active']}</b>\n"
        f"💳 Платных: <b>{stats['paid']}</b>\n"
        f"🎁 Trial: <b>{stats['trial']}</b>\n"
        f"⛔ Истекли: <b>{stats['expired']}</b>\n"
        f"🚫 Заблокированы: <b>{stats['blocked']}</b>\n\n"
        f"💰 Оплаты сегодня: <b>{money_today}</b> сум\n"
        f"💰 Оплаты за 30 дней: <b>{money_30}</b> сум\n\n"
        f"🧾 Чеков на проверке: <b>{stats['pending_receipts']}</b>\n\n"
        "Быстрые команды:\n"
        "• <code>/payments</code> — проверить новые чеки\n"
        "• <code>/paid1 ID</code> — 1 месяц / 300 000 сум\n"
        "• <code>/paid3 ID</code> — 3 месяца / 800 000 сум\n"
        "• <code>/paid6 ID</code> — 6 месяцев / 1 500 000 сум\n"
        "• <code>/expiring</code> — кто скоро заканчивается"
        + backup_hint,
        reply_markup=admin_menu_for_message(message),
    )


@dp.message(Command("check"))
async def check_connection(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    ensure_subscription(telegram_id)
    user = db.get_user(telegram_id)
    client = get_uzum_for_user(telegram_id)
    shop_id = db.get_default_shop_id(telegram_id)
    lines = [
        "✅ <b>Проверка подключения</b>",
        f"Telegram ID: <code>{telegram_id}</code>",
        f"Подписка: {subscription_status_text(telegram_id)}",
        f"Uzum API: {'✅ подключён' if client else '❌ не подключён'}",
        f"Активный магазин: {f'<code>{shop_id}</code>' if shop_id else '—'}",
    ]
    if client is None:
        lines.append("\nЧто делать: нажмите <code>/connect</code> и отправьте Uzum API-ключ.")
        await message.answer("\n".join(lines), reply_markup=menu_for_message(message))
        return
    try:
        data = await client.get_shops()
        shops = extract_items(data)
        encrypted = db.get_encrypted_token(telegram_id)
        if encrypted and shops:
            db.save_connection(telegram_id, encrypted, shops)
        lines.append(f"Магазинов найдено: <b>{len(shops)}</b>")
    except Exception as e:
        lines.append("Магазины: ❌ ошибка")
        lines.append(f"Причина: <code>{escape(str(e))[:500]}</code>")
        await message.answer("\n".join(lines), reply_markup=menu_for_message(message))
        return
    if shop_id:
        try:
            rows = await load_sku_rows(client, int(shop_id), max_pages=1)
            lines.append(f"Остатки/товары: ✅ доступно, SKU строк: <b>{len(rows)}</b>")
        except Exception as e:
            lines.append("Остатки/товары: ❌ ошибка")
            lines.append(f"Причина: <code>{escape(str(e))[:500]}</code>")
        try:
            date_from, date_to = _today_range_ms()
            sales_rows, _ = await _load_finance_orders(client, int(shop_id), date_from_ms=date_from, date_to_ms=date_to, max_pages=1, page_size=20)
            lines.append(f"Finance API: ✅ доступно, продаж сегодня: <b>{len(sales_rows)}</b>")
        except Exception as e:
            lines.append("Finance API: ❌ ошибка")
            lines.append(f"Причина: <code>{escape(str(e))[:500]}</code>")
    lines.append("\nЕсли здесь всё ✅ — бот готов к работе.")
    await message.answer("\n".join(lines), reply_markup=menu_for_message(message))


def subscription_queue_markup(rows: list[dict[str, Any]]) -> InlineKeyboardMarkup | None:
    buttons = []
    for row in rows[:10]:
        queue_id = int(row.get("id") or 0)
        label = _renewal_user_label(row)
        buttons.append([
            InlineKeyboardButton(
                text=f"👁 #{queue_id} · {label}"[:60],
                callback_data=f"renewal_preview:{queue_id}",
            )
        ])
    return InlineKeyboardMarkup(inline_keyboard=buttons) if buttons else None


def subscription_reminder_review_markup(queue_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="✅ Подтвердить и отправить",
                    callback_data=f"renewal_send:{int(queue_id)}",
                )
            ],
            [
                InlineKeyboardButton(
                    text="🗑 Не отправлять",
                    callback_data=f"renewal_dismiss:{int(queue_id)}",
                )
            ],
        ]
    )


def subscription_reminder_preview_text(row: dict[str, Any]) -> str:
    last_error = str(row.get("last_error") or "").strip()
    error_text = f"\n\n⚠️ Последняя ошибка: <code>{escape(last_error)}</code>" if last_error else ""
    return (
        "👁 <b>Проверка напоминания</b>\n\n"
        f"Очередь: <b>#{int(row.get('id') or 0)}</b>\n"
        f"Клиент: <code>{int(row.get('telegram_id') or 0)}</code> — {escape(_renewal_user_label(row))}\n"
        f"Доступ до: <b>{_fmt_dt(row.get('active_until'))}</b>\n\n"
        "<b>Текст, который получит клиент:</b>\n\n"
        f"{row.get('draft_text') or '—'}"
        f"{error_text}\n\n"
        "Сообщение будет отправлено только после нажатия кнопки подтверждения."
    )


@dp.message(Command("expiring", "renewals"))
async def admin_expiring(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    summary = refresh_subscription_reminder_queue()
    rows = list_pending_subscription_reminders(100)
    await message.answer(
        build_subscription_action_digest(rows, int(summary.get("created") or 0)),
        reply_markup=subscription_queue_markup(rows) or admin_menu_for_message(message),
    )


@dp.message(Command("reminder"))
async def admin_subscription_reminder_preview(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    args = parse_args(message.text or "").split()
    if not args or not args[0].isdigit():
        await message.answer(
            "Откройте список <code>/renewals</code>, затем напишите <code>/reminder НОМЕР</code>.",
            reply_markup=admin_menu_for_message(message),
        )
        return
    row = get_subscription_reminder(int(args[0]), pending_only=True)
    if not row:
        await message.answer(
            "Этот черновик уже обработан или не найден. Обновите список: <code>/renewals</code>.",
            reply_markup=admin_menu_for_message(message),
        )
        return
    await message.answer(
        subscription_reminder_preview_text(row),
        reply_markup=subscription_reminder_review_markup(int(row["id"])),
    )


@dp.callback_query(F.data.startswith("renewal_preview:"))
async def subscription_reminder_preview_callback(callback: CallbackQuery) -> None:
    admin_id = int(callback.from_user.id) if callback.from_user else 0
    if not admin_only(admin_id):
        await callback.answer("Только для администратора", show_alert=True)
        return
    raw_id = str(callback.data or "").split(":", 1)[-1]
    row = get_subscription_reminder(int(raw_id), pending_only=True) if raw_id.isdigit() else None
    if not row:
        await callback.answer("Черновик уже обработан или не найден", show_alert=True)
        return
    await callback.answer()
    if callback.message:
        await callback.message.answer(
            subscription_reminder_preview_text(row),
            reply_markup=subscription_reminder_review_markup(int(row["id"])),
        )


@dp.callback_query(F.data.startswith("renewal_send:"))
async def subscription_reminder_send_callback(callback: CallbackQuery) -> None:
    admin_id = int(callback.from_user.id) if callback.from_user else 0
    if not admin_only(admin_id):
        await callback.answer("Только для администратора", show_alert=True)
        return
    raw_id = str(callback.data or "").split(":", 1)[-1]
    if not raw_id.isdigit():
        await callback.answer("Некорректный номер", show_alert=True)
        return
    queue_id = int(raw_id)
    row = get_subscription_reminder(queue_id, pending_only=True)
    if not row or not claim_subscription_reminder(queue_id, admin_id):
        await callback.answer("Черновик уже обработан", show_alert=True)
        return

    target = int(row.get("telegram_id") or 0)
    try:
        await bot.send_message(
            target,
            str(row.get("draft_text") or ""),
            reply_markup=main_menu_for_user(target),
        )
        finish_subscription_reminder(queue_id, sent=True)
    except Exception as exc:
        logging.exception("Confirmed subscription reminder failed: queue=%s user=%s", queue_id, target)
        finish_subscription_reminder(queue_id, sent=False, error=str(exc))
        await callback.answer("Не удалось отправить. Черновик сохранён.", show_alert=True)
        return

    await callback.answer("Напоминание отправлено")
    if callback.message:
        await callback.message.answer(
            f"✅ Напоминание <b>#{queue_id}</b> отправлено пользователю <code>{target}</code> после вашего подтверждения.",
            reply_markup=admin_menu_for_user(admin_id),
        )


@dp.callback_query(F.data.startswith("renewal_dismiss:"))
async def subscription_reminder_dismiss_callback(callback: CallbackQuery) -> None:
    admin_id = int(callback.from_user.id) if callback.from_user else 0
    if not admin_only(admin_id):
        await callback.answer("Только для администратора", show_alert=True)
        return
    raw_id = str(callback.data or "").split(":", 1)[-1]
    if not raw_id.isdigit() or not dismiss_subscription_reminder(int(raw_id), admin_id):
        await callback.answer("Черновик уже обработан или не найден", show_alert=True)
        return
    await callback.answer("Черновик отклонён")
    if callback.message:
        await callback.message.answer(
            f"🗑 Черновик <b>#{int(raw_id)}</b> отмечен как «не отправлять».",
            reply_markup=admin_menu_for_user(admin_id),
        )


@dp.message(Command("blocked"))
async def admin_blocked_users(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    rows = list_blocked_users(50)
    if not rows:
        await message.answer("⛔ Заблокированных пользователей нет.", reply_markup=admin_menu_for_message(message))
        return
    await message.answer(
        "⛔ <b>Заблокированные пользователи</b>\n\n" + "\n".join(subscription_compact_line(r) for r in rows),
        reply_markup=admin_menu_for_message(message),
    )


@dp.message(Command("staff_connections"))
async def staff_connections_admin(message: Message) -> None:
    await message.answer(
        "🔑 Подключение через сотрудника отключено. Сейчас используется официальный способ через API-ключ: /connect",
        reply_markup=admin_menu_for_message(message),
    )


@dp.message(F.text == "🔌 Через сотрудника")
@dp.message(F.text == "🔌 Xodim orqali")
async def staff_connections_button(message: Message) -> None:
    await staff_connections_admin(message)



@dp.message(F.text.in_({"🔑 Подключение API", "🔑 API ulanishi"}))
async def admin_api_connections(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not admin_only(telegram_id):
        await message.answer(tr_user(telegram_id, "admin_only"))
        return

    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT u.telegram_id, u.username, u.first_name, u.default_shop_id,
                   u.uzum_token_encrypted,
                   s.trial_until, s.subscription_until, s.blocked
            FROM users u
            LEFT JOIN subscriptions s ON s.telegram_id = u.telegram_id
            ORDER BY CASE WHEN u.uzum_token_encrypted IS NOT NULL THEN 0 ELSE 1 END,
                     u.updated_at DESC
            """
        ).fetchall()

    connected = [dict(r) for r in rows if r["uzum_token_encrypted"]]
    not_connected = [dict(r) for r in rows if not r["uzum_token_encrypted"]]

    lines = [
        "🔑 <b>Подключение Uzum API</b>",
        "",
        f"👥 Всего пользователей: <b>{len(rows)}</b>",
        f"✅ Подключили API: <b>{len(connected)}</b>",
        f"❌ Не подключили API: <b>{len(not_connected)}</b>",
        "",
        "✅ <b>API подключён</b>",
    ]

    def user_line(row: dict[str, Any], ok: bool) -> str:
        username = str(row.get("username") or "").strip()
        name = str(row.get("first_name") or "").strip()
        label = f"@{username}" if username else (name or "без имени")
        shop = row.get("default_shop_id")
        shop_text = str(shop) if shop else "магазин не выбран"
        return (
            f"{'✅' if ok else '❌'} <code>{int(row['telegram_id'])}</code> — "
            f"{escape(label)} | 🏪 {escape(shop_text)}"
        )

    if connected:
        lines.extend(user_line(row, True) for row in connected)
    else:
        lines.append("— пока никого")

    lines.extend(["", "❌ <b>API не подключён</b>"])
    if not_connected:
        lines.extend(user_line(row, False) for row in not_connected)
    else:
        lines.append("— все подключили API")

    # Telegram ограничивает одно сообщение примерно 4096 символами.
    chunk = ""
    for line in lines:
        candidate = f"{chunk}\n{line}" if chunk else line
        if len(candidate) > 3800:
            await message.answer(chunk, reply_markup=ADMIN_PANEL_MENU_RU)
            chunk = line
        else:
            chunk = candidate
    if chunk:
        await message.answer(chunk, reply_markup=ADMIN_PANEL_MENU_RU)


@dp.message(Command("users"))
async def admin_users(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not admin_only(telegram_id):
        return
    rows = list_subscription_users(30)
    if not rows:
        await message.answer("Пользователей пока нет.", reply_markup=menu_for_message(message))
        return
    lines = [subscription_compact_line(row) for row in rows]
    await message.answer(
        "👥 <b>Пользователи</b>\n\n"
        + "\n".join(lines)
        + "\n\nКоманды: <code>/renewals</code>, <code>/extend ID 30</code>, <code>/paid ID сумма дни</code>, <code>/payments</code>",
        reply_markup=menu_for_message(message),
    )


@dp.message(Command("user"))
async def admin_user_info(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "")
    if not arg.split() or not arg.split()[0].isdigit():
        await message.answer("Напишите так: <code>/user TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg.split()[0])
    row = ensure_subscription(target)
    user = db.get_user(target)
    username_text = f"@{escape(str(user['username']))}" if user and user['username'] else "—"
    shop_text = f"<code>{user['default_shop_id']}</code>" if user and user['default_shop_id'] else "—"
    payments = list_payments(target, 5)
    payments_text = "\n".join(payment_line(p) for p in payments) if payments else "—"
    await message.answer(
        "👤 <b>Пользователь</b>\n\n"
        f"ID: <code>{target}</code>\n"
        f"Username: {username_text}\n"
        f"Uzum API: {'✅ подключён' if user and user['uzum_token_encrypted'] else '❌ не подключён'}\n"
        f"Магазин: {shop_text}\n"
        f"Статус: {subscription_status_text(target)}\n"
        f"Trial до: <b>{_fmt_dt(row.get('trial_until'))}</b>\n"
        f"Оплачено до: <b>{_fmt_dt(row.get('subscription_until'))}</b>\n\n"
        f"💳 Последние оплаты:\n{payments_text}",
        reply_markup=menu_for_message(message),
    )


@dp.message(Command("extend"))
async def admin_extend(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    parts = parse_args(message.text or "").split()
    if len(parts) < 2 or not parts[0].isdigit() or not parts[1].isdigit():
        await message.answer("Напишите так: <code>/extend TELEGRAM_ID 30</code>", reply_markup=menu_for_message(message))
        return
    target = int(parts[0])
    days = int(parts[1])
    new_until = extend_subscription_days(target, days)
    await message.answer(
        f"✅ Доступ продлён для <code>{target}</code> на {days} дней.\n"
        f"Активен до: <b>{_fmt_dt(new_until)}</b>",
        reply_markup=menu_for_message(message),
    )
    try:
        await bot.send_message(
            target,
            f"✅ Ваша подписка продлена на {days} дней.\nАктивна до: <b>{_fmt_dt(new_until)}</b>",
            reply_markup=main_menu_for_user(target),
        )
    except Exception:
        pass




@dp.message(Command("paid"))
async def admin_paid(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    parts = parse_args(message.text or "").split(maxsplit=3)
    if len(parts) < 3 or not parts[0].isdigit() or not parts[1].replace("_", "").isdigit() or not parts[2].isdigit():
        await message.answer(
            "Напишите так: <code>/paid TELEGRAM_ID СУММА ДНИ комментарий</code>\n"
            "Пример: <code>/paid 123456789 300000 30 чек Click</code>",
            reply_markup=menu_for_message(message),
        )
        return
    target = int(parts[0])
    amount = int(parts[1].replace("_", ""))
    days = int(parts[2])
    comment = parts[3] if len(parts) > 3 else "ручная оплата"
    new_until = extend_subscription_days(target, days)
    payment_id = record_payment(target, amount, days, admin_id, comment)
    amount_text = f"{amount:,}".replace(",", " ")
    await message.answer(
        f"✅ Оплата записана #{payment_id}\n"
        f"Пользователь: <code>{target}</code>\n"
        f"Сумма: <b>{amount_text}</b> сум\n"
        f"Продление: <b>{days}</b> дней\n"
        f"Доступ до: <b>{_fmt_dt(new_until)}</b>",
        reply_markup=menu_for_message(message),
    )
    try:
        await bot.send_message(
            target,
            "✅ <b>Оплата подтверждена</b>\n\n"
            f"Подписка продлена на <b>{days}</b> дней.\n"
            f"Доступ активен до: <b>{_fmt_dt(new_until)}</b>\n\n"
            "Спасибо за оплату!",
            reply_markup=main_menu_for_user(target),
        )
    except Exception:
        pass


@dp.message(Command("paid1"))
async def admin_paid_1_month(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "").split(maxsplit=1)
    if not arg or not arg[0].isdigit():
        await message.answer("Напишите так: <code>/paid1 TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg[0])
    new_until = extend_subscription_days(target, 30)
    payment_id = record_payment(target, 300000, 30, admin_id, "1 месяц")
    await message.answer(f"✅ Оплата #{payment_id}: 300 000 сум. Доступ для <code>{target}</code> до <b>{_fmt_dt(new_until)}</b>", reply_markup=menu_for_message(message))
    try:
        await bot.send_message(target, f"✅ Оплата подтверждена. Подписка продлена на 1 месяц. Доступ до: <b>{_fmt_dt(new_until)}</b>", reply_markup=main_menu_for_user(target))
    except Exception:
        pass


@dp.message(Command("paid3"))
async def admin_paid_3_months(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "").split(maxsplit=1)
    if not arg or not arg[0].isdigit():
        await message.answer("Напишите так: <code>/paid3 TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg[0])
    new_until = extend_subscription_days(target, 90)
    payment_id = record_payment(target, 800000, 90, admin_id, "3 месяца")
    await message.answer(f"✅ Оплата #{payment_id}: 800 000 сум. Доступ для <code>{target}</code> до <b>{_fmt_dt(new_until)}</b>", reply_markup=menu_for_message(message))
    try:
        await bot.send_message(target, f"✅ Оплата подтверждена. Подписка продлена на 3 месяца. Доступ до: <b>{_fmt_dt(new_until)}</b>", reply_markup=main_menu_for_user(target))
    except Exception:
        pass


@dp.message(Command("paid6"))
async def admin_paid_6_months(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "").split(maxsplit=1)
    if not arg or not arg[0].isdigit():
        await message.answer("Напишите так: <code>/paid6 TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg[0])
    new_until = extend_subscription_days(target, 180)
    payment_id = record_payment(target, 1500000, 180, admin_id, "6 месяцев")
    await message.answer(f"✅ Оплата #{payment_id}: 1 500 000 сум. Доступ для <code>{target}</code> до <b>{_fmt_dt(new_until)}</b>", reply_markup=menu_for_message(message))
    try:
        await bot.send_message(target, f"✅ Оплата подтверждена. Подписка продлена на 6 месяцев. Доступ до: <b>{_fmt_dt(new_until)}</b>", reply_markup=main_menu_for_user(target))
    except Exception:
        pass


def pending_payment_requests_markup(rows: list[dict[str, Any]]) -> InlineKeyboardMarkup | None:
    buttons = []
    for row in rows[:10]:
        request_id = int(row.get("id") or 0)
        label = payment_request_user_label(row)
        amount = _payment_amount_text(int(row.get("amount") or 0))
        buttons.append([
            InlineKeyboardButton(
                text=f"🧾 #{request_id} · {label} · {amount}"[:60],
                callback_data=f"payview:{request_id}",
            )
        ])
    return InlineKeyboardMarkup(inline_keyboard=buttons) if buttons else None


@dp.message(Command("payments"))
async def admin_payments(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    args = parse_args(message.text or "").split()
    target = int(args[0]) if args and args[0].isdigit() else None
    pending = list_payment_requests("pending_review", 50)
    rows = list_payments(target, 20)
    if pending:
        pending_lines = [
            f"🧾 <b>#{int(row['id'])}</b> · <code>{int(row['telegram_id'])}</code> — "
            f"{escape(payment_request_user_label(row))} · "
            f"{int(row['plan_months'])} мес. · {_payment_amount_text(int(row['amount']))} сум"
            for row in pending[:20]
        ]
        await message.answer(
            f"🕐 <b>Чеки ожидают проверки: {len(pending)}</b>\n\n" + "\n".join(pending_lines),
            reply_markup=pending_payment_requests_markup(pending),
        )
    if rows:
        title = f"💳 <b>Оплаты пользователя <code>{target}</code></b>" if target else "💳 <b>Последние подтверждённые оплаты</b>"
        await message.answer(title + "\n\n" + "\n".join(payment_line(row) for row in rows), reply_markup=menu_for_message(message))
    elif not pending:
        await message.answer("💳 Чеков на проверке и подтверждённых оплат пока нет.", reply_markup=menu_for_message(message))


def create_sqlite_backup(source_path: str | Path, destination_path: str | Path) -> Path:
    """Create a transactionally consistent SQLite snapshot, including WAL data."""
    source = Path(source_path)
    destination = Path(destination_path)
    if not source.is_file():
        raise FileNotFoundError(str(source))
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.unlink(missing_ok=True)

    source_conn: sqlite3.Connection | None = None
    destination_conn: sqlite3.Connection | None = None
    try:
        source_conn = sqlite3.connect(
            str(source),
            timeout=SQLITE_BUSY_TIMEOUT_MS / 1000,
        )
        destination_conn = sqlite3.connect(
            str(destination),
            timeout=SQLITE_BUSY_TIMEOUT_MS / 1000,
        )
        source_conn.backup(destination_conn, pages=1000, sleep=0.05)
        destination_conn.commit()
        integrity = destination_conn.execute("PRAGMA integrity_check").fetchone()
        if not integrity or str(integrity[0]).lower() != "ok":
            raise RuntimeError(f"SQLite backup integrity check failed: {integrity!r}")
    except Exception:
        destination.unlink(missing_ok=True)
        raise
    finally:
        if destination_conn is not None:
            destination_conn.close()
        if source_conn is not None:
            source_conn.close()
    try:
        destination.chmod(0o600)
    except OSError:
        pass
    return destination


@dp.message(Command("backup_db"))
async def admin_backup_db(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not owner_only(admin_id):
        await message.answer(
            "⛔ Резервную копию базы может скачать только владелец бота.",
            reply_markup=admin_menu_for_message(message) if admin_only(admin_id) else menu_for_message(message),
        )
        return
    path = Path(DB_PATH)
    if not path.exists():
        await message.answer(f"❌ База не найдена: <code>{escape(str(path))}</code>", reply_markup=menu_for_message(message))
        return
    await message.answer("📦 Создаю целостную резервную копию базы. Храните файл аккуратно — там данные пользователей.", reply_markup=menu_for_message(message))
    backup_path = Path(tempfile.gettempdir()) / f"bot_backup_{admin_id}_{time.time_ns()}.db"
    try:
        await asyncio.to_thread(create_sqlite_backup, path, backup_path)
        await message.answer_document(
            FSInputFile(
                str(backup_path),
                filename=f"bot_backup_{datetime.now(UZT).strftime('%Y%m%d_%H%M')}.db",
            )
        )
    except Exception as e:
        await send_api_error(message, e)
    finally:
        backup_path.unlink(missing_ok=True)

@dp.message(Command("trial"))
async def admin_trial(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    parts = parse_args(message.text or "").split()
    if len(parts) < 2 or not parts[0].isdigit() or not parts[1].isdigit():
        await message.answer("Напишите так: <code>/trial TELEGRAM_ID 3</code>", reply_markup=menu_for_message(message))
        return
    target = int(parts[0])
    days = int(parts[1])
    new_until = set_trial_days(target, days)
    await message.answer(
        f"🎁 Пользователю <code>{target}</code> добавлено <b>{max(1, days)}</b> дн. trial. "
        f"Новый срок: <b>{_fmt_dt(new_until)}</b>",
        reply_markup=menu_for_message(message),
    )


@dp.message(Command("block"))
async def admin_block(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "").split()
    if not arg or not arg[0].isdigit():
        await message.answer("Напишите так: <code>/block TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg[0])
    set_blocked(target, True)
    await message.answer(f"⛔ Пользователь <code>{target}</code> заблокирован.", reply_markup=menu_for_message(message))


@dp.message(Command("unblock"))
async def admin_unblock(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "").split()
    if not arg or not arg[0].isdigit():
        await message.answer("Напишите так: <code>/unblock TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg[0])
    set_blocked(target, False)
    await message.answer(f"✅ Пользователь <code>{target}</code> разблокирован.", reply_markup=menu_for_message(message))


@dp.message(Command("broadcast"))
async def admin_broadcast(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    text = parse_args(message.text or "")
    if not text:
        await message.answer("Напишите так: <code>/broadcast текст рассылки</code>", reply_markup=menu_for_message(message))
        return
    rows = list_subscription_users(500)
    sent = 0
    for row in rows:
        target = int(row["telegram_id"])
        try:
            await bot.send_message(target, "📢 <b>Сообщение от администратора</b>\n\n" + text, reply_markup=main_menu_for_user(target))
            sent += 1
            await asyncio.sleep(0.05)
        except Exception:
            pass
    await message.answer(f"✅ Рассылка завершена. Отправлено: {sent}", reply_markup=menu_for_message(message))


@dp.message(Command("debug_product"))
async def debug_product(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req

    try:
        data = await client.get_products(shop_id, page=0, size=1)
        items = extract_items(data)
        if not items:
            await message.answer(
                "Товар для debug не найден. Ответ API:\n<code>"
                + escape(compact_json_preview(data, limit=3000))
                + "</code>",
                reply_markup=menu_for_message(message),
            )
            return

        await message.answer(
            "🧪 <b>Первый товар — сырой JSON</b>\n\n<code>"
            + escape(compact_json_preview(items[0], limit=3200))
            + "</code>",
            reply_markup=menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("export_products"))
async def export_products(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req

    try:
        rows = await load_sku_rows(client, shop_id, max_pages=50)
        if not rows:
            await message.answer("SKU-остатки для экспорта не найдены.", reply_markup=menu_for_message(message))
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Stocks"
        ws.append([
            "Product ID",
            "SKU ID",
            "Barcode",
            "Seller code",
            "Category",
            "Product title",
            "SKU title",
            "Цена продажи",
            "Себестоимость Uzum",
            "ИКПУ / МХИК",
            "Платное хранение",
            "Тариф хранения / SKU",
            "Начислено хранения",
            "FBO / склад Uzum",
            "FBS/DBS / склад продавца",
            "Итого доступно",
            "Активно",
            "Продано",
            "Возвраты",
            "Недостача",
            "Брак",
            "Ожидает",
            "Статус",
        ])

        for r in rows:
            ws.append([
                excel_value(r.get("product_id")),
                excel_value(r.get("sku_id")),
                excel_value(r.get("barcode")),
                excel_value(r.get("seller_item_code")),
                excel_value(r.get("category")),
                excel_value(r.get("product_title")),
                excel_value(r.get("sku_full_title") or r.get("sku_title")),
                excel_value(r.get("price")),
                excel_value(r.get("purchase_price")),
                excel_value(r.get("ikpu")),
                "Да" if r.get("paid_storage") else "Нет",
                excel_value(r.get("paid_storage_price_item")),
                excel_value(r.get("paid_storage_amount")),
                excel_value(r.get("fbo")),
                excel_value(r.get("fbs")),
                excel_value(r.get("total")),
                excel_value(r.get("active")),
                excel_value(r.get("sold")),
                excel_value(r.get("returned")),
                excel_value(r.get("missing")),
                excel_value(r.get("defected")),
                excel_value(r.get("pending")),
                excel_value(status_display(r.get("status")) if r.get("status") else ""),
            ])

        for column in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 70)

        tmp_dir = Path(tempfile.gettempdir())
        filename = tmp_dir / f"uzum_stocks_{shop_id}.xlsx"
        wb.save(filename)
        await message.answer(f"✅ Экспортировано SKU-остатков: {len(rows)}", reply_markup=menu_for_message(message))
        await message.answer_document(FSInputFile(filename))
    except Exception as e:
        await send_api_error(message, e)


# --- Управленческий Excel-отчёт ---
def _datetime_for_report(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, (int, float)):
        timestamp = float(value)
        if timestamp > 10_000_000_000:
            timestamp /= 1000.0
        try:
            dt = datetime.fromtimestamp(timestamp, UZT)
        except (OverflowError, OSError, ValueError):
            return None
    else:
        text = str(value).strip()
        if not text:
            return None
        if text.isdigit():
            return _datetime_for_report(int(text))
        normalized = text.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(normalized)
        except ValueError:
            for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d"):
                try:
                    dt = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
            else:
                return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UZT)
    return dt.astimezone(UZT).replace(tzinfo=None)


def _finance_order_datetime(item: dict[str, Any]) -> datetime | None:
    """Date when Uzum accepted/created the finance order line.

    ``date`` is the documented Seller Finance API field.  Keeping it ahead of
    generic creation/update timestamps prevents a later metadata update from
    moving a sale to another reporting day.
    """
    value = pick(
        item,
        "date",
        "orderDate",
        "createdAt",
        "dateCreated",
        "created",
        "paymentDate",
        "updatedAt",
    )
    if isinstance(value, dict):
        value = pick(value, "date", "value", "createdAt")
    return _datetime_for_report(value)


def _finance_issued_datetime(item: dict[str, Any]) -> datetime | None:
    """Date when the order was issued/picked up by the customer."""
    value = pick(
        item,
        "dateIssued",
        "issuedDate",
        "dateDelivered",
        "completedDate",
        default=None,
    )
    if isinstance(value, dict):
        value = pick(value, "date", "value", "dateIssued", default=None)
    return _datetime_for_report(value)


def _finance_datetime_for_report(item: dict[str, Any]) -> datetime | None:
    # Compatibility name used by existing reports: their primary period is the
    # order/acceptance day.  Issued sales use `_finance_issued_datetime`.
    return _finance_order_datetime(item)


def _normalize_finance_row(item: dict[str, Any]) -> dict[str, Any]:
    status = _finance_status(item)
    gross = _finance_gross_revenue(item)
    commission = _finance_commission(item)
    logistics = _finance_logistics(item)
    direct = _finance_payout_direct(item)
    payout = direct if direct is not None else max(0.0, gross - commission - logistics)
    kind = "cancel" if _is_cancelled_status(status) else "return" if _is_returned_status(status) else "sale"
    return {
        "date": _finance_datetime_for_report(item),
        "date_issued": _finance_issued_datetime(item),
        "kind": kind,
        "status": status,
        "order_id": _finance_order_key_for_stats(item),
        "title": _finance_title(item),
        "sku": _finance_sku_key_for_stats(item),
        "qty": _finance_qty(item),
        "revenue": gross,
        "commission": commission,
        "logistics": logistics,
        "payout": max(0.0, payout),
        "withdrawn": _finance_withdrawn(item),
    }


def _daily_report_rows(
    rows: list[dict[str, Any]],
    date_from_ms: int,
    date_to_ms: int,
) -> list[dict[str, Any]]:
    buckets: dict[str, list[dict[str, Any]]] = {}
    for item in rows:
        dt = _finance_datetime_for_report(item)
        if dt is None:
            continue
        buckets.setdefault(dt.strftime("%Y-%m-%d"), []).append(item)

    start = datetime.fromtimestamp(date_from_ms / 1000, UZT).date()
    end = datetime.fromtimestamp(date_to_ms / 1000, UZT).date()
    result: list[dict[str, Any]] = []
    current = start
    while current <= end:
        day_rows = buckets.get(current.isoformat(), [])
        stats = _build_noorza_today_stats(day_rows)
        result.append({
            "date": datetime(current.year, current.month, current.day),
            "orders": int(stats.get("orders") or 0),
            "units": float(stats.get("units") or 0),
            "cancelled": int(stats.get("cancelled") or 0),
            "returns": float(stats.get("returns") or 0),
            "revenue": float(stats.get("revenue") or 0),
            "commission": float(stats.get("commission") or 0),
            "logistics": float(stats.get("logistics") or 0),
            "payout": float(stats.get("payout_total") or 0),
        })
        current += timedelta(days=1)
    return result


def _period_payload(
    key: str,
    rows: list[dict[str, Any]],
    costs: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    stats = _build_noorza_today_stats(rows)
    unit_rows = _build_unit_rows_from_finance(rows, costs)
    profit = _profit_summary_from_unit_rows(unit_rows, stats)
    return {
        "key": key,
        **stats,
        "profit": float(profit.get("profit") or 0) if float(profit.get("known_revenue") or 0) > 0 else None,
        "cost_coverage": float(profit.get("coverage") or 0),
    }


def _build_stock_report_rows(
    stock_rows: list[dict[str, Any]],
    sales_7: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    sold_qty: dict[str, float] = {}
    for item in sales_7:
        status = _finance_status(item)
        if _is_cancelled_status(status) or _is_returned_status(status):
            continue
        for key in _sale_match_keys(item):
            sold_qty[key] = sold_qty.get(key, 0.0) + _finance_qty(item)

    unique_keys_by_row = _unique_stock_match_keys(stock_rows)
    normalized: list[dict[str, Any]] = []
    lost: list[dict[str, Any]] = []
    for row, row_keys in zip(stock_rows, unique_keys_by_row):
        total = float(_num_any(row.get("total")))
        fbo = float(_num_any(row.get("fbo")))
        fbs = float(_num_any(row.get("fbs")))
        price = float(_num_any(row.get("price")))
        missing = float(_num_any(row.get("missing")))
        defected = float(_num_any(row.get("defected")))
        qty_7 = max([sold_qty.get(key, 0.0) for key in row_keys] or [0.0])
        avg_day = qty_7 / 7.0
        days_left = round(total / avg_day, 1) if avg_day > 0 and total > 0 else None
        if total <= 0:
            action_ru, action_uz = "Пополнить остаток или скрыть карточку", "Qoldiqni to‘ldirish yoki kartani yashirish"
        elif missing > 0 or defected > 0:
            action_ru, action_uz = "Сверить склад, акт и компенсацию Uzum", "Ombor, dalolatnoma va Uzum kompensatsiyasini tekshirish"
        elif days_left is not None and days_left <= SMART_LOW_STOCK_DAYS:
            action_ru, action_uz = "Срочно запланировать поставку", "Zudlik bilan yetkazib berishni rejalashtirish"
        elif total <= LOW_STOCK_THRESHOLD:
            action_ru, action_uz = "Запланировать пополнение", "Qoldiqni to‘ldirishni rejalashtirish"
        elif qty_7 <= 0 and total > 0:
            action_ru, action_uz = "Проверить цену, контент и продвижение", "Narx, kontent va reklamani tekshirish"
        else:
            action_ru, action_uz = "Контроль не требуется", "Nazorat talab qilinmaydi"
        item = {
            "product_id": row.get("product_id"),
            "sku_id": row.get("sku_id"),
            "barcode": row.get("barcode"),
            "title": row.get("product_title") or row.get("sku_full_title") or row.get("sku_title") or "Без названия",
            "sku": row.get("sku_full_title") or row.get("sku_title") or row.get("sku_id") or row.get("barcode"),
            "price": price,
            "purchase_price": row.get("purchase_price"),
            "ikpu": row.get("ikpu"),
            "paid_storage": bool(row.get("paid_storage")),
            "paid_storage_price_item": row.get("paid_storage_price_item"),
            "paid_storage_amount": row.get("paid_storage_amount"),
            "fbo": fbo,
            "fbs": fbs,
            "total": total,
            "sold_7": qty_7,
            "days_left": days_left,
            "missing": missing,
            "defected": defected,
            "status": status_display(row.get("status")) if row.get("status") else "",
            "archived": bool(row.get("archived")),
            "loss_all_time": bool(row.get("loss_all_time")),
            "loss_only": bool(row.get("loss_only")),
            "action_ru": action_ru,
            "action_uz": action_uz,
        }
        normalized.append(item)
        if missing > 0 or defected > 0:
            lost.append({
                **item,
                "estimated_loss": price * (missing + defected),
            })
    normalized.sort(key=lambda value: (float(value.get("total") or 0), str(value.get("title") or "")))
    lost.sort(key=lambda value: float(value.get("estimated_loss") or 0), reverse=True)
    return normalized, lost


def _build_premium_actions(
    stock: list[dict[str, Any]],
    products: list[dict[str, Any]],
    sales_stats: dict[str, Any],
) -> list[dict[str, Any]]:
    actions: list[dict[str, Any]] = []

    def add(
        priority: str,
        category_ru: str,
        category_uz: str,
        title: str,
        sku: str,
        problem_ru: str,
        problem_uz: str,
        recommendation_ru: str,
        recommendation_uz: str,
        amount: float = 0.0,
        source: str = "",
    ) -> None:
        actions.append({
            "priority": priority,
            "category_ru": category_ru,
            "category_uz": category_uz,
            "title": title,
            "sku": sku,
            "problem_ru": problem_ru,
            "problem_uz": problem_uz,
            "recommendation_ru": recommendation_ru,
            "recommendation_uz": recommendation_uz,
            "amount": amount,
            "source": source,
        })

    for item in stock:
        total = float(item.get("total") or 0)
        missing = float(item.get("missing") or 0)
        defected = float(item.get("defected") or 0)
        price = float(item.get("price") or 0)
        days_left = item.get("days_left")
        title = str(item.get("title") or "—")
        sku = str(item.get("sku") or "")
        if missing > 0 or defected > 0:
            all_time = bool(item.get("loss_all_time"))
            add(
                "critical", "Потери склада", "Ombor yo‘qotishlari", title, sku,
                f"За весь период потеряно {missing:.0f}, брак {defected:.0f} шт." if all_time else f"Потеряно {missing:.0f}, брак {defected:.0f} шт.",
                f"Barcha davrda yo‘qolgan {missing:.0f}, yaroqsiz {defected:.0f} dona." if all_time else f"Yo‘qolgan {missing:.0f}, yaroqsiz {defected:.0f} dona.",
                "Проверить акты, остатки и возможную компенсацию.",
                "Dalolatnoma, qoldiq va kompensatsiyani tekshiring.",
                price * (missing + defected), "Products API — накопительно" if all_time else "Остатки",
            )
        if item.get("loss_only") or item.get("archived"):
            continue
        if total <= 0:
            recommended = int(item.get("recommended_qty") or 0)
            add(
                "critical", "Нет в наличии", "Qoldiq yo‘q", title, sku,
                "Карточка активна без доступного остатка.", "Kartada mavjud qoldiq yo‘q.",
                (f"Поставить около {recommended} шт. или временно скрыть карточку." if recommended else "Проверить спрос и временно скрыть карточку."),
                (f"Taxminan {recommended} dona yetkazing yoki kartani vaqtincha yashiring." if recommended else "Talabni tekshiring va kartani vaqtincha yashiring."),
                float(item.get("risk_value") or 0), "Остатки + прогноз",
            )
        elif isinstance(days_left, (int, float)) and float(days_left) <= SMART_LOW_STOCK_DAYS:
            recommended = int(item.get("recommended_qty") or 0)
            reorder_date = item.get("reorder_date")
            reorder_text = reorder_date.strftime("%d.%m.%Y") if isinstance(reorder_date, datetime) else "сейчас"
            add(
                "critical", "Скоро закончится", "Tez tugaydi", title, sku,
                f"Остатка примерно на {float(days_left):.1f} дня.", f"Qoldiq taxminan {float(days_left):.1f} kunga yetadi.",
                f"Заказать {recommended} шт. до {reorder_text}." if recommended else "Срочно проверить план поставки.",
                f"{recommended} dona buyurtma qiling." if recommended else "Yetkazib berish rejasini tekshiring.",
                float(item.get("risk_value") or 0), "Прогноз поставки",
            )
        elif total <= float(item.get("low_stock_threshold") or LOW_STOCK_THRESHOLD):
            recommended = int(item.get("recommended_qty") or 0)
            add(
                "warning", "Низкий остаток", "Kam qoldiq", title, sku,
                f"Осталось {total:.0f} шт.", f"{total:.0f} dona qoldi.",
                f"Запланировать поставку {recommended} шт." if recommended else "Проверить необходимость пополнения.",
                f"{recommended} dona yetkazishni rejalashtiring." if recommended else "To‘ldirish zarurligini tekshiring.",
                float(item.get("risk_value") or 0), "Остатки + прогноз",
            )
        elif float(item.get("sold_7") or 0) <= 0:
            add(
                "warning", "Нет продаж", "Savdo yo‘q", title, sku,
                "Есть остаток, но не было продаж за 7 дней.", "Qoldiq bor, lekin 7 kunda savdo bo‘lmagan.",
                "Проверить цену, фото, описание и рекламу.", "Narx, rasm, tavsif va reklamani tekshiring.",
                price * total, "Продажи + остатки",
            )

    for item in products:
        title = str(item.get("title") or "—")
        sku = str(item.get("sku") or "")
        revenue = float(item.get("revenue") or 0)
        if item.get("cost_per_unit") is None:
            add(
                "warning", "Нет себестоимости", "Tannarx yo‘q", title, sku,
                "Прибыль по товару нельзя рассчитать.", "Tovar foydasini hisoblab bo‘lmaydi.",
                "Проверить purchasePrice в карточке или накладной Uzum.", "Uzum kartasi yoki yuk xatidagi purchasePrice ni tekshiring.",
                revenue, "Себестоимость Uzum",
            )
        else:
            profit = float(item.get("profit") or 0)
            margin = float(item.get("margin") or 0)
            if profit < 0:
                add(
                    "critical", "Убыточный товар", "Zararli tovar", title, sku,
                    f"Расчётный убыток {_format_money(abs(profit))}.", f"Hisobiy zarar {_format_money(abs(profit))}.",
                    "Пересмотреть цену, себестоимость и участие в акциях.", "Narx, tannarx va aksiyalarni qayta ko‘rib chiqing.",
                    abs(profit), "Юнит-экономика",
                )
            elif margin < LOW_MARGIN_THRESHOLD_PERCENT:
                add(
                    "warning", "Низкая маржа", "Past marja", title, sku,
                    f"Маржа только {margin:.1f}%.", f"Marja atigi {margin:.1f}%.",
                    "Проверить цену, комиссию и логистику.", "Narx, komissiya va logistikani tekshiring.",
                    revenue, "Юнит-экономика",
                )

    cancellations = int(sales_stats.get("cancelled") or 0)
    if cancellations:
        add(
            "warning", "Отмены", "Bekor qilish", "Все товары", "",
            f"За 30 дней отменено позиций: {cancellations}.", f"30 kunda bekor qilingan pozitsiyalar: {cancellations}.",
            "Открыть лист «Отмены и возвраты» и проверить причины.", "«Bekor va qaytarish» varag‘ini ochib sabablarni tekshiring.",
            float(sales_stats.get("cancelled_value") or 0), "Finance API",
        )
    priority_order = {"critical": 0, "warning": 1, "info": 2}
    for action in actions:
        action["action_key"] = _business_action_key(action)
    actions.sort(key=lambda value: (-float(value.get("amount") or 0), priority_order.get(str(value.get("priority")), 9)))
    return actions


async def _build_full_excel_report(
    client: UzumClient,
    telegram_id: int,
    shop_id: int,
    *,
    lang: str = "ru",
) -> Path:
    generated_at = datetime.now(UZT).replace(tzinfo=None)
    period_ranges = {
        "today": _today_range_ms(),
        "yesterday": _yesterday_range_ms(),
        "7d": _last_7_days_range_ms(),
        "30d": _days_range_ms(30),
    }
    current_30_from, current_30_to = period_ranges["30d"]
    shift_30 = 30 * 24 * 60 * 60 * 1000
    period_ranges["prev30d"] = (current_30_from - shift_30, current_30_to - shift_30)

    finance_by_period: dict[str, list[dict[str, Any]]] = {}
    source_notes: list[str] = []
    for key, (date_from, date_to) in period_ranges.items():
        rows, _, source_info = await _load_finance_range_flexible(
            client,
            shop_id,
            date_from,
            date_to,
        )
        finance_by_period[key] = rows
        if "Достигнут защитный лимит" in source_info:
            source_notes.append(f"{key}: {source_info}")
        await asyncio.sleep(0.1)

    stock_raw = await load_sku_rows(client, shop_id, max_pages=50)
    await sync_uzum_sku_financials(
        client,
        telegram_id,
        shop_id,
        stock_rows=stock_raw,
    )
    costs = get_unit_cost_map(telegram_id, shop_id)
    period_payloads = [
        _period_payload(key, finance_by_period.get(key, []), costs)
        for key in ("today", "yesterday", "7d", "30d", "prev30d")
    ]
    rows_30 = finance_by_period.get("30d", [])
    stats_30 = _build_noorza_today_stats(rows_30)
    finance_settings = ensure_finance_settings(telegram_id, shop_id)
    product_rows = _build_unit_rows_from_finance(
        rows_30,
        costs,
        tax_percent=float(finance_settings.get("tax_percent") or 0),
    )
    profit_30 = _profit_summary_from_unit_rows(product_rows, stats_30)
    uzum_expenses = await load_uzum_expense_summary(
        client,
        shop_id,
        current_30_from,
        current_30_to,
    )
    business_profit = calculate_business_profit(
        profit_30,
        stats_30,
        finance_settings,
        days=30,
        uzum_expenses=uzum_expenses,
    )

    stock_rows, _ = _build_stock_report_rows(stock_raw, finance_by_period.get("7d", []))
    product_settings = ensure_product_settings(telegram_id)
    replenishment = build_replenishment_plan(stock_raw, rows_30, product_settings)
    replenishment_by_key: dict[str, dict[str, Any]] = {}
    for plan_item in replenishment:
        for key in _stock_match_keys(plan_item.get("row") or {}):
            replenishment_by_key[key] = plan_item
    for stock_item in stock_rows:
        stock_item["low_stock_threshold"] = int(product_settings.get("low_stock_threshold") or 0)
        plan_item = next(
            (
                replenishment_by_key[key]
                for key in _stock_match_keys(stock_item)
                if key in replenishment_by_key
            ),
            None,
        )
        if plan_item:
            stock_item.update({
                "risk_value": float(plan_item.get("risk_value") or 0),
                "recommended_qty": int(plan_item.get("recommended_qty") or 0),
                "reorder_date": plan_item.get("reorder_date"),
                "avg_daily": float(plan_item.get("avg_daily") or 0),
            })
    loss_raw, unavailable_loss_filters = await _load_all_time_loss_rows(client, shop_id)
    _, lost_rows = _build_stock_report_rows(loss_raw, finance_by_period.get("7d", []))
    stock_rows_for_actions = [
        {**row, "missing": 0, "defected": 0}
        for row in stock_rows
    ]
    stats_30["cancelled_value"] = sum(
        _finance_gross_revenue(item)
        for item in rows_30
        if _is_cancelled_status(_finance_status(item))
    )
    actions = _build_premium_actions(stock_rows_for_actions + lost_rows, product_rows, stats_30)

    notes: list[str] = []
    if source_notes:
        notes.extend(source_notes)
    notes.append(_t(
        lang,
        "Uzum API не передаёт даты отдельных потерь; лист «Потери за весь период» показывает накопительные quantityMissing и quantityDefected по SKU.",
        "Uzum API alohida yo‘qotish sanalarini bermaydi; «Barcha davr yo‘qotish» varag‘ida SKU bo‘yicha jamlangan quantityMissing va quantityDefected ko‘rsatiladi.",
    ))
    if unavailable_loss_filters:
        notes.append(_t(
            lang,
            "Часть фильтров товаров Uzum была временно недоступна: " + ", ".join(unavailable_loss_filters) + ". Список потерь может быть неполным.",
            "Uzum tovar filtrlarining bir qismi vaqtincha ishlamadi: " + ", ".join(unavailable_loss_filters) + ". Yo‘qotishlar ro‘yxati to‘liq bo‘lmasligi mumkin.",
        ))

    payload = {
        "shop_id": shop_id,
        "generated_at": generated_at,
        "cost_coverage": float(profit_30.get("coverage") or 0),
        "business_profit": business_profit,
        "finance_settings": finance_settings,
        "periods": period_payloads,
        "sales": [_normalize_finance_row(item) for item in rows_30],
        "daily": _daily_report_rows(rows_30, current_30_from, current_30_to),
        "products": product_rows,
        "stock": stock_rows,
        "actions": actions,
        "lost": lost_rows,
        "notes": notes,
    }
    timestamp = generated_at.strftime("%Y-%m-%d_%H-%M")
    filename = Path(tempfile.gettempdir()) / f"uzum_premium_report_{shop_id}_{timestamp}.xlsx"
    return await asyncio.to_thread(build_premium_workbook, payload, filename, lang=lang)


@dp.message(Command("report_excel"))
@dp.message(Command("report"))
@dp.message(Command("full_report"))
async def report_excel(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)

    await message.answer(
        (
            "⌛ Pullik boshqaruv Excel hisoboti tayyorlanmoqda...\n"
            "Barcha savdolar, foyda, qoldiq, yo‘qotishlar va tayyor amallar yig‘ilmoqda."
            if lang == "uz"
            else "⌛ Готовлю управленческий Excel-отчёт...\n"
            "Собираю все продажи, прибыль, остатки, потери и готовый список действий."
        ),
        reply_markup=menu_for_message(message),
    )
    filename: Path | None = None
    try:
        filename = await _build_full_excel_report(
            client,
            telegram_id,
            shop_id,
            lang=lang,
        )
        await message.answer_document(
            FSInputFile(filename),
            caption=(
                "✅ <b>Boshqaruv Excel hisoboti tayyor</b>\n\n"
                "Ichida: xulosa va grafiklar, barcha savdolar, kunlik dinamika, "
                "tovarlar foydasi, bekor qilish va qaytarishlar, qoldiq prognozi, "
                "yo‘qotishlar hamda tayyor amallar ro‘yxati."
                if lang == "uz"
                else "✅ <b>Управленческий Excel-отчёт готов</b>\n\n"
                "Внутри: сводка и графики, все продажи, динамика по дням, "
                "прибыль по товарам, отмены и возвраты, прогноз остатков, "
                "потери и готовый список действий."
            ),
            reply_markup=menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)
    finally:
        if filename is not None:
            try:
                filename.unlink(missing_ok=True)
            except OSError:
                pass


# --- Четырёхстраничный управленческий PDF-отчёт ---
PDF_REPORT_PERIODS: dict[str, tuple[str, str, int]] = {
    "today": ("Сегодня", "Bugun", 1),
    "7d": ("7 дней", "7 kun", 7),
    "30d": ("30 дней", "30 kun", 30),
}
_PDF_REPORT_IN_PROGRESS: set[tuple[int, int]] = set()


def _pdf_report_period_markup(lang: str) -> InlineKeyboardMarkup:
    uz = normalize_lang(lang) == "uz"
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="Bugun" if uz else "Сегодня",
                    callback_data="pdfreport:today",
                ),
                InlineKeyboardButton(
                    text="7 kun" if uz else "7 дней",
                    callback_data="pdfreport:7d",
                ),
                InlineKeyboardButton(
                    text="30 kun" if uz else "30 дней",
                    callback_data="pdfreport:30d",
                ),
            ]
        ]
    )


def _pdf_report_range(period_key: str) -> tuple[int, int, int]:
    if period_key == "today":
        date_from, date_to = _today_range_ms()
        return date_from, date_to, 1
    if period_key == "7d":
        date_from, date_to = _last_7_days_range_ms()
        return date_from, date_to, 7
    if period_key == "30d":
        date_from, date_to = _days_range_ms(30)
        return date_from, date_to, 30
    raise ValueError("Неизвестный период PDF-отчёта")


def _pdf_report_period_label(date_from_ms: int, date_to_ms: int) -> str:
    date_from = datetime.fromtimestamp(date_from_ms / 1000, UZT)
    date_to = datetime.fromtimestamp(date_to_ms / 1000, UZT)
    if date_from.date() == date_to.date():
        return date_from.strftime("%d.%m.%Y")
    return f"{date_from:%d.%m.%Y} - {date_to:%d.%m.%Y}"


def _pdf_problem_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return _problem_finance_rows(rows)


def _pdf_management_actions(
    stock_rows: list[dict[str, Any]],
    stats: dict[str, Any],
    profit: dict[str, Any],
    defect_events: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    actions: list[dict[str, Any]] = []
    zero_count = sum(
        1
        for row in stock_rows
        if not row.get("loss_only") and float(row.get("total") or 0) <= 0
    )
    low_count = sum(
        1
        for row in stock_rows
        if float(row.get("total") or 0) > 0
        and (
            float(row.get("total") or 0)
            <= float(row.get("low_stock_threshold") or 5)
            or (
                row.get("days_left") is not None
                and float(row.get("days_left") or 0) <= 7
            )
        )
    )
    if zero_count or low_count:
        actions.append({
            "priority": "critical" if zero_count else "warning",
            "title_ru": "Устранить дефицит",
            "title_uz": "Qoldiq xavfini bartaraf etish",
            "body_ru": f"Нет в наличии: {zero_count} SKU; низкий остаток: {low_count}. Сначала подготовьте поставку по этим позициям.",
            "body_uz": f"Qoldiq yo‘q: {zero_count} SKU; kam qoldiq: {low_count}. Avval shu pozitsiyalar yetkazib berishini tayyorlang.",
        })

    cancellations = int(stats.get("cancelled") or 0)
    returns = float(stats.get("returns") or 0)
    if cancellations or returns:
        actions.append({
            "priority": "warning",
            "title_ru": "Разобрать отмены и возвраты",
            "title_uz": "Bekor va qaytarishni tahlil qilish",
            "body_ru": f"Отменено строк: {cancellations}; возвращено: {returns:g} шт. Проверьте причины и проблемные SKU на странице 4.",
            "body_uz": f"Bekor qatorlar: {cancellations}; qaytarilgan: {returns:g} dona. Sabab va muammoli SKUlarni 4-sahifada tekshiring.",
        })

    defected_delta = sum(int(item.get("defected_delta") or 0) for item in defect_events)
    missing_delta = sum(int(item.get("missing_delta") or 0) for item in defect_events)
    if defected_delta or missing_delta:
        actions.append({
            "priority": "critical" if defected_delta else "warning",
            "title_ru": "Проверить потери и брак FBO",
            "title_uz": "FBO yo‘qotish va brakni tekshirish",
            "body_ru": f"За период зафиксировано: брак +{defected_delta}, потери +{missing_delta}. Сверьте акты и компенсации Uzum.",
            "body_uz": f"Davrda qayd etildi: brak +{defected_delta}, yo‘qotish +{missing_delta}. Uzum dalolatnoma va kompensatsiyasini tekshiring.",
        })

    missing_costs = int(profit.get("missing_count") or 0)
    if missing_costs:
        actions.append({
            "priority": "warning",
            "title_ru": "Проверить себестоимость в Uzum",
            "title_uz": "Uzum tannarxini tekshirish",
            "body_ru": f"Uzum не передал purchasePrice для {missing_costs} SKU. Прибыль показана только по известной части продаж.",
            "body_uz": f"Uzum {missing_costs} SKU uchun purchasePrice bermadi. Foyda faqat ma’lum savdo qismi bo‘yicha ko‘rsatiladi.",
        })

    no_sales_count = sum(
        1
        for row in stock_rows
        if float(row.get("total") or 0) > 0 and float(row.get("sold_7") or 0) <= 0
    )
    if no_sales_count and len(actions) < 4:
        actions.append({
            "priority": "info",
            "title_ru": "Проверить товары без продаж",
            "title_uz": "Savdosiz tovarlarni tekshirish",
            "body_ru": f"У {no_sales_count} SKU есть остаток, но нет продаж за 7 дней. Проверьте цену, контент и продвижение.",
            "body_uz": f"{no_sales_count} SKUda qoldiq bor, lekin 7 kunda savdo yo‘q. Narx, kontent va reklamani tekshiring.",
        })
    return actions[:4]


async def _collect_seller_pdf_payload(
    client: UzumClient,
    telegram_id: int,
    shop_id: int,
    period_key: str,
) -> dict[str, Any]:
    date_from, date_to, period_days = _pdf_report_range(period_key)
    shift_ms = period_days * 24 * 60 * 60 * 1000
    previous_from, previous_to = date_from - shift_ms, date_to - shift_ms

    rows, _, source_info = await _load_finance_range_flexible(
        client,
        shop_id,
        date_from,
        date_to,
    )
    await asyncio.sleep(0.1)
    comparison_available = True
    try:
        previous_rows, _, _ = await _load_finance_range_flexible(
            client,
            shop_id,
            previous_from,
            previous_to,
        )
    except Exception:
        comparison_available = False
        previous_rows = []
        logging.exception(
            "PDF report: previous period unavailable user=%s shop=%s period=%s",
            telegram_id,
            shop_id,
            period_key,
        )

    if period_key == "7d":
        sales_7 = rows
    else:
        await asyncio.sleep(0.1)
        try:
            week_from, week_to = _last_7_days_range_ms()
            sales_7, _, _ = await _load_finance_range_flexible(
                client,
                shop_id,
                week_from,
                week_to,
            )
        except Exception:
            sales_7 = []
            logging.exception(
                "PDF report: seven-day stock velocity unavailable user=%s shop=%s",
                telegram_id,
                shop_id,
            )

    stats = _build_noorza_today_stats(rows)
    previous_stats = _build_noorza_today_stats(previous_rows)
    product_settings = ensure_product_settings(telegram_id)

    stock_rows: list[dict[str, Any]] = []
    stock_data_available = True
    cost_data_stale = False
    try:
        stock_raw = await load_sku_rows(client, shop_id, max_pages=50)
        cost_status = await sync_uzum_sku_financials(
            client,
            telegram_id,
            shop_id,
            stock_rows=stock_raw,
        )
        cost_data_stale = bool(cost_status.get("stale"))
        stock_rows, _ = _build_stock_report_rows(stock_raw, sales_7)
        threshold = int(product_settings.get("low_stock_threshold") or 5)
        for row in stock_rows:
            row["low_stock_threshold"] = threshold
    except Exception:
        stock_data_available = False
        cost_data_stale = True
        logging.exception(
            "PDF report: stock section unavailable user=%s shop=%s",
            telegram_id,
            shop_id,
        )

    costs = get_unit_cost_map(telegram_id, shop_id)
    finance_settings = ensure_finance_settings(telegram_id, shop_id)
    products = _build_unit_rows_from_finance(
        rows,
        costs,
        tax_percent=float(finance_settings.get("tax_percent") or 0),
    )
    profit = _profit_summary_from_unit_rows(products, stats)
    uzum_expenses = await load_uzum_expense_summary(
        client,
        shop_id,
        date_from,
        date_to,
    )
    business_profit = calculate_business_profit(
        profit,
        stats,
        finance_settings,
        days=period_days,
        uzum_expenses=uzum_expenses,
    )
    if cost_data_stale:
        business_profit["complete"] = False

    cumulative_defects: list[dict[str, Any]] = []
    loss_data_available = True
    unavailable_loss_filters: list[str] = []
    try:
        loss_raw, unavailable_loss_filters = await _load_all_time_loss_rows(
            client,
            shop_id,
        )
        _, cumulative_losses = _build_stock_report_rows(loss_raw, sales_7)
        cumulative_defects = [
            row for row in cumulative_losses if float(row.get("defected") or 0) > 0
        ]
    except Exception:
        loss_data_available = False
        logging.exception(
            "PDF report: cumulative loss section unavailable user=%s shop=%s",
            telegram_id,
            shop_id,
        )

    defect_events = list_loss_defect_events(
        telegram_id,
        shop_id,
        date_from,
        date_to,
    )
    daily = _daily_report_rows(rows, date_from, date_to)
    for row in daily:
        date_value = row.get("date")
        row["label"] = (
            date_value.strftime("%d.%m")
            if isinstance(date_value, datetime)
            else str(date_value or "")
        )

    problems = _pdf_problem_rows(rows)
    data_notes: list[str] = []
    if "Достигнут защитный лимит" in source_info:
        data_notes.append("finance_rows_truncated")
    if unavailable_loss_filters:
        data_notes.append("loss_filters_unavailable")
    if not comparison_available:
        data_notes.append("comparison_unavailable")
    if not stock_data_available:
        data_notes.append("stock_unavailable")
    if not loss_data_available:
        data_notes.append("loss_data_unavailable")
    if cost_data_stale:
        data_notes.append("cost_data_stale")
    if not bool(business_profit.get("uzum_expenses_available")):
        data_notes.append("expenses_unavailable")

    return {
        "shop_id": shop_id,
        "generated_at": datetime.now(UZT).replace(tzinfo=None),
        "period_key": period_key,
        "period_days": period_days,
        "period_label": _pdf_report_period_label(date_from, date_to),
        "stats": stats,
        "previous_stats": previous_stats,
        "comparison_available": comparison_available,
        "profit": profit,
        "business_profit": business_profit,
        "finance_settings": finance_settings,
        "daily": daily,
        "products": products,
        "stock": stock_rows,
        "stock_data_available": stock_data_available,
        "problems": problems,
        "defect_events": defect_events,
        "cumulative_defects": cumulative_defects,
        "loss_data_available": loss_data_available,
        "actions": _pdf_management_actions(
            stock_rows,
            stats,
            profit,
            defect_events,
        ),
        "data_notes": data_notes,
    }


async def _build_seller_pdf_for_user(
    client: UzumClient,
    telegram_id: int,
    shop_id: int,
    period_key: str,
    *,
    lang: str,
) -> tuple[Path, dict[str, Any]]:
    payload = await _collect_seller_pdf_payload(
        client,
        telegram_id,
        shop_id,
        period_key,
    )
    regular_font, bold_font = _fbo_pdf_font_paths()
    timestamp = datetime.now(UZT).strftime("%Y-%m-%d_%H-%M-%S")
    filename = (
        Path(tempfile.gettempdir())
        / f"sellerpro_pdf_{shop_id}_{period_key}_{timestamp}.pdf"
    )
    await asyncio.to_thread(
        build_seller_pdf_report,
        payload,
        filename,
        lang=lang,
        regular_font_path=regular_font,
        bold_font_path=bold_font,
    )
    return filename, payload


@dp.message(Command("pdf_report"))
@dp.message(Command("management_pdf"))
async def seller_pdf_report_menu(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not await require_premium_subscription(message, telegram_id):
        return
    req = await require_connection(message)
    if req is None:
        return
    lang = get_user_language(telegram_id)
    await message.answer(
        (
            "📄 <b>Boshqaruv PDF hisoboti</b>\n\n"
            "Davrni tanlang. Hisobotda savdo, foyda, qoldiq, tavsiyalar, "
            "bekor qilish, qaytarish va brak bo‘ladi."
            if lang == "uz"
            else "📄 <b>Управленческий PDF-отчёт</b>\n\n"
            "Выберите период. В отчёте будут продажи, прибыль, остатки, "
            "рекомендации, отмены, возвраты и брак."
        ),
        reply_markup=_pdf_report_period_markup(lang),
    )


@dp.callback_query(F.data.startswith("pdfreport:"))
async def seller_pdf_report_callback(callback: CallbackQuery) -> None:
    telegram_id = int(callback.from_user.id)
    lang = get_user_language(telegram_id)
    if not has_paid_subscription(telegram_id):
        if subscription_access_level(telegram_id) == "trial":
            await send_trial_premium_locked(callback, telegram_id)
        else:
            await callback.answer(
                "Obuna tugagan" if lang == "uz" else "Подписка закончилась",
                show_alert=True,
            )
        return

    period_key = str(callback.data or "").partition(":")[2]
    if period_key not in PDF_REPORT_PERIODS:
        await callback.answer(
            "Noto‘g‘ri davr" if lang == "uz" else "Неизвестный период",
            show_alert=True,
        )
        return
    client = get_uzum_for_user(telegram_id)
    shop_id = db.get_default_shop_id(telegram_id)
    if client is None or shop_id is None:
        await callback.answer(
            "Avval do‘konni ulang" if lang == "uz" else "Сначала подключите магазин",
            show_alert=True,
        )
        return
    shop_id = int(shop_id)
    progress_key = (telegram_id, shop_id)
    if progress_key in _PDF_REPORT_IN_PROGRESS:
        await callback.answer(
            "Hisobot allaqachon tayyorlanmoqda"
            if lang == "uz"
            else "Отчёт уже формируется",
            show_alert=True,
        )
        return
    if callback.message is None:
        await callback.answer()
        return

    await callback.answer(
        "Tayyorlashni boshladim" if lang == "uz" else "Начал подготовку"
    )
    await callback.message.answer(
        (
            "⌛ Ma’lumotlarni yig‘ib, PDF tayyorlayapman. Bu biroz vaqt olishi mumkin..."
            if lang == "uz"
            else "⌛ Собираю данные и формирую PDF. Это может занять немного времени..."
        )
    )
    _PDF_REPORT_IN_PROGRESS.add(progress_key)
    filename: Path | None = None
    try:
        filename, payload = await _build_seller_pdf_for_user(
            client,
            telegram_id,
            shop_id,
            period_key,
            lang=lang,
        )
        stats = dict(payload.get("stats") or {})
        profit = dict(payload.get("profit") or {})
        coverage = float(profit.get("coverage") or 0)
        if float(stats.get("revenue") or 0) <= 0:
            quality_note = (
                "Davrda savdo topilmadi."
                if lang == "uz"
                else "За период продажи не найдены."
            )
        elif coverage >= 0.999:
            quality_note = (
                "✅ Tannarx qamrovi: 100%."
                if lang == "uz"
                else "✅ Покрытие себестоимостью: 100%."
            )
        else:
            quality_note = (
                f"⚠️ Tannarx qamrovi: {coverage * 100:.1f}%. Foyda faqat ma’lum qism bo‘yicha."
                if lang == "uz"
                else f"⚠️ Покрытие себестоимостью: {coverage * 100:.1f}%. Прибыль показана только по известной части."
            )
        data_notes = set(payload.get("data_notes") or [])
        if "finance_rows_truncated" in data_notes:
            quality_note += (
                "\n⚠️ Finance qatorlari texnik limitga yetdi; davrni qisqartiring."
                if lang == "uz"
                else "\n⚠️ Достигнут технический лимит строк Finance; выберите меньший период."
            )
        if "loss_filters_unavailable" in data_notes:
            quality_note += (
                "\n⚠️ Uzumning ayrim brak filtrlari vaqtincha ishlamadi."
                if lang == "uz"
                else "\n⚠️ Часть фильтров брака Uzum была временно недоступна."
            )
        if "stock_unavailable" in data_notes:
            quality_note += (
                "\n⚠️ Qoldiq bo‘limi Uzum API xatosi sababli mavjud emas."
                if lang == "uz"
                else "\n⚠️ Раздел остатков недоступен из-за ответа Uzum API."
            )
        if "loss_data_unavailable" in data_notes:
            quality_note += (
                "\n⚠️ Jamlangan brak ma’lumoti vaqtincha mavjud emas."
                if lang == "uz"
                else "\n⚠️ Накопительные данные брака временно недоступны."
            )
        if "comparison_unavailable" in data_notes:
            quality_note += (
                "\n⚠️ Oldingi davr bilan taqqoslash vaqtincha mavjud emas."
                if lang == "uz"
                else "\n⚠️ Сравнение с предыдущим периодом временно недоступно."
            )
        public_name = f"SellerPro_{shop_id}_{period_key}.pdf"
        await callback.message.answer_document(
            FSInputFile(str(filename), filename=public_name),
            caption=(
                "✅ <b>Boshqaruv PDF hisoboti tayyor</b>\n"
                f"Davr: {escape(str(payload.get('period_label') or '-'))}\n"
                f"{quality_note}"
                if lang == "uz"
                else "✅ <b>Управленческий PDF-отчёт готов</b>\n"
                f"Период: {escape(str(payload.get('period_label') or '-'))}\n"
                f"{quality_note}"
            ),
            reply_markup=report_menu_for_user(telegram_id),
        )
    except Exception as error:
        logging.exception(
            "PDF report failed user=%s shop=%s period=%s",
            telegram_id,
            shop_id,
            period_key,
        )
        raw = escape(str(error))
        await callback.message.answer(
            (
                "⚠️ <b>PDF hisobotni yaratib bo‘lmadi</b>\n"
                f"<code>{raw[:900]}</code>"
                if lang == "uz"
                else "⚠️ <b>Не удалось сформировать PDF-отчёт</b>\n"
                f"<code>{raw[:900]}</code>"
            ),
            reply_markup=report_menu_for_user(telegram_id),
        )
    finally:
        _PDF_REPORT_IN_PROGRESS.discard(progress_key)
        if filename is not None:
            try:
                filename.unlink(missing_ok=True)
            except OSError:
                pass


# --- Уведомления о новых заказах ---
# Логика простая и безопасная:
# 1) при первом запуске бот запоминает текущие CREATED-заказы и не спамит ими;
# 2) дальше каждые ORDER_CHECK_INTERVAL_SECONDS секунд проверяет новые CREATED-заказы;
# 3) если появился новый заказ, пишет продавцу в Telegram.
_seen_order_keys_by_scope: dict[tuple[int, int], set[str]] = {}
_orders_watch_initialized_scopes: set[tuple[int, int]] = set()


def order_key(order: Any) -> str:
    """Делаем стабильный ключ заказа из ID. Если ID в ответе API спрятан, берём hash JSON."""
    if isinstance(order, dict):
        for key in (
            "id",
            "orderId",
            "order_id",
            "shipmentId",
            "shipment_id",
            "postingNumber",
            "posting_number",
            "number",
            "barcode",
        ):
            value = order.get(key)
            if value not in (None, ""):
                return f"{key}:{value}"

        # Иногда ID лежит внутри вложенного объекта.
        for value in order.values():
            if isinstance(value, dict):
                nested = order_key(value)
                if nested:
                    return nested

    raw = json.dumps(order, ensure_ascii=False, sort_keys=True, default=str)
    return "hash:" + hashlib.sha1(raw.encode("utf-8")).hexdigest()


def connected_users_for_order_watch(access_feature: str = "premium") -> list[dict[str, Any]]:
    """Return one watcher row for every connected shop available to a user.

    ``users.default_shop_id`` is only the shop selected for interactive menu
    commands.  Background notifications must not silently inherit that UI
    choice: a seller who connected several shops expects every one of them to
    be monitored.  The LEFT JOIN keeps older installations working when a
    connection exists but the ``shops`` table has not been populated yet.
    """
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT
                u.telegram_id,
                COALESCE(s.shop_id, u.default_shop_id) AS shop_id,
                COALESCE(s.shop_id, u.default_shop_id) AS default_shop_id,
                u.default_shop_id AS selected_shop_id,
                COALESCE(s.title, '') AS shop_title,
                u.uzum_token_encrypted
            FROM users AS u
            LEFT JOIN shops AS s
              ON s.telegram_id = u.telegram_id
            WHERE u.uzum_token_encrypted IS NOT NULL
              AND COALESCE(s.shop_id, u.default_shop_id) IS NOT NULL
            ORDER BY
                u.telegram_id,
                CASE WHEN s.shop_id = u.default_shop_id THEN 0 ELSE 1 END,
                COALESCE(s.shop_id, u.default_shop_id)
            """
        ).fetchall()
    allowed_by_user: dict[int, bool] = {}
    result: list[dict[str, Any]] = []
    seen_scopes: set[tuple[int, int]] = set()
    for raw_row in rows:
        row = dict(raw_row)
        telegram_id = int(row["telegram_id"])
        shop_id = int(row["shop_id"])
        scope = (telegram_id, shop_id)
        if scope in seen_scopes:
            continue
        if telegram_id not in allowed_by_user:
            allowed_by_user[telegram_id] = feature_access_allowed(
                telegram_id,
                access_feature,
            )
        if not allowed_by_user[telegram_id]:
            continue
        seen_scopes.add(scope)
        result.append(row)
    return result


def connected_shop_ids_for_user(telegram_id: int) -> list[int]:
    """Return every stored shop id, with a legacy default-shop fallback."""
    shop_ids = {
        int(row["shop_id"])
        for row in db.list_shops(int(telegram_id))
        if row["shop_id"] is not None
    }
    default_shop_id = db.get_default_shop_id(int(telegram_id))
    if not shop_ids and default_shop_id is not None:
        shop_ids.add(int(default_shop_id))
    return sorted(shop_ids)


def connected_shop_titles_for_user(telegram_id: int) -> dict[int, str]:
    """Return safe shop labels for multi-shop digest sections."""
    result: dict[int, str] = {}
    for row in db.list_shops(int(telegram_id)):
        shop_id = int(row["shop_id"])
        title = str(row["title"] or "").strip()
        if title and title != "—":
            result[shop_id] = title
    return result


async def check_new_orders_once() -> None:
    users = connected_users_for_order_watch()
    for row in users:
        telegram_id = int(row["telegram_id"])
        shop_id = int(row["default_shop_id"])
        encrypted_token = row["uzum_token_encrypted"]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            data = await client.get_fbs_orders(shop_id, status="CREATED", page=0, size=20)
            items = extract_items(data)
        except Exception:
            logging.exception("Order watcher: failed to check orders for %s", telegram_id)
            continue

        keys_now = [order_key(item) for item in items]
        scope = (telegram_id, shop_id)
        known = _seen_order_keys_by_scope.setdefault(scope, set())

        # Первый проход: просто запоминаем текущие заказы, чтобы не прислать старые как новые.
        if scope not in _orders_watch_initialized_scopes:
            known.update(keys_now)
            _orders_watch_initialized_scopes.add(scope)
            logging.info(
                "Order watcher initialized for user=%s shop=%s orders=%s",
                telegram_id,
                shop_id,
                len(keys_now),
            )
            continue

        new_items = [item for item, key in zip(items, keys_now) if key not in known]
        known.update(keys_now)

        # Чтобы память не росла бесконечно.
        if len(known) > 1000:
            _seen_order_keys_by_scope[scope] = set(keys_now)

        if not new_items:
            continue

        lines = [format_order_line(item) for item in new_items[:5]]
        more = "" if len(new_items) <= 5 else f"\n\nЕщё новых заказов: {len(new_items) - 5}"
        text = (
            f"🔔 <b>Новый заказ CREATED</b>\n"
            f"Магазин: <code>{shop_id}</code>\n"
            f"Новых заказов: <b>{len(new_items)}</b>\n\n"
            + "\n\n".join(lines)
            + more
            + "\n\nОткрыть список: <code>/orders</code>"
        )

        try:
            await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
        except Exception:
            logging.exception("Order watcher: failed to send notification to %s", telegram_id)


async def order_watch_loop() -> None:
    await asyncio.sleep(10)
    logging.info(
        "Order watcher started. Interval: %s seconds. Enabled: %s",
        ORDER_CHECK_INTERVAL_SECONDS,
        NEW_ORDER_NOTIFICATIONS,
    )
    if not NEW_ORDER_NOTIFICATIONS:
        logging.info("Order watcher disabled globally; no FBS order API requests will be made")
        return
    while True:
        try:
            await check_new_orders_once()
        except Exception:
            logging.exception("Order watcher loop error")
        await asyncio.sleep(max(60, ORDER_CHECK_INTERVAL_SECONDS))


@dp.message(Command("notify_status"))
async def notify_status(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    enabled = product_setting_enabled(telegram_id, "notify_orders")
    shop_ids = connected_shop_ids_for_user(telegram_id)
    initialized = bool(shop_ids) and all(
        (telegram_id, watched_shop_id) in _orders_watch_initialized_scopes
        for watched_shop_id in shop_ids
    )
    await message.answer(
        "🔔 <b>Уведомления о новых заказах</b>\n\n"
        f"Статус: {'✅ включены' if enabled else '❌ выключены'}\n"
        f"Магазинов под наблюдением: <b>{len(shop_ids) or 1}</b>\n"
        f"Активный в меню: <code>{shop_id}</code>\n"
        f"Проверка каждые: <b>{max(60, ORDER_CHECK_INTERVAL_SECONDS)}</b> сек.\n"
        f"Состояние: {'заказы уже запомнены' if initialized else 'инициализация при следующей проверке'}\n\n"
        "Бот уведомит, когда появится новый заказ со статусом <code>CREATED</code>.",
        reply_markup=menu_for_message(message),
    )


# --- Уведомления о низких остатках ---
# Логика:
# 1) при первом запуске бот запоминает текущие SKU с остатком ниже порога и не спамит ими;
# 2) дальше проверяет остатки каждые LOW_STOCK_CHECK_INTERVAL_SECONDS секунд;
# 3) если SKU впервые стал ниже/равен порогу, бот присылает уведомление.
_seen_low_stock_keys_by_user: dict[int, set[str]] = {}
_low_stock_watch_initialized: set[int] = set()


def stock_row_key(row: dict[str, Any]) -> str:
    for key in ("sku_id", "barcode", "seller_item_code", "product_id"):
        value = row.get(key)
        if value not in (None, ""):
            return f"{key}:{value}"
    raw = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
    return "hash:" + hashlib.sha1(raw.encode("utf-8")).hexdigest()


async def check_low_stock_once() -> None:
    users = connected_users_for_order_watch()
    threshold = max(0, LOW_STOCK_THRESHOLD)

    for row in users:
        telegram_id = int(row["telegram_id"])
        shop_id = int(row["default_shop_id"])
        encrypted_token = row["uzum_token_encrypted"]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows = await load_sku_rows(client, shop_id, max_pages=50)
        except Exception:
            logging.exception("Low stock watcher: failed to check stocks for %s", telegram_id)
            continue

        low_rows = [
            r
            for r in rows
            if r.get("total") is not None and int(r.get("total") or 0) <= threshold
        ]
        low_keys_now = [stock_row_key(r) for r in low_rows]
        known = _seen_low_stock_keys_by_user.setdefault(telegram_id, set())

        # Первый проход: запоминаем текущие низкие остатки, чтобы не прислать старые как новые.
        if telegram_id not in _low_stock_watch_initialized:
            known.update(low_keys_now)
            _low_stock_watch_initialized.add(telegram_id)
            logging.info(
                "Low stock watcher initialized for user=%s shop=%s low_skus=%s threshold=%s",
                telegram_id,
                shop_id,
                len(low_keys_now),
                threshold,
            )
            continue

        new_low_rows = [r for r, key in zip(low_rows, low_keys_now) if key not in known]
        known.update(low_keys_now)

        # Если товар восстановился выше порога, удаляем его из known.
        # Тогда при повторном падении ниже порога бот снова уведомит.
        _seen_low_stock_keys_by_user[telegram_id] = set(low_keys_now)

        if not new_low_rows:
            continue

        lines = [format_sku_stock_line(item, mode="all") for item in new_low_rows[:10]]
        more = "" if len(new_low_rows) <= 10 else f"\n\nЕщё SKU с низким остатком: {len(new_low_rows) - 10}"
        text = (
            f"📉 <b>Товар заканчивается</b>\n"
            f"Магазин: <code>{shop_id}</code>\n"
            f"Порог: ≤ <b>{threshold}</b> шт.\n"
            f"Новых позиций с низким остатком: <b>{len(new_low_rows)}</b>\n\n"
            + "\n\n".join(lines)
            + more
            + f"\n\nПоказать все низкие остатки: <code>/lowstock {threshold}</code>"
        )

        try:
            await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
        except Exception:
            logging.exception("Low stock watcher: failed to send notification to %s", telegram_id)


async def low_stock_watch_loop() -> None:
    await asyncio.sleep(20)
    logging.info(
        "Low stock watcher started. Interval: %s seconds. Threshold: %s. Enabled: %s",
        LOW_STOCK_CHECK_INTERVAL_SECONDS,
        LOW_STOCK_THRESHOLD,
        LOW_STOCK_NOTIFICATIONS,
    )
    while True:
        try:
            await check_low_stock_once()
        except Exception:
            logging.exception("Low stock watcher loop error")
        await asyncio.sleep(max(300, LOW_STOCK_CHECK_INTERVAL_SECONDS))


@dp.message(Command("lowstock_notify_status"))
async def lowstock_notify_status(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    settings = ensure_product_settings(telegram_id)
    enabled = bool(int(settings.get("notify_low_stock") or 0))
    threshold = int(settings.get("low_stock_threshold") or 0)
    shop_ids = connected_shop_ids_for_user(telegram_id)
    initialized = bool(shop_ids) and all(
        _load_stock_watch_snapshot(telegram_id, watched_shop_id, "low") is not None
        for watched_shop_id in shop_ids
    )
    await message.answer(
        "📉 <b>Уведомления о низких остатках</b>\n\n"
        f"Статус: {'✅ включены' if enabled else '❌ выключены'}\n"
        f"Магазинов под наблюдением: <b>{len(shop_ids) or 1}</b>\n"
        f"Активный в меню: <code>{shop_id}</code>\n"
        f"Порог: ≤ <b>{threshold}</b> шт.\n"
        f"Проверка каждые: <b>{max(300, LOW_STOCK_CHECK_INTERVAL_SECONDS)}</b> сек.\n"
        f"Состояние: {'остатки уже запомнены' if initialized else 'инициализация при следующей проверке'}\n\n"
        "Бот уведомит, когда товар впервые опустится до порога или ниже.",
        reply_markup=menu_for_message(message),
    )


# --- Уведомления о товарах с нулевым остатком ---
# Логика:
# 1) при первом запуске бот запоминает текущие SKU с остатком 0 и не спамит ими;
# 2) дальше проверяет остатки каждые OUT_OF_STOCK_CHECK_INTERVAL_SECONDS секунд;
# 3) если SKU впервые стал равен 0, бот присылает отдельное срочное уведомление.
_seen_out_of_stock_keys_by_user: dict[int, set[str]] = {}
_out_of_stock_watch_initialized: set[int] = set()


async def check_out_of_stock_once() -> None:
    users = connected_users_for_order_watch()

    for row in users:
        telegram_id = int(row["telegram_id"])
        shop_id = int(row["default_shop_id"])
        encrypted_token = row["uzum_token_encrypted"]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows = await load_sku_rows(client, shop_id, max_pages=50)
        except Exception:
            logging.exception("Out of stock watcher: failed to check stocks for %s", telegram_id)
            continue

        zero_rows = [
            r
            for r in rows
            if r.get("total") is not None and int(r.get("total") or 0) == 0
        ]
        zero_keys_now = [stock_row_key(r) for r in zero_rows]
        known = _seen_out_of_stock_keys_by_user.setdefault(telegram_id, set())

        # Первый проход: запоминаем текущие нулевые остатки, чтобы не прислать старые как новые.
        if telegram_id not in _out_of_stock_watch_initialized:
            known.update(zero_keys_now)
            _out_of_stock_watch_initialized.add(telegram_id)
            logging.info(
                "Out of stock watcher initialized for user=%s shop=%s zero_skus=%s",
                telegram_id,
                shop_id,
                len(zero_keys_now),
            )
            continue

        new_zero_rows = [r for r, key in zip(zero_rows, zero_keys_now) if key not in known]

        # Если товар снова появился в наличии, удаляем его из known.
        # Тогда при повторном падении до 0 бот снова уведомит.
        _seen_out_of_stock_keys_by_user[telegram_id] = set(zero_keys_now)

        if not new_zero_rows:
            continue

        lines = [format_sku_stock_line(item, mode="all") for item in new_zero_rows[:10]]
        more = "" if len(new_zero_rows) <= 10 else f"\n\nЕщё SKU с нулевым остатком: {len(new_zero_rows) - 10}"
        text = (
            f"❌ <b>Товар закончился</b>\n"
            f"Магазин: <code>{shop_id}</code>\n"
            f"Новых позиций с остатком 0: <b>{len(new_zero_rows)}</b>\n\n"
            + "\n\n".join(lines)
            + more
            + "\n\nПоказать товары с низким остатком: <code>/lowstock 0</code>"
        )

        try:
            await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
        except Exception:
            logging.exception("Out of stock watcher: failed to send notification to %s", telegram_id)


async def out_of_stock_watch_loop() -> None:
    await asyncio.sleep(30)
    logging.info(
        "Out of stock watcher started. Interval: %s seconds. Enabled: %s",
        OUT_OF_STOCK_CHECK_INTERVAL_SECONDS,
        OUT_OF_STOCK_NOTIFICATIONS,
    )
    while True:
        try:
            await check_out_of_stock_once()
        except Exception:
            logging.exception("Out of stock watcher loop error")
        await asyncio.sleep(max(300, OUT_OF_STOCK_CHECK_INTERVAL_SECONDS))


@dp.message(Command("outofstock_notify_status"))
async def outofstock_notify_status(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    enabled = product_setting_enabled(telegram_id, "notify_out_of_stock")
    shop_ids = connected_shop_ids_for_user(telegram_id)
    initialized = bool(shop_ids) and all(
        _load_stock_watch_snapshot(telegram_id, watched_shop_id, "zero") is not None
        for watched_shop_id in shop_ids
    )
    await message.answer(
        "❌ <b>Уведомления о нулевых остатках</b>\n\n"
        f"Статус: {'✅ включены' if enabled else '❌ выключены'}\n"
        f"Магазинов под наблюдением: <b>{len(shop_ids) or 1}</b>\n"
        f"Активный в меню: <code>{shop_id}</code>\n"
        f"Проверка каждые: <b>{max(300, OUT_OF_STOCK_CHECK_INTERVAL_SECONDS)}</b> сек.\n"
        f"Состояние: {'нулевые остатки уже запомнены' if initialized else 'инициализация при следующей проверке'}\n\n"
        "Бот уведомит, когда товар впервые опустится до остатка <b>0</b>.",
        reply_markup=menu_for_message(message),
    )


# --- Уведомления о новых продажах через Finance API ---
# Важно: старое уведомление "Новые заказы" смотрит только FBS/DBS заказы CREATED.
# Если продажа была FBO или уже не имеет статус CREATED, она может не попасть в этот список.
# Этот watcher смотрит финансовые строки продаж за сегодня и присылает уведомление о новых строках.
_seen_sale_keys_by_user: dict[int, set[str]] = {}
_sales_watch_initialized: set[int] = set()


def sale_key(item: dict[str, Any]) -> str:
    for key in (
        "id",
        "orderItemId",
        "orderItem_id",
        "orderId",
        "order_id",
        "skuId",
        "sku_id",
        "postingNumber",
        "number",
        "barcode",
    ):
        value = item.get(key)
        if value not in (None, ""):
            # Добавляем дату/сумму, чтобы разные продажи одного SKU не слиплись.
            date_part = ""
            for dk in ("date", "orderDate", "createdAt", "operationDate", "saleDate"):
                dv = item.get(dk)
                if dv not in (None, ""):
                    date_part = str(dv)
                    break
            amount = _finance_revenue(item)
            qty = _finance_qty(item)
            return f"{key}:{value}|{date_part}|{qty}|{amount}"
    raw = json.dumps(item, ensure_ascii=False, sort_keys=True, default=str)
    return "hash:" + hashlib.sha1(raw.encode("utf-8")).hexdigest()


def _deep_pick_value(obj: Any, names: tuple[str, ...]) -> Any:
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k in names and v not in (None, ""):
                return v
        for v in obj.values():
            found = _deep_pick_value(v, names)
            if found not in (None, ""):
                return found
    elif isinstance(obj, list):
        for v in obj:
            found = _deep_pick_value(v, names)
            if found not in (None, ""):
                return found
    return None


def _finance_sku_title(item: dict[str, Any]) -> str:
    value = _deep_pick_value(item, ("skuTitle", "skuName", "skuFullTitle", "offerName", "barcode", "skuId"))
    if isinstance(value, dict):
        value = pick(value, "title", "name", "value", "id")
    return str(value or "-")


def _finance_order_id(item: dict[str, Any]) -> str:
    return str(_deep_pick_value(item, ("orderId", "order_id", "orderNumber", "postingNumber")) or "-")


def _finance_sale_id(item: dict[str, Any]) -> str:
    return str(_deep_pick_value(item, ("id", "saleId", "operationId", "orderItemId", "orderItem_id")) or "-")


def _finance_date_value(item: dict[str, Any]) -> Any:
    return _deep_pick_value(item, ("date", "saleDate", "operationDate", "createdAt", "orderDate", "createdDate"))


def _format_finance_date(value: Any) -> str:
    if value in (None, ""):
        return "-"
    if isinstance(value, (int, float)):
        ts = float(value)
        if ts > 10_000_000_000:
            ts = ts / 1000
        try:
            return datetime.fromtimestamp(ts, UZT).strftime("%d.%m.%Y %H:%M")
        except Exception:
            return str(value)
    text_value = str(value).strip()
    try:
        iso = text_value.replace("Z", "+00:00")
        dt = datetime.fromisoformat(iso)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UZT)
        return dt.astimezone(UZT).strftime("%d.%m.%Y %H:%M")
    except Exception:
        return text_value[:16]


def format_sale_line(item: dict[str, Any]) -> str:
    title = escape(_finance_title(item))
    qty = _finance_qty(item)
    amount = _finance_revenue(item)
    status = escape(_finance_status(item))
    return (
        f"• <b>{title}</b>\n"
        f"  Штук: <b>{qty:g}</b> | Сумма: <b>{_format_money(amount)}</b> | Статус: <code>{status}</code>"
    )


def build_new_sale_message(item: dict[str, Any], shop_id: int | None = None, lang: str = "ru") -> str:
    lang = normalize_lang(lang)
    title = escape(_finance_title(item))
    sku = escape(_finance_sku_title(item))
    qty = _finance_qty(item)

    # Для уведомления показываем цену за штуку, как в Noorza Bot.
    unit_price = _deep_pick_number(item, ("sellPrice", "soldPrice", "price", "skuPrice", "productPrice"))
    if unit_price is None:
        unit_price = _finance_gross_revenue(item) / max(1.0, qty)

    commission = _finance_commission(item)
    logistics = _finance_logistics(item)
    payout_direct = _finance_payout_direct(item)
    payout = payout_direct if payout_direct is not None else max(0.0, _finance_gross_revenue(item) - commission - logistics)

    if lang == "uz":
        shop_line = f"🏪 Do‘kon: <code>{shop_id}</code>\n" if shop_id is not None else ""
        return (
            "🛒 <b>Yangi savdo</b>\n\n"
            + shop_line +
            f"📦 Tovar: <b>{title}</b>\n"
            f"🔖 SKU: <code>{sku}</code>\n"
            f"🔢 Soni: <b>{qty:g} dona</b>\n\n"
            f"💵 Narx: <b>{_format_money(float(unit_price or 0))}</b>\n"
            f"🏷 Komissiya: <b>{_format_money(float(commission))}</b>\n"
            f"🚚 Logistika: <b>{_format_money(float(logistics))}</b>\n"
            f"✅ To‘lovga: <b>{_format_money(float(payout))}</b>\n\n"
            f"🆔 Buyurtma: <code>{escape(_finance_order_id(item))}</code>\n"
            f"📌 Status: <code>{escape(_finance_status(item))}</code>\n"
            f"🕒 Sana: {escape(_format_finance_date(_finance_date_value(item)))}"
        )

    shop_line = f"🏪 Магазин: <code>{shop_id}</code>\n" if shop_id is not None else ""
    return (
        "🛒 <b>Новая продажа</b>\n\n"
        + shop_line +
        f"📦 Товар: <b>{title}</b>\n"
        f"🔖 SKU: <code>{sku}</code>\n"
        f"🔢 Кол-во: <b>{qty:g} шт.</b>\n\n"
        f"💵 Цена: <b>{_format_money(float(unit_price or 0))}</b>\n"
        f"🏷 Комиссия: <b>{_format_money(float(commission))}</b>\n"
        f"🚚 Логистика: <b>{_format_money(float(logistics))}</b>\n"
        f"✅ К выплате: <b>{_format_money(float(payout))}</b>\n\n"
        f"🆔 Заказ: <code>{escape(_finance_order_id(item))}</code>\n"
        f"📌 Статус: <code>{escape(_finance_status(item))}</code>\n"
        f"🕒 Дата: {escape(_format_finance_date(_finance_date_value(item)))}"
    )


async def check_new_sales_once() -> None:
    users = connected_users_for_order_watch()
    date_from, date_to = _today_range_ms()

    for row in users:
        telegram_id = int(row["telegram_id"])
        shop_id = int(row["default_shop_id"])
        encrypted_token = row["uzum_token_encrypted"]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows, first_response = await _load_finance_orders(
                client,
                shop_id,
                date_from_ms=date_from,
                date_to_ms=date_to,
                max_pages=5,
                page_size=100,
            )
        except Exception:
            logging.exception("Sales watcher: failed to check sales for %s", telegram_id)
            continue

        keys_now = [sale_key(item) for item in rows]
        known = _seen_sale_keys_by_user.setdefault(telegram_id, set())

        # Первый проход: запоминаем текущие продажи за сегодня, чтобы не прислать старые.
        if telegram_id not in _sales_watch_initialized:
            known.update(keys_now)
            _sales_watch_initialized.add(telegram_id)
            logging.info(
                "Sales watcher initialized for user=%s shop=%s sales_rows=%s",
                telegram_id,
                shop_id,
                len(keys_now),
            )
            continue

        new_rows = [item for item, key in zip(rows, keys_now) if key not in known]
        known.update(keys_now)

        if len(known) > 3000:
            _seen_sale_keys_by_user[telegram_id] = set(keys_now)

        if not new_rows:
            continue

        # Отправляем каждую новую продажу отдельным сообщением в стиле Noorza Bot.
        for item in new_rows[:10]:
            try:
                await bot.send_message(
                    telegram_id,
                    build_new_sale_message(item, shop_id=shop_id, lang=get_user_language(telegram_id)),
                    reply_markup=main_menu_for_user(telegram_id),
                )
                await asyncio.sleep(0.15)
            except Exception:
                logging.exception("Sales watcher: failed to send sale notification to %s", telegram_id)

        if len(new_rows) > 10:
            try:
                await bot.send_message(
                    telegram_id,
                    f"➕ Ещё новых строк продаж: <b>{len(new_rows) - 10}</b>\nПодробно: <code>/balance</code>",
                    reply_markup=main_menu_for_user(telegram_id),
                )
            except Exception:
                logging.exception("Sales watcher: failed to send summary notification to %s", telegram_id)


async def sales_watch_loop() -> None:
    await asyncio.sleep(40)
    logging.info(
        "Sales watcher started. Interval: %s seconds. Enabled: %s",
        SALE_CHECK_INTERVAL_SECONDS,
        SALE_NOTIFICATIONS,
    )
    while True:
        try:
            await check_new_sales_once()
        except Exception:
            logging.exception("Sales watcher loop error")
        await asyncio.sleep(max(60, SALE_CHECK_INTERVAL_SECONDS))


@dp.message(Command("sales_notify_status"))
async def sales_notify_status(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    enabled = product_setting_enabled(telegram_id, "notify_sales")
    shop_ids = connected_shop_ids_for_user(telegram_id)
    initialized = bool(shop_ids) and all(
        _watch_is_initialized(telegram_id, watched_shop_id, "finance")
        for watched_shop_id in shop_ids
    )
    await message.answer(
        "💸 <b>Уведомления о новых продажах</b>\n\n"
        f"Статус: {'✅ включены' if enabled else '❌ выключены'}\n"
        f"Магазинов под наблюдением: <b>{len(shop_ids) or 1}</b>\n"
        f"Активный в меню: <code>{shop_id}</code>\n"
        f"Проверка каждые: <b>{max(60, SALE_CHECK_INTERVAL_SECONDS)}</b> сек.\n"
        f"Состояние: {'все магазины инициализированы' if initialized else 'инициализация магазинов при следующей проверке'}\n\n"
        "Бот смотрит Finance API за сегодня. Если Finance API отдаёт продажу с задержкой, уведомление тоже придёт с задержкой.",
        reply_markup=menu_for_message(message),
    )



# --- Уведомления об изменении остатков: FBO + FBS/DBS ---
# Это нужно для FBO-продаж: FBO-заказ может не появиться в FBS/DBS CREATED,
# но остаток на складе Uzum уменьшается. Бот сравнивает общий остаток, FBO и FBS.
_stock_snapshot_by_user: dict[int, dict[str, dict[str, Any]]] = {}
_stock_change_watch_initialized: set[int] = set()


def _stock_qty(value: Any) -> int:
    num = _num_from_value(value)
    if num is None:
        return 0
    return int(num)


def _stock_snapshot_item(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "total": _stock_qty(row.get("total")),
        "fbo": _stock_qty(row.get("fbo")),
        "fbs": _stock_qty(row.get("fbs")),
        "title": str(
            row.get("title")
            or row.get("productTitle")
            or row.get("skuTitle")
            or row.get("name")
            or row.get("product_name")
            or "SKU"
        ),
        "price": row.get("price"),
        "row": row,
    }


def _stock_change_snapshot(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    snapshot: dict[str, dict[str, Any]] = {}
    for row in rows:
        snapshot[stock_row_key(row)] = _stock_snapshot_item(row)
    return snapshot


def _format_stock_change_line(key: str, before: dict[str, Any], after: dict[str, Any]) -> str:
    title = escape(str(after.get("title") or before.get("title") or key))
    old_total = int(before.get("total") or 0)
    new_total = int(after.get("total") or 0)
    old_fbo = int(before.get("fbo") or 0)
    new_fbo = int(after.get("fbo") or 0)
    old_fbs = int(before.get("fbs") or 0)
    new_fbs = int(after.get("fbs") or 0)

    parts = []
    if old_total != new_total:
        parts.append(f"Итого: <b>{old_total}</b> → <b>{new_total}</b>")
    if old_fbo != new_fbo:
        parts.append(f"FBO: <b>{old_fbo}</b> → <b>{new_fbo}</b>")
    if old_fbs != new_fbs:
        parts.append(f"FBS/DBS: <b>{old_fbs}</b> → <b>{new_fbs}</b>")
    if not parts:
        parts.append("остаток изменился")

    delta = new_total - old_total
    delta_text = f" | Разница: <b>{delta}</b> шт" if delta else ""
    return f"• <b>{title}</b>\n  " + " | ".join(parts) + delta_text


async def check_stock_change_once() -> None:
    users = connected_users_for_order_watch()

    for row in users:
        telegram_id = int(row["telegram_id"])
        shop_id = int(row["default_shop_id"])
        encrypted_token = row["uzum_token_encrypted"]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows = await load_sku_rows(client, shop_id, max_pages=20)
            snapshot_now = _stock_change_snapshot(rows)
        except Exception:
            logging.exception("Stock change watcher: failed to check stocks for %s", telegram_id)
            continue

        previous = _stock_snapshot_by_user.setdefault(telegram_id, {})

        # Первый проход: только запоминаем, чтобы не прислать старые изменения.
        if telegram_id not in _stock_change_watch_initialized:
            _stock_snapshot_by_user[telegram_id] = snapshot_now
            _stock_change_watch_initialized.add(telegram_id)
            logging.info(
                "Stock change watcher initialized for user=%s shop=%s skus=%s",
                telegram_id,
                shop_id,
                len(snapshot_now),
            )
            continue

        decreased: list[tuple[str, dict[str, Any], dict[str, Any]]] = []
        for key, after in snapshot_now.items():
            before = previous.get(key)
            if not before:
                continue
            before_total = int(before.get("total") or 0)
            after_total = int(after.get("total") or 0)
            before_fbo = int(before.get("fbo") or 0)
            after_fbo = int(after.get("fbo") or 0)
            before_fbs = int(before.get("fbs") or 0)
            after_fbs = int(after.get("fbs") or 0)

            if after_total < before_total or after_fbo < before_fbo or after_fbs < before_fbs:
                decreased.append((key, before, after))

        _stock_snapshot_by_user[telegram_id] = snapshot_now

        if not decreased:
            continue

        lines = [_format_stock_change_line(key, before, after) for key, before, after in decreased[:10]]
        more = "" if len(decreased) <= 10 else f"\n\nЕщё изменений: {len(decreased) - 10}"
        text = (
            "📦 <b>Изменение остатков</b>\n"
            f"Магазин: <code>{shop_id}</code>\n"
            "Уменьшился остаток по SKU. Это может быть продажа, резерв, списание или изменение склада.\n\n"
            + "\n\n".join(lines)
            + more
            + "\n\nПроверить остатки: <code>/stock</code>"
        )

        try:
            await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
        except Exception:
            logging.exception("Stock change watcher: failed to send notification to %s", telegram_id)


async def stock_change_watch_loop() -> None:
    await asyncio.sleep(70)
    logging.info(
        "Stock change watcher started. Interval: %s seconds. Enabled: %s",
        STOCK_CHANGE_CHECK_INTERVAL_SECONDS,
        STOCK_CHANGE_NOTIFICATIONS,
    )
    while True:
        try:
            await check_stock_change_once()
        except Exception:
            logging.exception("Stock change watcher loop error")
        await asyncio.sleep(max(60, STOCK_CHANGE_CHECK_INTERVAL_SECONDS))


@dp.message(Command("stock_change_notify_status"))
async def stock_change_notify_status(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    enabled = product_setting_enabled(telegram_id, "notify_stock_change")
    initialized = telegram_id in _stock_change_watch_initialized
    await message.answer(
        "📦 <b>Уведомления об изменении остатков</b>\n\n"
        f"Статус: {'✅ включены' if enabled else '❌ выключены'}\n"
        f"Магазин: <code>{shop_id}</code>\n"
        f"Проверка каждые: <b>{max(60, STOCK_CHANGE_CHECK_INTERVAL_SECONDS)}</b> сек.\n"
        f"Состояние: {'остатки уже запомнены' if initialized else 'инициализация при следующей проверке'}\n\n"
        "Бот сравнивает <b>FBO</b>, <b>FBS/DBS</b> и <b>общий остаток</b>. "
        "Так можно поймать FBO-продажи, даже если Finance API показывает нули.",
        reply_markup=menu_for_message(message),
    )



# --- Умный раздел: что требует внимания ---
def _attention_recommendation_ru(low_count: int, zero_count: int, dead_count: int, missing_cost: int, low_margin: int, cancels_today: int) -> str:
    if zero_count > 0:
        return "Начните с товаров, которые закончились: они не смогут продаваться, пока не пополнятся остатки."
    if low_count > 0:
        return "Сначала проверьте товары, которые скоро закончатся, особенно если они хорошо продаются."
    if missing_cost > 0:
        return "Проверьте purchasePrice в карточках Uzum: бот берёт себестоимость только из Uzum и не подставляет её вручную."
    if low_margin > 0:
        return "Проверьте товары с низкой маржой: возможно, цена или себестоимость указаны невыгодно."
    if dead_count > 0:
        return "Посмотрите товары без продаж: возможно, стоит изменить цену, фото или вывести товар из оборота."
    if cancels_today > 0:
        return "Проверьте сегодняшние отмены и товары, по которым они произошли."
    return "Критичных проблем не видно. Можно посмотреть продажи и прибыль за 30 дней."


def _attention_recommendation_uz(low_count: int, zero_count: int, dead_count: int, missing_cost: int, low_margin: int, cancels_today: int) -> str:
    if zero_count > 0:
        return "Avval qoldig‘i tugagan tovarlarni tekshiring: qoldiq bo‘lmasa, savdo ham bo‘lmaydi."
    if low_count > 0:
        return "Avval tez tugayotgan tovarlarni tekshiring, ayniqsa ular yaxshi sotilayotgan bo‘lsa."
    if missing_cost > 0:
        return "Uzum kartalaridagi purchasePrice qiymatini tekshiring: bot tannarxni faqat Uzumdan oladi va qo‘lda almashtirmaydi."
    if low_margin > 0:
        return "Past marjali tovarlarni tekshiring: narx yoki tannarx foydasiz bo‘lishi mumkin."
    if dead_count > 0:
        return "30 kun sotilmagan tovarlarni ko‘rib chiqing: narx, rasm yoki aylanmani tekshirish kerak."
    if cancels_today > 0:
        return "Bugungi bekor qilishlarni tekshiring."
    return "Jiddiy muammo ko‘rinmayapti. 30 kunlik savdo va foydani ko‘rishingiz mumkin."


async def _build_attention_summary(message: Message) -> str | None:
    req = await require_connection(message)
    if req is None:
        return None
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)

    # 1) Остатки
    stock_rows = await load_sku_rows(client, shop_id, max_pages=50)
    low_count = 0
    zero_count = 0
    for row in stock_rows:
        total = _stock_row_total(row)
        if total <= 0:
            zero_count += 1
        elif total <= LOW_STOCK_THRESHOLD:
            low_count += 1

    # 2) Накопительные потери и брак по всем SKU, включая архивные карточки
    try:
        loss_rows, _ = await _load_all_time_loss_rows(client, shop_id)
        missing_count = len(loss_rows)
    except Exception:
        missing_count = 0

    # 3) Продажи/отмены сегодня и товары без продаж за DEAD_STOCK_DAYS
    date_today_from, date_today_to = _today_range_ms()
    today_rows, _ = await _load_finance_orders(client, shop_id, date_from_ms=date_today_from, date_to_ms=date_today_to, max_pages=3, page_size=100)
    cancels_today = sum(1 for item in today_rows if _is_cancelled_status(_finance_status(item)))

    date_from, date_to = _days_range_ms(DEAD_STOCK_DAYS)
    sales_rows, _ = await _load_finance_orders(client, shop_id, date_from_ms=date_from, date_to_ms=date_to, max_pages=5, page_size=100)
    sold_keys: set[str] = set()
    for item in sales_rows:
        if not _is_cancelled_status(_finance_status(item)):
            sold_keys.update(_sale_match_keys(item))
    dead_count = 0
    for row in stock_rows:
        if _stock_row_total(row) <= 0:
            continue
        keys = _stock_match_keys(row)
        if keys and not keys.intersection(sold_keys):
            dead_count += 1

    # 4) Юнит-экономика: нет себестоимости и низкая маржа
    try:
        unit_rows, _stats, _saved_costs = await _unit_economy_for_shop(client, telegram_id, shop_id, days=30)
    except Exception:
        unit_rows = []
    missing_cost = sum(1 for r in unit_rows if r.get("cost_per_unit") is None)
    low_margin = sum(
        1
        for r in unit_rows
        if r.get("cost_per_unit") is not None
        and r.get("profit") is not None
        and float(r.get("margin") or 0) < LOW_MARGIN_THRESHOLD_PERCENT
    )

    if lang == "uz":
        rec = _attention_recommendation_uz(low_count, zero_count, dead_count, missing_cost, low_margin, cancels_today)
        return (
            "🧠 <b>Nimani tekshirish kerak</b>\n\n"
            f"🏪 Do‘kon: <code>{shop_id}</code>\n\n"
            f"⚠️ Tez tugayotgan tovarlar: <b>{low_count}</b>\n"
            f"❌ Qoldig‘i tugagan: <b>{zero_count}</b>\n"
            f"🐢 {DEAD_STOCK_DAYS} kun sotilmagan: <b>{dead_count}</b>\n"
            f"🧾 Tannarxi kiritilmagan: <b>{missing_cost}</b>\n"
            f"📉 Past marja: <b>{low_margin}</b>\n"
            f"❌ Bugungi bekor qilishlar: <b>{cancels_today}</b>\n"
            f"🧭 Yo‘qolganlar: <b>{missing_count}</b>\n\n"
            f"💡 <b>Tavsiya:</b> {escape(rec)}\n\n"
            "Pastdagi tugmalar orqali kerakli bo‘limni oching 👇"
        )

    rec = _attention_recommendation_ru(low_count, zero_count, dead_count, missing_cost, low_margin, cancels_today)
    return (
        "🧠 <b>Что требует внимания</b>\n\n"
        f"🏪 Магазин: <code>{shop_id}</code>\n\n"
        f"⚠️ Скоро закончится: <b>{low_count}</b>\n"
        f"❌ Закончились: <b>{zero_count}</b>\n"
        f"🐢 Без продаж {DEAD_STOCK_DAYS} дней: <b>{dead_count}</b>\n"
        f"🧾 Без себестоимости: <b>{missing_cost}</b>\n"
        f"📉 Низкая маржа: <b>{low_margin}</b>\n"
        f"❌ Отмены сегодня: <b>{cancels_today}</b>\n"
        f"🧭 Потерянные: <b>{missing_count}</b>\n\n"
        f"💡 <b>Рекомендация:</b> {escape(rec)}\n\n"
        "Ниже можете сразу открыть нужный раздел 👇"
    )


@dp.message(Command("attention", "check_attention", "what_check"))
async def attention_report(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    wait_text = "⌛ Do‘konni tekshiryapman..." if lang == "uz" else "⌛ Проверяю магазин..."
    await message.answer(wait_text, reply_markup=attention_menu_for_message(message))
    try:
        text = await _build_attention_summary(message)
        if text:
            await message.answer(text, reply_markup=attention_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)

# --- Главное меню в стиле Noorza Bot ---
@dp.message(F.text == "👑 Admin")
@dp.message(F.text == "👑 Админ")
async def button_admin_panel(message: Message) -> None:
    await admin_panel(message)


@dp.message(F.text == "👥 Foydalanuvchilar")
@dp.message(F.text == "👥 Пользователи")
async def button_admin_users(message: Message) -> None:
    await admin_users(message)


@dp.message(F.text == "💳 To‘lovlar")
@dp.message(F.text == "💳 Оплаты")
async def button_admin_payments(message: Message) -> None:
    await admin_payments(message)


@dp.message(F.text == "⏳ Tugayotganlar")
@dp.message(F.text == "⏳ Скоро заканчиваются")
async def button_admin_expiring(message: Message) -> None:
    await admin_expiring(message)


@dp.message(F.text == "⛔ Bloklanganlar")
@dp.message(F.text == "⛔ Заблокированные")
async def button_admin_blocked(message: Message) -> None:
    await admin_blocked_users(message)


@dp.message(F.text == "💰 Savdo")
@dp.message(F.text == "💰 Продажи")
async def button_sales_section_simple(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    text = "💰 <b>Savdo bo‘limi</b>\nKerakli davr yoki hisobotni tanlang 👇" if lang == "uz" else "💰 <b>Продажи</b>\nВыберите, что посмотреть 👇"
    await message.answer(text, reply_markup=sales_menu_for_message(message))


@dp.message(F.text.in_({"✨ Ещё по продажам", "✨ Yana savdo tahlili"}))
async def button_sales_more(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    text = (
        "✨ <b>Qo‘shimcha savdo tahlili</b>\nKamroq ishlatiladigan batafsil vositalar."
        if get_user_language(telegram_id) == "uz"
        else "✨ <b>Дополнительная аналитика</b>\nЗдесь собраны подробные инструменты, которые нужны не каждый день."
    )
    await message.answer(text, reply_markup=sales_more_menu_for_message(message))


@dp.message(F.text.in_({"⬅️ Продажи", "⬅️ Savdo"}))
async def button_back_to_sales(message: Message) -> None:
    await button_sales_section_simple(message)


@dp.message(F.text == "📦 Ombor")
@dp.message(F.text == "📦 Склад")
async def button_stock_section_simple(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    text = "📦 <b>Ombor</b>\nQoldiq, prognoz yoki yo‘qotishlarni tanlang 👇" if lang == "uz" else "📦 <b>Склад</b>\nОстатки, прогноз и потерянные товары 👇"
    await message.answer(text, reply_markup=stock_menu_for_message(message))


@dp.message(F.text == "🔔 Xabarnomalar")
@dp.message(F.text == "🔔 Уведомления")
async def button_notifications_section_simple(message: Message) -> None:
    await notification_hub_screen(message)


@dp.message(F.text == "📊 Hisobotlar")
@dp.message(F.text == "📊 Отчёты")
async def button_reports_section_simple(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    text = "📊 <b>Hisobotlar</b>\nPDF, Excel, foyda va tayyor hisobotlar 👇" if lang == "uz" else "📊 <b>Отчёты</b>\nPDF, Excel, прибыль и готовые отчёты 👇"
    await message.answer(text, reply_markup=report_menu_for_message(message))


@dp.message(F.text.in_({"🏠 Обзор магазина", "🏠 Do‘kon holati"}))
async def button_store_overview(message: Message) -> None:
    await dashboard(message)


@dp.message(F.text.in_({"🚨 Важно сейчас", "🚨 Hozir muhim"}))
async def button_important_now(message: Message) -> None:
    await business_control_center(message)


@dp.message(F.text.in_({"🧮 Финансы", "🧮 Moliya", "💰 Себестоимость и расходы", "💰 Tannarx va xarajat"}))
async def button_finance_section(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    text = (
        "🧮 <b>Moliya</b>\nTannarx, xarajatlar va haqiqiy foydani boshqaring."
        if get_user_language(telegram_id) == "uz"
        else "🧮 <b>Финансы</b>\nСебестоимость, расходы и реальная прибыль — в одном месте."
    )
    await message.answer(text, reply_markup=finance_menu_for_message(message))


@dp.message(F.text.in_({"🔐 Подключение Uzum", "🔐 Uzum ulanishi", "🔐 API и подключение", "🔐 API va ulanish"}))
async def button_connection_section(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    connected = db.has_uzum_connection(telegram_id)
    if get_user_language(telegram_id) == "uz":
        text = (
            "🔐 <b>Uzum ulanishi</b>\n\n"
            f"Holat: {'✅ ulangan' if connected else '❌ ulanmagan'}\n"
            "Bu yerda ulanishni tekshirish yoki API-kalitni xavfsiz yangilash mumkin."
        )
    else:
        text = (
            "🔐 <b>Подключение Uzum</b>\n\n"
            f"Статус: {'✅ подключено' if connected else '❌ не подключено'}\n"
            "Здесь можно проверить соединение или безопасно обновить API-ключ."
        )
    await message.answer(text, reply_markup=connection_menu_for_message(message))


@dp.message(F.text == "🧠 Tekshirish")
@dp.message(F.text == "🧠 Что проверить")
async def button_attention_section(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    if lang == "uz":
        text = (
            "🧠 <b>Nimani tekshirish kerak</b>\n\n"
            "Bot do‘koningizdagi muhim joylarni tekshiradi:\n"
            "⚠️ tez tugayotgan qoldiqlar, 🐢 sotilmayotgan tovarlar, "
            "🧾 tannarxi kiritilmagan SKU va 📉 past marja.\n\n"
            "Boshlash uchun <b>🔍 Hozir tekshirish</b> tugmasini bosing 👇"
        )
    else:
        text = (
            "🧠 <b>Что проверить</b>\n\n"
            "Бот быстро покажет, где есть проблемы:\n"
            "⚠️ заканчивающиеся остатки, 🐢 товары без продаж, "
            "🧾 SKU без себестоимости и 📉 низкую маржу.\n\n"
            "Чтобы начать, нажмите <b>🔍 Проверить сейчас</b> 👇"
        )
    await message.answer(text, reply_markup=attention_menu_for_message(message))


@dp.message(F.text == "🔍 Hozir tekshirish")
@dp.message(F.text == "🔍 Проверить сейчас")
@dp.message(F.text == "🔍 Do‘konni tekshirish")
@dp.message(F.text == "🔍 Проверить магазин")
async def button_attention_now(message: Message) -> None:
    await attention_report(message)


@dp.message(F.text == "⚠️ Qoldiqlar")
@dp.message(F.text == "⚠️ Остатки")
async def button_attention_stock(message: Message) -> None:
    await smart_lowstock(message)


@dp.message(F.text == "🐢 Sotuv yo‘q")
@dp.message(F.text == "🐢 Без продаж")
async def button_attention_dead_stock(message: Message) -> None:
    await dead_stock(message)


@dp.message(F.text == "🧾 Tannarx yo‘q")
@dp.message(F.text == "🧾 Нет себестоимости")
async def button_attention_missing_cost(message: Message) -> None:
    await unit_economy(message)


@dp.message(F.text == "📉 Past foyda")
@dp.message(F.text == "📉 Низкая прибыль")
async def button_attention_low_margin(message: Message) -> None:
    await profit_report(message)


@dp.message(Command("ikpu", "mxik", "catalog_codes"))
@dp.message(F.text.in_({"🧾 ИКПУ / МХИК", "🧾 IKPU / MXIK"}))
async def ikpu_catalog_report(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    if not await require_premium_subscription(message, telegram_id):
        return
    lang = get_user_language(telegram_id)
    await message.answer(
        "⌛ IKPU / MXIK kodlarini Uzumdan tekshiryapman..."
        if lang == "uz"
        else "⌛ Проверяю ИКПУ / МХИК по данным Uzum...",
        reply_markup=stock_menu_for_message(message),
    )
    try:
        status = await sync_uzum_sku_financials(
            client,
            telegram_id,
            shop_id,
            force=True,
        )
        rows = list_uzum_sku_financials(telegram_id, shop_id)
        missing_rows = [row for row in rows if not str(row.get("ikpu") or "").strip()]
        total = int(status.get("total") or 0)
        with_ikpu = int(status.get("with_ikpu") or 0)
        coverage = with_ikpu / total * 100.0 if total else 0.0
        title = "🧾 <b>IKPU / MXIK tekshiruvi</b>" if lang == "uz" else "🧾 <b>Проверка ИКПУ / МХИК</b>"
        summary = [
            f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
            f"📦 SKU jami: <b>{total}</b>" if lang == "uz" else f"📦 Всего SKU: <b>{total}</b>",
            f"✅ Kod bor: <b>{with_ikpu}</b> | ⚠️ Kod yo‘q: <b>{len(missing_rows)}</b>" if lang == "uz" else f"✅ Код заполнен: <b>{with_ikpu}</b> | ⚠️ Нет кода: <b>{len(missing_rows)}</b>",
            f"📊 To‘ldirilgan: <b>{coverage:.1f}%</b>" if lang == "uz" else f"📊 Заполнено: <b>{coverage:.1f}%</b>",
            "Manba: Uzum Product API dagi <code>ikpu</code> maydoni." if lang == "uz" else "Источник: поле <code>ikpu</code> в Uzum Product API.",
        ]
        if not missing_rows:
            await message.answer(
                title + "\n\n" + "\n".join(summary) + (
                    "\n\n✅ Barcha SKUlarda IKPU / MXIK kodi bor."
                    if lang == "uz"
                    else "\n\n✅ У всех SKU заполнен код ИКПУ / МХИК."
                ),
                reply_markup=stock_menu_for_message(message),
            )
            return
        items: list[str] = []
        for index, row in enumerate(missing_rows, start=1):
            product = escape(_short_text(str(row.get("product_title") or row.get("sku_title") or "—"), 80))
            sku = escape(_short_text(str(row.get("sku_title") or row.get("sku_id") or "—"), 65))
            article = escape(_short_text(str(row.get("seller_item_code") or "—"), 45))
            items.append(
                (
                    f"{index}. <b>{product}</b>\n🔖 SKU: <code>{sku}</code> | Artikul: <code>{article}</code>\n⚠️ IKPU / MXIK yo‘q"
                    if lang == "uz"
                    else f"{index}. <b>{product}</b>\n🔖 SKU: <code>{sku}</code> | Артикул: <code>{article}</code>\n⚠️ ИКПУ / МХИК не заполнен"
                )
            )
        await send_paginated_list(
            message,
            kind="ikpu_missing",
            title=title,
            summary=summary,
            items=items,
            section="stock",
            reply_markup=stock_menu_for_message(message),
        )
    except Exception as error:
        await send_api_error(message, error)


@dp.message(F.text == "❌ Bekor qilishlar")
@dp.message(F.text == "❌ Отмены")
async def button_attention_cancel(message: Message) -> None:
    await cancellations_report(message)


@dp.message(F.text == "🔐 Xavfsizlik")
@dp.message(F.text == "🔐 Безопасность")
async def button_security_simple(message: Message) -> None:
    await security(message)


@dp.message(F.text == "💸 Yangi savdolar")
async def button_sales_notify_status_uz(message: Message) -> None:
    await sales_notify_status(message)


@dp.message(F.text == "🚫 Bekor qilishlar")
@dp.message(F.text == "🚫 Отмены заказов")
async def button_cancel_notify_status(message: Message) -> None:
    await cancel_notify_status(message)


@dp.message(F.text == "📉 Kam qoldiq")
async def button_lowstock_notify_status_uz(message: Message) -> None:
    await lowstock_notify_status(message)


@dp.message(F.text == "❌ Qoldiq tugagan")
async def button_outofstock_notify_status_uz(message: Message) -> None:
    await outofstock_notify_status(message)


@dp.message(F.text == "⚙️ Holat")
async def button_status_uz(message: Message) -> None:
    await status(message)


@dp.message(F.text == "📦 Baza zaxirasi")
@dp.message(F.text == "📦 Бэкап базы")
async def button_admin_backup(message: Message) -> None:
    await admin_backup_db(message)


@dp.message(F.text == "📢 Xabar yuborish")
@dp.message(F.text == "📢 Рассылка")
async def button_admin_broadcast_help(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    await message.answer(
        "📢 <b>Рассылка</b>\n\n"
        "Чтобы отправить сообщение всем пользователям, напишите:\n"
        "<code>/broadcast ваш текст</code>\n\n"
        "Пример:\n"
        "<code>/broadcast Завтра в 09:00 будет обновление бота.</code>",
        reply_markup=admin_menu_for_message(message),
    )


@dp.message(F.text == "✅ Ulanishni tekshirish")
@dp.message(F.text == "✅ Проверить подключение")
async def button_check_connection(message: Message) -> None:
    await check_connection(message)


@dp.message(F.text == "⬅️ Asosiy menyu")
@dp.message(F.text == "⬅️ Главное меню")
@dp.message(F.text == "🏠 Asosiy")
@dp.message(F.text == "🏠 Главное")
@dp.message(F.text == "Menyu")
@dp.message(F.text == "Меню")
async def button_main_menu(message: Message) -> None:
    await message.answer(tr_user(upsert_from_message(message), "main_menu"), reply_markup=menu_for_message(message))



# --- Печать SKU-этикеток ---------------------------------------------------
# Официальные методы Swagger:
# GET  /v1/product/barcodes/types
# POST /v1/product/shop/{shopId}/barcodes/print


def _barcode_types_from_response(data: Any) -> list[dict[str, Any]]:
    """Извлекает справочник размеров этикеток из прямого или обёрнутого ответа API."""
    found: list[dict[str, Any]] = []

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            direct = value.get("barcodeLabelTypes")
            if isinstance(direct, list):
                for item in direct:
                    if isinstance(item, dict) and item.get("id") is not None:
                        found.append(item)
            for key in ("payload", "data", "result", "content"):
                nested = value.get(key)
                if isinstance(nested, (dict, list)):
                    walk(nested)
        elif isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    if item.get("id") is not None and any(k in item for k in ("title", "printType")):
                        found.append(item)
                    else:
                        walk(item)

    walk(data)
    unique: list[dict[str, Any]] = []
    seen: set[int] = set()
    for item in found:
        try:
            type_id = int(item.get("id"))
        except (TypeError, ValueError):
            continue
        if type_id in seen:
            continue
        seen.add(type_id)
        unique.append(item)
    return unique


def _barcode_type_markup(types: list[dict[str, Any]], lang: str) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for item in types:
        try:
            type_id = int(item.get("id"))
        except (TypeError, ValueError):
            continue
        title = str(item.get("title") or item.get("printType") or f"ID {type_id}").strip()
        print_type = str(item.get("printType") or "").strip()
        label = title if not print_type or print_type.lower() in title.lower() else f"{title} — {print_type}"
        rows.append([InlineKeyboardButton(text=f"🏷 {label}"[:60], callback_data=f"barcode_type:{type_id}")])
    rows.append([
        InlineKeyboardButton(
            text="❌ Bekor qilish" if normalize_lang(lang) == "uz" else "❌ Отмена",
            callback_data="barcode_cancel",
        )
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _parse_barcode_items(raw: str) -> tuple[list[dict[str, int]], str | None]:
    """Принимает строки вида `SKU_ID количество`; одинаковые SKU объединяет."""
    import re

    raw = (raw or "").strip()
    if not raw:
        return [], "empty"

    # Поддерживаем несколько строк, а также разделитель `;`.
    lines = [part.strip() for part in raw.replace(";", "\n").splitlines() if part.strip()]
    if len(lines) > BARCODE_MAX_SKUS:
        return [], "too_many_skus"

    merged: dict[int, int] = {}
    for line in lines:
        # Допустимо: 123456 10, 123456:10, 123456,10
        parts = [p for p in re.split(r"[\s,:=]+", line.strip()) if p]
        if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
            return [], f"bad_line:{line}"
        sku_id = int(parts[0])
        amount = int(parts[1])
        if sku_id <= 0:
            return [], f"bad_sku:{line}"
        if amount < 1 or amount > BARCODE_MAX_AMOUNT_PER_SKU:
            return [], f"bad_amount:{line}"
        merged[sku_id] = merged.get(sku_id, 0) + amount
        if merged[sku_id] > BARCODE_MAX_AMOUNT_PER_SKU:
            return [], f"bad_amount:{line}"

    if len(merged) > BARCODE_MAX_SKUS:
        return [], "too_many_skus"
    total = sum(merged.values())
    if total > BARCODE_MAX_TOTAL_LABELS:
        return [], f"too_many_total:{total}"

    return [{"skuId": sku_id, "amount": amount} for sku_id, amount in merged.items()], None


async def _download_product_barcodes_pdf(
    client: UzumClient,
    shop_id: int,
    data: list[dict[str, int]],
) -> bytes:
    """Делает бинарный POST и возвращает PDF; Accept специально меняется на application/pdf."""
    import httpx

    url = f"{client.base_url}/v1/product/shop/{int(shop_id)}/barcodes/print"
    headers = dict(getattr(client, "headers", {}) or {})
    headers["Accept"] = "application/pdf"
    headers["Content-Type"] = "application/json"
    timeout = httpx.Timeout(120.0, connect=15.0)

    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as http:
        response = await http.post(url, headers=headers, json={"data": data})

    if response.status_code >= 400:
        try:
            body: Any = response.json()
            body_text = json.dumps(body, ensure_ascii=False, default=str)
        except Exception:
            body_text = response.text[:1500]
        raise RuntimeError(f"Uzum API error {response.status_code}: {body_text}")

    content = response.content
    content_type = (response.headers.get("content-type") or "").lower()
    if not content:
        raise RuntimeError("Uzum API вернул пустой файл этикеток")
    if not content.startswith(b"%PDF") and "application/pdf" not in content_type:
        preview = content[:1200].decode("utf-8", errors="replace")
        raise RuntimeError(f"Uzum API вернул не PDF: {preview}")
    return content


@dp.message(Command("labels", "barcodes", "sku_labels"))
async def sku_labels_start(message: Message, state: FSMContext) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    await state.clear()

    try:
        raw_types = await client._request("GET", "/v1/product/barcodes/types")
        types = _barcode_types_from_response(raw_types)
        if not types:
            text = (
                "⚠️ Uzum API etiketka o‘lchamlarini qaytarmadi. Keyinroq qayta urinib ko‘ring."
                if normalize_lang(lang) == "uz"
                else "⚠️ Uzum API не вернул доступные размеры этикеток. Попробуйте позже."
            )
            await message.answer(text, reply_markup=stock_menu_for_message(message))
            return

        await state.update_data(barcode_shop_id=shop_id)
        if normalize_lang(lang) == "uz":
            text = (
                "🏷 <b>SKU etiketkalarini chop etish</b>\n\n"
                f"Do‘kon: <code>{shop_id}</code>\n"
                "Avval etiketka o‘lchamini tanlang 👇"
            )
        else:
            text = (
                "🏷 <b>Печать этикеток SKU</b>\n\n"
                f"Магазин: <code>{shop_id}</code>\n"
                "Сначала выберите размер этикетки 👇"
            )
        await message.answer(text, reply_markup=_barcode_type_markup(types, lang))
    except Exception as error:
        await state.clear()
        await send_api_error(message, error)


@dp.callback_query(F.data == "barcode_cancel")
async def sku_labels_cancel(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    await callback.answer()
    if callback.message:
        lang = get_user_language(callback.from_user.id)
        text = "Amal bekor qilindi." if normalize_lang(lang) == "uz" else "Печать этикеток отменена."
        await callback.message.answer(text, reply_markup=main_menu_for_user(callback.from_user.id))


@dp.callback_query(F.data.startswith("barcode_type:"))
async def sku_label_type_selected(callback: CallbackQuery, state: FSMContext) -> None:
    telegram_id = callback.from_user.id
    lang = get_user_language(telegram_id)
    try:
        type_id = int((callback.data or "").split(":", 1)[1])
    except (IndexError, TypeError, ValueError):
        await callback.answer("Неверный размер этикетки", show_alert=True)
        return

    client = get_uzum_for_user(telegram_id)
    shop_id = db.get_default_shop_id(telegram_id)
    if client is None or shop_id is None:
        await state.clear()
        await callback.answer("Сначала подключите магазин", show_alert=True)
        return
    if not has_active_subscription(telegram_id):
        await state.clear()
        await callback.answer("Подписка закончилась", show_alert=True)
        return

    await state.set_state(BarcodePrintStates.waiting_for_items)
    await state.update_data(barcode_type_id=type_id, barcode_shop_id=int(shop_id))
    await callback.answer()
    if not callback.message:
        return

    if normalize_lang(lang) == "uz":
        text = (
            "✅ O‘lcham tanlandi.\n\n"
            "Har bir qatorda <b>SKU ID va etiketkalar sonini</b> yuboring:\n"
            "<code>12345678 10</code>\n"
            "<code>87654321 5</code>\n\n"
            "Bir SKU uchun 1–100 dona. Bir nechta SKU alohida qatorda.\n"
            "Bekor qilish: <code>/cancel</code>"
        )
    else:
        text = (
            "✅ Размер выбран.\n\n"
            "Отправьте <b>ID SKU и количество этикеток</b>, каждую позицию с новой строки:\n"
            "<code>12345678 10</code>\n"
            "<code>87654321 5</code>\n\n"
            "Для одного SKU — от 1 до 100 этикеток.\n"
            "Отмена: <code>/cancel</code>"
        )
    await callback.message.answer(text)


@dp.message(BarcodePrintStates.waiting_for_items)
async def sku_label_items_received(message: Message, state: FSMContext) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    items, error = _parse_barcode_items(message.text or "")
    if error:
        if normalize_lang(lang) == "uz":
            if error == "too_many_skus":
                reason = f"Bir so‘rovda ko‘pi bilan {BARCODE_MAX_SKUS} ta SKU mumkin."
            elif error.startswith("too_many_total:"):
                total = error.split(":", 1)[1]
                reason = f"Jami {total} ta etiketka juda ko‘p. Bir martada ko‘pi bilan {BARCODE_MAX_TOTAL_LABELS} ta yuboring."
            elif error.startswith("bad_amount:"):
                reason = "Har bir SKU uchun son 1 dan 100 gacha bo‘lishi kerak."
            else:
                reason = "Format noto‘g‘ri. Har qatorda faqat SKU ID va son bo‘lishi kerak."
            text = (
                f"⚠️ {reason}\n\n"
                "To‘g‘ri misol:\n<code>12345678 10\n87654321 5</code>\n\n"
                "Bekor qilish: <code>/cancel</code>"
            )
        else:
            if error == "too_many_skus":
                reason = f"За один запрос можно отправить максимум {BARCODE_MAX_SKUS} SKU."
            elif error.startswith("too_many_total:"):
                total = error.split(":", 1)[1]
                reason = f"Всего указано {total} этикеток. Отправьте не более {BARCODE_MAX_TOTAL_LABELS} за один раз."
            elif error.startswith("bad_amount:"):
                reason = "Количество для каждого SKU должно быть от 1 до 100."
            else:
                reason = "Неверный формат. В каждой строке должны быть только ID SKU и количество."
            text = (
                f"⚠️ {reason}\n\n"
                "Правильный пример:\n<code>12345678 10\n87654321 5</code>\n\n"
                "Отмена: <code>/cancel</code>"
            )
        await message.answer(text)
        return

    saved = await state.get_data()
    try:
        barcode_type_id = int(saved.get("barcode_type_id"))
    except (TypeError, ValueError):
        await state.clear()
        await message.answer(
            "Размер этикетки потерян. Запустите <code>/labels</code> ещё раз.",
            reply_markup=stock_menu_for_message(message),
        )
        return

    req = await require_connection(message)
    if req is None:
        await state.clear()
        return
    _, client, shop_id = req
    payload = [
        {
            "skuId": int(item["skuId"]),
            "amount": int(item["amount"]),
            "barcodeTypeId": barcode_type_id,
        }
        for item in items
    ]
    total_labels = sum(int(item["amount"]) for item in items)

    wait_text = (
        f"⏳ {total_labels} ta etiketkali PDF tayyorlanmoqda..."
        if normalize_lang(lang) == "uz"
        else f"⏳ Формирую PDF на {total_labels} этикеток..."
    )
    await message.answer(wait_text)

    path: str | None = None
    try:
        pdf = await _download_product_barcodes_pdf(client, shop_id, payload)
        with tempfile.NamedTemporaryFile(prefix=f"uzum_labels_{shop_id}_", suffix=".pdf", delete=False) as tmp:
            tmp.write(pdf)
            path = tmp.name

        caption = (
            f"🏷 SKU etiketkalari tayyor\nDo‘kon: <code>{shop_id}</code>\nSKU: <b>{len(items)}</b> | Etiketka: <b>{total_labels}</b>"
            if normalize_lang(lang) == "uz"
            else f"🏷 Этикетки SKU готовы\nМагазин: <code>{shop_id}</code>\nSKU: <b>{len(items)}</b> | Этикеток: <b>{total_labels}</b>"
        )
        await message.answer_document(
            FSInputFile(path, filename=f"uzum_sku_labels_{shop_id}.pdf"),
            caption=caption,
            reply_markup=stock_menu_for_message(message),
        )
        await state.clear()
    except Exception as error:
        # Оставляем выбранный размер и состояние: пользователь может исправить SKU и повторить.
        await send_api_error(message, error)
    finally:
        if path:
            try:
                os.remove(path)
            except OSError:
                pass


@dp.message(F.text == "🏷 Этикетки SKU")
@dp.message(F.text == "🏷 SKU etiketkalari")
async def button_sku_labels(message: Message, state: FSMContext) -> None:
    await sku_labels_start(message, state)


@dp.message(F.text == "💰 Balans")
@dp.message(F.text == "💰 Баланс")
@dp.message(F.text == "📊 Biznes xulosa")
@dp.message(F.text == "📊 Бизнес-сводка")
async def button_balance(message: Message) -> None:
    await balance(message)


@dp.message(F.text == "📊 Bugun")
@dp.message(F.text == "📊 Сегодня")
async def button_today(message: Message) -> None:
    await today_sales(message)


@dp.message(F.text == "📆 Kecha")
@dp.message(F.text == "📆 Вчера")
async def button_yesterday(message: Message) -> None:
    await yesterday_sales(message)


@dp.message(F.text == "🗓 7 kun")
@dp.message(F.text == "🗓 7 дней")
async def button_week(message: Message) -> None:
    await week_sales(message)


@dp.message(F.text == "📅 30 kun")
@dp.message(F.text == "📅 30 дней")
async def button_30_days(message: Message) -> None:
    await balance(message)


@dp.message(F.text == "📦 Qoldiq")
@dp.message(F.text == "📦 Barcha qoldiq")
@dp.message(F.text == "📦 Остатки")
@dp.message(F.text == "📦 Все остатки")
async def button_stock_short(message: Message) -> None:
    await stock(message)


@dp.message(F.text == "⚠️ Заканчивается")
@dp.message(F.text == "⚠️ Заканчиваются")
async def button_lowstock_short(message: Message) -> None:
    await lowstock(message)


@dp.message(F.text == "🧭 Yo‘qolganlar")
@dp.message(F.text == "🧭 Потерянные")
@dp.message(F.text == "🧭 Yo‘qolgan tovarlar")
@dp.message(F.text == "🧭 Yo‘qotish va brak")
@dp.message(F.text == "🧭 Потерянные товары")
@dp.message(F.text == "🧭 Потери и брак")
async def button_lost(message: Message) -> None:
    await lost_goods(message)


@dp.message(F.text == "💎 Obuna")
@dp.message(F.text == "💎 Подписка")
@dp.message(F.text == "💎 To‘liq versiya")
@dp.message(F.text == "💎 Полная версия")
async def button_subscription(message: Message) -> None:
    await subscribe(message)


@dp.message(F.text == "ℹ️ Yordam")
@dp.message(F.text == "ℹ️ Помощь")
@dp.message(F.text == "❓ Помощь")
async def button_help(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    connected = db.has_uzum_connection(telegram_id)
    trial = subscription_access_level(telegram_id) == "trial"
    if lang == "uz":
        if connected:
            if trial:
                text = (
                    "ℹ️ <b>Sinov davri</b>\n\n"
                    "💰 <b>Savdo</b> — bugungi savdolar\n"
                    "💸 <b>Savdo xabarlari</b> — yangi savdolarni darhol yoki soatlik olish\n"
                    "🌙 <b>Ertalabki hisobot</b> — kunlik qisqa hisobot\n\n"
                    "Qolgan imkoniyatlar to‘liq obunada ochiladi."
                )
            else:
                text = (
                    "ℹ️ <b>Qayerga bosish kerak?</b>\n\n"
                    "🏠 <b>Do‘kon holati</b> — asosiy ko‘rsatkichlar\n"
                    "💰 <b>Savdo</b> — davr, foyda va top tovarlar\n"
                    "📦 <b>Ombor</b> — qoldiq, yo‘qotish va yetkazish rejasi\n"
                    "🚨 <b>Hozir muhim</b> — birinchi navbatdagi muammolar\n"
                    "📊 <b>Hisobotlar</b> — Excel va tayyor xulosalar\n"
                    "⚙️ <b>Sozlamalar</b> — xabarnomalar, moliya va ulanish\n\n"
                    "Muammo bo‘lsa, pastdagi tugma orqali administratorga yozing."
                )
        else:
            text = (
                "ℹ️ <b>Do‘konni ulash</b>\n\n"
                "1. Pastdagi videoni ko‘ring.\n"
                "2. <b>🔌 Do‘konni ulash</b> tugmasini bosing.\n"
                "3. Uzum Seller kabinetidagi API-kalitni yuboring.\n\n"
                "Bot kalitni tekshiradi va asosiy menyuni avtomatik ochadi."
            )
    else:
        if connected:
            if trial:
                text = (
                    "ℹ️ <b>Пробный период</b>\n\n"
                    "💰 <b>Продажи</b> — продажи за сегодня\n"
                    "💸 <b>Уведомления о продажах</b> — сразу или сводкой раз в час\n"
                    "🌙 <b>Утренний отчёт</b> — ежедневная краткая сводка\n\n"
                    "Остальные возможности открываются в полной подписке."
                )
            else:
                text = (
                    "ℹ️ <b>Куда нажимать?</b>\n\n"
                    "🏠 <b>Обзор магазина</b> — главные показатели\n"
                    "💰 <b>Продажи</b> — периоды, прибыль и топ товаров\n"
                    "📦 <b>Склад</b> — остатки, потери и план поставки\n"
                    "🚨 <b>Важно сейчас</b> — проблемы, требующие действий\n"
                    "📊 <b>Отчёты</b> — Excel и готовые сводки\n"
                    "⚙️ <b>Настройки</b> — уведомления, финансы и подключение\n\n"
                    "Если что-то не получается, напишите администратору кнопкой ниже."
                )
        else:
            text = (
                "ℹ️ <b>Как подключить магазин</b>\n\n"
                "1. Посмотрите видео по кнопке ниже.\n"
                "2. Нажмите <b>🔌 Подключить магазин</b>.\n"
                "3. Отправьте API-ключ из кабинета Uzum Seller.\n\n"
                "Бот проверит ключ и автоматически откроет рабочее меню."
            )
    await message.answer(text, reply_markup=help_links_markup(lang) or settings_menu_for_message(message))


@dp.message(F.text.in_({"🎥 Видеоинструкция", "🎥 Как подключить", "🎥 API ulash videosi", "🎥 Qanday ulash kerak"}))
async def button_video_instruction(message: Message) -> None:
    await video_instruction(message)


# Старые красивые кнопки оставлены для совместимости, если они остались у пользователя в Telegram.
@dp.message(F.text == "📊 Аналитика")
async def section_analytics(message: Message) -> None:
    await message.answer(tr_user(upsert_from_message(message), "main_menu"), reply_markup=menu_for_message(message))


@dp.message(F.text == "📦 Товары")
async def section_products(message: Message) -> None:
    await stock(message)


@dp.message(F.text == "🛒 Заказы/продажи")
async def section_orders(message: Message) -> None:
    await orders(message)


@dp.message(F.text == "🔔 Уведомления старое")
async def section_notifications(message: Message) -> None:
    await notify_status(message)


@dp.message(F.text.in_({"⚙️ Настройки", "⚙️ Sozlamalar", "⬅️ Настройки", "⬅️ Sozlamalar"}))
async def section_settings(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    trial = subscription_access_level(telegram_id) == "trial"
    if get_user_language(telegram_id) == "uz":
        text = (
            "⚙️ <b>Sinov sozlamalari</b>\n\nSavdo xabarlari, do‘konlar va Uzum ulanishini boshqaring."
            if trial
            else "⚙️ <b>Sozlamalar</b>\n\nXabarnomalar, do‘konlar, moliya va Uzum ulanishini shu yerda boshqaring."
        )
    else:
        text = (
            "⚙️ <b>Настройки пробного периода</b>\n\nЗдесь можно настроить уведомления о продажах, магазины и подключение Uzum."
            if trial
            else "⚙️ <b>Настройки</b>\n\nУведомления, магазины, финансы и подключение Uzum собраны здесь."
        )
    await message.answer(text, reply_markup=settings_menu_for_message(message))


@dp.message(F.text == "⭐ Отзывы")
@dp.message(F.text == "⭐ Sharhlar")
async def button_reviews(message: Message) -> None:
    await reviews(message)


@dp.message(F.text == "📈 Сводка")
@dp.message(F.text == "📈 Сводка FBO/FBS")
async def button_dashboard(message: Message) -> None:
    await dashboard(message)


@dp.message(F.text == "📊 Сводка заказов")
@dp.message(F.text == "📊 Сводка FBS/DBS")
async def button_orders_summary(message: Message) -> None:
    await orders_summary(message)


@dp.message(F.text == "💰 Продажи Finance")
async def button_sales(message: Message) -> None:
    await sales(message)


@dp.message(F.text == "📦 Все товары")
async def button_products(message: Message) -> None:
    await products(message)


@dp.message(F.text == "📊 Все остатки")
async def button_stock(message: Message) -> None:
    await stock(message)


@dp.message(F.text == "🏬 Остатки FBO")
async def button_fbo_stock(message: Message) -> None:
    await fbo(message)


@dp.message(F.text == "🚚 Остатки FBS/DBS")
async def button_fbs_stock(message: Message) -> None:
    await fbs(message)


@dp.message(F.text == "🛒 Новые заказы")
@dp.message(F.text == "🛒 FBS/DBS заказы")
async def button_orders(message: Message) -> None:
    await orders(message)


@dp.message(F.text == "📊 Excel hisobot")
@dp.message(F.text == "📊 Excel отчёт")
@dp.message(F.text == "📊 Excel-отчёт")
@dp.message(F.text == "📄 Excel-отчёт")
async def button_excel_report(message: Message) -> None:
    await report_excel(message)


@dp.message(F.text == "📄 PDF-отчёт")
@dp.message(F.text == "📄 PDF hisobot")
async def button_pdf_report(message: Message) -> None:
    await seller_pdf_report_menu(message)


@dp.message(F.text == "⚙️ Статус")
async def button_status(message: Message) -> None:
    await status(message)


@dp.message(F.text == "🏪 Do‘konlar")
@dp.message(F.text == "🏪 Магазины")
async def button_shops(message: Message) -> None:
    await shops(message)


@dp.message(F.text == "🔔 Новые заказы")
@dp.message(F.text == "🔔 FBS/DBS новые заказы")
async def button_notify_status(message: Message) -> None:
    await notify_status(message)


@dp.message(F.text == "💸 Новые продажи")
@dp.message(F.text == "💸 Продажи Finance")
@dp.message(F.text == "💸 Уведомления о продажах")
@dp.message(F.text == "💸 Savdo xabarlari")
async def button_sales_notify_status(message: Message) -> None:
    if str(message.text or "") in {"💸 Уведомления о продажах", "💸 Savdo xabarlari"}:
        await sales_mode_screen(message)
    else:
        await sales_notify_status(message)


@dp.message(F.text == "📦 Изменение остатков")
@dp.message(F.text == "📦 Изменение FBO/FBS")
@dp.message(F.text == "📦 FBO/FBS движение")
@dp.message(F.text == "📦 FBO/FBS движение остатков")
async def button_stock_change_notify_status(message: Message) -> None:
    await stock_change_notify_status(message)


@dp.message(F.text == "📉 Низкие остатки")
@dp.message(F.text == "📉 Низкие остатки FBO/FBS")
async def button_lowstock_notify_status(message: Message) -> None:
    await lowstock_notify_status(message)


@dp.message(F.text == "❌ Нет в наличии")
async def button_outofstock_notify_status(message: Message) -> None:
    await outofstock_notify_status(message)





# --- PRO FEATURES: multi-shop, analytics, reports, reminders ---
def _shop_id_from_any(shop: Any) -> int | None:
    if isinstance(shop, dict):
        for key in ("shopId", "shop_id", "id", "value"):
            value = shop.get(key)
            try:
                if value not in (None, ""):
                    return int(value)
            except Exception:
                pass
        for value in shop.values():
            if isinstance(value, dict):
                found = _shop_id_from_any(value)
                if found is not None:
                    return found
    else:
        for attr in ("shop_id", "shopId", "id"):
            value = getattr(shop, attr, None)
            try:
                if value not in (None, ""):
                    return int(value)
            except Exception:
                pass
    return None


def _shop_name_from_any(shop: Any) -> str:
    if isinstance(shop, dict):
        value = pick(shop, "title", "name", "shopName", "storeName", "legalName", "displayName")
        if value not in (None, ""):
            return str(value)
        for v in shop.values():
            if isinstance(v, dict):
                nested = _shop_name_from_any(v)
                if nested != "":
                    return nested
    return ""


async def _user_shop_list(telegram_id: int, client: UzumClient | None = None) -> list[dict[str, Any]]:
    shops_raw = db.list_shops(telegram_id) or []
    if not shops_raw and client is not None:
        try:
            data = await client.get_shops()
            items = extract_items(data)
            encrypted = db.get_encrypted_token(telegram_id)
            if encrypted and items:
                db.save_connection(telegram_id, encrypted, items)
            shops_raw = db.list_shops(telegram_id) or items
        except Exception:
            shops_raw = []

    result: list[dict[str, Any]] = []
    seen: set[int] = set()
    for item in shops_raw:
        sid = _shop_id_from_any(item)
        if sid is None or sid in seen:
            continue
        seen.add(sid)
        result.append({"shop_id": sid, "name": _shop_name_from_any(item), "raw": item})
    return result


def _stock_row_title(row: dict[str, Any]) -> str:
    value = pick(row, "skuTitle", "sku_title", "title", "productTitle", "product_title", "name")
    if value in (None, ""):
        value = _deep_pick_value(row, ("skuTitle", "productTitle", "title", "name"))
    return str(value or "Без названия")


def _stock_row_sku(row: dict[str, Any]) -> str:
    value = pick(row, "sku", "skuId", "sku_id", "barcode", "offerId", "shopSku")
    if value in (None, ""):
        value = _deep_pick_value(row, ("sku", "skuId", "barcode", "offerId"))
    return str(value or "")


def _stock_row_total(row: dict[str, Any]) -> int:
    value = _num_from_value(pick(row, "total", "quantity", "available", "stock", "qty"))
    if value is None:
        value = _deep_pick_number(row, ("total", "quantity", "available", "stock", "qty"))
    return int(value or 0)


def _stock_row_price(row: dict[str, Any]) -> float:
    value = _num_from_value(pick(row, "sellPrice", "price", "purchasePrice", "oldPrice"))
    if value is None:
        value = _deep_pick_number(row, ("sellPrice", "price", "purchasePrice", "oldPrice"))
    return float(value or 0.0)


def _match_key(kind: str, value: Any) -> str | None:
    if isinstance(value, dict):
        value = pick(value, "id", "value", "title", "name", "code")
    if isinstance(value, (list, tuple, set, dict)):
        return None
    normalized = " ".join(str(value or "").strip().lower().split())
    if not normalized or normalized in {"-", "—", "none", "null", "0"}:
        return None
    return f"{kind}:{normalized}"


def _sale_match_keys(item: dict[str, Any]) -> set[str]:
    """Return typed Finance API identifiers for safe SKU matching."""
    keys: set[str] = set()

    def add(kind: str, *names: str) -> None:
        for name in names:
            value = _deep_pick_value(item, (name,))
            key = _match_key(kind, value)
            if key:
                keys.add(key)

    add("id", "skuId", "sku_id", "offerId")
    add("barcode", "barcode")
    add("seller", "sellerSku", "sellerItemCode", "article")
    add("variant", "skuTitle", "skuName", "skuFullTitle", "offerName")
    add("product", "productTitle", "productName")

    # Some Finance response variants expose only a generic `sku` or `title`.
    generic_sku = _deep_pick_value(item, ("sku",))
    for kind in ("id", "variant"):
        key = _match_key(kind, generic_sku)
        if key:
            keys.add(key)
    title_key = _match_key("product", _finance_title(item))
    if title_key:
        keys.add(title_key)
    return keys


def _stock_match_keys(row: dict[str, Any]) -> set[str]:
    """Return identifiers compatible with `_sale_match_keys`."""
    keys: set[str] = set()

    def add(kind: str, *fields: str) -> None:
        for field in fields:
            key = _match_key(kind, row.get(field))
            if key:
                keys.add(key)

    add("id", "sku_id", "skuId", "offerId")
    add("barcode", "barcode")
    add("seller", "seller_item_code", "sellerItemCode", "article")
    add("variant", "sku_full_title", "sku_title", "skuFullTitle", "skuTitle", "sku")
    add("product", "product_title", "productTitle", "title")
    return keys


def _unique_stock_match_keys(stock_rows: list[dict[str, Any]]) -> list[set[str]]:
    """Only keep aliases that identify one SKU within the current shop.

    A product title can be shared by several size/color variants.  Matching a
    sale by that non-unique title would incorrectly give the same sale to every
    variant and inflate the forecast.
    """
    all_keys = [_stock_match_keys(row) for row in stock_rows]
    counts: dict[str, int] = {}
    for keys in all_keys:
        for key in keys:
            counts[key] = counts.get(key, 0) + 1
    return [{key for key in keys if counts.get(key) == 1} for keys in all_keys]


def _merge_noorza_stats(stats_list: list[dict[str, Any]]) -> dict[str, Any]:
    result = {
        "rows": 0.0,
        "orders": 0.0,
        "cancelled": 0.0,
        "returned_rows": 0.0,
        "units": 0.0,
        "returns": 0.0,
        "revenue": 0.0,
        "commission": 0.0,
        "logistics": 0.0,
        "payout_total": 0.0,
        "withdrawn": 0.0,
        "left_to_withdraw": 0.0,
        "statuses": {},
    }
    statuses: dict[str, int] = {}
    products: dict[str, dict[str, Any]] = {}
    for stats in stats_list:
        for key in ("rows", "orders", "cancelled", "returned_rows", "units", "returns", "revenue", "commission", "logistics", "payout_total", "withdrawn", "left_to_withdraw"):
            result[key] = float(result.get(key) or 0) + float(stats.get(key) or 0)
        for status, count in (stats.get("statuses") or {}).items():
            statuses[str(status)] = statuses.get(str(status), 0) + int(count or 0)
        for product in stats.get("top_products") or []:
            key = str(product.get("sku") or product.get("title") or "—").strip().lower()
            entry = products.setdefault(key, {
                "title": product.get("title"),
                "sku": product.get("sku"),
                "qty": 0.0,
                "revenue": 0.0,
                "payout": 0.0,
            })
            for field in ("qty", "revenue", "payout"):
                entry[field] = float(entry.get(field) or 0) + float(product.get(field) or 0)
    result["statuses"] = statuses
    revenue = float(result.get("revenue") or 0)
    orders = float(result.get("orders") or 0)
    units = float(result.get("units") or 0)
    total_outcomes = float(result.get("rows") or 0) + float(result.get("cancelled") or 0) + float(result.get("returned_rows") or 0)
    result["average_order"] = revenue / max(1.0, orders)
    result["average_unit"] = revenue / max(1.0, units)
    result["commission_rate"] = float(result.get("commission") or 0) / revenue if revenue > 0 else 0.0
    result["logistics_rate"] = float(result.get("logistics") or 0) / revenue if revenue > 0 else 0.0
    result["cancellation_rate"] = float(result.get("cancelled") or 0) / max(1.0, total_outcomes)
    result["top_products"] = sorted(products.values(), key=lambda value: float(value.get("revenue") or 0), reverse=True)[:5]
    return result


def _format_all_shops_balance(days_title: str, shops_count: int, stats: dict[str, Any], per_shop: list[str]) -> str:
    text = (
        f"🌐 <b>Баланс по всем магазинам {escape(days_title)}</b>\n\n"
        f"🏪 Магазинов: <b>{shops_count}</b>\n"
        f"🛒 Заказов: <b>{int(stats.get('orders') or 0)}</b>\n"
        f"🧾 Позиций: <b>{int(stats['rows'])}</b>\n"
        f"📦 Товаров продано: <b>{float(stats['units']):.0f} шт.</b>\n"
        f"❌ Отмен: <b>{int(stats.get('cancelled') or 0)}</b>\n"
        f"↩️ Возвратов: <b>{float(stats['returns']):.0f} шт.</b>\n\n"
        f"💵 Выручка: <b>{_format_money(float(stats['revenue']))}</b>\n"
        f"🏷 Комиссия Uzum: <b>{_format_money(float(stats['commission']))}</b>\n"
        f"🚚 Логистика: <b>{_format_money(float(stats['logistics']))}</b>\n\n"
        f"✅ К выплате: <b>{_format_money(float(stats['payout_total']))}</b>\n"
        f"💳 Уже выведено: <b>{_format_money(float(stats['withdrawn']))}</b>\n"
        f"🧾 Остаток к выплате: <b>{_format_money(float(stats['left_to_withdraw']))}</b>"
    )
    if per_shop:
        text += "\n\n<b>🏪 По магазинам:</b>\n" + "\n".join(per_shop[:20])
    return text


async def _all_shops_finance_stats(telegram_id: int, client: UzumClient, date_from: int, date_to: int) -> tuple[dict[str, Any], list[str], int]:
    lang = get_user_language(telegram_id)
    shops_list = await _user_shop_list(telegram_id, client)
    if not shops_list:
        default_shop = db.get_default_shop_id(telegram_id)
        if default_shop:
            shops_list = [{"shop_id": int(default_shop), "name": "", "raw": {}}]
    stats_list: list[dict[str, Any]] = []
    per_shop: list[str] = []
    for shop in shops_list:
        sid = int(shop["shop_id"])
        try:
            rows, _, _ = await _load_finance_range_flexible(client, sid, date_from, date_to)
            stats = _build_noorza_today_stats(rows)
            stats_list.append(stats)
            name = f" — {escape(shop['name'])}" if shop.get("name") else ""
            if lang == "uz":
                per_shop.append(
                    f"• <code>{sid}</code>{name}: {_format_money(float(stats['revenue']))}, "
                    f"{float(stats['units']):.0f} dona, to‘lovga {_format_money(float(stats['payout_total']))}"
                )
            else:
                per_shop.append(
                    f"• <code>{sid}</code>{name}: {_format_money(float(stats['revenue']))}, "
                    f"{float(stats['units']):.0f} шт., к выплате {_format_money(float(stats['payout_total']))}"
                )
            await asyncio.sleep(0.2)
        except Exception as e:
            logging.exception("All shops balance: failed for shop=%s", sid)
            per_shop.append(f"• <code>{sid}</code>: ошибка API — {escape(str(e)[:80])}")
    return _merge_noorza_stats(stats_list), per_shop, len(shops_list)


async def _all_shops_business_stats(
    telegram_id: int,
    client: UzumClient,
    date_from: int,
    date_to: int,
    *,
    days: int,
) -> tuple[dict[str, Any], list[dict[str, Any]], int]:
    """Build one honest profit view for every connected shop.

    Finance, product costs and marketplace expenses are isolated per shop.
    A failed shop remains visible in the result instead of silently lowering
    the aggregate.  Cost is read only from Uzum ``purchasePrice`` cache.
    """
    shops_list = await _user_shop_list(telegram_id, client)
    if not shops_list:
        default_shop = db.get_default_shop_id(telegram_id)
        if default_shop:
            shops_list = [{"shop_id": int(default_shop), "name": "", "raw": {}}]

    stats_list: list[dict[str, Any]] = []
    per_shop: list[dict[str, Any]] = []
    for shop in shops_list:
        sid = int(shop["shop_id"])
        result: dict[str, Any] = {
            "shop_id": sid,
            "name": str(shop.get("name") or ""),
            "error": None,
        }
        try:
            rows, _, _ = await _load_finance_range_flexible(
                client,
                sid,
                date_from,
                date_to,
            )
            stats = _build_noorza_today_stats(rows)
            stats_list.append(stats)

            cost_cache_stale = False
            try:
                cost_status = await sync_uzum_sku_financials(
                    client,
                    telegram_id,
                    sid,
                )
                cost_cache_stale = bool(cost_status.get("stale"))
            except Exception:
                cost_cache_stale = True
                logging.exception(
                    "All shops business: Uzum purchasePrice sync failed shop=%s",
                    sid,
                )

            settings = ensure_finance_settings(telegram_id, sid)
            costs = get_unit_cost_map(telegram_id, sid)
            products = _build_unit_rows_from_finance(
                rows,
                costs,
                tax_percent=float(settings.get("tax_percent") or 0),
            )
            cost_summary = _profit_summary_from_unit_rows(products, stats)
            expense_summary = await load_uzum_expense_summary(
                client,
                sid,
                date_from,
                date_to,
            )
            business = calculate_business_profit(
                cost_summary,
                stats,
                settings,
                days=days,
                uzum_expenses=expense_summary,
            )
            if cost_cache_stale:
                business["complete"] = False
            result.update(
                {
                    "stats": stats,
                    "cost_summary": cost_summary,
                    "business": business,
                    "cost_cache_stale": cost_cache_stale,
                }
            )
        except Exception as error:
            result["error"] = str(error)
            logging.exception("All shops business: failed for shop=%s", sid)
        per_shop.append(result)
        await asyncio.sleep(0.15)

    totals = _merge_noorza_stats(stats_list)
    successful = [row for row in per_shop if not row.get("error")]
    businesses = [dict(row.get("business") or {}) for row in successful]
    cost_summaries = [dict(row.get("cost_summary") or {}) for row in successful]
    total_revenue = float(totals.get("revenue") or 0)
    known_revenue = sum(float(row.get("known_revenue") or 0) for row in cost_summaries)
    coverage = known_revenue / total_revenue if total_revenue > 0 else 0.0
    totals.update(
        {
            "cost_total": sum(float(row.get("cost_total") or 0) for row in businesses),
            "known_profit": sum(float(row.get("known_profit") or 0) for row in businesses),
            "calculation_revenue": sum(
                float(row.get("calculation_revenue") or 0) for row in businesses
            ),
            "calculation_payout": sum(
                float(row.get("calculation_payout") or 0) for row in businesses
            ),
            "calculation_commission": sum(
                float(row.get("calculation_commission") or 0) for row in businesses
            ),
            "calculation_logistics": sum(
                float(row.get("calculation_logistics") or 0) for row in businesses
            ),
            "other_payout_deductions": sum(
                float(row.get("other_payout_deductions") or 0) for row in businesses
            ),
            "payout_adjustment": sum(
                float(row.get("payout_adjustment") or 0) for row in businesses
            ),
            "tax_expense": sum(float(row.get("tax_expense") or 0) for row in businesses),
            "uzum_expense_total": sum(
                float(row.get("uzum_expense_total") or 0) for row in businesses
            ),
            "uzum_expense_deductions": sum(
                float(row.get("uzum_expense_deductions") or 0) for row in businesses
            ),
            "uzum_expense_refunds": sum(
                float(row.get("uzum_expense_refunds") or 0) for row in businesses
            ),
            "external_expense_total": sum(
                float(row.get("advertising_expense") or 0)
                + float(row.get("storage_expense") or 0)
                + float(row.get("other_expense") or 0)
                for row in businesses
            ),
            "net_profit": sum(float(row.get("net_profit") or 0) for row in businesses),
            "coverage": max(0.0, min(1.0, coverage)),
            "missing_count": sum(
                int(row.get("missing_count") or 0) for row in businesses
            ),
            "expenses_available": bool(successful)
            and all(bool(row.get("uzum_expenses_available")) for row in businesses),
            "complete": len(successful) == len(shops_list)
            and bool(successful)
            and all(bool(row.get("complete")) for row in businesses),
            "successful_shops": len(successful),
            "failed_shops": max(0, len(shops_list) - len(successful)),
        }
    )
    return totals, per_shop, len(shops_list)


def _format_all_shops_business(
    telegram_id: int,
    *,
    days: int,
    shops_count: int,
    totals: dict[str, Any],
    per_shop: list[dict[str, Any]],
) -> str:
    lang = get_user_language(telegram_id)
    uz = normalize_lang(lang) == "uz"
    complete = bool(totals.get("complete"))
    coverage = float(totals.get("coverage") or 0)
    result_label = (
        "Sof foyda"
        if uz and complete
        else "Ma’lum ma’lumotlar bo‘yicha natija"
        if uz
        else "Чистая прибыль"
        if complete
        else "Результат по известным данным"
    )
    if uz:
        lines = [
            f"🌐 <b>Barcha do‘konlar — {days} kun</b>",
            f"🏪 Do‘konlar: <b>{shops_count}</b>",
            "",
            f"🛒 Buyurtma: <b>{int(totals.get('orders') or 0)}</b> | 📦 Sotildi: <b>{float(totals.get('units') or 0):.0f}</b>",
            f"❌ Bekor: <b>{int(totals.get('cancelled') or 0)}</b> | ↩️ Qaytarish: <b>{float(totals.get('returns') or 0):.0f}</b>",
            "",
            *_format_profit_bridge_lines(totals, lang=lang),
            f"📌 Tannarx qamrovi: <b>{coverage * 100:.1f}%</b>",
        ]
    else:
        lines = [
            f"🌐 <b>Все магазины — {days} дней</b>",
            f"🏪 Магазинов: <b>{shops_count}</b>",
            "",
            f"🛒 Заказов: <b>{int(totals.get('orders') or 0)}</b> | 📦 Продано: <b>{float(totals.get('units') or 0):.0f}</b>",
            f"❌ Отмен: <b>{int(totals.get('cancelled') or 0)}</b> | ↩️ Возвратов: <b>{float(totals.get('returns') or 0):.0f}</b>",
            "",
            *_format_profit_bridge_lines(totals, lang=lang),
            f"📌 Покрытие себестоимостью: <b>{coverage * 100:.1f}%</b>",
        ]

    if not complete:
        warning_parts: list[str] = []
        missing = int(totals.get("missing_count") or 0)
        if missing:
            warning_parts.append(
                f"Uzum purchasePrice bermagan SKU: {missing}"
                if uz
                else f"SKU без purchasePrice от Uzum: {missing}"
            )
        if not bool(totals.get("expenses_available")):
            warning_parts.append(
                "Uzum xarajatlari to‘liq olinmadi"
                if uz
                else "расходы Uzum загружены не полностью"
            )
        if int(totals.get("failed_shops") or 0):
            warning_parts.append(
                f"API xatosi bo‘lgan do‘konlar: {int(totals['failed_shops'])}"
                if uz
                else f"магазинов с ошибкой API: {int(totals['failed_shops'])}"
            )
        if warning_parts:
            lines.extend(["", "⚠️ " + "; ".join(warning_parts) + "."])

    lines.extend(["", "<b>🏪 Do‘konlar bo‘yicha:</b>" if uz else "<b>🏪 По магазинам:</b>"])
    for row in per_shop[:20]:
        sid = int(row.get("shop_id") or 0)
        name = escape(_short_text(str(row.get("name") or ""), 45))
        name_part = f" — {name}" if name else ""
        if row.get("error"):
            lines.append(
                f"• <code>{sid}</code>{name_part}: API xatosi"
                if uz
                else f"• <code>{sid}</code>{name_part}: ошибка API"
            )
            continue
        stats = dict(row.get("stats") or {})
        business = dict(row.get("business") or {})
        shop_complete = bool(business.get("complete"))
        shop_label = (
            "sof foyda"
            if uz and shop_complete
            else "ma’lum natija"
            if uz
            else "чистая прибыль"
            if shop_complete
            else "известный результат"
        )
        lines.append(
            f"• <code>{sid}</code>{name_part}: {_format_money(float(stats.get('revenue') or 0))}, "
            f"{shop_label} <b>{_format_money(float(business.get('net_profit') or 0))}</b>, "
            f"{float(business.get('coverage') or 0) * 100:.0f}%"
        )

    lines.extend(
        [
            "",
            (
                "ℹ️ Tannarx faqat Uzum purchasePrice’dan olinadi; yetishmagan qiymatlar taxmin qilinmaydi."
                if uz
                else "ℹ️ Себестоимость берётся только из Uzum purchasePrice; отсутствующие значения не подставляются."
            ),
        ]
    )
    return "\n".join(lines)


@dp.message(Command("balance_all", "all_balance", "allshops", "all_shops"))
async def balance_all_shops(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not await require_active_subscription(message, telegram_id):
        return
    client = get_uzum_for_user(telegram_id)
    if client is None:
        await message.answer("Сначала подключите Uzum API-токен: <code>/connect</code>", reply_markup=menu_for_message(message))
        return
    lang = get_user_language(telegram_id)
    await message.answer(
        "⌛ Barcha do‘konlar bo‘yicha savdo, tannarx va xarajatlar hisoblanmoqda..."
        if normalize_lang(lang) == "uz"
        else "⌛ Считаю продажи, себестоимость и расходы по всем магазинам...",
        reply_markup=menu_for_message(message),
    )
    try:
        date_from, date_to = _days_range_ms(30)
        totals, per_shop, shops_count = await _all_shops_business_stats(
            telegram_id,
            client,
            date_from,
            date_to,
            days=30,
        )
        await message.answer(
            _format_all_shops_business(
                telegram_id,
                days=30,
                shops_count=shops_count,
                totals=totals,
                per_shop=per_shop,
            ),
            reply_markup=menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)


async def _top_products_for_shop(client: UzumClient, shop_id: int, days: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    date_from, date_to = _days_range_ms(days)
    rows, _ = await _load_finance_orders(client, shop_id, date_from_ms=date_from, date_to_ms=date_to, max_pages=5, page_size=100)
    groups: dict[str, dict[str, Any]] = {}
    for item in rows:
        if _is_cancelled_status(_finance_status(item)):
            continue
        title = _finance_title(item)
        sku = _finance_sku_title(item)
        key = (sku or title or "-").strip().lower()
        if not key:
            key = title.strip().lower()
        entry = groups.setdefault(key, {"title": title, "sku": sku, "qty": 0.0, "revenue": 0.0, "payout": 0.0})
        gross = _finance_gross_revenue(item)
        commission = _finance_commission(item)
        logistics = _finance_logistics(item)
        payout_direct = _finance_payout_direct(item)
        payout = payout_direct if payout_direct is not None else max(0.0, gross - commission - logistics)
        entry["qty"] += _finance_qty(item)
        entry["revenue"] += gross
        entry["payout"] += max(0.0, payout)
    top = sorted(groups.values(), key=lambda x: float(x.get("revenue") or 0), reverse=True)
    return top, _build_noorza_today_stats(rows)


@dp.message(Command("top", "top_products"))
async def top_products(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    telegram_id = message.from_user.id if message.from_user else 0
    lang = get_user_language(telegram_id)
    days = TOP_PRODUCTS_DAYS
    await message.answer(f"⌛ {days} kunlik top tovarlarni hisoblayapman..." if lang == "uz" else f"⌛ Считаю топ товаров за {days} дней...", reply_markup=sales_menu_for_message(message))
    try:
        top, stats = await _top_products_for_shop(client, shop_id, days)
        if not top:
            text = f"🏆 <b>{days} kunlik top tovarlar</b>\n🏪 Do‘kon: <code>{shop_id}</code>\n\nSavdolar topilmadi." if lang == "uz" else f"🏆 <b>Топ товаров за {days} дней</b>\n🏪 Магазин: <code>{shop_id}</code>\n\nПродаж не найдено."
            await message.answer(text, reply_markup=sales_menu_for_message(message))
            return
        title = f"🏆 <b>{days} kunlik top tovarlar</b>" if lang == "uz" else f"🏆 <b>Топ товаров за {days} дней</b>"
        summary = [
            f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
            f"📦 Sotilgan: <b>{float(stats['units']):.0f} dona</b>" if lang == "uz" else f"📦 Всего продано: <b>{float(stats['units']):.0f} шт.</b>",
            f"💵 Tushum: <b>{_format_money(float(stats['revenue']))}</b>" if lang == "uz" else f"💵 Выручка: <b>{_format_money(float(stats['revenue']))}</b>",
        ]
        items: list[str] = []
        for idx, item in enumerate(top, start=1):
            title_item = escape(_short_text(item.get("title"), 85))
            sku = escape(_short_text(item.get("sku"), 60))
            sku_line = f"\n🔖 SKU: <code>{sku}</code>" if sku and sku != "-" else ""
            if lang == "uz":
                items.append(
                    f"{idx}. <b>{title_item}</b>{sku_line}\n"
                    f"🔢 Soni: <b>{float(item.get('qty') or 0):.0f} dona</b> | "
                    f"💵 Tushum: <b>{_format_money(float(item.get('revenue') or 0))}</b> | "
                    f"✅ To‘lovga: <b>{_format_money(float(item.get('payout') or 0))}</b>"
                )
            else:
                items.append(
                    f"{idx}. <b>{title_item}</b>{sku_line}\n"
                    f"🔢 Продано: <b>{float(item.get('qty') or 0):.0f} шт.</b> | "
                    f"💵 Выручка: <b>{_format_money(float(item.get('revenue') or 0))}</b> | "
                    f"✅ К выплате: <b>{_format_money(float(item.get('payout') or 0))}</b>"
                )
        await send_paginated_list(message, kind="top", title=title, summary=summary, items=items, section="sales", reply_markup=sales_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


@dp.message(Command("deadstock", "no_sales", "stuck"))
async def dead_stock(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    telegram_id = message.from_user.id if message.from_user else 0
    lang = get_user_language(telegram_id)
    days = DEAD_STOCK_DAYS
    await message.answer(f"⌛ {days} kun sotilmagan tovarlarni qidiryapman..." if lang == "uz" else f"⌛ Ищу товары с остатком, но без продаж за {days} дней...", reply_markup=sales_menu_for_message(message))
    try:
        date_from, date_to = _days_range_ms(days)
        sales_rows, _ = await _load_finance_orders(client, shop_id, date_from_ms=date_from, date_to_ms=date_to, max_pages=5, page_size=100)
        sold_keys: set[str] = set()
        for item in sales_rows:
            if not _is_cancelled_status(_finance_status(item)):
                sold_keys.update(_sale_match_keys(item))
        stock_rows = await load_sku_rows(client, shop_id, max_pages=50)
        candidates: list[dict[str, Any]] = []
        for row in stock_rows:
            total = _stock_row_total(row)
            if total <= 0:
                continue
            keys = _stock_match_keys(row)
            if keys and not keys.intersection(sold_keys):
                price = _stock_row_price(row)
                candidates.append({"row": row, "total": total, "price": price, "value": total * price})
        candidates.sort(key=lambda x: float(x.get("value") or 0), reverse=True)
        if not candidates:
            text = f"🐢 <b>{days} kun sotilmagan tovarlar</b>\n🏪 Do‘kon: <code>{shop_id}</code>\n\nQoldiqda turib sotilmayotgan tovarlar topilmadi." if lang == "uz" else f"🐢 <b>Товары без продаж за {days} дней</b>\n🏪 Магазин: <code>{shop_id}</code>\n\nНе нашёл товаров с остатком и нулевыми продажами."
            await message.answer(text, reply_markup=sales_menu_for_message(message))
            return
        total_value = sum(float(x.get("value") or 0) for x in candidates)
        title = f"🐢 <b>{days} kun sotilmagan tovarlar</b>" if lang == "uz" else f"🐢 <b>Товары без продаж за {days} дней</b>"
        summary = [
            f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
            f"📦 Pozitsiyalar: <b>{len(candidates)}</b>" if lang == "uz" else f"📦 Позиций: <b>{len(candidates)}</b>",
            f"💰 Taxminan muzlagan summa: <b>{_format_money(total_value)}</b>" if lang == "uz" else f"💰 Примерно заморожено: <b>{_format_money(total_value)}</b>",
        ]
        items: list[str] = []
        for idx, item in enumerate(candidates, start=1):
            row = item["row"]
            title_item = escape(_short_text(_stock_row_title(row), 85))
            sku = escape(_short_text(_stock_row_sku(row), 60))
            sku_line = f"\n🔖 SKU: <code>{sku}</code>" if sku else ""
            if lang == "uz":
                items.append(f"{idx}. <b>{title_item}</b>{sku_line}\n📦 Qoldiq: <b>{int(item['total'])} dona</b> | 💵 Narx: {_format_money(float(item['price']))} | 💰 Summa: <b>{_format_money(float(item['value']))}</b>")
            else:
                items.append(f"{idx}. <b>{title_item}</b>{sku_line}\n📦 Остаток: <b>{int(item['total'])} шт.</b> | 💵 Цена: {_format_money(float(item['price']))} | 💰 Сумма: <b>{_format_money(float(item['value']))}</b>")
        await send_paginated_list(message, kind="dead", title=title, summary=summary, items=items, section="sales", reply_markup=sales_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


def build_replenishment_plan(
    stock_rows: list[dict[str, Any]],
    sales_rows: list[dict[str, Any]],
    settings: dict[str, Any],
    *,
    now: datetime | None = None,
) -> list[dict[str, Any]]:
    """Build a seller-facing replenishment plan from 7/14/30-day velocity."""
    now_local = now or datetime.now(UZT)
    if now_local.tzinfo is None:
        now_local = now_local.replace(tzinfo=UZT)
    else:
        now_local = now_local.astimezone(UZT)

    velocity: dict[str, dict[int, float]] = {}
    for item in sales_rows:
        status = _finance_status(item)
        if _is_cancelled_status(status) or _is_returned_status(status):
            continue
        qty = max(0.0, float(_finance_qty(item)))
        if qty <= 0:
            continue
        dt = _finance_datetime_for_report(item)
        if dt is not None:
            dt_local = dt.replace(tzinfo=UZT) if dt.tzinfo is None else dt.astimezone(UZT)
            age_days = max(0.0, (now_local - dt_local).total_seconds() / 86400.0)
        else:
            age_days = None
        windows = (7, 14, 30) if age_days is None else tuple(day for day in (7, 14, 30) if age_days <= day)
        for key in _sale_match_keys(item):
            bucket = velocity.setdefault(key, {7: 0.0, 14: 0.0, 30: 0.0})
            for day in windows:
                bucket[day] += qty

    unique_keys_by_row = _unique_stock_match_keys(stock_rows)

    threshold = max(0, int(settings.get("low_stock_threshold") or LOW_STOCK_THRESHOLD))
    lead_days = max(0, int(settings.get("lead_time_days") or 0))
    safety_days = max(0, int(settings.get("safety_days") or 0))
    target_days = max(1, int(settings.get("target_cover_days") or 30))
    plan: list[dict[str, Any]] = []

    for row, keys in zip(stock_rows, unique_keys_by_row):
        total = max(0, _stock_row_total(row))
        quantities = {
            day: max([velocity.get(key, {}).get(day, 0.0) for key in keys] or [0.0])
            for day in (7, 14, 30)
        }
        avg_7 = quantities[7] / 7.0
        avg_14 = quantities[14] / 14.0
        avg_30 = quantities[30] / 30.0
        available = [(avg_7, 0.5), (avg_14, 0.3), (avg_30, 0.2)]
        positive = [(value, weight) for value, weight in available if value > 0]
        finance_avg_daily = (
            sum(value * weight for value, weight in positive) / sum(weight for _, weight in positive)
            if positive
            else 0.0
        )
        api_avg_daily = max(0.0, float(_num_from_value(row.get("avg_daily_sales_api")) or 0.0))
        if finance_avg_daily > 0:
            avg_daily = finance_avg_daily
            forecast_source = "finance_30d"
        elif api_avg_daily > 0:
            # Official Product API fallback.  It prevents a false "no sales"
            # forecast when Finance API uses a different SKU representation.
            avg_daily = api_avg_daily
            forecast_source = "uzum_avgdsales"
        else:
            avg_daily = 0.0
            forecast_source = "no_sales"
        days_left = (float(total) / avg_daily) if avg_daily > 0 else None
        reorder_in_days = (
            max(0.0, float(days_left) - lead_days - safety_days)
            if days_left is not None
            else None
        )
        stockout_date = now_local + timedelta(days=days_left) if days_left is not None else None
        reorder_date = now_local + timedelta(days=reorder_in_days) if reorder_in_days is not None else None
        target_units = math.ceil(avg_daily * (target_days + lead_days + safety_days)) if avg_daily > 0 else total
        recommended_qty = max(0, target_units - total)
        if avg_daily <= 0 and total <= threshold:
            recommended_qty = 0
        price = max(0.0, _stock_row_price(row))
        risk_days = max(0.0, float(lead_days + safety_days) - float(days_left or 0)) if avg_daily > 0 else 0.0
        if total <= 0 and avg_daily > 0:
            risk_days = max(risk_days, min(7.0, float(target_days)))
        risk_value = avg_daily * risk_days * price
        trend_percent = (
            (avg_7 / avg_30 - 1.0) * 100.0
            if forecast_source == "finance_30d" and avg_7 > 0 and avg_30 > 0
            else None
        )
        urgent = bool(
            total <= 0
            or (days_left is not None and days_left <= lead_days + safety_days)
            or total <= threshold
        )
        if recommended_qty <= 0 and not urgent:
            continue
        plan.append({
            "row": row,
            "title": _stock_row_title(row),
            "sku": _stock_row_sku(row),
            "total": total,
            "sold_7": quantities[7],
            "sold_14": quantities[14],
            "sold_30": quantities[30],
            "avg_daily": avg_daily,
            "finance_avg_daily": finance_avg_daily,
            "api_avg_daily": api_avg_daily,
            "forecast_source": forecast_source,
            "days_left": days_left,
            "stockout_date": stockout_date,
            "reorder_date": reorder_date,
            "recommended_qty": recommended_qty,
            "risk_value": risk_value,
            "trend_percent": trend_percent,
            "urgent": urgent,
            "price": price,
        })

    plan.sort(key=lambda item: (
        0 if item.get("urgent") else 1,
        -float(item.get("risk_value") or 0),
        float(item.get("days_left") if item.get("days_left") is not None else 999999),
    ))
    return plan


def _split_replenishment_plan(
    plan: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Separate real supply actions from low-stock rows without demand."""
    actionable = [
        item
        for item in plan
        if float(item.get("avg_daily") or 0.0) > 0
        and int(item.get("recommended_qty") or 0) > 0
    ]
    no_demand = [
        item
        for item in plan
        if float(item.get("avg_daily") or 0.0) <= 0
    ]
    return actionable, no_demand


@dp.message(Command("smart_lowstock", "forecast_stock", "supply_plan", "replenishment"))
async def smart_lowstock(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    telegram_id = message.from_user.id if message.from_user else 0
    lang = get_user_language(telegram_id)
    await message.answer("⌛ Qoldiq necha kunga yetishini hisoblayapman..." if lang == "uz" else "⌛ Считаю, на сколько дней хватит остатков...", reply_markup=stock_menu_for_message(message))
    try:
        date_from, date_to = _days_range_ms(30)
        sales_rows, _, _ = await _load_finance_range_flexible(client, shop_id, date_from, date_to)
        stock_rows = await load_sku_rows(client, shop_id, max_pages=50)
        settings = ensure_product_settings(telegram_id)
        raw_plan = build_replenishment_plan(stock_rows, sales_rows, settings)
        alerts, no_demand = _split_replenishment_plan(raw_plan)
        lead_days = int(settings.get("lead_time_days") or 0)
        safety_days = int(settings.get("safety_days") or 0)
        target_days = int(settings.get("target_cover_days") or 30)
        urgent_count = sum(
            1
            for item in alerts
            if int(item.get("total") or 0) <= 0
            or (
                item.get("days_left") is not None
                and float(item.get("days_left") or 0) <= lead_days + safety_days
            )
        )
        finance_count = sum(1 for item in alerts if item.get("forecast_source") == "finance_30d")
        api_average_count = sum(1 for item in alerts if item.get("forecast_source") == "uzum_avgdsales")

        if not alerts:
            no_demand_note = (
                f"\n\nℹ️ Past qoldiqli, ammo savdosiz SKU: <b>{len(no_demand)}</b>. "
                "Ularga yetkazib berish tavsiya qilinmaydi."
                if lang == "uz" and no_demand
                else f"\n\nℹ️ SKU с низким остатком, но без спроса: <b>{len(no_demand)}</b>. "
                "Поставлять их сейчас не нужно."
                if no_demand
                else ""
            )
            text = (
                f"✅ <b>Yetkazib berish hozir kerak emas</b>\n"
                f"🏪 Do‘kon: <code>{shop_id}</code>\n\n"
                "Bot savdo tezligi va joriy qoldiq bo‘yicha yetkazib berish zaruratini topmadi."
                f"{no_demand_note}"
                if lang == "uz"
                else f"✅ <b>Поставка сейчас не требуется</b>\n"
                f"🏪 Магазин: <code>{shop_id}</code>\n\n"
                "По скорости продаж и текущему остатку товаров для пополнения не найдено."
                f"{no_demand_note}"
            )
            await message.answer(text, reply_markup=stock_menu_for_message(message))
            return

        title = "🚚 <b>Yetkazib berish rejasi</b>" if lang == "uz" else "🚚 <b>План поставки</b>"
        summary = [
            f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
            (
                f"🔴 Hozir buyurtma berish: <b>{urgent_count}</b> | "
                f"Jami yetkazish: <b>{len(alerts)}</b> SKU"
                if lang == "uz"
                else f"🔴 Заказать сейчас: <b>{urgent_count}</b> | "
                f"Всего к поставке: <b>{len(alerts)}</b> SKU"
            ),
            (
                f"Hisob: yetkazish {lead_days} kun + xavfsizlik {safety_days} kun; "
                f"yetkazilgandan keyin {target_days} kunlik zaxira."
                if lang == "uz"
                else f"Расчёт: доставка {lead_days} дн. + страховка {safety_days} дн.; "
                f"после поставки — запас на {target_days} дн."
            ),
            (
                f"Manba: Finance bo‘yicha {finance_count}, Uzum o‘rtacha tezligi bo‘yicha {api_average_count}."
                if lang == "uz"
                else f"Источник скорости: Finance — {finance_count}, средняя Uzum — {api_average_count}."
            ),
            (
                f"ℹ️ Savdosiz past qoldiq: {len(no_demand)} SKU — yetkazishga kiritilmadi."
                if lang == "uz" and no_demand
                else f"ℹ️ Без спроса: {len(no_demand)} SKU — в поставку не включены."
                if no_demand
                else ""
            ),
        ]
        items: list[str] = []
        for idx, item in enumerate(alerts, start=1):
            title_item = escape(_short_text(item.get("title"), 85))
            sku = escape(_short_text(item.get("sku"), 60))
            days_left = item.get("days_left")
            stockout = item.get("stockout_date")
            reorder = item.get("reorder_date")
            stockout_text = stockout.strftime("%d.%m.%Y") if isinstance(stockout, datetime) else "—"
            reorder_now = (
                not isinstance(reorder, datetime)
                or reorder.date() <= datetime.now(UZT).date()
            )
            reorder_text = (
                ("hozir" if lang == "uz" else "сейчас")
                if reorder_now
                else reorder.strftime("%d.%m.%Y")
            )
            trend = item.get("trend_percent")
            trend_line = ""
            if trend is not None and abs(float(trend)) >= 5:
                if lang == "uz":
                    trend_line = f"\n📈 So‘nggi trend: <b>{float(trend):+.0f}%</b>"
                else:
                    direction = "рост" if float(trend) > 0 else "снижение"
                    trend_line = f"\n📈 Тренд спроса: <b>{direction} {abs(float(trend)):.0f}%</b>"
            row = item.get("row") if isinstance(item.get("row"), dict) else {}
            fbo = max(0, int(_num_from_value(row.get("fbo")) or 0))
            fbs = max(0, int(_num_from_value(row.get("fbs")) or 0))
            source = item.get("forecast_source")
            risk_value = float(item.get("risk_value") or 0)
            risk_line = ""
            if risk_value > 0:
                risk_line = (
                    f"\n💸 Sotuv xavfi: <b>{_format_money(risk_value)}</b>"
                    if lang == "uz"
                    else f"\n💸 Возможная упущенная выручка: <b>{_format_money(risk_value)}</b>"
                )
            if lang == "uz":
                days_text = f"taxminan {float(days_left):.1f} kun"
                sku_line = f"\n🔖 SKU: <code>{sku}</code>" if sku else ""
                status_text = (
                    "🔴 Qoldiq tugagan"
                    if int(item.get("total") or 0) <= 0
                    else "🔴 Hozir buyurtma bering"
                    if float(days_left or 0) <= lead_days + safety_days
                    else "🟡 Yetkazib berishni rejalashtiring"
                )
                speed_text = (
                    f"{float(item['avg_daily']):.2f} dona/kun (Uzum o‘rtachasi)"
                    if source == "uzum_avgdsales"
                    else f"{float(item['avg_daily']):.2f} dona/kun; 7 kunda {float(item.get('sold_7') or 0):.0f}, 30 kunda {float(item.get('sold_30') or 0):.0f}"
                )
                items.append(
                    f"{idx}. {status_text}\n<b>{title_item}</b>{sku_line}\n"
                    f"📦 Hozir: <b>{int(item['total'])} dona</b> (Uzum {fbo} + FBS {fbs})\n"
                    f"🛒 Savdo tezligi: <b>{speed_text}</b>\n"
                    f"⏳ {days_text} yetadi — <b>{stockout_text}</b> gacha\n"
                    f"✅ Amal: <b>{int(item['recommended_qty'])} dona</b> {reorder_text} buyurtma qiling"
                    f"{trend_line}{risk_line}"
                )
            else:
                days_text = f"примерно на {float(days_left):.1f} дн."
                sku_line = f"\n🔖 SKU: <code>{sku}</code>" if sku else ""
                status_text = (
                    "🔴 Товар уже закончился"
                    if int(item.get("total") or 0) <= 0
                    else "🔴 Заказать сейчас"
                    if float(days_left or 0) <= lead_days + safety_days
                    else "🟡 Запланировать поставку"
                )
                speed_text = (
                    f"{float(item['avg_daily']):.2f} шт./день (средняя Uzum)"
                    if source == "uzum_avgdsales"
                    else f"{float(item['avg_daily']):.2f} шт./день; продано 7 дн.: {float(item.get('sold_7') or 0):.0f}, 30 дн.: {float(item.get('sold_30') or 0):.0f}"
                )
                items.append(
                    f"{idx}. {status_text}\n<b>{title_item}</b>{sku_line}\n"
                    f"📦 Сейчас: <b>{int(item['total'])} шт.</b> (Uzum {fbo} + FBS {fbs})\n"
                    f"🛒 Скорость продаж: <b>{speed_text}</b>\n"
                    f"⏳ Хватит {days_text} — до <b>{stockout_text}</b>\n"
                    f"✅ Действие: заказать <b>{int(item['recommended_qty'])} шт.</b> — <b>{reorder_text}</b>"
                    f"{trend_line}{risk_line}"
                )
        await send_paginated_list(message, kind="forecast", title=title, summary=summary, items=items, section="stock", reply_markup=stock_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


def _finance_rows_in_window(
    rows: list[dict[str, Any]],
    start: datetime,
    end: datetime,
    *,
    issued: bool = False,
) -> list[dict[str, Any]]:
    start_naive = start.astimezone(UZT).replace(tzinfo=None) if start.tzinfo else start
    end_naive = end.astimezone(UZT).replace(tzinfo=None) if end.tzinfo else end
    result: list[dict[str, Any]] = []
    for item in rows:
        value = _finance_issued_datetime(item) if issued else _finance_order_datetime(item)
        if value is not None and start_naive <= value < end_naive:
            result.append(item)
    return result


def _expense_report_parts(summary: dict[str, Any]) -> dict[str, Any]:
    additional = 0.0
    refunds = 0.0
    ledger_deductions = 0.0
    ledger_refunds = 0.0
    categories: dict[str, float] = {}
    names: dict[str, float] = {}
    for item in summary.get("rows") or []:
        signed = float(item.get("signed_amount") or 0)
        category = str(item.get("category") or "other")
        categories[category] = categories.get(category, 0.0) + signed
        name = " ".join(str(item.get("name") or category or "Uzum").split())
        names[name] = names.get(name, 0.0) + signed
        if signed >= 0:
            ledger_deductions += signed
            if bool(item.get("included_in_profit")):
                additional += signed
        else:
            ledger_refunds += abs(signed)
            if bool(item.get("included_in_profit")):
                refunds += abs(signed)
    return {
        "additional_expenses": additional,
        "refunds": refunds,
        "ledger_deductions": ledger_deductions,
        "ledger_refunds": ledger_refunds,
        "expense_categories": categories,
        "expense_names": names,
    }


async def _daily_fbs_scheme_index(
    client: UzumClient,
    shop_id: int,
    start: datetime,
    end: datetime,
) -> dict[str, dict[str, Any]]:
    """Best-effort exact FBS/DBS mapping without per-order API spam."""
    result: dict[str, dict[str, Any]] = {}
    max_pages = max(1, min(20, int(os.getenv("DAILY_FBS_SCHEME_MAX_PAGES", "6") or "6")))
    for status in ("COMPLETED", "CANCELED", "RETURNED"):
        try:
            for page in range(max_pages):
                data = await client.get_fbs_orders(
                    shop_id,
                    status=status,
                    page=page,
                    size=50,
                    date_from=int(start.timestamp() * 1000),
                    date_to=int(end.timestamp() * 1000),
                )
                items = [item for item in _extract_list_any(data) if isinstance(item, dict)]
                for item in items:
                    order_id = str(
                        _deep_pick_value(item, ("orderId", "id", "orderNumber")) or ""
                    ).strip()
                    if order_id:
                        result[order_id] = item
                if len(items) < 50:
                    break
                await asyncio.sleep(0.04)
        except Exception as error:
            logging.info(
                "Daily FBS/DBS scheme mapping unavailable shop=%s status=%s: %s",
                shop_id,
                status,
                str(error)[:160],
            )
    return result


def _market_daily_shop_payload(
    rows: list[dict[str, Any]],
    *,
    start: datetime,
    end: datetime,
    costs: dict[str, dict[str, Any]],
    settings: dict[str, Any],
    expense_summary: dict[str, Any],
    shop_id: int,
    shop_name: str = "",
) -> dict[str, Any]:
    accepted_rows = _finance_rows_in_window(rows, start, end)
    issued_rows = _finance_rows_in_window(rows, start, end, issued=True)
    previous_issued_rows = _finance_rows_in_window(
        rows,
        start - timedelta(days=1),
        start,
        issued=True,
    )
    tax_percent = float(settings.get("tax_percent") or 0)
    accepted_stats = _build_noorza_today_stats(accepted_rows)
    issued_stats = _build_noorza_today_stats(issued_rows)
    previous_issued_stats = _build_noorza_today_stats(previous_issued_rows)
    products = _build_unit_rows_from_finance(
        accepted_rows,
        costs,
        tax_percent=tax_percent,
    )
    for product in products:
        product["shop_id"] = int(shop_id)
        product["shop_name"] = str(shop_name or "")
    profit = _profit_summary_from_unit_rows(products, accepted_stats)

    cancellation_orders: set[str] = set()
    cancellation_units = 0.0
    cancellation_value = 0.0
    cancellation_reasons: dict[str, int] = {}
    for item in accepted_rows:
        cancelled_qty = _finance_cancelled_qty(item)
        if cancelled_qty <= 0 and not _has_cancel_event_status(_finance_status(item)):
            continue
        cancellation_orders.add(_finance_order_key_for_stats(item))
        cancellation_units += cancelled_qty
        cancellation_value += _finance_cancelled_revenue(item, cancelled_qty)
        reason = _finance_cancel_reason(item) or "UZUM_REASON_NOT_PROVIDED"
        cancellation_reasons[reason] = cancellation_reasons.get(reason, 0) + 1

    order_schemes: dict[str, str] = {}
    for item in accepted_rows:
        if _finance_qty(item) <= 0:
            continue
        order_id = _finance_order_key_for_stats(item)
        scheme = _finance_scheme(item)
        current = order_schemes.get(order_id, "—")
        if current == "—" or scheme != "—":
            order_schemes[order_id] = scheme
    schemes = {"FBO": 0, "FBS": 0, "DBS": 0, "UNKNOWN": 0}
    for scheme in order_schemes.values():
        key = scheme if scheme in {"FBO", "FBS", "DBS"} else "UNKNOWN"
        schemes[key] += 1

    expense_parts = _expense_report_parts(expense_summary)
    business = calculate_business_profit(
        profit,
        accepted_stats,
        settings,
        days=1,
        uzum_expenses=expense_summary,
    )
    external = float(business.get("external_expense_total") or 0)
    known_net = sum(
        float(item.get("net_profit") or 0)
        for item in products
        if item.get("net_profit") is not None
    )
    result = float(business.get("net_profit") or 0)
    revenue = float(accepted_stats.get("revenue") or 0)
    coverage = float(profit.get("coverage") or 0)
    complete = bool(
        (coverage >= 0.999 or revenue <= 0)
        and expense_summary.get("available")
    )
    return {
        "shop_id": int(shop_id),
        "shop_name": str(shop_name or ""),
        "accepted": accepted_stats,
        "issued": issued_stats,
        "previous_issued": previous_issued_stats,
        "products": products,
        "profit": profit,
        "business": business,
        "tax_percent": tax_percent,
        "known_net_product_profit": known_net,
        "result": result,
        "cost_coverage": coverage,
        "complete": complete,
        "cancellations": len({value for value in cancellation_orders if value not in {"", "-", "—"}}),
        "cancellation_units": cancellation_units,
        "cancellation_value": cancellation_value,
        "cancellation_reasons": cancellation_reasons,
        "schemes": schemes,
        "external_expenses": external,
        "expenses_available": bool(expense_summary.get("available")),
        **expense_parts,
    }


async def _collect_market_daily_report(
    telegram_id: int,
    client: UzumClient,
    report_date: date,
) -> dict[str, Any]:
    start = datetime(report_date.year, report_date.month, report_date.day, tzinfo=UZT)
    end = start + timedelta(days=1)
    # Issued orders can have been accepted earlier.  The lookback keeps those
    # rows available before the exact `dateIssued` filter is applied locally.
    query_start = start - timedelta(days=60)
    shops = await _user_shop_list(telegram_id, client)
    if not shops:
        default_shop = db.get_default_shop_id(telegram_id)
        if default_shop:
            shops = [{"shop_id": int(default_shop), "name": "", "raw": {}}]
    per_shop: list[dict[str, Any]] = []
    errors: list[str] = []
    for shop in shops:
        shop_id = int(shop["shop_id"])
        try:
            rows, _, _ = await _load_finance_range_flexible(
                client,
                shop_id,
                int(query_start.timestamp() * 1000),
                int(end.timestamp() * 1000),
            )
            scheme_index = await _daily_fbs_scheme_index(client, shop_id, start, end)
            if scheme_index:
                rows = [
                    {
                        **item,
                        "_fbs_order_detail": scheme_index.get(_finance_order_id(item)),
                    }
                    if _finance_order_id(item) in scheme_index
                    else item
                    for item in rows
                ]
            try:
                await sync_uzum_sku_financials(client, telegram_id, shop_id)
            except Exception:
                # Historical Finance purchasePrice remains usable.  The exact
                # coverage below will reveal any SKU that still lacks a cost.
                logging.exception("Daily report cost fallback sync failed shop=%s", shop_id)
            costs = get_unit_cost_map(telegram_id, shop_id)
            settings = ensure_finance_settings(telegram_id, shop_id)
            expenses = await load_uzum_expense_summary(
                client,
                shop_id,
                int(start.timestamp() * 1000),
                int(end.timestamp() * 1000),
            )
            per_shop.append(
                _market_daily_shop_payload(
                    rows,
                    start=start,
                    end=end,
                    costs=costs,
                    settings=settings,
                    expense_summary=expenses,
                    shop_id=shop_id,
                    shop_name=str(shop.get("name") or ""),
                )
            )
        except Exception as error:
            errors.append(f"{shop_id}: {str(error)[:160]}")
            logging.exception("Market-style daily report failed shop=%s", shop_id)
        await asyncio.sleep(0.15)

    accepted = _merge_noorza_stats([dict(item.get("accepted") or {}) for item in per_shop])
    issued = _merge_noorza_stats([dict(item.get("issued") or {}) for item in per_shop])
    previous_issued = _merge_noorza_stats(
        [dict(item.get("previous_issued") or {}) for item in per_shop]
    )
    products = [product for item in per_shop for product in item.get("products") or []]
    products.sort(key=lambda item: float(item.get("revenue") or 0), reverse=True)
    total_revenue = float(accepted.get("revenue") or 0)
    known_revenue = sum(float((item.get("profit") or {}).get("known_revenue") or 0) for item in per_shop)
    coverage = known_revenue / total_revenue if total_revenue > 0 else 0.0
    schemes = {"FBO": 0, "FBS": 0, "DBS": 0, "UNKNOWN": 0}
    reasons: dict[str, int] = {}
    categories: dict[str, float] = {}
    expense_names: dict[str, float] = {}
    for item in per_shop:
        for key in schemes:
            schemes[key] += int((item.get("schemes") or {}).get(key) or 0)
        for key, count in (item.get("cancellation_reasons") or {}).items():
            reasons[str(key)] = reasons.get(str(key), 0) + int(count or 0)
        for key, amount in (item.get("expense_categories") or {}).items():
            categories[str(key)] = categories.get(str(key), 0.0) + float(amount or 0)
        for key, amount in (item.get("expense_names") or {}).items():
            expense_names[str(key)] = expense_names.get(str(key), 0.0) + float(amount or 0)
    complete = bool(per_shop) and not errors and all(bool(item.get("complete")) for item in per_shop)
    business_fields = (
        "calculation_revenue",
        "calculation_payout",
        "calculation_commission",
        "calculation_logistics",
        "other_payout_deductions",
        "payout_adjustment",
        "cost_total",
        "known_profit",
        "tax_expense",
        "uzum_expense_deductions",
        "uzum_expense_refunds",
        "uzum_expense_total",
        "external_expense_total",
        "net_profit",
    )
    aggregate_business = {
        field: sum(
            float((item.get("business") or {}).get(field) or 0)
            for item in per_shop
        )
        for field in business_fields
    }
    aggregate_business.update(
        {
            "complete": complete,
            "coverage": max(0.0, min(1.0, coverage)),
            "missing_count": sum(
                int((item.get("business") or {}).get("missing_count") or 0)
                for item in per_shop
            ),
            "uzum_expenses_available": bool(per_shop)
            and all(
                bool((item.get("business") or {}).get("uzum_expenses_available"))
                for item in per_shop
            ),
        }
    )
    return {
        "date": report_date,
        "shop_id": per_shop[0]["shop_id"] if len(per_shop) == 1 else None,
        "shop_label": (
            str(per_shop[0].get("shop_name") or per_shop[0]["shop_id"])
            if len(per_shop) == 1
            else f"{len(per_shop)} shops"
        ),
        "shops_count": len(shops),
        "successful_shops": len(per_shop),
        "accepted": accepted,
        "issued": issued,
        "previous_issued": previous_issued,
        "products": products,
        "per_shop": per_shop,
        "cost_coverage": max(0.0, min(1.0, coverage)),
        "complete": complete,
        "business": aggregate_business,
        "result": sum(float(item.get("result") or 0) for item in per_shop),
        "known_net_product_profit": sum(float(item.get("known_net_product_profit") or 0) for item in per_shop),
        "additional_expenses": sum(float(item.get("additional_expenses") or 0) for item in per_shop),
        "refunds": sum(float(item.get("refunds") or 0) for item in per_shop),
        "external_expenses": sum(float(item.get("external_expenses") or 0) for item in per_shop),
        "ledger_deductions": sum(float(item.get("ledger_deductions") or 0) for item in per_shop),
        "ledger_refunds": sum(float(item.get("ledger_refunds") or 0) for item in per_shop),
        "cancellations": sum(int(item.get("cancellations") or 0) for item in per_shop),
        "cancellation_units": sum(float(item.get("cancellation_units") or 0) for item in per_shop),
        "cancellation_value": sum(float(item.get("cancellation_value") or 0) for item in per_shop),
        "cancellation_reasons": reasons,
        "schemes": schemes,
        "expense_categories": categories,
        "expense_names": expense_names,
        "errors": errors,
    }


def _format_market_daily_report(payload: dict[str, Any], *, lang: str) -> str:
    uz = normalize_lang(lang) == "uz"
    accepted = dict(payload.get("accepted") or {})
    issued = dict(payload.get("issued") or {})
    previous_issued = dict(payload.get("previous_issued") or {})
    complete = bool(payload.get("complete"))
    business = dict(payload.get("business") or {})
    if uz:
        lines = [
            "🟣 <b>Seller.pro.uz · kunlik hisobot</b>",
            f"📅 <b>{payload.get('date').strftime('%d.%m.%Y')}</b> · 🏪 {int(payload.get('shops_count') or 0)} ta do‘kon",
            "",
            f"📥 <b>Qabul qilingan buyurtmalar: {int(accepted.get('orders') or 0)}</b>",
            f"🛍 Sotilgan tovarlar: <b>{float(accepted.get('units') or 0):.0f} dona</b>",
            "",
            *_format_profit_bridge_lines(business, lang=lang),
            "",
            f"📤 <b>Olib ketilgan buyurtmalar: {int(issued.get('orders') or 0)}</b> <i>(ma’lumot uchun)</i>",
            f"🛍 Sotilgan tovarlar: <b>{float(issued.get('units') or 0):.0f} dona</b>",
            f"💰 Daromad: <b>{_format_money(float(issued.get('revenue') or 0))}</b>",
            f"💳 Chiqarishga: <b>{_format_money(float(issued.get('payout_total') or 0))}</b>",
            "",
            f"❌ Bekor qilingan buyurtmalar: <b>{int(payload.get('cancellations') or 0)}</b>",
            f"💸 Bekor qilingan summa: <b>{_format_money(float(payload.get('cancellation_value') or 0))}</b>",
        ]
    else:
        lines = [
            "🟣 <b>Seller.pro.uz · дневной отчёт</b>",
            f"📅 <b>{payload.get('date').strftime('%d.%m.%Y')}</b> · 🏪 магазинов: {int(payload.get('shops_count') or 0)}",
            "",
            f"📥 <b>Принято заказов: {int(accepted.get('orders') or 0)}</b>",
            f"🛍 Продано товаров: <b>{float(accepted.get('units') or 0):.0f} шт.</b>",
            "",
            *_format_profit_bridge_lines(business, lang=lang),
            "",
            f"📤 <b>Выдано заказов: {int(issued.get('orders') or 0)}</b> <i>(справочно, второй раз в прибыль не входит)</i>",
            f"🛍 Выдано товаров: <b>{float(issued.get('units') or 0):.0f} шт.</b>",
            f"💰 Выручка: <b>{_format_money(float(issued.get('revenue') or 0))}</b>",
            f"💳 К выплате: <b>{_format_money(float(issued.get('payout_total') or 0))}</b>",
            "",
            f"❌ Отменено заказов: <b>{int(payload.get('cancellations') or 0)}</b>",
            f"💸 Сумма отмен: <b>{_format_money(float(payload.get('cancellation_value') or 0))}</b>",
        ]

    report_date = payload.get("date")
    previous_date = report_date - timedelta(days=1)
    today = datetime.now(UZT).date()
    current_payout = float(issued.get("payout_total") or 0)
    previous_payout = float(previous_issued.get("payout_total") or 0)
    if report_date == today:
        if uz:
            lines.extend([
                "",
                f"💵 Kecha chiqarilgan buyurtmalar bo‘yicha taxminiy to‘lov: <b>{_format_money(previous_payout)}</b> · {int(previous_issued.get('orders') or 0)} ta buyurtma",
                f"💵 Bugun chiqarilgan buyurtmalar bo‘yicha taxminiy to‘lov: <b>{_format_money(current_payout)}</b> · {int(issued.get('orders') or 0)} ta buyurtma",
            ])
        else:
            lines.extend([
                "",
                f"💵 Вчера ориентировочно к выплате по выданным заказам: <b>{_format_money(previous_payout)}</b> · заказов: {int(previous_issued.get('orders') or 0)}",
                f"💵 Сегодня ориентировочно к выплате по выданным заказам: <b>{_format_money(current_payout)}</b> · заказов: {int(issued.get('orders') or 0)}",
            ])
    else:
        if uz:
            lines.extend([
                "",
                f"💵 {previous_date.strftime('%d.%m.%Y')} uchun taxminiy to‘lov: <b>{_format_money(previous_payout)}</b> · {int(previous_issued.get('orders') or 0)} ta buyurtma",
                f"💵 {report_date.strftime('%d.%m.%Y')} uchun taxminiy to‘lov: <b>{_format_money(current_payout)}</b> · {int(issued.get('orders') or 0)} ta buyurtma",
            ])
        else:
            lines.extend([
                "",
                f"💵 Ориентировочно к выплате за {previous_date.strftime('%d.%m.%Y')}: <b>{_format_money(previous_payout)}</b> · заказов: {int(previous_issued.get('orders') or 0)}",
                f"💵 Ориентировочно к выплате за {report_date.strftime('%d.%m.%Y')}: <b>{_format_money(current_payout)}</b> · заказов: {int(issued.get('orders') or 0)}",
            ])

    reasons = sorted(
        (payload.get("cancellation_reasons") or {}).items(),
        key=lambda pair: int(pair[1] or 0),
        reverse=True,
    )
    if reasons:
        lines.append("💬 <b>Sabablar:</b>" if uz else "💬 <b>Причины:</b>")
        for raw, count in reasons[:5]:
            reason_item = {"cancelReason": "" if raw == "UZUM_REASON_NOT_PROVIDED" else raw}
            label = _format_cancel_reason(reason_item, lang=lang)
            lines.append(f"• {escape(label)}: <b>{int(count or 0)}</b>")

    schemes = dict(payload.get("schemes") or {})
    lines.extend([
        "",
        "🚚 <b>FBO / FBS bo‘yicha:</b>" if uz else "🚚 <b>По схемам FBO / FBS:</b>",
        f"• FBO: <b>{int(schemes.get('FBO') or 0)}</b> · FBS: <b>{int(schemes.get('FBS') or 0)}</b> · DBS: <b>{int(schemes.get('DBS') or 0)}</b>",
    ])
    if int(schemes.get("UNKNOWN") or 0):
        lines.append(
            f"• Uzum sxemani bermadi: <b>{int(schemes['UNKNOWN'])}</b>"
            if uz
            else f"• Uzum не передал схему: <b>{int(schemes['UNKNOWN'])}</b>"
        )

    lines.extend([
        "",
        f"➖ Kecha yechib olindi: <b>{_format_money(float(payload.get('ledger_deductions') or 0))}</b>" if uz else f"➖ Списано за день: <b>{_format_money(float(payload.get('ledger_deductions') or 0))}</b>",
    ])
    expense_names = sorted(
        (payload.get("expense_names") or {}).items(),
        key=lambda pair: abs(float(pair[1] or 0)),
        reverse=True,
    )
    for name, amount in expense_names[:8]:
        amount_value = float(amount or 0)
        if amount_value > 0:
            lines.append(f"  ➤ {escape(_short_text(name, 55))}: {_format_money(amount_value)}")
    lines.append(
        f"➕ Kecha qaytarildi: <b>{_format_money(float(payload.get('ledger_refunds') or 0))}</b>"
        if uz
        else f"➕ Возвращено за день: <b>{_format_money(float(payload.get('ledger_refunds') or 0))}</b>"
    )
    for name, amount in expense_names[:8]:
        amount_value = float(amount or 0)
        if amount_value < 0:
            lines.append(f"  ➤ {escape(_short_text(name, 55))}: {_format_money(abs(amount_value))}")
    lines.append(
        f"📌 Tannarx qamrovi: <b>{float(payload.get('cost_coverage') or 0) * 100:.1f}%</b>"
        if uz
        else f"📌 Покрытие себестоимостью: <b>{float(payload.get('cost_coverage') or 0) * 100:.1f}%</b>"
    )
    if not complete:
        lines.append(
            "⚠️ Bu yakuniy sof foyda emas: ayrim purchasePrice/xarajat ma’lumotlari yetishmaydi."
            if uz
            else "⚠️ Это не окончательная чистая прибыль: не хватает части purchasePrice/расходов."
        )
    if payload.get("errors"):
        lines.append(
            f"⚠️ API xatosi bo‘lgan do‘konlar: {len(payload['errors'])}."
            if uz
            else f"⚠️ Магазинов с ошибкой API: {len(payload['errors'])}."
        )
    return "\n".join(lines)


async def _build_morning_report_text(telegram_id: int, client: UzumClient) -> str:
    report_date = (datetime.now(UZT) - timedelta(days=1)).date()
    payload = await _collect_market_daily_report(telegram_id, client, report_date)
    return _format_market_daily_report(payload, lang=get_user_language(telegram_id))


@dp.message(Command("morning_report", "daily_report"))
async def morning_report(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not await require_active_subscription(message, telegram_id):
        return
    client = get_uzum_for_user(telegram_id)
    if client is None:
        await message.answer("Сначала подключите Uzum API-токен: <code>/connect</code>", reply_markup=menu_for_message(message))
        return
    await message.answer("⌛ Готовлю утренний отчёт за вчера...", reply_markup=menu_for_message(message))
    try:
        await message.answer(await _build_morning_report_text(telegram_id, client), reply_markup=menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


def _requested_daily_report_date(text: str, *, now: datetime | None = None) -> date:
    today = (now or datetime.now(UZT)).astimezone(UZT).date()
    argument = parse_args(text or "").strip().lower()
    if not argument or argument in {"today", "bugun", "сегодня"}:
        return today
    if argument in {"yesterday", "kecha", "вчера"}:
        return today - timedelta(days=1)
    try:
        selected = datetime.strptime(argument, "%Y-%m-%d").date()
    except ValueError as error:
        raise ValueError("Используйте дату в формате YYYY-MM-DD, today или yesterday.") from error
    if selected > today:
        raise ValueError("Нельзя сформировать отчёт за будущую дату.")
    if selected < today - timedelta(days=60):
        raise ValueError("Доступен период не старше 60 дней.")
    return selected


@dp.message(Command("market_report", "daily_finance_report", "seller_daily"))
@dp.message(F.text.in_({"📋 Дневной отчёт", "📋 Kunlik hisobot"}))
async def market_style_daily_report(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not await require_active_subscription(message, telegram_id):
        return
    client = get_uzum_for_user(telegram_id)
    if client is None:
        await message.answer(
            "Avval Uzum API tokenini ulang: <code>/connect</code>"
            if get_user_language(telegram_id) == "uz"
            else "Сначала подключите Uzum API-токен: <code>/connect</code>",
            reply_markup=report_menu_for_message(message),
        )
        return
    lang = get_user_language(telegram_id)
    try:
        report_date = _requested_daily_report_date(str(message.text or ""))
    except ValueError as error:
        await message.answer(
            "Sana: <code>/market_report 2026-07-21</code>, <code>today</code> yoki <code>yesterday</code>."
            if lang == "uz"
            else f"{escape(str(error))}\nПример: <code>/market_report 2026-07-21</code>.",
            reply_markup=report_menu_for_message(message),
        )
        return

    await message.answer(
        f"⌛ {report_date.strftime('%d.%m.%Y')} uchun kunlik hisobot, PDF va Excel tayyorlanmoqda..."
        if lang == "uz"
        else f"⌛ Готовлю дневной отчёт, PDF и Excel за {report_date.strftime('%d.%m.%Y')}...",
        reply_markup=report_menu_for_message(message),
    )
    files: list[Path] = []
    try:
        payload = await _collect_market_daily_report(telegram_id, client, report_date)
        await message.answer(
            _format_market_daily_report(payload, lang=lang),
            reply_markup=report_menu_for_message(message),
        )
        stamp = report_date.strftime("%Y-%m-%d")
        xlsx = await asyncio.to_thread(
            build_market_daily_workbook,
            payload,
            Path(tempfile.gettempdir()) / f"sellerpro_daily_{stamp}.xlsx",
            lang=lang,
        )
        pdf = await asyncio.to_thread(
            build_market_daily_pdf,
            payload,
            Path(tempfile.gettempdir()) / f"sellerpro_daily_{stamp}.pdf",
            lang=lang,
        )
        files.extend([Path(pdf), Path(xlsx)])
        await message.answer_document(
            FSInputFile(str(pdf), filename=f"sellerpro_daily_{stamp}.pdf"),
            caption=("📄 Tovarlar bo‘yicha foyda va ROI" if lang == "uz" else "📄 Прибыль и ROI по товарам"),
            reply_markup=report_menu_for_message(message),
        )
        await message.answer_document(
            FSInputFile(str(xlsx), filename=f"sellerpro_daily_{stamp}.xlsx"),
            caption=("📊 Tekshiriladigan Excel hisoboti" if lang == "uz" else "📊 Проверяемый Excel-отчёт"),
            reply_markup=report_menu_for_message(message),
        )
    except Exception as error:
        await send_api_error(message, error)
    finally:
        for path in files:
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass


@dp.message(Command("extend1", "extend_month"))
async def admin_extend_1_month(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "").split()
    if not arg or not arg[0].isdigit():
        await message.answer("Напишите так: <code>/extend1 TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg[0])
    new_until = extend_subscription_days(target, 30)
    await message.answer(f"✅ Продлено на 1 месяц для <code>{target}</code>. До: <b>{_fmt_dt(new_until)}</b>", reply_markup=menu_for_message(message))


@dp.message(Command("extend3"))
async def admin_extend_3_months(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "").split()
    if not arg or not arg[0].isdigit():
        await message.answer("Напишите так: <code>/extend3 TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg[0])
    new_until = extend_subscription_days(target, 90)
    await message.answer(f"✅ Продлено на 3 месяца для <code>{target}</code>. До: <b>{_fmt_dt(new_until)}</b>", reply_markup=menu_for_message(message))


@dp.message(Command("extend6"))
async def admin_extend_6_months(message: Message) -> None:
    admin_id = upsert_from_message(message)
    if not admin_only(admin_id):
        return
    arg = parse_args(message.text or "").split()
    if not arg or not arg[0].isdigit():
        await message.answer("Напишите так: <code>/extend6 TELEGRAM_ID</code>", reply_markup=menu_for_message(message))
        return
    target = int(arg[0])
    new_until = extend_subscription_days(target, 180)
    await message.answer(f"✅ Продлено на 6 месяцев для <code>{target}</code>. До: <b>{_fmt_dt(new_until)}</b>", reply_markup=menu_for_message(message))




# --- Юнит-экономика ---
def _unit_group_key(item: dict[str, Any]) -> str:
    sku = _finance_sku_title(item)
    if sku and sku != "-":
        return _unit_sku_key(sku)
    return _unit_sku_key(_finance_title(item))


def _unit_cost_lookup(costs: dict[str, dict[str, Any]], item: dict[str, Any]) -> float | None:
    for key in (_finance_sku_title(item), _finance_title(item), str(_deep_pick_value(item, ("skuId", "sku", "barcode", "offerId")) or "")):
        sku_key = _unit_sku_key(key)
        if sku_key and sku_key in costs:
            try:
                return float(costs[sku_key].get("cost") or 0)
            except Exception:
                return None
    return None


def _build_unit_rows_from_finance(
    rows: list[dict[str, Any]],
    costs: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Aggregate finance rows by SKU and attach saved seller costs."""
    groups: dict[str, dict[str, Any]] = {}
    for item in rows:
        status = _finance_status(item)
        if _is_cancelled_status(status) or _is_returned_status(status):
            continue
        key = _unit_group_key(item)
        if not key:
            continue
        qty = _finance_qty(item)
        revenue = _finance_gross_revenue(item)
        commission = _finance_commission(item)
        logistics = _finance_logistics(item)
        payout_direct = _finance_payout_direct(item)
        payout = payout_direct if payout_direct is not None else max(0.0, revenue - commission - logistics)
        cost_per_unit = _unit_cost_lookup(costs, item)
        entry = groups.setdefault(key, {
            "sku": _finance_sku_title(item),
            "title": _finance_title(item),
            "qty": 0.0,
            "revenue": 0.0,
            "commission": 0.0,
            "logistics": 0.0,
            "payout": 0.0,
            "cost_per_unit": cost_per_unit,
            "cost_total": 0.0,
            "profit": None,
        })
        if entry.get("cost_per_unit") is None and cost_per_unit is not None:
            entry["cost_per_unit"] = cost_per_unit
        entry["qty"] += qty
        entry["revenue"] += revenue
        entry["commission"] += commission
        entry["logistics"] += logistics
        entry["payout"] += max(0.0, payout)
        if cost_per_unit is not None:
            entry["cost_total"] += float(cost_per_unit) * qty
    for entry in groups.values():
        if entry.get("cost_per_unit") is not None:
            entry["profit"] = float(entry.get("payout") or 0) - float(entry.get("cost_total") or 0)
            revenue = float(entry.get("revenue") or 0)
            entry["margin"] = (float(entry["profit"]) / revenue * 100.0) if revenue > 0 else 0.0
        else:
            entry["margin"] = None
    return sorted(groups.values(), key=lambda x: float(x.get("revenue") or 0), reverse=True)


async def _unit_economy_for_shop(client: UzumClient, telegram_id: int, shop_id: int, days: int = 30) -> tuple[list[dict[str, Any]], dict[str, Any], int]:
    date_from, date_to = _days_range_ms(days)
    rows, _, _ = await _load_finance_range_flexible(
        client,
        shop_id,
        date_from,
        date_to,
    )
    cost_status = await sync_uzum_sku_financials(client, telegram_id, shop_id)
    costs = get_unit_cost_map(telegram_id, shop_id)
    finance_settings = ensure_finance_settings(telegram_id, shop_id)
    top = _build_unit_rows_from_finance(
        rows,
        costs,
        tax_percent=float(finance_settings.get("tax_percent") or 0),
    )
    return top, _build_noorza_today_stats(rows), int(cost_status.get("with_cost") or 0)


def _format_unit_economy(shop_id: int, days: int, rows: list[dict[str, Any]], stats: dict[str, Any], saved_costs: int, lang: str = "ru") -> str:
    lang = normalize_lang(lang)
    if lang == "uz":
        if not rows:
            return (
                f"🧾 <b>Unit iqtisodiyot</b>\n🏪 Do‘kon: <code>{shop_id}</code>\n\n"
                "Savdolar topilmadi. Avval 30 kunlik savdolar bo‘yicha ma’lumot bo‘lishi kerak."
            )
        total_known_profit = sum(float(x.get("profit") or 0) for x in rows if x.get("profit") is not None)
        known_items = sum(1 for x in rows if x.get("profit") is not None)
        lines = [
            f"🧾 <b>Unit iqtisodiyot — {days} kun</b>",
            f"🏪 Do‘kon: <code>{shop_id}</code>",
            f"📦 Sotilgan: <b>{float(stats.get('units') or 0):.0f} dona</b>",
            f"💵 Tushum: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
            f"✅ To‘lovga: <b>{_format_money(float(stats.get('payout_total') or 0))}</b>",
            f"🔄 Uzum bergan tannarxlar: <b>{saved_costs}</b>",
        ]
        if known_items:
            lines.append(f"💰 Taxminiy sof foyda: <b>{_format_money(total_known_profit)}</b>")
        lines.append("\n<b>Top tovarlar:</b>")
        for idx, item in enumerate(rows[:10], start=1):
            sku = escape(_short_text(item.get("sku"), 55))
            title = escape(_short_text(item.get("title"), 70))
            cost = item.get("cost_per_unit")
            if cost is None:
                hint = "\n   ⚠️ Uzum bu SKU uchun tannarx bermagan"
            else:
                profit = float(item.get("profit") or 0)
                margin = float(item.get("margin") or 0)
                hint = f"\n   🧾 Tannarx: <b>{_format_money(float(cost))}</b> | 💰 Foyda: <b>{_format_money(profit)}</b> | 📈 Marja: <b>{margin:.1f}%</b>"
            lines.append(
                f"{idx}. <b>{title}</b>\n"
                f"   🔖 SKU: <code>{sku}</code>\n"
                f"   🔢 Soni: <b>{float(item.get('qty') or 0):.0f} dona</b> | "
                f"💵 Tushum: <b>{_format_money(float(item.get('revenue') or 0))}</b> | "
                f"✅ To‘lovga: <b>{_format_money(float(item.get('payout') or 0))}</b>" + hint
            )
        lines.append("\nTannarx faqat Uzum ma’lumotidan olinadi.")
        return "\n\n".join(lines)

    if not rows:
        return (
            f"🧾 <b>Юнит-экономика</b>\n🏪 Магазин: <code>{shop_id}</code>\n\n"
            "Продаж не найдено. Сначала должны быть продажи за выбранный период."
        )
    total_known_profit = sum(float(x.get("profit") or 0) for x in rows if x.get("profit") is not None)
    known_items = sum(1 for x in rows if x.get("profit") is not None)
    lines = [
        f"🧾 <b>Юнит-экономика за {days} дней</b>",
        f"🏪 Магазин: <code>{shop_id}</code>",
        f"📦 Продано: <b>{float(stats.get('units') or 0):.0f} шт.</b>",
        f"💵 Выручка: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
        f"✅ К выплате: <b>{_format_money(float(stats.get('payout_total') or 0))}</b>",
        f"🔄 Себестоимостей получено от Uzum: <b>{saved_costs}</b>",
    ]
    if known_items:
        lines.append(f"💰 Примерная чистая прибыль: <b>{_format_money(total_known_profit)}</b>")
    lines.append("\n<b>Топ товаров:</b>")
    for idx, item in enumerate(rows[:10], start=1):
        sku = escape(_short_text(item.get("sku"), 55))
        title = escape(_short_text(item.get("title"), 70))
        cost = item.get("cost_per_unit")
        if cost is None:
            hint = "\n   ⚠️ Uzum не передал себестоимость для этого SKU"
        else:
            profit = float(item.get("profit") or 0)
            margin = float(item.get("margin") or 0)
            hint = f"\n   🧾 Себестоимость: <b>{_format_money(float(cost))}</b> | 💰 Прибыль: <b>{_format_money(profit)}</b> | 📈 Маржа: <b>{margin:.1f}%</b>"
        lines.append(
            f"{idx}. <b>{title}</b>\n"
            f"   🔖 SKU: <code>{sku}</code>\n"
            f"   🔢 Кол-во: <b>{float(item.get('qty') or 0):.0f} шт.</b> | "
            f"💵 Выручка: <b>{_format_money(float(item.get('revenue') or 0))}</b> | "
            f"✅ К выплате: <b>{_format_money(float(item.get('payout') or 0))}</b>" + hint
        )
    lines.append("\nСебестоимость берётся только из данных Uzum.")
    return "\n\n".join(lines)


@dp.message(Command("unit", "unit_economy"))
async def unit_economy(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, client, shop_id = req
    lang = get_user_language(telegram_id)
    await message.answer("⌛ Hisoblayapman..." if lang == "uz" else "⌛ Считаю юнит-экономику...", reply_markup=sales_menu_for_message(message))
    try:
        rows, stats, saved_costs = await _unit_economy_for_shop(client, telegram_id, shop_id, days=30)
        if not rows:
            text = f"🧾 <b>Unit iqtisodiyot</b>\n🏪 Do‘kon: <code>{shop_id}</code>\n\nSavdolar topilmadi." if lang == "uz" else f"🧾 <b>Юнит-экономика</b>\n🏪 Магазин: <code>{shop_id}</code>\n\nПродаж не найдено."
            await message.answer(text, reply_markup=sales_menu_for_message(message))
            return
        total_known_profit = sum(float(x.get("profit") or 0) for x in rows if x.get("profit") is not None)
        known_items = sum(1 for x in rows if x.get("profit") is not None)
        title = "🧾 <b>Unit iqtisodiyot — 30 kun</b>" if lang == "uz" else "🧾 <b>Юнит-экономика за 30 дней</b>"
        summary = [
            f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
            f"📦 Sotilgan: <b>{float(stats.get('units') or 0):.0f} dona</b> | 💵 Tushum: <b>{_format_money(float(stats.get('revenue') or 0))}</b>" if lang == "uz" else f"📦 Продано: <b>{float(stats.get('units') or 0):.0f} шт.</b> | 💵 Выручка: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
            f"🔄 Uzum tannarxlari: <b>{saved_costs}</b>" if lang == "uz" else f"🔄 Себестоимостей из Uzum: <b>{saved_costs}</b>",
        ]
        if known_items:
            summary.append(f"💰 Taxminiy sof foyda: <b>{_format_money(total_known_profit)}</b>" if lang == "uz" else f"💰 Примерная чистая прибыль: <b>{_format_money(total_known_profit)}</b>")
        items: list[str] = []
        for idx, item in enumerate(rows, start=1):
            sku = escape(_short_text(item.get("sku"), 55))
            title_item = escape(_short_text(item.get("title"), 70))
            cost = item.get("cost_per_unit")
            if cost is None:
                hint = "\n⚠️ Uzum bu SKU uchun tannarx bermagan" if lang == "uz" else "\n⚠️ Uzum не передал себестоимость для этого SKU"
            else:
                profit = float(item.get("profit") or 0)
                margin = float(item.get("margin") or 0)
                hint = f"\n🧾 Tannarx: <b>{_format_money(float(cost))}</b> | 💰 Foyda: <b>{_format_money(profit)}</b> | 📈 Marja: <b>{margin:.1f}%</b>" if lang == "uz" else f"\n🧾 Себестоимость: <b>{_format_money(float(cost))}</b> | 💰 Прибыль: <b>{_format_money(profit)}</b> | 📈 Маржа: <b>{margin:.1f}%</b>"
            if lang == "uz":
                items.append(f"{idx}. <b>{title_item}</b>\n🔖 SKU: <code>{sku}</code>\n🔢 Soni: <b>{float(item.get('qty') or 0):.0f} dona</b> | 💵 Tushum: <b>{_format_money(float(item.get('revenue') or 0))}</b> | ✅ To‘lovga: <b>{_format_money(float(item.get('payout') or 0))}</b>{hint}")
            else:
                items.append(f"{idx}. <b>{title_item}</b>\n🔖 SKU: <code>{sku}</code>\n🔢 Кол-во: <b>{float(item.get('qty') or 0):.0f} шт.</b> | 💵 Выручка: <b>{_format_money(float(item.get('revenue') or 0))}</b> | ✅ К выплате: <b>{_format_money(float(item.get('payout') or 0))}</b>{hint}")
        await send_paginated_list(message, kind="unit", title=title, summary=summary, items=items, section="sales", reply_markup=sales_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


async def show_uzum_cost_status(message: Message, *, force: bool = True) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    await message.answer(
        "⌛ Uzum tannarxlarini yangilayapman..."
        if lang == "uz"
        else "⌛ Обновляю себестоимость из Uzum...",
        reply_markup=finance_menu_for_message(message),
    )
    try:
        status = await sync_uzum_sku_financials(
            client,
            telegram_id,
            shop_id,
            force=force,
        )
        total = int(status.get("total") or 0)
        with_cost = int(status.get("with_cost") or 0)
        missing = int(status.get("missing_cost") or 0)
        coverage = with_cost / total * 100.0 if total else 0.0
        updated = _fmt_dt(status.get("fetched_at"))
        stale_note = (
            "\n⚠️ Uzum vaqtincha javob bermadi; oxirgi saqlangan ma’lumot ko‘rsatildi."
            if lang == "uz" and status.get("stale")
            else "\n⚠️ Uzum временно не ответил; показаны последние сохранённые данные."
            if status.get("stale")
            else ""
        )
        if lang == "uz":
            text_value = (
                "🔄 <b>Uzum tannarxi</b>\n\n"
                f"🏪 Do‘kon: <code>{shop_id}</code>\n"
                f"📦 SKU jami: <b>{total}</b>\n"
                f"✅ Tannarx bor: <b>{with_cost}</b>\n"
                f"⚠️ Uzum tannarx bermagan: <b>{missing}</b>\n"
                f"📊 Qamrov: <b>{coverage:.1f}%</b>\n"
                f"🕒 Yangilandi: <b>{updated}</b>\n\n"
                "Tannarx faqat Uzum API dagi <code>purchasePrice</code> maydonidan olinadi. "
                "Excel va qo‘lda kiritilgan narxlar foyda hisobida ishlatilmaydi."
                f"{stale_note}"
            )
        else:
            text_value = (
                "🔄 <b>Себестоимость из Uzum</b>\n\n"
                f"🏪 Магазин: <code>{shop_id}</code>\n"
                f"📦 Всего SKU: <b>{total}</b>\n"
                f"✅ С себестоимостью: <b>{with_cost}</b>\n"
                f"⚠️ Uzum не передал цену: <b>{missing}</b>\n"
                f"📊 Покрытие: <b>{coverage:.1f}%</b>\n"
                f"🕒 Обновлено: <b>{updated}</b>\n\n"
                "Источник — только документированное поле Uzum <code>purchasePrice</code>. "
                "Excel и ручные значения в расчёте прибыли не используются."
                f"{stale_note}"
            )
        await message.answer(text_value, reply_markup=finance_menu_for_message(message))
    except Exception as error:
        await send_api_error(message, error)


@dp.message(Command("cost", "setcost", "set_cost"))
async def set_cost_command(message: Message) -> None:
    await show_uzum_cost_status(message)
    return
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    lang = get_user_language(telegram_id)
    parsed = _parse_cost_command_args(message.text or "")
    if not parsed:
        text = (
            "Tannarxni shunday kiriting:\n<code>/cost SKU 60000</code>\n\nMasalan:\n<code>/cost NOORZA-NR751-BEJEV-XXL 60000</code>"
            if lang == "uz" else
            "Укажите себестоимость так:\n<code>/cost SKU 60000</code>\n\nНапример:\n<code>/cost NOORZA-NR751-БЕЖЕВ-XXL 60000</code>"
        )
        await message.answer(text, reply_markup=menu_for_message(message))
        return
    sku, cost = parsed
    save_unit_cost(telegram_id, shop_id, sku, cost, title=sku)
    if lang == "uz":
        await message.answer(f"✅ Saqlandi\n🔖 SKU: <code>{escape(sku)}</code>\n🧾 Tannarx: <b>{_format_money(cost)}</b>", reply_markup=menu_for_message(message))
    else:
        await message.answer(f"✅ Сохранено\n🔖 SKU: <code>{escape(sku)}</code>\n🧾 Себестоимость: <b>{_format_money(cost)}</b>", reply_markup=menu_for_message(message))


@dp.message(Command("costs"))
async def costs_command(message: Message) -> None:
    await show_uzum_cost_status(message)
    return
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    lang = get_user_language(telegram_id)
    rows = list_unit_costs(telegram_id, shop_id, limit=50)
    if not rows:
        text = "Hali tannarxlar kiritilmagan. Qo‘shish: <code>/cost SKU 60000</code>" if lang == "uz" else "Себестоимость ещё не указана. Добавить: <code>/cost SKU 60000</code>"
        await message.answer(text, reply_markup=menu_for_message(message))
        return
    title = "🧾 <b>Saqlangan tannarxlar</b>" if lang == "uz" else "🧾 <b>Сохранённая себестоимость</b>"
    lines = [title, f"🏪 <code>{shop_id}</code>"]
    for r in rows:
        lines.append(f"• <code>{escape(str(r.get('sku_key') or ''))}</code> — <b>{_format_money(float(r.get('cost') or 0))}</b>")
    await message.answer("\n".join(lines), reply_markup=menu_for_message(message))


@dp.message(Command("delcost", "delete_cost"))
async def delete_cost_command(message: Message) -> None:
    await show_uzum_cost_status(message)
    return
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    sku = parse_args(message.text or "").strip()
    lang = get_user_language(telegram_id)
    if not sku:
        await message.answer("Напишите так: <code>/delcost SKU</code>", reply_markup=menu_for_message(message))
        return
    ok = delete_unit_cost(telegram_id, shop_id, sku)
    if lang == "uz":
        await message.answer("✅ O‘chirildi" if ok else "Topilmadi", reply_markup=menu_for_message(message))
    else:
        await message.answer("✅ Удалено" if ok else "Не найдено", reply_markup=menu_for_message(message))



def _build_cost_template(path: str, lang: str = "ru") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Себестоимость" if lang != "uz" else "Tannarx"
    headers = ["SKU", "Себестоимость", "Название (необязательно)"] if lang != "uz" else ["SKU", "Tannarx", "Nomi (ixtiyoriy)"]
    ws.append(headers)
    ws.append(["NOORZA-NR751-BEJEV-XXL", 60000, "Пример товара" if lang != "uz" else "Tovar namunasi"])
    ws.append(["NOORZA-KOR101-SERIY", 45000, ""])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 38
    ws.freeze_panes = "A2"
    wb.save(path)


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = text.replace(" ", "_").replace("-", "_")
    return text


def _parse_excel_cost_value(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) if float(value) >= 0 else None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("сум", "").replace("so'm", "").replace("so‘m", "").replace("uzs", "")
    s = s.replace(" ", "").replace("\u00a0", "").replace(",", ".")
    try:
        val = float(s)
        return val if val >= 0 else None
    except Exception:
        return None


def _detect_cost_columns(ws) -> tuple[int, int, int | None, int]:
    header_values = [cell.value for cell in ws[1]]
    normalized = [_normalize_header(v) for v in header_values]
    sku_col = cost_col = title_col = None
    for idx, h in enumerate(normalized, start=1):
        if h in {"sku", "артикул", "sku_id", "sku_title", "шк", "barcode", "offerid", "offer_id"}:
            sku_col = idx
        if h in {"себестоимость", "sebestoimost", "tannarx", "cost", "purchase_price", "закуп", "закупочная_цена", "tan_narx"}:
            cost_col = idx
        if h in {"название", "name", "title", "nomi", "товар", "product", "product_title", "название_(необязательно)", "nomi_(ixtiyoriy)"}:
            title_col = idx
    if sku_col and cost_col:
        return sku_col, cost_col, title_col, 2
    return 1, 2, 3, 1


def _validate_cost_workbook_file(path: str | Path) -> None:
    workbook_path = Path(path)
    if workbook_path.stat().st_size > COST_IMPORT_MAX_FILE_BYTES:
        raise ValueError(
            f"Файл слишком большой. Максимум: {COST_IMPORT_MAX_FILE_BYTES // 1_048_576} МБ."
        )
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            members = archive.infolist()
            if any(member.flag_bits & 0x1 for member in members):
                raise ValueError("Защищённые паролем Excel-файлы не поддерживаются.")
            unpacked_size = sum(max(0, int(member.file_size)) for member in members)
            if unpacked_size > COST_IMPORT_MAX_UNCOMPRESSED_BYTES:
                raise ValueError("Excel-файл содержит слишком большой объём распакованных данных.")
    except zipfile.BadZipFile as error:
        raise ValueError("Файл повреждён или не является корректным .xlsx.") from error


def _parse_costs_workbook(path: str) -> tuple[list[dict[str, Any]], list[str]]:
    _validate_cost_workbook_file(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        sku_col, cost_col, title_col, start_row = _detect_cost_columns(ws)
        imported: list[dict[str, Any]] = []
        errors: list[str] = []
        seen: set[str] = set()
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=start_row, values_only=True),
            start=start_row,
        ):
            if row_idx - start_row >= COST_IMPORT_MAX_ROWS:
                errors.append(
                    f"Достигнут лимит импорта: {COST_IMPORT_MAX_ROWS} строк. Остальные строки пропущены."
                )
                break
            sku_val = row[sku_col - 1] if len(row) >= sku_col else None
            cost_val = row[cost_col - 1] if len(row) >= cost_col else None
            title_val = row[title_col - 1] if title_col and len(row) >= title_col else ""
            sku = str(sku_val or "").strip()
            if not sku:
                continue
            cost = _parse_excel_cost_value(cost_val)
            if cost is None:
                errors.append(f"Строка {row_idx}: неверная себестоимость для {sku}")
                continue
            key = _unit_sku_key(sku)
            if not key:
                errors.append(f"Строка {row_idx}: неверный SKU")
                continue
            if key in seen:
                continue
            seen.add(key)
            imported.append(
                {
                    "sku": sku,
                    "cost": float(cost),
                    "title": str(title_val or "").strip(),
                }
            )
        return imported, errors
    finally:
        wb.close()


def _save_unit_costs_bulk(
    telegram_id: int,
    shop_id: int,
    items: list[dict[str, Any]],
) -> int:
    init_unit_economy_tables()
    now_text = _dt_to_db(_utc_now()) or ""
    values = [
        (
            int(telegram_id),
            int(shop_id),
            _unit_sku_key(item.get("sku")),
            str(item.get("title") or item.get("sku") or "").strip(),
            float(item.get("cost") or 0),
            now_text,
        )
        for item in items
        if _unit_sku_key(item.get("sku"))
    ]
    with db.connect() as conn:
        conn.executemany(
            """
            INSERT INTO unit_costs (telegram_id, shop_id, sku_key, title, cost, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id, sku_key) DO UPDATE SET
                title = excluded.title,
                cost = excluded.cost,
                updated_at = excluded.updated_at
            """,
            values,
        )
        conn.commit()
    return len(values)


@dp.message(Command("cost_template", "costs_template", "template_costs"))
async def cost_template_command(message: Message, state: FSMContext | None = None) -> None:
    if state is not None:
        await state.clear()
    await show_uzum_cost_status(message)
    return
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    if state is not None:
        await state.set_state(CostImportStates.waiting_for_file)
    filename = f"unit_costs_template_{telegram_id}.xlsx"
    file_path = str(Path(tempfile.gettempdir()) / filename)
    _build_cost_template(file_path, lang=lang)
    if lang == "uz":
        text = (
            "📥 <b>Tannarxlarni Excel orqali yuklash</b>\n\n"
            "1. Shablonni yuklab oling.\n"
            "2. SKU va tannarxlarni to‘ldiring.\n"
            "3. Tayyor Excel faylni shu chatga yuboring.\n\n"
            "Ustunlar: <b>SKU</b> va <b>Tannarx</b>.\n"
            "Bekor qilish: <code>/cancel</code>"
        )
    else:
        text = (
            "📥 <b>Загрузка себестоимости через Excel</b>\n\n"
            "1. Скачайте шаблон.\n"
            "2. Заполните SKU и себестоимость.\n"
            "3. Отправьте готовый Excel-файл сюда в чат.\n\n"
            "Обязательные колонки: <b>SKU</b> и <b>Себестоимость</b>.\n"
            "Отмена: <code>/cancel</code>"
        )
    await message.answer(text, reply_markup=menu_for_message(message))
    await message.answer_document(FSInputFile(file_path, filename="unit_costs_template.xlsx"))


@dp.message(Command("import_costs", "upload_costs"))
async def import_costs_command(message: Message, state: FSMContext) -> None:
    await state.clear()
    await show_uzum_cost_status(message)
    return
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    await state.set_state(CostImportStates.waiting_for_file)
    text = (
        "📥 Excel faylni yuboring. Shablon kerak bo‘lsa: <code>/cost_template</code>\nBekor qilish: <code>/cancel</code>"
        if lang == "uz"
        else "📥 Отправьте Excel-файл с себестоимостью. Шаблон: <code>/cost_template</code>\nОтмена: <code>/cancel</code>"
    )
    await message.answer(text, reply_markup=menu_for_message(message))


@dp.message(CostImportStates.waiting_for_file, F.document)
async def receive_costs_excel(message: Message, state: FSMContext) -> None:
    await state.clear()
    await show_uzum_cost_status(message)
    return
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    document = message.document
    if document is None:
        return
    filename = (document.file_name or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
        await message.answer(
            "Excel fayl .xlsx formatida bo‘lishi kerak." if lang == "uz" else "Нужен Excel-файл в формате .xlsx.",
            reply_markup=menu_for_message(message),
        )
        return
    if document.file_size and int(document.file_size) > COST_IMPORT_MAX_FILE_BYTES:
        await message.answer(
            (
                f"Excel fayl juda katta. Maksimum: {COST_IMPORT_MAX_FILE_BYTES // 1_048_576} MB."
                if lang == "uz"
                else f"Excel-файл слишком большой. Максимум: {COST_IMPORT_MAX_FILE_BYTES // 1_048_576} МБ."
            ),
            reply_markup=menu_for_message(message),
        )
        return
    tmp_path = str(
        Path(tempfile.gettempdir()) / f"costs_{telegram_id}_{time.time_ns()}.xlsx"
    )
    try:
        await bot.download(document, destination=tmp_path)
        items, errors = await asyncio.to_thread(_parse_costs_workbook, tmp_path)
        if not items:
            await message.answer(
                "Faylda saqlash uchun SKU va tannarx topilmadi." if lang == "uz" else "В файле не нашёл SKU и себестоимость для сохранения.",
                reply_markup=menu_for_message(message),
            )
            return
        saved_count = await asyncio.to_thread(
            _save_unit_costs_bulk,
            telegram_id,
            int(shop_id),
            items,
        )
        await state.clear()
        preview = "\n".join([f"• <code>{escape(i['sku'])}</code> — <b>{_format_money(float(i['cost']))}</b>" for i in items[:10]])
        more = max(0, len(items) - 10)
        if lang == "uz":
            text = f"✅ <b>Tannarxlar saqlandi</b>\n\n🏪 Do‘kon: <code>{shop_id}</code>\n📦 Saqlandi: <b>{saved_count}</b> SKU"
            if preview:
                text += "\n\n" + preview
            if more:
                text += f"\n...yana {more} ta"
            text += "\n\nEndi <b>🧾 Unit iqtisodiyot</b> yoki <b>💰 Foyda</b> bo‘limini tekshiring."
        else:
            text = f"✅ <b>Себестоимость сохранена</b>\n\n🏪 Магазин: <code>{shop_id}</code>\n📦 Сохранено: <b>{saved_count}</b> SKU"
            if preview:
                text += "\n\n" + preview
            if more:
                text += f"\n...ещё {more} шт."
            text += "\n\nТеперь проверьте <b>🧾 Юнит-экономику</b> или <b>💰 Прибыль</b>."
        if errors:
            text += ("\n\n⚠️ Ошибки: " if lang != "uz" else "\n\n⚠️ Xatolar: ") + str(len(errors))
            text += "\n" + "\n".join(escape(e) for e in errors[:5])
        await message.answer(text, reply_markup=sales_menu_for_message(message))
    except ValueError as error:
        await message.answer(
            (
                f"❌ Excel faylni yuklab bo‘lmadi: {escape(str(error))}"
                if lang == "uz"
                else f"❌ Не удалось импортировать Excel: {escape(str(error))}"
            ),
            reply_markup=menu_for_message(message),
        )
    except Exception as e:
        await send_api_error(message, e)
    finally:
        try:
            Path(tmp_path).unlink(missing_ok=True)
        except Exception:
            pass


@dp.message(CostImportStates.waiting_for_file)
async def receive_costs_excel_wrong(message: Message) -> None:
    await show_uzum_cost_status(message)


def _profit_summary_from_unit_rows(rows: list[dict[str, Any]], stats: dict[str, Any]) -> dict[str, Any]:
    known_rows = [r for r in rows if float(r.get("known_cost_qty") or (r.get("qty") if r.get("cost_per_unit") is not None else 0) or 0) > 0]
    complete_rows = [r for r in rows if bool(r.get("cost_complete", r.get("cost_per_unit") is not None))]
    cost_total = sum(float(r.get("cost_total") or 0) for r in known_rows)
    known_profit = sum(float(r.get("profit") or 0) for r in known_rows)
    known_net_profit = sum(float(r.get("net_profit") if r.get("net_profit") is not None else r.get("profit") or 0) for r in known_rows)
    known_revenue = sum(
        float(r.get("known_revenue") if r.get("known_revenue") is not None else r.get("revenue") or 0)
        for r in known_rows
    )
    known_payout = sum(
        float(r.get("known_payout") if r.get("known_payout") is not None else r.get("payout") or 0)
        for r in known_rows
    )
    known_commission = sum(float(r.get("known_commission") or 0) for r in known_rows)
    known_logistics = sum(float(r.get("known_logistics") or 0) for r in known_rows)
    missing = [r for r in rows if not bool(r.get("cost_complete", r.get("cost_per_unit") is not None))]
    total_revenue = float(stats.get("revenue") or 0)
    margin = (known_profit / known_revenue * 100.0) if known_revenue > 0 else 0.0
    coverage = known_revenue / total_revenue if total_revenue > 0 else 0.0
    return {
        "cost_total": cost_total,
        "profit": known_profit,
        "net_profit_after_tax": known_net_profit,
        "margin": margin,
        "coverage": max(0.0, min(1.0, coverage)),
        "missing_count": len(missing),
        "known_count": len(complete_rows),
        "total_count": len(rows),
        "total_revenue": total_revenue,
        "known_revenue": known_revenue,
        "known_payout": known_payout,
        "known_commission": known_commission,
        "known_logistics": known_logistics,
        "missing": missing,
    }


def calculate_business_profit(
    cost_summary: dict[str, Any],
    stats: dict[str, Any],
    settings: dict[str, Any],
    *,
    days: int = 30,
    uzum_expenses: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Apply tax, actual Uzum expenses and external seller expenses.

    The result deliberately keeps cost coverage visible: if some SKUs have no
    cost, net profit is an incomplete estimate and must not be presented as an
    accounting fact.
    """
    period_ratio = max(0.0, float(days)) / 30.0
    revenue = float(stats.get("revenue") or 0)
    tax_percent = max(0.0, min(100.0, float(settings.get("tax_percent") or 0)))
    coverage = max(0.0, min(1.0, float(cost_summary.get("coverage") or 0)))
    missing_count = int(cost_summary.get("missing_count") or 0)
    cost_data_complete = coverage >= 0.999 or (revenue <= 0 and missing_count == 0)
    # When purchasePrice is incomplete, presenting tax for all revenue against
    # profit from only known-cost rows mixes two different bases.  Keep the
    # visible result internally consistent and expose the full-period tax
    # separately for audit.
    known_revenue = float(cost_summary.get("known_revenue") or 0)
    tax_basis_revenue = revenue if cost_data_complete else known_revenue
    calculation_payout = (
        float(stats.get("payout_total") or 0)
        if cost_data_complete
        else float(cost_summary.get("known_payout") or 0)
    )
    calculation_commission = (
        float(stats.get("commission") or 0)
        if cost_data_complete
        else float(cost_summary.get("known_commission") or 0)
    )
    calculation_logistics = (
        float(stats.get("logistics") or 0)
        if cost_data_complete
        else float(cost_summary.get("known_logistics") or 0)
    )
    payout_residual = (
        tax_basis_revenue
        - calculation_payout
        - calculation_commission
        - calculation_logistics
    )
    other_payout_deductions = max(0.0, payout_residual)
    payout_adjustment = max(0.0, -payout_residual)
    tax_expense = tax_basis_revenue * tax_percent / 100.0
    full_period_tax_expense = revenue * tax_percent / 100.0
    advertising = max(0.0, float(settings.get("advertising_monthly") or 0)) * period_ratio
    storage = max(0.0, float(settings.get("storage_monthly") or 0)) * period_ratio
    other = max(0.0, float(settings.get("other_monthly") or 0)) * period_ratio
    expense_summary = dict(uzum_expenses or {})
    uzum_expenses_available = bool(expense_summary.get("available", uzum_expenses is None))
    signed_expense_total = float(expense_summary.get("total") or 0)
    uzum_expense_deductions = float(
        expense_summary.get("deductions")
        if expense_summary.get("deductions") is not None
        else max(0.0, signed_expense_total)
    )
    uzum_expense_refunds = float(
        expense_summary.get("refunds")
        if expense_summary.get("refunds") is not None
        else max(0.0, -signed_expense_total)
    )
    uzum_expense_total = uzum_expense_deductions - uzum_expense_refunds
    cost_total = float(cost_summary.get("cost_total") or 0)
    known_profit = calculation_payout - cost_total
    external_expense_total = advertising + storage + other
    net_profit = (
        known_profit
        - tax_expense
        - uzum_expense_deductions
        + uzum_expense_refunds
        - external_expense_total
    )
    return {
        "days": int(days),
        "revenue": revenue,
        "commission": float(stats.get("commission") or 0),
        "logistics": float(stats.get("logistics") or 0),
        "payout_total": float(stats.get("payout_total") or 0),
        "calculation_revenue": tax_basis_revenue,
        "calculation_payout": calculation_payout,
        "calculation_commission": calculation_commission,
        "calculation_logistics": calculation_logistics,
        "other_payout_deductions": other_payout_deductions,
        "payout_adjustment": payout_adjustment,
        "cost_total": cost_total,
        "known_profit": known_profit,
        "tax_expense": tax_expense,
        "full_period_tax_expense": full_period_tax_expense,
        "tax_basis_revenue": tax_basis_revenue,
        "uzum_expense_total": uzum_expense_total,
        "uzum_expense_deductions": uzum_expense_deductions,
        "uzum_expense_refunds": uzum_expense_refunds,
        "uzum_storage_expense": float(expense_summary.get("storage") or 0),
        "uzum_advertising_expense": float(expense_summary.get("advertising") or 0),
        "uzum_penalty_expense": float(expense_summary.get("penalty") or 0),
        "uzum_other_expense": float(expense_summary.get("other") or 0),
        "uzum_order_charge_already_in_payout": float(
            expense_summary.get("order_charge") or 0
        ),
        "uzum_expenses_available": uzum_expenses_available,
        "advertising_expense": advertising,
        "storage_expense": storage,
        "other_expense": other,
        "external_expense_total": external_expense_total,
        "operating_expenses": (
            tax_expense + uzum_expense_total + advertising + storage + other
        ),
        "net_profit": net_profit,
        "net_margin": (net_profit / revenue * 100.0) if revenue > 0 else 0.0,
        "coverage": coverage,
        "complete": cost_data_complete and uzum_expenses_available,
        "missing_count": missing_count,
    }


def _format_profit_bridge_lines(
    business: dict[str, Any],
    *,
    lang: str = "ru",
) -> list[str]:
    """Return one reconciled, human-readable path from revenue to result."""
    uz = normalize_lang(lang) == "uz"
    complete = bool(business.get("complete"))
    result_label = (
        "Sof foyda"
        if uz and complete
        else "Ma’lum ma’lumotlar bo‘yicha natija"
        if uz
        else "Чистая прибыль"
        if complete
        else "Результат по известным данным"
    )
    lines = [
        "🧮 <b>Natija qanday hisoblandi</b>" if uz else "🧮 <b>Как получился результат</b>",
        f"1. Tushum: <b>{_format_money(float(business.get('calculation_revenue') or 0))}</b>"
        if uz
        else f"1. Выручка: <b>{_format_money(float(business.get('calculation_revenue') or 0))}</b>",
        f"2. − Uzum komissiyasi: <b>{_format_money(float(business.get('calculation_commission') or 0))}</b>"
        if uz
        else f"2. − Комиссия Uzum: <b>{_format_money(float(business.get('calculation_commission') or 0))}</b>",
        f"3. − Logistika: <b>{_format_money(float(business.get('calculation_logistics') or 0))}</b>"
        if uz
        else f"3. − Логистика: <b>{_format_money(float(business.get('calculation_logistics') or 0))}</b>",
    ]
    other_payout = float(business.get("other_payout_deductions") or 0)
    payout_adjustment = float(business.get("payout_adjustment") or 0)
    lines.append(
        f"4. − To‘lov ichidagi boshqa ushlanmalar: <b>{_format_money(other_payout)}</b>"
        if uz
        else f"4. − Другие удержания внутри выплаты: <b>{_format_money(other_payout)}</b>"
    )
    if payout_adjustment > 0.5:
        lines.append(
            f"4a. + Uzum to‘lov tuzatishi: <b>{_format_money(payout_adjustment)}</b>"
            if uz
            else f"4а. + Корректировка выплаты Uzum: <b>{_format_money(payout_adjustment)}</b>"
        )
    lines.extend(
        [
            f"= To‘lovga: <b>{_format_money(float(business.get('calculation_payout') or 0))}</b>"
            if uz
            else f"= К выплате: <b>{_format_money(float(business.get('calculation_payout') or 0))}</b>",
            f"5. − Tannarx: <b>{_format_money(float(business.get('cost_total') or 0))}</b>"
            if uz
            else f"5. − Себестоимость: <b>{_format_money(float(business.get('cost_total') or 0))}</b>",
            f"= Soliqdan oldingi foyda: <b>{_format_money(float(business.get('known_profit') or 0))}</b>"
            if uz
            else f"= Прибыль до налога: <b>{_format_money(float(business.get('known_profit') or 0))}</b>",
            f"6. − Soliq: <b>{_format_money(float(business.get('tax_expense') or 0))}</b>"
            if uz
            else f"6. − Налог: <b>{_format_money(float(business.get('tax_expense') or 0))}</b>",
            f"7. − Uzum qo‘shimcha xarajatlari: <b>{_format_money(float(business.get('uzum_expense_deductions') or 0))}</b>"
            if uz
            else f"7. − Доп. расходы Uzum: <b>{_format_money(float(business.get('uzum_expense_deductions') or 0))}</b>",
            f"8. + Uzum qaytargan mablag‘: <b>{_format_money(float(business.get('uzum_expense_refunds') or 0))}</b>"
            if uz
            else f"8. + Возвраты от Uzum: <b>{_format_money(float(business.get('uzum_expense_refunds') or 0))}</b>",
            f"9. − Tashqi xarajatlar: <b>{_format_money(float(business.get('external_expense_total') or 0))}</b>"
            if uz
            else f"9. − Внешние расходы: <b>{_format_money(float(business.get('external_expense_total') or 0))}</b>",
            f"<b>= {result_label}: {_format_money(float(business.get('net_profit') or 0))}</b>",
        ]
    )
    lines.append(
        "<i>Komissiya va logistika «to‘lovga» summasida allaqachon hisobga olingan va ikkinchi marta ayrilmaydi.</i>"
        if uz
        else "<i>Комиссия и логистика уже учтены в «к выплате» и второй раз не вычитаются.</i>"
    )
    if not complete:
        lines.append(
            "⚠️ Hisob faqat tannarxi ma’lum bo‘lgan savdolar bo‘yicha."
            if uz
            else "⚠️ Расчёт относится только к продажам с известной себестоимостью."
        )
    return lines


def _format_profit_report(shop_id: int, rows: list[dict[str, Any]], stats: dict[str, Any], lang: str = "ru") -> str:
    summary = _profit_summary_from_unit_rows(rows, stats)
    top_profit = sorted([r for r in rows if r.get("cost_per_unit") is not None], key=lambda r: float(r.get("profit") or 0), reverse=True)[:10]
    if lang == "uz":
        lines = [
            "💰 <b>30 kunlik foyda</b>",
            f"🏪 Do‘kon: <code>{shop_id}</code>",
            "",
            f"💵 Tushum: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
            f"🏷 Uzum komissiyasi: <b>{_format_money(float(stats.get('commission') or 0))}</b>",
            f"🚚 Logistika: <b>{_format_money(float(stats.get('logistics') or 0))}</b>",
            f"📦 Tannarx: <b>{_format_money(float(summary['cost_total']))}</b>",
            "",
            f"✅ To‘lovga: <b>{_format_money(float(stats.get('payout_total') or 0))}</b>",
            f"💰 Hisobiy foyda: <b>{_format_money(float(summary['profit']))}</b>",
            f"📈 Marja: <b>{float(summary['margin']):.1f}%</b>",
            f"📌 Tannarx bilan qamrov: <b>{float(summary['coverage']) * 100:.1f}%</b>",
        ]
        if summary["missing_count"]:
            lines.append(f"\n⚠️ Uzum tannarx bermagan SKU: <b>{summary['missing_count']}</b>")
            lines.append("Tannarx manbasi: faqat Uzum purchasePrice.")
        if top_profit:
            lines.append("\n🏆 <b>Foyda bo‘yicha top tovarlar:</b>")
            for idx, r in enumerate(top_profit, start=1):
                lines.append(f"{idx}. {escape(_short_text(str(r.get('title') or r.get('sku') or '-'), 55))} — <b>{_format_money(float(r.get('profit') or 0))}</b>")
        return "\n".join(lines)
    lines = [
        "💰 <b>Прибыль за 30 дней</b>",
        f"🏪 Магазин: <code>{shop_id}</code>",
        "",
        f"💵 Выручка: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
        f"🏷 Комиссия Uzum: <b>{_format_money(float(stats.get('commission') or 0))}</b>",
        f"🚚 Логистика: <b>{_format_money(float(stats.get('logistics') or 0))}</b>",
        f"📦 Себестоимость: <b>{_format_money(float(summary['cost_total']))}</b>",
        "",
        f"✅ К выплате: <b>{_format_money(float(stats.get('payout_total') or 0))}</b>",
        f"💰 Расчётная прибыль: <b>{_format_money(float(summary['profit']))}</b>",
        f"📈 Маржа: <b>{float(summary['margin']):.1f}%</b>",
        f"📌 Покрытие себестоимостью: <b>{float(summary['coverage']) * 100:.1f}%</b>",
    ]
    if summary["missing_count"]:
        lines.append(f"\n⚠️ Uzum не передал себестоимость: <b>{summary['missing_count']}</b> SKU")
        lines.append("Источник себестоимости: только Uzum purchasePrice.")
    if top_profit:
        lines.append("\n🏆 <b>Топ товаров по прибыли:</b>")
        for idx, r in enumerate(top_profit, start=1):
            lines.append(f"{idx}. {escape(_short_text(str(r.get('title') or r.get('sku') or '-'), 55))} — <b>{_format_money(float(r.get('profit') or 0))}</b>")
    return "\n".join(lines)


@dp.message(Command("profit", "unit_profit", "profit_report"))
async def profit_report(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    await message.answer("⌛ Foydani hisoblayapman..." if lang == "uz" else "⌛ Считаю прибыль...", reply_markup=sales_menu_for_message(message))
    try:
        rows, stats, _ = await _unit_economy_for_shop(client, telegram_id, shop_id, days=30)
        summary_stats = _profit_summary_from_unit_rows(rows, stats)
        finance_settings = ensure_finance_settings(telegram_id, shop_id)
        expense_from, expense_to = _days_range_ms(30)
        uzum_expenses = await load_uzum_expense_summary(
            client,
            shop_id,
            expense_from,
            expense_to,
        )
        business_profit = calculate_business_profit(
            summary_stats,
            stats,
            finance_settings,
            days=30,
            uzum_expenses=uzum_expenses,
        )
        known = sorted(
            [r for r in rows if r.get("cost_per_unit") is not None],
            key=lambda r: float(r.get("net_profit") or 0),
            reverse=True,
        )
        complete_profit = bool(business_profit.get("complete"))
        title = (
            "💰 <b>30 kunlik sof foyda</b>" if complete_profit else "💰 <b>Ma’lum tannarx bo‘yicha hisob</b>"
        ) if lang == "uz" else (
            "💰 <b>Чистая прибыль за 30 дней</b>" if complete_profit else "💰 <b>Расчёт по известной себестоимости</b>"
        )
        summary = [
            f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
            "",
            *_format_profit_bridge_lines(business_profit, lang=lang),
            f"📌 Tannarx qamrovi: <b>{float(summary_stats['coverage']) * 100:.1f}%</b>"
            if lang == "uz"
            else f"📌 Покрытие себестоимостью: <b>{float(summary_stats['coverage']) * 100:.1f}%</b>",
            (
                "ℹ️ Tovar foydasi soliqdan keyin, lekin umumiy Uzum va tashqi xarajatlarni taqsimlamasdan ko‘rsatiladi."
                if lang == "uz"
                else "ℹ️ Прибыль товара показана после налога, но без распределения общих расходов Uzum и внешних расходов."
            ),
        ]
        if summary_stats["missing_count"]:
            summary.append(f"⚠️ Uzum tannarx bermagan SKU: <b>{summary_stats['missing_count']}</b>" if lang == "uz" else f"⚠️ Uzum не передал себестоимость: <b>{summary_stats['missing_count']}</b> SKU")
            summary.append(
                "ℹ️ Natija yakuniy sof foyda emas: Uzum barcha SKU tannarxini bermagan."
                if lang == "uz"
                else "ℹ️ Это не окончательная чистая прибыль: Uzum передал себестоимость не для всех SKU."
            )
        if not business_profit.get("uzum_expenses_available"):
            summary.append(
                "⚠️ Uzum xarajatlari vaqtincha olinmadi; sof foyda to‘liq emas."
                if lang == "uz"
                else "⚠️ Расходы Uzum временно недоступны; чистая прибыль неполная."
            )
        summary.append(
            "⚙️ Xarajatlarni o‘zgartirish: <code>/finance_settings</code>"
            if lang == "uz"
            else "⚙️ Изменить налоги и расходы: <code>/finance_settings</code>"
        )
        items: list[str] = []
        for idx, r in enumerate(known, start=1):
            title_item = escape(_short_text(str(r.get("title") or r.get("sku") or "-"), 70))
            sku = escape(_short_text(str(r.get("sku") or ""), 55))
            net_profit = float(r.get("net_profit") or 0)
            roi = r.get("roi")
            roi_text = "—" if roi is None else f"{float(roi):.1f}%"
            if lang == "uz":
                items.append(
                    f"{idx}. <b>{title_item}</b>\n"
                    f"🔖 SKU: <code>{sku}</code>\n"
                    f"Tushum: {_format_money(float(r.get('revenue') or 0))} · To‘lovga: {_format_money(float(r.get('payout') or 0))}\n"
                    f"Tannarx: {_format_money(float(r.get('cost_total') or 0))} · Soliq: {_format_money(float(r.get('tax_expense') or 0))}\n"
                    f"💰 Tovar foydasi: <b>{_format_money(net_profit)}</b> · ROI: <b>{roi_text}</b>"
                )
            else:
                items.append(
                    f"{idx}. <b>{title_item}</b>\n"
                    f"🔖 SKU: <code>{sku}</code>\n"
                    f"Выручка: {_format_money(float(r.get('revenue') or 0))} · К выплате: {_format_money(float(r.get('payout') or 0))}\n"
                    f"Себестоимость: {_format_money(float(r.get('cost_total') or 0))} · Налог: {_format_money(float(r.get('tax_expense') or 0))}\n"
                    f"💰 Прибыль товара: <b>{_format_money(net_profit)}</b> · ROI: <b>{roi_text}</b>"
                )
        if not items:
            items = ["Uzum tannarx bergan savdolar hali yo‘q." if lang == "uz" else "Пока нет продаж, для которых Uzum передал себестоимость."]
        await send_paginated_list(message, kind="profit", title=title, summary=summary, items=items, section="sales", reply_markup=sales_menu_for_message(message))
    except Exception as e:
        await send_api_error(message, e)


@dp.message(F.text == "💰 Foyda")
@dp.message(F.text == "💰 Прибыль")
async def button_profit_report_near_unit(message: Message) -> None:
    await profit_report(message)


@dp.message(F.text == "📥 Tannarx Excel")
@dp.message(F.text == "📥 Себестоимость Excel")
@dp.message(F.text == "🔄 Uzum tannarxi")
@dp.message(F.text == "🔄 Себестоимость Uzum")
async def button_costs_excel_near_unit(message: Message, state: FSMContext) -> None:
    await state.clear()
    await show_uzum_cost_status(message)


@dp.message(F.text == "🧾 Unit iqtisodiyot")
@dp.message(F.text == "🧾 Юнит-экономика")
async def button_unit_economy(message: Message) -> None:
    await unit_economy(message)

@dp.message(F.text == "🌐 Barcha do‘konlar")
@dp.message(F.text == "🌐 Все магазины")
async def button_all_shops(message: Message) -> None:
    await balance_all_shops(message)


@dp.message(F.text == "🏆 Top tovarlar")
@dp.message(F.text == "🏆 Топ товаров")
async def button_top_products(message: Message) -> None:
    await top_products(message)


@dp.message(F.text == "🐢 Sotilmayapti")
@dp.message(F.text == "🐢 Не продаётся")
async def button_dead_stock(message: Message) -> None:
    await dead_stock(message)


@dp.message(F.text == "🌙 Ertalabki hisobot")
@dp.message(F.text == "🌙 Утренний отчёт")
@dp.message(F.text == "🌙 Qisqa hisobot")
@dp.message(F.text == "🌙 Краткий отчёт")
async def button_morning_report(message: Message) -> None:
    await morning_report(message)


@dp.message(F.text == "⚠️ Qoldiq prognozi")
@dp.message(F.text == "⚠️ Прогноз остатков")
async def button_smart_lowstock(message: Message) -> None:
    await smart_lowstock(message)


_daily_report_sent: set[tuple[int, str]] = set()


def init_daily_report_delivery_table() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS daily_report_delivery (
                telegram_id INTEGER NOT NULL,
                report_date TEXT NOT NULL,
                sent_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, report_date)
            )
            """
        )
        conn.commit()


def daily_report_was_sent(telegram_id: int, report_date: str) -> bool:
    init_daily_report_delivery_table()
    with db.connect() as conn:
        row = conn.execute(
            "SELECT 1 FROM daily_report_delivery WHERE telegram_id = ? AND report_date = ?",
            (int(telegram_id), str(report_date)),
        ).fetchone()
    return row is not None


def mark_daily_report_sent(telegram_id: int, report_date: str) -> None:
    init_daily_report_delivery_table()
    with db.connect() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO daily_report_delivery (telegram_id, report_date, sent_at)
            VALUES (?, ?, ?)
            """,
            (int(telegram_id), str(report_date), _dt_to_db(_utc_now()) or ""),
        )
        conn.execute(
            "DELETE FROM daily_report_delivery WHERE report_date < ?",
            ((datetime.now(UZT) - timedelta(days=60)).strftime("%Y-%m-%d"),),
        )
        conn.commit()


init_daily_report_delivery_table()


def _connected_users_basic() -> list[dict[str, Any]]:
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT telegram_id, uzum_token_encrypted, default_shop_id
            FROM users
            WHERE uzum_token_encrypted IS NOT NULL
            """
        ).fetchall()
    return [dict(row) for row in rows if has_active_subscription(int(row["telegram_id"]))]


async def _build_scheduled_period_text(
    telegram_id: int,
    client: UzumClient,
    *,
    days: int,
    kind: str,
) -> str:
    lang = get_user_language(telegram_id)
    now = datetime.now(UZT)
    date_to = int(now.timestamp() * 1000)
    date_from = int((now - timedelta(days=days)).timestamp() * 1000)
    shift = days * 24 * 60 * 60 * 1000
    stats, per_shop, shops_count = await _all_shops_finance_stats(telegram_id, client, date_from, date_to)
    previous, _, _ = await _all_shops_finance_stats(
        telegram_id,
        client,
        date_from - shift,
        date_to - shift,
    )
    revenue_change = _format_change(
        _percent_change(float(stats.get("revenue") or 0), float(previous.get("revenue") or 0)),
        lang=lang,
    )
    payout_change = _format_change(
        _percent_change(float(stats.get("payout_total") or 0), float(previous.get("payout_total") or 0)),
        lang=lang,
    )
    top = list(stats.get("top_products") or [])[:3]
    if lang == "uz":
        title = "📅 <b>Haftalik boshqaruv hisoboti</b>" if kind == "weekly" else "🗓 <b>Oylik boshqaruv hisoboti</b>"
        lines = [
            title,
            f"Davr: <b>{days} kun</b> | Do‘konlar: <b>{shops_count}</b>",
            "",
            f"🛒 Buyurtma: <b>{int(stats.get('orders') or 0)}</b> | 📦 Sotildi: <b>{float(stats.get('units') or 0):.0f}</b>",
            f"❌ Bekor: <b>{int(stats.get('cancelled') or 0)}</b> | ↩️ Qaytarish: <b>{float(stats.get('returns') or 0):.0f}</b>",
            f"💵 Tushum: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
            f"✅ To‘lovga: <b>{_format_money(float(stats.get('payout_total') or 0))}</b>",
            f"🔄 Oldingi davrga: tushum {revenue_change}, to‘lov {payout_change}",
        ]
        if per_shop:
            lines.extend(["", "<b>Do‘konlar bo‘yicha:</b>", *per_shop[:10]])
        if top:
            lines.extend(["", "🏆 <b>Top-3:</b>"])
            lines.extend(
                f"{index}. {escape(_short_text(str(item.get('title') or '—'), 55))} — {_format_money(float(item.get('revenue') or 0))}"
                for index, item in enumerate(top, start=1)
            )
        lines.extend(["", "💼 Keyingi qadamlar: /control_center"])
        return "\n".join(lines)

    title = "📅 <b>Еженедельный управленческий отчёт</b>" if kind == "weekly" else "🗓 <b>Ежемесячный управленческий отчёт</b>"
    lines = [
        title,
        f"Период: <b>{days} дней</b> | Магазинов: <b>{shops_count}</b>",
        "",
        f"🛒 Заказов: <b>{int(stats.get('orders') or 0)}</b> | 📦 Продано: <b>{float(stats.get('units') or 0):.0f}</b>",
        f"❌ Отмен: <b>{int(stats.get('cancelled') or 0)}</b> | ↩️ Возвратов: <b>{float(stats.get('returns') or 0):.0f}</b>",
        f"💵 Выручка: <b>{_format_money(float(stats.get('revenue') or 0))}</b>",
        f"✅ К выплате: <b>{_format_money(float(stats.get('payout_total') or 0))}</b>",
        f"🔄 К предыдущему периоду: выручка {revenue_change}, выплата {payout_change}",
    ]
    if per_shop:
        lines.extend(["", "<b>По магазинам:</b>", *per_shop[:10]])
    if top:
        lines.extend(["", "🏆 <b>Топ-3:</b>"])
        lines.extend(
            f"{index}. {escape(_short_text(str(item.get('title') or '—'), 55))} — {_format_money(float(item.get('revenue') or 0))}"
            for index, item in enumerate(top, start=1)
        )
    lines.extend(["", "💼 Следующие действия: /control_center"])
    return "\n".join(lines)


async def daily_report_loop() -> None:
    await asyncio.sleep(90)
    logging.info("Personal scheduled reports loop started")
    while True:
        try:
            now = datetime.now(UZT)
            for row in _connected_users_basic():
                telegram_id = int(row["telegram_id"])
                access_level = subscription_access_level(telegram_id)
                premium = access_level in {"admin", "paid"}
                settings = ensure_product_settings(telegram_id)
                due: list[tuple[str, str, int]] = []
                if int(settings.get("daily_enabled") or 0) and now.hour >= int(settings.get("daily_hour") or 0):
                    due.append(("daily", now.strftime("%Y-%m-%d"), 1))
                if (
                    premium
                    and int(settings.get("weekly_enabled") or 0)
                    and now.weekday() == int(settings.get("weekly_weekday") or 0)
                    and now.hour >= int(settings.get("weekly_hour") or 0)
                ):
                    iso = now.isocalendar()
                    due.append(("weekly", f"{iso.year}-W{iso.week:02d}", 7))
                if (
                    premium
                    and int(settings.get("monthly_enabled") or 0)
                    and now.day == int(settings.get("monthly_day") or 1)
                    and now.hour >= int(settings.get("monthly_hour") or 0)
                ):
                    due.append(("monthly", now.strftime("%Y-%m"), 30))
                if not due:
                    continue
                try:
                    token = cipher.decrypt(row["uzum_token_encrypted"])
                    client = UzumClient(token, UZUM_API_BASE_URL)
                except Exception:
                    logging.exception("Scheduled report: token failed user=%s", telegram_id)
                    continue
                for kind, period_key, days in due:
                    if scheduled_report_was_sent(telegram_id, kind, period_key):
                        continue
                    try:
                        text = (
                            await _build_morning_report_text(telegram_id, client)
                            if kind == "daily"
                            else await _build_scheduled_period_text(telegram_id, client, days=days, kind=kind)
                        )
                        await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
                        mark_scheduled_report_sent(telegram_id, kind, period_key)
                        if kind == "daily":
                            mark_daily_report_sent(telegram_id, period_key)
                        await asyncio.sleep(0.5)
                    except Exception:
                        logging.exception("Scheduled report failed user=%s kind=%s", telegram_id, kind)
        except Exception:
            logging.exception("Scheduled report loop error")
        await asyncio.sleep(600)


async def subscription_reminder_loop() -> None:
    await asyncio.sleep(120)
    logging.info(
        "Subscription draft loop started. Enabled: %s. Admin digest hour UZT: %s. Windows: %s",
        SUBSCRIPTION_REMINDERS,
        SUBSCRIPTION_ADMIN_DIGEST_HOUR_UZT,
        SUBSCRIPTION_REMINDER_DAYS,
    )
    while True:
        try:
            if SUBSCRIPTION_REMINDERS:
                now_uzt = datetime.now(UZT)
                today_key = now_uzt.strftime("%Y-%m-%d")
                last_digest_day = _subscription_automation_state_get("last_admin_digest_day")
                if now_uzt.hour >= SUBSCRIPTION_ADMIN_DIGEST_HOUR_UZT and last_digest_day != today_key:
                    summary = refresh_subscription_reminder_queue()
                    rows = list_pending_subscription_reminders(100)
                    digest = build_subscription_action_digest(rows, int(summary.get("created") or 0))
                    sent_to_admin = False
                    for admin_id in sorted(ADMIN_IDS):
                        try:
                            await bot.send_message(
                                admin_id,
                                digest,
                                reply_markup=subscription_queue_markup(rows) or ADMIN_PANEL_MENU_RU,
                            )
                            sent_to_admin = True
                            await asyncio.sleep(0.2)
                        except Exception:
                            logging.exception("Subscription action digest failed for admin=%s", admin_id)
                    if sent_to_admin:
                        _subscription_automation_state_set("last_admin_digest_day", today_key)
                    elif not ADMIN_IDS:
                        logging.error("Subscription action digest skipped: ADMIN_IDS is empty")
        except Exception:
            logging.exception("Subscription draft loop error")
        await asyncio.sleep(3600)


# --- OPTIMIZED WATCHERS: защита от 429 Too Many Requests ---
# Если один и тот же Uzum API-токен / магазин подключён у нескольких Telegram-пользователей
# (например, владелец и жена), старые watcher-функции делали одинаковый запрос для каждого пользователя.
# Ниже мы переопределяем check_*_once: один запрос на связку token+shop, потом рассылка всем пользователям группы.
def connected_watch_groups(
    *setting_fields: str,
    access_feature: str = "premium",
) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for row in connected_users_for_order_watch(access_feature):
        telegram_id = int(row["telegram_id"])
        if setting_fields and not any(
            product_setting_enabled(telegram_id, field)
            for field in setting_fields
        ):
            continue
        shop_id = int(row.get("shop_id") or row["default_shop_id"])
        shop_title = str(row.get("shop_title") or "").strip()
        encrypted_token = row["uzum_token_encrypted"]
        # Fernet шифрует один и тот же токен каждый раз по-разному, поэтому
        # группировка по зашифрованной строке не убирала дубли. Сравниваем только
        # безопасные отпечатки расшифрованных токенов и сам токен нигде не сохраняем.
        try:
            raw_token = cipher.decrypt(encrypted_token)
            token_fingerprint = hashlib.sha256(raw_token.encode("utf-8")).hexdigest()
        except Exception:
            token_fingerprint = hashlib.sha256(str(encrypted_token).encode("utf-8")).hexdigest()
        key = hashlib.sha1(f"{shop_id}:{token_fingerprint}".encode("utf-8")).hexdigest()
        if key not in groups:
            groups[key] = {
                "watch_key": key,
                "shop_id": shop_id,
                "shop_title": shop_title,
                "uzum_token_encrypted": encrypted_token,
                "telegram_ids": [],
                "shop_titles_by_user": {},
            }
        if telegram_id not in groups[key]["telegram_ids"]:
            groups[key]["telegram_ids"].append(telegram_id)
        if shop_title:
            groups[key]["shop_titles_by_user"][telegram_id] = shop_title
    return list(groups.values())


_WATCH_STOCK_CACHE: dict[str, tuple[float, list[dict[str, Any]]]] = {}
_WATCH_STOCK_CACHE_LOCKS: dict[str, tuple[Any, asyncio.Lock]] = {}


def _watch_group_key(group: dict[str, Any]) -> str:
    stored = str(group.get("watch_key") or "")
    if stored:
        return stored
    return hashlib.sha1(
        f"{group.get('shop_id')}:{group.get('uzum_token_encrypted')}".encode("utf-8")
    ).hexdigest()


def _watch_stock_cache_key(client: UzumClient, shop_id: int) -> str:
    headers = getattr(client, "headers", {}) or {}
    token = str(headers.get("Authorization") or "")
    token_fingerprint = hashlib.sha1(token.encode("utf-8")).hexdigest() if token else str(id(client))
    return f"{int(shop_id)}:{token_fingerprint}"


def _watch_stock_cache_lock(cache_key: str) -> asyncio.Lock:
    loop = asyncio.get_running_loop()
    stored = _WATCH_STOCK_CACHE_LOCKS.get(cache_key)
    if stored is None or stored[0] is not loop:
        lock = asyncio.Lock()
        _WATCH_STOCK_CACHE_LOCKS[cache_key] = (loop, lock)
        return lock
    return stored[1]


async def _load_watch_stock_rows(
    client: UzumClient,
    shop_id: int,
) -> list[dict[str, Any]]:
    cache_key = _watch_stock_cache_key(client, shop_id)
    now = time.monotonic()
    cached = _WATCH_STOCK_CACHE.get(cache_key)
    if cached and now - cached[0] < WATCH_STOCK_CACHE_SECONDS:
        return cached[1]

    async with _watch_stock_cache_lock(cache_key):
        now = time.monotonic()
        cached = _WATCH_STOCK_CACHE.get(cache_key)
        if cached and now - cached[0] < WATCH_STOCK_CACHE_SECONDS:
            return cached[1]
        rows = await load_sku_rows(client, shop_id, max_pages=50)
        _WATCH_STOCK_CACHE[cache_key] = (time.monotonic(), rows)
        if len(_WATCH_STOCK_CACHE) > 100:
            stale_before = time.monotonic() - WATCH_STOCK_CACHE_SECONDS * 3
            for key, (saved_at, _) in list(_WATCH_STOCK_CACHE.items()):
                if saved_at < stale_before:
                    _WATCH_STOCK_CACHE.pop(key, None)
                    _WATCH_STOCK_CACHE_LOCKS.pop(key, None)
        logging.info(
            "Watcher stock snapshot refreshed shop=%s skus=%s cache=%ss",
            shop_id,
            len(rows),
            WATCH_STOCK_CACHE_SECONDS,
        )
        return rows


async def check_new_orders_once() -> None:
    for group in connected_watch_groups("notify_orders"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(x) for x in group["telegram_ids"]]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            data = await client.get_fbs_orders(shop_id, status="CREATED", page=0, size=20)
            items = extract_items(data)
        except Exception as error:
            _log_watcher_api_failure(
                "Order watcher",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        keys_now = [order_key(item) for item in items]
        for telegram_id in telegram_ids:
            scope = (telegram_id, shop_id)
            known = _seen_order_keys_by_scope.setdefault(scope, set())
            if scope not in _orders_watch_initialized_scopes:
                known.update(keys_now)
                _orders_watch_initialized_scopes.add(scope)
                logging.info(
                    "Order watcher initialized for user=%s shop=%s orders=%s",
                    telegram_id, shop_id, len(keys_now),
                )
                continue

            new_items = [item for item, key in zip(items, keys_now) if key not in known]
            known.update(keys_now)
            if len(known) > 1000:
                _seen_order_keys_by_scope[scope] = set(keys_now)
            if not new_items:
                continue

            lines = [format_order_line(item) for item in new_items[:5]]
            more = "" if len(new_items) <= 5 else f"\n\nЕщё новых заказов: {len(new_items) - 5}"
            text = (
                f"🔔 <b>Новый заказ CREATED</b>\n"
                f"Магазин: <code>{shop_id}</code>\n"
                f"Новых заказов: <b>{len(new_items)}</b>\n\n"
                + "\n\n".join(lines)
                + more
                + "\n\nОткрыть список: <code>/orders</code>"
            )
            try:
                await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
                await asyncio.sleep(0.15)
            except Exception:
                logging.exception("Order watcher: failed to send notification to %s", telegram_id)
        await asyncio.sleep(0.5)


async def check_low_stock_once() -> None:
    for group in connected_watch_groups("notify_low_stock"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(x) for x in group["telegram_ids"]]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows = await _load_watch_stock_rows(client, shop_id)
        except Exception as error:
            _log_watcher_api_failure(
                "Low stock watcher",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        for telegram_id in telegram_ids:
            threshold = max(0, int(ensure_product_settings(telegram_id).get("low_stock_threshold") or 0))
            low_rows = [r for r in rows if r.get("total") is not None and int(r.get("total") or 0) <= threshold]
            low_keys_now = [stock_row_key(r) for r in low_rows]
            known = _seen_low_stock_keys_by_user.setdefault(telegram_id, set())
            if telegram_id not in _low_stock_watch_initialized:
                known.update(low_keys_now)
                _low_stock_watch_initialized.add(telegram_id)
                logging.info(
                    "Low stock watcher initialized for user=%s shop=%s low_skus=%s threshold=%s",
                    telegram_id, shop_id, len(low_keys_now), threshold,
                )
                continue

            new_low_rows = [r for r, key in zip(low_rows, low_keys_now) if key not in known]
            _seen_low_stock_keys_by_user[telegram_id] = set(low_keys_now)
            if not new_low_rows:
                continue

            lines = [format_sku_stock_line(item, mode="all") for item in new_low_rows[:10]]
            more = "" if len(new_low_rows) <= 10 else f"\n\nЕщё SKU с низким остатком: {len(new_low_rows) - 10}"
            text = (
                f"📉 <b>Товар заканчивается</b>\n"
                f"Магазин: <code>{shop_id}</code>\n"
                f"Порог: ≤ <b>{threshold}</b> шт.\n"
                f"Новых позиций с низким остатком: <b>{len(new_low_rows)}</b>\n\n"
                + "\n\n".join(lines)
                + more
                + f"\n\nПоказать все низкие остатки: <code>/lowstock {threshold}</code>"
            )
            try:
                await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
                await asyncio.sleep(0.15)
            except Exception:
                logging.exception("Low stock watcher: failed to send notification to %s", telegram_id)
        await asyncio.sleep(0.5)


async def check_out_of_stock_once() -> None:
    for group in connected_watch_groups("notify_out_of_stock"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(x) for x in group["telegram_ids"]]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows = await _load_watch_stock_rows(client, shop_id)
        except Exception as error:
            _log_watcher_api_failure(
                "Out of stock watcher",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        zero_rows = [r for r in rows if r.get("total") is not None and int(r.get("total") or 0) == 0]
        zero_keys_now = [stock_row_key(r) for r in zero_rows]

        for telegram_id in telegram_ids:
            known = _seen_out_of_stock_keys_by_user.setdefault(telegram_id, set())
            if telegram_id not in _out_of_stock_watch_initialized:
                known.update(zero_keys_now)
                _out_of_stock_watch_initialized.add(telegram_id)
                logging.info(
                    "Out of stock watcher initialized for user=%s shop=%s zero_skus=%s",
                    telegram_id, shop_id, len(zero_keys_now),
                )
                continue

            new_zero_rows = [r for r, key in zip(zero_rows, zero_keys_now) if key not in known]
            _seen_out_of_stock_keys_by_user[telegram_id] = set(zero_keys_now)
            if not new_zero_rows:
                continue

            lines = [format_sku_stock_line(item, mode="all") for item in new_zero_rows[:10]]
            more = "" if len(new_zero_rows) <= 10 else f"\n\nЕщё SKU с нулевым остатком: {len(new_zero_rows) - 10}"
            text = (
                "❌ <b>Товар закончился</b>\n"
                f"Магазин: <code>{shop_id}</code>\n"
                f"Новых позиций с остатком 0: <b>{len(new_zero_rows)}</b>\n\n"
                + "\n\n".join(lines)
                + more
                + "\n\nПоказать остатки: <code>/stock</code>"
            )
            try:
                await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
                await asyncio.sleep(0.15)
            except Exception:
                logging.exception("Out of stock watcher: failed to send notification to %s", telegram_id)
        await asyncio.sleep(0.5)


async def check_new_sales_once() -> None:
    date_from, date_to = _today_range_ms()
    for group in connected_watch_groups("notify_sales", "notify_cancellations"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(x) for x in group["telegram_ids"]]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows, first_response = await _load_finance_orders(
                client,
                shop_id,
                date_from_ms=date_from,
                date_to_ms=date_to,
                max_pages=5,
                page_size=100,
            )
        except Exception:
            logging.exception("Sales watcher optimized: failed to check shop=%s users=%s", shop_id, telegram_ids)
            await asyncio.sleep(3)
            continue

        keys_now = [sale_key(item) for item in rows]
        for telegram_id in telegram_ids:
            sales_enabled = product_setting_enabled(telegram_id, "notify_sales")
            cancellations_enabled = product_setting_enabled(telegram_id, "notify_cancellations")
            known = _seen_sale_keys_by_user.setdefault(telegram_id, set())
            if telegram_id not in _sales_watch_initialized:
                known.update(keys_now)
                _sales_watch_initialized.add(telegram_id)
                logging.info(
                    "Sales watcher initialized for user=%s shop=%s sales_rows=%s",
                    telegram_id, shop_id, len(keys_now),
                )
                continue

            new_rows = [item for item, key in zip(rows, keys_now) if key not in known]
            known.update(keys_now)
            if len(known) > 3000:
                _seen_sale_keys_by_user[telegram_id] = set(keys_now)
            if not new_rows:
                continue

            for item in new_rows[:10]:
                try:
                    await bot.send_message(telegram_id, build_new_sale_message(item, shop_id=shop_id, lang=get_user_language(telegram_id)), reply_markup=main_menu_for_user(telegram_id))
                    await asyncio.sleep(0.15)
                except Exception:
                    logging.exception("Sales watcher: failed to send sale notification to %s", telegram_id)

            if len(new_rows) > 10:
                try:
                    await bot.send_message(
                        telegram_id,
                        f"➕ Ещё новых строк продаж: <b>{len(new_rows) - 10}</b>\nПодробно: <code>/balance</code>",
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                except Exception:
                    logging.exception("Sales watcher: failed to send summary notification to %s", telegram_id)
        await asyncio.sleep(0.5)


async def check_stock_change_once() -> None:
    for group in connected_watch_groups("notify_stock_change"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(x) for x in group["telegram_ids"]]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows = await _load_watch_stock_rows(client, shop_id)
            snapshot_now = _stock_change_snapshot(rows)
        except Exception as error:
            _log_watcher_api_failure(
                "Stock change watcher",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        for telegram_id in telegram_ids:
            previous = _stock_snapshot_by_user.setdefault(telegram_id, {})
            if telegram_id not in _stock_change_watch_initialized:
                _stock_snapshot_by_user[telegram_id] = snapshot_now
                _stock_change_watch_initialized.add(telegram_id)
                logging.info(
                    "Stock change watcher initialized for user=%s shop=%s skus=%s",
                    telegram_id, shop_id, len(snapshot_now),
                )
                continue

            decreased: list[tuple[str, dict[str, Any], dict[str, Any]]] = []
            for key, after in snapshot_now.items():
                before = previous.get(key)
                if not before:
                    continue
                before_total = int(before.get("total") or 0)
                after_total = int(after.get("total") or 0)
                before_fbo = int(before.get("fbo") or 0)
                after_fbo = int(after.get("fbo") or 0)
                before_fbs = int(before.get("fbs") or 0)
                after_fbs = int(after.get("fbs") or 0)
                if after_total < before_total or after_fbo < before_fbo or after_fbs < before_fbs:
                    decreased.append((key, before, after))

            _stock_snapshot_by_user[telegram_id] = snapshot_now
            if not decreased:
                continue

            lines = [_format_stock_change_line(key, before, after) for key, before, after in decreased[:10]]
            more = "" if len(decreased) <= 10 else f"\n\nЕщё изменений: {len(decreased) - 10}"
            text = (
                "📦 <b>Изменение остатков</b>\n"
                f"Магазин: <code>{shop_id}</code>\n"
                "Уменьшился остаток по SKU. Это может быть продажа, резерв, списание или изменение склада.\n\n"
                + "\n\n".join(lines)
                + more
                + "\n\nПроверить остатки: <code>/stock</code>"
            )
            try:
                await bot.send_message(telegram_id, text, reply_markup=main_menu_for_user(telegram_id))
                await asyncio.sleep(0.15)
            except Exception:
                logging.exception("Stock change watcher: failed to send notification to %s", telegram_id)
        await asyncio.sleep(0.5)


# --- Уведомления об отменах через Finance API ---
# Отмена в Uzum чаще всего не появляется как новый заказ, а меняет статус уже существующей
# финансовой строки на CANCELED / PARTIALLY_CANCELED. Поэтому обычный watcher новых продаж
# может её не прислать. Этот блок отслеживает именно изменение статуса.
CANCEL_NOTIFICATIONS = os.getenv("CANCEL_NOTIFICATIONS", "1").strip().lower() in {"1", "true", "yes", "on", "да"}
_sale_status_by_user: dict[int, dict[str, str]] = {}


def finance_identity_key(item: dict[str, Any]) -> str:
    """Стабильный ключ строки продажи без количества/суммы/статуса.

    Нужен, чтобы увидеть изменение статуса PROCESSING -> CANCELED у той же продажи.
    """
    parts: list[str] = []
    order_id = _finance_order_id(item)
    sale_id = _finance_sale_id(item)
    sku = str(_deep_pick_value(item, ("skuId", "sku_id", "skuTitle", "skuName", "barcode")) or "-")
    title = _finance_title(item)
    for label, value in (("order", order_id), ("sale", sale_id), ("sku", sku), ("title", title)):
        if value not in (None, "", "-"):
            parts.append(f"{label}:{value}")
    if parts:
        return "|".join(parts)
    raw = json.dumps(item, ensure_ascii=False, sort_keys=True, default=str)
    return "hash:" + hashlib.sha1(raw.encode("utf-8")).hexdigest()


def build_cancel_message(item: dict[str, Any], shop_id: int | None = None, lang: str = "ru") -> str:
    lang = normalize_lang(lang)
    title = escape(_finance_title(item))
    sku = escape(_finance_sku_title(item))
    qty = _finance_qty(item)
    unit_price = _deep_pick_number(item, ("sellPrice", "soldPrice", "price", "skuPrice", "productPrice"))
    status = escape(_finance_status(item))
    order_id = escape(_finance_order_id(item))
    sale_id = escape(_finance_sale_id(item))
    date_text = escape(_format_finance_date(_finance_date_value(item)))

    if lang == "uz":
        shop_line = f"🏪 Do‘kon: <code>{shop_id}</code>\n" if shop_id is not None else ""
        price_line = f"💵 Summa: <b>{_format_money(float(unit_price or 0))}</b>\n" if unit_price is not None else ""
        return (
            "❌ <b>Buyurtma bekor qilindi</b>\n\n"
            + shop_line +
            f"📦 Tovar: <b>{title}</b>\n"
            f"🔖 SKU: <code>{sku}</code>\n"
            f"🔢 Soni: <b>{qty:g} dona</b>\n\n"
            + price_line +
            f"🆔 Buyurtma: <code>{order_id}</code>\n"
            f"📌 Status: <code>{status}</code>\n"
            f"🕒 Sana: {date_text}"
        )

    shop_line = f"🏪 Магазин: <code>{shop_id}</code>\n" if shop_id is not None else ""
    price_line = f"💵 Сумма: <b>{_format_money(float(unit_price or 0))}</b>\n" if unit_price is not None else ""
    return (
        "❌ <b>Отмена заказа</b>\n\n"
        + shop_line +
        f"📦 Товар: <b>{title}</b>\n"
        f"🔖 SKU: <code>{sku}</code>\n"
        f"🔢 Кол-во: <b>{qty:g} шт.</b>\n\n"
        + price_line +
        f"🆔 Заказ: <code>{order_id}</code>\n"
        f"📌 Статус: <code>{status}</code>\n"
        f"🕒 Дата: {date_text}"
    )


# Переопределяем watcher продаж: теперь он присылает и новые продажи, и новые отмены.
async def check_new_sales_once() -> None:
    date_from, date_to = _today_range_ms()
    for group in connected_watch_groups():
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(x) for x in group["telegram_ids"]]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows, first_response = await _load_finance_orders(
                client,
                shop_id,
                date_from_ms=date_from,
                date_to_ms=date_to,
                max_pages=5,
                page_size=100,
            )
        except Exception:
            logging.exception("Sales watcher optimized: failed to check shop=%s users=%s", shop_id, telegram_ids)
            await asyncio.sleep(3)
            continue

        keys_now = [sale_key(item) for item in rows]
        identity_status_now = {finance_identity_key(item): _finance_status(item) for item in rows}

        for telegram_id in telegram_ids:
            known = _seen_sale_keys_by_user.setdefault(telegram_id, set())
            status_memory = _sale_status_by_user.setdefault(telegram_id, {})

            # Первый проход: запоминаем текущее состояние, чтобы не прислать старые продажи/отмены.
            if telegram_id not in _sales_watch_initialized:
                known.update(keys_now)
                status_memory.update(identity_status_now)
                _sales_watch_initialized.add(telegram_id)
                logging.info(
                    "Sales watcher initialized for user=%s shop=%s sales_rows=%s",
                    telegram_id, shop_id, len(keys_now),
                )
                continue

            cancel_rows: list[dict[str, Any]] = []
            if CANCEL_NOTIFICATIONS:
                for item in rows:
                    ident = finance_identity_key(item)
                    current_status = _finance_status(item)
                    previous_status = status_memory.get(ident)
                    # Уведомляем, если строка стала отменённой после предыдущей проверки.
                    if _is_cancelled_status(current_status) and not _is_cancelled_status(str(previous_status or "")):
                        cancel_rows.append(item)
                status_memory.update(identity_status_now)

            # Новые строки продаж. Отменённые строки не отправляем как "новая продажа".
            new_rows = [
                item for item, key in zip(rows, keys_now)
                if key not in known and not _is_cancelled_status(_finance_status(item))
            ]
            known.update(keys_now)
            if len(known) > 3000:
                _seen_sale_keys_by_user[telegram_id] = set(keys_now)

            for item in new_rows[:10]:
                try:
                    await bot.send_message(
                        telegram_id,
                        build_new_sale_message(item, shop_id=shop_id, lang=get_user_language(telegram_id)),
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                    await asyncio.sleep(0.15)
                except Exception:
                    logging.exception("Sales watcher: failed to send sale notification to %s", telegram_id)

            for item in cancel_rows[:10]:
                try:
                    await bot.send_message(
                        telegram_id,
                        build_cancel_message(item, shop_id=shop_id, lang=get_user_language(telegram_id)),
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                    await asyncio.sleep(0.15)
                except Exception:
                    logging.exception("Sales watcher: failed to send cancel notification to %s", telegram_id)

            total_extra = max(0, len(new_rows) - 10) + max(0, len(cancel_rows) - 10)
            if total_extra:
                try:
                    await bot.send_message(
                        telegram_id,
                        f"➕ Ещё новых событий: <b>{total_extra}</b>\nПодробно: <code>/balance</code>",
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                except Exception:
                    logging.exception("Sales watcher: failed to send summary notification to %s", telegram_id)
        await asyncio.sleep(0.5)


@dp.message(Command("cancel_notify_status", "cancellations_notify_status"))
async def cancel_notify_status(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    req = await require_connection(message)
    if req is None:
        return
    _, _, shop_id = req
    lang = get_user_language(telegram_id)
    enabled = product_setting_enabled(telegram_id, "notify_cancellations")
    if normalize_lang(lang) == "uz":
        text = (
            "❌ <b>Bekor qilingan buyurtmalar xabarnomasi</b>\n\n"
            f"Holat: {'✅ yoqilgan' if enabled else '❌ o‘chirilgan'}\n"
            f"Do‘kon: <code>{shop_id}</code>\n"
            f"Tekshiruv: har <b>{max(60, SALE_CHECK_INTERVAL_SECONDS)}</b> soniyada\n\n"
            "Bot yangi bekor qilingan buyurtmalarni Finance API orqali kuzatadi."
        )
    else:
        text = (
            "❌ <b>Уведомления об отменах</b>\n\n"
            f"Статус: {'✅ включены' if enabled else '❌ выключены'}\n"
            f"Магазин: <code>{shop_id}</code>\n"
            f"Проверка каждые: <b>{max(60, SALE_CHECK_INTERVAL_SECONDS)}</b> сек.\n\n"
            "Бот отслеживает новые отмены через Finance API."
        )
    await message.answer(text, reply_markup=main_menu_for_user(telegram_id))


# --- Персональные уведомления, расписание, расходы и поставка ---
AUTOMATION_BOOL_FIELDS = {
    "notify_orders",
    "notify_sales",
    "notify_low_stock",
    "notify_out_of_stock",
    "notify_cancellations",
    "notify_reviews",
    "notify_stock_change",
    "notify_losses",
    "notify_defects",
    "notify_fbo_acceptance",
    "notify_supply_reminders",
    "notify_return_pickup",
    "daily_enabled",
    "weekly_enabled",
    "monthly_enabled",
}

PRODUCT_NUMERIC_LIMITS: dict[str, tuple[float, float]] = {
    "daily_hour": (0, 23),
    "weekly_weekday": (1, 7),
    "weekly_hour": (0, 23),
    "monthly_day": (1, 28),
    "monthly_hour": (0, 23),
    "low_stock_threshold": (0, 100000),
    "lead_time_days": (0, 90),
    "safety_days": (0, 90),
    "target_cover_days": (1, 365),
}

FINANCE_NUMERIC_LIMITS: dict[str, tuple[float, float]] = {
    "tax_percent": (0, 100),
    "advertising_monthly": (0, 10_000_000_000),
    "storage_monthly": (0, 10_000_000_000),
    "other_monthly": (0, 10_000_000_000),
}


def _toggle_icon(value: Any) -> str:
    return "✅" if bool(int(value or 0)) else "▫️"


def _weekday_name(value: Any, lang: str) -> str:
    index = max(0, min(6, int(value or 0)))
    ru = ("Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс")
    uz = ("Du", "Se", "Ch", "Pa", "Ju", "Sh", "Ya")
    return (uz if normalize_lang(lang) == "uz" else ru)[index]


def _sales_mode_label(settings: dict[str, Any], lang: str, *, compact: bool = False) -> str:
    mode = _sales_notification_mode_from_settings(settings)
    uz = normalize_lang(lang) == "uz"
    labels = {
        "instant": ("⚡ Сразу", "⚡ Darhol"),
        "hourly": ("🕐 Сводка раз в час", "🕐 Har soatda hisobot"),
        "off": ("🔕 Выключены", "🔕 O‘chirilgan"),
    }
    ru_text, uz_text = labels[mode]
    text = uz_text if uz else ru_text
    if compact and mode == "hourly":
        return "🕐 За час" if not uz else "🕐 Har soat"
    return text


NOTIFICATION_SECTION_FIELDS: dict[str, tuple[str, ...]] = {
    "sales": ("notify_cancellations", "notify_orders"),
    "stock": (
        "notify_low_stock",
        "notify_out_of_stock",
        "notify_stock_change",
        "notify_losses",
        "notify_defects",
        "notify_fbo_acceptance",
        "notify_supply_reminders",
        "notify_return_pickup",
    ),
    "reports": ("daily_enabled", "weekly_enabled", "monthly_enabled"),
}


def _enabled_count(settings: dict[str, Any], fields: Iterable[str]) -> tuple[int, int]:
    field_list = list(fields)
    enabled = sum(1 for field in field_list if bool(int(settings.get(field) or 0)))
    return enabled, len(field_list)


def notification_hub_text(telegram_id: int) -> str:
    lang = get_user_language(telegram_id)
    row = ensure_product_settings(telegram_id)
    shop_count = len(connected_shop_ids_for_user(telegram_id))
    stock_on, stock_total = _enabled_count(row, NOTIFICATION_SECTION_FIELDS["stock"])
    reports_on, reports_total = _enabled_count(row, NOTIFICATION_SECTION_FIELDS["reports"])
    if lang == "uz":
        return (
            "🔔 <b>Xabarnomalar</b>\n\n"
            f"💸 Savdo: <b>{_sales_mode_label(row, lang)}</b>\n"
            f"📦 Ombor va yetkazish: <b>{stock_on}/{stock_total} yoqilgan</b>\n"
            f"📅 Avtohisobotlar: <b>{reports_on}/{reports_total} yoqilgan</b>\n"
            f"⭐ Sharhlar: {_toggle_icon(row.get('notify_reviews'))}\n\n"
            f"🏪 Tezkor xabarnomalar barcha ulangan do‘konlarda ishlaydi: <b>{shop_count}</b>\n\n"
            "Sozlash uchun kerakli guruhni tanlang."
        )
    return (
        "🔔 <b>Уведомления</b>\n\n"
        f"💸 Продажи: <b>{_sales_mode_label(row, lang)}</b>\n"
        f"📦 Склад и поставки: <b>{stock_on}/{stock_total} включено</b>\n"
        f"📅 Автоотчёты: <b>{reports_on}/{reports_total} включено</b>\n"
        f"⭐ Отзывы: {_toggle_icon(row.get('notify_reviews'))}\n\n"
        f"🏪 Оперативные уведомления работают для всех подключённых магазинов: <b>{shop_count}</b>\n\n"
        "Выберите группу — увидите только относящиеся к ней настройки."
    )


def notification_hub_markup(telegram_id: int) -> InlineKeyboardMarkup:
    lang = get_user_language(telegram_id)
    row = ensure_product_settings(telegram_id)
    uz = lang == "uz"
    stock_on, stock_total = _enabled_count(row, NOTIFICATION_SECTION_FIELDS["stock"])
    reports_on, reports_total = _enabled_count(row, NOTIFICATION_SECTION_FIELDS["reports"])
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text=(f"💸 Savdo · {_sales_mode_label(row, lang, compact=True)}" if uz else f"💸 Продажи · {_sales_mode_label(row, lang, compact=True)}"),
            callback_data="notifyhub:sales",
        )],
        [
            InlineKeyboardButton(
                text=(f"📦 Ombor · {stock_on}/{stock_total}" if uz else f"📦 Склад · {stock_on}/{stock_total}"),
                callback_data="notifyhub:stock",
            ),
            InlineKeyboardButton(
                text=(f"📅 Hisobot · {reports_on}/{reports_total}" if uz else f"📅 Отчёты · {reports_on}/{reports_total}"),
                callback_data="notifyhub:reports",
            ),
        ],
        [InlineKeyboardButton(
            text=f"{_toggle_icon(row.get('notify_reviews'))} {'Sharhlar' if uz else 'Отзывы'}",
            callback_data="notifytoggle:main:notify_reviews",
        )],
    ])


def notification_section_text(telegram_id: int, section: str) -> str:
    lang = get_user_language(telegram_id)
    row = ensure_product_settings(telegram_id)
    uz = lang == "uz"
    if section == "sales":
        if uz:
            return (
                "💸 <b>Savdo va buyurtmalar</b>\n\n"
                f"Savdo xabarlari: <b>{_sales_mode_label(row, lang)}</b>\n"
                f"Bekor qilishlar: {_toggle_icon(row.get('notify_cancellations'))}\n"
                f"Yangi FBS buyurtmalar: {_toggle_icon(row.get('notify_orders'))}\n\n"
                "Kichik do‘kon uchun «Darhol», ko‘p savdo bo‘lsa «Har soat» rejimini tanlang."
            )
        return (
            "💸 <b>Продажи и заказы</b>\n\n"
            f"Продажи: <b>{_sales_mode_label(row, lang)}</b>\n"
            f"Отмены: {_toggle_icon(row.get('notify_cancellations'))}\n"
            f"Новые FBS-заказы: {_toggle_icon(row.get('notify_orders'))}\n\n"
            "Для небольшого магазина выберите «Сразу», для большого — «Сводка раз в час»."
        )
    if section == "stock":
        if uz:
            return (
                "📦 <b>Ombor va yetkazish</b>\n\n"
                f"Kam qoldiq: {_toggle_icon(row.get('notify_low_stock'))} (≤ {int(row.get('low_stock_threshold') or 0)})\n"
                f"Tovar tugashi: {_toggle_icon(row.get('notify_out_of_stock'))}\n"
                f"Qoldiq o‘zgarishi: {_toggle_icon(row.get('notify_stock_change'))}\n"
                f"Yo‘qotish: {_toggle_icon(row.get('notify_losses'))}\n"
                f"Yaroqsiz tovar: {_toggle_icon(row.get('notify_defects'))}\n"
                f"FBO qabuli va farqlar PDF: {_toggle_icon(row.get('notify_fbo_acceptance'))}\n"
                f"Yetkazish vaqti eslatmasi: {_toggle_icon(row.get('notify_supply_reminders'))}\n"
                f"Qaytarmani olib ketish eslatmasi: {_toggle_icon(row.get('notify_return_pickup'))}"
            )
        return (
            "📦 <b>Склад и поставки</b>\n\n"
            f"Низкий остаток: {_toggle_icon(row.get('notify_low_stock'))} (≤ {int(row.get('low_stock_threshold') or 0)})\n"
            f"Товар закончился: {_toggle_icon(row.get('notify_out_of_stock'))}\n"
            f"Изменение остатка: {_toggle_icon(row.get('notify_stock_change'))}\n"
            f"Потери: {_toggle_icon(row.get('notify_losses'))}\n"
            f"Брак: {_toggle_icon(row.get('notify_defects'))}\n"
            f"Приёмка FBO и PDF с расхождениями: {_toggle_icon(row.get('notify_fbo_acceptance'))}\n"
            f"Срок поставки: {_toggle_icon(row.get('notify_supply_reminders'))}\n"
            f"Забрать возврат: {_toggle_icon(row.get('notify_return_pickup'))}"
        )
    if uz:
        return (
            "📅 <b>Avtohisobotlar</b>\n\n"
            f"Kunlik: {_toggle_icon(row.get('daily_enabled'))} {int(row.get('daily_hour') or 0):02d}:00\n"
            f"Haftalik: {_toggle_icon(row.get('weekly_enabled'))} {_weekday_name(row.get('weekly_weekday'), lang)} {int(row.get('weekly_hour') or 0):02d}:00\n"
            f"Oylik: {_toggle_icon(row.get('monthly_enabled'))} {int(row.get('monthly_day') or 1)}-kun {int(row.get('monthly_hour') or 0):02d}:00\n\n"
            "Vaqt zonasi: Asia/Tashkent."
        )
    return (
        "📅 <b>Автоматические отчёты</b>\n\n"
        f"Ежедневный: {_toggle_icon(row.get('daily_enabled'))} {int(row.get('daily_hour') or 0):02d}:00\n"
        f"Еженедельный: {_toggle_icon(row.get('weekly_enabled'))} {_weekday_name(row.get('weekly_weekday'), lang)} {int(row.get('weekly_hour') or 0):02d}:00\n"
        f"Ежемесячный: {_toggle_icon(row.get('monthly_enabled'))} {int(row.get('monthly_day') or 1)} числа в {int(row.get('monthly_hour') or 0):02d}:00\n\n"
        "Часовой пояс: Asia/Tashkent."
    )


def notification_section_markup(telegram_id: int, section: str) -> InlineKeyboardMarkup:
    lang = get_user_language(telegram_id)
    row = ensure_product_settings(telegram_id)
    uz = lang == "uz"

    def toggle(field: str, ru: str, uz_text: str) -> InlineKeyboardButton:
        return InlineKeyboardButton(
            text=f"{_toggle_icon(row.get(field))} {uz_text if uz else ru}",
            callback_data=f"notifytoggle:{section}:{field}",
        )

    back = InlineKeyboardButton(
        text="⬅️ Xabarnomalar" if uz else "⬅️ Все уведомления",
        callback_data="notifyhub:main",
    )
    if section == "sales":
        rows = [
            [InlineKeyboardButton(
                text=f"💸 {_sales_mode_label(row, lang, compact=True)}",
                callback_data="salesmode:menu:sales",
            )],
            [
                toggle("notify_cancellations", "Отмены", "Bekor qilish"),
                toggle("notify_orders", "FBS-заказы", "FBS buyurtma"),
            ],
            [back],
        ]
    elif section == "stock":
        rows = [
            [
                toggle("notify_low_stock", "Низкий остаток", "Kam qoldiq"),
                toggle("notify_out_of_stock", "Закончился", "Tugadi"),
            ],
            [toggle("notify_stock_change", "Изменение остатка", "Qoldiq o‘zgarishi")],
            [toggle("notify_losses", "Потери", "Yo‘qotish"), toggle("notify_defects", "Брак", "Yaroqsiz")],
            [toggle("notify_fbo_acceptance", "Приёмка FBO", "FBO qabuli")],
            [
                toggle("notify_supply_reminders", "Срок поставки", "Yetkazish vaqti"),
                toggle("notify_return_pickup", "Забрать возврат", "Qaytarma"),
            ],
            [InlineKeyboardButton(
                text=("📉 Qoldiq chegarasi" if uz else "📉 Порог низкого остатка"),
                callback_data="notifyedit:stock:low_stock_threshold",
            )],
            [back],
        ]
    else:
        rows = [
            [toggle("daily_enabled", "Ежедневный", "Kunlik")],
            [toggle("weekly_enabled", "Еженедельный", "Haftalik")],
            [toggle("monthly_enabled", "Ежемесячный", "Oylik")],
            [InlineKeyboardButton(text=("🕘 Kunlik vaqt" if uz else "🕘 Время ежедневного"), callback_data="notifyedit:reports:daily_hour")],
            [
                InlineKeyboardButton(text=("📅 Hafta kuni" if uz else "📅 День недели"), callback_data="notifyedit:reports:weekly_weekday"),
                InlineKeyboardButton(text=("🕘 Hafta vaqti" if uz else "🕘 Время недельного"), callback_data="notifyedit:reports:weekly_hour"),
            ],
            [
                InlineKeyboardButton(text=("🗓 Oy kuni" if uz else "🗓 День месяца"), callback_data="notifyedit:reports:monthly_day"),
                InlineKeyboardButton(text=("🕘 Oy vaqti" if uz else "🕘 Время месячного"), callback_data="notifyedit:reports:monthly_hour"),
            ],
            [back],
        ]
    return InlineKeyboardMarkup(inline_keyboard=rows)


def automation_settings_text(telegram_id: int) -> str:
    lang = get_user_language(telegram_id)
    row = ensure_product_settings(telegram_id)
    if lang == "uz":
        return (
            "⚙️ <b>Shaxsiy avtomatlashtirish</b>\n\n"
            "Har bir tugma faqat sizning Telegram hisobingiz uchun ishlaydi.\n\n"
            f"💸 Savdo xabarlari: <b>{_sales_mode_label(row, lang)}</b>\n"
            f"🚫 Bekor qilish: {_toggle_icon(row.get('notify_cancellations'))}\n"
            f"🧭 Yangi yo‘qotish: {_toggle_icon(row.get('notify_losses'))}\n"
            f"🧪 Yangi yaroqsiz: {_toggle_icon(row.get('notify_defects'))}\n"
            f"🚚 FBO qabuli: {_toggle_icon(row.get('notify_fbo_acceptance'))}\n"
            f"⏰ Yetkazish vaqti: {_toggle_icon(row.get('notify_supply_reminders'))}\n"
            f"↩️ Qaytarmani olish: {_toggle_icon(row.get('notify_return_pickup'))}\n"
            f"📉 Kam qoldiq: {_toggle_icon(row.get('notify_low_stock'))} (≤ {int(row.get('low_stock_threshold') or 0)})\n"
            f"❌ Qoldiq tugashi: {_toggle_icon(row.get('notify_out_of_stock'))}\n"
            f"⭐ Sharhlar: {_toggle_icon(row.get('notify_reviews'))}\n"
            f"🔔 Yangi FBS buyurtma: {_toggle_icon(row.get('notify_orders'))}\n"
            f"📦 Qoldiq o‘zgarishi: {_toggle_icon(row.get('notify_stock_change'))}\n\n"
            f"🌅 Kunlik: {_toggle_icon(row.get('daily_enabled'))} {int(row.get('daily_hour') or 0):02d}:00\n"
            f"📅 Haftalik: {_toggle_icon(row.get('weekly_enabled'))} {_weekday_name(row.get('weekly_weekday'), lang)} {int(row.get('weekly_hour') or 0):02d}:00\n"
            f"🗓 Oylik: {_toggle_icon(row.get('monthly_enabled'))} {int(row.get('monthly_day') or 1)}-kun {int(row.get('monthly_hour') or 0):02d}:00\n\n"
            "Vaqt zonasi: <b>Asia/Tashkent</b>."
        )
    return (
        "⚙️ <b>Персональная автоматизация</b>\n\n"
        "Каждая настройка действует только для вашего Telegram-аккаунта.\n\n"
        f"💸 Уведомления о продажах: <b>{_sales_mode_label(row, lang)}</b>\n"
        f"🚫 Отмены: {_toggle_icon(row.get('notify_cancellations'))}\n"
        f"🧭 Новые потери: {_toggle_icon(row.get('notify_losses'))}\n"
        f"🧪 Новый брак: {_toggle_icon(row.get('notify_defects'))}\n"
        f"🚚 Приёмка FBO: {_toggle_icon(row.get('notify_fbo_acceptance'))}\n"
        f"⏰ Срок поставки: {_toggle_icon(row.get('notify_supply_reminders'))}\n"
        f"↩️ Забрать возврат: {_toggle_icon(row.get('notify_return_pickup'))}\n"
        f"📉 Низкие остатки: {_toggle_icon(row.get('notify_low_stock'))} (≤ {int(row.get('low_stock_threshold') or 0)})\n"
        f"❌ Товар закончился: {_toggle_icon(row.get('notify_out_of_stock'))}\n"
        f"⭐ Отзывы: {_toggle_icon(row.get('notify_reviews'))}\n"
        f"🔔 Новые FBS-заказы: {_toggle_icon(row.get('notify_orders'))}\n"
        f"📦 Изменения остатков: {_toggle_icon(row.get('notify_stock_change'))}\n\n"
        f"🌅 Ежедневный отчёт: {_toggle_icon(row.get('daily_enabled'))} {int(row.get('daily_hour') or 0):02d}:00\n"
        f"📅 Еженедельный: {_toggle_icon(row.get('weekly_enabled'))} {_weekday_name(row.get('weekly_weekday'), lang)} {int(row.get('weekly_hour') or 0):02d}:00\n"
        f"🗓 Ежемесячный: {_toggle_icon(row.get('monthly_enabled'))} {int(row.get('monthly_day') or 1)} числа в {int(row.get('monthly_hour') or 0):02d}:00\n\n"
        "Часовой пояс: <b>Asia/Tashkent</b>."
    )


def automation_settings_markup(telegram_id: int) -> InlineKeyboardMarkup:
    lang = get_user_language(telegram_id)
    row = ensure_product_settings(telegram_id)
    uz = lang == "uz"

    def toggle(field: str, ru: str, uz_text: str) -> InlineKeyboardButton:
        return InlineKeyboardButton(
            text=f"{_toggle_icon(row.get(field))} {uz_text if uz else ru}",
            callback_data=f"autotoggle:{field}",
        )

    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text=f"💸 {_sales_mode_label(row, lang, compact=True)}",
            callback_data="salesmode:menu",
        ), toggle("notify_cancellations", "Отмены", "Bekor")],
        [toggle("notify_losses", "Потери", "Yo‘qotish"), toggle("notify_defects", "Брак", "Yaroqsiz")],
        [toggle("notify_fbo_acceptance", "Приёмка FBO", "FBO qabuli")],
        [toggle("notify_supply_reminders", "Срок поставки", "Yetkazish vaqti"),
         toggle("notify_return_pickup", "Забрать возврат", "Qaytarma")],
        [toggle("notify_low_stock", "Низкий остаток", "Kam qoldiq"), toggle("notify_out_of_stock", "Закончился", "Tugadi")],
        [toggle("notify_reviews", "Отзывы", "Sharh"), toggle("notify_orders", "FBS-заказы", "FBS buyurtma")],
        [toggle("notify_stock_change", "Движение склада", "Ombor harakati")],
        [toggle("daily_enabled", "Ежедневный", "Kunlik"), toggle("weekly_enabled", "Еженедельный", "Haftalik")],
        [toggle("monthly_enabled", "Ежемесячный", "Oylik")],
        [InlineKeyboardButton(text=("🕘 Kunlik vaqt" if uz else "🕘 Время ежедневного"), callback_data="autoedit:daily_hour"),
         InlineKeyboardButton(text=("📅 Hafta kuni" if uz else "📅 День недели"), callback_data="autoedit:weekly_weekday")],
        [InlineKeyboardButton(text=("🕘 Hafta vaqti" if uz else "🕘 Время недельного"), callback_data="autoedit:weekly_hour"),
         InlineKeyboardButton(text=("🗓 Oy kuni" if uz else "🗓 День месяца"), callback_data="autoedit:monthly_day")],
        [InlineKeyboardButton(text=("🕘 Oy vaqti" if uz else "🕘 Время месячного"), callback_data="autoedit:monthly_hour"),
         InlineKeyboardButton(text=("📉 Qoldiq chegarasi" if uz else "📉 Порог остатка"), callback_data="autoedit:low_stock_threshold")],
    ])


def sales_mode_selection_text(telegram_id: int) -> str:
    lang = get_user_language(telegram_id)
    settings = ensure_product_settings(telegram_id)
    shop_count = len(connected_shop_ids_for_user(telegram_id))
    trial = subscription_access_level(telegram_id) == "trial"
    if normalize_lang(lang) == "uz":
        return (
            "💸 <b>Savdo xabarlari rejimi</b>\n\n"
            f"🏪 Kuzatuvda: <b>barcha ulangan do‘konlar ({shop_count})</b>\n"
            f"Hozir: <b>{_sales_mode_label(settings, lang)}</b>\n\n"
            "⚡ <b>Darhol</b> — har bir yangi savdo alohida xabar bo‘lib keladi. "
            "Savdosi kam do‘konlar uchun qulay.\n\n"
            "🕐 <b>Har soatda hisobot</b> — bot barcha do‘konlardagi savdolarni yig‘adi va "
            "soatiga bitta umumiy xabar yuboradi: buyurtmalar, dona, tushum, komissiya, "
            "logistika, to‘lov va do‘konlar kesimi."
            + (
                "\n\n🚫 Bekor qilingan buyurtmalar tanlangan rejimdan qat’i nazar alohida tezkor xabar bo‘lib keladi."
                if not trial
                else ""
            )
        )
    return (
        "💸 <b>Режим уведомлений о продажах</b>\n\n"
        f"🏪 Под наблюдением: <b>все подключённые магазины ({shop_count})</b>\n"
        f"Сейчас: <b>{_sales_mode_label(settings, lang)}</b>\n\n"
        "⚡ <b>Сразу</b> — каждая новая продажа приходит отдельным сообщением. "
        "Удобно для небольших магазинов.\n\n"
        "🕐 <b>Сводка раз в час</b> — бот объединяет продажи всех магазинов в одно сообщение: "
        "заказы, количество товаров, выручка, комиссия, логистика, выплата и разбивка по магазинам."
        + (
            "\n\n🚫 Отмены остаются отдельными срочными уведомлениями независимо от выбранного режима."
            if not trial
            else ""
        )
    )


def sales_mode_selection_markup(telegram_id: int, *, back_to: str = "automation") -> InlineKeyboardMarkup:
    lang = get_user_language(telegram_id)
    uz = normalize_lang(lang) == "uz"
    mode = get_sales_notification_mode(telegram_id)

    def option(value: str, ru: str, uz_text: str) -> InlineKeyboardButton:
        marker = "✅" if mode == value else "▫️"
        return InlineKeyboardButton(
            text=f"{marker} {uz_text if uz else ru}",
            callback_data=f"salesmode:{value}:{back_to}",
        )

    return InlineKeyboardMarkup(inline_keyboard=[
        [option("instant", "⚡ Сразу", "⚡ Darhol"), option("hourly", "🕐 За час", "🕐 Har soat")],
        [option("off", "🔕 Выключить", "🔕 O‘chirish")],
        [InlineKeyboardButton(text="⬅️ Ortga" if uz else "⬅️ Назад", callback_data=f"salesmode:back:{back_to}")],
    ])


def supply_settings_text(telegram_id: int) -> str:
    lang = get_user_language(telegram_id)
    row = ensure_product_settings(telegram_id)
    if lang == "uz":
        return (
            "🚚 <b>Yetkazib berish hisob-kitobi</b>\n\n"
            f"Yetkazish muddati: <b>{int(row.get('lead_time_days') or 0)} kun</b>\n"
            f"Xavfsizlik zaxirasi: <b>{int(row.get('safety_days') or 0)} kun</b>\n"
            f"Maqsadli qamrov: <b>{int(row.get('target_cover_days') or 30)} kun</b>\n"
            f"Kam qoldiq chegarasi: <b>{int(row.get('low_stock_threshold') or 0)} dona</b>\n\n"
            "Bot 7/14/30 kunlik savdo tezligini hisobga oladi."
        )
    return (
        "🚚 <b>Параметры планирования поставки</b>\n\n"
        f"Срок поставки: <b>{int(row.get('lead_time_days') or 0)} дн.</b>\n"
        f"Страховой запас: <b>{int(row.get('safety_days') or 0)} дн.</b>\n"
        f"Целевой запас: <b>{int(row.get('target_cover_days') or 30)} дн.</b>\n"
        f"Порог низкого остатка: <b>{int(row.get('low_stock_threshold') or 0)} шт.</b>\n\n"
        "Прогноз учитывает скорость продаж за 7, 14 и 30 дней."
    )


def supply_settings_markup(telegram_id: int) -> InlineKeyboardMarkup:
    lang = get_user_language(telegram_id)
    uz = lang == "uz"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=("🚚 Yetkazish kuni" if uz else "🚚 Срок поставки"), callback_data="autoedit:lead_time_days")],
        [InlineKeyboardButton(text=("🛡 Xavfsizlik zaxirasi" if uz else "🛡 Страховой запас"), callback_data="autoedit:safety_days")],
        [InlineKeyboardButton(text=("📦 Maqsadli zaxira" if uz else "📦 Целевой запас"), callback_data="autoedit:target_cover_days")],
        [InlineKeyboardButton(text=("📉 Kam qoldiq" if uz else "📉 Порог остатка"), callback_data="autoedit:low_stock_threshold")],
    ])


def finance_settings_text(telegram_id: int, shop_id: int) -> str:
    lang = get_user_language(telegram_id)
    row = ensure_finance_settings(telegram_id, shop_id)
    tax_percent = float(row.get("tax_percent") or 0)
    tax_name = _tax_setting_label(tax_percent, lang=lang)
    if lang == "uz":
        return (
            "🧮 <b>Sof foyda sozlamalari</b>\n\n"
            f"🏪 Do‘kon: <code>{shop_id}</code>\n"
            f"🏛 Soliq: <b>{escape(tax_name)}</b>\n"
            f"📣 Tashqi reklama / oy: <b>{_format_money(float(row.get('advertising_monthly') or 0))}</b>\n"
            f"🏬 Tashqi saqlash / oy: <b>{_format_money(float(row.get('storage_monthly') or 0))}</b>\n"
            f"🧾 Boshqa tashqi / oy: <b>{_format_money(float(row.get('other_monthly') or 0))}</b>\n\n"
            "Uzum xarajatlari API orqali avtomatik olinadi. Bu yerga faqat Uzumdan tashqaridagi xarajatlarni kiriting."
        )
    return (
        "🧮 <b>Настройки чистой прибыли</b>\n\n"
        f"🏪 Магазин: <code>{shop_id}</code>\n"
        f"🏛 Налог: <b>{escape(tax_name)}</b>\n"
        f"📣 Внешняя реклама в месяц: <b>{_format_money(float(row.get('advertising_monthly') or 0))}</b>\n"
        f"🏬 Внешнее хранение в месяц: <b>{_format_money(float(row.get('storage_monthly') or 0))}</b>\n"
        f"🧾 Другие внешние расходы: <b>{_format_money(float(row.get('other_monthly') or 0))}</b>\n\n"
        "Расходы внутри Uzum загружаются автоматически. Здесь указывайте только расходы вне Uzum."
    )


def finance_settings_markup(telegram_id: int) -> InlineKeyboardMarkup:
    lang = get_user_language(telegram_id)
    uz = lang == "uz"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=("🏛 Soliq turini tanlash" if uz else "🏛 Выбрать налог"), callback_data="taxmenu:open")],
        [InlineKeyboardButton(text=("📣 Tashqi reklama" if uz else "📣 Внешняя реклама"), callback_data="autoedit:advertising_monthly")],
        [InlineKeyboardButton(text=("🏬 Tashqi saqlash" if uz else "🏬 Внешнее хранение"), callback_data="autoedit:storage_monthly")],
        [InlineKeyboardButton(text=("🧾 Boshqa tashqi" if uz else "🧾 Другие внешние"), callback_data="autoedit:other_monthly")],
    ])


def _tax_setting_label(percent: float, *, lang: str) -> str:
    value = round(max(0.0, min(100.0, float(percent or 0))), 4)
    uz = normalize_lang(lang) == "uz"
    if value == 0:
        return "Soliqsiz — 0%" if uz else "Без налога — 0%"
    if value == 1:
        return "Aylanmadan olinadigan soliq — 1%" if uz else "Налог с оборота — 1%"
    if value == 6:
        return "Qo‘shilgan qiymat solig‘i — 6%" if uz else "НДС — 6%"
    if value == 12:
        return "Qo‘shilgan qiymat solig‘i — 12%" if uz else "НДС — 12%"
    return f"Soliq — {value:g}%" if uz else f"Налог — {value:g}%"


def tax_selection_markup(telegram_id: int, current: float) -> InlineKeyboardMarkup:
    uz = get_user_language(telegram_id) == "uz"

    def choice(value: float, ru: str, uz_text: str) -> InlineKeyboardButton:
        marker = "✅" if abs(float(current or 0) - value) < 0.0001 else "▫️"
        return InlineKeyboardButton(
            text=f"{marker} {uz_text if uz else ru}",
            callback_data=f"taxset:{value:g}",
        )

    return InlineKeyboardMarkup(inline_keyboard=[
        [choice(0, "Без налога · 0%", "Soliqsiz · 0%")],
        [choice(1, "С оборота · 1%", "Aylanmadan · 1%")],
        [choice(6, "НДС · 6%", "QQS · 6%"), choice(12, "НДС · 12%", "QQS · 12%")],
        [InlineKeyboardButton(text="✍️ Boshqa foiz" if uz else "✍️ Другой процент", callback_data="taxset:custom")],
        [InlineKeyboardButton(text="⬅️ Ortga" if uz else "⬅️ Назад", callback_data="taxset:back")],
    ])


@dp.callback_query(F.data == "taxmenu:open")
async def tax_selection_open(callback: CallbackQuery) -> None:
    if not callback.from_user:
        return
    telegram_id = int(callback.from_user.id)
    shop_id = db.get_default_shop_id(telegram_id)
    if shop_id is None:
        await callback.answer("Avval do‘konni ulang" if get_user_language(telegram_id) == "uz" else "Сначала подключите магазин", show_alert=True)
        return
    current = float(ensure_finance_settings(telegram_id, int(shop_id)).get("tax_percent") or 0)
    text = (
        "🏛 <b>Soliq turini tanlang</b>\n\nSoliq har bir tovar tushumidan hisoblanadi. Kerak bo‘lsa boshqa foizni qo‘lda kiriting."
        if get_user_language(telegram_id) == "uz"
        else "🏛 <b>Выберите налог</b>\n\nНалог рассчитывается с выручки каждого товара. При необходимости можно указать другой процент."
    )
    if callback.message:
        await callback.message.edit_text(text, reply_markup=tax_selection_markup(telegram_id, current))
    await callback.answer()


@dp.callback_query(F.data.startswith("taxset:"))
async def tax_selection_set(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user:
        return
    telegram_id = int(callback.from_user.id)
    lang = get_user_language(telegram_id)
    action = str(callback.data or "").split(":", 1)[-1]
    shop_id = db.get_default_shop_id(telegram_id)
    if shop_id is None:
        await callback.answer("Avval do‘konni ulang" if lang == "uz" else "Сначала подключите магазин", show_alert=True)
        return
    if action == "back":
        if callback.message:
            await callback.message.edit_text(
                finance_settings_text(telegram_id, int(shop_id)),
                reply_markup=finance_settings_markup(telegram_id),
            )
        await callback.answer()
        return
    if action == "custom":
        await state.set_state(ProductSettingsStates.waiting_for_value)
        await state.update_data(product_setting_field="tax_percent", product_setting_return_section="finance")
        if callback.message:
            await callback.message.answer(
                "Soliq foizini kiriting: 0 dan 100 gacha." if lang == "uz" else "Введите налог в процентах: от 0 до 100.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(text="❌ Bekor qilish" if lang == "uz" else "❌ Отмена", callback_data="settingeditcancel")
                ]]),
            )
        await callback.answer()
        return
    try:
        value = float(action)
    except ValueError:
        await callback.answer("Неизвестный налог", show_alert=True)
        return
    if value not in {0.0, 1.0, 6.0, 12.0}:
        await callback.answer("Неизвестный налог", show_alert=True)
        return
    update_finance_setting(telegram_id, int(shop_id), "tax_percent", value)
    if callback.message:
        await callback.message.edit_text(
            finance_settings_text(telegram_id, int(shop_id)),
            reply_markup=finance_settings_markup(telegram_id),
        )
    await callback.answer("Saqlandi" if lang == "uz" else "Сохранено")


@dp.message(Command("automation", "notification_settings", "auto_reports"))
@dp.message(F.text.in_({"⚙️ Настроить уведомления", "⚙️ Xabarnomalarni sozlash", "📅 Автоотчёты", "📅 Avtohisobotlar"}))
async def automation_settings_screen(message: Message) -> None:
    section = "reports" if str(message.text or "") in {"📅 Автоотчёты", "📅 Avtohisobotlar"} else "main"
    await notification_hub_screen(message, section=section)


async def notification_hub_screen(message: Message, *, section: str = "main") -> None:
    telegram_id = upsert_from_message(message)
    if not await require_active_subscription(message, telegram_id):
        return
    if section in NOTIFICATION_SECTION_FIELDS:
        text = notification_section_text(telegram_id, section)
        markup = notification_section_markup(telegram_id, section)
    else:
        text = notification_hub_text(telegram_id)
        markup = notification_hub_markup(telegram_id)
    await message.answer(text, reply_markup=markup)


@dp.callback_query(F.data.startswith("notifyhub:"))
async def notification_hub_callback(callback: CallbackQuery) -> None:
    if not callback.from_user:
        return
    telegram_id = int(callback.from_user.id)
    section = str(callback.data or "").split(":", 1)[-1]
    if section != "main" and section not in NOTIFICATION_SECTION_FIELDS:
        await callback.answer("Неизвестный раздел", show_alert=True)
        return
    if callback.message:
        if section == "main":
            await callback.message.edit_text(
                notification_hub_text(telegram_id),
                reply_markup=notification_hub_markup(telegram_id),
            )
        else:
            await callback.message.edit_text(
                notification_section_text(telegram_id, section),
                reply_markup=notification_section_markup(telegram_id, section),
            )
    await callback.answer()


@dp.callback_query(F.data.startswith("notifytoggle:"))
async def notification_hub_toggle_callback(callback: CallbackQuery) -> None:
    if not callback.from_user:
        return
    telegram_id = int(callback.from_user.id)
    parts = str(callback.data or "").split(":")
    if len(parts) != 3:
        await callback.answer("Неизвестная настройка", show_alert=True)
        return
    section, field = parts[1], parts[2]
    allowed = field == "notify_reviews" and section == "main"
    if section in NOTIFICATION_SECTION_FIELDS and field in NOTIFICATION_SECTION_FIELDS[section]:
        allowed = True
    if not allowed or field not in AUTOMATION_BOOL_FIELDS:
        await callback.answer("Неизвестная настройка", show_alert=True)
        return
    row = ensure_product_settings(telegram_id)
    update_product_setting(telegram_id, field, 0 if int(row.get(field) or 0) else 1)
    if callback.message:
        if section == "main":
            await callback.message.edit_text(
                notification_hub_text(telegram_id),
                reply_markup=notification_hub_markup(telegram_id),
            )
        else:
            await callback.message.edit_text(
                notification_section_text(telegram_id, section),
                reply_markup=notification_section_markup(telegram_id, section),
            )
    await callback.answer("Saqlandi" if get_user_language(telegram_id) == "uz" else "Сохранено")


@dp.message(Command("sales_notify_mode", "sales_notifications"))
async def sales_mode_screen(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not await require_active_subscription(message, telegram_id):
        return
    back_to = "trial" if subscription_access_level(telegram_id) == "trial" else "automation"
    await message.answer(
        sales_mode_selection_text(telegram_id),
        reply_markup=sales_mode_selection_markup(telegram_id, back_to=back_to),
    )


@dp.callback_query(F.data.startswith("salesmode:"))
async def sales_mode_callback(callback: CallbackQuery) -> None:
    if not callback.from_user:
        return
    telegram_id = int(callback.from_user.id)
    parts = str(callback.data or "").split(":")
    action = parts[1] if len(parts) > 1 else ""
    back_to = parts[2] if len(parts) > 2 else "automation"
    lang = get_user_language(telegram_id)
    if subscription_access_level(telegram_id) == "trial":
        back_to = "trial"

    if action == "menu":
        if callback.message:
            await callback.message.edit_text(
                sales_mode_selection_text(telegram_id),
                reply_markup=sales_mode_selection_markup(telegram_id, back_to=back_to),
            )
        await callback.answer()
        return

    if action == "back":
        if callback.message:
            if back_to == "trial":
                await callback.message.edit_text(
                    "💸 Savdo xabarlari sozlamalari saqlandi."
                    if normalize_lang(lang) == "uz"
                    else "💸 Настройки уведомлений о продажах сохранены."
                )
                await callback.message.answer(
                    "Kerakli sinov funksiyasini tanlang 👇"
                    if normalize_lang(lang) == "uz"
                    else "Выберите доступную функцию пробного периода 👇",
                    reply_markup=sales_menu_for_user(telegram_id),
                )
            elif back_to == "sales":
                await callback.message.edit_text(
                    notification_section_text(telegram_id, "sales"),
                    reply_markup=notification_section_markup(telegram_id, "sales"),
                )
            else:
                await callback.message.edit_text(
                    automation_settings_text(telegram_id),
                    reply_markup=automation_settings_markup(telegram_id),
                )
        await callback.answer()
        return

    if action not in SALES_NOTIFICATION_MODES:
        await callback.answer(
            "Noma’lum rejim" if normalize_lang(lang) == "uz" else "Неизвестный режим",
            show_alert=True,
        )
        return

    set_sales_notification_mode(telegram_id, action)
    reset_user_sales_digest_schedule(
        telegram_id,
        now=_utc_now(),
        clear_queue=True,
    )
    if callback.message:
        if back_to == "trial":
            await callback.message.edit_text(
                sales_mode_selection_text(telegram_id),
                reply_markup=sales_mode_selection_markup(telegram_id, back_to="trial"),
            )
        elif back_to == "sales":
            await callback.message.edit_text(
                notification_section_text(telegram_id, "sales"),
                reply_markup=notification_section_markup(telegram_id, "sales"),
            )
        else:
            await callback.message.edit_text(
                automation_settings_text(telegram_id),
                reply_markup=automation_settings_markup(telegram_id),
            )
    labels = {
        "instant": ("Режим: продажи сразу", "Rejim: savdolar darhol"),
        "hourly": ("Режим: сводка раз в час", "Rejim: har soatda hisobot"),
        "off": ("Уведомления о продажах выключены", "Savdo xabarlari o‘chirildi"),
    }
    ru_text, uz_text = labels[action]
    await callback.answer(uz_text if normalize_lang(lang) == "uz" else ru_text)


@dp.message(Command("loss_defect_notify_status", "warehouse_loss_notifications"))
@dp.message(F.text == "🧭 Yo‘qotish va yaroqsiz")
async def loss_defect_notification_status(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, _, shop_id = req
    settings = ensure_product_settings(telegram_id)
    lang = get_user_language(telegram_id)
    initialized = operational_watcher_initialized(telegram_id, shop_id, "loss_defect")
    if lang == "uz":
        text = (
            "🧭 <b>Yo‘qotish va yaroqsiz tovar xabarnomalari</b>\n\n"
            f"Do‘kon: <code>{shop_id}</code>\n"
            f"Yo‘qotish: {_toggle_icon(settings.get('notify_losses'))}\n"
            f"Yaroqsiz: {_toggle_icon(settings.get('notify_defects'))}\n"
            f"Tekshiruv: har <b>{max(300, LOSS_DEFECT_CHECK_INTERVAL_SECONDS)}</b> soniyada\n"
            f"Holat: {'boshlang‘ich ma’lumot saqlangan' if initialized else 'keyingi tekshiruvda boshlanadi'}\n\n"
            "Bot quantityMissing va quantityDefected oshganda xabar beradi."
        )
    else:
        text = (
            "🧭 <b>Уведомления о потерях и браке</b>\n\n"
            f"Магазин: <code>{shop_id}</code>\n"
            f"Потери: {_toggle_icon(settings.get('notify_losses'))}\n"
            f"Брак: {_toggle_icon(settings.get('notify_defects'))}\n"
            f"Проверка каждые: <b>{max(300, LOSS_DEFECT_CHECK_INTERVAL_SECONDS)}</b> сек.\n"
            f"Состояние: {'исходные значения сохранены' if initialized else 'инициализация при следующей проверке'}\n\n"
            "Бот уведомит при увеличении quantityMissing или quantityDefected."
        )
    await message.answer(text, reply_markup=notify_menu_for_message(message))


@dp.message(Command("fbo_acceptance_notify_status", "fbo_acceptance_notifications"))
@dp.message(F.text.in_({"🚚 Приёмка FBO", "🚚 FBO qabuli"}))
async def fbo_acceptance_notification_status(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, _, shop_id = req
    settings = ensure_product_settings(telegram_id)
    lang = get_user_language(telegram_id)
    initialized = operational_watcher_initialized(telegram_id, shop_id, "fbo_acceptance")
    if lang == "uz":
        text = (
            "🚚 <b>FBO qabul xabarnomalari</b>\n\n"
            f"Do‘kon: <code>{shop_id}</code>\n"
            f"Holat: {_toggle_icon(settings.get('notify_fbo_acceptance'))}\n"
            f"Tekshiruv: har <b>{max(300, FBO_ACCEPTANCE_CHECK_INTERVAL_SECONDS)}</b> soniyada\n"
            f"Nazorat: {'faol' if initialized else 'keyingi tekshiruvda boshlanadi'}\n\n"
            "To‘liq qabul qilinsa qisqa xabar, farq bo‘lsa muammoli tovarlar PDF fayli keladi."
        )
    else:
        text = (
            "🚚 <b>Уведомления о приёмке FBO</b>\n\n"
            f"Магазин: <code>{shop_id}</code>\n"
            f"Статус: {_toggle_icon(settings.get('notify_fbo_acceptance'))}\n"
            f"Проверка каждые: <b>{max(300, FBO_ACCEPTANCE_CHECK_INTERVAL_SECONDS)}</b> сек.\n"
            f"Контроль: {'активен' if initialized else 'инициализация при следующей проверке'}\n\n"
            "При полной приёмке придёт краткое уведомление, при расхождении — PDF со списком товаров."
        )
    await message.answer(text, reply_markup=notify_menu_for_message(message))


@dp.message(Command("supply_settings", "replenishment_settings"))
@dp.message(F.text.in_({
    "⚙️ Срок поставки",
    "⚙️ Yetkazish muddati",
    "⚙️ Параметры поставки",
    "⚙️ Yetkazish sozlamasi",
    "🚚 Настройки поставки",
    "🚚 Yetkazish sozlamalari",
}))
async def supply_settings_screen(message: Message) -> None:
    telegram_id = upsert_from_message(message)
    if not await require_active_subscription(message, telegram_id):
        return
    await message.answer(supply_settings_text(telegram_id), reply_markup=supply_settings_markup(telegram_id))


@dp.message(Command("finance_settings", "expenses_settings"))
@dp.message(F.text.in_({"🧮 Расходы и налоги", "🧮 Xarajat va soliq"}))
async def finance_settings_screen(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, _, shop_id = req
    await message.answer(
        finance_settings_text(telegram_id, shop_id),
        reply_markup=finance_settings_markup(telegram_id),
    )


@dp.message(Command("uzum_expenses", "expenses_uzum"))
@dp.message(F.text.in_({"🧾 Расходы Uzum", "🧾 Uzum xarajatlari"}))
async def uzum_expenses_report(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    if not await require_premium_subscription(message, telegram_id):
        return
    lang = get_user_language(telegram_id)
    await message.answer(
        "⌛ Uzum xarajatlarini 30 kun uchun yuklayapman..."
        if lang == "uz"
        else "⌛ Загружаю расходы Uzum за 30 дней...",
        reply_markup=finance_menu_for_message(message),
    )
    date_from, date_to = _days_range_ms(30)
    summary = await load_uzum_expense_summary(
        client,
        shop_id,
        date_from,
        date_to,
        force=True,
    )
    if not summary.get("available"):
        await message.answer(
            (
                "⚠️ <b>Uzum xarajatlari vaqtincha olinmadi</b>\n\n"
                "Foyda hisobida bu ma’lumot yetishmayotgani ko‘rsatiladi. Keyinroq qayta urinib ko‘ring."
                if lang == "uz"
                else "⚠️ <b>Расходы Uzum временно недоступны</b>\n\n"
                "В расчёте прибыли это будет отмечено как неполные данные. Попробуйте обновить позже."
            ),
            reply_markup=finance_menu_for_message(message),
        )
        return

    total = float(summary.get("total") or 0)
    title = "🧾 <b>Uzum xarajatlari — 30 kun</b>" if lang == "uz" else "🧾 <b>Расходы Uzum за 30 дней</b>"
    header = [
        f"🏪 Do‘kon: <code>{shop_id}</code>" if lang == "uz" else f"🏪 Магазин: <code>{shop_id}</code>",
        f"💳 Foydadan ayriladigan qo‘shimcha xarajatlar: <b>{_format_money(total)}</b>" if lang == "uz" else f"💳 Дополнительные расходы для расчёта прибыли: <b>{_format_money(total)}</b>",
        (
            f"🏬 Saqlash: <b>{_format_money(float(summary.get('storage') or 0))}</b> | "
            f"📣 Reklama: <b>{_format_money(float(summary.get('advertising') or 0))}</b>"
            if lang == "uz"
            else f"🏬 Хранение: <b>{_format_money(float(summary.get('storage') or 0))}</b> | "
            f"📣 Реклама: <b>{_format_money(float(summary.get('advertising') or 0))}</b>"
        ),
        (
            f"⚠️ Jarima: <b>{_format_money(float(summary.get('penalty') or 0))}</b> | "
            f"🧾 Boshqa: <b>{_format_money(float(summary.get('other') or 0))}</b>"
            if lang == "uz"
            else f"⚠️ Штрафы: <b>{_format_money(float(summary.get('penalty') or 0))}</b> | "
            f"🧾 Прочие: <b>{_format_money(float(summary.get('other') or 0))}</b>"
        ),
    ]
    if int(summary.get("pending_count") or 0):
        header.append(
            f"⏳ Hali tasdiqlanmagan: <b>{int(summary['pending_count'])}</b>"
            if lang == "uz"
            else f"⏳ Ещё не подтверждено: <b>{int(summary['pending_count'])}</b>"
        )
    if abs(float(summary.get("order_charge") or 0)) > 0.001:
        header.append(
            f"ℹ️ Komissiya/logistika: <b>{_format_money(float(summary['order_charge']))}</b> — to‘lovda hisobga olingan"
            if lang == "uz"
            else f"ℹ️ Комиссия/логистика: <b>{_format_money(float(summary['order_charge']))}</b> — уже учтены в выплате"
        )
    items: list[str] = []
    for index, row in enumerate(list(summary.get("rows") or []), start=1):
        date_text = _fmt_dt(row.get("date"))
        name = escape(_short_text(str(row.get("name") or "Uzum"), 85))
        source = escape(_short_text(str(row.get("source") or "—"), 45))
        amount = float(row.get("signed_amount") or 0)
        included = bool(row.get("included_in_profit", True))
        accounting_note = (
            "\nℹ️ To‘lovda hisobga olingan, qayta ayrilmaydi."
            if lang == "uz" and not included
            else "\nℹ️ Уже учтено в выплате, повторно не вычитается."
            if not included
            else ""
        )
        items.append(
            (
                f"{index}. <b>{name}</b>\n🕒 {date_text} | Manba: <code>{source}</code>\n"
                f"💳 <b>{_format_money(amount)}</b>{accounting_note}"
                if lang == "uz"
                else f"{index}. <b>{name}</b>\n🕒 {date_text} | Источник: <code>{source}</code>\n"
                f"💳 <b>{_format_money(amount)}</b>{accounting_note}"
            )
        )
    if not items:
        items = [
            "Tanlangan davrda tasdiqlangan Uzum xarajatlari topilmadi."
            if lang == "uz"
            else "За выбранный период подтверждённых расходов Uzum не найдено."
        ]
    await send_paginated_list(
        message,
        kind="uzum_expenses",
        title=title,
        summary=header,
        items=items,
        section="sales",
        reply_markup=finance_menu_for_message(message),
    )


@dp.callback_query(F.data.startswith("autotoggle:"))
async def automation_toggle_callback(callback: CallbackQuery) -> None:
    if not callback.from_user:
        return
    telegram_id = int(callback.from_user.id)
    field = str(callback.data or "").split(":", 1)[-1]
    if field not in AUTOMATION_BOOL_FIELDS:
        await callback.answer("Неизвестная настройка", show_alert=True)
        return
    row = ensure_product_settings(telegram_id)
    update_product_setting(telegram_id, field, 0 if int(row.get(field) or 0) else 1)
    if callback.message:
        await callback.message.edit_text(
            automation_settings_text(telegram_id),
            reply_markup=automation_settings_markup(telegram_id),
        )
    await callback.answer("Сохранено")


@dp.callback_query(F.data.startswith("autoedit:"))
@dp.callback_query(F.data.startswith("notifyedit:"))
async def automation_edit_callback(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user:
        return
    telegram_id = int(callback.from_user.id)
    parts = str(callback.data or "").split(":")
    if parts and parts[0] == "notifyedit" and len(parts) == 3:
        return_section = parts[1]
        field = parts[2]
    else:
        field = parts[1] if len(parts) > 1 else ""
        if field in {"lead_time_days", "safety_days", "target_cover_days"}:
            return_section = "supply"
        elif field in FINANCE_NUMERIC_LIMITS:
            return_section = "finance"
        else:
            return_section = "automation"
    if field not in PRODUCT_NUMERIC_LIMITS and field not in FINANCE_NUMERIC_LIMITS:
        await callback.answer("Неизвестная настройка", show_alert=True)
        return
    lang = get_user_language(telegram_id)
    prompts_ru = {
        "daily_hour": "Введите час ежедневного отчёта: от 0 до 23.",
        "weekly_weekday": "Введите день недели: 1 — понедельник, 7 — воскресенье.",
        "weekly_hour": "Введите час еженедельного отчёта: от 0 до 23.",
        "monthly_day": "Введите день месяца: от 1 до 28.",
        "monthly_hour": "Введите час ежемесячного отчёта: от 0 до 23.",
        "low_stock_threshold": "Введите порог низкого остатка в штуках.",
        "lead_time_days": "Через сколько дней товар обычно приезжает на склад?",
        "safety_days": "На сколько дней нужен страховой запас?",
        "target_cover_days": "На сколько дней продаж формировать рекомендуемую поставку?",
        "tax_percent": "Введите налог в процентах, например 4.",
        "advertising_monthly": "Введите средние расходы на рекламу за месяц в сумах.",
        "storage_monthly": "Введите расходы на хранение за месяц в сумах.",
        "other_monthly": "Введите другие ежемесячные расходы в сумах.",
    }
    prompts_uz = {
        "daily_hour": "Kunlik hisobot soatini kiriting: 0 dan 23 gacha.",
        "weekly_weekday": "Hafta kunini kiriting: 1 — dushanba, 7 — yakshanba.",
        "weekly_hour": "Haftalik hisobot soatini kiriting: 0 dan 23 gacha.",
        "monthly_day": "Oy kunini kiriting: 1 dan 28 gacha.",
        "monthly_hour": "Oylik hisobot soatini kiriting: 0 dan 23 gacha.",
        "low_stock_threshold": "Kam qoldiq chegarasini dona bilan kiriting.",
        "lead_time_days": "Tovar omborga odatda necha kunda yetib keladi?",
        "safety_days": "Necha kunlik xavfsizlik zaxirasi kerak?",
        "target_cover_days": "Tavsiya etilgan yetkazish necha kunlik savdoga yetsin?",
        "tax_percent": "Soliq foizini kiriting, masalan 4.",
        "advertising_monthly": "Bir oylik reklama xarajatini so‘mda kiriting.",
        "storage_monthly": "Bir oylik saqlash xarajatini so‘mda kiriting.",
        "other_monthly": "Boshqa oylik xarajatlarni so‘mda kiriting.",
    }
    await state.set_state(ProductSettingsStates.waiting_for_value)
    await state.update_data(product_setting_field=field, product_setting_return_section=return_section)
    if callback.message:
        await callback.message.answer(
            (prompts_uz if lang == "uz" else prompts_ru)[field],
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(
                    text="❌ Bekor qilish" if lang == "uz" else "❌ Отмена",
                    callback_data="settingeditcancel",
                )
            ]]),
        )
    await callback.answer()


@dp.callback_query(F.data == "settingeditcancel")
async def automation_edit_cancel_callback(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user:
        return
    telegram_id = int(callback.from_user.id)
    await state.clear()
    if callback.message:
        await callback.message.edit_text(
            "Bekor qilindi." if get_user_language(telegram_id) == "uz" else "Изменение отменено."
        )
    await callback.answer()


@dp.message(ProductSettingsStates.waiting_for_value)
async def product_setting_value_received(message: Message, state: FSMContext) -> None:
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    data = await state.get_data()
    field = str(data.get("product_setting_field") or "")
    return_section = str(data.get("product_setting_return_section") or "automation")
    raw = str(message.text or "").strip().replace(" ", "").replace(",", ".")
    try:
        value = float(raw)
    except ValueError:
        await message.answer("Son kiriting." if lang == "uz" else "Введите число.")
        return
    limits = PRODUCT_NUMERIC_LIMITS.get(field) or FINANCE_NUMERIC_LIMITS.get(field)
    if not limits or not (limits[0] <= value <= limits[1]):
        await message.answer(
            f"Qiymat {limits[0]:g} dan {limits[1]:g} gacha bo‘lishi kerak."
            if lang == "uz" and limits
            else f"Допустимое значение: от {limits[0]:g} до {limits[1]:g}." if limits
            else "Неизвестная настройка.",
        )
        return
    if field == "weekly_weekday":
        value -= 1
    if field in PRODUCT_NUMERIC_LIMITS:
        update_product_setting(telegram_id, field, int(value))
        await state.clear()
        if field in {"lead_time_days", "safety_days", "target_cover_days"}:
            await message.answer(supply_settings_text(telegram_id), reply_markup=supply_settings_markup(telegram_id))
        elif return_section in NOTIFICATION_SECTION_FIELDS:
            await message.answer(
                notification_section_text(telegram_id, return_section),
                reply_markup=notification_section_markup(telegram_id, return_section),
            )
        else:
            await message.answer(automation_settings_text(telegram_id), reply_markup=automation_settings_markup(telegram_id))
        return
    shop_id = db.get_default_shop_id(telegram_id)
    if shop_id is None:
        await state.clear()
        await message.answer("Avval do‘konni ulang." if lang == "uz" else "Сначала подключите магазин.", reply_markup=menu_for_message(message))
        return
    update_finance_setting(telegram_id, int(shop_id), field, value)
    await state.clear()
    await message.answer(
        finance_settings_text(telegram_id, int(shop_id)),
        reply_markup=finance_settings_markup(telegram_id),
    )


# --- Денежный центр действий и прозрачная оценка пользы ---
def _finance_rows_for_last_days(
    rows: list[dict[str, Any]],
    days: int,
    *,
    now: datetime | None = None,
) -> list[dict[str, Any]]:
    now_local = now or datetime.now(UZT)
    if now_local.tzinfo is not None:
        now_naive = now_local.astimezone(UZT).replace(tzinfo=None)
    else:
        now_naive = now_local
    cutoff = now_naive - timedelta(days=max(1, int(days)))
    dated: list[dict[str, Any]] = []
    unknown: list[dict[str, Any]] = []
    for item in rows:
        dt = _finance_datetime_for_report(item)
        if dt is None:
            unknown.append(item)
        elif dt >= cutoff:
            dated.append(item)
    return dated if dated else unknown


async def _collect_business_control_data(
    telegram_id: int,
    client: UzumClient,
    shop_id: int,
) -> dict[str, Any]:
    date_from, date_to = _days_range_ms(30)
    finance_rows, _, source_info = await _load_finance_range_flexible(
        client,
        shop_id,
        date_from,
        date_to,
    )
    stats = _build_noorza_today_stats(finance_rows)
    stats["cancelled_value"] = sum(
        _finance_gross_revenue(item)
        for item in finance_rows
        if _is_cancelled_status(_finance_status(item))
    )
    stock_raw = await load_sku_rows(client, shop_id, max_pages=50)
    await sync_uzum_sku_financials(
        client,
        telegram_id,
        shop_id,
        stock_rows=stock_raw,
    )
    costs = get_unit_cost_map(telegram_id, shop_id)
    product_rows = _build_unit_rows_from_finance(finance_rows, costs)
    sales_7 = _finance_rows_for_last_days(finance_rows, 7)
    stock_rows, _ = _build_stock_report_rows(stock_raw, sales_7)
    settings = ensure_product_settings(telegram_id)
    replenishment = build_replenishment_plan(stock_raw, finance_rows, settings)
    plan_by_key: dict[str, dict[str, Any]] = {}
    for plan_item in replenishment:
        for key in _stock_match_keys(plan_item.get("row") or {}):
            plan_by_key[key] = plan_item
    for stock_item in stock_rows:
        stock_item["low_stock_threshold"] = int(settings.get("low_stock_threshold") or 0)
        plan_item = next(
            (plan_by_key[key] for key in _stock_match_keys(stock_item) if key in plan_by_key),
            None,
        )
        if plan_item:
            stock_item.update({
                "risk_value": float(plan_item.get("risk_value") or 0),
                "recommended_qty": int(plan_item.get("recommended_qty") or 0),
                "reorder_date": plan_item.get("reorder_date"),
                "avg_daily": float(plan_item.get("avg_daily") or 0),
                "days_left": plan_item.get("days_left"),
            })

    loss_filters: list[str] = []
    try:
        loss_raw, loss_filters = await _load_all_time_loss_rows(client, shop_id)
        _, lost_rows = _build_stock_report_rows(loss_raw, sales_7)
    except Exception:
        logging.exception("Control center: all-time losses failed user=%s shop=%s", telegram_id, shop_id)
        lost_rows = []

    stock_for_actions = [{**row, "missing": 0, "defected": 0} for row in stock_rows]
    actions = _build_premium_actions(stock_for_actions + lost_rows, product_rows, stats)
    visible_actions = sync_business_actions(telegram_id, shop_id, actions)
    return {
        "actions": visible_actions,
        "all_actions": actions,
        "replenishment": replenishment,
        "stats": stats,
        "products": product_rows,
        "stock": stock_rows,
        "lost": lost_rows,
        "source_info": source_info,
        "loss_filters": loss_filters,
        "generated_at": datetime.now(UZT),
    }


def _business_center_text(
    telegram_id: int,
    shop_id: int,
    data: dict[str, Any],
) -> str:
    lang = get_user_language(telegram_id)
    actions = list(data.get("actions") or [])
    total_effect = sum(max(0.0, float(item.get("amount") or 0)) for item in actions)
    critical = sum(1 for item in actions if item.get("priority") == "critical")
    warnings = sum(1 for item in actions if item.get("priority") == "warning")
    generated = data.get("generated_at")
    generated_text = generated.strftime("%d.%m.%Y %H:%M") if isinstance(generated, datetime) else "—"
    if lang == "uz":
        lines = [
            "🚨 <b>Hozir muhim</b>",
            f"🏪 Do‘kon: <code>{shop_id}</code>",
            f"🕒 Yangilandi: {generated_text}",
            "",
            f"💰 Potensial xavf summasi: <b>{_format_money(total_effect)}</b>",
            f"🔴 Jiddiy: <b>{critical}</b> | 🟡 Diqqat: <b>{warnings}</b>",
        ]
        if not actions:
            lines.extend(["", "✅ Hozir faol muammo topilmadi yoki ular vaqtincha yashirilgan."])
        else:
            lines.extend(["", "<b>Avval nima qilish kerak:</b>"])
            for index, action in enumerate(actions[:8], start=1):
                amount = float(action.get("amount") or 0)
                amount_text = _format_money(amount) if amount > 0 else "baholanmagan"
                lines.extend([
                    "",
                    f"{index}. <b>{escape(str(action.get('category_uz') or 'Vazifa'))}</b> — <b>{amount_text}</b>",
                    f"📦 {escape(_short_text(str(action.get('title') or '—'), 75))}",
                    f"⚠️ {escape(str(action.get('problem_uz') or ''))}",
                    f"➡️ {escape(str(action.get('recommendation_uz') or ''))}",
                ])
        lines.extend(["", "ℹ️ Summa — kafolatlangan tejash emas, mavjud ma’lumotlar bo‘yicha potensial ta’sir."])
        return "\n".join(lines)

    lines = [
        "🚨 <b>Важно сейчас</b>",
        f"🏪 Магазин: <code>{shop_id}</code>",
        f"🕒 Обновлено: {generated_text}",
        "",
        f"💰 Потенциально под риском: <b>{_format_money(total_effect)}</b>",
        f"🔴 Критично: <b>{critical}</b> | 🟡 Внимание: <b>{warnings}</b>",
    ]
    if not actions:
        lines.extend(["", "✅ Активных проблем сейчас не найдено либо они временно отложены."])
    else:
        lines.extend(["", "<b>Что сделать в первую очередь:</b>"])
        for index, action in enumerate(actions[:8], start=1):
            amount = float(action.get("amount") or 0)
            amount_text = _format_money(amount) if amount > 0 else "эффект не оценён"
            lines.extend([
                "",
                f"{index}. <b>{escape(str(action.get('category_ru') or 'Задача'))}</b> — <b>{amount_text}</b>",
                f"📦 {escape(_short_text(str(action.get('title') or '—'), 75))}",
                f"⚠️ {escape(str(action.get('problem_ru') or ''))}",
                f"➡️ {escape(str(action.get('recommendation_ru') or ''))}",
            ])
    lines.extend(["", "ℹ️ Сумма — потенциальный эффект по доступным данным, а не гарантированная экономия."])
    return "\n".join(lines)


def _business_center_markup(
    telegram_id: int,
    shop_id: int,
    actions: list[dict[str, Any]],
) -> InlineKeyboardMarkup:
    lang = get_user_language(telegram_id)
    uz = lang == "uz"
    rows: list[list[InlineKeyboardButton]] = []
    for index, action in enumerate(actions[:3], start=1):
        key = str(action.get("action_key") or "")
        rows.append([
            InlineKeyboardButton(
                text=f"✅ {index} {'bajarildi' if uz else 'решено'}",
                callback_data=f"biz:resolved:{shop_id}:{key}",
            ),
            InlineKeyboardButton(
                text=f"⏰ {index} {'3 kun' if uz else 'на 3 дня'}",
                callback_data=f"biz:snoozed:{shop_id}:{key}",
            ),
        ])
    rows.append([
        InlineKeyboardButton(text=("🔄 Yangilash" if uz else "🔄 Обновить"), callback_data=f"biz:refresh:{shop_id}"),
        InlineKeyboardButton(text=("📈 Bot foydasi" if uz else "📈 Польза бота"), callback_data=f"biz:value:{shop_id}"),
    ])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def _render_business_center(
    target: Message,
    telegram_id: int,
    client: UzumClient,
    shop_id: int,
    *,
    edit: bool = False,
) -> dict[str, Any]:
    data = await _collect_business_control_data(telegram_id, client, shop_id)
    text = _business_center_text(telegram_id, shop_id, data)
    markup = _business_center_markup(telegram_id, shop_id, list(data.get("actions") or []))
    if edit:
        try:
            await target.edit_text(text, reply_markup=markup)
        except Exception:
            await target.answer(text, reply_markup=markup)
    else:
        await target.answer(text, reply_markup=markup)
    return data


@dp.message(Command("control_center", "business_actions", "money_actions"))
@dp.message(F.text.in_({"💼 Центр действий", "💼 Amallar markazi"}))
async def business_control_center(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    await message.answer(
        "⌛ Do‘kon tekshirilmoqda..." if lang == "uz" else "⌛ Проверяю магазин и расставляю приоритеты...",
        reply_markup=attention_menu_for_message(message),
    )
    try:
        await _render_business_center(message, telegram_id, client, shop_id)
    except Exception as exc:
        await send_api_error(message, exc)


def _business_value_text(
    telegram_id: int,
    shop_id: int,
    data: dict[str, Any],
) -> str:
    lang = get_user_language(telegram_id)
    actions = list(data.get("actions") or [])
    active_total = sum(max(0.0, float(item.get("amount") or 0)) for item in actions)
    resolved_total = resolved_business_value(telegram_id, shop_id)
    categories: dict[str, float] = {}
    for item in actions:
        name = str(
            (item.get("category_uz") if lang == "uz" else item.get("category_ru"))
            or "—"
        )
        categories[name] = categories.get(name, 0.0) + max(0.0, float(item.get("amount") or 0))
    top_categories = sorted(categories.items(), key=lambda pair: pair[1], reverse=True)[:5]
    if lang == "uz":
        lines = [
            "📈 <b>Botning biznes uchun foydasi</b>",
            f"🏪 Do‘kon: <code>{shop_id}</code>",
            "",
            f"🎯 Hozir nazoratdagi potensial summa: <b>{_format_money(active_total)}</b>",
            f"✅ Shu oy bajarilgan vazifalar summasi: <b>{_format_money(resolved_total)}</b>",
            f"📋 Faol vazifalar: <b>{len(actions)}</b>",
        ]
        if top_categories:
            lines.extend(["", "<b>Asosiy manbalar:</b>"])
            lines.extend(f"• {escape(name)}: <b>{_format_money(amount)}</b>" for name, amount in top_categories)
        lines.extend(["", "ℹ️ Bu ko‘rsatkichlar potensial ta’sir va bajarilgan vazifalarni ko‘rsatadi; bankdagi haqiqiy tejash emas."])
        return "\n".join(lines)
    lines = [
        "📈 <b>Польза бота для бизнеса</b>",
        f"🏪 Магазин: <code>{shop_id}</code>",
        "",
        f"🎯 Сейчас под контролем: <b>{_format_money(active_total)}</b>",
        f"✅ Сумма отмеченных решёнными задач за месяц: <b>{_format_money(resolved_total)}</b>",
        f"📋 Активных задач: <b>{len(actions)}</b>",
    ]
    if top_categories:
        lines.extend(["", "<b>Основные источники потенциального эффекта:</b>"])
        lines.extend(f"• {escape(name)}: <b>{_format_money(amount)}</b>" for name, amount in top_categories)
    lines.extend(["", "ℹ️ Это оценка потенциального влияния и выполненных задач, а не подтверждённая банковская экономия."])
    return "\n".join(lines)


@dp.message(Command("bot_value", "value_report", "roi"))
@dp.message(F.text.in_({"📈 Польза бота", "📈 Bot foydasi", "📈 Эффект рекомендаций", "📈 Tavsiyalar ta’siri"}))
async def business_value_report(message: Message) -> None:
    req = await require_connection(message)
    if req is None:
        return
    telegram_id, client, shop_id = req
    lang = get_user_language(telegram_id)
    await message.answer("⌛ Foyda hisoblanmoqda..." if lang == "uz" else "⌛ Считаю потенциальную пользу...", reply_markup=attention_menu_for_message(message))
    try:
        data = await _collect_business_control_data(telegram_id, client, shop_id)
        await message.answer(
            _business_value_text(telegram_id, shop_id, data),
            reply_markup=_business_center_markup(telegram_id, shop_id, list(data.get("actions") or [])),
        )
    except Exception as exc:
        await send_api_error(message, exc)


def _user_has_shop(telegram_id: int, shop_id: int) -> bool:
    return any(int(row["shop_id"]) == int(shop_id) for row in db.list_shops(telegram_id))


@dp.callback_query(F.data.startswith("biz:"))
async def business_action_callback(callback: CallbackQuery) -> None:
    if not callback.from_user or not callback.message:
        return
    telegram_id = int(callback.from_user.id)
    parts = str(callback.data or "").split(":")
    if len(parts) < 3 or not parts[2].isdigit():
        await callback.answer("Некорректная команда", show_alert=True)
        return
    operation = parts[1]
    shop_id = int(parts[2])
    if not _user_has_shop(telegram_id, shop_id):
        await callback.answer("Магазин недоступен", show_alert=True)
        return
    lang = get_user_language(telegram_id)
    if operation in {"resolved", "snoozed"}:
        if len(parts) < 4:
            await callback.answer("Задача не найдена", show_alert=True)
            return
        ok = update_business_action_state(telegram_id, shop_id, parts[3], operation)
        if not ok:
            await callback.answer(
                "Vazifa topilmadi yoki allaqachon yangilangan"
                if lang == "uz"
                else "Задача не найдена или уже обновлена",
                show_alert=True,
            )
            return
        await callback.answer(
            ("Bajarildi" if operation == "resolved" else "3 kunga qoldirildi")
            if lang == "uz"
            else ("Отмечено решённым" if operation == "resolved" else "Отложено на 3 дня"),
        )
    else:
        await callback.answer("Yangilanmoqda..." if lang == "uz" else "Обновляю...")
    client = get_uzum_for_user(telegram_id)
    if client is None:
        return
    try:
        data = await _collect_business_control_data(telegram_id, client, shop_id)
        if operation == "value":
            text = _business_value_text(telegram_id, shop_id, data)
        else:
            text = _business_center_text(telegram_id, shop_id, data)
        await callback.message.edit_text(
            text,
            reply_markup=_business_center_markup(telegram_id, shop_id, list(data.get("actions") or [])),
        )
    except Exception:
        logging.exception("Business action callback failed user=%s shop=%s", telegram_id, shop_id)


@dp.message(F.text.in_({
    "🚚 Что заказать",
    "🚚 Nima buyurtma qilish",
    "🚚 План поставки",
    "🚚 Yetkazib berish rejasi",
}))
async def supply_plan_button(message: Message) -> None:
    await smart_lowstock(message)


@dp.message(F.text)
async def friendly_auto_start(message: Message, state: FSMContext) -> None:
    """Показывает понятное стартовое меню, если новый клиент написал любое сообщение вместо /start."""
    telegram_id = upsert_from_message(message)
    lang = get_user_language(telegram_id)
    current_state = await state.get_state()

    # Если пользователь был в процессе подключения/импорта, не сбрасываем состояние без причины.
    if current_state:
        if lang == "uz":
            await message.answer(
                "Hozir bot avvalgi amal uchun ma’lumot kutmoqda. Oldingi ko‘rsatmaga muvofiq javob bering "
                "yoki ekrandagi <b>Bekor qilish</b> tugmasini bosing.",
                reply_markup=menu_for_message(message),
            )
        else:
            await message.answer(
                "Сейчас бот ждёт данные для предыдущего действия. Ответьте по подсказке выше "
                "или нажмите кнопку <b>Отмена</b>, если она показана.",
                reply_markup=menu_for_message(message),
            )
        return

    ensure_subscription(telegram_id)
    connected = db.has_uzum_connection(telegram_id)
    if lang == "uz":
        text = (
            "Bu xabarni amal sifatida tanimadim. Kerakli bo‘limni pastdagi tugmalar orqali tanlang 👇"
            if connected
            else "Avval <b>🎥 Qanday ulash kerak</b>, so‘ng <b>🔌 Do‘konni ulash</b> tugmasini bosing 👇"
        )
    else:
        text = (
            "Не распознал это как действие. Выберите нужный раздел кнопкой внизу 👇"
            if connected
            else "Сначала нажмите <b>🎥 Как подключить</b>, затем <b>🔌 Подключить магазин</b> 👇"
        )
    await message.answer(text, reply_markup=menu_for_message(message))


@dp.message(F.photo)
@dp.message(F.document)
async def recover_payment_receipt_after_restart(message: Message, state: FSMContext) -> None:
    """Resume a persisted payment request if the in-memory FSM was lost on restart."""
    telegram_id = upsert_from_message(message)
    pending = latest_awaiting_payment_request(telegram_id)
    if not pending:
        return
    await state.set_state(PaymentStates.waiting_for_receipt)
    await state.update_data(payment_request_id=int(pending["id"]))
    await payment_receipt_received(message, state)


# --- Финальная чистка узбекских сообщений для клиентов ---
# Убирает оставшиеся русские фразы в разделах уведомлений, подписки и помощи.
_FINAL_UZ_TRANSLATE_RUNTIME_TEXT_TO_UZ = translate_runtime_text_to_uz


def translate_runtime_text_to_uz(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text
    text = _FINAL_UZ_TRANSLATE_RUNTIME_TEXT_TO_UZ(text)

    fixes = [
        # Уведомления о низких остатках
        ("📉 <b>Уведомления о низких остатках</b>", "📉 <b>Kam qoldiq xabarnomalari</b>"),
        ("Уведомления о низких остатках", "Kam qoldiq xabarnomalari"),
        ("📉 <b>Низкие остатки</b>", "📉 <b>Kam qoldiq</b>"),
        ("Низкие остатки", "Kam qoldiq"),
        ("Порог:", "Chegara:"),
        ("Chegara: ≤", "Chegara: ≤"),
        ("шт.", "dona"),
        ("Проверка каждые:", "Tekshiruv har"),
        ("Tekshiruv har: <b>", "Tekshiruv har <b>"),
        ("Tekshiruv har <b>1800</b> soniya", "Tekshiruv har <b>1800</b> soniyada"),
        ("Tekshiruv har <b>300</b> soniya", "Tekshiruv har <b>300</b> soniyada"),
        ("Tekshiruv har <b>60</b> soniya", "Tekshiruv har <b>60</b> soniyada"),
        ("Состояние: остатки уже запомнены", "Holat: qoldiqlar allaqachon eslab qolingan"),
        ("Состояние: нулевые остатки уже запомнены", "Holat: nol qoldiqlar allaqachon eslab qolingan"),
        ("Состояние: инициализация при следующей проверке", "Holat: keyingi tekshiruvda ishga tushadi"),
        ("Holat: остатки уже запомнены", "Holat: qoldiqlar allaqachon eslab qolingan"),
        ("Holat: нулевые остатки уже запомнены", "Holat: nol qoldiqlar allaqachon eslab qolingan"),
        ("Holat: инициализация при следующей проверке", "Holat: keyingi tekshiruvda ishga tushadi"),
        ("Бот уведомит, когда товар впервые опустится до порога или ниже.", "Tovar birinchi marta belgilangan chegaragacha yoki undan pastga tushganda bot xabar beradi."),
        ("Бот уведомит, когда товар впервые опустится до остатка <b>0</b>.", "Tovar qoldig‘i birinchi marta <b>0</b> bo‘lganda bot xabar beradi."),

        # Подписка / тарифы
        ("💎 <b>Uzum Seller Assistant obunasi</b>", "💎 <b>Uzum Seller Assistant obunasi</b>"),
        ("Nimalar kiradi:", "Nimalar kiradi:"),
        ("1 месяц", "1 oy"),
        ("3 месяца", "3 oy"),
        ("6 месяцев", "6 oy"),
        ("1 oy — 300 000 so‘m | 3 oy — 800 000 so‘m | 6 oy — 1 500 000 so‘m | Без ограничений по количеству магазинов", "1 oy — 300 000 so‘m | 3 oy — 800 000 so‘m | 6 oy — 1 500 000 so‘m | Do‘konlar soni cheklanmagan"),
        ("Без ограничений по количеству магазинов", "Do‘konlar soni cheklanmagan"),
        ("Без ограничений по количеству do‘konlar", "Do‘konlar soni cheklanmagan"),
        ("To‘lov uchun administratorga yozing:", "To‘lov uchun administratorga yozing:"),
        ("Нажмите кнопку ниже, напишите администратору и отправьте чек. После проверки доступ будет продлён.", "Quyidagi tugmani bosing, administratorga yozing va chekni yuboring. Tekshiruvdan keyin kirish uzaytiriladi."),
        ("Нажмите кнопку ниже, напишите администратору и отправьте чек.", "Quyidagi tugmani bosing, administratorga yozing va chekni yuboring."),
        ("После проверки доступ будет продлён.", "Tekshiruvdan keyin kirish uzaytiriladi."),
        ("Чек tekshirilgach, administrator kirishni uzaytiradi.", "Chek tekshirilgach, administrator kirishni uzaytiradi."),
        ("Подписка", "Obuna"),
        ("подписка", "obuna"),

        # Остатки / продажи / статусы
        ("Статус:", "Holat:"),
        ("Status:", "Holat:"),
        ("Holat: ✅ включены", "Holat: ✅ yoqilgan"),
        ("Holat: ❌ выключены", "Holat: ❌ o‘chirilgan"),
        ("включены", "yoqilgan"),
        ("выключены", "o‘chirilgan"),
        ("Магазин:", "Do‘kon:"),
        ("Do‘kon:", "Do‘kon:"),
        ("сум", "so‘m"),
        ("сек", "soniya"),

        # Команды и подсказки
        ("Нажмите кнопку ниже", "Quyidagi tugmani bosing"),
        ("напишите администратору", "administratorga yozing"),
        ("отправьте чек", "chekni yuboring"),
        ("Проверка", "Tekshiruv"),
        ("Состояние", "Holat"),
        ("Порог", "Chegara"),
    ]
    for old, new in fixes:
        text = text.replace(old, new)

    # Небольшая нормализация после замен
    text = text.replace("soniya.", "soniya")
    text = text.replace("soniyaiya", "soniya")
    text = text.replace("Holat::", "Holat:")
    text = text.replace("Chegara::", "Chegara:")
    text = text.replace("Tekshiruv har: har", "Tekshiruv har")
    text = text.replace("Tekshiruv har har", "Tekshiruv har")
    return text


# --- FULL CHECK: финальная узбекская чистка перед запуском ---
# Этот слой специально стоит самым последним: исправляет смешанные русско-узбекские
# фразы, которые появляются из старых русских отчётов после автоматической замены.
_FULL_CHECK_TRANSLATE_RUNTIME_TEXT_TO_UZ = translate_runtime_text_to_uz


def translate_runtime_text_to_uz(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text
    text = _FULL_CHECK_TRANSLATE_RUNTIME_TEXT_TO_UZ(text)

    fixes = [
        # Частые артефакты от замен "день/дней" внутри слова "сегодня/средняя"
        ("сегоkun", "bugun"),
        ("Сегоkun", "Bugun"),
        ("сегодня", "bugun"),
        ("Сегодня", "Bugun"),
        ("Среkunя строка", "O‘rtacha savdo"),
        ("Средняя строка", "O‘rtacha savdo"),
        ("O‘rtacha строка", "O‘rtacha savdo"),
        ("Отменённых qatorlar", "Bekor qilinganlar"),
        ("Отмененных qatorlar", "Bekor qilinganlar"),
        ("Отменённых строк", "Bekor qilinganlar"),
        ("Отмененных строк", "Bekor qilinganlar"),

        # Подключение API
        ("🔌 <b>Подключение магазина</b>", "🔌 <b>Do‘konni ulash</b>"),
        ("Чтобы подключить магазин, создайте API-ключ в кабинете Uzum Seller и отправьте его сюда.", "Do‘konni ulash uchun Uzum Seller kabinetida API-kalit yarating va shu yerga yuboring."),
        ("🎥 Видеоинструкция", "🎥 Videoqo‘llanma"),
        ("📌 Текстовая инструкция", "📌 Matnli qo‘llanma"),
        ("Отправьте API-ключ следующим сообщением.", "API-kalitni keyingi xabarda yuboring."),
        ("Отмена:", "Bekor qilish:"),
        ("Похоже, это не Uzum API-токен.", "Bu Uzum API-kalitiga o‘xshamaydi."),
        ("Отправьте полный токен или нажмите", "To‘liq kalitni yuboring yoki bosing"),
        ("✅ <b>Магазин уже подключён</b>", "✅ <b>Do‘kon allaqachon ulangan</b>"),
        ("Если вы случайно нажали <b>🔌 Подключить</b>, ничего страшного — старый API-ключ не удалён и не слетит.", "Agar <b>🔌 Ulash</b> tugmasini tasodifan bosgan bo‘lsangiz, xavotir olmang — eski API-kalit o‘chirilmaydi."),
        ("Чтобы заменить API-ключ, используйте только команду", "API-kalitni almashtirish uchun faqat quyidagi buyruqdan foydalaning:"),
        ("Чтобы полностью удалить подключение:", "Ulanishni butunlay o‘chirish uchun:"),
        ("🔐 <b>Uzum API-ключ не принят</b>", "🔐 <b>Uzum API-kalit qabul qilinmadi</b>"),
        ("Возможно, ключ неверный, удалён или истёк.", "Kalit noto‘g‘ri, o‘chirilgan yoki muddati tugagan bo‘lishi mumkin."),
        ("Создайте новый ключ в кабинете Uzum Seller и подключите его через", "Uzum Seller kabinetida yangi kalit yarating va uni ulang:"),
        ("✅ Подключение к Uzum API удалено. Можно подключить заново через", "✅ Uzum API ulanishi o‘chirildi. Qayta ulash uchun:"),

        # Видео / инструкция / безопасность / поддержка
        ("🎥 <b>Видеоинструкция по подключению API</b>", "🎥 <b>API ulash bo‘yicha videoqo‘llanma</b>"),
        ("В видео коротко показано:", "Videoda qisqa ko‘rsatilgan:"),
        ("Где в кабинете Uzum Seller находятся ключи API.", "Uzum Seller kabinetida API kalitlari qayerda joylashgani."),
        ("Как создать новый ключ.", "Yangi kalitni qanday yaratish."),
        ("Как подключить ключ к боту через", "Kalitni botga qanday ulash:"),
        ("Нажмите кнопку ниже, чтобы открыть видео", "Videoni ochish uchun quyidagi tugmani bosing"),
        ("▶️ Смотреть видео", "▶️ Videoni ko‘rish"),
        ("🔑 <b>Как подключить Uzum API к боту</b>", "🔑 <b>Uzum API-kalitni botga qanday ulash mumkin</b>"),
        ("API-ключ создаётся только в вашем кабинете Uzum Seller.", "API-kalit faqat sizning Uzum Seller kabinetingizda yaratiladi."),
        ("Это не пароль от кабинета, ключ можно удалить в любой момент.", "Bu kabinet paroli emas, kalitni istalgan vaqtda o‘chirishingiz mumkin."),
        ("Где взять API-ключ:", "API-kalitni qayerdan olish mumkin:"),
        ("Нажмите на профиль / аватарку в правом верхнем углу.", "O‘ng yuqori burchakdagi profil / avatarkani bosing."),
        ("Откройте", "Oching:"),
        ("Ключи API", "API kalitlari"),
        ("Создать ключ", "Kalit yaratish"),
        ("Скопируйте ключ.", "Kalitni nusxalang."),
        ("Вернитесь в бот и отправьте ключ через", "Botga qayting va kalitni yuboring:"),
        ("🔐 <b>Безопасность API-ключа</b>", "🔐 <b>API-kalit xavfsizligi</b>"),
        ("Ваш Uzum API-ключ не показывается в боте и не отправляется обратно сообщением.", "Uzum API-kalitingiz botda ko‘rsatilmaydi va qayta xabar qilib yuborilmaydi."),
        ("После подключения бот старается удалить сообщение, где был отправлен ключ.", "Ulangandan keyin bot kalit yuborilgan xabarni o‘chirishga harakat qiladi."),
        ("В базе хранится только защищённая версия ключа.", "Bazaga kalitning faqat himoyalangan ko‘rinishi saqlanadi."),
        ("Вы можете в любой момент удалить подключение командой", "Ulanishni istalgan vaqtda quyidagi buyruq bilan o‘chirishingiz mumkin:"),
        ("🛟 <b>Поддержка Uzum Seller Assistant</b>", "🛟 <b>Uzum Seller Assistant yordami</b>"),
        ("Ваш Telegram ID:", "Telegram ID’ingiz:"),
        ("Если бот не показывает данные, проверьте:", "Agar bot ma’lumot ko‘rsatmasa, tekshiring:"),
        ("API-ключ активен в кабинете Uzum Seller.", "API-kalit Uzum Seller kabinetida faol."),
        ("У ключа есть доступ к нужному магазину.", "Kalit kerakli do‘konga kirish huquqiga ega."),
        ("В кабинете Uzum есть продажи за выбранный период.", "Uzum kabinetida tanlangan davr uchun savdolar bor."),
        ("Если меняли API-ключ — нажмите", "Agar API-kalitni almashtirgan bo‘lsangiz, bosing:"),

        # Подписка
        ("💎 <b>Подписка Uzum Seller Assistant</b>", "💎 <b>Uzum Seller Assistant obunasi</b>"),
        ("Что входит:", "Nimalar kiradi:"),
        ("продажи FBO/FBS за сегодня, вчера, 7 и 30 дней", "bugun, kecha, 7 va 30 kunlik FBO/FBS savdolar"),
        ("остатки и товары, которые заканчиваются", "qoldiqlar va tugab borayotgan tovarlar"),
        ("потерянные товары, если Uzum отдаёт их в API", "Uzum API bersa, yo‘qolgan tovarlar"),
        ("уведомления о новых продажах", "yangi savdolar haqida xabarnomalar"),
        ("работа с несколькими магазинами", "bir nechta do‘kon bilan ishlash"),
        ("Excel-отчёт и утренний отчёт", "Excel hisobot va ertalabki hisobot"),
        ("для нового пользователя", "yangi foydalanuvchi uchun"),
        ("💰 <b>Тарифы</b>", "💰 <b>Tariflar</b>"),
        ("Для оплаты напишите администратору", "To‘lov uchun administratorga yozing"),
        ("После проверки чекa администратор продлит доступ.", "Chek tekshirilgach, administrator kirishni uzaytiradi."),
        ("Проверить статус", "Holatni tekshirish"),
        ("Trial до:", "Trial muddati:"),
        ("Оплачено до:", "To‘langan sana:"),
        ("Тарифы:", "Tariflar:"),
        ("История оплат:", "To‘lovlar tarixi:"),
        ("Поддержка:", "Yordam:"),
        ("Заменить API-ключ:", "API-kalitni almashtirish:"),
        ("Удалить API-ключ:", "API-kalitni o‘chirish:"),
        ("⛔ Obuna закончилась", "⛔ Obuna muddati tugagan"),
        ("⛔ Подписка закончилась", "⛔ Obuna muddati tugagan"),
        ("👑 Админ-доступ: без ограничений", "👑 Admin kirish: cheklovsiz"),
        ("⛔ Пользователь заблокирован", "⛔ Foydalanuvchi bloklangan"),

        # Разделы / меню / общее
        ("Русский", "Rus tili"),
        ("Выберите действие", "Amalni tanlang"),
        ("Выберите раздел", "Bo‘limni tanlang"),
        ("Главное меню", "Asosiy menyu"),
        ("Действие отменено.", "Amal bekor qilindi."),
        ("Язык интерфейса", "Interfeys tili"),
        ("Выберите язык, на котором бот будет показывать меню и основные подсказки.", "Bot menyu va asosiy ko‘rsatmalarni qaysi tilda ko‘rsatishini tanlang."),
        ("Язык изменён", "Til o‘zgartirildi"),
        ("Админ-панель доступна только владельцу бота.", "Admin panel faqat bot egasi uchun."),
        ("Доступ ограничен", "Kirish cheklangan"),
        ("Trial или подписка закончились.", "Trial yoki obuna muddati tugagan."),
        ("Ваш Uzum-токен и настройки сохранены — после продления всё снова заработает.", "Uzum tokeningiz va sozlamalaringiz saqlanadi — obuna uzaytirilgach hammasi yana ishlaydi."),
        ("Проверить подписку", "Obunani tekshirish"),
        ("Оплата", "To‘lov"),
        ("Сначала подключите магазин", "Avval do‘konni ulang"),
        ("Сначала подключите Uzum API-токен", "Avval Uzum API-kalitini ulang"),
        ("Продажи", "Savdo"),
        ("Склад", "Ombor"),
        ("Уведомления", "Xabarnomalar"),
        ("Отчёты", "Hisobotlar"),
        ("Что проверить", "Tekshirish"),
        ("Помощь", "Yordam"),
        ("Подключить", "Ulash"),
        ("Магазины", "Do‘konlar"),
        ("Пользователь", "Foydalanuvchi"),
        ("Доступ", "Kirish"),
        ("Проверить подключение", "Ulanishni tekshirish"),
        ("Что делать", "Nima qilish kerak"),
        ("отправьте Uzum API-ключ", "Uzum API-kalitini yuboring"),

        # Продажи / финансы / отчёты
        ("💰 <b>Продажи за bugun</b>", "💰 <b>Bugungi savdo</b>"),
        ("💰 <b>Продажи за Bugun</b>", "💰 <b>Bugungi savdo</b>"),
        ("💰 <b>Продажи ", "💰 <b>Savdo "),
        ("Продаж:", "Savdolar:"),
        ("Товаров продано:", "Sotilgan tovarlar:"),
        ("Возвратов:", "Qaytarilganlar:"),
        ("Продаж не найдено", "Savdolar topilmadi"),
        ("Пока продаж tanlangan davr uchun не найдено.", "Tanlangan davr uchun hali savdolar topilmadi."),
        ("Если продажа только появилась в кабинете, она может отобразиться чуть позже.", "Agar savdo kabinetda endi paydo bo‘lgan bo‘lsa, botda biroz keyin ko‘rinishi mumkin."),
        ("Ответ Finance API пришёл, но строки продаж не найдены.", "Finance API javob berdi, lekin savdo qatorlari topilmadi."),
        ("Фрагмент ответа", "Javobdan parcha"),
        ("Показаны первые", "Birinchi"),
        ("позиций из", "pozitsiya ko‘rsatildi, jami"),
        ("Новая продажа", "Yangi savdo"),
        ("Заказ", "Buyurtma"),
        ("Цена:", "Narx:"),
        ("Чистая прибыль", "Sof foyda"),
        ("Топ товаров по прибыли", "Foyda bo‘yicha top tovarlar"),
        ("Top tovarlar по прибыли", "Foyda bo‘yicha top tovarlar"),
        ("Загрузить себестоимость", "Tannarxni yuklash"),
        ("Себестоимостей сохранено", "Saqlangan tannarxlar"),
        ("Себестоимость ещё не указана", "Tannarx hali ko‘rsatilmagan"),
        ("Добавить", "Qo‘shish"),
        ("Укажите себестоимость так", "Tannarxni shunday kiriting"),
        ("Например", "Masalan"),
        ("Теперь проверьте", "Endi tekshiring"),
        ("Юнит-экономику", "Unit iqtisodiyot"),
        ("Unit iqtisodiyot за", "Unit iqtisodiyot"),

        # Склад / товары / накладные
        ("📦 <b>Товары магазина</b>", "📦 <b>Do‘kon tovarlari</b>"),
        ("Tovarы магазина", "Do‘kon tovarlari"),
        ("📊 <b>Остатки FBO / склад Uzum</b>", "📊 <b>FBO qoldiqlari / Uzum ombori</b>"),
        ("Qoldiq FBO / склад Uzum", "FBO qoldiqlari / Uzum ombori"),
        ("Проверяю потерянные товары", "Yo‘qolgan tovarlar tekshirilmoqda"),
        ("Раздел использует поле quantityMissing из Products API. Если в кабинете Uzum потери считаются по актам иначе, сумма может отличаться.", "Bo‘lim Products API’dagi quantityMissing maydonidan foydalanadi. Agar Uzum kabinetida yo‘qotishlar boshqacha hisoblangan bo‘lsa, summa farq qilishi mumkin."),
        ("Загружаю FBO-накладные поставки", "FBO yuk xatlari yuklanmoqda"),
        ("Чтобы посмотреть состав, отправьте", "Tarkibini ko‘rish uchun yuboring"),
        ("Например:", "Masalan:"),
        ("Заказов по основным статусам не найдено", "Asosiy statuslar bo‘yicha buyurtmalar topilmadi"),
        ("Собираю общую сводку магазина", "Do‘kon bo‘yicha umumiy xulosa tayyorlanmoqda"),
        ("остаток изменился", "qoldiq o‘zgardi"),
        ("Показать товары с низким остатком", "Kam qoldiqdagi tovarlarni ko‘rsatish"),
        ("Готовлю подробный Excel-отчёт", "Batafsil Excel hisobot tayyorlanmoqda"),
        ("Это может занять 20–60 секунд", "Bu 20–60 soniya vaqt olishi mumkin"),
        ("собираю продажи, остатки и FBO-накладные", "savdolar, qoldiqlar va FBO yuk xatlari yig‘ilmoqda"),

        # Уведомления / отмены
        ("❌ <b>Уведомления об отменах</b>", "❌ <b>Bekor qilishlar xabarnomalari</b>"),
        ("Бот отслеживает новые отмены через Finance API.", "Bot yangi bekor qilishlarni Finance API orqali kuzatadi."),
        ("🛒 <b>Новая продажа</b>", "🛒 <b>Yangi savdo</b>"),

        # Что требует внимания
        ("Критичных проблем не видно. Можно посмотреть продажи и прибыль 30 kun uchun.", "Jiddiy muammo ko‘rinmayapti. 30 kunlik savdo va foydani ko‘rishingiz mumkin."),
        ("Начните с товаров, которые закончились: они не смогут продаваться, пока не пополнятся остатки.", "Avval tugagan tovarlardan boshlang: qoldiq to‘ldirilmaguncha ular sotilmaydi."),
        ("Сначала проверьте товары, которые скоро закончатся, особенно если они хорошо продаются.", "Avval tez tugaydigan tovarlarni tekshiring, ayniqsa ular yaxshi sotilayotgan bo‘lsa."),
        ("Проверьте purchasePrice в карточках Uzum: бот берёт себестоимость только из Uzum и не подставляет её вручную.", "Uzum kartalaridagi purchasePrice qiymatini tekshiring: bot tannarxni faqat Uzumdan oladi va qo‘lda almashtirmaydi."),
        ("Проверьте товары с низкой маржой: возможно, цена или себестоимость указаны невыгодно.", "Past marjali tovarlarni tekshiring: narx yoki tannarx foydasiz bo‘lishi mumkin."),
        ("Посмотрите товары без продаж: возможно, стоит изменить цену, фото или вывести товар из оборота.", "Sotilmayotgan tovarlarni ko‘ring: narxni, rasmlarni o‘zgartirish yoki tovarni chiqarish kerak bo‘lishi mumkin."),
        ("Проверьте сегодняшние отмены и товары, по которым они произошли.", "Bugungi bekor qilishlar va ular bo‘lgan tovarlarni tekshiring."),
        ("Ниже можете сразу открыть нужный раздел", "Quyida kerakli bo‘limni darhol ochishingiz mumkin"),
        ("Проверяю магазин", "Do‘kon tekshirilmoqda"),
        ("Скоро закончится", "Tez tugaydi"),
        ("Закончились", "Tugagan"),
        ("Без продаж", "Sotuv yo‘q"),
        ("Без себестоимости", "Tannarx yo‘q"),
        ("Низкая маржа", "Past marja"),
        ("Низкая прибыль", "Past foyda"),
        ("Отмены сегодня", "Bugungi bekor qilishlar"),
        ("Потерянные", "Yo‘qolganlar"),
        ("Рекомендация", "Tavsiya"),

        # Excel / выгрузки
        ("Сводка", "Xulosa"),
        ("Показатель", "Ko‘rsatkich"),
        ("Значение", "Qiymat"),
        ("Дата создания отчёта", "Hisobot yaratilgan sana"),
        ("Период продаж в деталях", "Savdolar davri batafsil"),
        ("SKU в остатках", "Qoldiqdagi SKU"),
        ("SKU заканчиваются", "Tugayotgan SKU"),
        ("SKU с потерями", "Yo‘qotishli SKU"),
        ("FBO накладных найдено", "FBO yuk xatlari topildi"),
        ("Состав накладных загружен", "Yuk xatlari tarkibi yuklandi"),
        ("Период", "Davr"),
        ("Позиций", "Pozitsiyalar"),
        ("Товаров, dona", "Tovarlar, dona"),
        ("ID заказа/операции", "Buyurtma/operatsiya ID"),
        ("SKU/код", "SKU/kod"),
        ("Выведено", "Chiqarilgan"),
        ("Сырой фрагмент", "Xom parcha"),
        ("Код продавца", "Sotuvchi kodi"),
        ("Категория", "Kategoriya"),
        ("Цена", "Narx"),
        ("Активно", "Faol"),
        ("Потеряно", "Yo‘qolgan"),
        ("Брак", "Yaroqsiz"),
        ("Ожидает", "Kutilmoqda"),
        ("Примерная so‘mма", "Taxminiy summa"),
        ("Номер", "Raqam"),
        ("Создана", "Yaratilgan"),
        ("Создан", "Yaratilgan"),
        ("Окно от", "Oyna boshi"),
        ("Окно до", "Oyna oxiri"),
        ("Принята", "Qabul qilingan"),
        ("По накладной", "Yuk xati bo‘yicha"),
        ("Закупочная цена", "Xarid narxi"),
        ("Сумма по накладной", "Yuk xati summasi"),

        # Админка — чтобы в узбекском режиме тоже не было каши, но команды оставляем как есть
        ("👑 Админ", "👑 Admin"),
        ("👥 Пользователи", "👥 Foydalanuvchilar"),
        ("💳 Оплаты", "💳 To‘lovlar"),
        ("⏳ Скоро заканчиваются", "⏳ Tugayotganlar"),
        ("⛔ Заблокированные", "⛔ Bloklanganlar"),
        ("📦 Бэкап базы", "📦 Baza zaxirasi"),
        ("📢 Рассылка", "📢 Xabar yuborish"),
        ("⬅️ Главное меню", "⬅️ Asosiy menyu"),
        ("Чтобы отправить сообщение всем пользователям, напишите", "Barcha foydalanuvchilarga xabar yuborish uchun yozing"),
        ("Пример", "Masalan"),
        ("Отправляю резервную копию базы. Храните файл аккуратно — там данные пользователей.", "Baza zaxirasi yuborilmoqda. Faylni ehtiyot saqlang — unda foydalanuvchilar ma’lumotlari bor."),
    ]

    for old, new in fixes:
        text = text.replace(old, new)

    # Yakuniy normalizatsiya: eng ko‘p uchraydigan aralashmalar
    import re
    text = re.sub(r"(?<=\d)\s*сум\b", " so‘m", text)
    text = re.sub(r"(?<=\d)\s*шт\.?\b", " dona", text)
    text = text.replace("so‘mма", "summa")
    text = text.replace("so‘mмы", "summaning")
    text = text.replace("Tovarов", "Tovarlar")
    text = text.replace("Qaytarilganов", "Qaytarilganlar")
    text = text.replace("Savdo 30 kun uchun", "30 kunlik savdo")
    text = text.replace("Savdo 7 kun uchun", "7 kunlik savdo")
    text = text.replace("Savdo bugun uchun", "Bugungi savdo")
    text = text.replace("Savdo kecha uchun", "Kechagi savdo")
    text = text.replace("за bugun", "bugun uchun")
    text = text.replace("за Kecha", "kecha uchun")
    text = text.replace("за 7 kun", "7 kun uchun")
    text = text.replace("за 30 kun", "30 kun uchun")
    text = text.replace(" | 1 oy", " | 1 oy")
    text = text.replace("soniyaунд", "soniya")
    text = text.replace("soniyaия", "soniya")
    text = text.replace("Holat::", "Holat:")
    text = text.replace("Chegara::", "Chegara:")
    text = text.replace("Tekshiruv har: ", "Tekshiruv har ")
    text = text.replace("Tekshiruv har <b>300</b> soniya", "Tekshiruv har <b>300</b> soniyada")
    text = text.replace("Tekshiruv har <b>1800</b> soniya", "Tekshiruv har <b>1800</b> soniyada")
    return text


# --- FINAL AUDIT LAYER: исправления после полной проверки примеров ---
_AUDIT_TRANSLATE_RUNTIME_TEXT_TO_UZ = translate_runtime_text_to_uz


def translate_runtime_text_to_uz(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text
    text = _AUDIT_TRANSLATE_RUNTIME_TEXT_TO_UZ(text)
    fixes = [
        ("💎 <b>Obuna Uzum Seller Assistant</b>", "💎 <b>Uzum Seller Assistant obunasi</b>"),
        ("✅ продажи FBO/FBS bugun uchun, вчера, 7 и 30 kun", "✅ bugun, kecha, 7 va 30 kunlik FBO/FBS savdolar"),
        ("✅ продажи FBO/FBS bugun uchun, kecha, 7 va 30 kun", "✅ bugun, kecha, 7 va 30 kunlik FBO/FBS savdolar"),
        ("продажи FBO/FBS bugun uchun, вчера, 7 и 30 kun", "bugun, kecha, 7 va 30 kunlik FBO/FBS savdolar"),
        ("Для оплаты administratorga yozing", "To‘lov uchun administratorga yozing"),
        ("Для оплаты", "To‘lov uchun"),
        ("Tovarlar продано", "Sotilgan tovarlar"),
        ("Tovarlar sotildi", "Sotilgan tovarlar"),
        ("Sotilgan tovarlar:", "Sotilgan tovarlar:"),
        ("✅ <b>Do‘kon уже подключён</b>", "✅ <b>Do‘kon allaqachon ulangan</b>"),
        ("Do‘kon уже подключён", "Do‘kon allaqachon ulangan"),
        ("Savdo bugun uchun", "Bugungi savdo"),
        ("Savdo kecha uchun", "Kechagi savdo"),
        ("Savdo 7 kun uchun", "7 kunlik savdo"),
        ("Savdo 30 kun uchun", "30 kunlik savdo"),
        ("Продажи bugun uchun", "Bugungi savdo"),
        ("Продажи kecha uchun", "Kechagi savdo"),
        ("Продажи 7 kun uchun", "7 kunlik savdo"),
        ("Продажи 30 kun uchun", "30 kunlik savdo"),
        ("Sotuvlar bugun uchun", "Bugungi savdo"),
        ("Sotuvlar kecha uchun", "Kechagi savdo"),
        ("Sotuvlar 7 kun uchun", "7 kunlik savdo"),
        ("Sotuvlar 30 kun uchun", "30 kunlik savdo"),
        ("вчера", "kecha"),
        ("7 и 30", "7 va 30"),
        ("и 30", "va 30"),
        ("уже", "allaqachon"),
        ("подключён", "ulangan"),
        ("подключен", "ulangan"),
        ("продано", "sotilgan"),
        ("продажи", "savdolar"),
        ("Продажи", "Savdolar"),
        ("товары", "tovarlar"),
        ("Товары", "Tovarlar"),
        ("остатки", "qoldiqlar"),
        ("Остатки", "Qoldiqlar"),
        ("которые заканчиваются", "tugab borayotgan"),
        ("уведомления", "xabarnomalar"),
        ("Уведомления", "Xabarnomalar"),
        ("новых", "yangi"),
        ("несколькими магазинами", "bir nechta do‘kon"),
        ("работа с", "ishlash:"),
        ("потерянные", "yo‘qolgan"),
        ("Потерянные", "Yo‘qolgan"),
        ("если Uzum отдаёт их в API", "agar Uzum API’da bersa"),
        ("для", "uchun"),
        ("нового пользователя", "yangi foydalanuvchi"),
    ]
    for old, new in fixes:
        text = text.replace(old, new)
    # Последняя нормализация заголовков и фраз после всех замен
    text = text.replace("💰 <b>Savdo bugun uchun</b>", "💰 <b>Bugungi savdo</b>")
    text = text.replace("💰 <b>Savdo kecha uchun</b>", "💰 <b>Kechagi savdo</b>")
    text = text.replace("💰 <b>Savdo 7 kun uchun</b>", "💰 <b>7 kunlik savdo</b>")
    text = text.replace("💰 <b>Savdo 30 kun uchun</b>", "💰 <b>30 kunlik savdo</b>")
    text = text.replace("💰 <b>Bugungi savdo uchun</b>", "💰 <b>Bugungi savdo</b>")
    text = text.replace("💰 <b>Kechagi savdo uchun</b>", "💰 <b>Kechagi savdo</b>")
    text = text.replace("savdolar FBO/FBS bugun uchun, kecha, 7 va 30 kun", "bugun, kecha, 7 va 30 kunlik FBO/FBS savdolar")
    text = text.replace("Tovarlar sotilgan", "Sotilgan tovarlar")
    return text


# --- Надёжный watcher отмен: хранит статусы в БД и проверяет последние 7 дней ---
# Причина: память Python очищается при SIGTERM/перезапуске BotHost. Без БД отмена,
# произошедшая во время перезапуска, могла быть принята за исходное состояние и не отправлялась.
def init_sale_status_watch_table() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sale_status_watch (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                identity_key TEXT NOT NULL,
                status TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, identity_key)
            )
            """
        )
        conn.commit()


def load_saved_sale_statuses(telegram_id: int, shop_id: int) -> dict[str, str]:
    init_sale_status_watch_table()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT identity_key, status
            FROM sale_status_watch
            WHERE telegram_id = ? AND shop_id = ?
            """,
            (int(telegram_id), int(shop_id)),
        ).fetchall()
    return {str(row["identity_key"]): str(row["status"] or "") for row in rows}


def save_sale_statuses(telegram_id: int, shop_id: int, statuses: dict[str, str]) -> None:
    if not statuses:
        return
    init_sale_status_watch_table()
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        conn.executemany(
            """
            INSERT INTO sale_status_watch
                (telegram_id, shop_id, identity_key, status, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id, identity_key) DO UPDATE SET
                status = excluded.status,
                updated_at = excluded.updated_at
            """,
            [
                (int(telegram_id), int(shop_id), str(identity), str(status), now_text)
                for identity, status in statuses.items()
            ],
        )
        conn.execute(
            """
            DELETE FROM sale_status_watch
            WHERE telegram_id = ? AND shop_id = ? AND updated_at < ?
            """,
            (
                int(telegram_id),
                int(shop_id),
                _dt_to_db(_utc_now() - timedelta(days=30)) or "",
            ),
        )
        conn.commit()


init_sale_status_watch_table()


def _sales_digest_product_title(item: dict[str, Any]) -> str:
    for field in ("productTitle", "productName", "offerName", "title", "name", "skuTitle"):
        value = _deep_pick_value(item, (field,))
        if isinstance(value, dict):
            value = pick(value, "title", "name", "value")
        if value not in (None, ""):
            return str(value)
    return _finance_title(item)


def _sales_digest_event(
    item: dict[str, Any],
    *,
    detected_at: datetime | None = None,
) -> dict[str, Any]:
    normalized = _normalize_finance_row(item)
    sold_at = normalized.get("date")
    if isinstance(sold_at, datetime) and sold_at.tzinfo is None:
        sold_at = sold_at.replace(tzinfo=UZT)
    detected = detected_at or _utc_now()
    payout = max(0.0, float(normalized.get("payout") or 0))
    return {
        "event_key": sale_key(item),
        "identity_key": finance_identity_key(item),
        "order_id": str(normalized.get("order_id") or "-"),
        "product_title": _sales_digest_product_title(item),
        "sku_title": str(normalized.get("sku") or _finance_sku_title(item) or "-"),
        "quantity": max(0.0, float(normalized.get("qty") or 0)),
        "revenue": max(0.0, float(normalized.get("revenue") or 0)),
        "commission": max(0.0, float(normalized.get("commission") or 0)),
        "logistics": max(0.0, float(normalized.get("logistics") or 0)),
        "payout": payout,
        "sold_at": _dt_to_db(sold_at) if isinstance(sold_at, datetime) else None,
        "detected_at": _dt_to_db(detected) or "",
    }


def enqueue_sales_digest_events(
    telegram_id: int,
    shop_id: int,
    items: list[dict[str, Any]],
    *,
    detected_at: datetime | None = None,
) -> int:
    if not items:
        return 0
    init_product_value_tables()
    detected = detected_at or _utc_now()
    events = [_sales_digest_event(item, detected_at=detected) for item in items]
    cutoff = _dt_to_db(detected - timedelta(days=7)) or ""
    with db.connect() as conn:
        before = conn.total_changes
        conn.executemany(
            """
            INSERT OR IGNORE INTO sales_digest_queue (
                telegram_id, shop_id, event_key, identity_key, order_id,
                product_title, sku_title, quantity, revenue, commission,
                logistics, payout, sold_at, detected_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    int(telegram_id),
                    int(shop_id),
                    event["event_key"],
                    event["identity_key"],
                    event["order_id"],
                    event["product_title"],
                    event["sku_title"],
                    event["quantity"],
                    event["revenue"],
                    event["commission"],
                    event["logistics"],
                    event["payout"],
                    event["sold_at"],
                    event["detected_at"],
                )
                for event in events
            ],
        )
        inserted = conn.total_changes - before
        conn.execute(
            """
            DELETE FROM sales_digest_queue
            WHERE telegram_id = ? AND shop_id = ? AND detected_at < ?
            """,
            (int(telegram_id), int(shop_id), cutoff),
        )
        conn.commit()
    return max(0, int(inserted))


def discard_sales_digest_events(
    telegram_id: int,
    shop_id: int,
    identity_keys: Iterable[str],
) -> int:
    values = sorted({str(value) for value in identity_keys if str(value)})
    if not values:
        return 0
    init_product_value_tables()
    deleted = 0
    with db.connect() as conn:
        for offset in range(0, len(values), 500):
            chunk = values[offset:offset + 500]
            placeholders = ",".join("?" for _ in chunk)
            cursor = conn.execute(
                f"""
                DELETE FROM sales_digest_queue
                WHERE telegram_id = ? AND shop_id = ?
                  AND identity_key IN ({placeholders})
                """,
                (int(telegram_id), int(shop_id), *chunk),
            )
            deleted += max(0, int(cursor.rowcount or 0))
        conn.commit()
    return deleted


def reset_sales_digest_schedule(
    telegram_id: int,
    shop_id: int,
    *,
    now: datetime | None = None,
    clear_queue: bool = False,
) -> None:
    init_product_value_tables()
    now_text = _dt_to_db(now or _utc_now()) or ""
    with db.connect() as conn:
        if clear_queue:
            conn.execute(
                "DELETE FROM sales_digest_queue WHERE telegram_id = ? AND shop_id = ?",
                (int(telegram_id), int(shop_id)),
            )
        conn.execute(
            """
            INSERT INTO sales_digest_state
                (telegram_id, shop_id, last_sent_at, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id) DO UPDATE SET
                last_sent_at = excluded.last_sent_at,
                updated_at = excluded.updated_at
            """,
            (int(telegram_id), int(shop_id), now_text, now_text),
        )
        conn.commit()


def get_sales_digest_last_sent(
    telegram_id: int,
    shop_id: int,
    *,
    now: datetime | None = None,
) -> datetime:
    init_product_value_tables()
    current = now or _utc_now()
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT last_sent_at
            FROM sales_digest_state
            WHERE telegram_id = ? AND shop_id = ?
            """,
            (int(telegram_id), int(shop_id)),
        ).fetchone()
    parsed = _dt_from_db(row["last_sent_at"]) if row else None
    if parsed is not None:
        return parsed
    reset_sales_digest_schedule(telegram_id, shop_id, now=current)
    return current


def _sales_digest_summary_from_events(
    events: list[dict[str, Any]],
    *,
    shop_id: int,
    period_start: datetime,
    period_end: datetime,
) -> dict[str, Any]:
    order_ids = {
        str(event.get("order_id"))
        for event in events
        if str(event.get("order_id") or "") not in {"", "-", "—"}
    }
    products: dict[tuple[str, str], dict[str, Any]] = {}
    for event in events:
        title = str(event.get("product_title") or "Без названия")
        sku = str(event.get("sku_title") or "-")
        product = products.setdefault(
            (title, sku),
            {"title": title, "sku": sku, "quantity": 0.0, "revenue": 0.0},
        )
        product["quantity"] = float(product["quantity"]) + float(event.get("quantity") or 0)
        product["revenue"] = float(product["revenue"]) + float(event.get("revenue") or 0)
    positions = len(events)
    revenue = sum(float(event.get("revenue") or 0) for event in events)
    orders = len(order_ids) or positions
    return {
        "shop_id": int(shop_id),
        "positions": positions,
        "orders": orders,
        "units": sum(float(event.get("quantity") or 0) for event in events),
        "revenue": revenue,
        "commission": sum(float(event.get("commission") or 0) for event in events),
        "logistics": sum(float(event.get("logistics") or 0) for event in events),
        "payout": sum(float(event.get("payout") or 0) for event in events),
        "average_check": revenue / max(1, orders),
        "product_count": len(products),
        "top_products": sorted(
            products.values(),
            key=lambda value: (float(value.get("revenue") or 0), float(value.get("quantity") or 0)),
            reverse=True,
        )[:5],
        "period_start": period_start,
        "period_end": period_end,
    }


def summarize_sales_digest_items(
    items: list[dict[str, Any]],
    *,
    shop_id: int,
    period_start: datetime,
    period_end: datetime,
) -> dict[str, Any]:
    events = [_sales_digest_event(item, detected_at=period_end) for item in items]
    return _sales_digest_summary_from_events(
        events,
        shop_id=shop_id,
        period_start=period_start,
        period_end=period_end,
    )


def load_sales_digest_summary(
    telegram_id: int,
    shop_id: int,
    *,
    period_start: datetime,
    period_end: datetime,
) -> dict[str, Any]:
    init_product_value_tables()
    params = (int(telegram_id), int(shop_id))
    with db.connect() as conn:
        totals = conn.execute(
            """
            SELECT
                COUNT(*) AS positions,
                COUNT(DISTINCT CASE
                    WHEN order_id IS NOT NULL AND order_id NOT IN ('', '-', '—')
                    THEN order_id END
                ) AS orders,
                COALESCE(SUM(quantity), 0) AS units,
                COALESCE(SUM(revenue), 0) AS revenue,
                COALESCE(SUM(commission), 0) AS commission,
                COALESCE(SUM(logistics), 0) AS logistics,
                COALESCE(SUM(payout), 0) AS payout,
                COUNT(DISTINCT COALESCE(product_title, '') || char(31) || COALESCE(sku_title, ''))
                    AS product_count
            FROM sales_digest_queue
            WHERE telegram_id = ? AND shop_id = ?
            """,
            params,
        ).fetchone()
        top_rows = conn.execute(
            """
            SELECT
                COALESCE(product_title, 'Без названия') AS title,
                COALESCE(sku_title, '-') AS sku,
                COALESCE(SUM(quantity), 0) AS quantity,
                COALESCE(SUM(revenue), 0) AS revenue
            FROM sales_digest_queue
            WHERE telegram_id = ? AND shop_id = ?
            GROUP BY product_title, sku_title
            ORDER BY revenue DESC, quantity DESC
            LIMIT 5
            """,
            params,
        ).fetchall()
    positions = int(totals["positions"] or 0) if totals else 0
    orders = int(totals["orders"] or 0) if totals else 0
    revenue = float(totals["revenue"] or 0) if totals else 0.0
    return {
        "shop_id": int(shop_id),
        "positions": positions,
        "orders": orders or positions,
        "units": float(totals["units"] or 0) if totals else 0.0,
        "revenue": revenue,
        "commission": float(totals["commission"] or 0) if totals else 0.0,
        "logistics": float(totals["logistics"] or 0) if totals else 0.0,
        "payout": float(totals["payout"] or 0) if totals else 0.0,
        "average_check": revenue / max(1, orders or positions),
        "product_count": int(totals["product_count"] or 0) if totals else 0,
        "top_products": [dict(row) for row in top_rows],
        "period_start": period_start,
        "period_end": period_end,
    }


def _sales_digest_period_text(start: datetime, end: datetime, lang: str) -> str:
    start_local = start.astimezone(UZT) if start.tzinfo else start.replace(tzinfo=UZT)
    end_local = end.astimezone(UZT) if end.tzinfo else end.replace(tzinfo=UZT)
    if start_local.date() == end_local.date():
        return f"{start_local:%d.%m.%Y}, {start_local:%H:%M}–{end_local:%H:%M}"
    return f"{start_local:%d.%m %H:%M}–{end_local:%d.%m %H:%M}"


def build_sales_digest_message(
    summary: dict[str, Any],
    *,
    lang: str = "ru",
    burst: bool = False,
) -> str:
    uz = normalize_lang(lang) == "uz"
    period = _sales_digest_period_text(
        summary.get("period_start") or _utc_now(),
        summary.get("period_end") or _utc_now(),
        lang,
    )
    positions = int(summary.get("positions") or 0)
    orders = int(summary.get("orders") or positions)
    units = float(summary.get("units") or 0)
    product_count = int(summary.get("product_count") or 0)
    top_products = list(summary.get("top_products") or [])
    if uz:
        lines = [
            "🔥 <b>Savdolar oqimi bo‘yicha hisobot</b>" if burst else "🕐 <b>Soatlik savdo hisoboti</b>",
            f"🏪 Do‘kon: <code>{int(summary.get('shop_id') or 0)}</code>",
            f"🕒 Davr: <b>{period}</b>",
            "",
            f"🧾 Buyurtmalar: <b>{orders}</b> | Pozitsiyalar: <b>{positions}</b>",
            f"📦 Sotildi: <b>{units:g} dona</b>",
            f"💵 Tushum: <b>{_format_money(float(summary.get('revenue') or 0))}</b>",
            f"🧮 O‘rtacha chek: <b>{_format_money(float(summary.get('average_check') or 0))}</b>",
            f"🏷 Komissiya: <b>{_format_money(float(summary.get('commission') or 0))}</b>",
            f"🚚 Logistika: <b>{_format_money(float(summary.get('logistics') or 0))}</b>",
            f"✅ To‘lovga: <b>{_format_money(float(summary.get('payout') or 0))}</b>",
        ]
        if top_products:
            lines.extend(["", "🏆 <b>Top tovarlar:</b>"])
            for index, product in enumerate(top_products, start=1):
                title = escape(_short_text(str(product.get("title") or product.get("sku") or "-"), 55))
                lines.append(
                    f"{index}. {title} — <b>{float(product.get('quantity') or 0):g} dona</b> · "
                    f"{_format_money(float(product.get('revenue') or 0))}"
                )
        if product_count > len(top_products):
            lines.append(f"Yana tovarlar: <b>{product_count - len(top_products)}</b>")
        return "\n".join(lines)

    lines = [
        "🔥 <b>Сводка по потоку продаж</b>" if burst else "🕐 <b>Продажи за час</b>",
        f"🏪 Магазин: <code>{int(summary.get('shop_id') or 0)}</code>",
        f"🕒 Период: <b>{period}</b>",
        "",
        f"🧾 Заказов: <b>{orders}</b> | Позиций: <b>{positions}</b>",
        f"📦 Продано: <b>{units:g} шт.</b>",
        f"💵 Выручка: <b>{_format_money(float(summary.get('revenue') or 0))}</b>",
        f"🧮 Средний чек: <b>{_format_money(float(summary.get('average_check') or 0))}</b>",
        f"🏷 Комиссия: <b>{_format_money(float(summary.get('commission') or 0))}</b>",
        f"🚚 Логистика: <b>{_format_money(float(summary.get('logistics') or 0))}</b>",
        f"✅ К выплате: <b>{_format_money(float(summary.get('payout') or 0))}</b>",
    ]
    if top_products:
        lines.extend(["", "🏆 <b>Топ товаров:</b>"])
        for index, product in enumerate(top_products, start=1):
            title = escape(_short_text(str(product.get("title") or product.get("sku") or "-"), 55))
            lines.append(
                f"{index}. {title} — <b>{float(product.get('quantity') or 0):g} шт.</b> · "
                f"{_format_money(float(product.get('revenue') or 0))}"
            )
    if product_count > len(top_products):
        lines.append(f"Ещё товаров: <b>{product_count - len(top_products)}</b>")
    return "\n".join(lines)


def mark_sales_digest_sent(
    telegram_id: int,
    shop_id: int,
    *,
    sent_at: datetime | None = None,
    clear_queue: bool = True,
) -> None:
    init_product_value_tables()
    now_text = _dt_to_db(sent_at or _utc_now()) or ""
    with db.connect() as conn:
        if clear_queue:
            conn.execute(
                "DELETE FROM sales_digest_queue WHERE telegram_id = ? AND shop_id = ?",
                (int(telegram_id), int(shop_id)),
            )
        conn.execute(
            """
            INSERT INTO sales_digest_state
                (telegram_id, shop_id, last_sent_at, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id) DO UPDATE SET
                last_sent_at = excluded.last_sent_at,
                updated_at = excluded.updated_at
            """,
            (int(telegram_id), int(shop_id), now_text, now_text),
        )
        conn.commit()


async def maybe_send_hourly_sales_digest(
    telegram_id: int,
    shop_id: int,
    *,
    now: datetime | None = None,
) -> bool:
    current = now or _utc_now()
    last_sent = get_sales_digest_last_sent(telegram_id, shop_id, now=current)
    if current - last_sent < timedelta(seconds=SALES_DIGEST_INTERVAL_SECONDS):
        return False
    summary = load_sales_digest_summary(
        telegram_id,
        shop_id,
        period_start=last_sent,
        period_end=current,
    )
    if int(summary.get("positions") or 0) <= 0:
        mark_sales_digest_sent(
            telegram_id,
            shop_id,
            sent_at=current,
            clear_queue=False,
        )
        return False
    try:
        await bot.send_message(
            telegram_id,
            build_sales_digest_message(
                summary,
                lang=get_user_language(telegram_id),
            ),
            reply_markup=main_menu_for_user(telegram_id),
        )
    except Exception:
        logging.exception(
            "Hourly sales digest: delivery failed user=%s shop=%s events=%s",
            telegram_id,
            shop_id,
            summary.get("positions"),
        )
        return False
    mark_sales_digest_sent(telegram_id, shop_id, sent_at=current)
    logging.info(
        "Hourly sales digest sent user=%s shop=%s orders=%s positions=%s units=%s revenue=%s",
        telegram_id,
        shop_id,
        summary.get("orders"),
        summary.get("positions"),
        summary.get("units"),
        summary.get("revenue"),
    )
    return True


def reset_user_sales_digest_schedule(
    telegram_id: int,
    *,
    now: datetime | None = None,
    clear_queue: bool = False,
) -> None:
    """Reset the single hourly schedule shared by all of a user's shops."""
    init_product_value_tables()
    now_text = _dt_to_db(now or _utc_now()) or ""
    with db.connect() as conn:
        if clear_queue:
            conn.execute(
                "DELETE FROM sales_digest_queue WHERE telegram_id = ?",
                (int(telegram_id),),
            )
        conn.execute(
            """
            INSERT INTO sales_digest_user_state
                (telegram_id, last_sent_at, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(telegram_id) DO UPDATE SET
                last_sent_at = excluded.last_sent_at,
                updated_at = excluded.updated_at
            """,
            (int(telegram_id), now_text, now_text),
        )
        conn.execute(
            """
            UPDATE sales_digest_state
            SET last_sent_at = ?, updated_at = ?
            WHERE telegram_id = ?
            """,
            (now_text, now_text, int(telegram_id)),
        )
        conn.commit()


def get_user_sales_digest_last_sent(
    telegram_id: int,
    *,
    now: datetime | None = None,
) -> datetime:
    """Get the shared multi-shop digest schedule without sending old events."""
    init_product_value_tables()
    current = now or _utc_now()
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT last_sent_at
            FROM sales_digest_user_state
            WHERE telegram_id = ?
            """,
            (int(telegram_id),),
        ).fetchone()
    parsed = _dt_from_db(row["last_sent_at"]) if row else None
    if parsed is not None:
        return parsed
    # On the first r13 cycle start the clock now. Pending rows stay queued and
    # will be included in the first normal hourly digest instead of flooding a
    # seller immediately after a deployment.
    reset_user_sales_digest_schedule(telegram_id, now=current)
    return current


def load_user_sales_digest_summaries(
    telegram_id: int,
    *,
    period_start: datetime,
    period_end: datetime,
) -> list[dict[str, Any]]:
    """Load pending hourly summaries grouped by shop for one Telegram user."""
    init_product_value_tables()
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT shop_id
            FROM sales_digest_queue
            WHERE telegram_id = ?
            ORDER BY shop_id
            """,
            (int(telegram_id),),
        ).fetchall()
    summaries = [
        load_sales_digest_summary(
            telegram_id,
            int(row["shop_id"]),
            period_start=period_start,
            period_end=period_end,
        )
        for row in rows
    ]
    return [summary for summary in summaries if int(summary.get("positions") or 0) > 0]


def build_multi_shop_sales_digest_message(
    summaries: list[dict[str, Any]],
    *,
    lang: str = "ru",
    watched_shop_count: int | None = None,
    shop_titles: dict[int, str] | None = None,
) -> str:
    """Build one phone-friendly hourly message for every monitored shop."""
    active = [summary for summary in summaries if int(summary.get("positions") or 0) > 0]
    if not active:
        return ""
    if len(active) == 1:
        return build_sales_digest_message(active[0], lang=lang)

    uz = normalize_lang(lang) == "uz"
    period_start = min(
        summary.get("period_start") or _utc_now()
        for summary in active
    )
    period_end = max(
        summary.get("period_end") or _utc_now()
        for summary in active
    )
    period = _sales_digest_period_text(period_start, period_end, lang)
    positions = sum(int(summary.get("positions") or 0) for summary in active)
    orders = sum(int(summary.get("orders") or 0) for summary in active)
    units = sum(float(summary.get("units") or 0) for summary in active)
    revenue = sum(float(summary.get("revenue") or 0) for summary in active)
    commission = sum(float(summary.get("commission") or 0) for summary in active)
    logistics = sum(float(summary.get("logistics") or 0) for summary in active)
    payout = sum(float(summary.get("payout") or 0) for summary in active)
    total_watched = max(len(active), int(watched_shop_count or 0))
    titles = shop_titles or {}

    if uz:
        lines = [
            "🕐 <b>Barcha do‘konlar bo‘yicha soatlik savdo</b>",
            f"🏪 Savdo bo‘lgan do‘konlar: <b>{len(active)}</b> / kuzatuvda: <b>{total_watched}</b>",
            f"🕒 Davr: <b>{period}</b>",
            "",
            f"🧾 Buyurtmalar: <b>{orders}</b> | Pozitsiyalar: <b>{positions}</b>",
            f"📦 Sotildi: <b>{units:g} dona</b>",
            f"💵 Jami tushum: <b>{_format_money(revenue)}</b>",
            f"🧮 O‘rtacha chek: <b>{_format_money(revenue / max(1, orders))}</b>",
            f"🏷 Komissiya: <b>{_format_money(commission)}</b>",
            f"🚚 Logistika: <b>{_format_money(logistics)}</b>",
            f"✅ Jami to‘lovga: <b>{_format_money(payout)}</b>",
            "",
            "🏪 <b>Do‘konlar bo‘yicha:</b>",
        ]
    else:
        lines = [
            "🕐 <b>Продажи за час по всем магазинам</b>",
            f"🏪 Магазинов с продажами: <b>{len(active)}</b> / отслеживается: <b>{total_watched}</b>",
            f"🕒 Период: <b>{period}</b>",
            "",
            f"🧾 Заказов: <b>{orders}</b> | Позиций: <b>{positions}</b>",
            f"📦 Продано: <b>{units:g} шт.</b>",
            f"💵 Общая выручка: <b>{_format_money(revenue)}</b>",
            f"🧮 Средний чек: <b>{_format_money(revenue / max(1, orders))}</b>",
            f"🏷 Комиссия: <b>{_format_money(commission)}</b>",
            f"🚚 Логистика: <b>{_format_money(logistics)}</b>",
            f"✅ Всего к выплате: <b>{_format_money(payout)}</b>",
            "",
            "🏪 <b>По магазинам:</b>",
        ]

    for summary in active[:20]:
        shop_id = int(summary.get("shop_id") or 0)
        title = _short_text(str(titles.get(shop_id) or "").strip(), 30)
        label = (
            f"<b>{escape(title)}</b> · <code>{shop_id}</code>"
            if title
            else f"<code>{shop_id}</code>"
        )
        if uz:
            lines.append(
                f"• {label} — <b>{int(summary.get('orders') or 0)}</b> buyurtma, "
                f"<b>{float(summary.get('units') or 0):g}</b> dona, "
                f"{_format_money(float(summary.get('revenue') or 0))}"
            )
        else:
            lines.append(
                f"• {label} — <b>{int(summary.get('orders') or 0)}</b> зак., "
                f"<b>{float(summary.get('units') or 0):g}</b> шт., "
                f"{_format_money(float(summary.get('revenue') or 0))}"
            )
    if len(active) > 20:
        remaining = len(active) - 20
        lines.append(
            f"Yana do‘konlar: <b>{remaining}</b>"
            if uz
            else f"Ещё магазинов: <b>{remaining}</b>"
        )
    return "\n".join(lines)


def mark_user_sales_digest_sent(
    telegram_id: int,
    *,
    sent_at: datetime | None = None,
    clear_queue: bool = True,
) -> None:
    """Atomically advance one user's digest and clear all included shops."""
    init_product_value_tables()
    current = sent_at or _utc_now()
    now_text = _dt_to_db(current) or ""
    shop_ids = connected_shop_ids_for_user(telegram_id)
    with db.connect() as conn:
        if clear_queue:
            conn.execute(
                "DELETE FROM sales_digest_queue WHERE telegram_id = ?",
                (int(telegram_id),),
            )
        conn.execute(
            """
            INSERT INTO sales_digest_user_state
                (telegram_id, last_sent_at, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(telegram_id) DO UPDATE SET
                last_sent_at = excluded.last_sent_at,
                updated_at = excluded.updated_at
            """,
            (int(telegram_id), now_text, now_text),
        )
        conn.executemany(
            """
            INSERT INTO sales_digest_state
                (telegram_id, shop_id, last_sent_at, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id) DO UPDATE SET
                last_sent_at = excluded.last_sent_at,
                updated_at = excluded.updated_at
            """,
            [
                (int(telegram_id), int(shop_id), now_text, now_text)
                for shop_id in shop_ids
            ],
        )
        conn.commit()


async def maybe_send_hourly_sales_digest_all_shops(
    telegram_id: int,
    *,
    now: datetime | None = None,
) -> bool:
    """Send at most one hourly sales message across all connected shops."""
    current = now or _utc_now()
    last_sent = get_user_sales_digest_last_sent(telegram_id, now=current)
    if current - last_sent < timedelta(seconds=SALES_DIGEST_INTERVAL_SECONDS):
        return False
    summaries = load_user_sales_digest_summaries(
        telegram_id,
        period_start=last_sent,
        period_end=current,
    )
    if not summaries:
        mark_user_sales_digest_sent(
            telegram_id,
            sent_at=current,
            clear_queue=False,
        )
        return False
    message = build_multi_shop_sales_digest_message(
        summaries,
        lang=get_user_language(telegram_id),
        watched_shop_count=len(connected_shop_ids_for_user(telegram_id)),
        shop_titles=connected_shop_titles_for_user(telegram_id),
    )
    try:
        await bot.send_message(
            telegram_id,
            message,
            reply_markup=main_menu_for_user(telegram_id),
        )
    except Exception:
        logging.exception(
            "Hourly multi-shop digest: delivery failed user=%s shops=%s events=%s",
            telegram_id,
            len(summaries),
            sum(int(summary.get("positions") or 0) for summary in summaries),
        )
        return False
    mark_user_sales_digest_sent(telegram_id, sent_at=current)
    logging.info(
        "Hourly multi-shop digest sent user=%s shops=%s orders=%s positions=%s revenue=%s",
        telegram_id,
        len(summaries),
        sum(int(summary.get("orders") or 0) for summary in summaries),
        sum(int(summary.get("positions") or 0) for summary in summaries),
        sum(float(summary.get("revenue") or 0) for summary in summaries),
    )
    return True


_sales_watch_initialized_scopes: set[tuple[int, int]] = set()
_sales_full_scan_at_by_group: dict[str, float] = {}


# Финальное переопределение. sales_watch_loop использует именно эту версию.
async def check_new_sales_once() -> None:
    now = _utc_now()
    date_to = int(now.timestamp() * 1000)

    for group in connected_watch_groups("notify_sales", "notify_cancellations"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(x) for x in group["telegram_ids"]]
        watch_key = _watch_group_key(group)
        last_full_scan = _sales_full_scan_at_by_group.get(watch_key)
        full_scan = (
            last_full_scan is None
            or time.monotonic() - last_full_scan >= SALES_WATCH_FULL_SCAN_INTERVAL_SECONDS
        )
        if full_scan:
            # Полная сверка нужна для отмен старых заказов, но раз в час достаточно.
            date_from = int((now - timedelta(days=7)).timestamp() * 1000)
            max_pages = 10
        else:
            # Частый опрос смотрит только свежее окно и создаёт заметно меньше страниц API.
            date_from = int(
                (now - timedelta(hours=SALES_WATCH_FAST_LOOKBACK_HOURS)).timestamp() * 1000
            )
            max_pages = 5

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows, _ = await _load_finance_orders(
                client,
                shop_id,
                date_from_ms=date_from,
                date_to_ms=date_to,
                max_pages=max_pages,
                page_size=100,
            )
            if full_scan:
                _sales_full_scan_at_by_group[watch_key] = time.monotonic()
                logging.info(
                    "Sales watcher full reconciliation completed shop=%s rows=%s next_full_in=%ss",
                    shop_id,
                    len(rows),
                    SALES_WATCH_FULL_SCAN_INTERVAL_SECONDS,
                )
        except Exception as error:
            _log_watcher_api_failure(
                "Sales/cancel watcher",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        identity_status_now = {
            finance_identity_key(item): _finance_status(item)
            for item in rows
        }

        for telegram_id in telegram_ids:
            sales_mode = get_sales_notification_mode(telegram_id)
            cancellations_enabled = product_setting_enabled(telegram_id, "notify_cancellations")
            status_memory = _sale_status_by_user.setdefault(telegram_id, {})

            # Загружаем статусы из SQLite. Они не пропадают после SIGTERM.
            saved_statuses = load_saved_sale_statuses(telegram_id, shop_id)
            if saved_statuses:
                status_memory.update(saved_statuses)

            # Первый снимок фиксируется отдельно, поэтому даже магазин без единой продажи
            # не потеряет свою первую будущую продажу после перезапуска бота.
            sales_baseline_ready = operational_watcher_initialized(
                telegram_id,
                shop_id,
                "sales",
            )
            first_ever_snapshot = not saved_statuses and not sales_baseline_ready

            cancel_rows: list[dict[str, Any]] = []
            if cancellations_enabled and not first_ever_snapshot:
                for item in rows:
                    ident = finance_identity_key(item)
                    current_status = _finance_status(item)
                    previous_status = status_memory.get(ident)
                    if (
                        _is_cancelled_status(current_status)
                        and (
                            previous_status is None
                            or not _is_cancelled_status(str(previous_status))
                        )
                    ):
                        cancel_rows.append(item)

            if first_ever_snapshot:
                new_rows: list[dict[str, Any]] = []
            else:
                new_rows = [
                    item
                    for item in rows
                    if finance_identity_key(item) not in saved_statuses
                    and not _is_cancelled_status(_finance_status(item))
                    and not _is_returned_status(_finance_status(item))
                ]

            scope = (telegram_id, shop_id)
            if scope not in _sales_watch_initialized_scopes:
                _sales_watch_initialized_scopes.add(scope)
                _sales_watch_initialized.add(telegram_id)
                logging.info(
                    "Sales watcher initialized for user=%s shop=%s sales_rows=%s "
                    "saved_statuses=%s mode=%s",
                    telegram_id,
                    shop_id,
                    len(rows),
                    len(saved_statuses),
                    sales_mode,
                )

            no_longer_sales = [
                finance_identity_key(item)
                for item in rows
                if _is_cancelled_status(_finance_status(item))
                or _is_returned_status(_finance_status(item))
            ]
            if no_longer_sales:
                discard_sales_digest_events(
                    telegram_id,
                    shop_id,
                    no_longer_sales,
                )

            if sales_mode == "hourly" and new_rows:
                inserted = enqueue_sales_digest_events(
                    telegram_id,
                    shop_id,
                    new_rows,
                    detected_at=now,
                )
                logging.info(
                    "Hourly sales digest queued user=%s shop=%s detected=%s inserted=%s",
                    telegram_id,
                    shop_id,
                    len(new_rows),
                    inserted,
                )

            status_memory.update(identity_status_now)
            save_sale_statuses(telegram_id, shop_id, identity_status_now)
            if not sales_baseline_ready:
                mark_operational_watcher_initialized(
                    telegram_id,
                    shop_id,
                    "sales",
                )

            if sales_mode == "instant":
                for item in new_rows[:INSTANT_SALE_BURST_LIMIT]:
                    try:
                        await bot.send_message(
                            telegram_id,
                            build_new_sale_message(
                                item,
                                shop_id=shop_id,
                                lang=get_user_language(telegram_id),
                            ),
                            reply_markup=main_menu_for_user(telegram_id),
                        )
                        await asyncio.sleep(0.15)
                    except Exception:
                        logging.exception(
                            "Sales watcher: failed to send sale notification to %s",
                            telegram_id,
                        )

                overflow = new_rows[INSTANT_SALE_BURST_LIMIT:]
                if overflow:
                    summary = summarize_sales_digest_items(
                        overflow,
                        shop_id=shop_id,
                        period_start=now - timedelta(seconds=max(60, SALE_CHECK_INTERVAL_SECONDS)),
                        period_end=now,
                    )
                    try:
                        await bot.send_message(
                            telegram_id,
                            build_sales_digest_message(
                                summary,
                                lang=get_user_language(telegram_id),
                                burst=True,
                            ),
                            reply_markup=main_menu_for_user(telegram_id),
                        )
                    except Exception:
                        logging.exception(
                            "Sales watcher: failed to send burst summary to %s",
                            telegram_id,
                        )

            for item in cancel_rows[:10]:
                try:
                    await bot.send_message(
                        telegram_id,
                        build_cancel_message(
                            item,
                            shop_id=shop_id,
                            lang=get_user_language(telegram_id),
                        ),
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                    logging.info(
                        "Cancel notification sent user=%s shop=%s order=%s status=%s",
                        telegram_id,
                        shop_id,
                        _finance_order_id(item),
                        _finance_status(item),
                    )
                    await asyncio.sleep(0.15)
                except Exception:
                    logging.exception(
                        "Sales watcher: failed to send cancel notification to %s",
                        telegram_id,
                    )

            if len(cancel_rows) > 10:
                lang = get_user_language(telegram_id)
                extra = len(cancel_rows) - 10
                text = (
                    f"➕ Yana bekor qilingan buyurtmalar: <b>{extra}</b>\n"
                    "Batafsil: <code>/balance</code>"
                    if normalize_lang(lang) == "uz"
                    else f"➕ Ещё отмен за эту проверку: <b>{extra}</b>\n"
                    "Подробно: <code>/balance</code>"
                )
                try:
                    await bot.send_message(
                        telegram_id,
                        text,
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                except Exception:
                    logging.exception(
                        "Sales watcher: failed to send cancel overflow to %s",
                        telegram_id,
                    )

            if sales_mode == "hourly":
                await maybe_send_hourly_sales_digest(
                    telegram_id,
                    shop_id,
                    now=now,
                )

        await asyncio.sleep(0.5)


# --- Операционные уведомления: потери/брак и приёмка FBO ---
def _loss_watch_snapshot(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    snapshot: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = _loss_row_key(row)
        snapshot[key] = {
            "sku_key": key,
            "product_title": str(
                row.get("product_title")
                or row.get("sku_full_title")
                or row.get("sku_title")
                or "Без названия"
            ),
            "sku_title": str(row.get("sku_full_title") or row.get("sku_title") or ""),
            "sku_id": str(row.get("sku_id") or row.get("seller_item_code") or ""),
            "barcode": str(row.get("barcode") or ""),
            "missing_qty": _loss_qty(row, "missing"),
            "defected_qty": _loss_qty(row, "defected"),
            "price": max(0.0, float(_num_from_value(row.get("price")) or 0)),
        }
    return snapshot


def calculate_loss_defect_changes(
    previous: dict[str, dict[str, Any]],
    current: dict[str, dict[str, Any]],
    *,
    notify_losses: bool,
    notify_defects: bool,
) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    for key, item in current.items():
        old = previous.get(key) or {}
        missing_delta = max(
            0,
            int(item.get("missing_qty") or 0) - int(old.get("missing_qty") or 0),
        )
        defected_delta = max(
            0,
            int(item.get("defected_qty") or 0) - int(old.get("defected_qty") or 0),
        )
        visible_missing = missing_delta if notify_losses else 0
        visible_defected = defected_delta if notify_defects else 0
        if visible_missing <= 0 and visible_defected <= 0:
            continue
        price = max(0.0, float(item.get("price") or 0))
        changes.append({
            **item,
            "sku_key": key,
            "missing_delta": visible_missing,
            "defected_delta": visible_defected,
            "estimated_value": price * (visible_missing + visible_defected),
        })
    changes.sort(
        key=lambda item: (
            -float(item.get("estimated_value") or 0),
            -(int(item.get("missing_delta") or 0) + int(item.get("defected_delta") or 0)),
            str(item.get("product_title") or ""),
        )
    )
    return changes


def build_loss_defect_notification(
    shop_id: int,
    changes: list[dict[str, Any]],
    *,
    lang: str = "ru",
) -> str:
    shown = changes[:15]
    total_missing = sum(int(item.get("missing_delta") or 0) for item in changes)
    total_defected = sum(int(item.get("defected_delta") or 0) for item in changes)
    total_value = sum(float(item.get("estimated_value") or 0) for item in changes)
    if normalize_lang(lang) == "uz":
        lines = [
            "🚨 <b>FBO omborida yangi yo‘qotish yoki yaroqsiz tovar</b>",
            f"🏪 Do‘kon: <code>{shop_id}</code>",
            f"🧭 Yangi yo‘qotish: <b>+{total_missing} dona</b>",
            f"🧪 Yangi yaroqsiz: <b>+{total_defected} dona</b>",
        ]
        if total_value > 0:
            lines.append(f"💰 Sotuv narxi bo‘yicha baho: <b>{_format_money(total_value)}</b>")
        for index, item in enumerate(shown, start=1):
            sku = item.get("sku_id") or item.get("barcode") or item.get("sku_title") or "—"
            lines.extend([
                "",
                f"{index}. <b>{escape(_short_text(item.get('product_title'), 70))}</b>",
                f"SKU: <code>{escape(_short_text(sku, 80))}</code>",
            ])
            if int(item.get("missing_delta") or 0) > 0:
                lines.append(
                    f"🧭 Yo‘qotildi: <b>+{int(item['missing_delta'])}</b> "
                    f"(jami {int(item.get('missing_qty') or 0)})"
                )
            if int(item.get("defected_delta") or 0) > 0:
                lines.append(
                    f"🧪 Yaroqsiz: <b>+{int(item['defected_delta'])}</b> "
                    f"(jami {int(item.get('defected_qty') or 0)})"
                )
        if len(changes) > len(shown):
            lines.extend(["", f"Yana o‘zgarishlar: <b>{len(changes) - len(shown)}</b>"])
        lines.extend(["", "Barcha yig‘ma yo‘qotishlar: <code>/lost</code>"])
        return "\n".join(lines)

    lines = [
        "🚨 <b>Новые потери или брак на складе FBO</b>",
        f"🏪 Магазин: <code>{shop_id}</code>",
        f"🧭 Новые потери: <b>+{total_missing} шт.</b>",
        f"🧪 Новый брак: <b>+{total_defected} шт.</b>",
    ]
    if total_value > 0:
        lines.append(f"💰 Оценка по цене продажи: <b>{_format_money(total_value)}</b>")
    for index, item in enumerate(shown, start=1):
        sku = item.get("sku_id") or item.get("barcode") or item.get("sku_title") or "—"
        lines.extend([
            "",
            f"{index}. <b>{escape(_short_text(item.get('product_title'), 70))}</b>",
            f"SKU: <code>{escape(_short_text(sku, 80))}</code>",
        ])
        if int(item.get("missing_delta") or 0) > 0:
            lines.append(
                f"🧭 Потеряно: <b>+{int(item['missing_delta'])}</b> "
                f"(всего {int(item.get('missing_qty') or 0)})"
            )
        if int(item.get("defected_delta") or 0) > 0:
            lines.append(
                f"🧪 Брак: <b>+{int(item['defected_delta'])}</b> "
                f"(всего {int(item.get('defected_qty') or 0)})"
            )
    if len(changes) > len(shown):
        lines.extend(["", f"Ещё изменений: <b>{len(changes) - len(shown)}</b>"])
    lines.extend(["", "Все накопительные потери: <code>/lost</code>"])
    return "\n".join(lines)


def _fbo_first_value(item: dict[str, Any], paths: tuple[str, ...]) -> tuple[Any, bool]:
    for path in paths:
        value = _value_by_path(item, path)
        if value not in (None, ""):
            return value, True
    return None, False


def _fbo_invoice_watch_key(invoice: dict[str, Any]) -> str:
    invoice_id = _invoice_id(invoice)
    if invoice_id not in (None, ""):
        return f"id:{invoice_id}"
    raw = "|".join(
        str(value or "")
        for value in (
            _invoice_number(invoice),
            _value_by_path(invoice, "dateCreated", "createdAt", "creationDate"),
            _value_by_path(invoice, "totalToStock", "quantityToStock", "totalQuantity"),
        )
    )
    return "hash:" + hashlib.sha1(raw.encode("utf-8")).hexdigest()[:20]


def _fbo_invoice_quantities(invoice: dict[str, Any]) -> tuple[float, float, bool]:
    planned_value, _ = _fbo_first_value(
        invoice,
        ("totalToStock", "quantityToStock", "totalQuantity", "plannedQuantity"),
    )
    accepted_value, accepted_known = _fbo_first_value(
        invoice,
        ("totalAccepted", "quantityAccepted", "acceptedQuantity"),
    )
    return (
        max(0.0, _num_any(planned_value)),
        max(0.0, _num_any(accepted_value)),
        accepted_known,
    )


def fbo_invoice_acceptance_is_terminal(invoice: dict[str, Any]) -> bool:
    accepted_date = _value_by_path(invoice, "dateAccepted", "acceptedAt", "acceptanceDate")
    if accepted_date not in (None, ""):
        return True
    status = _invoice_status(invoice).strip().upper()
    if not status or status == "—":
        return False
    if any(token in status for token in ("CANCEL", "ОТМЕН", "BEKOR")):
        return False
    return any(
        token in status
        for token in (
            "ACCEPTED",
            "PARTIALLY_ACCEPTED",
            "COMPLETED",
            "FINISHED",
            "CLOSED",
            "RECEIVED",
            "ПРИНЯТ",
            "ЗАВЕРШ",
            "ПОЛУЧЕН",
            "QABUL",
            "YAKUN",
        )
    )


def _fbo_acceptance_items(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for product in products:
        nested = _value_by_path(product, "skuForInvoiceDtoList", "skuList", "skus")
        sources: list[tuple[dict[str, Any], dict[str, Any]]] = []
        if isinstance(nested, list) and nested:
            for sku in nested:
                if isinstance(sku, dict):
                    merged = dict(product)
                    merged.update(sku)
                    qty_source = merged if len(nested) == 1 else sku
                    sources.append((merged, qty_source))
        else:
            sources.append((product, product))

        for source, qty_source in sources:
            planned_value, planned_known = _fbo_first_value(
                qty_source,
                ("quantityToStock", "toStock", "quantity", "plannedQuantity", "totalToStock"),
            )
            accepted_value, accepted_known = _fbo_first_value(
                qty_source,
                ("quantityAccepted", "accepted", "acceptedQuantity", "totalAccepted"),
            )
            defected_value, _ = _fbo_first_value(
                qty_source,
                (
                    "quantityDefected",
                    "defectedQuantity",
                    "defectiveQuantity",
                    "damagedQuantity",
                ),
            )
            rejected_value, _ = _fbo_first_value(
                qty_source,
                (
                    "quantityNotAccepted",
                    "notAcceptedQuantity",
                    "quantityRejected",
                    "rejectedQuantity",
                    "rejectQuantity",
                ),
            )
            missing_value, _ = _fbo_first_value(
                qty_source,
                ("quantityMissing", "missingQuantity", "shortageQuantity"),
            )
            planned = max(0.0, _num_any(planned_value))
            accepted = max(0.0, _num_any(accepted_value))
            defected = max(0.0, _num_any(defected_value))
            explicit_rejected = max(
                max(0.0, _num_any(rejected_value)),
                max(0.0, _num_any(missing_value)),
                defected,
            )
            if not accepted_known and planned_known and explicit_rejected > 0:
                accepted = max(0.0, planned - explicit_rejected)
                accepted_known = True
            difference = max(0.0, planned - accepted) if accepted_known else 0.0
            not_accepted = max(difference, explicit_rejected)
            title = str(
                _value_by_path(source, "productTitle", "title", "product.name", "name")
                or "Без названия"
            )
            sku = str(
                _value_by_path(
                    source,
                    "skuId",
                    "sku.id",
                    "skuTitle",
                    "skuName",
                    "sellerItemCode",
                    "barcode",
                )
                or "—"
            )
            barcode = str(_value_by_path(source, "barcode", "sku.barcode") or "")
            reason_value = _value_by_path(
                source,
                "defectReason",
                "rejectionReason",
                "rejectReason",
                "reason",
                "comment",
                "status",
            )
            reason = _status_text_any(reason_value) if isinstance(reason_value, dict) else str(reason_value or "")
            if planned <= 0 and accepted <= 0 and not_accepted <= 0:
                continue
            rows.append({
                "title": title,
                "sku": sku,
                "barcode": barcode,
                "planned_qty": planned,
                "accepted_qty": accepted,
                "accepted_known": accepted_known,
                "not_accepted_qty": not_accepted,
                "explicit_problem_qty": explicit_rejected,
                "defected_qty": defected,
                "reason": reason,
            })
    rows.sort(
        key=lambda item: (
            -float(item.get("not_accepted_qty") or 0),
            str(item.get("title") or ""),
        )
    )
    return rows


def summarize_fbo_acceptance(
    invoice: dict[str, Any],
    products: list[dict[str, Any]],
    *,
    shop_id: int,
) -> dict[str, Any]:
    items = _fbo_acceptance_items(products)
    invoice_planned, invoice_accepted, invoice_accepted_known = _fbo_invoice_quantities(invoice)
    product_planned = sum(float(item.get("planned_qty") or 0) for item in items)
    known_items = [item for item in items if item.get("accepted_known")]
    product_accepted = sum(float(item.get("accepted_qty") or 0) for item in known_items)
    product_not_accepted = sum(float(item.get("not_accepted_qty") or 0) for item in items)
    product_explicit_problem = sum(float(item.get("explicit_problem_qty") or 0) for item in items)
    product_defected = sum(float(item.get("defected_qty") or 0) for item in items)

    invoice_problem_value, _ = _fbo_first_value(
        invoice,
        (
            "totalNotAccepted",
            "quantityNotAccepted",
            "totalRejected",
            "rejectedQuantity",
            "totalDefected",
            "quantityDefected",
        ),
    )
    invoice_problem = max(0.0, _num_any(invoice_problem_value))
    # Invoice-list totals and invoice-product details are updated by Uzum at
    # different moments.  Accepted quantity is monotonic, therefore the most
    # complete observed value must win instead of letting a stale list-level 0
    # overwrite `quantityAccepted=100` from product details.
    planned = max(invoice_planned, product_planned)
    accepted_known = invoice_accepted_known or bool(known_items)
    accepted = max(
        invoice_accepted if invoice_accepted_known else 0.0,
        product_accepted if known_items else 0.0,
    )
    if planned > 0:
        accepted = min(planned, accepted)
    difference = max(0.0, planned - accepted) if accepted_known else 0.0
    # A stale per-SKU zero is not explicit evidence of a rejected/defective
    # item.  Only documented problem counters may override a newer accepted
    # total.  The aggregate difference is still reported when it is non-zero.
    not_accepted = max(difference, product_explicit_problem, invoice_problem)
    zero_without_evidence = bool(
        planned > 0
        and accepted_known
        and accepted <= 0
        and invoice_problem <= 0
        and product_explicit_problem <= 0
        and not any(str(item.get("reason") or "").strip() for item in items)
    )
    problem_items = [
        item
        for item in items
        if float(item.get("explicit_problem_qty") or 0) > 0
        or (
            not zero_without_evidence
            and float(item.get("not_accepted_qty") or 0) > 0
        )
    ]
    details_complete = not_accepted <= 0 or bool(problem_items)
    if not_accepted > 0 and not problem_items:
        problem_items = [{
            "title": "Итого по накладной",
            "sku": "—",
            "barcode": "",
            "planned_qty": planned,
            "accepted_qty": accepted,
            "accepted_known": accepted_known,
            "not_accepted_qty": not_accepted,
            "defected_qty": invoice_problem,
            "reason": "Uzum API не передал разбивку расхождения по SKU",
        }]

    outcome = (
        "unknown"
        if zero_without_evidence
        else "discrepancy"
        if not_accepted > 0
        else "success"
        if planned > 0 and accepted_known
        else "unknown"
    )
    acceptance_rate = (accepted / planned * 100.0) if planned > 0 else 0.0
    return {
        "shop_id": int(shop_id),
        "invoice_id": str(_invoice_id(invoice) or ""),
        "invoice_number": _invoice_number(invoice),
        "status": _invoice_status(invoice),
        "accepted_date": _date_text_any(
            _value_by_path(invoice, "dateAccepted", "acceptedAt", "acceptanceDate")
        ),
        "planned_qty": planned,
        "accepted_qty": accepted,
        "not_accepted_qty": not_accepted,
        "defected_qty": product_defected if product_defected > 0 else invoice_problem,
        "positions": len(items),
        "items": items,
        "problem_items": problem_items,
        "details_complete": details_complete,
        "zero_without_evidence": zero_without_evidence,
        "product_reported_not_accepted_qty": product_not_accepted,
        "acceptance_rate": acceptance_rate,
        "outcome": outcome,
    }


def build_fbo_acceptance_notification(summary: dict[str, Any], *, lang: str = "ru") -> str:
    uz = normalize_lang(lang) == "uz"
    outcome = str(summary.get("outcome") or "unknown")
    shop_id = int(summary.get("shop_id") or 0)
    number = escape(str(summary.get("invoice_number") or summary.get("invoice_id") or "—"))
    planned = _fmt_qty(summary.get("planned_qty"))
    accepted = _fmt_qty(summary.get("accepted_qty"))
    rejected = _fmt_qty(summary.get("not_accepted_qty"))
    positions = int(summary.get("positions") or 0)
    accepted_date = escape(str(summary.get("accepted_date") or "—"))
    rate = float(summary.get("acceptance_rate") or 0)
    positions_uz = (
        f"📦 SKU pozitsiyalari: <b>{positions}</b>\n"
        if positions > 0
        else "📦 Tarkib: <b>Uzum API tafsilot bermadi</b>\n"
    )
    positions_ru = (
        f"📦 SKU-позиций: <b>{positions}</b>\n"
        if positions > 0
        else "📦 Состав: <b>Uzum API не передал детализацию</b>\n"
    )

    if uz and outcome == "success":
        return (
            "✅ <b>FBO yetkazib berish to‘liq qabul qilindi</b>\n\n"
            f"🏪 Do‘kon: <code>{shop_id}</code>\n"
            f"📄 Yuk xati: <b>№{number}</b>\n"
            f"{positions_uz}"
            f"🚚 Yetkazildi: <b>{planned} dona</b>\n"
            f"✅ Qabul qilindi: <b>{accepted} dona</b>\n"
            f"📊 Qabul darajasi: <b>{rate:.1f}%</b>\n"
            f"🕒 Yakunlandi: {accepted_date}\n\n"
            "Barcha tovarlar farq va yaroqsiz holatsiz qabul qilindi."
        )
    if uz and outcome == "discrepancy":
        details = "" if summary.get("details_complete") else "\n⚠️ Uzum SKU bo‘yicha to‘liq tafsilotni bermadi."
        return (
            "⚠️ <b>FBO qabuli farq bilan yakunlandi</b>\n\n"
            f"🏪 Do‘kon: <code>{shop_id}</code>\n"
            f"📄 Yuk xati: <b>№{number}</b>\n"
            f"🚚 Yetkazildi: <b>{planned} dona</b>\n"
            f"✅ Qabul qilindi: <b>{accepted} dona</b>\n"
            f"🧪 Qabul qilinmadi / yaroqsiz: <b>{rejected} dona</b>\n"
            f"📊 Qabul darajasi: <b>{rate:.1f}%</b>\n\n"
            "📎 Muammoli tovarlar PDF faylda ko‘rsatilgan."
            + details
        )
    if uz:
        return (
            "ℹ️ <b>FBO qabuli yakunlandi</b>\n\n"
            f"🏪 Do‘kon: <code>{shop_id}</code>\n"
            f"📄 Yuk xati: <b>№{number}</b>\n"
            f"📌 Status: <b>{escape(str(summary.get('status') or '—'))}</b>\n\n"
            "Uzum API yakuniy miqdorlarni bermadi; tafsilotlar keyingi tekshiruvda yangilanadi."
        )

    if outcome == "success":
        return (
            "✅ <b>Поставка FBO принята полностью</b>\n\n"
            f"🏪 Магазин: <code>{shop_id}</code>\n"
            f"📄 Накладная: <b>№{number}</b>\n"
            f"{positions_ru}"
            f"🚚 Передано: <b>{planned} шт.</b>\n"
            f"✅ Принято: <b>{accepted} шт.</b>\n"
            f"📊 Приёмка: <b>{rate:.1f}%</b>\n"
            f"🕒 Завершено: {accepted_date}\n\n"
            "Все товары приняты без расхождений и брака."
        )
    if outcome == "discrepancy":
        details = "" if summary.get("details_complete") else "\n⚠️ Uzum не передал полную разбивку расхождения по SKU."
        return (
            "⚠️ <b>Приёмка FBO завершена с расхождениями</b>\n\n"
            f"🏪 Магазин: <code>{shop_id}</code>\n"
            f"📄 Накладная: <b>№{number}</b>\n"
            f"🚚 Передано: <b>{planned} шт.</b>\n"
            f"✅ Принято: <b>{accepted} шт.</b>\n"
            f"🧪 Не принято / брак: <b>{rejected} шт.</b>\n"
            f"📊 Приёмка: <b>{rate:.1f}%</b>\n\n"
            "📎 Список проблемных товаров приложен в PDF."
            + details
        )
    return (
        "ℹ️ <b>Приёмка FBO завершена</b>\n\n"
        f"🏪 Магазин: <code>{shop_id}</code>\n"
        f"📄 Накладная: <b>№{number}</b>\n"
        f"📌 Статус: <b>{escape(str(summary.get('status') or '—'))}</b>\n\n"
        "Uzum API не передал итоговые количества; данные обновятся при следующей проверке."
    )


def _fbo_pdf_font_paths() -> tuple[Path, Path]:
    configured = str(os.getenv("PDF_FONT_PATH", "") or "").strip()
    app_dir = Path(__file__).resolve().parent
    regular_candidates = [
        Path(configured) if configured else None,
        app_dir / "DejaVuSans.ttf",
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
    ]
    regular = next((path for path in regular_candidates if path and path.exists()), None)
    if regular is None:
        raise RuntimeError("Не найден Unicode-шрифт для PDF. Укажите PDF_FONT_PATH.")
    bold_candidates = [
        regular.with_name(regular.name.replace(".ttf", "-Bold.ttf")),
        app_dir / "DejaVuSans-Bold.ttf",
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
    ]
    bold = next((path for path in bold_candidates if path.exists()), regular)
    return regular, bold


def build_fbo_acceptance_pdf(
    summary: dict[str, Any],
    output: str | Path | None = None,
    *,
    lang: str = "ru",
) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    regular_path, bold_path = _fbo_pdf_font_paths()
    if "UzumReport" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("UzumReport", str(regular_path)))
    if "UzumReportBold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("UzumReportBold", str(bold_path)))

    if output is None:
        with tempfile.NamedTemporaryFile(prefix="fbo_acceptance_", suffix=".pdf", delete=False) as tmp:
            output_path = Path(tmp.name)
    else:
        output_path = Path(output)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    uz = normalize_lang(lang) == "uz"
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "FboTitle",
        parent=styles["Title"],
        fontName="UzumReportBold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#17365D"),
        alignment=TA_LEFT,
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "FboSubtitle",
        parent=styles["BodyText"],
        fontName="UzumReport",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#52657A"),
    )
    cell_style = ParagraphStyle(
        "FboCell",
        parent=styles["BodyText"],
        fontName="UzumReport",
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )
    cell_center = ParagraphStyle(
        "FboCellCenter",
        parent=cell_style,
        alignment=TA_CENTER,
    )
    header_style = ParagraphStyle(
        "FboHeader",
        parent=cell_center,
        fontName="UzumReportBold",
        textColor=colors.white,
    )

    invoice_number = escape(str(summary.get("invoice_number") or summary.get("invoice_id") or "—"))
    title = "FBO qabulidagi farqlar" if uz else "Расхождения при приёмке FBO"
    subtitle = (
        f"Do‘kon {int(summary.get('shop_id') or 0)} | Yuk xati №{invoice_number} | "
        f"Yaratildi {datetime.now(UZT).strftime('%d.%m.%Y %H:%M')}"
        if uz
        else f"Магазин {int(summary.get('shop_id') or 0)} | Накладная №{invoice_number} | "
        f"Сформировано {datetime.now(UZT).strftime('%d.%m.%Y %H:%M')}"
    )
    story: list[Any] = [Paragraph(title, title_style), Paragraph(subtitle, subtitle_style), Spacer(1, 5 * mm)]

    summary_data = [
        [
            Paragraph("Yetkazildi" if uz else "Передано", header_style),
            Paragraph("Qabul qilindi" if uz else "Принято", header_style),
            Paragraph("Qabul qilinmadi" if uz else "Не принято", header_style),
            Paragraph("Qabul foizi" if uz else "Процент приёмки", header_style),
        ],
        [
            Paragraph(_fmt_qty(summary.get("planned_qty")), cell_center),
            Paragraph(_fmt_qty(summary.get("accepted_qty")), cell_center),
            Paragraph(_fmt_qty(summary.get("not_accepted_qty")), cell_center),
            Paragraph(f"{float(summary.get('acceptance_rate') or 0):.1f}%", cell_center),
        ],
    ]
    summary_table = Table(summary_data, colWidths=[46 * mm] * 4, repeatRows=1)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EEF4FA")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C9DD")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B7C9DD")),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.extend([summary_table, Spacer(1, 6 * mm)])

    headers = (
        ("№", "Tovar", "SKU / shtrix", "Yetkazildi", "Qabul", "Qabul qilinmadi", "Yaroqsiz", "Sabab / status")
        if uz
        else ("№", "Товар", "SKU / штрихкод", "Передано", "Принято", "Не принято", "Брак", "Причина / статус")
    )
    table_data: list[list[Any]] = [[Paragraph(value, header_style) for value in headers]]
    for index, item in enumerate(list(summary.get("problem_items") or []), start=1):
        sku_text = str(item.get("sku") or "—")
        if item.get("barcode") and str(item.get("barcode")) not in sku_text:
            sku_text += f"\n{item['barcode']}"
        reason = str(item.get("reason") or ("API sababni ko‘rsatmadi" if uz else "Причина не указана API"))
        table_data.append([
            Paragraph(str(index), cell_center),
            Paragraph(escape(str(item.get("title") or "—")), cell_style),
            Paragraph(escape(sku_text).replace("\n", "<br/>"), cell_style),
            Paragraph(_fmt_qty(item.get("planned_qty")), cell_center),
            Paragraph(_fmt_qty(item.get("accepted_qty")), cell_center),
            Paragraph(_fmt_qty(item.get("not_accepted_qty")), cell_center),
            Paragraph(_fmt_qty(item.get("defected_qty")), cell_center),
            Paragraph(escape(reason), cell_style),
        ])
    detail_table = Table(
        table_data,
        colWidths=[10 * mm, 62 * mm, 34 * mm, 21 * mm, 21 * mm, 25 * mm, 19 * mm, 55 * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    detail_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7FA")]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#AABBCD")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CAD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(detail_table)
    if not summary.get("details_complete"):
        note = (
            "Eslatma: Uzum API farqni SKU bo‘yicha to‘liq bermadi; umumiy ko‘rsatkich ko‘rsatildi."
            if uz
            else "Примечание: Uzum API не передал полную разбивку расхождения по SKU; показан общий итог."
        )
        story.extend([Spacer(1, 4 * mm), Paragraph(note, subtitle_style)])

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=13 * mm,
        title=title,
        author="Uzum Seller Assistant",
    )

    def footer(canvas: Any, document: Any) -> None:
        canvas.saveState()
        canvas.setFont("UzumReport", 8)
        canvas.setFillColor(colors.HexColor("#6B7C8F"))
        page_text = f"Sahifa {document.page}" if uz else f"Страница {document.page}"
        canvas.drawRightString(landscape(A4)[0] - 10 * mm, 6 * mm, page_text)
        canvas.drawString(10 * mm, 6 * mm, "Uzum Seller Assistant")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return output_path


async def check_loss_defect_once() -> None:
    for group in connected_watch_groups("notify_losses", "notify_defects"):
        shop_id = int(group["shop_id"])
        telegram_ids = [int(value) for value in group["telegram_ids"]]
        try:
            token = cipher.decrypt(group["uzum_token_encrypted"])
            client = UzumClient(token, UZUM_API_BASE_URL)
            rows, unavailable_filters = await _load_all_time_loss_rows(client, shop_id)
            current = _loss_watch_snapshot(rows)
        except Exception as error:
            _log_watcher_api_failure(
                "Loss/defect watcher",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        for telegram_id in telegram_ids:
            initialized = operational_watcher_initialized(
                telegram_id,
                shop_id,
                "loss_defect",
            )
            previous = load_product_loss_snapshot(telegram_id, shop_id)
            if not initialized:
                save_product_loss_snapshot(
                    telegram_id,
                    shop_id,
                    current,
                    reset_absent=not unavailable_filters,
                )
                mark_operational_watcher_initialized(
                    telegram_id,
                    shop_id,
                    "loss_defect",
                )
                logging.info(
                    "Loss/defect watcher initialized user=%s shop=%s skus=%s",
                    telegram_id,
                    shop_id,
                    len(current),
                )
                continue

            all_changes = calculate_loss_defect_changes(
                previous,
                current,
                notify_losses=True,
                notify_defects=True,
            )
            changes = calculate_loss_defect_changes(
                previous,
                current,
                notify_losses=product_setting_enabled(telegram_id, "notify_losses"),
                notify_defects=product_setting_enabled(telegram_id, "notify_defects"),
            )
            delivered = not changes
            if changes:
                try:
                    await bot.send_message(
                        telegram_id,
                        build_loss_defect_notification(
                            shop_id,
                            changes,
                            lang=get_user_language(telegram_id),
                        ),
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                    delivered = True
                    logging.info(
                        "Loss/defect notification sent user=%s shop=%s changes=%s",
                        telegram_id,
                        shop_id,
                        len(changes),
                    )
                    await asyncio.sleep(0.15)
                except Exception:
                    logging.exception(
                        "Loss/defect watcher: delivery failed user=%s shop=%s",
                        telegram_id,
                        shop_id,
                    )
            if delivered:
                if all_changes:
                    try:
                        record_loss_defect_events(
                            telegram_id,
                            shop_id,
                            all_changes,
                        )
                    except Exception:
                        logging.exception(
                            "Loss/defect watcher: event history failed user=%s shop=%s",
                            telegram_id,
                            shop_id,
                        )
                save_product_loss_snapshot(
                    telegram_id,
                    shop_id,
                    current,
                    reset_absent=not unavailable_filters,
                )
        await asyncio.sleep(0.5)


async def loss_defect_watch_loop() -> None:
    await asyncio.sleep(85)
    logging.info(
        "Loss/defect watcher started. Interval: %s seconds",
        LOSS_DEFECT_CHECK_INTERVAL_SECONDS,
    )
    while True:
        try:
            await check_loss_defect_once()
        except Exception:
            logging.exception("Loss/defect watcher loop error")
        await asyncio.sleep(max(300, LOSS_DEFECT_CHECK_INTERVAL_SECONDS))


_fbo_full_scan_at_by_group: dict[str, float] = {}


async def check_fbo_acceptance_once() -> None:
    for group in connected_watch_groups("notify_fbo_acceptance"):
        shop_id = int(group["shop_id"])
        telegram_ids = [int(value) for value in group["telegram_ids"]]
        watch_key = _watch_group_key(group)
        baseline_required = any(
            not operational_watcher_initialized(user_id, shop_id, "fbo_acceptance")
            for user_id in telegram_ids
        )
        last_full_scan = _fbo_full_scan_at_by_group.get(watch_key)
        full_scan = (
            baseline_required
            or last_full_scan is None
            or time.monotonic() - last_full_scan >= FBO_ACCEPTANCE_FULL_SCAN_INTERVAL_SECONDS
        )
        invoice_pages = (
            FBO_ACCEPTANCE_INVOICE_PAGES
            if full_scan
            else FBO_ACCEPTANCE_FAST_PAGES
        )
        try:
            token = cipher.decrypt(group["uzum_token_encrypted"])
            client = UzumClient(token, UZUM_API_BASE_URL)
            invoices, _ = await _load_fbo_invoices(
                client,
                shop_id,
                max_pages=invoice_pages,
                page_size=20,
            )
            if full_scan:
                _fbo_full_scan_at_by_group[watch_key] = time.monotonic()
                logging.info(
                    "FBO acceptance full reconciliation completed shop=%s invoices=%s "
                    "next_full_in=%ss",
                    shop_id,
                    len(invoices),
                    FBO_ACCEPTANCE_FULL_SCAN_INTERVAL_SECONDS,
                )
        except Exception as error:
            _log_watcher_api_failure(
                "FBO acceptance watcher invoice list",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        products_cache: dict[str, list[dict[str, Any]]] = {}
        products_error: set[str] = set()
        products_retry_later: set[str] = set()
        for telegram_id in telegram_ids:
            initialized = operational_watcher_initialized(
                telegram_id,
                shop_id,
                "fbo_acceptance",
            )
            if not initialized:
                for invoice in invoices:
                    invoice_key = _fbo_invoice_watch_key(invoice)
                    planned, accepted, _ = _fbo_invoice_quantities(invoice)
                    terminal = fbo_invoice_acceptance_is_terminal(invoice)
                    save_fbo_acceptance_watch_state(
                        telegram_id,
                        shop_id,
                        invoice_key,
                        invoice_id=_invoice_id(invoice),
                        invoice_number=_invoice_number(invoice),
                        status=_invoice_status(invoice),
                        planned_qty=planned,
                        accepted_qty=accepted,
                        terminal=terminal,
                        baseline_notified=terminal,
                    )
                mark_operational_watcher_initialized(
                    telegram_id,
                    shop_id,
                    "fbo_acceptance",
                )
                logging.info(
                    "FBO acceptance watcher initialized user=%s shop=%s invoices=%s",
                    telegram_id,
                    shop_id,
                    len(invoices),
                )
                continue

            for invoice in invoices:
                invoice_key = _fbo_invoice_watch_key(invoice)
                previous = get_fbo_acceptance_watch_state(
                    telegram_id,
                    shop_id,
                    invoice_key,
                )
                terminal = fbo_invoice_acceptance_is_terminal(invoice)
                planned, accepted, _ = _fbo_invoice_quantities(invoice)
                save_fbo_acceptance_watch_state(
                    telegram_id,
                    shop_id,
                    invoice_key,
                    invoice_id=_invoice_id(invoice),
                    invoice_number=_invoice_number(invoice),
                    status=_invoice_status(invoice),
                    planned_qty=planned,
                    accepted_qty=accepted,
                    terminal=terminal,
                )
                state = get_fbo_acceptance_watch_state(
                    telegram_id,
                    shop_id,
                    invoice_key,
                ) or {}
                if not terminal or state.get("notified_at"):
                    continue
                if previous and previous.get("notified_at"):
                    continue

                if invoice_key not in products_cache and invoice_key not in products_error:
                    invoice_id = _invoice_id(invoice)
                    try:
                        numeric_invoice_id = int(str(invoice_id).strip())
                        raw_products = await _request_fbo_invoice_products(
                            client,
                            shop_id,
                            numeric_invoice_id,
                        )
                        products_cache[invoice_key] = [
                            item
                            for item in _extract_list_any(raw_products)
                            if isinstance(item, dict)
                        ]
                    except Exception as error:
                        products_error.add(invoice_key)
                        products_cache[invoice_key] = []
                        if _is_uzum_rate_limit_error(error):
                            products_retry_later.add(invoice_key)
                            _fbo_full_scan_at_by_group.pop(watch_key, None)
                            _log_watcher_api_failure(
                                f"FBO acceptance products invoice={invoice_id}",
                                error,
                                shop_id=shop_id,
                                telegram_ids=telegram_ids,
                            )
                            # Не помечаем накладную обработанной: детали будут
                            # повторно загружены на следующем цикле.
                            continue
                        logging.exception(
                            "FBO acceptance watcher: products failed shop=%s invoice=%s",
                            shop_id,
                            invoice_id,
                        )

                if invoice_key in products_retry_later:
                    continue

                summary = summarize_fbo_acceptance(
                    invoice,
                    products_cache.get(invoice_key, []),
                    shop_id=shop_id,
                )
                if summary.get("outcome") == "unknown":
                    logging.warning(
                        "FBO acceptance watcher: terminal invoice without totals shop=%s invoice=%s",
                        shop_id,
                        _invoice_id(invoice),
                    )
                    continue

                lang = get_user_language(telegram_id)
                text = build_fbo_acceptance_notification(summary, lang=lang)
                pdf_path: Path | None = None
                try:
                    if summary.get("outcome") == "discrepancy":
                        try:
                            pdf_path = await asyncio.to_thread(build_fbo_acceptance_pdf, summary, lang=lang)
                            safe_invoice = "".join(
                                char if char.isalnum() else "_"
                                for char in str(summary.get("invoice_number") or summary.get("invoice_id") or "invoice")
                            )[:40]
                            await bot.send_document(
                                telegram_id,
                                FSInputFile(
                                    str(pdf_path),
                                    filename=f"fbo_acceptance_{safe_invoice or 'invoice'}.pdf",
                                ),
                                caption=text,
                                reply_markup=main_menu_for_user(telegram_id),
                            )
                        except Exception:
                            logging.exception(
                                "FBO acceptance PDF failed; sending text fallback user=%s shop=%s invoice=%s",
                                telegram_id,
                                shop_id,
                                _invoice_id(invoice),
                            )
                            fallback_note = (
                                "\n\n⚠️ PDF yaratilmadi, lekin tafovut saqlandi va tekshirish talab etiladi."
                                if normalize_lang(lang) == "uz"
                                else "\n\n⚠️ PDF не сформирован, но расхождение сохранено и требует проверки."
                            )
                            await bot.send_message(
                                telegram_id,
                                text + fallback_note,
                                reply_markup=main_menu_for_user(telegram_id),
                            )
                    else:
                        await bot.send_message(
                            telegram_id,
                            text,
                            reply_markup=main_menu_for_user(telegram_id),
                        )
                    mark_fbo_acceptance_notified(
                        telegram_id,
                        shop_id,
                        invoice_key,
                    )
                    logging.info(
                        "FBO acceptance notification sent user=%s shop=%s invoice=%s outcome=%s",
                        telegram_id,
                        shop_id,
                        _invoice_id(invoice),
                        summary.get("outcome"),
                    )
                    await asyncio.sleep(0.2)
                except Exception:
                    logging.exception(
                        "FBO acceptance watcher: delivery failed user=%s shop=%s invoice=%s",
                        telegram_id,
                        shop_id,
                        _invoice_id(invoice),
                    )
                finally:
                    if pdf_path is not None:
                        try:
                            pdf_path.unlink(missing_ok=True)
                        except OSError:
                            pass
        await asyncio.sleep(0.5)


async def fbo_acceptance_watch_loop() -> None:
    await asyncio.sleep(110)
    logging.info(
        "FBO acceptance watcher started. Interval: %s seconds",
        FBO_ACCEPTANCE_CHECK_INTERVAL_SECONDS,
    )
    while True:
        try:
            await check_fbo_acceptance_once()
        except Exception:
            logging.exception("FBO acceptance watcher loop error")
        await asyncio.sleep(max(300, FBO_ACCEPTANCE_CHECK_INTERVAL_SECONDS))


# --- Deadline reminders for FBO supplies and marketplace returns ---
def _logistics_status_code(value: Any) -> str:
    return "_".join(
        _status_text_any(value).strip().upper().replace("-", " ").split()
    )


def _logistics_is_terminal(value: Any) -> bool:
    status = _logistics_status_code(value)
    return any(
        marker in status
        for marker in (
            "CANCEL",
            "CLOSED",
            "COMPLETED",
            "FINISHED",
            "REJECTED",
            "DECLINED",
            "ACCEPTED",
        )
    )


def _logistics_location(item: dict[str, Any]) -> str:
    value = _value_by_path(
        item,
        "warehouse.name",
        "warehouseName",
        "logisticWarehouse.name",
        "logisticWarehouseName",
        "stock.name",
        "stockName",
    )
    return str(value or "—")


def _invoice_slot(item: dict[str, Any]) -> tuple[datetime | None, datetime | None]:
    return (
        parse_api_datetime(
            _value_by_path(item, "timeSlotReservation.timeFrom", "timeFrom")
        ),
        parse_api_datetime(
            _value_by_path(item, "timeSlotReservation.timeTo", "timeTo")
        ),
    )


def _return_id(item: dict[str, Any]) -> str:
    return str(_value_by_path(item, "id", "returnId", "return.id") or "").strip()


def _return_number(item: dict[str, Any]) -> str:
    return str(
        _value_by_path(
            item,
            "returnNumber",
            "number",
            "return.number",
            "deliveryCertificate",
        )
        or _return_id(item)
        or "—"
    )


def _return_status(item: dict[str, Any]) -> str:
    return _status_text_any(
        _value_by_path(item, "returnStatus", "status", "state", "returnStatus.value")
    )


def _return_paid_storage(item: dict[str, Any]) -> tuple[datetime | None, str]:
    start_at = parse_api_datetime(
        _value_by_path(
            item,
            "paidStorage.startDate",
            "paidStorage.dateFrom",
            "paidStorageStartDate",
            "storageStartDate",
        )
    )
    status = _status_text_any(
        _value_by_path(
            item,
            "paidStorage.status",
            "paidStorageStatus",
            "storageStatus",
        )
    )
    return start_at, status


def _return_pickup_slot(item: dict[str, Any]) -> tuple[datetime | None, datetime | None]:
    return (
        parse_api_datetime(
            _value_by_path(item, "timeSlotReservation.timeFrom", "timeFrom")
        ),
        parse_api_datetime(
            _value_by_path(item, "timeSlotReservation.timeTo", "timeTo")
        ),
    )


def _logistics_interval_text(start_at: datetime | None, end_at: datetime | None) -> str:
    if start_at is None:
        return "—"
    start_text = _fmt_dt(start_at)
    if end_at is None:
        return start_text
    if start_at.date() == end_at.date():
        return f"{start_text}–{end_at.astimezone(UZT).strftime('%H:%M')}"
    return f"{start_text}–{_fmt_dt(end_at)}"


def build_supply_reminder_message(
    invoice: dict[str, Any],
    *,
    shop_id: int,
    bucket: str,
    lang: str,
) -> str:
    start_at, end_at = _invoice_slot(invoice)
    number = escape(_short_text(_invoice_number(invoice), 80))
    location = escape(_short_text(_logistics_location(invoice), 90))
    slot = escape(_logistics_interval_text(start_at, end_at))
    status = escape(_short_text(_invoice_status(invoice), 60))
    if normalize_lang(lang) == "uz":
        urgency = "3 soatdan kam qoldi" if bucket == "3h" else "24 soatdan kam qoldi"
        return (
            f"⏰ <b>Yetkazish vaqti — {urgency}</b>\n\n"
            f"🏪 Do‘kon: <code>{shop_id}</code>\n"
            f"📄 Yuk xati: <b>{number}</b>\n"
            f"🏬 Ombor: {location}\n"
            f"🕒 Vaqt: <b>{slot}</b>\n"
            f"📌 Holat: {status}\n\n"
            "Uzum Seller’da yuk tayyorligini va belgilangan vaqtni tekshiring."
        )
    urgency = "осталось не более 3 часов" if bucket == "3h" else "осталось не более 24 часов"
    return (
        f"⏰ <b>Срок поставки — {urgency}</b>\n\n"
        f"🏪 Магазин: <code>{shop_id}</code>\n"
        f"📄 Накладная: <b>{number}</b>\n"
        f"🏬 Склад: {location}\n"
        f"🕒 Слот: <b>{slot}</b>\n"
        f"📌 Статус: {status}\n\n"
        "Проверьте готовность груза и назначенный слот в Uzum Seller."
    )


def build_return_pickup_reminder_message(
    item: dict[str, Any],
    *,
    shop_id: int,
    bucket: str,
    lang: str,
) -> str:
    start_at, storage_status = _return_paid_storage(item)
    pickup_from, pickup_to = _return_pickup_slot(item)
    number = escape(_short_text(_return_number(item), 80))
    location = escape(_short_text(_logistics_location(item), 90))
    start_text = escape(_fmt_dt(start_at))
    pickup_text = escape(_logistics_interval_text(pickup_from, pickup_to))
    status = escape(_short_text(_return_status(item), 60))
    storage_status_text = escape(_short_text(storage_status, 60))
    if normalize_lang(lang) == "uz":
        urgency = {
            "3d": "3 kun qoldi",
            "2d": "2 kun qoldi",
            "1d": "1 kun qoldi",
            "active": "pullik saqlash boshlandi",
        }.get(bucket, "muddat yaqin")
        lines = [
            f"↩️ <b>Qaytarmani olib ketish — {urgency}</b>",
            "",
            f"🏪 Do‘kon: <code>{shop_id}</code>",
            f"📄 Qaytarma: <b>{number}</b>",
            f"🏬 Ombor: {location}",
            f"💳 Pullik saqlash boshlanishi: <b>{start_text}</b>",
            f"📌 Holat: {status} / {storage_status_text}",
        ]
        if pickup_from is not None:
            lines.append(f"🕒 Olib ketish vaqti: <b>{pickup_text}</b>")
        lines.extend(["", "Qo‘shimcha saqlash xarajati bo‘lmasligi uchun Uzum Seller’da qaytarmani tekshiring."])
        return "\n".join(lines)
    urgency = {
        "3d": "осталось 3 дня",
        "2d": "осталось 2 дня",
        "1d": "остался 1 день",
        "active": "платное хранение началось",
    }.get(bucket, "срок приближается")
    lines = [
        f"↩️ <b>Заберите возврат — {urgency}</b>",
        "",
        f"🏪 Магазин: <code>{shop_id}</code>",
        f"📄 Возврат: <b>{number}</b>",
        f"🏬 Склад: {location}",
        f"💳 Начало платного хранения: <b>{start_text}</b>",
        f"📌 Статус: {status} / {storage_status_text}",
    ]
    if pickup_from is not None:
        lines.append(f"🕒 Слот получения: <b>{pickup_text}</b>")
    lines.extend(["", "Проверьте возврат в Uzum Seller, чтобы избежать лишних расходов на хранение."])
    return "\n".join(lines)


async def _load_return_invoices(
    client: UzumClient,
    shop_id: int,
    *,
    max_pages: int,
    page_size: int = 20,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for page in range(max(1, int(max_pages))):
        data = await client.get_returns(shop_id, page=page, size=page_size)
        items = [row for row in _extract_list_any(data) if isinstance(row, dict)]
        if not items:
            break
        for item in items:
            identity = _return_id(item) or hashlib.sha256(
                json.dumps(item, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
            ).hexdigest()[:24]
            if identity not in seen:
                seen.add(identity)
                rows.append(item)
        if len(items) < page_size:
            break
    return rows


async def check_logistics_reminders_once() -> None:
    now = _utc_now()
    for group in connected_watch_groups(
        "notify_supply_reminders",
        "notify_return_pickup",
    ):
        shop_id = int(group["shop_id"])
        telegram_ids = [int(value) for value in group["telegram_ids"]]
        supply_users = [
            user_id
            for user_id in telegram_ids
            if product_setting_enabled(user_id, "notify_supply_reminders")
        ]
        return_users = [
            user_id
            for user_id in telegram_ids
            if product_setting_enabled(user_id, "notify_return_pickup")
        ]
        try:
            token = cipher.decrypt(group["uzum_token_encrypted"])
            client = UzumClient(token, UZUM_API_BASE_URL)
        except Exception as error:
            _log_watcher_api_failure(
                "Logistics reminder token",
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            continue

        invoices: list[dict[str, Any]] = []
        returns: list[dict[str, Any]] = []
        if supply_users:
            try:
                invoices, _ = await _load_fbo_invoices(
                    client,
                    shop_id,
                    max_pages=LOGISTICS_REMINDER_INVOICE_PAGES,
                    page_size=20,
                )
            except Exception as error:
                _log_watcher_api_failure(
                    "Supply deadline reminder",
                    error,
                    shop_id=shop_id,
                    telegram_ids=supply_users,
                )
        if return_users:
            try:
                returns = await _load_return_invoices(
                    client,
                    shop_id,
                    max_pages=LOGISTICS_REMINDER_RETURN_PAGES,
                )
            except Exception as error:
                _log_watcher_api_failure(
                    "Return pickup reminder",
                    error,
                    shop_id=shop_id,
                    telegram_ids=return_users,
                )

        for invoice in invoices:
            if _logistics_is_terminal(_invoice_status(invoice)):
                continue
            start_at, _ = _invoice_slot(invoice)
            bucket = supply_reminder_bucket(start_at, now)
            if bucket is None:
                continue
            invoice_id = str(_invoice_id(invoice) or _invoice_number(invoice)).strip()
            slot_key = start_at.isoformat() if start_at else "unknown"
            for telegram_id in supply_users:
                _enqueue_notification(
                    "supply_reminder",
                    telegram_id,
                    shop_id,
                    f"invoice:{invoice_id}:slot:{slot_key}:{bucket}",
                    {
                        "text": build_supply_reminder_message(
                            invoice,
                            shop_id=shop_id,
                            bucket=bucket,
                            lang=get_user_language(telegram_id),
                        )
                    },
                )

        for item in returns:
            if _logistics_is_terminal(_return_status(item)):
                continue
            start_at, storage_status = _return_paid_storage(item)
            bucket = return_reminder_bucket(
                start_at,
                now,
                storage_status=storage_status,
            )
            if bucket is None:
                continue
            return_id = _return_id(item) or _return_number(item)
            storage_key = start_at.isoformat() if start_at else "unknown"
            for telegram_id in return_users:
                _enqueue_notification(
                    "return_pickup_reminder",
                    telegram_id,
                    shop_id,
                    f"return:{return_id}:storage:{storage_key}:{bucket}",
                    {
                        "text": build_return_pickup_reminder_message(
                            item,
                            shop_id=shop_id,
                            bucket=bucket,
                            lang=get_user_language(telegram_id),
                        )
                    },
                )
        await asyncio.sleep(0.2)

    await _deliver_pending_notifications()


async def logistics_reminder_watch_loop() -> None:
    await asyncio.sleep(135)
    logging.info(
        "Logistics reminder watcher started. Interval: %s seconds",
        LOGISTICS_REMINDER_CHECK_INTERVAL_SECONDS,
    )
    while True:
        try:
            await check_logistics_reminders_once()
        except Exception:
            logging.exception("Logistics reminder watcher loop error")
        await asyncio.sleep(LOGISTICS_REMINDER_CHECK_INTERVAL_SECONDS)

# =============================================================================
# RELEASE HARDENING 2026-07
# Correct Finance math, partial cancellations, SKU isolation, persistent watcher
# state, retryable Telegram notifications, rate-limit-safe loss report, and
# preservation of pending Telegram updates. Subscription logic is unchanged.
# =============================================================================

RELEASE_VERSION = "2026.07.17-r1"
DROP_PENDING_UPDATES = (
    os.getenv("DROP_PENDING_UPDATES", "0").strip().lower()
    in {"1", "true", "yes", "on", "да"}
)
SALE_WATCH_LOOKBACK_DAYS = max(
    1, min(30, int(os.getenv("SALE_WATCH_LOOKBACK_DAYS", "7") or "7"))
)
SALE_WATCH_MAX_PAGES = max(
    5, min(200, int(os.getenv("SALE_WATCH_MAX_PAGES", "50") or "50"))
)
SALE_WATCH_BASELINE_PAGES = max(
    1, min(SALE_WATCH_MAX_PAGES, int(os.getenv("SALE_WATCH_BASELINE_PAGES", "5") or "5"))
)
SALE_WATCH_PAGE_SIZE = max(
    20, min(100, int(os.getenv("SALE_WATCH_PAGE_SIZE", "100") or "100"))
)
# Finance /v1/finance/orders does not document a status filter.  Older builds
# sent statuses=CANCELLED/CANCELED/... and received repeated HTTP 400 responses.
# Cancellation changes are now detected from normal Finance rows, with a deeper
# unfiltered scan at a low frequency for older orders.
CANCEL_WATCH_DEEP_SCAN_INTERVAL_SECONDS = max(
    900,
    int(os.getenv("CANCEL_WATCH_DEEP_SCAN_INTERVAL_SECONDS", "3600") or "3600"),
)
CANCEL_WATCH_DEEP_SCAN_MAX_PAGES = max(
    1,
    min(
        SALE_WATCH_MAX_PAGES,
        int(os.getenv("CANCEL_WATCH_DEEP_SCAN_MAX_PAGES", "10") or "10"),
    ),
)
SALE_NOTIFICATION_BATCH_LIMIT = max(
    1, min(100, int(os.getenv("SALE_NOTIFICATION_BATCH_LIMIT", "30") or "30"))
)
SALE_DIGEST_THRESHOLD = max(
    0, min(1000, int(os.getenv("SALE_DIGEST_THRESHOLD", "25") or "25"))
)
NOTIFICATION_MAX_ATTEMPTS = max(
    0, min(1000, int(os.getenv("NOTIFICATION_MAX_ATTEMPTS", "10") or "10"))
)
WATCH_STATE_RETENTION_DAYS = max(
    7, min(365, int(os.getenv("WATCH_STATE_RETENTION_DAYS", "45") or "45"))
)
STOCK_WATCH_CACHE_SECONDS = max(
    15, min(600, int(os.getenv("STOCK_WATCH_CACHE_SECONDS", "90") or "90"))
)
LOSS_REPORT_CACHE_SECONDS = max(
    60, min(86400, int(os.getenv("LOSS_REPORT_CACHE_SECONDS", "3600") or "3600"))
)
LOSS_REPORT_MAX_REQUESTS = max(
    2, min(100, int(os.getenv("LOSS_REPORT_MAX_REQUESTS", "40") or "40"))
)
LOSS_REPORT_FILTERS = tuple(
    value.strip().upper()
    for value in os.getenv("LOSS_REPORT_FILTERS", "ALL,ARCHIVE,DEFECTED").replace(";", ",").split(",")
    if value.strip()
) or ("ALL", "ARCHIVE", "DEFECTED")
# Kept only as an in-memory scheduler.  No unsupported status query parameters
# are sent to Uzum Finance API.
_CANCEL_DEEP_SCAN_LAST_ATTEMPT: dict[int, float] = {}


def _finance_original_qty(item: dict[str, Any]) -> float:
    """Quantity from Uzum without inventing 1 when the API explicitly returns 0."""
    value = _deep_pick_number(
        item,
        (
            "amount",
            "quantity",
            "count",
            "qty",
            "skuAmount",
            "productAmount",
            "quantityPurchased",
            "orderedQuantity",
            "orderAmount",
        ),
    )
    if value is None:
        return 0.0
    return max(0.0, float(value))


def _status_upper(status: Any) -> str:
    return str(status or "").strip().upper().replace("-", "_").replace(" ", "_")


def _is_partial_cancel_status(status: Any) -> bool:
    value = _status_upper(status)
    return "PARTIAL" in value and ("CANCEL" in value or "ОТМЕН" in value)


def _is_full_cancelled_status(status: Any) -> bool:
    value = _status_upper(status)
    if _is_partial_cancel_status(value):
        return False
    return "CANCEL" in value or "ОТМЕН" in value


def _has_cancel_event_status(status: Any) -> bool:
    return _is_full_cancelled_status(status) or _is_partial_cancel_status(status)


# Compatibility: existing calculations use this predicate to decide whether the
# whole line must be excluded. A partial cancellation must not exclude it.
def _is_cancelled_status(status: str) -> bool:
    return _is_full_cancelled_status(status)


def _finance_cancelled_qty(item: dict[str, Any], status: str | None = None) -> float:
    ordered = _finance_original_qty(item)
    explicit = _deep_pick_number(
        item,
        (
            # Documented SellerOrderItemDto field.
            "cancelled",
            "amountCancelled",
            "amountCanceled",
            "cancelledAmount",
            "canceledAmount",
            "cancelledQuantity",
            "canceledQuantity",
            "quantityCancelled",
            "quantityCanceled",
            "cancelAmount",
            "cancelQty",
        ),
    )
    explicit_qty = max(0.0, abs(float(explicit))) if explicit is not None else 0.0
    current_status = status or _finance_status(item)
    if _is_full_cancelled_status(current_status):
        return ordered if ordered > 0 else explicit_qty
    if ordered > 0:
        return min(ordered, explicit_qty)
    return explicit_qty


def _finance_return_qty(item: dict[str, Any], status: str | None = None) -> float:
    explicit = _deep_pick_number(
        item,
        (
            "amountReturns",
            "returnAmount",
            "returnedAmount",
            "quantityReturns",
            "returnedQuantity",
            "returnQuantity",
            "quantityReturned",
        ),
    )
    explicit_qty = max(0.0, abs(float(explicit))) if explicit is not None else 0.0
    current_status = status or _finance_status(item)
    ordered = _finance_original_qty(item)
    if _is_returned_status(current_status) and explicit_qty <= 0:
        return ordered
    if ordered > 0:
        return min(ordered, explicit_qty)
    return explicit_qty


def _finance_qty(item: dict[str, Any]) -> float:
    """Net sold quantity after partial/full cancellation and explicit returns."""
    ordered = _finance_original_qty(item)
    if ordered <= 0:
        return 0.0
    status = _finance_status(item)
    cancelled = _finance_cancelled_qty(item, status)
    returned = _finance_return_qty(item, status)
    if _is_returned_status(status) and returned <= 0:
        returned = ordered
    return max(0.0, ordered - cancelled - returned)


def _finance_unit_price(item: dict[str, Any]) -> float | None:
    """Uzum sellerPrice is a unit price, not a total for the whole row."""
    value = _deep_pick_number(
        item,
        (
            "sellerPrice",
            "sellPrice",
            "soldPrice",
            "priceWithDiscount",
            "productPrice",
            "skuPrice",
            "price",
        ),
    )
    if value is None:
        return None
    return max(0.0, float(value))


def _finance_explicit_total(item: dict[str, Any]) -> float | None:
    # Deliberately excludes sellerPrice and sellerAmount: those are not a safe
    # gross row total in the Seller OpenAPI finance schema.
    value = _deep_pick_number(
        item,
        (
            "totalSellerPrice",
            "totalPrice",
            "totalAmount",
            "totalSum",
            "grossAmount",
            "orderTotalPrice",
            "lineTotal",
        ),
    )
    if value is None:
        return None
    return max(0.0, float(value))


def _finance_gross_for_qty(item: dict[str, Any], qty: float) -> float:
    qty = max(0.0, float(qty))
    unit_price = _finance_unit_price(item)
    if unit_price is not None:
        return unit_price * qty

    explicit_total = _finance_explicit_total(item)
    if explicit_total is None:
        return 0.0

    ordered = _finance_original_qty(item)
    if ordered > 0 and qty < ordered:
        return explicit_total * (qty / ordered)
    return explicit_total if qty > 0 else 0.0


def _finance_gross_revenue(item: dict[str, Any]) -> float:
    return _finance_gross_for_qty(item, _finance_qty(item))


def _finance_revenue(item: dict[str, Any]) -> float:
    return _finance_gross_revenue(item)


def _finance_cancelled_revenue(item: dict[str, Any], cancel_qty: float | None = None) -> float:
    qty = _finance_cancelled_qty(item) if cancel_qty is None else max(0.0, float(cancel_qty))
    return _finance_gross_for_qty(item, qty)


def _finance_variant_key(item: dict[str, Any]) -> str:
    for names in (
        ("skuId", "sku_id"),
        ("sellerSku", "sellerSKU", "shopSku", "offerId"),
        ("barcode",),
        ("skuTitle", "skuFullTitle", "skuName"),
    ):
        value = _deep_pick_value(item, names)
        if isinstance(value, dict):
            value = pick(value, "id", "value", "title", "name", default=None)
        text = str(value or "").strip()
        if text and text not in {"-", "—"}:
            return _unit_sku_key(text)
    # Last resort only. This fallback is not mixed with stronger SKU identifiers.
    return _unit_sku_key(_finance_title(item))


def _finance_sku_key_for_stats(item: dict[str, Any]) -> str:
    return _finance_variant_key(item) or "—"


def _unit_group_key(item: dict[str, Any]) -> str:
    return _finance_variant_key(item)


def _finance_purchase_price(item: dict[str, Any]) -> float | None:
    """Historical unit cost documented on the Seller Finance order line."""
    value = _deep_pick_number(item, ("purchasePrice", "purchase_price"))
    if value is None or float(value) <= 0:
        return None
    return float(value)


def _finance_scheme(item: dict[str, Any]) -> str:
    value = _deep_pick_value(
        item,
        ("scheme", "fulfillmentScheme", "fulfilmentScheme", "deliveryScheme"),
    )
    if isinstance(value, dict):
        value = pick(value, "value", "code", "name", "title", default=None)
    normalized = _status_upper(value)
    if normalized in {"FBO", "FBS", "DBS"}:
        return normalized
    return "—"


def _unit_cost_with_source(
    costs: dict[str, dict[str, Any]],
    item: dict[str, Any],
) -> tuple[float | None, str | None]:
    historical = _finance_purchase_price(item)
    if historical is not None:
        return historical, "uzum_finance_purchasePrice"

    candidates: list[str] = []
    for names in (
        ("skuId", "sku_id"),
        ("sellerSku", "sellerSKU", "shopSku", "offerId"),
        ("barcode",),
        ("skuTitle", "skuFullTitle", "skuName"),
    ):
        value = _deep_pick_value(item, names)
        if isinstance(value, dict):
            value = pick(value, "id", "value", "title", "name", default=None)
        key = _unit_sku_key(value)
        if key:
            candidates.append(key)
    visible_key = _unit_sku_key(_finance_sku_title(item))
    if visible_key:
        candidates.append(visible_key)
    for key in candidates:
        if key not in costs:
            continue
        try:
            value = float(costs[key].get("cost") or 0)
        except (TypeError, ValueError):
            return None, None
        if value > 0:
            return value, "uzum_catalog_purchasePrice"
    return None, None


def _unit_cost_lookup(
    costs: dict[str, dict[str, Any]],
    item: dict[str, Any],
) -> float | None:
    value, _ = _unit_cost_with_source(costs, item)
    return value


def _sale_match_keys(item: dict[str, Any]) -> set[str]:
    strong: set[str] = set()
    for names in (
        ("skuId", "sku_id"),
        ("sellerSku", "sellerSKU", "shopSku", "offerId"),
        ("barcode",),
        ("skuTitle", "skuFullTitle", "skuName"),
    ):
        value = _deep_pick_value(item, names)
        if isinstance(value, dict):
            value = pick(value, "id", "value", "title", "name", default=None)
        key = str(value or "").strip().lower()
        if key and key not in {"-", "—"}:
            strong.add(key)
    if strong:
        return strong
    title = str(_finance_title(item) or "").strip().lower()
    return {title} if title else set()


def _stock_match_keys(row: dict[str, Any]) -> set[str]:
    strong: set[str] = set()
    for value in (
        row.get("sku_id"),
        row.get("seller_item_code"),
        row.get("barcode"),
        row.get("sku_full_title"),
        row.get("sku_title"),
    ):
        key = str(value or "").strip().lower()
        if key and key not in {"-", "—"}:
            strong.add(key)
    if strong:
        return strong
    title = str(row.get("product_title") or "").strip().lower()
    return {title} if title else set()


def _build_sales_stats(rows: list[dict[str, Any]]) -> dict[str, Any]:
    active_rows = 0
    cancelled_rows = 0
    cancelled_units = 0.0
    revenue = 0.0
    units = 0.0
    statuses: dict[str, int] = {}
    products: dict[str, dict[str, float | str]] = {}

    for item in rows:
        status = _finance_status(item)
        statuses[status] = statuses.get(status, 0) + 1
        cancelled_qty = _finance_cancelled_qty(item, status)
        if cancelled_qty > 0 or _has_cancel_event_status(status):
            cancelled_rows += 1
            cancelled_units += cancelled_qty

        qty = _finance_qty(item)
        if qty <= 0 or _is_returned_status(status):
            continue

        active_rows += 1
        amount = _finance_gross_revenue(item)
        revenue += amount
        units += qty

        sku = _finance_variant_key(item) or _finance_title(item)
        if sku not in products:
            products[sku] = {
                "title": _finance_title(item),
                "sku": sku,
                "qty": 0.0,
                "revenue": 0.0,
            }
        products[sku]["qty"] = float(products[sku]["qty"]) + qty
        products[sku]["revenue"] = float(products[sku]["revenue"]) + amount

    top_products = sorted(
        products.values(),
        key=lambda value: float(value.get("revenue") or 0),
        reverse=True,
    )[:5]
    return {
        "rows": len(rows),
        "cancelled": cancelled_rows,
        "cancelled_units": cancelled_units,
        "active_rows": active_rows,
        "revenue": revenue,
        "units": units,
        "avg": revenue / max(1, active_rows),
        "statuses": statuses,
        "top_products": top_products,
    }


def _build_noorza_today_stats(rows: list[dict[str, Any]]) -> dict[str, Any]:
    active_rows = 0
    cancelled_rows = 0
    cancelled_units = 0.0
    returned_rows = 0
    returns = 0.0
    units = 0.0
    revenue = 0.0
    commission = 0.0
    logistics = 0.0
    payout_total = 0.0
    withdrawn = 0.0
    statuses: dict[str, int] = {}
    order_keys: set[str] = set()
    products: dict[str, dict[str, Any]] = {}

    for item in rows:
        status = _finance_status(item)
        statuses[status] = statuses.get(status, 0) + 1

        cancelled_qty = _finance_cancelled_qty(item, status)
        if cancelled_qty > 0 or _has_cancel_event_status(status):
            cancelled_rows += 1
            cancelled_units += cancelled_qty

        returned_qty = _finance_return_qty(item, status)
        if returned_qty > 0:
            returns += returned_qty
        if _is_returned_status(status) and _finance_qty(item) <= 0:
            returned_rows += 1
            continue

        qty = _finance_qty(item)
        if qty <= 0:
            continue

        active_rows += 1
        order_keys.add(_finance_order_key_for_stats(item))
        units += qty

        gross = _finance_gross_revenue(item)
        comm = _finance_commission(item)
        logi = _finance_logistics(item)
        payout_direct = _finance_payout_direct(item)
        payout = payout_direct if payout_direct is not None else max(0.0, gross - comm - logi)

        revenue += gross
        commission += comm
        logistics += logi
        payout_total += max(0.0, payout)
        withdrawn += _finance_withdrawn(item)

        sku = _finance_variant_key(item) or "—"
        product = products.setdefault(
            sku,
            {
                "title": _finance_title(item),
                "sku": _finance_sku_title(item),
                "qty": 0.0,
                "revenue": 0.0,
                "payout": 0.0,
            },
        )
        product["qty"] = float(product["qty"]) + qty
        product["revenue"] = float(product["revenue"]) + gross
        product["payout"] = float(product["payout"]) + max(0.0, payout)

    orders = len(order_keys)
    top_products = sorted(
        products.values(),
        key=lambda value: float(value.get("revenue") or 0),
        reverse=True,
    )[:5]
    return {
        "source_rows": len(rows),
        "rows": active_rows,
        "orders": orders,
        "cancelled": cancelled_rows,
        "cancelled_units": cancelled_units,
        "returned_rows": returned_rows,
        "units": units,
        "returns": returns,
        "revenue": revenue,
        "commission": commission,
        "logistics": logistics,
        "payout_total": payout_total,
        "withdrawn": withdrawn,
        "left_to_withdraw": max(0.0, payout_total - withdrawn),
        "average_order": revenue / max(1, orders),
        "average_unit": revenue / max(1.0, units),
        "commission_rate": commission / revenue if revenue > 0 else 0.0,
        "logistics_rate": logistics / revenue if revenue > 0 else 0.0,
        "cancellation_rate": cancelled_rows / max(1, active_rows + cancelled_rows + returned_rows),
        "statuses": statuses,
        "top_products": top_products,
    }


def _build_unit_rows_from_finance(
    rows: list[dict[str, Any]],
    costs: dict[str, dict[str, Any]],
    *,
    tax_percent: float = 0.0,
) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    safe_tax_percent = max(0.0, min(100.0, float(tax_percent or 0)))
    for item in rows:
        status = _finance_status(item)
        qty = _finance_qty(item)
        if qty <= 0 or _is_returned_status(status):
            continue
        key = _unit_group_key(item)
        if not key:
            continue

        revenue = _finance_gross_revenue(item)
        commission = _finance_commission(item)
        logistics = _finance_logistics(item)
        payout_direct = _finance_payout_direct(item)
        payout = payout_direct if payout_direct is not None else max(0.0, revenue - commission - logistics)
        cost_per_unit, cost_source = _unit_cost_with_source(costs, item)
        tax_expense = revenue * safe_tax_percent / 100.0
        scheme = _finance_scheme(item)

        entry = groups.setdefault(
            key,
            {
                "sku": _finance_sku_title(item),
                "title": _finance_title(item),
                "qty": 0.0,
                "revenue": 0.0,
                "commission": 0.0,
                "logistics": 0.0,
                "payout": 0.0,
                "cost_per_unit": None,
                "cost_total": 0.0,
                "known_cost_qty": 0.0,
                "known_revenue": 0.0,
                "known_payout": 0.0,
                "known_commission": 0.0,
                "known_logistics": 0.0,
                "known_tax_expense": 0.0,
                "tax_expense": 0.0,
                "missing_cost_qty": 0.0,
                "cost_sources": set(),
                "schemes": set(),
                "profit": None,
            },
        )
        entry["qty"] += qty
        entry["revenue"] += revenue
        entry["commission"] += commission
        entry["logistics"] += logistics
        entry["payout"] += max(0.0, payout)
        entry["tax_expense"] += tax_expense
        if scheme != "—":
            entry["schemes"].add(scheme)
        if cost_per_unit is not None:
            entry["cost_total"] += float(cost_per_unit) * qty
            entry["known_cost_qty"] += qty
            entry["known_revenue"] += revenue
            entry["known_payout"] += max(0.0, payout)
            entry["known_commission"] += commission
            entry["known_logistics"] += logistics
            entry["known_tax_expense"] += tax_expense
            if cost_source:
                entry["cost_sources"].add(cost_source)
        else:
            entry["missing_cost_qty"] += qty

    for entry in groups.values():
        known_qty = float(entry.get("known_cost_qty") or 0)
        known_revenue = float(entry.get("known_revenue") or 0)
        cost_total = float(entry.get("cost_total") or 0)
        if known_qty > 0:
            entry["cost_per_unit"] = cost_total / known_qty
            entry["profit"] = float(entry.get("known_payout") or 0) - cost_total
            entry["net_profit"] = float(entry["profit"]) - float(entry.get("known_tax_expense") or 0)
            entry["margin"] = float(entry["profit"]) / known_revenue * 100.0 if known_revenue > 0 else 0.0
            entry["net_margin"] = float(entry["net_profit"]) / known_revenue * 100.0 if known_revenue > 0 else 0.0
            entry["roi"] = float(entry["net_profit"]) / cost_total * 100.0 if cost_total > 0 else None
        else:
            entry["margin"] = None
            entry["net_profit"] = None
            entry["net_margin"] = None
            entry["roi"] = None
        entry["cost_complete"] = float(entry.get("missing_cost_qty") or 0) <= 0
        sources = sorted(entry.pop("cost_sources", set()))
        entry["cost_source"] = ",".join(sources) if sources else None
        schemes = sorted(entry.pop("schemes", set()))
        entry["scheme"] = "/".join(schemes) if schemes else "—"
    return sorted(groups.values(), key=lambda value: float(value.get("revenue") or 0), reverse=True)


def _normalized_finance_piece(
    item: dict[str, Any],
    *,
    kind: str,
    qty: float,
    revenue: float,
) -> dict[str, Any]:
    gross_net = _finance_gross_revenue(item)
    commission = _finance_commission(item)
    logistics = _finance_logistics(item)
    direct = _finance_payout_direct(item)
    payout_net = direct if direct is not None else max(0.0, gross_net - commission - logistics)

    # For cancellation/return rows the payout field represents the associated
    # lost payout estimate. It is prorated only for display, without changing
    # the raw Uzum commission/logistics values used in sale totals.
    ordered = _finance_original_qty(item)
    ratio = min(1.0, qty / ordered) if ordered > 0 else 0.0
    if kind == "sale":
        piece_commission = commission
        piece_logistics = logistics
        piece_payout = max(0.0, payout_net)
        withdrawn = _finance_withdrawn(item)
    else:
        piece_commission = commission * ratio
        piece_logistics = logistics * ratio
        piece_payout = max(0.0, revenue - piece_commission - piece_logistics)
        withdrawn = 0.0

    return {
        "date": _finance_datetime_for_report(item),
        "kind": kind,
        "status": _finance_status(item),
        "order_id": _finance_order_key_for_stats(item),
        "title": _finance_title(item),
        "sku": _finance_sku_title(item),
        "qty": max(0.0, qty),
        "revenue": max(0.0, revenue),
        "commission": max(0.0, piece_commission),
        "logistics": max(0.0, piece_logistics),
        "payout": max(0.0, piece_payout),
        "withdrawn": max(0.0, withdrawn),
        "reason": _finance_cancel_reason(item) or str(_deep_pick_value(item, ("returnCause", "comment")) or ""),
        "scheme": _finance_scheme(item),
    }


def _normalize_finance_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for item in rows:
        net_qty = _finance_qty(item)
        cancelled_qty = _finance_cancelled_qty(item)
        returned_qty = _finance_return_qty(item)

        if net_qty > 0:
            result.append(
                _normalized_finance_piece(
                    item,
                    kind="sale",
                    qty=net_qty,
                    revenue=_finance_gross_for_qty(item, net_qty),
                )
            )
        if cancelled_qty > 0:
            result.append(
                _normalized_finance_piece(
                    item,
                    kind="cancel",
                    qty=cancelled_qty,
                    revenue=_finance_gross_for_qty(item, cancelled_qty),
                )
            )
        if returned_qty > 0:
            result.append(
                _normalized_finance_piece(
                    item,
                    kind="return",
                    qty=returned_qty,
                    revenue=_finance_gross_for_qty(item, returned_qty),
                )
            )
        if net_qty <= 0 and cancelled_qty <= 0 and returned_qty <= 0:
            result.append(
                _normalized_finance_piece(item, kind="sale", qty=0.0, revenue=0.0)
            )
    return result


def _normalize_finance_row(item: dict[str, Any]) -> dict[str, Any]:
    pieces = _normalize_finance_rows([item])
    return pieces[0] if pieces else _normalized_finance_piece(
        item, kind="sale", qty=0.0, revenue=0.0
    )


def sale_key(item: dict[str, Any]) -> str:
    """Stable sale-line key that does not change when status/qty changes."""
    return finance_identity_key(item)


def finance_identity_key(item: dict[str, Any]) -> str:
    parts: list[str] = []
    for label, value in (
        ("order", _finance_order_id(item)),
        ("sale", _finance_sale_id(item)),
        ("sku", _finance_variant_key(item)),
    ):
        text = str(value or "").strip()
        if text and text not in {"-", "—"}:
            parts.append(f"{label}:{text}")
    if parts:
        return "|".join(parts)

    date_value = _finance_date_value(item)
    title = _finance_title(item)
    raw_key = f"{date_value}|{title}|{_finance_original_qty(item)}|{_finance_unit_price(item)}"
    return "hash:" + hashlib.sha256(raw_key.encode("utf-8")).hexdigest()


def build_new_sale_message(
    item: dict[str, Any],
    shop_id: int | None = None,
    lang: str = "ru",
) -> str:
    lang = normalize_lang(lang)
    title = escape(_finance_title(item))
    sku = escape(_finance_sku_title(item))
    qty = _finance_qty(item)
    unit_price = _finance_unit_price(item)
    total = _finance_gross_revenue(item)
    if unit_price is None and qty > 0:
        unit_price = total / qty

    commission = _finance_commission(item)
    logistics = _finance_logistics(item)
    payout_direct = _finance_payout_direct(item)
    payout = payout_direct if payout_direct is not None else max(0.0, total - commission - logistics)

    if lang == "uz":
        shop_line = f"🏪 Do‘kon: <code>{shop_id}</code>\n" if shop_id is not None else ""
        return (
            "🛒 <b>Yangi savdo</b>\n\n"
            + shop_line
            + f"📦 Tovar: <b>{title}</b>\n"
            + f"🔖 SKU: <code>{sku}</code>\n"
            + f"🔢 Soni: <b>{qty:g} dona</b>\n\n"
            + f"💵 Dona narxi: <b>{_format_money(float(unit_price or 0))}</b>\n"
            + f"💰 Jami: <b>{_format_money(total)}</b>\n"
            + f"🏷 Komissiya: <b>{_format_money(commission)}</b>\n"
            + f"🚚 Logistika: <b>{_format_money(logistics)}</b>\n"
            + f"✅ To‘lovga: <b>{_format_money(float(payout))}</b>\n\n"
            + f"🆔 Buyurtma: <code>{escape(_finance_order_id(item))}</code>\n"
            + f"📌 Status: <code>{escape(_finance_status(item))}</code>\n"
            + f"🕒 Sana: {escape(_format_finance_date(_finance_date_value(item)))}"
        )

    shop_line = f"🏪 Магазин: <code>{shop_id}</code>\n" if shop_id is not None else ""
    return (
        "🛒 <b>Новая продажа</b>\n\n"
        + shop_line
        + f"📦 Товар: <b>{title}</b>\n"
        + f"🔖 SKU: <code>{sku}</code>\n"
        + f"🔢 Кол-во: <b>{qty:g} шт.</b>\n\n"
        + f"💵 Цена за 1 шт.: <b>{_format_money(float(unit_price or 0))}</b>\n"
        + f"💰 Сумма: <b>{_format_money(total)}</b>\n"
        + f"🏷 Комиссия: <b>{_format_money(commission)}</b>\n"
        + f"🚚 Логистика: <b>{_format_money(logistics)}</b>\n"
        + f"✅ К выплате: <b>{_format_money(float(payout))}</b>\n\n"
        + f"🆔 Заказ: <code>{escape(_finance_order_id(item))}</code>\n"
        + f"📌 Статус: <code>{escape(_finance_status(item))}</code>\n"
        + f"🕒 Дата: {escape(_format_finance_date(_finance_date_value(item)))}"
    )


def _finance_cancel_reason(item: dict[str, Any]) -> str:
    value = _deep_pick_value(
        item,
        (
            "cancelReason",
            "cancellationReason",
            "returnCause",
            "reason",
            "comment",
        ),
    )
    if isinstance(value, dict):
        value = pick(value, "value", "code", "name", "title", "text", default=None)
    return " ".join(str(value or "").strip().split())


_CANCEL_REASON_LABELS: dict[str, tuple[str, str]] = {
    "NOT_FOUND_ON_WAREHOUSE": ("Не найдено на складе", "Omborda topilmadi"),
    "NOT_FOUND_IN_WAREHOUSE": ("Не найдено на складе", "Omborda topilmadi"),
    "OUT_OF_STOCK": ("Нет товара на складе", "Omborda tovar yo‘q"),
    "SELLER_CANCELLED": ("Отменено продавцом", "Sotuvchi bekor qildi"),
    "SELLER_CANCELED": ("Отменено продавцом", "Sotuvchi bekor qildi"),
    "CUSTOMER_CANCELLED": ("Отменено покупателем", "Xaridor bekor qildi"),
    "CUSTOMER_CANCELED": ("Отменено покупателем", "Xaridor bekor qildi"),
    "BUYER_CANCELLED": ("Отменено покупателем", "Xaridor bekor qildi"),
    "BUYER_CANCELED": ("Отменено покупателем", "Xaridor bekor qildi"),
    "DELIVERY_FAILED": ("Не удалось доставить", "Yetkazib berilmadi"),
    "PICKUP_EXPIRED": ("Не забрано из пункта выдачи", "Olib ketish punktidan olinmadi"),
    "NOT_PICKED_UP": ("Не забрано покупателем", "Xaridor olib ketmadi"),
    "OTHER": ("Другая причина", "Boshqa sabab"),
}


def _format_cancel_reason(item: dict[str, Any], *, lang: str) -> str:
    raw = _finance_cancel_reason(item)
    if not raw:
        return (
            "Uzum API sababni bermadi"
            if normalize_lang(lang) == "uz"
            else "Uzum API не передал причину"
        )
    normalized = _status_upper(raw)
    translated = _CANCEL_REASON_LABELS.get(normalized)
    if translated:
        return translated[1] if normalize_lang(lang) == "uz" else translated[0]
    return raw


def build_cancel_message(
    item: dict[str, Any],
    shop_id: int | None = None,
    lang: str = "ru",
    cancel_qty: float | None = None,
) -> str:
    lang = normalize_lang(lang)
    title = escape(_finance_title(item))
    sku = escape(_finance_sku_title(item))
    qty = _finance_cancelled_qty(item) if cancel_qty is None else max(0.0, float(cancel_qty))
    unit_price = _finance_unit_price(item)
    total = _finance_gross_for_qty(item, qty)
    status = escape(_finance_status(item))
    order_id = escape(_finance_order_id(item))
    date_text = escape(_format_finance_date(_finance_date_value(item)))
    scheme = escape(_finance_scheme(item))
    reason = escape(_format_cancel_reason(item, lang=lang))
    qty_text_uz = f"<b>{qty:g} dona</b>" if qty > 0 else "<b>Uzum aniq sonni bermadi</b>"
    qty_text_ru = f"<b>{qty:g} шт.</b>" if qty > 0 else "<b>Uzum не передал точное количество</b>"

    if lang == "uz":
        shop_line = f"🏪 Do‘kon: <code>{shop_id}</code>\n" if shop_id is not None else ""
        return (
            "❌ <b>Buyurtma bekor qilindi</b>\n\n"
            + shop_line
            + f"📦 Tovar: <b>{title}</b>\n"
            + f"🔖 SKU: <code>{sku}</code>\n"
            + f"🔢 Bekor qilindi: {qty_text_uz}\n"
            + f"🚚 Sxema: <b>{scheme}</b>\n"
            + f"💬 Sabab: <b>{reason}</b>\n\n"
            + f"💵 Dona narxi: <b>{_format_money(float(unit_price or 0))}</b>\n"
            + f"💰 Bekor qilingan summa: <b>{_format_money(total)}</b>\n"
            + f"🆔 Buyurtma: <code>{order_id}</code>\n"
            + f"📌 Status: <code>{status}</code>\n"
            + f"🕒 Sana: {date_text}"
        )

    shop_line = f"🏪 Магазин: <code>{shop_id}</code>\n" if shop_id is not None else ""
    return (
        "❌ <b>Отмена заказа</b>\n\n"
        + shop_line
        + f"📦 Товар: <b>{title}</b>\n"
        + f"🔖 SKU: <code>{sku}</code>\n"
        + f"🔢 Отменено: {qty_text_ru}\n"
        + f"🚚 Схема: <b>{scheme}</b>\n"
        + f"💬 Причина: <b>{reason}</b>\n\n"
        + f"💵 Цена за 1 шт.: <b>{_format_money(float(unit_price or 0))}</b>\n"
        + f"💰 Сумма отмены: <b>{_format_money(total)}</b>\n"
        + f"🆔 Заказ: <code>{order_id}</code>\n"
        + f"📌 Статус: <code>{status}</code>\n"
        + f"🕒 Дата: {date_text}"
    )


def init_release_watch_tables() -> None:
    with db.connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS finance_watch_state (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                identity_key TEXT NOT NULL,
                status TEXT NOT NULL,
                ordered_qty REAL NOT NULL DEFAULT 0,
                cancelled_qty REAL NOT NULL DEFAULT 0,
                returned_qty REAL NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, identity_key)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS watcher_meta (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                watch_type TEXT NOT NULL,
                initialized INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, watch_type)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS notification_outbox (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT NOT NULL,
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                event_key TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                attempts INTEGER NOT NULL DEFAULT 0,
                next_attempt_at TEXT,
                last_error TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                sent_at TEXT,
                UNIQUE (event_type, telegram_id, shop_id, event_key)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_notification_outbox_pending
            ON notification_outbox (status, next_attempt_at, id)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS stock_watch_snapshot (
                telegram_id INTEGER NOT NULL,
                shop_id INTEGER NOT NULL,
                watch_type TEXT NOT NULL,
                snapshot_json TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (telegram_id, shop_id, watch_type)
            )
            """
        )
        conn.commit()


init_release_watch_tables()


def _watch_is_initialized(telegram_id: int, shop_id: int, watch_type: str) -> bool:
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT initialized
            FROM watcher_meta
            WHERE telegram_id = ? AND shop_id = ? AND watch_type = ?
            """,
            (int(telegram_id), int(shop_id), str(watch_type)),
        ).fetchone()
    return bool(row and int(row["initialized"] or 0) == 1)


def _set_watch_initialized(telegram_id: int, shop_id: int, watch_type: str) -> None:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        conn.execute(
            """
            INSERT INTO watcher_meta (telegram_id, shop_id, watch_type, initialized, updated_at)
            VALUES (?, ?, ?, 1, ?)
            ON CONFLICT(telegram_id, shop_id, watch_type) DO UPDATE SET
                initialized = 1,
                updated_at = excluded.updated_at
            """,
            (int(telegram_id), int(shop_id), str(watch_type), now_text),
        )
        conn.commit()


def _load_finance_watch_state(
    telegram_id: int,
    shop_id: int,
) -> dict[str, dict[str, Any]]:
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT identity_key, status, ordered_qty, cancelled_qty, returned_qty
            FROM finance_watch_state
            WHERE telegram_id = ? AND shop_id = ?
            """,
            (int(telegram_id), int(shop_id)),
        ).fetchall()
    return {
        str(row["identity_key"]): {
            "status": str(row["status"] or ""),
            "ordered_qty": float(row["ordered_qty"] or 0),
            "cancelled_qty": float(row["cancelled_qty"] or 0),
            "returned_qty": float(row["returned_qty"] or 0),
        }
        for row in rows
    }


def _save_finance_watch_rows(
    telegram_id: int,
    shop_id: int,
    rows: list[dict[str, Any]],
) -> None:
    if not rows:
        return
    now_text = _dt_to_db(_utc_now()) or ""
    values: list[tuple[Any, ...]] = []
    for item in rows:
        values.append(
            (
                int(telegram_id),
                int(shop_id),
                finance_identity_key(item),
                _finance_status(item),
                _finance_original_qty(item),
                _finance_cancelled_qty(item),
                _finance_return_qty(item),
                now_text,
            )
        )
    with db.connect() as conn:
        conn.executemany(
            """
            INSERT INTO finance_watch_state
                (telegram_id, shop_id, identity_key, status, ordered_qty,
                 cancelled_qty, returned_qty, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id, identity_key) DO UPDATE SET
                status = excluded.status,
                ordered_qty = excluded.ordered_qty,
                cancelled_qty = excluded.cancelled_qty,
                returned_qty = excluded.returned_qty,
                updated_at = excluded.updated_at
            """,
            values,
        )
        conn.commit()


def _enqueue_notification(
    event_type: str,
    telegram_id: int,
    shop_id: int,
    event_key: str,
    payload: dict[str, Any],
) -> bool:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        cursor = conn.execute(
            """
            INSERT OR IGNORE INTO notification_outbox
                (event_type, telegram_id, shop_id, event_key, payload_json,
                 status, attempts, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, 'pending', 0, ?, ?)
            """,
            (
                str(event_type),
                int(telegram_id),
                int(shop_id),
                str(event_key),
                json.dumps(payload, ensure_ascii=False, default=str),
                now_text,
                now_text,
            ),
        )
        conn.commit()
    return bool(cursor.rowcount)


def _pending_notifications(limit: int) -> list[dict[str, Any]]:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM notification_outbox
            WHERE status = 'pending'
              AND (next_attempt_at IS NULL OR next_attempt_at <= ?)
            ORDER BY id ASC
            LIMIT ?
            """,
            (now_text, max(1, int(limit))),
        ).fetchall()
    return [dict(row) for row in rows]


def _mark_notification_sent(notification_id: int) -> None:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        conn.execute(
            """
            UPDATE notification_outbox
            SET status = 'sent', sent_at = ?, updated_at = ?, last_error = NULL
            WHERE id = ?
            """,
            (now_text, now_text, int(notification_id)),
        )
        conn.commit()


def _mark_notification_failed(notification_id: int, attempts: int, error: str) -> None:
    now = _utc_now()
    new_attempts = int(attempts) + 1
    if NOTIFICATION_MAX_ATTEMPTS > 0 and new_attempts >= NOTIFICATION_MAX_ATTEMPTS:
        status = "dead"
        next_attempt = None
    else:
        status = "pending"
        delay_seconds = min(3600, 30 * (2 ** min(new_attempts, 7)))
        next_attempt = _dt_to_db(now + timedelta(seconds=delay_seconds))
    with db.connect() as conn:
        conn.execute(
            """
            UPDATE notification_outbox
            SET status = ?, attempts = ?, next_attempt_at = ?, last_error = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                status,
                new_attempts,
                next_attempt,
                str(error)[:1000],
                _dt_to_db(now) or "",
                int(notification_id),
            ),
        )
        conn.commit()



def _sale_digest_payload(items: list[dict[str, Any]]) -> dict[str, Any]:
    stats = _build_sales_stats(items)
    order_ids = {
        _finance_order_key_for_stats(item)
        for item in items
        if _finance_order_key_for_stats(item) not in {"", "-", "—"}
    }
    return {
        "rows": int(stats.get("active_rows") or 0),
        "orders": len(order_ids),
        "units": float(stats.get("units") or 0),
        "revenue": float(stats.get("revenue") or 0),
        "top_products": list(stats.get("top_products") or [])[:5],
    }


def build_sale_digest_message(
    summary: dict[str, Any],
    *,
    shop_id: int,
    lang: str = "ru",
) -> str:
    rows = int(summary.get("rows") or 0)
    orders = int(summary.get("orders") or 0)
    units = float(summary.get("units") or 0)
    revenue = float(summary.get("revenue") or 0)
    products = list(summary.get("top_products") or [])

    if normalize_lang(lang) == "uz":
        lines = [
            "📈 <b>Yangi savdolar — umumiy xabar</b>",
            f"🏪 Do‘kon: <code>{shop_id}</code>",
            "",
            f"🧾 Yangi pozitsiyalar: <b>{rows}</b>",
            f"🛍 Buyurtmalar: <b>{orders}</b>",
            f"📦 Sotilgan: <b>{units:g} dona</b>",
            f"💰 Tushum: <b>{_format_money(revenue)}</b>",
        ]
        if products:
            lines.append("\n🏆 <b>Top tovarlar:</b>")
            for product in products:
                title = escape(_short_text(str(product.get("title") or product.get("sku") or "—"), 55))
                lines.append(
                    f"• {title}: {float(product.get('qty') or 0):g} dona — "
                    f"<b>{_format_money(float(product.get('revenue') or 0))}</b>"
                )
        lines.append("\nKo‘p savdo bo‘lgani uchun bot ularni bitta xabarga birlashtirdi.")
        return "\n".join(lines)

    lines = [
        "📈 <b>Новые продажи — сводное уведомление</b>",
        f"🏪 Магазин: <code>{shop_id}</code>",
        "",
        f"🧾 Новых позиций: <b>{rows}</b>",
        f"🛍 Заказов: <b>{orders}</b>",
        f"📦 Продано: <b>{units:g} шт.</b>",
        f"💰 Выручка: <b>{_format_money(revenue)}</b>",
    ]
    if products:
        lines.append("\n🏆 <b>Топ товаров:</b>")
        for product in products:
            title = escape(_short_text(str(product.get("title") or product.get("sku") or "—"), 55))
            lines.append(
                f"• {title}: {float(product.get('qty') or 0):g} шт. — "
                f"<b>{_format_money(float(product.get('revenue') or 0))}</b>"
            )
    lines.append("\nИз-за большого количества продаж бот объединил их в одно сообщение.")
    return "\n".join(lines)


def _single_order_payload(data: Any) -> dict[str, Any] | None:
    if not isinstance(data, dict):
        return None
    for key in ("payload", "data", "result"):
        value = data.get(key)
        if isinstance(value, dict):
            nested = _single_order_payload(value)
            return nested or value
    return data


async def _enrich_cancel_item_from_fbs(
    telegram_id: int,
    item: dict[str, Any],
) -> dict[str, Any]:
    """Best-effort FBS/DBS detail lookup for exact scheme and cancel reason.

    Finance rows do not document a fulfilment scheme.  The one-order endpoint
    does, so cancellation alerts use it when available.  A 404 usually means
    the line is outside FBS/DBS; that must not delay the important alert.
    """
    order_id = _finance_order_id(item)
    try:
        numeric_order_id = int(str(order_id).strip())
    except (TypeError, ValueError):
        return item
    client = get_uzum_for_user(telegram_id)
    if client is None:
        return item
    try:
        detail_raw = await client.get_fbs_order(numeric_order_id)
        detail = _single_order_payload(detail_raw)
    except Exception as error:
        logging.info(
            "Cancellation detail unavailable user=%s order=%s: %s",
            telegram_id,
            order_id,
            str(error)[:160],
        )
        return item
    if not detail:
        return item
    return {**item, "_fbs_order_detail": detail}


async def _deliver_pending_notifications() -> None:
    rows = _pending_notifications(SALE_NOTIFICATION_BATCH_LIMIT)
    for row in rows:
        notification_id = int(row["id"])
        try:
            payload = json.loads(str(row.get("payload_json") or "{}"))
            telegram_id = int(row["telegram_id"])
            shop_id = int(row["shop_id"])
            event_type = str(row["event_type"])
            required_feature = (
                "sales_notifications"
                if event_type in {"sale", "sale_digest"}
                else "premium"
            )
            if not feature_access_allowed(telegram_id, required_feature):
                # Do not retry an event that is no longer allowed after a
                # subscription transition.  The durable row remains auditable.
                _mark_notification_sent(notification_id)
                logging.info(
                    "Notification suppressed by access type=%s user=%s shop=%s",
                    event_type,
                    telegram_id,
                    shop_id,
                )
                continue

            if event_type == "sale_digest":
                summary = payload.get("summary")
                if not isinstance(summary, dict):
                    raise ValueError("Outbox digest payload has no summary")
                text = build_sale_digest_message(
                    summary,
                    shop_id=shop_id,
                    lang=get_user_language(telegram_id),
                )
            elif event_type in {"supply_reminder", "return_pickup_reminder"}:
                text = str(payload.get("text") or "").strip()
                if not text:
                    raise ValueError("Outbox reminder payload has no text")
                if len(text) > 4000:
                    raise ValueError("Outbox reminder text is too long")
            else:
                item = payload.get("item")
                if not isinstance(item, dict):
                    raise ValueError("Outbox payload has no finance item")
                if event_type == "sale":
                    text = build_new_sale_message(
                        item,
                        shop_id=shop_id,
                        lang=get_user_language(telegram_id),
                    )
                elif event_type == "cancel":
                    item = await _enrich_cancel_item_from_fbs(telegram_id, item)
                    text = build_cancel_message(
                        item,
                        shop_id=shop_id,
                        lang=get_user_language(telegram_id),
                        cancel_qty=float(payload.get("cancel_qty") or 0),
                    )
                else:
                    raise ValueError(f"Unknown outbox event type: {event_type}")

            await bot.send_message(
                telegram_id,
                text,
                reply_markup=main_menu_for_user(telegram_id),
            )
            _mark_notification_sent(notification_id)
            logging.info(
                "Notification sent type=%s user=%s shop=%s key=%s",
                event_type,
                telegram_id,
                shop_id,
                row.get("event_key"),
            )
            await asyncio.sleep(0.12)
        except Exception as exc:
            _mark_notification_failed(
                notification_id,
                int(row.get("attempts") or 0),
                str(exc),
            )
            logging.exception(
                "Notification outbox delivery failed id=%s user=%s",
                notification_id,
                row.get("telegram_id"),
            )
            # Telegram/network outage usually affects subsequent messages too.
            break


async def _load_finance_watch_pages(
    client: UzumClient,
    shop_id: int,
    *,
    known_identities: set[str],
    baseline: bool,
) -> list[dict[str, Any]]:
    now = _utc_now()
    date_from_ms = int((now - timedelta(days=SALE_WATCH_LOOKBACK_DAYS)).timestamp() * 1000)
    date_to_ms = int(now.timestamp() * 1000)
    max_pages = SALE_WATCH_BASELINE_PAGES if baseline else SALE_WATCH_MAX_PAGES

    result: list[dict[str, Any]] = []
    seen_rows: set[str] = set()
    consecutive_known_pages = 0
    for page in range(max_pages):
        data = await _finance_orders_request(
            client,
            shop_id,
            date_from_ms=date_from_ms,
            date_to_ms=date_to_ms,
            page=page,
            size=SALE_WATCH_PAGE_SIZE,
        )
        items = _deep_items(data)
        if not items:
            break

        page_ids: list[str] = []
        for item in items:
            identity = finance_identity_key(item)
            signature = (
                f"{identity}|{_finance_status(item)}|"
                f"{_finance_cancelled_qty(item)}|{_finance_return_qty(item)}"
            )
            if signature in seen_rows:
                continue
            seen_rows.add(signature)
            page_ids.append(identity)
            result.append(item)

        if known_identities and page_ids and all(value in known_identities for value in page_ids):
            consecutive_known_pages += 1
        else:
            consecutive_known_pages = 0

        if len(items) < SALE_WATCH_PAGE_SIZE:
            break
        if not baseline and consecutive_known_pages >= 2:
            break
        await asyncio.sleep(0.04)

    if len(result) >= max_pages * SALE_WATCH_PAGE_SIZE:
        logging.warning(
            "Finance watcher reached safety limit shop=%s rows=%s pages=%s",
            shop_id,
            len(result),
            max_pages,
        )
    return result


async def _load_cancel_status_rows(
    client: UzumClient,
    shop_id: int,
) -> list[dict[str, Any]]:
    """Periodically search older cancellations without unsupported filters.

    Uzum's documented Finance endpoint accepts date/shop/pagination parameters,
    but not ``statuses``.  We therefore load ordinary Finance pages and filter
    cancellation rows locally.  The deep scan runs at most once per configured
    interval for each shop, which avoids both HTTP 400 log spam and unnecessary
    pressure that previously caused HTTP 429 responses.
    """

    now_mono = time.monotonic()
    last_attempt = float(_CANCEL_DEEP_SCAN_LAST_ATTEMPT.get(int(shop_id), 0.0))
    if now_mono - last_attempt < CANCEL_WATCH_DEEP_SCAN_INTERVAL_SECONDS:
        return []

    # Record the attempt before network calls.  A temporary API failure should
    # not cause this expensive scan to repeat every minute.
    _CANCEL_DEEP_SCAN_LAST_ATTEMPT[int(shop_id)] = now_mono

    now = _utc_now()
    date_from_ms = int((now - timedelta(days=SALE_WATCH_LOOKBACK_DAYS)).timestamp() * 1000)
    date_to_ms = int(now.timestamp() * 1000)
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    pages_loaded = 0

    try:
        for page in range(CANCEL_WATCH_DEEP_SCAN_MAX_PAGES):
            data = await _finance_orders_request(
                client,
                shop_id,
                date_from_ms=date_from_ms,
                date_to_ms=date_to_ms,
                page=page,
                size=SALE_WATCH_PAGE_SIZE,
            )
            pages_loaded += 1
            items = _deep_items(data)
            if not items:
                break

            for item in items:
                status = _finance_status(item)
                cancelled_qty = _finance_cancelled_qty(item)
                if not _has_cancel_event_status(status) and cancelled_qty <= 0:
                    continue
                signature = (
                    f"{finance_identity_key(item)}|{_status_upper(status)}|"
                    f"{cancelled_qty:g}|{_finance_return_qty(item):g}"
                )
                if signature in seen:
                    continue
                seen.add(signature)
                rows.append(item)

            if len(items) < SALE_WATCH_PAGE_SIZE:
                break
            await asyncio.sleep(0.08)
    except Exception as exc:
        if _is_uzum_rate_limit_error(exc):
            logging.warning(
                "Cancel deep scan postponed by Uzum rate limit shop=%s pages=%s; "
                "recent cancellation monitoring remains active",
                shop_id,
                pages_loaded,
            )
        else:
            logging.warning(
                "Cancel deep scan failed shop=%s pages=%s: %s",
                shop_id,
                pages_loaded,
                exc,
            )
        return []

    logging.info(
        "Cancel deep scan completed shop=%s pages=%s cancellation_rows=%s next_in=%ss",
        shop_id,
        pages_loaded,
        len(rows),
        CANCEL_WATCH_DEEP_SCAN_INTERVAL_SECONDS,
    )
    return rows


async def check_new_sales_once() -> None:
    for group in connected_watch_groups():
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(value) for value in group["telegram_ids"]]

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)

            state_by_user = {
                telegram_id: _load_finance_watch_state(telegram_id, shop_id)
                for telegram_id in telegram_ids
            }
            initialized_by_user = {
                telegram_id: _watch_is_initialized(telegram_id, shop_id, "finance")
                for telegram_id in telegram_ids
            }
            known_union: set[str] = set()
            for state in state_by_user.values():
                known_union.update(state.keys())
            baseline_group = not any(initialized_by_user.values())

            recent_rows = await _load_finance_watch_pages(
                client,
                shop_id,
                known_identities=known_union,
                baseline=baseline_group,
            )
            cancel_rows = await _load_cancel_status_rows(client, shop_id)

            merged: dict[str, dict[str, Any]] = {}
            for item in recent_rows + cancel_rows:
                identity = finance_identity_key(item)
                current = merged.get(identity)
                if current is None:
                    merged[identity] = item
                    continue
                current_score = (
                    _finance_cancelled_qty(current),
                    _finance_return_qty(current),
                    str(_finance_status(current)),
                )
                item_score = (
                    _finance_cancelled_qty(item),
                    _finance_return_qty(item),
                    str(_finance_status(item)),
                )
                if item_score > current_score:
                    merged[identity] = item
            rows = list(merged.values())

        except Exception:
            logging.exception(
                "Release finance watcher failed shop=%s users=%s",
                shop_id,
                telegram_ids,
            )
            await asyncio.sleep(3)
            continue

        for telegram_id in telegram_ids:
            previous = state_by_user.get(telegram_id, {})
            initialized = initialized_by_user.get(telegram_id, False)

            if not initialized:
                _save_finance_watch_rows(telegram_id, shop_id, rows)
                _set_watch_initialized(telegram_id, shop_id, "finance")
                _sales_watch_initialized.add(telegram_id)
                logging.info(
                    "Sales watcher baseline user=%s shop=%s rows=%s",
                    telegram_id,
                    shop_id,
                    len(rows),
                )
                continue

            new_sales: list[dict[str, Any]] = []
            for item in rows:
                identity = finance_identity_key(item)
                before = previous.get(identity)
                net_qty = _finance_qty(item)
                cancelled_qty = _finance_cancelled_qty(item)
                status = _finance_status(item)

                if before is None:
                    if net_qty > 0:
                        new_sales.append(item)
                    if cancelled_qty > 0 or _has_cancel_event_status(status):
                        _enqueue_notification(
                            "cancel",
                            telegram_id,
                            shop_id,
                            f"{identity}:cancel:{cancelled_qty:g}:{_status_upper(status)}",
                            {"item": item, "cancel_qty": cancelled_qty},
                        )
                    continue

                previous_cancelled = float(before.get("cancelled_qty") or 0)
                if cancelled_qty > previous_cancelled:
                    delta = cancelled_qty - previous_cancelled
                    _enqueue_notification(
                        "cancel",
                        telegram_id,
                        shop_id,
                        f"{identity}:cancel:{cancelled_qty:g}:{_status_upper(status)}",
                        {"item": item, "cancel_qty": delta},
                    )
                elif (
                    _has_cancel_event_status(status)
                    and not _has_cancel_event_status(before.get("status"))
                ):
                    # Do not invent a cancelled quantity for partial cancellation
                    # when the API omitted it. The message will explicitly say that
                    # the exact quantity was not supplied.
                    fallback_qty = (
                        _finance_original_qty(item)
                        if _is_full_cancelled_status(status)
                        else 0.0
                    )
                    _enqueue_notification(
                        "cancel",
                        telegram_id,
                        shop_id,
                        f"{identity}:cancel-status:{_status_upper(status)}",
                        {"item": item, "cancel_qty": fallback_qty},
                    )

            if (
                SALE_DIGEST_THRESHOLD > 0
                and len(new_sales) > SALE_DIGEST_THRESHOLD
            ):
                digest_ids = sorted(finance_identity_key(item) for item in new_sales)
                digest_hash = hashlib.sha256(
                    "|".join(digest_ids).encode("utf-8")
                ).hexdigest()[:32]
                _enqueue_notification(
                    "sale_digest",
                    telegram_id,
                    shop_id,
                    f"digest:{digest_hash}",
                    {"summary": _sale_digest_payload(new_sales)},
                )
                logging.info(
                    "High-volume sale digest queued user=%s shop=%s rows=%s",
                    telegram_id,
                    shop_id,
                    len(new_sales),
                )
            else:
                for item in new_sales:
                    identity = finance_identity_key(item)
                    _enqueue_notification(
                        "sale",
                        telegram_id,
                        shop_id,
                        identity,
                        {"item": item},
                    )

            # State is committed only after durable outbox rows were inserted.
            # A crash before this point causes harmless deduplicated re-enqueueing.
            _save_finance_watch_rows(telegram_id, shop_id, rows)
            _sales_watch_initialized.add(telegram_id)

        await asyncio.sleep(0.2)

    await _deliver_pending_notifications()


def _minimal_stock_snapshot(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = stock_row_key(row)
        result[key] = {
            "total": int(float(row.get("total") or 0)),
            "fbo": int(float(row.get("fbo") or 0)),
            "fbs": int(float(row.get("fbs") or 0)),
        }
    return result


def _load_stock_watch_snapshot(
    telegram_id: int,
    shop_id: int,
    watch_type: str,
) -> dict[str, dict[str, Any]] | None:
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT snapshot_json
            FROM stock_watch_snapshot
            WHERE telegram_id = ? AND shop_id = ? AND watch_type = ?
            """,
            (int(telegram_id), int(shop_id), str(watch_type)),
        ).fetchone()
    if not row:
        return None
    try:
        value = json.loads(str(row["snapshot_json"] or "{}"))
        return value if isinstance(value, dict) else None
    except Exception:
        return None


def _save_stock_watch_snapshot(
    telegram_id: int,
    shop_id: int,
    watch_type: str,
    snapshot: dict[str, dict[str, Any]],
) -> None:
    now_text = _dt_to_db(_utc_now()) or ""
    with db.connect() as conn:
        conn.execute(
            """
            INSERT INTO stock_watch_snapshot
                (telegram_id, shop_id, watch_type, snapshot_json, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(telegram_id, shop_id, watch_type) DO UPDATE SET
                snapshot_json = excluded.snapshot_json,
                updated_at = excluded.updated_at
            """,
            (
                int(telegram_id),
                int(shop_id),
                str(watch_type),
                json.dumps(snapshot, ensure_ascii=False),
                now_text,
            ),
        )
        conn.commit()


_STOCK_WATCH_FETCH_CACHE: dict[str, tuple[float, list[dict[str, Any]]]] = {}
_STOCK_WATCH_FETCH_LOCK = asyncio.Lock()


async def _stock_rows_for_watch(
    encrypted_token: str,
    shop_id: int,
) -> list[dict[str, Any]]:
    cache_key = hashlib.sha1(f"{shop_id}:{encrypted_token}".encode("utf-8")).hexdigest()
    now_mono = time.monotonic()
    cached = _STOCK_WATCH_FETCH_CACHE.get(cache_key)
    if cached and now_mono - cached[0] <= STOCK_WATCH_CACHE_SECONDS:
        return cached[1]

    async with _STOCK_WATCH_FETCH_LOCK:
        cached = _STOCK_WATCH_FETCH_CACHE.get(cache_key)
        now_mono = time.monotonic()
        if cached and now_mono - cached[0] <= STOCK_WATCH_CACHE_SECONDS:
            return cached[1]
        token = cipher.decrypt(encrypted_token)
        client = UzumClient(token, UZUM_API_BASE_URL)
        rows = await load_sku_rows(client, shop_id, max_pages=50)
        _STOCK_WATCH_FETCH_CACHE[cache_key] = (time.monotonic(), rows)
        return rows


async def check_low_stock_once() -> None:
    for group in connected_watch_groups("notify_low_stock"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(value) for value in group["telegram_ids"]]
        try:
            rows = await _stock_rows_for_watch(encrypted_token, shop_id)
        except Exception:
            logging.exception("Low stock watcher failed shop=%s", shop_id)
            continue

        current = _minimal_stock_snapshot(rows)
        rows_by_key = {stock_row_key(row): row for row in rows}
        for telegram_id in telegram_ids:
            settings = ensure_product_settings(telegram_id)
            threshold = max(
                0,
                int(settings.get("low_stock_threshold") or LOW_STOCK_THRESHOLD),
            )
            previous = _load_stock_watch_snapshot(telegram_id, shop_id, "low")
            if previous is None:
                _save_stock_watch_snapshot(telegram_id, shop_id, "low", current)
                _set_watch_initialized(telegram_id, shop_id, "low")
                _low_stock_watch_initialized.add(telegram_id)
                logging.info(
                    "Low stock watcher baseline user=%s shop=%s skus=%s threshold=%s",
                    telegram_id,
                    shop_id,
                    len(current),
                    threshold,
                )
                continue

            event_rows: list[dict[str, Any]] = []
            for key, after in current.items():
                total = int(after.get("total") or 0)
                before_total = int((previous.get(key) or {}).get("total") or 0)
                if 0 < total <= threshold and (
                    key not in previous or before_total <= 0 or before_total > threshold
                ):
                    row = rows_by_key.get(key)
                    if row:
                        event_rows.append(row)

            delivered = True
            if event_rows:
                lines = [format_sku_stock_line(row, mode="all") for row in event_rows[:10]]
                more = "" if len(event_rows) <= 10 else f"\n\nЕщё SKU: {len(event_rows) - 10}"
                text = (
                    "📉 <b>Товар заканчивается</b>\n"
                    f"Магазин: <code>{shop_id}</code>\n"
                    f"Порог: от 1 до <b>{threshold}</b> шт.\n"
                    f"Новых позиций: <b>{len(event_rows)}</b>\n\n"
                    + "\n\n".join(lines)
                    + more
                    + f"\n\nПоказать все: <code>/lowstock {threshold}</code>"
                )
                try:
                    await bot.send_message(
                        telegram_id,
                        text,
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                except Exception:
                    delivered = False
                    logging.exception("Low stock watcher send failed user=%s shop=%s", telegram_id, shop_id)
            if delivered:
                _save_stock_watch_snapshot(telegram_id, shop_id, "low", current)
                _low_stock_watch_initialized.add(telegram_id)
        await asyncio.sleep(0.2)


async def check_out_of_stock_once() -> None:
    for group in connected_watch_groups("notify_out_of_stock"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(value) for value in group["telegram_ids"]]
        try:
            rows = await _stock_rows_for_watch(encrypted_token, shop_id)
        except Exception:
            logging.exception("Out-of-stock watcher failed shop=%s", shop_id)
            continue

        current = _minimal_stock_snapshot(rows)
        rows_by_key = {stock_row_key(row): row for row in rows}
        for telegram_id in telegram_ids:
            previous = _load_stock_watch_snapshot(telegram_id, shop_id, "zero")
            if previous is None:
                _save_stock_watch_snapshot(telegram_id, shop_id, "zero", current)
                _set_watch_initialized(telegram_id, shop_id, "zero")
                _out_of_stock_watch_initialized.add(telegram_id)
                logging.info(
                    "Out-of-stock watcher baseline user=%s shop=%s skus=%s",
                    telegram_id,
                    shop_id,
                    len(current),
                )
                continue

            event_rows: list[dict[str, Any]] = []
            for key, after in current.items():
                total = int(after.get("total") or 0)
                before = previous.get(key)
                before_total = int((before or {}).get("total") or 0)
                if total == 0 and (before is None or before_total > 0):
                    row = rows_by_key.get(key)
                    if row:
                        event_rows.append(row)

            delivered = True
            if event_rows:
                lines = [format_sku_stock_line(row, mode="all") for row in event_rows[:10]]
                more = "" if len(event_rows) <= 10 else f"\n\nЕщё SKU: {len(event_rows) - 10}"
                text = (
                    "❌ <b>Товар закончился</b>\n"
                    f"Магазин: <code>{shop_id}</code>\n"
                    f"Новых позиций с остатком 0: <b>{len(event_rows)}</b>\n\n"
                    + "\n\n".join(lines)
                    + more
                    + "\n\nПоказать остатки: <code>/stock</code>"
                )
                try:
                    await bot.send_message(
                        telegram_id,
                        text,
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                except Exception:
                    delivered = False
                    logging.exception("Out-of-stock watcher send failed user=%s shop=%s", telegram_id, shop_id)
            if delivered:
                _save_stock_watch_snapshot(telegram_id, shop_id, "zero", current)
                _out_of_stock_watch_initialized.add(telegram_id)
        await asyncio.sleep(0.2)


async def check_stock_change_once() -> None:
    for group in connected_watch_groups("notify_stock_change"):
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(value) for value in group["telegram_ids"]]
        try:
            rows = await _stock_rows_for_watch(encrypted_token, shop_id)
        except Exception:
            logging.exception("Stock change watcher failed shop=%s", shop_id)
            continue

        current = _minimal_stock_snapshot(rows)
        rows_by_key = {stock_row_key(row): row for row in rows}
        for telegram_id in telegram_ids:
            previous = _load_stock_watch_snapshot(telegram_id, shop_id, "change")
            if previous is None:
                _save_stock_watch_snapshot(telegram_id, shop_id, "change", current)
                _set_watch_initialized(telegram_id, shop_id, "change")
                _stock_change_watch_initialized.add(telegram_id)
                logging.info(
                    "Stock change watcher baseline user=%s shop=%s skus=%s",
                    telegram_id,
                    shop_id,
                    len(current),
                )
                continue

            decreased: list[tuple[str, dict[str, Any], dict[str, Any]]] = []
            for key, after in current.items():
                before = previous.get(key)
                if not before:
                    continue
                if (
                    int(after.get("total") or 0) < int(before.get("total") or 0)
                    or int(after.get("fbo") or 0) < int(before.get("fbo") or 0)
                    or int(after.get("fbs") or 0) < int(before.get("fbs") or 0)
                ):
                    before_display = {
                        **before,
                        "title": (
                            rows_by_key.get(key, {}).get("sku_full_title")
                            or rows_by_key.get(key, {}).get("sku_title")
                            or rows_by_key.get(key, {}).get("product_title")
                            or key
                        ),
                    }
                    after_display = {
                        **after,
                        "title": before_display["title"],
                    }
                    decreased.append((key, before_display, after_display))

            delivered = True
            if decreased:
                lines = [
                    _format_stock_change_line(key, before, after)
                    for key, before, after in decreased[:10]
                ]
                more = "" if len(decreased) <= 10 else f"\n\nЕщё изменений: {len(decreased) - 10}"
                text = (
                    "📦 <b>Изменение остатков</b>\n"
                    f"Магазин: <code>{shop_id}</code>\n"
                    "Уменьшился остаток по SKU.\n\n"
                    + "\n\n".join(lines)
                    + more
                    + "\n\nПроверить остатки: <code>/stock</code>"
                )
                try:
                    await bot.send_message(
                        telegram_id,
                        text,
                        reply_markup=main_menu_for_user(telegram_id),
                    )
                except Exception:
                    delivered = False
                    logging.exception("Stock change watcher send failed user=%s shop=%s", telegram_id, shop_id)
            if delivered:
                _save_stock_watch_snapshot(telegram_id, shop_id, "change", current)
                _stock_change_watch_initialized.add(telegram_id)
        await asyncio.sleep(0.2)


_LOSS_REPORT_CACHE: dict[int, tuple[float, list[dict[str, Any]], list[str]]] = {}


async def _load_all_time_loss_rows(
    client: UzumClient,
    shop_id: int,
    *,
    force_refresh: bool = False,
) -> tuple[list[dict[str, Any]], list[str]]:
    cached = _LOSS_REPORT_CACHE.get(int(shop_id))
    if (
        not force_refresh
        and cached
        and time.monotonic() - cached[0] <= LOSS_REPORT_CACHE_SECONDS
    ):
        import copy
        logging.info(
            "Loss report cache hit shop=%s rows=%s unavailable_filters=%s",
            shop_id,
            len(cached[1]),
            ",".join(cached[2]) or "none",
        )
        return copy.deepcopy(cached[1]), list(cached[2])

    merged: dict[str, dict[str, Any]] = {}
    unavailable_filters: list[str] = []
    first_error: Exception | None = None
    successful_filters = 0
    filters = LOSS_REPORT_FILTERS
    pages_per_filter = max(1, LOSS_REPORT_MAX_REQUESTS // max(1, len(filters)))

    for product_filter in filters:
        try:
            products = await load_products(
                client,
                shop_id,
                max_pages=pages_per_filter,
                page_size=100,
                product_filter=product_filter,
            )
            successful_filters += 1
            logging.info(
                "Loss report scan shop=%s filter=%s products=%s max_pages=%s",
                shop_id,
                product_filter,
                len(products),
                pages_per_filter,
            )
        except Exception as exc:
            first_error = first_error or exc
            unavailable_filters.append(product_filter)
            logging.warning(
                "Loss report filter unavailable filter=%s shop=%s: %s",
                product_filter,
                shop_id,
                exc,
            )
            continue

        for product in products:
            if not isinstance(product, dict):
                continue
            product_archived = bool(product.get("archived")) or product_filter == "ARCHIVE"
            for source_row in flatten_sku_rows([product]):
                row = dict(source_row)
                raw = row.get("raw") if isinstance(row.get("raw"), dict) else {}
                row["archived"] = product_archived or bool(raw.get("archived"))
                row["loss_all_time"] = True
                row["loss_only"] = True
                row["missing"] = _loss_qty(row, "missing")
                row["defected"] = _loss_qty(row, "defected")
                if row["missing"] <= 0 and row["defected"] <= 0:
                    continue

                key = _loss_row_key(row)
                current = merged.get(key)
                if current is None:
                    row["source_filters"] = [product_filter]
                    merged[key] = row
                    continue

                current["missing"] = max(_loss_qty(current, "missing"), row["missing"])
                current["defected"] = max(_loss_qty(current, "defected"), row["defected"])
                current["archived"] = bool(current.get("archived")) or bool(row.get("archived"))
                source_filters = list(current.get("source_filters") or [])
                if product_filter not in source_filters:
                    source_filters.append(product_filter)
                current["source_filters"] = source_filters
                for field in (
                    "price",
                    "commission",
                    "total",
                    "active",
                    "fbo",
                    "fbs",
                    "status",
                    "product_id",
                    "sku_id",
                    "barcode",
                    "seller_item_code",
                ):
                    if current.get(field) in (None, "", "—") and row.get(field) not in (None, "", "—"):
                        current[field] = row.get(field)

    if successful_filters == 0 and first_error is not None:
        raise first_error

    rows = list(merged.values())
    rows.sort(
        key=lambda row: (
            -(_loss_qty(row, "missing") + _loss_qty(row, "defected")),
            str(row.get("product_title") or row.get("sku_full_title") or ""),
        )
    )
    logging.info(
        "Loss report completed shop=%s filters=%s rows=%s unavailable_filters=%s force_refresh=%s",
        shop_id,
        ",".join(filters),
        len(rows),
        ",".join(unavailable_filters) or "none",
        force_refresh,
    )
    # Не сохраняем пустой или частичный результат: временный ответ Uzum не
    # должен превращаться в ложное «потерь нет» на весь срок кэша.
    if rows and not unavailable_filters:
        import copy
        _LOSS_REPORT_CACHE[int(shop_id)] = (
            time.monotonic(),
            copy.deepcopy(rows),
            [],
        )
    else:
        _LOSS_REPORT_CACHE.pop(int(shop_id), None)
    return rows, unavailable_filters


def _attach_current_stock_to_losses(
    loss_rows: list[dict[str, Any]],
    current_stock_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_key: dict[str, dict[str, Any]] = {}
    for stock_row in current_stock_rows:
        for key in _stock_match_keys(stock_row):
            by_key.setdefault(key, stock_row)

    result: list[dict[str, Any]] = []
    for loss_row in loss_rows:
        updated = dict(loss_row)
        current: dict[str, Any] | None = None
        for key in _stock_match_keys(loss_row):
            if key in by_key:
                current = by_key[key]
                break
        if current is not None:
            for field in ("total", "active", "fbo", "fbs", "status"):
                updated[field] = current.get(field)
            if current.get("price") not in (None, "", "—"):
                updated["price"] = current.get("price")
            updated["current_stock_found"] = True
        else:
            # Archived/loss-only SKU: no current active stock row exists.
            updated["total"] = 0
            updated["fbo"] = 0
            updated["fbs"] = 0
            updated["current_stock_found"] = False
        result.append(updated)
    return result


def _cleanup_release_state() -> None:
    cutoff = _dt_to_db(_utc_now() - timedelta(days=WATCH_STATE_RETENTION_DAYS)) or ""
    with db.connect() as conn:
        conn.execute(
            "DELETE FROM notification_outbox WHERE status IN ('sent', 'dead') AND updated_at < ?",
            (cutoff,),
        )
        conn.execute(
            "DELETE FROM finance_watch_state WHERE updated_at < ?",
            (cutoff,),
        )
        conn.commit()


_cleanup_release_state()


# =============================================================================
# PREMIUM MERGE 2026-07-18
# Preserves per-user instant/hourly modes while using the durable outbox and
# adaptive Finance pagination from RELEASE_HARDENING.
# =============================================================================
PREMIUM_RELEASE_VERSION = "2026.07.22-premium-r17-clear-profit-bridge"

WATCHER_ACCESS_BACKOFF_SECONDS = max(
    300,
    int(os.getenv("WATCHER_ACCESS_BACKOFF_SECONDS", "3600") or "3600"),
)
_WATCHER_ACCESS_BACKOFF_UNTIL: dict[tuple[str, str], float] = {}


def _watcher_access_backoff_key(watcher: str, group: dict[str, Any]) -> tuple[str, str]:
    return str(watcher), _watch_group_key(group)


def _watcher_access_is_paused(watcher: str, group: dict[str, Any]) -> bool:
    key = _watcher_access_backoff_key(watcher, group)
    until = float(_WATCHER_ACCESS_BACKOFF_UNTIL.get(key, 0.0))
    if until <= time.monotonic():
        _WATCHER_ACCESS_BACKOFF_UNTIL.pop(key, None)
        return False
    return True


def _watcher_access_pause(watcher: str, group: dict[str, Any], error: BaseException) -> None:
    if _uzum_access_error_kind(error):
        _WATCHER_ACCESS_BACKOFF_UNTIL[_watcher_access_backoff_key(watcher, group)] = (
            time.monotonic() + WATCHER_ACCESS_BACKOFF_SECONDS
        )


def _watcher_access_resume(watcher: str, group: dict[str, Any]) -> None:
    _WATCHER_ACCESS_BACKOFF_UNTIL.pop(_watcher_access_backoff_key(watcher, group), None)


async def check_new_sales_once() -> None:
    hourly_users: set[int] = set()
    watcher_name = "Premium finance watcher"
    for group in connected_watch_groups(
        "notify_sales",
        "notify_cancellations",
        access_feature="sales_notifications",
    ):
        if _watcher_access_is_paused(watcher_name, group):
            continue
        shop_id = int(group["shop_id"])
        encrypted_token = group["uzum_token_encrypted"]
        telegram_ids = [int(value) for value in group["telegram_ids"]]
        hourly_users.update(
            telegram_id
            for telegram_id in telegram_ids
            if get_sales_notification_mode(telegram_id) == "hourly"
        )

        try:
            token = cipher.decrypt(encrypted_token)
            client = UzumClient(token, UZUM_API_BASE_URL)
            state_by_user = {
                telegram_id: _load_finance_watch_state(telegram_id, shop_id)
                for telegram_id in telegram_ids
            }
            initialized_by_user = {
                telegram_id: _watch_is_initialized(telegram_id, shop_id, "finance")
                for telegram_id in telegram_ids
            }
            known_union: set[str] = set()
            for state in state_by_user.values():
                known_union.update(state.keys())
            baseline_group = not any(initialized_by_user.values())
            recent_rows = await _load_finance_watch_pages(
                client,
                shop_id,
                known_identities=known_union,
                baseline=baseline_group,
            )
            need_cancellations = any(
                has_paid_subscription(telegram_id)
                and product_setting_enabled(telegram_id, "notify_cancellations")
                for telegram_id in telegram_ids
            )
            cancel_rows = (
                await _load_cancel_status_rows(client, shop_id)
                if need_cancellations
                else []
            )

            merged: dict[str, dict[str, Any]] = {}
            for item in recent_rows + cancel_rows:
                identity = finance_identity_key(item)
                current = merged.get(identity)
                if current is None:
                    merged[identity] = item
                    continue
                current_score = (
                    _finance_cancelled_qty(current),
                    _finance_return_qty(current),
                    str(_finance_status(current)),
                )
                item_score = (
                    _finance_cancelled_qty(item),
                    _finance_return_qty(item),
                    str(_finance_status(item)),
                )
                if item_score > current_score:
                    merged[identity] = item
            rows = list(merged.values())
            _watcher_access_resume(watcher_name, group)
        except Exception as error:
            _watcher_access_pause(watcher_name, group, error)
            _log_watcher_api_failure(
                watcher_name,
                error,
                shop_id=shop_id,
                telegram_ids=telegram_ids,
            )
            await asyncio.sleep(2)
            continue

        now = _utc_now()
        for telegram_id in telegram_ids:
            previous = state_by_user.get(telegram_id, {})
            initialized = initialized_by_user.get(telegram_id, False)
            sales_mode = get_sales_notification_mode(telegram_id)
            cancellations_enabled = (
                has_paid_subscription(telegram_id)
                and product_setting_enabled(
                    telegram_id,
                    "notify_cancellations",
                )
            )

            if not initialized:
                _save_finance_watch_rows(telegram_id, shop_id, rows)
                _set_watch_initialized(telegram_id, shop_id, "finance")
                mark_operational_watcher_initialized(telegram_id, shop_id, "sales")
                logging.info(
                    "Sales watcher initialized user=%s shop=%s rows=%s mode=%s",
                    telegram_id,
                    shop_id,
                    len(rows),
                    sales_mode,
                )
                continue

            new_sales: list[dict[str, Any]] = []
            for item in rows:
                identity = finance_identity_key(item)
                before = previous.get(identity)
                net_qty = _finance_qty(item)
                cancelled_qty = _finance_cancelled_qty(item)
                status = _finance_status(item)

                if before is None:
                    if net_qty > 0:
                        new_sales.append(item)
                    if cancellations_enabled and (
                        cancelled_qty > 0 or _has_cancel_event_status(status)
                    ):
                        _enqueue_notification(
                            "cancel",
                            telegram_id,
                            shop_id,
                            f"{identity}:cancel:{cancelled_qty:g}:{_status_upper(status)}",
                            {"item": item, "cancel_qty": cancelled_qty},
                        )
                    continue

                if cancellations_enabled:
                    previous_cancelled = float(before.get("cancelled_qty") or 0)
                    if cancelled_qty > previous_cancelled:
                        delta = cancelled_qty - previous_cancelled
                        _enqueue_notification(
                            "cancel",
                            telegram_id,
                            shop_id,
                            f"{identity}:cancel:{cancelled_qty:g}:{_status_upper(status)}",
                            {"item": item, "cancel_qty": delta},
                        )
                    elif (
                        _has_cancel_event_status(status)
                        and not _has_cancel_event_status(before.get("status"))
                    ):
                        fallback_qty = (
                            _finance_original_qty(item)
                            if _is_full_cancelled_status(status)
                            else 0.0
                        )
                        _enqueue_notification(
                            "cancel",
                            telegram_id,
                            shop_id,
                            f"{identity}:cancel-status:{_status_upper(status)}",
                            {"item": item, "cancel_qty": fallback_qty},
                        )

            if sales_mode == "hourly":
                enqueue_sales_digest_events(
                    telegram_id,
                    shop_id,
                    new_sales,
                    detected_at=now,
                )
            elif sales_mode == "instant":
                if len(new_sales) > INSTANT_SALE_BURST_LIMIT:
                    direct = new_sales[:INSTANT_SALE_BURST_LIMIT]
                    overflow = new_sales[INSTANT_SALE_BURST_LIMIT:]
                else:
                    direct = new_sales
                    overflow = []
                for item in direct:
                    _enqueue_notification(
                        "sale",
                        telegram_id,
                        shop_id,
                        finance_identity_key(item),
                        {"item": item},
                    )
                if overflow:
                    digest_ids = sorted(finance_identity_key(item) for item in overflow)
                    digest_hash = hashlib.sha256(
                        "|".join(digest_ids).encode("utf-8")
                    ).hexdigest()[:32]
                    _enqueue_notification(
                        "sale_digest",
                        telegram_id,
                        shop_id,
                        f"burst:{digest_hash}",
                        {"summary": _sale_digest_payload(overflow)},
                    )

            # State is written only after notification data is durably stored.
            _save_finance_watch_rows(telegram_id, shop_id, rows)

        await asyncio.sleep(0.2)

    digest_now = _utc_now()
    for telegram_id in sorted(hourly_users):
        await maybe_send_hourly_sales_digest_all_shops(
            telegram_id,
            now=digest_now,
        )

    await _deliver_pending_notifications()


_BACKGROUND_TASKS: set[asyncio.Task[Any]] = set()


def _spawn_background_task(coro: Any, name: str) -> asyncio.Task[Any]:
    task = asyncio.create_task(coro, name=name)
    _BACKGROUND_TASKS.add(task)

    def _done(completed: asyncio.Task[Any]) -> None:
        _BACKGROUND_TASKS.discard(completed)
        if completed.cancelled():
            return
        try:
            error = completed.exception()
        except asyncio.CancelledError:
            return
        if error is not None:
            logging.error(
                "Background task stopped unexpectedly name=%s",
                completed.get_name(),
                exc_info=(type(error), error, error.__traceback__),
            )

    task.add_done_callback(_done)
    return task


async def main() -> None:
    logging.info(
        "PREMIUM_RELEASE_LOADED version=%s base=%s",
        PREMIUM_RELEASE_VERSION,
        APP_BUILD,
    )
    logging.info("DATABASE_READY path=%s wal=%s", DB_PATH, SQLITE_WAL)
    logging.info(
        "RELEASE_HARDENING_LOADED version=%s: finance math + persistent watchers + retry outbox",
        RELEASE_VERSION,
    )
    logging.info(
        "Premium config: sale_interval=%ss watch_pages=%s report_pages=%s drop_pending=%s "
        "API_min_interval=%.2fs cancel_deep_pages=%s cancel_deep_interval=%ss",
        SALE_CHECK_INTERVAL_SECONDS,
        SALE_WATCH_MAX_PAGES,
        FINANCE_REPORT_MAX_PAGES,
        DROP_PENDING_UPDATES,
        UZUM_API_MIN_INTERVAL_SECONDS,
        CANCEL_WATCH_DEEP_SCAN_MAX_PAGES,
        CANCEL_WATCH_DEEP_SCAN_INTERVAL_SECONDS,
    )
    logging.info("PREMIUM_UI_LOADED: compact main menu + guided sections + honest financial wording")
    logging.info("TRIAL_FEATURE_GATING_LOADED: sales today + sale alerts + morning report")
    logging.info("MANAGEMENT_PDF_LOADED: sales + profit + stock + cancellations + returns + defects")
    logging.info("STABILITY_SECURITY_LOADED: stock routes + watcher settings + safe backup + bounded Excel import")
    logging.info("UZUM_FINANCE_LOADED: purchasePrice-only cost + expense ledger + IKPU audit")
    logging.info("REPORT_RECONCILIATION_LOADED: historical purchasePrice + issued-date + cancelled quantity + tax/ROI")
    logging.info("PROFIT_BRIDGE_LOADED: revenue -> payout -> cost -> tax/expenses -> result")
    logging.info("FBO_ACCEPTANCE_RECONCILIATION_LOADED: invoice/product totals + stale-zero guard")
    logging.info("MARKET_STYLE_REPORTS_LOADED: daily PDF/Excel + loss/damage claim documents")
    logging.info("LOGISTICS_REMINDERS_LOADED: FBO slots + return paid-storage deadlines")
    logging.info("MULTI_SHOP_NOTIFICATIONS_LOADED: all connected shops + one combined hourly sales digest")
    logging.info("SALES_NOTIFICATION_DEFAULT_LOADED: hourly for all connected shops")
    try:
        pdf_regular_font, pdf_bold_font = _fbo_pdf_font_paths()
        logging.info(
            "PDF_FONT_READY: regular=%s bold=%s",
            pdf_regular_font,
            pdf_bold_font,
        )
    except RuntimeError as error:
        logging.warning("PDF_FONT_MISSING: %s", error)
    logging.info("LEGACY_CANCEL_STATUS_FILTER_SCAN_DISABLED: finance orders are scanned without unsupported status parameters")
    logging.info("SKU_LABELS_INTERFACE_LOADED: official barcode types + PDF SKU labels")
    logging.info("STOCK_TRUTH_LOADED: SKU-level FBO/FBS totals + supply forecast")

    init_language_tables()
    init_product_value_tables()
    await bot.delete_webhook(drop_pending_updates=DROP_PENDING_UPDATES)

    # Loops have staggered initial sleeps inside their implementations.
    _spawn_background_task(order_watch_loop(), "order_watch")
    _spawn_background_task(low_stock_watch_loop(), "low_stock_watch")
    _spawn_background_task(out_of_stock_watch_loop(), "out_of_stock_watch")
    _spawn_background_task(sales_watch_loop(), "sales_watch")
    if STOCK_CHANGE_NOTIFICATIONS:
        _spawn_background_task(stock_change_watch_loop(), "stock_change_watch")
    _spawn_background_task(loss_defect_watch_loop(), "loss_defect_watch")
    _spawn_background_task(fbo_acceptance_watch_loop(), "fbo_acceptance_watch")
    _spawn_background_task(logistics_reminder_watch_loop(), "logistics_reminder_watch")
    _spawn_background_task(daily_report_loop(), "daily_report")
    if SUBSCRIPTION_REMINDERS:
        _spawn_background_task(subscription_reminder_loop(), "subscription_reminder")
    _spawn_background_task(reviews_watch_loop(), "reviews_watch")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
